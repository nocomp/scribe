"""
app/api/capacite.py — Gestion capacitaire des lits — SCRIBE v1.3.0

Routes :
  GET  /api/v1/capacite/referentiel          → liste des unités avec dernière déclaration
  POST /api/v1/capacite/referentiel          → créer/modifier une unité (admin)
  POST /api/v1/capacite/declaration          → soumettre une déclaration (cadre)
  GET  /api/v1/capacite/declarations         → historique
  GET  /api/v1/capacite/synthese             → agrégation temps réel par pôle/site
  GET  /api/v1/capacite/evolution/{service}  → historique pour graphique
  GET  /api/v1/capacite/export-csv           → export pour main courante / archive
"""
import json
import logging
from datetime import datetime, timezone, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
import io
import pathlib
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    CapaciteReferentiel, CapaciteDeclaration, SitrepEntry
)
from app.api.auth import get_current_user, require_role

logger = logging.getLogger(__name__)
router = APIRouter()


# ── Schémas Pydantic ────────────────────────────────────────────────────────

class ReferentielCreate(BaseModel):
    service_nom:     str
    uf_code:         Optional[str] = None
    pole:            Optional[str] = None
    site:            Optional[str] = None
    capacite_totale: int = 0
    tension_1:       int = 0
    tension_2:       int = 0
    accept_homme:    bool = True
    accept_femme:    bool = True
    accept_indiffer: bool = True
    telephone_cadre: Optional[str] = None
    ordre_affichage: int = 99


class DeclarationCreate(BaseModel):
    referentiel_id:       int
    redacteur:            str
    point:                str = "matin"       # matin | aprem | soir
    lits_vides_h:         int = 0
    lits_vides_f:         int = 0
    lits_vides_i:         int = 0
    tension_activee:      int = 0
    lits_sup:             int = 0
    statut_lits:          str = "normal"
    statut_rh:            str = "complet"
    statut_materiel:      str = "ok"
    alerte_lits:          bool = False
    alerte_rh:            bool = False
    alerte_materiel:      bool = False
    commentaire_lits:     Optional[str] = None
    commentaire_rh:       Optional[str] = None
    commentaire_materiel: Optional[str] = None
    commentaire_general:  Optional[str] = None
    # Mode dégradé (v1.5.0)
    mode_degrade:   bool = False
    besoin_renfort: int  = 0
    peut_preter:    int  = 0


# ── Helpers ─────────────────────────────────────────────────────────────────

STATUT_POIDS = {
    "normal": 0, "complet": 0, "ok": 0,
    "tension": 1, "degrade": 1,
    "critique": 2, "insuffisant": 2,
    "ferme": 3, "hs": 3,
}

def _statut_global(ref: CapaciteReferentiel,
                   decl: Optional[CapaciteDeclaration]) -> str:
    """Calcule le statut global d'une unité depuis sa dernière déclaration."""
    if decl is None:
        return "inconnu"
    poids = max(
        STATUT_POIDS.get(decl.statut_lits, 0),
        STATUT_POIDS.get(decl.statut_rh, 0),
        STATUT_POIDS.get(decl.statut_materiel, 0),
    )
    if decl.alerte_lits or decl.alerte_rh or decl.alerte_materiel:
        poids = max(poids, 2)
    if poids >= 3: return "ferme"
    if poids >= 2: return "critique"
    if poids >= 1: return "tension"
    return "normal"


def _decl_to_dict(decl: CapaciteDeclaration) -> dict:
    return {
        "id":              decl.id,
        "referentiel_id":  decl.referentiel_id,
        "horodatage":      decl.horodatage.isoformat() if decl.horodatage else None,
        "redacteur":       decl.redacteur,
        "point":           decl.point,
        "lits_vides_h":    decl.lits_vides_h,
        "lits_vides_f":    decl.lits_vides_f,
        "lits_vides_i":    decl.lits_vides_i,
        "lits_vides_total": (decl.lits_vides_h or 0) + (decl.lits_vides_f or 0) + (decl.lits_vides_i or 0),
        "tension_activee": decl.tension_activee,
        "lits_sup":        decl.lits_sup,
        "statut_lits":     decl.statut_lits,
        "statut_rh":       decl.statut_rh,
        "statut_materiel": decl.statut_materiel,
        "alerte_lits":     decl.alerte_lits,
        "alerte_rh":       decl.alerte_rh,
        "alerte_materiel": decl.alerte_materiel,
        "commentaire_lits":     decl.commentaire_lits,
        "commentaire_rh":       decl.commentaire_rh,
        "commentaire_materiel": decl.commentaire_materiel,
        "commentaire_general":  decl.commentaire_general,
        "mode_degrade":    getattr(decl, "mode_degrade",   False),
        "besoin_renfort":  getattr(decl, "besoin_renfort",  0),
        "peut_preter":     getattr(decl, "peut_preter",     0),
        "incident_id":     decl.incident_id,
    }


def _ref_to_dict(ref: CapaciteReferentiel,
                 last_decl: Optional[CapaciteDeclaration] = None) -> dict:
    d = {
        "id":              ref.id,
        "service_nom":     ref.service_nom,
        "uf_code":         ref.uf_code,
        "pole":            ref.pole,
        "site":            ref.site,
        "capacite_totale": ref.capacite_totale,
        "tension_1":       ref.tension_1,
        "tension_2":       ref.tension_2,
        "accept_homme":    ref.accept_homme,
        "accept_femme":    ref.accept_femme,
        "accept_indiffer": ref.accept_indiffer,
        "telephone_cadre": ref.telephone_cadre,
        "ordre_affichage": ref.ordre_affichage,
        "actif":           ref.actif,
        "statut_global":   _statut_global(ref, last_decl),
        "derniere_declaration": _decl_to_dict(last_decl) if last_decl else None,
    }
    return d


# ── Routes ───────────────────────────────────────────────────────────────────

@router.get("/referentiel")
def get_referentiel(db: Session = Depends(get_db)):
    """Renvoie toutes les unités avec leur dernière déclaration."""
    refs = db.query(CapaciteReferentiel).filter(
        CapaciteReferentiel.actif == True
    ).order_by(CapaciteReferentiel.site, CapaciteReferentiel.pole,
               CapaciteReferentiel.ordre_affichage).all()

    result = []
    for ref in refs:
        last = (db.query(CapaciteDeclaration)
                  .filter(CapaciteDeclaration.referentiel_id == ref.id)
                  .order_by(CapaciteDeclaration.horodatage.desc())
                  .first())
        result.append(_ref_to_dict(ref, last))
    return result


@router.post("/referentiel")
def create_or_update_referentiel(
    payload: ReferentielCreate,
    db: Session = Depends(get_db),
    user=Depends(require_role("cadre_sante", "admin"))
):
    """Crée ou met à jour une unité de référence."""
    existing = db.query(CapaciteReferentiel).filter(
        CapaciteReferentiel.service_nom == payload.service_nom
    ).first()

    if existing:
        for k, v in payload.dict().items():
            setattr(existing, k, v)
        db.commit()
        db.refresh(existing)
        return _ref_to_dict(existing)
    else:
        ref = CapaciteReferentiel(**payload.dict())
        db.add(ref)
        db.commit()
        db.refresh(ref)
        return _ref_to_dict(ref)


@router.post("/declaration")
def submit_declaration(
    payload: DeclarationCreate,
    db: Session = Depends(get_db)
):
    """Soumet une déclaration capacitaire. Crée un incident automatiquement si alerte."""
    ref = db.query(CapaciteReferentiel).filter(
        CapaciteReferentiel.id == payload.referentiel_id
    ).first()
    if not ref:
        raise HTTPException(status_code=404, detail="Unité introuvable")

    decl = CapaciteDeclaration(**payload.dict())
    db.add(decl)
    db.flush()  # pour avoir l'id

    incident_id = None
    if payload.alerte_lits or payload.alerte_rh or payload.alerte_materiel:
        alertes = []
        if payload.alerte_lits:     alertes.append("Lits")
        if payload.alerte_rh:       alertes.append("RH")
        if payload.alerte_materiel: alertes.append("Matériel")
        alerte_str = " / ".join(alertes)

        # Déterminer l'urgence selon le statut le plus grave
        poids = max(
            STATUT_POIDS.get(payload.statut_lits, 0),
            STATUT_POIDS.get(payload.statut_rh, 0),
            STATUT_POIDS.get(payload.statut_materiel, 0),
        )
        urgency = 4 if poids >= 3 else 3 if poids >= 2 else 2

        commentaire_combined = " | ".join(filter(None, [
            payload.commentaire_lits,
            payload.commentaire_rh,
            payload.commentaire_materiel,
            payload.commentaire_general,
        ]))

        incident = SitrepEntry(
            declarant_nom=payload.redacteur,
            site_id=ref.site or ref.service_nom,
            type_crise="SANITAIRE",
            urgency=urgency,
            fait=f"[CAPACITÉ] Seuil d'alerte déclaré — {ref.service_nom} — {alerte_str}",
            analyse=commentaire_combined or f"Déclaration capacitaire {payload.point}",
            status="EN COURS",
            directeur_crise=None,
            jalons="[]",
        )
        db.add(incident)
        db.flush()
        incident_id = incident.id
        decl.incident_id = incident_id

    db.commit()
    db.refresh(decl)

    return {
        "ok": True,
        "declaration_id": decl.id,
        "incident_id": incident_id,
        "statut_global": _statut_global(ref, decl),
    }


@router.get("/declarations")
def get_declarations(
    referentiel_id: Optional[int] = None,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """Historique des déclarations."""
    q = db.query(CapaciteDeclaration).order_by(CapaciteDeclaration.horodatage.desc())
    if referentiel_id:
        q = q.filter(CapaciteDeclaration.referentiel_id == referentiel_id)
    return [_decl_to_dict(d) for d in q.limit(limit).all()]


@router.get("/synthese")
def get_synthese(db: Session = Depends(get_db)):
    """Agrégation par site et pôle — vue d'ensemble pour la cellule."""
    refs = db.query(CapaciteReferentiel).filter(
        CapaciteReferentiel.actif == True
    ).all()

    by_site = {}
    for ref in refs:
        site = ref.site or "Autre"
        pole = ref.pole or "Autre"
        if site not in by_site:
            by_site[site] = {}
        if pole not in by_site[site]:
            by_site[site][pole] = {
                "lits_total": 0, "lits_vides_h": 0, "lits_vides_f": 0,
                "lits_vides_i": 0, "alertes": 0, "non_declares": 0,
                "statuts": [], "services": []
            }

        last = (db.query(CapaciteDeclaration)
                  .filter(CapaciteDeclaration.referentiel_id == ref.id)
                  .order_by(CapaciteDeclaration.horodatage.desc())
                  .first())

        statut = _statut_global(ref, last)
        pg = by_site[site][pole]
        pg["lits_total"] += ref.capacite_totale or 0
        pg["statuts"].append(statut)
        pg["services"].append({
            "id": ref.id, "nom": ref.service_nom,
            "statut": statut,
            "lits_vides_h": last.lits_vides_h if last else None,
            "lits_vides_f": last.lits_vides_f if last else None,
            "lits_vides_i": last.lits_vides_i if last else None,
            "horodatage": last.horodatage.isoformat() if last and last.horodatage else None,
            "redacteur": last.redacteur if last else None,
            "alerte": (last.alerte_lits or last.alerte_rh or last.alerte_materiel) if last else False,
            "mode_degrade":   getattr(last, "mode_degrade",   False) if last else False,
            "besoin_renfort": getattr(last, "besoin_renfort",  0)    if last else 0,
            "peut_preter":    getattr(last, "peut_preter",     0)    if last else 0,
        })
        if statut == "inconnu":
            pg["non_declares"] += 1
        if last and (last.alerte_lits or last.alerte_rh or last.alerte_materiel):
            pg["alertes"] += 1
        if last:
            pg["lits_vides_h"] += last.lits_vides_h or 0
            pg["lits_vides_f"] += last.lits_vides_f or 0
            pg["lits_vides_i"] += last.lits_vides_i or 0

    # Calcul statut global par pôle
    POIDS_MAP = {"ferme": 3, "critique": 2, "tension": 1, "normal": 0, "inconnu": -1}
    result = {}
    for site, poles in by_site.items():
        result[site] = {}
        for pole, data in poles.items():
            poids_liste = [POIDS_MAP.get(s, -1) for s in data["statuts"]]
            max_p = max(poids_liste) if poids_liste else -1
            statut_pole = {3:"ferme", 2:"critique", 1:"tension", 0:"normal"}.get(max_p, "inconnu")
            result[site][pole] = {**data, "statut_pole": statut_pole}

    return result


@router.get("/evolution/{referentiel_id}")
def get_evolution(referentiel_id: int, jours: int = 3, db: Session = Depends(get_db)):
    """Historique pour graphique d'évolution (derniers N jours)."""
    depuis = datetime.now(timezone.utc) - timedelta(days=jours)
    decls = (db.query(CapaciteDeclaration)
               .filter(CapaciteDeclaration.referentiel_id == referentiel_id)
               .filter(CapaciteDeclaration.horodatage >= depuis)
               .order_by(CapaciteDeclaration.horodatage)
               .all())

    ref = db.query(CapaciteReferentiel).get(referentiel_id)
    return {
        "referentiel": _ref_to_dict(ref) if ref else None,
        "declarations": [_decl_to_dict(d) for d in decls],
    }


@router.get("/export-csv")
def export_capacite_csv(db: Session = Depends(get_db)):
    """Export CSV des déclarations — intégré dans la main courante."""
    import csv, io as _io
    from fastapi.responses import StreamingResponse

    decls = (db.query(CapaciteDeclaration)
               .order_by(CapaciteDeclaration.horodatage)
               .all())

    refs = {r.id: r for r in db.query(CapaciteReferentiel).all()}

    output = _io.StringIO()
    w = csv.writer(output, delimiter=";", quotechar='"', quoting=csv.QUOTE_ALL)
    w.writerow([
        "Horodatage", "Service", "Pôle", "Site", "Déclarant", "Point",
        "Statut lits", "Vides H", "Vides F", "Vides I", "Tension activée",
        "Statut RH", "Statut Matériel",
        "Alerte lits", "Alerte RH", "Alerte matériel",
        "Commentaire général"
    ])
    for d in decls:
        ref = refs.get(d.referentiel_id)
        w.writerow([
            d.horodatage.strftime("%d/%m/%Y %H:%M") if d.horodatage else "",
            ref.service_nom if ref else "",
            ref.pole if ref else "",
            ref.site if ref else "",
            d.redacteur or "",
            d.point or "",
            d.statut_lits or "",
            d.lits_vides_h or 0,
            d.lits_vides_f or 0,
            d.lits_vides_i or 0,
            d.tension_activee or 0,
            d.statut_rh or "",
            d.statut_materiel or "",
            "OUI" if d.alerte_lits else "",
            "OUI" if d.alerte_rh else "",
            "OUI" if d.alerte_materiel else "",
            d.commentaire_general or "",
        ])

    output.seek(0)
    now_str = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=capacite_{now_str}.csv"}
    )


@router.post("/import-xlsx")
async def import_from_xlsx(
    file: UploadFile = File(...),
    user=Depends(require_role("cadre_sante", "admin")),
    db: Session = Depends(get_db),
):
    """h151 — Import Excel du référentiel capacitaire.
    Colonnes attendues (ligne 5 = en-têtes, données dès ligne 6) :
      N° UF | Nom du service | Pôle | Site | Établissement |
      Lits total | Personnel soignant (ETP) | Accept H | Accept F | Ordre affichage
    Les UF existantes (même uf_code) sont mises à jour, les nouvelles créées.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl non installé — pip install openpyxl")
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        # Chercher la feuille Capacite_Lits ou la première
        ws = wb["Capacite_Lits"] if "Capacite_Lits" in wb.sheetnames else wb.active
        # Ligne 5 = en-têtes, données dès ligne 6
        header_row = 5
        data_start = 6
        rows_processed = 0
        rows_created = 0
        rows_updated = 0
        errors = []
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            # Ignorer les lignes vides ou les notes
            if not row or not row[0]:
                continue
            uf_code  = str(row[0]).strip() if row[0] is not None else ""
            svc_nom  = str(row[1]).strip() if row[1] is not None else ""
            pole     = str(row[2]).strip() if row[2] is not None else ""
            site     = str(row[3]).strip() if row[3] is not None else ""
            etab     = str(row[4]).strip() if row[4] is not None else ""
            lits     = int(row[5]) if row[5] is not None and str(row[5]).isdigit() else 0
            personnel = int(row[6]) if row[6] is not None and str(row[6]).strip().isdigit() else 0
            acc_h    = str(row[7]).strip().upper() != "NON" if row[7] else True
            acc_f    = str(row[8]).strip().upper() != "NON" if row[8] else True
            ordre    = int(row[9]) if row[9] is not None else 99
            if not svc_nom:
                continue
            rows_processed += 1
            existing = None
            if uf_code:
                existing = db.query(CapaciteReferentiel).filter(
                    CapaciteReferentiel.uf_code == uf_code
                ).first()
            if not existing and svc_nom:
                existing = db.query(CapaciteReferentiel).filter(
                    CapaciteReferentiel.service_nom == svc_nom
                ).first()
            if existing:
                existing.service_nom     = svc_nom
                existing.uf_code         = uf_code or existing.uf_code
                existing.pole            = pole or existing.pole
                existing.site            = site or existing.site
                existing.capacite_totale = lits
                existing.accept_homme    = acc_h
                existing.accept_femme    = acc_f
                existing.accept_indiffer = True
                existing.ordre_affichage = ordre
                rows_updated += 1
            else:
                new_ref = CapaciteReferentiel(
                    service_nom=svc_nom, uf_code=uf_code, pole=pole,
                    site=site, capacite_totale=lits,
                    accept_homme=acc_h, accept_femme=acc_f,
                    accept_indiffer=True, ordre_affichage=ordre, actif=True,
                )
                db.add(new_ref)
                rows_created += 1
        db.commit()
        return {
            "ok": True,
            "processed": rows_processed,
            "created": rows_created,
            "updated": rows_updated,
            "errors": errors,
            "message": f"{rows_created} service(s) créé(s), {rows_updated} mis à jour.",
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(400, f"Erreur import Excel : {e}")


@router.get("/template.xlsx")
def download_template():
    """h149 — Télécharge le template Excel vierge pour configurer les UF
    et faciliter la déclaration capacitaire.
    Cherche uf_modele.xlsx dans uploads/ puis à la racine du projet."""
    for candidate in [
        pathlib.Path(__file__).parent.parent.parent / "uploads" / "uf_modele.xlsx",
        pathlib.Path(__file__).parent.parent.parent / "uf_modele.xlsx",
    ]:
        if candidate.exists():
            return FileResponse(
                str(candidate),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename="SCRIBE_capacite_template.xlsx",
            )
    raise HTTPException(404, "Template non trouvé — déposez uf_modele.xlsx dans uploads/")

