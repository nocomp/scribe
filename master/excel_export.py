"""
master/excel_export.py — Génère un SCRIBE_config_etablissement.xlsx
====================================================================
Symétrique de excel_import.py — produit un xlsx au même format à partir
d'une instance configurée dans le master.

Sécurité (option hybride par défaut) :
  - Mots de passe hashés (bcrypt $2b$...) — l'admin devra changer à l'import
  - Clés API IA vidées
  - Tokens fédération vidés
  - SAUF si include_secrets=True (case à cocher utilisateur, à confirmer)
"""
from __future__ import annotations

import io
import logging
import pathlib
from typing import Any
from datetime import datetime

logger = logging.getLogger("scribe.master.excel_export")


def _hash_password_for_export(plain_password: str) -> str:
    """Hash bcrypt du mot de passe pour export (utilisateur le change à l'import)."""
    if not plain_password:
        return ""
    try:
        import bcrypt
        h = bcrypt.hashpw(plain_password.encode("utf-8"), bcrypt.gensalt())
        return h.decode("utf-8")
    except ImportError:
        # Fallback : sha256 si bcrypt absent (rare)
        import hashlib
        return "sha256:" + hashlib.sha256(plain_password.encode("utf-8")).hexdigest()


def export_instance_to_xlsx(
    config: dict[str, Any],
    db_data: dict[str, list] | None = None,
    *,
    include_secrets: bool = False,
) -> bytes:
    """Génère le xlsx en bytes pour une instance.

    Args:
        config: dict avec sigle, nom, adresse, latitude, longitude, admin_login,
                admin_password, etc. (issu de InstanceConfig.asdict)
        db_data: dict optionnel avec 'directeurs', 'telephonie', 'uf', 'capa'
                 lus de la DB de l'instance (si l'instance a déjà tourné).
                 Peut être None pour une instance jamais lancée.
        include_secrets: si True, exporte mdp en clair et clés API.
                         Sinon (défaut) hash le mdp et vide les secrets.

    Returns:
        Bytes du fichier xlsx prêt à télécharger.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ValueError("openpyxl n'est pas installé. Lancez : pip install openpyxl")

    db_data = db_data or {}
    wb = openpyxl.Workbook()
    # Suppr la sheet par défaut
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Styles communs
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="000091")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="666666")
    note_font = Font(italic=True, color="666666", size=9)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # ─────────────────────────────────────────────────────────
    # Onglet 1 — ETABLISSEMENT
    # ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("ETABLISSEMENT")
    ws.merge_cells("A1:E1")
    ws["A1"] = "⚙ Configuration de l'établissement SCRIBE"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws["A2"] = "Remplissez les informations dans la colonne 'Valeur'. Les notes en colonne C sont indicatives."
    ws["A2"].font = note_font
    ws.merge_cells("A2:E2")
    # En-têtes
    headers = ["Paramètre", "Valeur", "Description", "", ""]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    # Mot de passe selon include_secrets
    pwd_value = config.get("admin_password", "")
    pwd_note = "Mot de passe (à changer après installation)"
    if not include_secrets:
        pwd_value = _hash_password_for_export(pwd_value)
        pwd_note = "Mot de passe (HASH bcrypt — changez-le après import)"

    cle_ia = config.get("cle_api_ia", "") if include_secrets else ""
    cle_ia_note = (
        "Clé API du fournisseur IA"
        if include_secrets
        else "Clé API du fournisseur IA (vidée à l'export — à ressaisir)"
    )
    fed_token = config.get("fed_token", "") if include_secrets else ""

    rows = [
        ("NOM_ETABLISSEMENT",   config.get("nom", ""),               "Nom complet de l'établissement"),
        ("SIGLE",               config.get("sigle", ""),             "Sigle (3-5 lettres)"),
        ("FINESS",              config.get("finess", "000000000"),   "N° FINESS géographique"),
        ("LANGUE",              config.get("langue", "fr"),          "Langue interface : fr en de es it nl pl pt"),
        ("LOGIN_ADMIN",         config.get("admin_login", "dircrise"), "Login du compte administrateur"),
        ("MOT_DE_PASSE",        pwd_value,                           pwd_note),
        ("NOM_AFFICHE_ADMIN",   config.get("nom_affiche_admin", "Directeur de Crise"), "Nom affiché dans l'interface"),
        ("FOURNISSEUR_IA",      config.get("fournisseur_ia", ""),    "albert | openai | anthropic | mistral | ollama"),
        ("CLE_API_IA",          cle_ia,                              cle_ia_note),
        ("MODELE_IA",           config.get("modele_ia", ""),         "Modèle IA"),
        ("URL_BASE_IA",         config.get("url_base_ia", ""),       "URL API IA (laisser vide pour OpenAI/Anthropic)"),
        ("FEDERATION_ACTIVE",   "true" if config.get("federation_active") else "false", "true | false"),
        ("COLLECTEUR_URL",      config.get("collecteur_url", ""),    "http://IP:9000/api/push"),
        ("COLLECTEUR_TOKEN",    fed_token,                           "Token fédération (vidé à l'export)"),
        ("SYNC_CRISE",          "true" if config.get("sync_crise") else "false",      "Synchroniser incidents/KPIs"),
        ("SYNC_SANITAIRE",      "true" if config.get("sync_sanitaire") else "false",  "Synchroniser capacitaire"),
    ]
    for i, (k, v, d) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).border = border
        ws.cell(row=i, column=2, value=v).border = border
        ws.cell(row=i, column=3, value=d).border = border

    # Section sites
    sites_start_row = 4 + len(rows) + 1
    ws.cell(row=sites_start_row, column=1, value="📍 Sites géographiques").font = Font(bold=True, size=12, color="000091")
    ws.cell(row=sites_start_row + 1, column=1, value="Nom du site").font = header_font
    ws.cell(row=sites_start_row + 1, column=1).fill = header_fill
    ws.cell(row=sites_start_row + 1, column=2, value="Adresse").font = header_font
    ws.cell(row=sites_start_row + 1, column=2).fill = header_fill
    ws.cell(row=sites_start_row + 1, column=3, value="Latitude").font = header_font
    ws.cell(row=sites_start_row + 1, column=3).fill = header_fill
    ws.cell(row=sites_start_row + 1, column=4, value="Longitude").font = header_font
    ws.cell(row=sites_start_row + 1, column=4).fill = header_fill
    ws.cell(row=sites_start_row + 1, column=5, value="Téléphone garde").font = header_font
    ws.cell(row=sites_start_row + 1, column=5).fill = header_fill

    # Site principal (issu du config)
    if config.get("nom") or config.get("adresse"):
        ws.cell(row=sites_start_row + 2, column=1, value=config.get("nom", "Site Principal"))
        ws.cell(row=sites_start_row + 2, column=2, value=config.get("adresse", ""))
        ws.cell(row=sites_start_row + 2, column=3, value=config.get("latitude") or "")
        ws.cell(row=sites_start_row + 2, column=4, value=config.get("longitude") or "")
        ws.cell(row=sites_start_row + 2, column=5, value="")
    # Sites additionnels (si présents en DB)
    for j, site in enumerate(db_data.get("sites_additionnels", []), start=sites_start_row + 3):
        ws.cell(row=j, column=1, value=site.get("nom", ""))
        ws.cell(row=j, column=2, value=site.get("adresse", ""))
        ws.cell(row=j, column=3, value=site.get("latitude") or "")
        ws.cell(row=j, column=4, value=site.get("longitude") or "")
        ws.cell(row=j, column=5, value=site.get("telephone_garde", ""))

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18

    # ─────────────────────────────────────────────────────────
    # Onglet 2 — DIRECTEURS
    # ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("DIRECTEURS")
    ws.merge_cells("A1:E1")
    ws["A1"] = "👤 Directeurs d'astreinte"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28
    ws["A2"] = "Les directeurs apparaissent dans la liste de garde. Le premier est le directeur de crise par défaut."
    ws["A2"].font = note_font
    ws.merge_cells("A2:E2")
    cols = ["Nom Prénom", "Fonction", "Abréviation", "Téléphone portable", "Note"]
    for col, h in enumerate(cols, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    for i, d in enumerate(db_data.get("directeurs", []), start=4):
        ws.cell(row=i, column=1, value=d.get("nom", ""))
        ws.cell(row=i, column=2, value=d.get("fonction", ""))
        ws.cell(row=i, column=3, value=d.get("abrev", ""))
        ws.cell(row=i, column=4, value=d.get("telephone", ""))
        ws.cell(row=i, column=5, value=d.get("note", ""))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20

    # ─────────────────────────────────────────────────────────
    # Onglet 3 — TELEPHONIE
    # ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("TELEPHONIE")
    ws.merge_cells("A1:F1")
    ws["A1"] = "📞 Annuaire de crise — Contacts"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28
    ws["A2"] = "NORMAL = numéros IP/IPBX en fonctionnement standard. SECOURS = numéros mobiles ou alternatifs."
    ws["A2"].font = note_font
    ws.merge_cells("A2:F2")
    ws["A4"] = "CONTACTS NOMINAUX (réseau / fonctionnement standard)"
    ws["A4"].font = Font(bold=True, color="000091")
    ws.merge_cells("A4:F4")
    cols = ["Service", "N° Interne / IP", "N° Direct", "N° Mobile", "Site", "Note"]
    for col, h in enumerate(cols, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    for i, t in enumerate(db_data.get("telephonie", []), start=6):
        ws.cell(row=i, column=1, value=t.get("service", ""))
        ws.cell(row=i, column=2, value=t.get("interne", ""))
        ws.cell(row=i, column=3, value=t.get("direct", ""))
        ws.cell(row=i, column=4, value=t.get("mobile", ""))
        ws.cell(row=i, column=5, value=t.get("site", ""))
        ws.cell(row=i, column=6, value=t.get("note", ""))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 25

    # ─────────────────────────────────────────────────────────
    # Onglet 4 — UF_INCIDENTS
    # ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("UF_INCIDENTS")
    ws.merge_cells("A1:F1")
    ws["A1"] = "🏥 Unités Fonctionnelles — Liste des UF"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28
    ws["A2"] = "Ces UF apparaissent dans le formulaire de création d'incident."
    ws["A2"].font = note_font
    ws.merge_cells("A2:F2")
    cols = ["Code UF", "Libellé UF", "Pôle (SOINS)", "Site", "Actif (O/N)", "Note"]
    for col, h in enumerate(cols, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    for i, u in enumerate(db_data.get("uf", []), start=4):
        ws.cell(row=i, column=1, value=u.get("code", ""))
        ws.cell(row=i, column=2, value=u.get("libelle", ""))
        ws.cell(row=i, column=3, value=u.get("pole", ""))
        ws.cell(row=i, column=4, value=u.get("site", ""))
        ws.cell(row=i, column=5, value="O" if u.get("actif", True) else "N")
        ws.cell(row=i, column=6, value=u.get("note", ""))

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30

    # ─────────────────────────────────────────────────────────
    # Onglet 5 — SERVICES_CAPACITE
    # ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("SERVICES_CAPACITE")
    ws.merge_cells("A1:L1")
    ws["A1"] = "🛏️ Services — Gestion capacitaire"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28
    ws["A2"] = "Ces services apparaissent dans l'onglet Capacité."
    ws["A2"].font = note_font
    ws.merge_cells("A2:L2")
    cols = ["Service", "Code UF", "Pôle", "Site (groupe)", "Capa. totale",
            "Tension seuil 1", "Tension seuil 2",
            "Accepte H", "Accepte F", "Accepte I",
            "Tel. cadre", "Ordre"]
    for col, h in enumerate(cols, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    for i, s in enumerate(db_data.get("capa", []), start=4):
        ws.cell(row=i, column=1, value=s.get("service", ""))
        ws.cell(row=i, column=2, value=s.get("code_uf", ""))
        ws.cell(row=i, column=3, value=s.get("pole", ""))
        ws.cell(row=i, column=4, value=s.get("site", ""))
        ws.cell(row=i, column=5, value=s.get("capacite", 0))
        ws.cell(row=i, column=6, value=s.get("seuil_t1", 0))
        ws.cell(row=i, column=7, value=s.get("seuil_t2", 0))
        ws.cell(row=i, column=8, value="O" if s.get("accepte_h", True) else "N")
        ws.cell(row=i, column=9, value="O" if s.get("accepte_f", True) else "N")
        ws.cell(row=i, column=10, value="O" if s.get("accepte_i", False) else "N")
        ws.cell(row=i, column=11, value=s.get("tel_cadre", ""))
        ws.cell(row=i, column=12, value=s.get("ordre", 0))

    for col_letter, width in zip(
        ["A","B","C","D","E","F","G","H","I","J","K","L"],
        [22, 10, 16, 14, 12, 14, 14, 10, 10, 10, 18, 8],
    ):
        ws.column_dimensions[col_letter].width = width

    # Métadonnées de l'export (commentaire en haut de chaque onglet ?)
    # Pour l'instant on les met dans une note dans ETABLISSEMENT
    et = wb["ETABLISSEMENT"]
    last_row = et.max_row + 2
    et.cell(row=last_row, column=1, value="── EXPORT METADATA ──").font = Font(italic=True, color="999999", size=9)
    et.cell(row=last_row + 1, column=1, value="Date export").font = note_font
    et.cell(row=last_row + 1, column=2, value=datetime.now().isoformat(timespec="seconds"))
    et.cell(row=last_row + 2, column=1, value="Secrets inclus").font = note_font
    et.cell(row=last_row + 2, column=2, value="oui" if include_secrets else "non (mdp hashé, clé API vidée)")
    et.cell(row=last_row + 3, column=1, value="Source").font = note_font
    et.cell(row=last_row + 3, column=2, value="SCRIBE master export")

    # Sérialiser
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def export_ght_to_zip(
    instances_configs: list[dict[str, Any]],
    instances_db_data: dict[str, dict] | None = None,
    *,
    include_secrets: bool = False,
) -> bytes:
    """Génère un zip contenant 1 xlsx par instance (export GHT complet).

    Args:
        instances_configs: liste de dict config (un par instance)
        instances_db_data: dict {sigle: {directeurs, telephonie, uf, capa}}
        include_secrets: idem export_instance_to_xlsx

    Returns:
        Bytes du zip prêt à télécharger.
    """
    import zipfile

    instances_db_data = instances_db_data or {}
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        # Manifest
        manifest_lines = [
            "# SCRIBE — Export GHT",
            f"# Date : {datetime.now().isoformat(timespec='seconds')}",
            f"# Établissements : {len(instances_configs)}",
            f"# Secrets inclus : {'oui' if include_secrets else 'non (mdp hashé)'}",
            "",
            "# Liste des fichiers :",
        ]
        for cfg in instances_configs:
            sigle = cfg.get("sigle", f"port_{cfg.get('port', 'unknown')}")
            xlsx_name = f"{sigle}.xlsx"
            db_d = instances_db_data.get(sigle, {})
            xlsx_bytes = export_instance_to_xlsx(
                cfg, db_d, include_secrets=include_secrets,
            )
            zf.writestr(xlsx_name, xlsx_bytes)
            manifest_lines.append(f"#   {xlsx_name} (port :{cfg.get('port')})")
        zf.writestr("MANIFEST.txt", "\n".join(manifest_lines))

    out.seek(0)
    return out.read()
