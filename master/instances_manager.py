"""
master/instances_manager.py — Pilotage des instances SCRIBE depuis l'admin
==========================================================================
Permet à l'admin de la supervision de lancer/arrêter des instances SCRIBE
en subprocess depuis l'interface web, sans recourir à des scripts shell.

Inspiré du module `flotte` du projet CIAE, simplifié au strict nécessaire
pour SCRIBE :
  - Une instance = 1 port + 1 sigle + 1 admin login/pwd + 1 adresse géoloc
  - Stockage de l'état dans un fichier JSON (master_instances.json)
  - Subprocess détaché via start_new_session=True (survit au master si crash)
  - Auto-enrôlement dans le collecteur fédération (token généré au lancement)

Architecture :
  - InstanceConfig : dataclass de configuration d'une instance
  - InstanceState  : état runtime (PID, started_at, etc.)
  - InstanceManager: orchestrateur (start/stop/status, persistance)

Le master ne fait PAS d'édition multi-tenant : chaque instance est
indépendante, avec sa propre DB SQLite, son propre log, sa propre config XML.
"""
from __future__ import annotations

import json
import logging
import os
import pathlib
import secrets
import signal
import string
import subprocess
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from typing import Any

logger = logging.getLogger("scribe.master")

# Répertoire racine du projet (parent de master/)
PROJECT_ROOT = pathlib.Path(__file__).resolve().parent.parent
# h127 — chemins surchargeables par variable d'environnement, pour que les
# donnees d'instances et l'etat du master SURVIVENT aux changements de build
# (sinon, deployer un nouveau dossier de version repart sur des donnees vides).
DATA_DIR = pathlib.Path(os.environ.get("SCRIBE_DATA_DIR") or (PROJECT_ROOT / "data" / "instances"))
STATE_FILE = pathlib.Path(os.environ.get("SCRIBE_STATE_FILE") or (PROJECT_ROOT / "master" / "master_instances.json"))
PROFIL_BASE_XLSX = PROJECT_ROOT / "master" / "profil_base.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# Modèles de données
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class InstanceConfig:
    """Configuration d'une instance SCRIBE pilotée par le master."""
    port: int
    sigle: str = ""
    nom: str = ""
    admin_login: str = "dircrise"
    admin_password: str = ""
    # v3.4 (h38g) — Nom affiché de l'admin (display_name dans la DB).
    # Si vide, on garde un fallback pertinent au lieu du legacy
    # "Directeur de Crise" hardcodé qui rendait illisibles les
    # messages de chat quand le login n'était pas "dircrise" (ex: "rssi"
    # qui apparaissait comme "Directeur de Crise" dans les fils de chat).
    admin_display_name: str = ""
    adresse: str = ""
    latitude: float | None = None
    longitude: float | None = None
    timezone: str = ""  # v2.4.6 : "" = automatique (navigateur), sinon IANA
    # Si False, l'instance ne pousse PAS au collecteur de supervision.
    # Cas d'usage : établissement isolé, exercice solo, données sensibles HDS,
    # convention GHT non encore signée, etc.
    # L'instance reste pleinement fonctionnelle pour son propre usage interne.
    synchroniser: bool = True
    # v3.4 (h38h) — Plugins désactivés à la création (étape 5 du wizard).
    # Liste des plugin_id à mettre à False dans la table plugin_states
    # lors du premier bootstrap. Les autres restent activés par défaut.
    # Ce champ n'est consulté qu'au _bootstrap_db ; modifier ensuite via
    # /admin/plugins dans l'instance.
    plugins_disabled: list = field(default_factory=list)
    # v3.4 (h38k) — Langue par défaut de l'instance, choisie au wizard.
    # Stocké en config.js (consommé par scribe.js loadI18n() au boot).
    # Valeurs : fr, en, it, de, es, ou code ISO 2-letters d'une des 24 langues UE.
    langue: str = "fr"

    def __post_init__(self):
        # v3.4 (h38) — Strip systématique des champs textes pour éviter les
        # espaces traînants qui cassent les headers HTTP de la fédération
        # ("Illegal header value b'ch2 '"). Cf. bug supervision : un sigle
        # avec un espace en fin empêchait le push vers le collecteur, donc
        # l'instance n'apparaissait pas sur la carte territoriale.
        if isinstance(self.sigle, str):
            self.sigle = self.sigle.strip()
        if isinstance(self.nom, str):
            self.nom = self.nom.strip()
        if isinstance(self.admin_login, str):
            self.admin_login = self.admin_login.strip()
        if isinstance(self.admin_display_name, str):
            self.admin_display_name = self.admin_display_name.strip()
        if isinstance(self.adresse, str):
            self.adresse = self.adresse.strip()
        if isinstance(self.timezone, str):
            self.timezone = self.timezone.strip()
        if not self.sigle:
            self.sigle = f"Site_{self.port}"
        if not self.nom:
            self.nom = self.sigle
        if not self.admin_password:
            self.admin_password = generate_password()
        # v3.4 (h38g) — Fallback intelligent du display_name :
        # 1. Si le wizard a fourni un nom affiché, on l'utilise tel quel
        # 2. Sinon, on capitalize le login (rssi → "Rssi")
        # On ne met PLUS jamais "Directeur de Crise" en dur.
        if not self.admin_display_name:
            self.admin_display_name = self.admin_login.upper() if self.admin_login else "Admin"


@dataclass
class InstanceState:
    """État runtime d'une instance."""
    config: InstanceConfig
    statut: str = "arrete"  # arrete | actif | erreur
    pid: int | None = None
    started_at: str | None = None  # ISO format
    stopped_at: str | None = None
    db_path: str | None = None
    log_path: str | None = None
    fed_token: str | None = None  # token d'auto-enrôlement vers collecteur

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        d["config"] = asdict(self.config)
        return d

    @classmethod
    def from_dict(cls, d: dict[str, Any]) -> "InstanceState":
        cfg = InstanceConfig(**d.pop("config"))
        return cls(config=cfg, **d)


# ─────────────────────────────────────────────────────────────────────────────
# Utilitaires
# ─────────────────────────────────────────────────────────────────────────────

def generate_password(length: int = 12) -> str:
    """Génère un mot de passe lisible (sans caractères ambigus)."""
    # Exclut les caractères ambigus : I, l, O, 0, 1
    alphabet = (
        string.ascii_letters.replace("I", "").replace("l", "").replace("O", "")
        + "23456789"
        + "!@#$%&*"
    )
    return "".join(secrets.choice(alphabet) for _ in range(length))


def generate_token() -> str:
    """Token de fédération pour l'auto-enrôlement dans le collecteur."""
    return secrets.token_urlsafe(32)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _xml_esc(s: str) -> str:
    """Escape XML special chars."""
    if s is None:
        return ""
    return (str(s)
            .replace("&",  "&amp;")
            .replace("<",  "&lt;")
            .replace(">",  "&gt;")
            .replace('"',  "&quot;")
            .replace("'",  "&apos;"))


def _safe_path_segment(s: str, fallback: str = "instance") -> str:
    """Nettoie un sigle pour pouvoir l'utiliser comme nom de répertoire.

    Remplace les caractères problématiques par '_', strippe les espaces
    aux extrémités. Sur Windows, les espaces en fin de chemin et certains
    caractères posent problème à SQLite et au filesystem.

    Le sigle "humain" (cfg.sigle) reste intact pour affichage / JWT / config.js.
    Seul le chemin de répertoire est sanitizé.
    """
    if not s:
        return fallback
    # Whitelist : alphanum + tiret + underscore
    cleaned = "".join(c if (c.isalnum() or c in "-_") else "_"
                      for c in str(s).strip())
    # Compacter les underscores multiples
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    cleaned = cleaned.strip("_-") or fallback
    # Limite raisonnable
    return cleaned[:64]


# ─────────────────────────────────────────────────────────────────────────────
# Manager principal
# ─────────────────────────────────────────────────────────────────────────────

class InstanceManager:
    """Orchestrateur des instances pilotées."""

    def __init__(self):
        self.instances: dict[int, InstanceState] = {}  # clé = port
        self._load_state()
        # Ports pré-configurés par défaut (8000 à 8009)
        self._ensure_defaults()

    # ── État persisté ─────────────────────────────────────────────────────

    def _load_state(self) -> None:
        """Charge l'état depuis master_instances.json (si présent)."""
        if not STATE_FILE.exists():
            return
        try:
            with open(STATE_FILE, encoding="utf-8") as f:
                data = json.load(f)
            for d in data.get("instances", []):
                state = InstanceState.from_dict(d)
                # Au chargement, on remet le statut à "arrete" si le PID
                # n'existe plus. Le master ne re-lance pas auto les instances.
                if state.pid and not _pid_alive(state.pid):
                    state.statut = "arrete"
                    state.pid = None
                self.instances[state.config.port] = state
            logger.info(f"État rechargé : {len(self.instances)} instances")
            # v3.4 (h38) — Re-sauver l'état après chargement : __post_init__ a
            # appliqué un strip défensif sur sigle/nom/adresse, on persiste
            # cette correction pour qu'elle survive aux redémarrages.
            self._save_state()
        except Exception as e:
            logger.warning(f"Impossible de charger l'état : {e}")

    def _save_state(self) -> None:
        """Persiste l'état dans master_instances.json."""
        STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
        try:
            data = {
                "instances": [s.to_dict() for s in self.instances.values()],
                "saved_at": now_iso(),
            }
            with open(STATE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Impossible de sauvegarder l'état : {e}")

    def _ensure_defaults(self) -> None:
        """Crée les 10 instances par défaut (ports 8000-8009) si absentes."""
        for port in range(8000, 8010):
            if port not in self.instances:
                cfg = InstanceConfig(port=port)
                self.instances[port] = InstanceState(config=cfg)
        self._save_state()

    # ── API publique ──────────────────────────────────────────────────────

    def list_instances(self) -> list[dict[str, Any]]:
        """Retourne toutes les instances avec leur état actualisé."""
        out = []
        for state in sorted(self.instances.values(), key=lambda s: s.config.port):
            try:
                # Vérifier le PID en temps réel — si statut "actif" mais PID
                # mort (ou absent), on remet l'instance en "arrete" pour
                # éviter d'être bloqué (état zombie)
                if state.statut == "actif":
                    if not state.pid or not _pid_alive(state.pid):
                        logger.info(f"État zombie détecté pour {state.config.sigle} → reset arrete")
                        state.statut = "arrete"
                        state.pid = None
                        state.stopped_at = now_iso()
                d = state.to_dict()
                # Ne jamais renvoyer le mdp dans la liste — il faut une route
                # dédiée /credentials qui exige une auth admin
                d["config"]["admin_password"] = "***"
                out.append(d)
            except Exception as e:
                # Log mais ne casse pas la liste pour autant
                logger.error(f"Erreur sur instance port={getattr(state.config, 'port', '?')} : {e}")
                continue
        # Persister le reset éventuel des états zombies
        try: self._save_state()
        except Exception: pass
        return out

    def get_instance(self, port: int) -> InstanceState | None:
        return self.instances.get(port)

    def update_config(self, port: int, **fields) -> InstanceState:
        """Met à jour la config d'une instance (uniquement si arrêtée)."""
        state = self.instances.get(port)
        if not state:
            raise ValueError(f"Instance port {port} inconnue")
        if state.statut == "actif":
            raise ValueError(
                "Impossible de modifier une instance active. Arrêtez-la d'abord."
            )

        allowed = {
            "sigle", "nom", "admin_login", "admin_password",
            "admin_display_name",  # v3.4 (h38g)
            "plugins_disabled",    # v3.4 (h38h) — pré-désactivation plugins par wizard
            "langue",              # v3.4 (h38k) — langue par défaut posée par le wizard
            "adresse", "latitude", "longitude", "synchroniser",
            "timezone",
        }
        for k, v in fields.items():
            if k in allowed and v is not None:
                # Pour les champs string critiques, ignorer les chaînes vides
                # (sinon on écraserait des valeurs valides par "")
                if k in {"sigle", "nom", "admin_login"} and isinstance(v, str) and not v.strip():
                    continue
                setattr(state.config, k, v)
        # Reset explicite du mot de passe admin demande (regenerate / edition panneau) :
        # marque l'instance pour que _create_admin reapplique le mdp au prochain start.
        if fields.get("admin_password"):
            try:
                state.pending_admin_reset = True
            except Exception:
                pass
        self._save_state()
        return state

    def add_custom(self, port: int, **fields) -> InstanceState:
        """Ajoute une instance custom (port hors 8000-8009)."""
        if port in self.instances:
            raise ValueError(f"Port {port} déjà déclaré")
        if port < 1024 or port > 65535:
            raise ValueError("Port invalide (doit être entre 1024 et 65535)")
        cfg = InstanceConfig(port=port, **{
            k: v for k, v in fields.items()
            if k in {"sigle", "nom", "admin_login", "admin_password",
                     "adresse", "latitude", "longitude", "synchroniser"} and v is not None
        })
        state = InstanceState(config=cfg)
        self.instances[port] = state
        self._save_state()
        return state

    def remove_instance(self, port: int) -> None:
        """Supprime une instance (uniquement si arrêtée)."""
        state = self.instances.get(port)
        if not state:
            return
        if state.statut == "actif":
            raise ValueError("Arrêtez l'instance avant de la supprimer.")
        # Cleanup du dossier de l'instance
        if state.db_path:
            inst_dir = pathlib.Path(state.db_path).parent
            if inst_dir.exists() and "instances" in str(inst_dir):
                import shutil
                shutil.rmtree(inst_dir, ignore_errors=True)
        del self.instances[port]
        self._save_state()

    def get_credentials(self, port: int) -> dict[str, str]:
        """Retourne URL + login + mdp pour copier dans le presse-papier."""
        state = self.instances.get(port)
        if not state:
            raise ValueError(f"Instance port {port} inconnue")
        return {
            "url": f"http://localhost:{state.config.port}",
            "login": state.config.admin_login,
            "password": state.config.admin_password,
        }

    # ── Lancement / arrêt ─────────────────────────────────────────────────

    def start(self, port: int, collecteur_url: str = "http://localhost:9000") -> InstanceState:
        """Lance une instance en subprocess détaché."""
        state = self.instances.get(port)
        if not state:
            raise ValueError(f"Instance port {port} inconnue")

        # Reset auto si état zombie (statut actif mais PID mort) — évite
        # d'être bloqué après un crash silencieux
        if state.statut == "actif":
            if not state.pid or not _pid_alive(state.pid):
                logger.info(f"Reset zombie {state.config.sigle} avant relancement")
                state.statut = "arrete"
                state.pid = None

        if state.statut == "actif" and state.pid and _pid_alive(state.pid):
            raise ValueError(f"Instance déjà active (PID {state.pid})")

        # v3.0.0 — Libération prudente du port si occupé par un process SCRIBE
        # orphelin (instance mal arrêtée). Process tiers : non touchés.
        try:
            from master.port_cleanup import free_port_if_scribe
            r = free_port_if_scribe(port)
            if r["status"] == "foreign":
                raise ValueError(
                    f"Port {port} occupé par un process tiers "
                    f"(PID {r.get('pid', '?')}). "
                    f"SCRIBE ne touche pas aux process qu'il n'a pas lancés. "
                    f"Arrêtez-le manuellement avant de relancer l'instance."
                )
            if r["status"] == "failed_kill":
                raise ValueError(
                    f"Port {port} occupé par un process SCRIBE (PID {r.get('pid', '?')}) "
                    f"qui n'a pas pu être terminé. Tuez-le manuellement."
                )
            if r["status"] == "freed":
                logger.info(f"Port {port} libéré (instance SCRIBE orpheline terminée)")
        except ValueError:
            raise
        except Exception as _e:
            logger.warning(f"port_cleanup KO pour {port} : {_e}")
            # Fallback : check minimal sans cleanup
            if _port_in_use(port):
                raise ValueError(
                    f"Port {port} déjà utilisé par un autre processus. "
                    "Arrêtez-le ou choisissez un autre port."
                )

        cfg = state.config

        # 1. Préparer le répertoire de l'instance
        # Le sigle "humain" (cfg.sigle) peut contenir des espaces, accents,
        # chiffres, etc. Mais on ne peut pas l'utiliser tel quel comme nom
        # de répertoire sur Windows (un espace en fin de chemin casse SQLite,
        # certains caractères sont interdits).
        # → on utilise _safe_path_segment() qui produit un nom alphanum sûr
        #   pour le chemin, tout en gardant cfg.sigle intact pour l'affichage,
        #   le JWT, le config.js, le payload de fédération, etc.
        path_segment = _safe_path_segment(cfg.sigle, fallback=f"instance_{cfg.port}")
        inst_dir = DATA_DIR / path_segment
        inst_dir.mkdir(parents=True, exist_ok=True)
        state.db_path = str(inst_dir / "scribe.db")
        state.log_path = str(inst_dir / "scribe.log")

        # 2. Générer le token de fédération si pas déjà présent
        if not state.fed_token:
            state.fed_token = generate_token()

        # 3. Bootstrap de la DB depuis le profil de base (UF + capacité)
        try:
            self._bootstrap_db(state)
        except Exception as e:
            logger.error(f"Bootstrap échoué pour {cfg.sigle} : {e}")
            state.statut = "erreur"
            self._save_state()
            raise ValueError(f"Bootstrap DB échoué : {e}") from e

        # 3.5 Générer le config.js de l'instance (config.xml minimal aussi)
        # SCRIBE lit la fédération depuis config.js, PAS depuis l'env. Variables
        # SCRIBE_FED_URL/SCRIBE_FED_TOKEN sont ignorées par app/api/federation.py.
        # Ce qui compte : SCRIBE_CONFIG_JS pointant vers un config.js avec
        # federation.enabled=true / collecteur_url / token.
        instance_dir = pathlib.Path(state.db_path).parent
        config_js_path = instance_dir / "config.js"
        config_xml_path = instance_dir / "config.xml"
        try:
            self._generate_instance_config(
                state, collecteur_url, config_js_path, config_xml_path
            )
            logger.info(f"  config.js + config.xml générés pour {cfg.sigle}")
        except Exception as e:
            logger.warning(f"  Génération config.js échouée : {e}")

        # 3.6 Auto-enrôlement AVANT le subprocess Popen.
        # IMPORTANT : il faut que le token soit enregistré au collecteur AVANT
        # que l'instance fille ne fasse son 1er push (qui arrive 3-5s après
        # son démarrage). Sinon le push tombe dans le pending et la supervision
        # reste vide jusqu'à ce que l'admin clique "Accepter tout".
        # Cas SOLO : on saute, l'instance ne pousse rien.
        if cfg.synchroniser:
            try:
                self._auto_enrol(state, collecteur_url)
            except Exception as e:
                logger.warning(f"Auto-enrôlement échoué pour {cfg.sigle} : {e}")
                # Non bloquant : l'instance tourne, le push se fera plus tard.
                # L'admin peut accepter le pending si nécessaire.
        else:
            logger.info(
                f"  Instance {cfg.sigle} en mode SOLO (non synchronisée) — "
                f"aucun push vers la supervision"
            )

        # 4. Lancer le subprocess
        env = os.environ.copy()
        env.update({
            "SCRIBE_PORT": str(cfg.port),
            "SCRIBE_SIGLE": cfg.sigle,
            "SCRIBE_NOM": cfg.nom,
            "DATABASE_URL": f"sqlite:///{state.db_path}",
            # Chemin vers le config.js de cette instance (lu par federation.py)
            "SCRIBE_CONFIG_JS":   str(config_js_path),
            "SCRIBE_CONFIG_FILE": str(config_xml_path),
            "SCRIBE_ADMIN_LOGIN": cfg.admin_login,
            "SCRIBE_ADMIN_PWD":   cfg.admin_password,
            "SCRIBE_LATITUDE":    str(cfg.latitude or ""),
            "SCRIBE_LONGITUDE":   str(cfg.longitude or ""),
            "SCRIBE_ADRESSE":     cfg.adresse or "",
            # Encoding UTF-8 forcé pour Windows (évite cp1252 → cassage emojis)
            "PYTHONIOENCODING": "utf-8",
            "PYTHONUTF8": "1",
        })

        log_file = None
        try:
            log_file = open(state.log_path, "ab", buffering=0)
            popen_kwargs = {
                "cwd": str(PROJECT_ROOT),
                "env": env,
                "stdout": log_file,
                "stderr": log_file,
                "stdin":  subprocess.DEVNULL,
            }
            if os.name == "posix":
                # Linux/macOS : start_new_session détache du parent
                popen_kwargs["start_new_session"] = True
            else:
                # Windows : CREATE_NO_WINDOW seul.
                # NE PAS utiliser DETACHED_PROCESS : il est incompatible avec
                # passer stdout/stderr à un fichier (le subprocess crashe
                # silencieusement). CREATE_NEW_PROCESS_GROUP est aussi inutile
                # ici (on ne tue pas par CTRL_BREAK_EVENT).
                CREATE_NO_WINDOW = 0x08000000
                popen_kwargs["creationflags"] = CREATE_NO_WINDOW
            proc = subprocess.Popen(
                [sys.executable, "main.py"],
                **popen_kwargs,
            )
        except Exception as e:
            if log_file:
                try: log_file.close()
                except Exception: pass
            logger.error(f"Popen échoué pour {cfg.sigle} : {e}")
            state.statut = "erreur"
            self._save_state()
            raise ValueError(f"Lancement subprocess échoué : {e}") from e

        state.pid = proc.pid
        state.started_at = now_iso()
        state.stopped_at = None
        state.statut = "actif"

        # ── Vérification post-launch : le subprocess est-il vivant après 2s ?
        # Si non, on remonte l'erreur exacte du log de l'instance plutôt que
        # de laisser le softRefresh marquer "zombie" sans explication.
        import time
        time.sleep(2.0)
        rc = proc.poll()
        if rc is not None:
            # Le process est mort dans les 2s
            state.statut = "erreur"
            state.pid = None
            self._save_state()
            # Lire les dernières lignes du log pour aider au diagnostic
            log_tail = ""
            try:
                with open(state.log_path, "rb") as f:
                    content = f.read()
                log_tail = content.decode("utf-8", errors="replace")
                # Garder les 30 dernières lignes
                lines = log_tail.splitlines()[-30:]
                log_tail = "\n".join(lines)
            except Exception:
                pass
            logger.error(
                f"Subprocess {cfg.sigle} mort dans les 2s (returncode={rc}). "
                f"Voir {state.log_path}\n--- DERNIÈRES LIGNES DU LOG ---\n{log_tail}"
            )
            raise ValueError(
                f"L'instance a planté immédiatement (returncode={rc}). "
                f"Voir data/instances/{cfg.sigle}/scribe.log\n"
                f"Dernières lignes :\n{log_tail[-800:]}"
            )

        self._save_state()
        logger.info(f"Instance {cfg.sigle} lancée (PID {proc.pid}, port {port})")
        return state

    def stop(self, port: int) -> InstanceState:
        """Arrête une instance proprement (SIGTERM)."""
        state = self.instances.get(port)
        if not state:
            raise ValueError(f"Instance port {port} inconnue")
        if not state.pid:
            state.statut = "arrete"
            self._save_state()
            return state

        try:
            os.kill(state.pid, signal.SIGTERM)
            logger.info(f"SIGTERM envoyé à {state.config.sigle} (PID {state.pid})")
        except ProcessLookupError:
            logger.info(f"{state.config.sigle} : processus déjà disparu")
        except PermissionError:
            logger.error(f"{state.config.sigle} : permission refusée pour PID {state.pid}")
            raise

        state.pid = None
        state.stopped_at = now_iso()
        state.statut = "arrete"
        self._save_state()
        return state

    def stop_all(self) -> int:
        """Arrête toutes les instances actives. Retourne le nombre arrêté."""
        count = 0
        for port, state in list(self.instances.items()):
            if state.statut == "actif" and state.pid:
                try:
                    self.stop(port)
                    count += 1
                except Exception as e:
                    logger.warning(f"Erreur arrêt port {port} : {e}")
        return count

    # ── Internes ──────────────────────────────────────────────────────────

    def _generate_instance_config(
        self,
        state: "InstanceState",
        collecteur_url: str,
        config_js_path: pathlib.Path,
        config_xml_path: pathlib.Path,
    ) -> None:
        """Génère config.js (lu par app/api/federation.py) et config.xml minimal.

        SCRIBE charge la config de fédération depuis config.js puis surcharge
        avec config.xml si présent. Sans config.js / xml avec federation.enabled=true,
        l'instance ne pousse rien au collecteur (rien ne remonte en supervision).
        """
        cfg = state.config
        push_url = collecteur_url.rstrip("/") + "/api/push"

        # Mode "non synchronisée" : federation désactivée dans le config.js,
        # l'instance ne pousse rien au collecteur. L'admin peut quand même
        # ré-activer la fédération depuis l'UI Admin de l'instance si besoin
        # (config.xml prend la priorité sur config.js).
        fed_enabled_str = "true" if cfg.synchroniser else "false"

        # ── Peupler directeurs + annuaires depuis le profil xlsx s'il existe
        # Sinon : créer un directeur par défaut à partir du compte admin pour
        # que la dropdown "Directeur de crise" ne soit pas vide au premier login
        directeurs_list = []
        annuaire_normal_list = []
        annuaire_secours_list = []
        if PROFIL_BASE_XLSX.exists():
            try:
                d, an, as_ = self._read_directeurs_annuaire_from_xlsx(PROFIL_BASE_XLSX)
                directeurs_list = d
                annuaire_normal_list = an
                annuaire_secours_list = as_
            except Exception as e:
                logger.warning(f"  Lecture directeurs/annuaire xlsx KO (non bloquant) : {e}")

        # Fallback : si aucun directeur dans le xlsx, créer une entrée par défaut
        # basée sur le compte admin (sinon la dropdown VEILLE reste vide)
        if not directeurs_list:
            directeurs_list = [{
                "nom":          "Directeur de Crise",
                "abreviation":  "DDC",
                "telephone":    "",
                "fonction":     "Directeur de Crise",
            }]

        # ── 1. config.js ────────────────────────────────────────────────────
        # v3.4 (h38k) — Détecter la langue préférée pour cette instance.
        # Le wizard envoie payload.langue dans WizardInstanceCreate ;
        # on stocke cette préférence dans InstanceConfig.langue (mais le
        # champ n'existe pas encore — on le lit via getattr pour fallback "fr").
        instance_langue = getattr(cfg, "langue", "") or "fr"
        scribe_config = {
            "etablissement": {
                "nom":      cfg.nom,
                "sigle":    cfg.sigle,
                "timezone": cfg.timezone or "",  # v2.4.6 : IANA ou "" (auto)
            },
            # v3.4 (h38k) — Langue de l'instance posée par le wizard.
            # Le frontend scribe.js charge cette langue par défaut au boot
            # (cf. loadI18n() : SCRIBE_CONFIG.langue est lu si pas d'override admin).
            "langue":           instance_langue,
            "login_tagline": "",
            "admin": {
                "login":    cfg.admin_login,
                "password": cfg.admin_password,
            },
            "directeurs":       directeurs_list,
            "annuaire_normal":  annuaire_normal_list,
            "annuaire_secours": annuaire_secours_list,
            "ia": {
                "fournisseur": "",
                "cle":         "",
            },
            "federation": {
                "enabled":              fed_enabled_str,
                "collecteur_url":       push_url,
                "token":                state.fed_token,
                # 30s pour que la supervision se peuple rapidement après le
                # démarrage. Si tu veux moins de charge réseau en prod, monte
                # à 60-120s. Mais 30s reste raisonnable pour 7 instances locales.
                "intervalle_secondes":  30,
                "share_details":        "true",
                "share_min_urgency":    1,
                "sync_crise":           "true",
                "sync_sanitaire":       "true",
                "share_capacite_details": "true",
            },
            "exercice_mode":  False,
            "exercice_sigle": "",
        }

        config_js_path.parent.mkdir(parents=True, exist_ok=True)
        with open(config_js_path, "w", encoding="utf-8") as f:
            f.write("// Généré par master/instances_manager.py\n")
            f.write(f"// Instance {cfg.sigle} sur port {cfg.port}\n")
            f.write("const SCRIBE_CONFIG = ")
            f.write(json.dumps(scribe_config, ensure_ascii=False, indent=2))
            f.write(";\n")

        # ── 2. config.xml minimal ──────────────────────────────────────────
        # SCRIBE peut écrire dans ce fichier via Admin (modifier_config).
        # On le crée avec les mêmes infos que le config.js.
        # Format simple lu par federation.py et autres modules.
        xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<scribe_config>
  <etablissement>
    <nom>{_xml_esc(cfg.nom)}</nom>
    <sigle>{_xml_esc(cfg.sigle)}</sigle>
    <adresse>{_xml_esc(cfg.adresse or "")}</adresse>
    <latitude>{cfg.latitude or ""}</latitude>
    <longitude>{cfg.longitude or ""}</longitude>
  </etablissement>
  <admin>
    <login>{_xml_esc(cfg.admin_login)}</login>
    <password>{_xml_esc(cfg.admin_password)}</password>
  </admin>
  <federation>
    <enabled>{fed_enabled_str}</enabled>
    <collecteur_url>{push_url}</collecteur_url>
    <token>{state.fed_token}</token>
    <intervalle_secondes>30</intervalle_secondes>
  </federation>
</scribe_config>
"""
        config_xml_path.parent.mkdir(parents=True, exist_ok=True)
        with open(config_xml_path, "w", encoding="utf-8") as f:
            f.write(xml_content)

    def _bootstrap_db(self, state: InstanceState, force_reset: bool = False) -> None:
        """Initialise la DB de l'instance avec le profil de base — 100% in-process.

        Inspiré de CIAE/SubprocessDeployer : aucun subprocess, tout en SQLAlchemy
        direct. Évite les problèmes d'encoding console Windows et la cascade de
        subprocess imbriqués (master → import_config_xlsx → setup.py).

        Stratégie v2.4.6 :
          - Si la DB existe déjà ET force_reset=False (défaut) :
            on PRÉSERVE la DB, on assure juste que Hospital + admin existent
            (pour ne pas perdre les données saisies par l'utilisateur)
          - Si force_reset=True (bouton "🔄 Réinitialiser DB" dans le master)
            ou si la DB n'existe pas encore : recréation complète depuis le
            profil xlsx

        Étapes :
          1. Si force_reset : supprime la DB
          2. Crée la DB SQLite vide avec toutes les tables SCRIBE (si absente)
          3. Lit le profil xlsx avec openpyxl
          4. Crée Hospital + UF + CapaciteReferentiel directement (skip si existent)
          5. Crée le compte admin (login/mdp de la config) (skip si existe)
        """
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker

        # 0. Si force_reset : on supprime la DB pour repartir propre.
        # Sinon, on préserve les données existantes (déclarations, incidents, etc.)
        db_file = pathlib.Path(state.db_path)
        preserve_mode = db_file.exists() and not force_reset
        if force_reset and db_file.exists():
            try:
                db_file.unlink()
                logger.info(f"DB supprimée pour {state.config.sigle} (force_reset)")
            except Exception as e:
                logger.warning(f"Suppression DB échouée (pas bloquant) : {e}")
            # Idem pour les fichiers SQLite annexes (-wal, -shm)
            for suffix in ("-wal", "-shm", "-journal"):
                aux = pathlib.Path(state.db_path + suffix)
                if aux.exists():
                    try: aux.unlink()
                    except Exception: pass
        elif preserve_mode:
            logger.info(f"DB préservée pour {state.config.sigle} (mode normal)")

        # 1. Importer les modèles SCRIBE
        sys.path.insert(0, str(PROJECT_ROOT))
        from app.database import Base
        import app.models  # noqa: F401 enregistre tous les modèles dans Base.metadata

        url = f"sqlite:///{state.db_path}"
        eng = create_engine(url, connect_args={"check_same_thread": False})
        Base.metadata.create_all(bind=eng)
        logger.info(f"DB initialisée pour {state.config.sigle}")

        Sess = sessionmaker(bind=eng)
        sess = Sess()

        try:
            # 2. Créer le Hospital principal
            from app.models import Hospital
            # h153 — Éviter le fallback "Site_PORT" comme nom de Hospital :
            # il s'afficherait dans le select établissement de la création d'incident.
            # Si seul le sigle auto-généré est disponible, on le rend plus lisible.
            import re as _re
            _raw_nom   = (state.config.nom   or "").strip()
            _raw_sigle = (state.config.sigle or "").strip()
            _is_auto_sigle = bool(_re.match(r'^Site_\d+$', _raw_sigle, _re.IGNORECASE))
            if _raw_nom:
                hospital_nom = _raw_nom
            elif _raw_sigle and not _is_auto_sigle:
                hospital_nom = _raw_sigle
            else:
                # Dernier recours : utiliser le port pour créer un label lisible
                hospital_nom = f"Instance port {state.config.port}" 
            # Sécurité défensive : si pour une raison quelconque un Hospital
            # avec ce nom existe déjà (DB qui n'aurait pas pu être supprimée),
            # on le réutilise au lieu de planter.
            # Chercher aussi l'ancien nom auto-généré "Site_PORT" pour migration
            import re as _re2
            existing_h = sess.query(Hospital).filter(Hospital.nom == hospital_nom).first()
            if not existing_h:
                # Chercher le fallback auto-généré pour le migrer
                old_auto = f"Site_{state.config.port}"
                existing_h = sess.query(Hospital).filter(Hospital.nom == old_auto).first()
                if existing_h:
                    logger.info(f"  Migration Hospital : {old_auto!r} → {hospital_nom!r}")
                    existing_h.nom = hospital_nom
            if existing_h:
                hospital_id = existing_h.id
                existing_h.latitude = state.config.latitude or existing_h.latitude or 45.8992
                existing_h.longitude = state.config.longitude or existing_h.longitude or 6.1294
                logger.info(f"  Hospital existant réutilisé : {hospital_nom} (id={hospital_id})")
            else:
                hospital_principal = Hospital(
                    nom=hospital_nom,
                    latitude=state.config.latitude or 45.8992,    # Valmont par défaut
                    longitude=state.config.longitude or 6.1294,
                )
                sess.add(hospital_principal)
                sess.flush()  # pour avoir l'ID
                hospital_id = hospital_principal.id
                logger.info(f"  Hospital principal créé : {hospital_principal.nom} (id={hospital_id})")

            # 3. Importer UF + capacité depuis le profil xlsx (si présent)
            # En mode preserve, on skip l'import si la table contient déjà
            # des données (pour ne pas dupliquer ou écraser les modifs admin)
            from app.models import UniteFonctionnelle, CapaciteReferentiel
            uf_count = 0
            cap_count = 0
            already_has_uf = sess.query(UniteFonctionnelle).first() is not None
            already_has_capa = sess.query(CapaciteReferentiel).first() is not None
            should_import = not preserve_mode or (not already_has_uf and not already_has_capa)
            if should_import and PROFIL_BASE_XLSX.exists():
                try:
                    uf_count, cap_count = self._import_xlsx_inprocess(
                        sess, hospital_id, PROFIL_BASE_XLSX
                    )
                    logger.info(f"  Profil importé : {uf_count} UF + {cap_count} services capacité")
                except Exception as e:
                    logger.warning(f"  Import xlsx KO (non bloquant) : {e}")
            elif preserve_mode and (already_has_uf or already_has_capa):
                logger.info(f"  Mode preserve : UF/capa déjà présents, import xlsx skippé")
            elif not PROFIL_BASE_XLSX.exists():
                logger.warning(f"  Profil xlsx absent : {PROFIL_BASE_XLSX}")

            # 4. Créer le compte admin
            self._create_admin(sess, state, preserve=preserve_mode)

            # 5. v2.4.7 : pré-créer les services transverses depuis le xlsx
            # (sinon fallback hardcodé dans cartographie.py au 1er hit API)
            if PROFIL_BASE_XLSX.exists():
                try:
                    services = self._read_services_transverses_from_xlsx(PROFIL_BASE_XLSX)
                    if services:
                        from app.models import ServiceStatus
                        for s in services:
                            existing = sess.query(ServiceStatus).filter_by(
                                service_id=s["service_id"]
                            ).first()
                            if not existing:
                                sess.add(ServiceStatus(
                                    service_id=s["service_id"],
                                    libelle=s["libelle"],
                                    statut="OK",
                                ))
                        logger.info(f"  Services transverses xlsx : {len(services)} créés/vérifiés")
                except Exception as e:
                    logger.warning(f"  Services transverses xlsx KO (non bloquant) : {e}")

            sess.commit()
            logger.info(f"Bootstrap terminé pour {state.config.sigle}")
        except Exception as e:
            sess.rollback()
            logger.error(f"Bootstrap échoué pour {state.config.sigle} : {e}")
            raise
        finally:
            sess.close()
            eng.dispose()

    def _import_xlsx_inprocess(self, sess, hospital_id: int, xlsx_path) -> tuple[int, int]:
        """Lit le xlsx et insère UF + capacité dans la DB cible. In-process.

        Retourne (uf_count, capacite_count).
        """
        from openpyxl import load_workbook
        from app.models import UniteFonctionnelle, CapaciteReferentiel

        wb = load_workbook(xlsx_path, data_only=True)
        uf_count = 0
        cap_count = 0

        # ── UF_INCIDENTS ──
        if "UF_INCIDENTS" in wb.sheetnames:
            ws = wb["UF_INCIDENTS"]
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or not row[0] or not row[1]:
                    continue
                if str(row[0]).strip() == "Code UF":
                    continue
                # Colonne 5 (index 4) : actif "O"/"N"
                actif_col = row[4] if len(row) > 4 else None
                if actif_col and str(actif_col).strip().upper() == "N":
                    continue
                code = str(row[0]).strip()
                lib  = str(row[1]).strip()
                pole = str(row[2] or "").strip() if len(row) > 2 else ""

                exists = sess.query(UniteFonctionnelle).filter_by(
                    code_uf=code, hospital_id=hospital_id
                ).first()
                if not exists:
                    sess.add(UniteFonctionnelle(
                        code_uf=code, libelle=lib, pole=pole,
                        hospital_id=hospital_id,
                    ))
                    uf_count += 1
            sess.flush()

        # ── SERVICES_CAPACITE ──
        if "SERVICES_CAPACITE" in wb.sheetnames:
            ws = wb["SERVICES_CAPACITE"]
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or not row[0]:
                    continue
                if str(row[0]).strip() == "Service":
                    continue

                def _i(idx, default=0):
                    if idx >= len(row) or row[idx] is None:
                        return default
                    try:
                        return int(row[idx])
                    except (ValueError, TypeError):
                        return default

                def _s(idx, default=""):
                    if idx >= len(row) or row[idx] is None:
                        return default
                    return str(row[idx]).strip()

                def _b(idx, default=True):
                    val = _s(idx, "O")
                    return val.upper() == "O"

                nom = _s(0)
                code = _s(1) or None
                pole = _s(2)
                site = _s(3)
                capa = _i(4)
                t1   = _i(5)
                t2   = _i(6)
                h    = _b(7)
                f    = _b(8)
                ind  = _b(9)
                tel  = _s(10)
                ordre = _i(11, 99)

                exists = sess.query(CapaciteReferentiel).filter_by(
                    service_nom=nom, site=site
                ).first()
                if not exists:
                    sess.add(CapaciteReferentiel(
                        service_nom=nom, uf_code=code, pole=pole, site=site,
                        capacite_totale=capa, tension_1=t1, tension_2=t2,
                        accept_homme=h, accept_femme=f, accept_indiffer=ind,
                        telephone_cadre=tel, ordre_affichage=ordre,
                    ))
                    cap_count += 1
            sess.flush()

        return uf_count, cap_count

    def _read_directeurs_annuaire_from_xlsx(
        self, xlsx_path: pathlib.Path
    ) -> tuple[list[dict], list[dict], list[dict]]:
        """Lit les onglets DIRECTEURS et TELEPHONIE du xlsx pour peupler
        config.js (directeurs, annuaire_normal, annuaire_secours).

        Retourne (directeurs, annuaire_normal, annuaire_secours).

        Format attendu :
          - Onglet DIRECTEURS : Nom Prénom | Fonction | Abréviation | Téléphone | Note
            (commence ligne 4, en-têtes ligne 3)
          - Onglet TELEPHONIE : Service | Interne/IP | Direct | Mobile | Site | Note
            La distinction NORMAL/SECOURS se fait par une ligne d'en-tête de section
            "CONTACTS NOMINAUX" ou "CONTACTS SECOURS" (case-insensible)
        """
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, data_only=True)
        directeurs: list[dict] = []
        annuaire_normal: list[dict] = []
        annuaire_secours: list[dict] = []

        # ── DIRECTEURS ──
        if "DIRECTEURS" in wb.sheetnames:
            ws = wb["DIRECTEURS"]
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or not row[0]:
                    continue
                nom = str(row[0]).strip()
                if nom.lower() in ("nom prénom", "nom prenom", "nom"):
                    continue
                fonction = str(row[1] or "").strip() if len(row) > 1 else ""
                abrev = str(row[2] or "").strip() if len(row) > 2 else ""
                tel = str(row[3] or "").strip() if len(row) > 3 else ""
                directeurs.append({
                    "nom":         nom,
                    "abreviation": abrev or nom.split()[0][:3].upper(),
                    "telephone":   tel,
                    "fonction":    fonction,
                })

        # ── TELEPHONIE ──
        # Tolérant : si onglet "TELEPHONIE_SECOURS" existe séparément (v2.4.7),
        # on l'utilise pour les secours. Sinon on regarde les sections dans
        # un onglet "TELEPHONIE" unique.
        def _parse_telephonie_sheet(ws, target_normal, target_secours):
            current_section = "normal"
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                cell0 = str(row[0]).strip()
                if not cell0:
                    continue
                low = cell0.lower()
                # Détection en-tête de section : ligne où SEULE la première cellule
                # est remplie (les autres colonnes vides). C'est le cas des
                # titres de sections "CONTACTS NOMINAUX", "CONTACTS SECOURS"
                other_cells_filled = any(
                    row[i] is not None and str(row[i]).strip()
                    for i in range(1, min(len(row), 6))
                )
                if not other_cells_filled:
                    if "secours" in low:
                        current_section = "secours"
                        continue
                    if "nominaux" in low or "normal" in low or "principaux" in low:
                        current_section = "normal"
                        continue
                    # Ligne sans contact utile → skip
                    continue
                # Skip en-têtes répétés
                if low in ("service", "📞", "service ↓"):
                    continue
                # Skip lignes commentaire
                if low.startswith("===") or low.startswith("---"):
                    continue
                entry = {
                    "service": cell0,
                    "interne": str(row[1] or "").strip() if len(row) > 1 else "",
                    "direct":  str(row[2] or "").strip() if len(row) > 2 else "",
                    "mobile":  str(row[3] or "").strip() if len(row) > 3 else "",
                    "site":    str(row[4] or "").strip() if len(row) > 4 else "",
                    "note":    str(row[5] or "").strip() if len(row) > 5 else "",
                }
                if current_section == "secours":
                    target_secours.append(entry)
                else:
                    target_normal.append(entry)

        if "TELEPHONIE" in wb.sheetnames:
            _parse_telephonie_sheet(wb["TELEPHONIE"], annuaire_normal, annuaire_secours)

        # v2.4.7 : onglet dédié pour téléphonie de secours (sécurise vs détection
        # automatique par section). Si présent, écrase annuaire_secours.
        if "TELEPHONIE_SECOURS" in wb.sheetnames:
            secours_explicit: list[dict] = []
            _parse_telephonie_sheet(wb["TELEPHONIE_SECOURS"], secours_explicit, secours_explicit)
            if secours_explicit:
                annuaire_secours = secours_explicit

        return directeurs, annuaire_normal, annuaire_secours

    def _read_services_transverses_from_xlsx(
        self, xlsx_path: pathlib.Path
    ) -> list[dict]:
        """Lit l'onglet SERVICES_TRANSVERSES du xlsx. v2.4.7.

        Format attendu (en-têtes ligne 1) :
          Service ID | Libellé | Ordre

        Exemples :
          securite_physique | Sécurité physique | 1
          logistique        | Logistique        | 2
          si_dpi            | DPI (SI clinique) | 3
          si_msg            | Messagerie        | 4

        Retourne une liste de dicts {service_id, libelle}.
        Si l'onglet est absent ou vide, retourne [] (le caller utilisera le
        fallback hardcodé Sécurité physique + Logistique).
        """
        from openpyxl import load_workbook

        try:
            wb = load_workbook(xlsx_path, data_only=True)
        except Exception:
            return []

        if "SERVICES_TRANSVERSES" not in wb.sheetnames:
            return []

        ws = wb["SERVICES_TRANSVERSES"]
        services = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            sid = str(row[0]).strip()
            if not sid or sid.lower() in ("service id", "service_id", "id"):
                continue
            lib = str(row[1] or "").strip() if len(row) > 1 else ""
            if not lib:
                lib = sid.replace("_", " ").title()
            services.append({
                "service_id": sid,
                "libelle":    lib,
            })
        return services

    def _create_admin(self, sess, state: InstanceState, preserve: bool = False) -> None:
        """Crée ou met à jour le compte admin de l'instance."""
        from app.models import User

        # Hash bcrypt si disponible, sinon SHA-256 (cohérent avec app/api/auth.py
        # de SCRIBE qui accepte les deux en migration transparente)
        # bcrypt limite à 72 bytes — on tronque pour éviter les exceptions
        def _hash(pw: str) -> str:
            pw_bytes = pw.encode("utf-8")[:72]
            pw_safe = pw_bytes.decode("utf-8", errors="ignore")
            try:
                from passlib.context import CryptContext
                ctx = CryptContext(schemes=["bcrypt"], deprecated="auto",
                                   bcrypt__truncate_error=False)
                return ctx.hash(pw_safe)
            except Exception:
                import hashlib
                return hashlib.sha256(pw.encode("utf-8")).hexdigest()

        existing = sess.query(User).filter(
            User.username == state.config.admin_login
        ).first()
        # h60 — En mode EXERCICE, le mot de passe est fixe et non secret
        # (« Exercice2026! ») : on ne force JAMAIS son changement (sinon les
        # joueurs et le collecteur animateur sont bloqués à la 1ère connexion).
        # Détection sans import circulaire via le nom de la dataclass de config.
        _is_exo = type(state.config).__name__ == "ExerciceInstanceConfig"
        if existing:
            existing.role = "admin"
            existing.active = True
            # h120 — Ne reecrire le mot de passe QUE sur premiere init (base non
            # preservee) ou reset explicite (panneau). Une simple relance preserve
            # le mdp choisi par l'utilisateur a sa premiere connexion.
            _reset_pwd = (not preserve) or getattr(state, "pending_admin_reset", False)
            if _reset_pwd:
                existing.hashed_password = _hash(state.config.admin_password)
                try:
                    state.pending_admin_reset = False
                except Exception:
                    pass
            # v3.4 (h38g) — Aussi mettre à jour le display_name si défini
            # par le wizard. Évite que le legacy "Directeur de Crise"
            # créé par bootstrap_admin reste figé alors que l'utilisateur
            # avait renseigné un autre nom (ex: "RSSI").
            if state.config.admin_display_name:
                existing.display_name = state.config.admin_display_name
            # v3.4 (h38c) — Forcer le changement de mot de passe à la
            # première connexion (sauf si déjà déclaré non-nécessaire).
            # Le mdp initial étant généré par le wizard et potentiellement
            # transmis par email/chat, l'utilisateur doit le changer.
            if _reset_pwd:
                try:
                    existing.must_change_password = not _is_exo
                except Exception:
                    pass
            logger.info(f"  Compte admin {'reinitialise' if _reset_pwd else 'preserve'} : {state.config.admin_login}")
        else:
            new_admin = User(
                username=state.config.admin_login,
                display_name=state.config.admin_display_name or state.config.admin_login.upper(),
                role="admin",
                hashed_password=_hash(state.config.admin_password),
                active=True,
            )
            # v3.4 (h38c) — Forcer le changement de mot de passe à la
            # première connexion
            try:
                new_admin.must_change_password = not _is_exo
            except Exception:
                pass
            sess.add(new_admin)
            logger.info(f"  Compte admin créé : {state.config.admin_login} (display='{new_admin.display_name}', mdp à changer à la 1ère connexion)")
        # v3.4 (h38g) — Commit explicite ici pour être absolument certain
        # que must_change_password=True est persisté avant que l'instance
        # ne réponde à un éventuel login en attente. Sans ce commit, un
        # bootstrap concurrent ou un cache pouvait remettre le flag à False.
        try:
            sess.commit()
        except Exception as e:
            logger.warning(f"  Commit admin échoué (non bloquant) : {e}")

        # v3.4 (h38h) — Écriture des plugin_states selon les préférences
        # du wizard. Chaque plugin listé dans state.config.plugins_disabled
        # se voit attribuer enabled=False. Les autres restent activés par
        # défaut (config.PLUGINS = True). L'utilisateur peut ensuite modifier
        # ces choix via /admin/plugins après création.
        disabled = list(getattr(state.config, "plugins_disabled", []) or [])
        if disabled:
            try:
                from core.plugin_state_model import PluginState
                # S'assurer que la table existe (idempotent)
                from app.database import Base, engine
                Base.metadata.create_all(bind=engine, tables=[PluginState.__table__])

                n_disabled = 0
                for plugin_id in disabled:
                    if not isinstance(plugin_id, str) or not plugin_id.strip():
                        continue
                    pid = plugin_id.strip()
                    existing_ps = sess.query(PluginState).filter_by(plugin_id=pid).first()
                    if existing_ps:
                        if existing_ps.enabled:
                            existing_ps.enabled = False
                            n_disabled += 1
                    else:
                        sess.add(PluginState(plugin_id=pid, enabled=False))
                        n_disabled += 1
                sess.commit()
                logger.info(f"  Plugins désactivés via wizard : {n_disabled} ({disabled})")
            except Exception as e:
                # Non bloquant : si le modèle PluginState n'existe pas ou autre,
                # on log et on continue. L'instance fonctionnera juste avec
                # les plugins par défaut.
                logger.warning(f"  Écriture plugin_states KO (non bloquant) : {e}")

    def _auto_enrol(self, state: InstanceState, collecteur_url: str) -> None:
        """Pré-enregistre le token de l'instance auprès du collecteur.

        L'instance fille pushera ses données toutes les 30s avec son token.
        Le collecteur doit le connaître pour accepter les push.
        """
        try:
            import urllib.request
            import urllib.parse

            # Lire le token admin du collecteur (collecteur_admin.json)
            admin_file = PROJECT_ROOT / "collecteur" / "collecteur_admin.json"
            if not admin_file.exists():
                logger.warning("collecteur_admin.json introuvable, pas d'auto-enrôlement")
                return
            with open(admin_file, encoding="utf-8") as f:
                admin_data = json.load(f)
            admin_token = admin_data.get("admin_token")
            if not admin_token:
                logger.warning("admin_token absent dans collecteur_admin.json")
                return

            # POST /api/admin/tokens (route existante du collecteur)
            payload = json.dumps({
                "sigle": state.config.sigle,
                "token": state.fed_token,
                "nom": state.config.nom,
                "latitude": state.config.latitude,
                "longitude": state.config.longitude,
            }).encode("utf-8")
            req = urllib.request.Request(
                f"{collecteur_url}/api/admin/tokens",
                data=payload,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {admin_token}",
                },
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=5) as resp:
                if resp.status >= 300:
                    logger.warning(f"Auto-enrôlement HTTP {resp.status}")
                else:
                    logger.info(f"Auto-enrôlement OK pour {state.config.sigle}")
        except Exception as e:
            logger.warning(f"Auto-enrôlement échoué : {e}")
            # Non bloquant


# ─────────────────────────────────────────────────────────────────────────────
# Helpers OS-level
# ─────────────────────────────────────────────────────────────────────────────

def _pid_alive(pid: int) -> bool:
    """Vérifie qu'un PID existe (cross-platform Linux/macOS/Windows)."""
    if not pid:
        return False
    try:
        if os.name == "posix":
            # Unix : signal 0 ne tue pas, vérifie juste l'existence
            os.kill(pid, 0)
            return True
        else:
            # Windows : OpenProcess + GetExitCodeProcess
            # Plus fiable que os.kill(pid, 0) qui peut produire WinError 11
            import ctypes
            PROCESS_QUERY_LIMITED_INFORMATION = 0x1000
            STILL_ACTIVE = 259
            handle = ctypes.windll.kernel32.OpenProcess(
                PROCESS_QUERY_LIMITED_INFORMATION, False, pid
            )
            if not handle:
                return False
            try:
                exit_code = ctypes.c_ulong()
                ok = ctypes.windll.kernel32.GetExitCodeProcess(
                    handle, ctypes.byref(exit_code)
                )
                if not ok:
                    return False
                return exit_code.value == STILL_ACTIVE
            finally:
                ctypes.windll.kernel32.CloseHandle(handle)
    except ProcessLookupError:
        return False
    except PermissionError:
        # Le PID existe mais on n'a pas le droit (peu probable en local)
        return True
    except OSError as e:
        # WinError 11 ("Bad EXE format") ou autres : on considère vivant par
        # défaut pour ne pas perdre l'instance, mais on log
        logger.debug(f"_pid_alive({pid}) OSError : {e} → suppose vivant")
        return True


def _port_in_use(port: int) -> bool:
    """Vérifie si un port est déjà occupé sur localhost."""
    import socket
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(0.5)
    try:
        sock.bind(("127.0.0.1", port))
        return False
    except OSError:
        return True
    finally:
        sock.close()


# Instance globale — utilisée par les routes API du collecteur
_manager: InstanceManager | None = None


def get_manager() -> InstanceManager:
    """Retourne le singleton InstanceManager."""
    global _manager
    if _manager is None:
        _manager = InstanceManager()
    return _manager
