"""
master/master_routes.py — Routes API pour le pilotage d'instances
==================================================================
À importer depuis collecteur.py :

    from master.master_routes import router as master_router, lifecycle_register
    app.include_router(master_router)
    lifecycle_register(app)  # pour stop_all à l'arrêt

Routes exposées (toutes sous /api/master) :
  GET    /api/master/instances              Liste + statuts
  GET    /api/master/instances/{port}       Détails d'une instance
  PUT    /api/master/instances/{port}       Modifie config (si arrêtée)
  POST   /api/master/instances/{port}/start Lance l'instance
  POST   /api/master/instances/{port}/stop  Arrête l'instance
  GET    /api/master/instances/{port}/credentials  Récup login/mdp
  POST   /api/master/instances/custom       Ajoute instance custom
  DELETE /api/master/instances/{port}       Supprime instance
  POST   /api/master/instances/{port}/regenerate-password  Nouveau mdp
  GET    /api/master/instances/{port}/logs  Tail des logs
  GET    /api/master/geocode?q=...          Géocodage Nominatim
  GET    /api/master/profil/uf              Liste UF du profil de base
  POST   /api/master/profil/uf              Ajoute / modifie UF
  DELETE /api/master/profil/uf/{code}       Supprime UF
  GET    /api/master/profil/download        Download xlsx profil
  POST   /api/master/profil/upload          Upload xlsx profil
"""
from __future__ import annotations

import json
import logging
import os, pathlib
import re
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Request, Form
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from pydantic import BaseModel, Field


def now_iso() -> str:
    """Timestamp ISO 8601 UTC pour les flags (.onboarding_done, .wizard_force,
    rapport timestamps, etc.). v2.4.8.6 : était manquante, causait
    'name now_iso is not defined' au runtime sur tous les writes de flag."""
    return datetime.now(timezone.utc).isoformat()


from .instances_manager import (
    get_manager,
    InstanceConfig,
    PROFIL_BASE_XLSX,
    PROJECT_ROOT,
    generate_password,
)
from .geocoding import geocode

logger = logging.getLogger("scribe.master")

router = APIRouter(prefix="/api/master", tags=["master"])


# ─────────────────────────────────────────────────────────────────────────────
# Auth — réutilise require_admin du collecteur si possible, sinon fallback
# ─────────────────────────────────────────────────────────────────────────────

def _check_admin(request: Request) -> None:
    """Vérifie un token admin Bearer dans le header Authorization.

    Réutilise le admin_token du collecteur (collecteur/collecteur_admin.json).
    """
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(401, "Authorization Bearer requis")
    token = auth[7:].strip()

    admin_file = PROJECT_ROOT / "collecteur" / "collecteur_admin.json"
    if not admin_file.exists():
        raise HTTPException(500, "collecteur_admin.json introuvable")
    with open(admin_file, encoding="utf-8") as f:
        admin = json.load(f)
    if token != admin.get("admin_token"):
        raise HTTPException(403, "Token admin invalide")


# ─────────────────────────────────────────────────────────────────────────────
# Modèles Pydantic
# ─────────────────────────────────────────────────────────────────────────────

class InstanceUpdate(BaseModel):
    sigle:          str | None = None
    nom:            str | None = None
    admin_login:    str | None = None
    admin_password: str | None = None
    adresse:        str | None = None
    latitude:       float | None = None
    longitude:      float | None = None
    synchroniser:   bool | None = None


class InstanceCreate(BaseModel):
    port:           int = Field(ge=1024, le=65535)
    sigle:          str | None = None
    nom:            str | None = None
    admin_login:    str | None = None
    admin_password: str | None = None
    adresse:        str | None = None
    latitude:       float | None = None
    longitude:      float | None = None
    synchroniser:   bool | None = None


class UFEntry(BaseModel):
    code:    str
    libelle: str
    pole:    str = ""
    site:    str = "Site Principal"
    actif:   bool = True


# ─────────────────────────────────────────────────────────────────────────────
# Routes — Instances
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/instances")
def list_instances(request: Request):
    _check_admin(request)
    return {"instances": get_manager().list_instances()}


@router.get("/instances/{port:int}")
def get_instance(port: int, request: Request):
    _check_admin(request)
    state = get_manager().get_instance(port)
    if not state:
        raise HTTPException(404, f"Instance port {port} inconnue")
    d = state.to_dict()
    d["config"]["admin_password"] = "***"
    return d


@router.put("/instances/{port:int}")
def update_instance(port: int, payload: InstanceUpdate, request: Request):
    _check_admin(request)
    try:
        state = get_manager().update_config(port, **payload.model_dump(exclude_unset=True))
    except ValueError as e:
        raise HTTPException(400, str(e))
    d = state.to_dict()
    d["config"]["admin_password"] = "***"
    return d


@router.post("/instances/{port}/start")
def start_instance(port: int, request: Request):
    _check_admin(request)
    try:
        # v3000h17 — Priorité au hostname configuré (fichier hostname.conf),
        # sinon fallback sur le Host: de la requête courante (auto-détection).
        from master.hostname_config import get_external_host
        host = get_external_host(request=request, fallback="localhost")
        collecteur_port = request.url.port or 9000
        collecteur_url = f"http://{host}:{collecteur_port}"
        state = get_manager().start(port, collecteur_url=collecteur_url)
    except ValueError as e:
        # ValueError = problème métier (port occupé, instance inconnue...)
        raise HTTPException(400, str(e))
    except Exception as e:
        # Toute autre exception : log complet + 500 avec message lisible
        import traceback
        logger.error(f"Erreur start port {port} : {traceback.format_exc()}")
        logger.exception("Erreur lancement instance"); raise HTTPException(500, "Erreur interne lors du lancement (voir logs serveur)")
    return {
        "ok":      True,
        "pid":     state.pid,
        "url":     f"http://localhost:{state.config.port}",
        "started_at": state.started_at,
    }


@router.post("/instances/{port}/stop")
def stop_instance(port: int, request: Request):
    _check_admin(request)
    try:
        get_manager().stop(port)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "port": port, "statut": "arrete"}


@router.get("/instances/{port}/credentials")
def get_credentials(port: int, request: Request):
    _check_admin(request)
    try:
        return get_manager().get_credentials(port)
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.post("/instances/{port}/regenerate-password")
def regenerate_password(port: int, request: Request):
    _check_admin(request)
    mgr = get_manager()
    state = mgr.get_instance(port)
    if not state:
        raise HTTPException(404, f"Instance port {port} inconnue")
    if state.statut == "actif":
        raise HTTPException(400, "Arrêtez l'instance avant de régénérer le mot de passe")
    new_pwd = generate_password()
    mgr.update_config(port, admin_password=new_pwd)
    return {"ok": True, "password": new_pwd}


@router.post("/instances/{port}/reset-db")
def reset_instance_db(port: int, request: Request):
    """Réinitialise la DB de l'instance en repartant du profil xlsx.
    ATTENTION : toutes les données saisies (incidents, transferts, déclarations
    capacitaires, messages) seront PERDUES. L'instance doit être arrêtée."""
    _check_admin(request)
    mgr = get_manager()
    state = mgr.get_instance(port)
    if not state:
        raise HTTPException(404, f"Instance port {port} inconnue")
    if state.statut == "actif":
        raise HTTPException(400, "Arrêtez l'instance avant de réinitialiser la DB.")
    try:
        mgr._bootstrap_db(state, force_reset=True)
    except Exception as e:
        import traceback
        logger.error(f"Reset DB échoué pour port {port} : {traceback.format_exc()}")
        logger.exception("Erreur reset DB"); raise HTTPException(500, "Erreur interne lors du reset (voir logs serveur)")
    return {"ok": True, "sigle": state.config.sigle, "msg": "DB réinitialisée depuis le profil xlsx"}


@router.post("/free-ports")
def free_all_scribe_ports_route(request: Request):
    """v3.0.0 — Libère prudemment TOUS les ports SCRIBE étendus
    (master + instances + collecteurs + démo). Ne tue que les process
    clairement identifiables comme SCRIBE — les process tiers sont signalés
    mais jamais touchés.

    Utile en cas de bazar (process orphelins après crash, reboot incomplet,
    etc.) sans avoir à redémarrer la machine."""
    _check_admin(request)
    try:
        from master.port_cleanup import free_all_scribe_ports, summarize_results
        results = free_all_scribe_ports()
        return {
            "ok": True,
            "summary": summarize_results(results),
            "results": results,
        }
    except Exception as e:
        import traceback
        logger.error(f"free-ports échoué : {traceback.format_exc()}")
        raise HTTPException(500, "Erreur lors de la libération des ports (voir logs)")


@router.post("/instances/custom", status_code=201)
def add_custom(payload: InstanceCreate, request: Request):
    _check_admin(request)
    try:
        state = get_manager().add_custom(
            port=payload.port,
            **{k: v for k, v in payload.model_dump(exclude_unset=True).items() if k != "port"}
        )
    except ValueError as e:
        raise HTTPException(400, str(e))
    # v2.4.8 : marquer l'onboarding comme terminé dès qu'une instance custom
    # est créée (couvre le chemin "+ Nouveau" du master qui contourne le wizard)
    if state.config.sigle and not state.config.sigle.startswith("Site_"):
        _mark_onboarding_done()
    d = state.to_dict()
    d["config"]["admin_password"] = "***"
    return d


@router.delete("/instances/{port:int}", status_code=204)
def delete_instance(port: int, request: Request):
    _check_admin(request)
    try:
        get_manager().remove_instance(port)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return None


@router.get("/instances/{port}/logs")
def get_logs(port: int, request: Request, lines: int = 200):
    _check_admin(request)
    state = get_manager().get_instance(port)
    if not state or not state.log_path:
        return {"logs": "", "info": "Aucun log disponible"}
    log_file = pathlib.Path(state.log_path)
    if not log_file.exists():
        return {"logs": "", "info": "Fichier de log introuvable"}
    try:
        with open(log_file, "rb") as f:
            content = f.read()
        text = content.decode("utf-8", errors="replace")
        all_lines = text.splitlines()
        tail = "\n".join(all_lines[-lines:])
        return {"logs": tail, "lines_total": len(all_lines)}
    except Exception as e:
        return {"logs": "", "info": f"Erreur lecture logs : {e}"}


# ─────────────────────────────────────────────────────────────────────────────
# Routes — Géocodage
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/geocode")
def proxy_geocode(q: str, request: Request):
    """Proxy local vers Nominatim (évite la CSP côté navigateur).

    Pas de bias pays par défaut — accepte les adresses internationales
    (utile pour les hôpitaux frontaliers : Suisse, Belgique, Allemagne,
    Luxembourg, Monaco, Andorre). Si l'utilisateur tape "1202 Genève",
    il doit obtenir Genève en Suisse, pas un faux résultat en France.

    Retourne :
      - 200 + {lat, lon, display_name} si trouvé
      - 404 + {detail: ...} si vraiment aucun résultat
      - 502 + {detail: ...} si Nominatim injoignable / quota
    """
    _check_admin(request)
    if not q or not q.strip():
        raise HTTPException(400, "Paramètre q vide")

    # Tentative 1 : requête telle quelle, recherche mondiale
    result = geocode(q)
    if result:
        return result

    # Tentative 2 : si la requête est sans virgule, on tente avec
    # un format ville en extrayant les derniers mots
    q_clean = q.strip()
    if "," not in q_clean and " " in q_clean:
        words = q_clean.split()
        # On prend les 2 derniers mots comme indication de ville (ex: "Metz Tessy")
        if len(words) >= 2:
            ville_candidate = " ".join(words[-2:])
            result = geocode(ville_candidate)
            if result:
                return result
        # Sinon dernier mot
        result = geocode(words[-1])
        if result:
            return result

    raise HTTPException(
        404,
        "Aucun résultat. Essayez le format 'numéro rue, ville' ou 'ville, pays'."
    )


# ─────────────────────────────────────────────────────────────────────────────
# Routes — Profil de base (UF)
# ─────────────────────────────────────────────────────────────────────────────

def _read_profil_uf() -> list[dict]:
    """Lit la feuille UF_INCIDENTS du profil xlsx."""
    if not PROFIL_BASE_XLSX.exists():
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(PROFIL_BASE_XLSX, data_only=True)
        if "UF_INCIDENTS" not in wb.sheetnames:
            return []
        ws = wb["UF_INCIDENTS"]
        rows = []
        # Trouver la ligne d'en-têtes (cherche "Code UF" en colonne A)
        header_row = None
        for r in range(1, min(15, ws.max_row + 1)):
            v = ws.cell(r, 1).value
            if v and "code" in str(v).lower():
                header_row = r
                break
        if header_row is None:
            header_row = 4  # fallback

        for r in range(header_row + 1, ws.max_row + 1):
            code     = ws.cell(r, 1).value
            libelle  = ws.cell(r, 2).value
            pole     = ws.cell(r, 3).value or ""
            site     = ws.cell(r, 4).value or "Site Principal"
            actif    = ws.cell(r, 5).value
            if not code or not libelle:
                continue
            rows.append({
                "code":    str(code).strip(),
                "libelle": str(libelle).strip(),
                "pole":    str(pole).strip(),
                "site":    str(site).strip(),
                "actif":   str(actif or "").strip().upper() in ("O", "OUI", "Y", "YES", "TRUE", "1"),
            })
        return rows
    except Exception as e:
        logger.error(f"Lecture UF échouée : {e}")
        return []


def _write_profil_uf(rows: list[dict]) -> None:
    """Réécrit la feuille UF_INCIDENTS du profil xlsx."""
    if not PROFIL_BASE_XLSX.exists():
        raise HTTPException(500, "Profil xlsx absent")
    from openpyxl import load_workbook
    wb = load_workbook(PROFIL_BASE_XLSX)
    if "UF_INCIDENTS" not in wb.sheetnames:
        raise HTTPException(500, "Feuille UF_INCIDENTS absente")
    ws = wb["UF_INCIDENTS"]

    # Trouver la ligne d'en-têtes
    header_row = None
    for r in range(1, min(15, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if v and "code" in str(v).lower():
            header_row = r
            break
    if header_row is None:
        header_row = 4

    # Effacer les anciennes lignes
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)

    # Réécrire
    for i, row in enumerate(rows, start=header_row + 1):
        ws.cell(i, 1, value=row["code"])
        ws.cell(i, 2, value=row["libelle"])
        ws.cell(i, 3, value=row.get("pole", ""))
        ws.cell(i, 4, value=row.get("site", "Site Principal"))
        ws.cell(i, 5, value="O" if row.get("actif", True) else "N")

    wb.save(PROFIL_BASE_XLSX)


@router.get("/profil/uf")
def list_uf(request: Request):
    _check_admin(request)
    return {"uf": _read_profil_uf()}


@router.post("/profil/uf")
def save_uf(payload: dict, request: Request):
    """Remplace toute la liste des UF du profil de base."""
    _check_admin(request)
    rows = payload.get("uf")
    if not isinstance(rows, list):
        raise HTTPException(400, "Format attendu : { 'uf': [...] }")
    # Validation minimale
    cleaned = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        if not r.get("code") or not r.get("libelle"):
            continue
        cleaned.append({
            "code":    str(r["code"]).strip(),
            "libelle": str(r["libelle"]).strip(),
            "pole":    str(r.get("pole", "")).strip(),
            "site":    str(r.get("site", "Site Principal")).strip(),
            "actif":   bool(r.get("actif", True)),
        })
    _write_profil_uf(cleaned)
    return {"ok": True, "count": len(cleaned)}


@router.get("/profil/download")
def download_profil(request: Request):
    _check_admin(request)
    if not PROFIL_BASE_XLSX.exists():
        raise HTTPException(404, "Profil xlsx absent")
    return FileResponse(
        PROFIL_BASE_XLSX,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="SCRIBE_profil_base.xlsx",
    )


@router.get("/lang/{code}")
def get_lang(code: str):
    """Sert un fichier de langue depuis app/lang/. Pas d'auth nécessaire
    (les libellés UI ne sont pas sensibles)."""
    # Sécurité : whitelist stricte des codes langue
    if code not in {"fr", "en", "de", "es", "it", "nl", "pl", "pt"}:
        raise HTTPException(404, "Langue inconnue")
    lang_file = PROJECT_ROOT / "app" / "lang" / f"{code}.json"
    if not lang_file.exists():
        raise HTTPException(404, "Fichier langue absent")
    return FileResponse(
        lang_file,
        media_type="application/json",
        headers={"Cache-Control": "public, max-age=3600"},
    )


@router.post("/profil/upload")
async def upload_profil(request: Request, file: UploadFile = File(...)):
    _check_admin(request)
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Fichier xlsx attendu")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(413, "Fichier trop gros (max 10 MB)")
    PROFIL_BASE_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with open(PROFIL_BASE_XLSX, "wb") as f:
        f.write(content)
    return {"ok": True, "size": len(content)}


# ─────────────────────────────────────────────────────────────────────────────
# UI — Page HTML du panneau Instances
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/ui", response_class=HTMLResponse)
def master_ui():
    """Sert la page HTML du panneau Instances."""
    html_path = pathlib.Path(__file__).parent / "instances.html"
    if html_path.exists():
        return HTMLResponse(html_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>UI master non disponible</h1>", status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
# Routes mode EXERCICE — symétriques au mode prod, plage de ports 8660-8669,
# poussent vers le collecteur exercice :8565 au lieu de :9000
# ─────────────────────────────────────────────────────────────────────────────

from master.exercice_manager import get_exercice_manager  # noqa: E402


class ExerciceInstanceUpdate(BaseModel):
    sigle:          str | None = None
    nom:            str | None = None
    admin_login:    str | None = None
    admin_password: str | None = None
    adresse:        str | None = None
    latitude:       float | None = None
    longitude:      float | None = None


@router.get("/exercice/status")
def exercice_status(request: Request):
    """État global du mode exercice : collecteur :8565 + instances 8660-8669."""
    _check_admin(request)
    return get_exercice_manager().get_status()


@router.post("/exercice/collecteur/start")
def exercice_collecteur_start(request: Request):
    """Démarre le collecteur exercice :8565 en subprocess."""
    _check_admin(request)
    try:
        return get_exercice_manager().start_collecteur()
    except (ValueError, FileNotFoundError) as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        import traceback
        logger.error(f"Erreur start collecteur exo : {traceback.format_exc()}")
        logger.exception("Erreur lancement instance"); raise HTTPException(500, "Erreur interne lors du lancement (voir logs serveur)")


@router.post("/exercice/collecteur/stop")
def exercice_collecteur_stop(request: Request):
    """Arrête le collecteur exercice ET toutes les instances exercice."""
    _check_admin(request)
    try:
        return get_exercice_manager().stop_collecteur()
    except Exception as e:
        raise HTTPException(500, f"Erreur arrêt : {type(e).__name__}: {e}")


@router.get("/exercice/instances")
def exercice_list_instances(request: Request):
    """Liste les 10 slots exercice (8660-8669)."""
    _check_admin(request)
    return {"instances": get_exercice_manager().list_instances()}


@router.put("/exercice/instances/{port}")
def exercice_update_instance(
    port: int, payload: ExerciceInstanceUpdate, request: Request
):
    """Met à jour la config d'un slot exercice (sigle, login, adresse, etc.).
    Refusé si l'instance est active (arrêter d'abord)."""
    _check_admin(request)
    try:
        state = get_exercice_manager().update_config(
            port, **payload.model_dump(exclude_unset=True)
        )
    except ValueError as e:
        raise HTTPException(400, str(e))
    d = state.to_dict()
    d["config"]["admin_password"] = "***"
    return d


@router.post("/exercice/instances/{port}/start")
def exercice_start_instance(port: int, request: Request):
    """Démarre une instance exercice. Démarre aussi le collecteur :8565
    automatiquement s'il n'est pas déjà actif (commodité utilisateur)."""
    _check_admin(request)
    try:
        state = get_exercice_manager().start(port)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        import traceback
        logger.error(f"Erreur start exo port {port} : {traceback.format_exc()}")
        logger.exception("Erreur lancement instance"); raise HTTPException(500, "Erreur interne lors du lancement (voir logs serveur)")
    return {
        "ok":      True,
        "pid":     state.pid,
        "url":     f"http://localhost:{state.config.port}",
        "started_at": state.started_at,
    }


@router.post("/exercice/instances/{port}/stop")
def exercice_stop_instance(port: int, request: Request):
    """Arrête une instance exercice."""
    _check_admin(request)
    try:
        get_exercice_manager().stop(port)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "port": port, "statut": "arrete"}


@router.get("/exercice/instances/{port}/credentials")
def exercice_get_credentials(port: int, request: Request):
    """Retourne login + mot de passe en clair (pour copie animateur)."""
    _check_admin(request)
    mgr = get_exercice_manager()
    if port not in mgr.instances:
        raise HTTPException(404, f"Port {port} inconnu")
    state = mgr.instances[port]
    return {
        "login":    state.config.admin_login,
        "password": state.config.admin_password,
        "url":      f"http://localhost:{port}",
    }


@router.post("/exercice/reset")
def exercice_reset(request: Request):
    """Réinitialise toutes les DBs exercice + l'état du collecteur exercice.
    Toutes les instances doivent être arrêtées au préalable.
    Équivalent de `lancer_exercice.sh --reset` mais piloté depuis le master."""
    _check_admin(request)
    try:
        return get_exercice_manager().reset_all_dbs()
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        import traceback
        logger.error(f"Erreur reset exo : {traceback.format_exc()}")
        raise HTTPException(500, f"Erreur reset : {type(e).__name__}: {e}")


@router.get("/exercice/ui", response_class=HTMLResponse)
def exercice_ui():
    """Sert la page HTML de pilotage exercice."""
    html_path = pathlib.Path(__file__).parent / "exercice.html"
    if html_path.exists():
        return HTMLResponse(html_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>UI exercice non disponible</h1>", status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
# WIZARD D'ONBOARDING — splash screen, wizard 5 étapes, import/export xlsx
# ─────────────────────────────────────────────────────────────────────────────

ONBOARDING_FLAG = pathlib.Path(__file__).parent / ".onboarding_done"
# v2.4.8.3 : flag temporaire posé par /onboarding/reset pour forcer
# l'affichage du wizard une seule fois, malgré l'auto-repair de status.
# Consommé (supprimé) au premier appel de /status.
WIZARD_FORCE_FLAG = pathlib.Path(__file__).parent / ".wizard_force"


def _mark_onboarding_done():
    """Helper v2.4.8.4 : marque l'onboarding comme fait (crée le flag stable)
    ET nettoie le flag éphémère .wizard_force s'il existe.
    Utilisé partout où l'onboarding est terminé (finish, create-instance,
    instances/custom, instances/{port}/import-xlsx).
    Non bloquant : log un warning et continue en cas d'erreur disque."""
    try:
        ONBOARDING_FLAG.parent.mkdir(parents=True, exist_ok=True)
        ONBOARDING_FLAG.write_text(now_iso(), encoding="utf-8")
    except Exception as e:
        logger.warning(f"Création flag onboarding échouée (non bloquant) : {e}")
    if WIZARD_FORCE_FLAG.exists():
        try:
            WIZARD_FORCE_FLAG.unlink()
        except Exception as e:
            logger.warning(f"Suppression flag wizard_force échouée (non bloquant) : {e}")


@router.get("/onboarding/status")
def onboarding_status():
    """Retourne si l'utilisateur a déjà fini l'onboarding ou pas.
    Utilisé par le master.html pour décider d'afficher le splash."""
    mgr = get_manager()
    instances_configured = sum(
        1 for s in mgr.instances.values() if s.config.sigle and not s.config.sigle.startswith("Site_")
    )
    # v2.4.8.3 (corrigé v2.4.8.4) : si l'utilisateur a cliqué "🎯 Wizard"
    # (route /onboarding/reset), le flag .wizard_force est posé. Tant qu'il
    # existe, on force show_wizard=true.
    # IMPORTANT : ce flag N'EST PLUS consommé par /status (sinon il disparaît
    # au 1er appel et les appels suivants — il y en a souvent 2-3 pendant
    # le rechargement de la page — voient déjà l'état "normal", retournant
    # sur l'UI au lieu du wizard).
    # Le flag est consommé uniquement par /onboarding/finish (wizard terminé)
    # ou par un skip explicite, voir routes correspondantes.
    if WIZARD_FORCE_FLAG.exists():
        return {
            "done":                 False,
            "instances_count":      len(mgr.instances),
            "instances_configured": instances_configured,
            "show_wizard":          True,
            "forced":               True,  # debug-friendly
        }
    # v2.4.8.1 : on considère l'onboarding fini si soit (a) le flag existe,
    # soit (b) au moins une instance non-démo est configurée.
    flag_exists = ONBOARDING_FLAG.exists()
    done = flag_exists or instances_configured > 0
    # Auto-repair : si done implicite (instances déjà là) mais flag absent,
    # on le crée pour éviter de recalculer à chaque appel
    if done and not flag_exists:
        try:
            ONBOARDING_FLAG.parent.mkdir(parents=True, exist_ok=True)
            ONBOARDING_FLAG.write_text(now_iso(), encoding="utf-8")
        except Exception as e:
            logger.warning(f"Auto-création flag onboarding échouée (non bloquant) : {e}")
    return {
        "done":                 done,
        "instances_count":      len(mgr.instances),
        "instances_configured": instances_configured,
        "show_wizard":          not done,
    }


@router.post("/onboarding/finish")
def onboarding_finish(request: Request):
    """Marque l'onboarding comme terminé (création du flag .onboarding_done).
    v2.4.8.4 : nettoie aussi le flag wizard_force qui peut être actif si
    l'utilisateur arrive ici via le bouton 🎯 Wizard.
    v2.5.0 patch sécu E5 : protégé par _check_admin (sauf premier boot)."""
    # Si onboarding pas encore fait → autoriser (premier boot, pas encore d'admin)
    # Si déjà fait et bouton wizard utilisé → exiger admin
    if ONBOARDING_FLAG.exists():
        _check_admin(request)
    try:
        ONBOARDING_FLAG.parent.mkdir(parents=True, exist_ok=True)
        ONBOARDING_FLAG.write_text(now_iso(), encoding="utf-8")
        if WIZARD_FORCE_FLAG.exists():
            try: WIZARD_FORCE_FLAG.unlink()
            except Exception: pass
        return {"ok": True, "flag": str(ONBOARDING_FLAG)}
    except Exception as e:
        raise HTTPException(500, f"Impossible de marquer l'onboarding : {e}")


@router.post("/onboarding/reset")
def onboarding_reset(request: Request):
    """Réinitialise le flag d'onboarding et force l'affichage du wizard
    une fois (même si des instances sont déjà configurées).
    Appelé par le bouton 🎯 Wizard du master."""
    _check_admin(request)
    if ONBOARDING_FLAG.exists():
        ONBOARDING_FLAG.unlink()
    # v2.4.8.3 : poser un flag temporaire pour que le prochain /status renvoie
    # show_wizard=true sans déclencher l'auto-repair (qui sinon recrée le flag
    # immédiatement parce que des instances sont configurées).
    try:
        WIZARD_FORCE_FLAG.parent.mkdir(parents=True, exist_ok=True)
        WIZARD_FORCE_FLAG.write_text(now_iso(), encoding="utf-8")
    except Exception as e:
        logger.warning(f"Création flag wizard_force échouée : {e}")
    return {"ok": True}


@router.get("/onboarding/ui", response_class=HTMLResponse)
def onboarding_ui():
    """Sert la page HTML du wizard d'onboarding (splash + 5 étapes)."""
    html_path = pathlib.Path(__file__).parent / "onboarding.html"
    if html_path.exists():
        return HTMLResponse(html_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>Onboarding non disponible</h1>", status_code=500)


@router.get("/onboarding/template-xlsx")
def onboarding_template_xlsx():
    """Téléchargement du template Excel vierge."""
    template_path = pathlib.Path(__file__).parent.parent / "SCRIBE_config_etablissement.xlsx"
    if not template_path.exists():
        raise HTTPException(404, "Template SCRIBE_config_etablissement.xlsx introuvable")
    from fastapi.responses import FileResponse
    return FileResponse(
        path=str(template_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="SCRIBE_config_etablissement.xlsx",
    )


@router.post("/onboarding/export-config")
async def onboarding_export_config(request: Request):
    """v3.4 (h38) — Export d'une config partielle saisie dans le wizard,
    AVANT que l'instance soit créée. L'utilisateur peut continuer son
    wizard, ou interrompre et reprendre plus tard avec son xlsx.
    """
    _check_admin(request)
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(400, "JSON invalide")

    # Strip défensif sur les champs textes
    cfg = {}
    for k, v in payload.items():
        if isinstance(v, str):
            cfg[k] = v.strip()
        else:
            cfg[k] = v

    if not cfg.get("sigle") or cfg["sigle"].upper().startswith("SITE_"):
        raise HTTPException(400, "Sigle requis (3-5 lettres)")

    try:
        from master.excel_export import export_instance_to_xlsx
        # On exporte SANS secrets (mdp hashé, clés API vidées) car c'est
        # un fichier intermédiaire qui peut transiter par email/clé USB
        xlsx_bytes = export_instance_to_xlsx(cfg, db_data=None, include_secrets=False)
    except Exception as e:
        logger.exception("Erreur export config wizard")
        raise HTTPException(500, f"Erreur génération xlsx: {e}")

    from fastapi.responses import Response
    safe_sigle = re.sub(r"[^a-zA-Z0-9_-]", "", cfg["sigle"]) or "instance"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="SCRIBE_config_{safe_sigle}.xlsx"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Modèle de création d'instance depuis le wizard
# ─────────────────────────────────────────────────────────────────────────────

class WizardInstanceCreate(BaseModel):
    sigle:             str
    nom:               str
    finess:            str | None = None
    langue:            str = "fr"
    admin_login:       str = "dircrise"
    admin_password:    str
    nom_affiche_admin: str | None = "Directeur de Crise"
    adresse:           str
    latitude:          float | None = None
    longitude:         float | None = None
    timezone:          str | None = None  # v2.4.6 : "" ou IANA (Pacific/Tahiti)
    fournisseur_ia:    str | None = None
    cle_api_ia:        str | None = None
    modele_ia:         str | None = None
    url_base_ia:       str | None = None
    port:              int | None = None  # si None, premier port libre
    # v3.4 (h38h) — Plugins désactivés par le wizard (étape 5). Liste vide
    # par défaut → tous les plugins activés. Le wizard expose une checkbox
    # par plugin et envoie ici ceux que l'utilisateur a décochés.
    plugins_disabled:  list[str] | None = None


@router.post("/onboarding/create-instance", status_code=201)
def onboarding_create_instance(payload: WizardInstanceCreate, request: Request):
    """Crée une instance depuis le wizard — pré-remplit le premier slot libre
    (ou le port indiqué). Appelle ensuite manager.update_config()."""
    _check_admin(request)
    mgr = get_manager()

    # Validation : le sigle doit être présent, non-vide, et différent du default Site_NNNN
    sigle_clean = (payload.sigle or "").strip().upper()
    nom_clean = (payload.nom or "").strip()
    if not sigle_clean:
        raise HTTPException(400, "Le sigle est obligatoire.")
    if sigle_clean.startswith("SITE_"):
        raise HTTPException(400, "Le sigle ne peut pas commencer par 'Site_' (réservé aux slots vides).")
    if not nom_clean:
        raise HTTPException(400, "Le nom de l'établissement est obligatoire.")

    # Choix du port
    port = payload.port
    if port is None:
        # Premier slot vide (sigle commence par "Site_")
        for p, state in sorted(mgr.instances.items()):
            if not state.config.sigle or state.config.sigle.startswith("Site_"):
                port = p
                break
        if port is None:
            raise HTTPException(
                400,
                "Aucun slot libre. Supprimez une instance dans l'onglet Instances avant."
            )

    # Update config avec les données du wizard
    try:
        state = mgr.update_config(
            port,
            sigle=sigle_clean,
            nom=nom_clean,
            admin_login=(payload.admin_login or "dircrise").strip(),
            admin_password=payload.admin_password,
            # v3.4 (h38g) — propager le nom affiché choisi par l'utilisateur
            # dans le wizard. Sera utilisé comme User.display_name lors de
            # la création/update de l'admin dans la DB de l'instance.
            admin_display_name=(payload.nom_affiche_admin or "").strip(),
            adresse=(payload.adresse or "").strip(),
            latitude=payload.latitude,
            longitude=payload.longitude,
            timezone=(payload.timezone or "").strip(),
            # v3.4 (h38h) — Liste des plugins à désactiver à la création.
            # Le _bootstrap_db de l'instance écrira ces préférences dans
            # la table plugin_states avant le premier démarrage des plugins.
            plugins_disabled=list(payload.plugins_disabled or []),
            # v3.4 (h38k) — Langue par défaut de l'instance. Sera écrite dans
            # config.js, et scribe.js loadI18n() la chargera au boot.
            # Le wizard envoie un code ISO 2-letters parmi les 24 langues UE.
            langue=(payload.langue or "fr").strip()[:5] or "fr",
        )
    except ValueError as e:
        raise HTTPException(400, str(e))

    # Vérification post-condition : si après update le sigle reste "Site_NNNN",
    # c'est qu'on a un bug. On échoue de manière explicite plutôt que silencieuse.
    if state.config.sigle.startswith("Site_"):
        raise HTTPException(
            500,
            f"Erreur interne : le sigle '{state.config.sigle}' n'a pas été appliqué. "
            f"Contactez le support."
        )

    # v2.4.8 : marquer l'onboarding comme terminé automatiquement dès qu'une
    # instance non-démo a été configurée. Évite que le wizard se relance en
    # navigation privée quand l'utilisateur a déjà importé/configuré un
    # établissement (bug Polynésie : le client appelait /onboarding/finish
    # dans certains cas mais pas tous → flag manquant au redémarrage).
    # v2.4.8.4 : helper nettoie aussi wizard_force éventuel
    _mark_onboarding_done()

    return {
        "ok":   True,
        "port": port,
        "sigle": state.config.sigle,
        "nom":   state.config.nom,
        "url":   f"http://localhost:{port}",
    }


# ─────────────────────────────────────────────────────────────────────────────
# Import/Export Excel
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/onboarding/import-xlsx")
async def onboarding_import_xlsx(request: Request):
    """Upload d'un xlsx → parsing → preview JSON pour le wizard.
    v2.4.6 : le xlsx est aussi sauvegardé comme master/profil_base.xlsx pour
    qu'au lancement de l'instance, _bootstrap_db importe les UF/capacités
    directement (avant ce fix, il fallait re-uploader le xlsx dans
    "Profil de base" séparément — workflow cassé)."""
    _check_admin(request)
    from master.excel_import import parse_xlsx
    try:
        body = await request.body()
        if not body:
            raise HTTPException(400, "Fichier vide")
        cfg = parse_xlsx(body)
    except ValueError as e:
        raise HTTPException(400, f"Fichier invalide : {e}")
    except Exception as e:
        import traceback
        logger.error(f"Erreur parse xlsx : {traceback.format_exc()}")
        raise HTTPException(500, f"Erreur parsing : {e}")
    # v2.4.6 : sauvegarder aussi comme profil_base.xlsx pour les futurs bootstrap_db
    try:
        PROFIL_BASE_XLSX.parent.mkdir(parents=True, exist_ok=True)
        PROFIL_BASE_XLSX.write_bytes(body)
        logger.info(f"  Profil xlsx sauvegardé en {PROFIL_BASE_XLSX} ({len(body)} bytes)")
    except Exception as e:
        logger.warning(f"Sauvegarde profil_base.xlsx KO (non bloquant) : {e}")
    # Renvoyer un récap structuré
    first_site = cfg.sites[0] if cfg.sites else {}
    return {
        "nom":             cfg.nom,
        "sigle":           cfg.sigle,
        "finess":          cfg.finess,
        "langue":          cfg.langue,
        "admin_login":     cfg.login_admin,
        "admin_password":  cfg.mot_de_passe,
        "nom_affiche_admin": cfg.nom_affiche_admin,
        "adresse":         first_site.get("adresse", ""),
        "latitude":        first_site.get("latitude"),
        "longitude":       first_site.get("longitude"),
        "fournisseur_ia":  cfg.fournisseur_ia,
        "cle_api_ia":      cfg.cle_api_ia,
        "modele_ia":       cfg.modele_ia,
        "url_base_ia":     cfg.url_base_ia,
        "sites_count":     len(cfg.sites),
        "directeurs_count": len(cfg.directeurs),
        "telephonie_count": len(cfg.telephonie),
        "uf_count":        len(cfg.unites_fonctionnelles),
        "capacite_count":  len(cfg.services_capacite),
        "warnings":        cfg.warnings,
        "profil_saved":    PROFIL_BASE_XLSX.exists(),
    }


@router.get("/instances/{port}/export-xlsx")
def export_instance_xlsx(
    port: int, request: Request, include_secrets: bool = False
):
    """Exporte la config d'une instance au format xlsx.
    Par défaut hybride (mdp hashé, clés API vidées).
    include_secrets=true → tout en clair."""
    _check_admin(request)
    from master.excel_export import export_instance_to_xlsx
    from fastapi.responses import Response

    mgr = get_manager()
    if port not in mgr.instances:
        raise HTTPException(404, f"Port {port} inconnu")
    state = mgr.instances[port]
    cfg = state.config

    # Lecture des données DB si l'instance a déjà tourné
    db_data = {}
    if state.db_path and pathlib.Path(state.db_path).exists():
        try:
            db_data = _read_instance_db_data(state.db_path)
        except Exception as e:
            logger.warning(f"Lecture DB pour export {cfg.sigle} : {e}")

    config_dict = {
        "sigle": cfg.sigle, "nom": cfg.nom,
        "admin_login": cfg.admin_login, "admin_password": cfg.admin_password,
        "adresse": cfg.adresse,
        "latitude": cfg.latitude, "longitude": cfg.longitude,
        "fed_token": state.fed_token,
    }
    try:
        xlsx_bytes = export_instance_to_xlsx(
            config_dict, db_data, include_secrets=include_secrets,
        )
    except Exception as e:
        import traceback
        logger.error(f"Erreur export xlsx port {port} : {traceback.format_exc()}")
        raise HTTPException(500, f"Erreur export : {e}")

    safe_sigle = "".join(c if c.isalnum() else "_" for c in (cfg.sigle or f"port{port}"))
    filename = f"{safe_sigle}_export.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/instances/export-all-zip")
def export_all_instances_zip(request: Request, include_secrets: bool = False):
    """Exporte TOUTES les instances configurées en un seul zip (1 xlsx par instance).
    Pour le déploiement d'un GHT complet sur un autre serveur."""
    _check_admin(request)
    from master.excel_export import export_ght_to_zip
    from fastapi.responses import Response

    mgr = get_manager()
    instances_configs = []
    instances_db_data = {}
    for port, state in sorted(mgr.instances.items()):
        cfg = state.config
        # Skip slots vides
        if not cfg.sigle or cfg.sigle.startswith("Site_"):
            continue
        instances_configs.append({
            "port": port,
            "sigle": cfg.sigle, "nom": cfg.nom,
            "admin_login": cfg.admin_login, "admin_password": cfg.admin_password,
            "adresse": cfg.adresse,
            "latitude": cfg.latitude, "longitude": cfg.longitude,
            "fed_token": state.fed_token,
        })
        if state.db_path and pathlib.Path(state.db_path).exists():
            try:
                instances_db_data[cfg.sigle] = _read_instance_db_data(state.db_path)
            except Exception as e:
                logger.warning(f"Lecture DB pour export GHT {cfg.sigle} : {e}")

    if not instances_configs:
        raise HTTPException(400, "Aucune instance configurée à exporter")

    try:
        zip_bytes = export_ght_to_zip(
            instances_configs, instances_db_data, include_secrets=include_secrets,
        )
    except Exception as e:
        import traceback
        logger.error(f"Erreur export GHT zip : {traceback.format_exc()}")
        raise HTTPException(500, f"Erreur export : {e}")

    timestamp = now_iso()[:10]  # YYYY-MM-DD
    filename = f"scribe_ght_export_{timestamp}.zip"
    return Response(
        content=zip_bytes,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


from pydantic import BaseModel as _BaseModel
class _EncBackupIn(_BaseModel):
    password: str
    include_secrets: bool = True


@router.post("/instances/export-all-zip-encrypted")
def export_all_instances_zip_encrypted(body: _EncBackupIn, request: Request):
    """Backup GHT CHIFFRE (AES) : meme contenu que export-all-zip, dans une
    archive ZIP protegee par mot de passe (pyzipper). Le mot de passe transite
    dans le corps POST, jamais en query (pas de fuite dans les logs)."""
    _check_admin(request)
    if not body.password or len(body.password) < 4:
        raise HTTPException(400, "Mot de passe trop court (4 caracteres minimum).")
    from master.excel_export import export_ght_to_zip
    from fastapi.responses import Response
    mgr = get_manager()
    instances_configs = []
    instances_db_data = {}
    for port, state in sorted(mgr.instances.items()):
        cfg = state.config
        if not cfg.sigle or cfg.sigle.startswith("Site_"):
            continue
        instances_configs.append({
            "port": port, "sigle": cfg.sigle, "nom": cfg.nom,
            "admin_login": cfg.admin_login, "admin_password": cfg.admin_password,
            "adresse": cfg.adresse, "latitude": cfg.latitude, "longitude": cfg.longitude,
            "fed_token": state.fed_token,
        })
        if state.db_path and pathlib.Path(state.db_path).exists():
            try:
                instances_db_data[cfg.sigle] = _read_instance_db_data(state.db_path)
            except Exception as e:
                logger.warning(f"Lecture DB pour backup GHT {cfg.sigle} : {e}")
    if not instances_configs:
        raise HTTPException(400, "Aucune instance configuree a sauvegarder")
    try:
        inner_zip = export_ght_to_zip(
            instances_configs, instances_db_data, include_secrets=body.include_secrets,
        )
    except Exception as e:
        import traceback
        logger.error(f"Erreur backup GHT zip : {traceback.format_exc()}")
        raise HTTPException(500, f"Erreur backup : {e}")
    try:
        import pyzipper, io as _io
        ts = now_iso()[:10]
        buf = _io.BytesIO()
        with pyzipper.AESZipFile(buf, "w", compression=pyzipper.ZIP_DEFLATED,
                                 encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(body.password.encode("utf-8"))
            zf.writestr(f"scribe_ght_export_{ts}.zip", inner_zip)
        enc_bytes = buf.getvalue()
    except Exception as e:
        import traceback
        logger.error(f"Erreur chiffrement backup : {traceback.format_exc()}")
        raise HTTPException(500, f"Erreur chiffrement (pyzipper installe ?) : {e}")
    filename = f"scribe_ght_backup_{now_iso()[:10]}.zip"
    return Response(
        content=enc_bytes, media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _load_central_config():
    """Charge la config serveur (notifications SMTP/SMS) en CLAIR, ou {}."""
    try:
        import sys as _sys, os as _os
        _cdir = _os.path.join(PROJECT_ROOT, "collecteur")
        if _cdir not in _sys.path:
            _sys.path.insert(0, _cdir)
        import central_config_store as _ccs
        return _ccs.load_clear() or {}
    except Exception as e:
        logger.warning(f"Config serveur (lecture) indisponible : {e}")
        return {}


def _save_central_config(server_config):
    """Re-chiffre et persiste la config serveur restauree. Renvoie le nb de canaux."""
    if not server_config:
        return 0
    try:
        import sys as _sys, os as _os
        _cdir = _os.path.join(PROJECT_ROOT, "collecteur")
        if _cdir not in _sys.path:
            _sys.path.insert(0, _cdir)
        import central_config_store as _ccs
    except Exception as e:
        logger.warning(f"Config serveur (module) indisponible : {e}")
        return 0
    n = 0
    for domain, fields in server_config.items():
        if isinstance(fields, dict):
            try:
                _ccs.save(domain, fields, updated_by="restore")
                n += 1
            except Exception as e:
                logger.warning(f"Restauration config serveur {domain} : {e}")
    return n


_SERVER_FILE_PATTERNS = ["collecteur/collecteur_*.json", "master/profil_base.xlsx"]
_SERVER_DBS = {"collecteur_messagerie": "collecteur/collecteur_messagerie.db"}


def _proot():
    return os.path.abspath(str(PROJECT_ROOT))


def _safe_under_root(rel):
    dest = os.path.abspath(os.path.join(_proot(), rel))
    return dest if dest.startswith(_proot() + os.sep) else None


def _collect_server_files():
    """Tout l'etat serveur en fichiers (comptes supervision, tokens, donnees
    collecteur, profil UF), encode base64."""
    import os, base64, glob
    files = {}
    for pat in _SERVER_FILE_PATTERNS:
        for fp in glob.glob(os.path.join(_proot(), pat)):
            try:
                rel = os.path.relpath(fp, _proot())
                with open(fp, "rb") as fh:
                    files[rel] = base64.b64encode(fh.read()).decode("ascii")
            except Exception as e:
                logger.warning(f"Backup fichier serveur {fp} : {e}")
    return files


def _collect_server_dbs():
    """Bases SQLite serveur (messagerie du collecteur)."""
    import os
    from app.backup import dump_database
    dbs = {}
    for name, rel in _SERVER_DBS.items():
        path = os.path.join(_proot(), rel)
        if os.path.isfile(path):
            try:
                dbs[name] = {"path": rel, "dump": dump_database(path)}
            except Exception as e:
                logger.warning(f"Backup base serveur {name} : {e}")
    return dbs


def _restore_server_files(files):
    import os, base64
    n = 0
    for rel, b64 in (files or {}).items():
        dest = _safe_under_root(rel)
        if not dest:
            logger.warning(f"Chemin serveur refuse (hors racine) : {rel}")
            continue
        try:
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            with open(dest, "wb") as fh:
                fh.write(base64.b64decode(b64))
            n += 1
        except Exception as e:
            logger.warning(f"Restauration fichier serveur {rel} : {e}")
    return n


def _restore_server_dbs(dbs):
    import os
    from app.backup import restore_database
    n = 0
    for name, entry in (dbs or {}).items():
        rel = (entry or {}).get("path")
        dump = (entry or {}).get("dump")
        dest = _safe_under_root(rel) if rel else None
        if dest and dump:
            try:
                os.makedirs(os.path.dirname(dest), exist_ok=True)
                restore_database(dest, dump)
                n += 1
            except Exception as e:
                logger.warning(f"Restauration base serveur {name} : {e}")
    return n


# ── h144 — Sauvegarde INTEGRALE : arborescences de fichiers ──────────────────
# Dossiers serveur partages contenant du contenu lie aux instances mais STOCKE
# HORS scribe.db : pieces jointes (incidents + messagerie) et archives ZIP de
# rapports de crise. Sans eux, la sauvegarde n'est pas une image complete.
_SERVER_TREE_DIRS = ["uploads", "archives"]
# Fichiers a NE PAS embarquer depuis un dossier d'instance (la base est dumpee
# a part en JSON portable ; les -wal/-shm/-journal sont volatils ; les logs
# n'ont pas a etre restaures).
_INSTANCE_SKIP = {"scribe.db", "scribe.db-wal", "scribe.db-shm", "scribe.db-journal"}
_INSTANCE_SKIP_SUFFIX = (".log",)
# Garde-fou taille (octets) par fichier embarque, pour eviter d'embarquer par
# erreur un fichier geant ; 0 = pas de limite.
_MAX_BACKUP_FILE = int(os.environ.get("SCRIBE_BACKUP_MAX_FILE", str(200 * 1024 * 1024)))


def _collect_dir_tree(abs_dir):
    """Tous les fichiers sous abs_dir (recursif) -> {relpath: b64}. Vide si absent."""
    import os, base64
    out = {}
    if not abs_dir or not os.path.isdir(abs_dir):
        return out
    for root, _dirs, names in os.walk(abs_dir):
        for nm in names:
            fp = os.path.join(root, nm)
            try:
                if _MAX_BACKUP_FILE and os.path.getsize(fp) > _MAX_BACKUP_FILE:
                    logger.warning(f"Backup : fichier ignore (trop gros) {fp}")
                    continue
                rel = os.path.relpath(fp, abs_dir)
                with open(fp, "rb") as fh:
                    out[rel] = base64.b64encode(fh.read()).decode("ascii")
            except Exception as e:
                logger.warning(f"Backup arborescence {fp} : {e}")
    return out


def _restore_dir_tree(abs_dir, tree):
    """Re-ecrit {relpath: b64} sous abs_dir. Renvoie le nb de fichiers ecrits."""
    import os, base64
    if not tree or not abs_dir:
        return 0
    base = os.path.abspath(abs_dir)
    n = 0
    for rel, b64 in tree.items():
        dest = os.path.abspath(os.path.join(base, rel))
        # Anti path-traversal : la cible doit rester sous abs_dir
        if not dest.startswith(base + os.sep):
            logger.warning(f"Restauration arborescence refusee (hors dossier) : {rel}")
            continue
        try:
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            with open(dest, "wb") as fh:
                fh.write(base64.b64decode(b64))
            n += 1
        except Exception as e:
            logger.warning(f"Restauration fichier {rel} : {e}")
    return n


def _collect_instance_files(inst_dir):
    """Tous les fichiers du dossier d'instance SAUF la base (dumpee a part),
    ses annexes SQLite volatiles et les logs. -> {relpath: b64}."""
    import os, base64
    out = {}
    if not inst_dir or not os.path.isdir(inst_dir):
        return out
    for root, _dirs, names in os.walk(inst_dir):
        for nm in names:
            if nm in _INSTANCE_SKIP or nm.endswith(_INSTANCE_SKIP_SUFFIX):
                continue
            fp = os.path.join(root, nm)
            try:
                if _MAX_BACKUP_FILE and os.path.getsize(fp) > _MAX_BACKUP_FILE:
                    logger.warning(f"Backup : fichier instance ignore (trop gros) {fp}")
                    continue
                rel = os.path.relpath(fp, inst_dir)
                with open(fp, "rb") as fh:
                    out[rel] = base64.b64encode(fh.read()).decode("ascii")
            except Exception as e:
                logger.warning(f"Backup fichier instance {fp} : {e}")
    return out


def _collect_server_trees():
    """Arborescences serveur partagees (uploads, archives) -> {dir: {rel: b64}}.
    Honore SCRIBE_UPLOADS_DIR si defini hors racine projet."""
    import os
    trees = {}
    seen = set()
    for d in _SERVER_TREE_DIRS:
        abs_d = os.path.join(_proot(), d)
        tree = _collect_dir_tree(abs_d)
        if tree:
            trees[d] = tree
            seen.add(os.path.abspath(abs_d))
    # uploads redirige par variable d'environnement (messagerie)
    up_env = os.environ.get("SCRIBE_UPLOADS_DIR")
    if up_env:
        abs_up = os.path.abspath(up_env)
        if abs_up not in seen:
            tree = _collect_dir_tree(abs_up)
            if tree:
                trees["uploads"] = {**trees.get("uploads", {}), **tree}
    return trees


def _restore_server_trees(trees):
    """Restaure les arborescences serveur. Renvoie {dir: nb_fichiers}."""
    import os
    res = {}
    for d, tree in (trees or {}).items():
        # On restaure uniquement vers les dossiers connus, sous la racine projet
        if d not in _SERVER_TREE_DIRS:
            logger.warning(f"Arborescence serveur inconnue ignoree : {d}")
            continue
        abs_d = os.path.join(_proot(), d)
        res[d] = _restore_dir_tree(abs_d, tree)
    return res


@router.post("/instances/backup-full")
def backup_full(body: _EncBackupIn, request: Request):
    """Sauvegarde INTEGRALE plug-and-play (chiffree AES) : pour chaque instance,
    le contenu COMPLET de sa base SQLite + ses fichiers de config. Capture tout,
    contrairement a l'export xlsx partiel."""
    _check_admin(request)
    if not body.password or len(body.password) < 4:
        raise HTTPException(400, "Mot de passe trop court (4 caracteres minimum).")
    try:
        from app.backup import dump_database, make_encrypted_zip
    except Exception as e:
        raise HTTPException(500, f"Moteur de sauvegarde indisponible : {e}")
    from fastapi.responses import Response
    from .instances_manager import _safe_path_segment, DATA_DIR
    mgr = get_manager()
    instances = {}
    summary = {}
    for port, state in sorted(mgr.instances.items()):
        cfg = state.config
        if not cfg.sigle:
            continue
        # Resoudre le chemin de base meme si state.db_path n'est pas renseigne
        # (master redemarre, instance jamais relancee). On ne filtre PLUS sur le
        # nom du sigle : toute instance ayant une base reelle est sauvegardee.
        db_path = state.db_path or str(
            DATA_DIR / _safe_path_segment(cfg.sigle, fallback=f"instance_{port}") / "scribe.db")
        if not pathlib.Path(db_path).exists():
            continue  # pas de base => rien a sauvegarder pour cette instance
        entry = {
            "config": {
                "port": port, "sigle": cfg.sigle, "nom": cfg.nom,
                "admin_login": cfg.admin_login, "admin_password": cfg.admin_password,
                "adresse": cfg.adresse, "latitude": cfg.latitude,
                "longitude": cfg.longitude, "fed_token": state.fed_token,
            },
            "config_files": {},
            "instance_files": {},
            "database": None,
        }
        try:
            entry["database"] = dump_database(db_path)
        except Exception as e:
            logger.warning(f"Dump DB backup {cfg.sigle} : {e}")
        summary[cfg.sigle] = sum(len(v) for v in entry["database"]["data"].values()) if entry["database"] else 0
        inst_dir = pathlib.Path(db_path).parent
        for fn in ("config.xml", "config.js"):
            fp = inst_dir / fn
            if fp.exists():
                try:
                    entry["config_files"][fn] = fp.read_text(encoding="utf-8")
                except Exception:
                    pass
        # h144 — image COMPLETE du dossier d'instance : tout fichier present
        # (secrets de plugins, DB annexes, clefs, etc.) hors base/aux/logs.
        try:
            entry["instance_files"] = _collect_instance_files(str(inst_dir))
        except Exception as e:
            logger.warning(f"Backup fichiers instance {cfg.sigle} : {e}")
        instances[cfg.sigle] = entry
    server_config = _load_central_config()
    server_files = _collect_server_files()
    server_dbs = _collect_server_dbs()
    server_trees = _collect_server_trees()
    if not instances and not server_config and not server_files and not server_trees:
        raise HTTPException(400, "Aucune donnee a sauvegarder.")
    payload = {
        "manifest": {
            "scribe_version": "4.0.0", "kind": "ght-full-backup",
            "created_at": now_iso(), "instances": list(instances.keys()),
            "rows": summary,
            "instance_files": {s: len(i.get("instance_files") or {}) for s, i in instances.items()},
            "has_server_config": bool(server_config),
            "server_files": sorted(server_files.keys()),
            "server_dbs": sorted(server_dbs.keys()),
            "server_trees": {d: len(t) for d, t in server_trees.items()},
        },
        "instances": instances,
        "server_config": server_config,
        "server_files": server_files,
        "server_dbs": server_dbs,
        "server_trees": server_trees,
    }
    try:
        enc = make_encrypted_zip(payload, body.password)
    except Exception as e:
        import traceback
        logger.error(f"Erreur backup integral : {traceback.format_exc()}")
        raise HTTPException(500, f"Erreur chiffrement (pyzipper installe ?) : {e}")
    fn = f"scribe_backup_complet_{now_iso()[:10]}.zip"
    summary_str = "; ".join(f"{k}:{v}" for k, v in summary.items())
    return Response(content=enc, media_type="application/zip",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"',
                             "X-Scribe-Backup-Summary": summary_str,
                             "Access-Control-Expose-Headers": "X-Scribe-Backup-Summary"})


@router.post("/instances/restore-full")
async def restore_full(request: Request, file: UploadFile = File(...), password: str = Form(...)):
    """Restauration plug-and-play : dechiffre l'archive et recree chaque instance
    (config + base complete + fichiers de config). Instances cibles : ARRETEES."""
    _check_admin(request)
    if not password or len(password) < 4:
        raise HTTPException(400, "Mot de passe requis.")
    try:
        from app.backup import read_encrypted_zip, restore_database
    except Exception as e:
        raise HTTPException(500, f"Moteur de restauration indisponible : {e}")
    from .instances_manager import _safe_path_segment, DATA_DIR, InstanceState
    raw = await file.read()
    try:
        payload = read_encrypted_zip(raw, password)
    except Exception:
        raise HTTPException(400, "Mot de passe incorrect ou archive illisible.")
    insts = payload.get("instances")
    if not insts:
        raise HTTPException(400, "Archive invalide : aucune instance trouvee.")
    mgr = get_manager()
    running = [int(i["config"]["port"]) for i in insts.values()
               if mgr.instances.get(int(i["config"]["port"]))
               and getattr(mgr.instances[int(i["config"]["port"])], "statut", "") == "actif"]
    if running:
        raise HTTPException(409, f"Arretez d'abord les instances actives (ports {running}) avant de restaurer.")
    restored = []
    for sigle, inst in insts.items():
        c = inst["config"]; port = int(c["port"])
        cfg = InstanceConfig(
            port=port, sigle=c.get("sigle", ""), nom=c.get("nom", ""),
            admin_login=c.get("admin_login", "dircrise"),
            admin_password=c.get("admin_password", ""),
            adresse=c.get("adresse", ""),
            latitude=c.get("latitude"), longitude=c.get("longitude"),
        )
        state = mgr.instances.get(port)
        if state is None:
            state = InstanceState(config=cfg)
            mgr.instances[port] = state
        else:
            state.config = cfg
        if c.get("fed_token"):
            state.fed_token = c["fed_token"]
        # Chemin canonique = celui que start() recalculera au lancement
        seg = _safe_path_segment(cfg.sigle, fallback=f"instance_{port}")
        inst_dir = DATA_DIR / seg
        inst_dir.mkdir(parents=True, exist_ok=True)
        target = str(inst_dir / "scribe.db")
        # Purger d'eventuels fichiers SQLite annexes (-wal/-shm/-journal) perimes
        # de l'ancienne session : sinon un WAL non checkpointe peut masquer la
        # base restauree et l'instance semble "vide".
        for suf in ("-wal", "-shm", "-journal"):
            aux = pathlib.Path(target + suf)
            if aux.exists():
                try: aux.unlink()
                except Exception: pass
        state.db_path = target
        state.log_path = str(inst_dir / "scribe.log")
        applied_rows = 0
        if inst.get("database"):
            try:
                restore_database(target, inst["database"])
            except Exception as e:
                logger.error(f"Restauration DB {sigle} : {e}")
                raise HTTPException(500, f"Erreur restauration base {sigle} : {e}")
            # Verification REELLE sur disque (et pas le nombre attendu)
            try:
                import sqlite3 as _sq
                _c = _sq.connect(target)
                applied_rows = sum(
                    _c.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
                    for t in inst["database"].get("order", []))
                _c.close()
            except Exception as e:
                logger.warning(f"Verif post-restore {sigle} : {e}")
                applied_rows = sum(len(v) for v in inst["database"]["data"].values())
            # Re-purger le WAL cree par la restauration (checkpoint propre)
            for suf in ("-wal", "-shm", "-journal"):
                aux = pathlib.Path(target + suf)
                if aux.exists():
                    try: aux.unlink()
                    except Exception: pass
        for fn, content in (inst.get("config_files") or {}).items():
            try:
                (inst_dir / fn).write_text(content, encoding="utf-8")
            except Exception as e:
                logger.warning(f"Ecriture config {fn} ({sigle}) : {e}")
        # h144 — restaurer l'image complete du dossier d'instance (secrets de
        # plugins, DB annexes, etc.). Ecrit APRES config_files : fait foi.
        n_inst_files = 0
        try:
            n_inst_files = _restore_dir_tree(str(inst_dir), inst.get("instance_files") or {})
        except Exception as e:
            logger.warning(f"Restauration fichiers instance {sigle} : {e}")
        restored.append({"sigle": cfg.sigle, "port": port, "rows": applied_rows,
                         "files": n_inst_files})
    n_channels = _save_central_config(payload.get("server_config") or {})
    n_files = _restore_server_files(payload.get("server_files") or {})
    n_dbs = _restore_server_dbs(payload.get("server_dbs") or {})
    n_trees = _restore_server_trees(payload.get("server_trees") or {})
    mgr._save_state()
    return {"ok": True, "count": len(restored), "restored": restored,
            "server_channels": n_channels, "server_files": n_files,
            "server_dbs": n_dbs, "server_trees": n_trees}


@router.post("/instances/{port}/import-xlsx")
async def import_instance_xlsx(port: int, request: Request):
    """Importe un xlsx pour mettre à jour la config d'une instance.
    L'instance DOIT être arrêtée (refusé sinon).
    Écrase la config actuelle ET les UF/services en DB."""
    _check_admin(request)
    from master.excel_import import parse_xlsx

    mgr = get_manager()
    if port not in mgr.instances:
        raise HTTPException(404, f"Port {port} inconnu")
    state = mgr.instances[port]

    # SAFE-GUARD : refus si instance active
    if state.statut == "actif" and state.pid:
        raise HTTPException(
            400,
            f"L'instance {state.config.sigle} est active. "
            f"Arrêtez-la avant d'importer une nouvelle config."
        )

    try:
        body = await request.body()
        if not body:
            raise HTTPException(400, "Fichier vide")
        cfg = parse_xlsx(body)
    except ValueError as e:
        raise HTTPException(400, f"Fichier invalide : {e}")
    except Exception as e:
        import traceback
        logger.error(f"Erreur parse xlsx import : {traceback.format_exc()}")
        raise HTTPException(500, f"Erreur parsing : {e}")

    # Update config
    first_site = cfg.sites[0] if cfg.sites else {}
    try:
        mgr.update_config(
            port,
            sigle=cfg.sigle,
            nom=cfg.nom,
            admin_login=cfg.login_admin,
            adresse=first_site.get("adresse", ""),
            latitude=first_site.get("latitude"),
            longitude=first_site.get("longitude"),
        )
        # Mot de passe : ne mettre à jour QUE si présent en clair (pas un hash)
        if cfg.mot_de_passe and not cfg.mot_de_passe.startswith(("$2", "sha256:")):
            mgr.update_config(port, admin_password=cfg.mot_de_passe)
    except ValueError as e:
        raise HTTPException(400, str(e))

    # On stocke les UF/capa pour bootstrap au prochain start (state side-channel)
    # On les attache au state — sera consommé par _bootstrap_db si présent
    setattr(state, "_pending_uf",   cfg.unites_fonctionnelles)
    setattr(state, "_pending_capa", cfg.services_capacite)

    # v2.4.8 : marquer l'onboarding comme terminé (couvre le chemin
    # "import xlsx dans une instance existante" qui peut être le 1er geste
    # d'un utilisateur ouvrant SCRIBE). v2.4.8.4 : helper nettoie aussi wizard_force
    if cfg.sigle and not cfg.sigle.startswith("Site_"):
        _mark_onboarding_done()

    return {
        "ok":           True,
        "port":         port,
        "sigle":        cfg.sigle,
        "uf_imported":  len(cfg.unites_fonctionnelles),
        "capa_imported": len(cfg.services_capacite),
        "next":         "Cliquez ▶ Lancer pour démarrer l'instance avec la nouvelle config",
    }


@router.post("/onboarding/import-ght-zip")
async def import_ght_zip(request: Request):
    """Importe un zip GHT (export complet) — preview des établissements détectés.
    Ne crée PAS encore les instances ; renvoie un preview pour confirmation."""
    _check_admin(request)
    import zipfile
    import io as _io
    from master.excel_import import parse_xlsx

    body = await request.body()
    if not body:
        raise HTTPException(400, "Fichier vide")
    try:
        with zipfile.ZipFile(_io.BytesIO(body), "r") as zf:
            names = zf.namelist()
            xlsx_files = [n for n in names if n.lower().endswith(".xlsx")]
            if not xlsx_files:
                if any(n in ("backup.json", "manifest.json") for n in names):
                    raise HTTPException(400,
                        "Archive de sauvegarde complete chiffree detectee. Pour la restaurer, "
                        "utilisez le bouton « Restaurer » du panneau Instances (avec son mot de "
                        "passe) — l'import GHT n'accepte que les exports xlsx.")
                raise HTTPException(400, "Aucun .xlsx dans le zip")
            previews = []
            for fname in xlsx_files:
                try:
                    cfg = parse_xlsx(zf.read(fname))
                    previews.append({
                        "filename":      fname,
                        "sigle":         cfg.sigle,
                        "nom":           cfg.nom,
                        "uf_count":      len(cfg.unites_fonctionnelles),
                        "capa_count":    len(cfg.services_capacite),
                        "directeurs_count": len(cfg.directeurs),
                    })
                except Exception as e:
                    previews.append({"filename": fname, "error": str(e)})
    except zipfile.BadZipFile:
        raise HTTPException(400, "Zip invalide")
    return {"establishments": previews}


# Helper interne pour lire les données d'une DB d'instance
def _read_instance_db_data(db_path: str) -> dict[str, list]:
    """Lit les UF, capacités, directeurs, telephonie d'une DB d'instance.
    Best-effort : si une table manque on retourne juste vide."""
    import sqlite3
    out: dict[str, list] = {"uf": [], "capa": [], "directeurs": [], "telephonie": []}
    try:
        conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
        cur = conn.cursor()
        # UF
        try:
            cur.execute("SELECT code, libelle, pole, site, actif FROM unites_fonctionnelles ORDER BY code")
            for r in cur.fetchall():
                out["uf"].append({
                    "code": r[0], "libelle": r[1], "pole": r[2],
                    "site": r[3] or "", "actif": bool(r[4]),
                })
        except sqlite3.OperationalError:
            pass
        # Capacité
        try:
            cur.execute("""SELECT service, code_uf, pole, site, capacite_totale,
                                  seuil_t1, seuil_t2, accepte_h, accepte_f, accepte_i,
                                  tel_cadre, ordre
                           FROM capacite_referentiel ORDER BY ordre""")
            for r in cur.fetchall():
                out["capa"].append({
                    "service": r[0], "code_uf": r[1], "pole": r[2], "site": r[3],
                    "capacite": r[4], "seuil_t1": r[5], "seuil_t2": r[6],
                    "accepte_h": bool(r[7]), "accepte_f": bool(r[8]), "accepte_i": bool(r[9]),
                    "tel_cadre": r[10] or "", "ordre": r[11] or 0,
                })
        except sqlite3.OperationalError:
            pass
        conn.close()
    except Exception as e:
        logger.warning(f"Lecture DB {db_path} : {e}")
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Lifecycle — stop_all à l'arrêt du collecteur
# ─────────────────────────────────────────────────────────────────────────────

def lifecycle_register(app):
    """À appeler depuis collecteur.py pour stopper toutes les instances filles
    (mode prod ET mode exercice + collecteur exercice) quand le master est
    arrêté (Ctrl+C)."""
    @app.on_event("shutdown")
    def _on_shutdown():
        # Mode prod
        try:
            count = get_manager().stop_all()
            if count:
                logger.info(f"Master : {count} instance(s) arrêtée(s) à l'extinction")
        except Exception as e:
            logger.warning(f"Master shutdown prod : {e}")
        # Mode exercice (instances + collecteur :8565)
        try:
            count_exo = get_exercice_manager().stop_all()
            if count_exo:
                logger.info(f"Master : {count_exo} instance(s) exercice arrêtée(s)")
        except Exception as e:
            logger.warning(f"Master shutdown exercice : {e}")


# ─────────────────────────────────────────────────────────────────────────────
# v3000h17 — Configuration du hostname externe
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/setup-hostname")
def get_hostname_setup(request: Request):
    """Retourne l'état actuel du hostname + suggestions."""
    from master.hostname_config import (
        get_configured_hostname, suggest_hostnames, get_external_host,
    )
    return {
        "configured": get_configured_hostname(),
        "effective":  get_external_host(request=request, fallback="localhost"),
        "suggestions": suggest_hostnames(),
        "current_request_host": (request.headers.get("host") or "").split(":")[0],
    }


class HostnameUpdate(BaseModel):
    hostname: str


@router.post("/setup-hostname")
def set_hostname_setup(payload: HostnameUpdate, request: Request):
    """Configure le hostname externe (écrit hostname.conf à la racine)."""
    _check_admin(request)
    from master.hostname_config import set_configured_hostname
    try:
        set_configured_hostname(payload.hostname)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "hostname": payload.hostname,
            "message": "Hostname configuré. Redémarrer les instances pour appliquer."}


@router.get("/hostname-page", response_class=HTMLResponse)
def hostname_page(request: Request):
    """Page de configuration du hostname (HTML standalone, accessible sans auth)."""
    page_path = pathlib.Path(__file__).parent / "setup_hostname.html"
    if not page_path.exists():
        return HTMLResponse("<p>setup_hostname.html introuvable</p>", status_code=500)
    return HTMLResponse(page_path.read_text(encoding="utf-8"))
