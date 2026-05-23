"""
master/excel_import.py — Lecture du SCRIBE_config_etablissement.xlsx pour le master
==================================================================================
Wrapper léger autour de la logique de import_config_xlsx.py (qui lui modifie
directement scribe.db). Ici on EXTRAIT seulement les données structurées du
xlsx pour les passer au master, qui décide ensuite quoi en faire (créer un
slot, peupler la DB d'une instance précise, etc.).

Format attendu : 5 onglets ETABLISSEMENT, DIRECTEURS, TELEPHONIE,
UF_INCIDENTS, SERVICES_CAPACITE — voir SCRIBE_config_etablissement.xlsx.
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from typing import Any

logger = logging.getLogger("scribe.master.excel_import")


@dataclass
class ParsedConfig:
    """Résultat du parsing d'un fichier xlsx établissement."""
    # Onglet ETABLISSEMENT
    nom: str = ""
    sigle: str = ""
    finess: str = ""
    langue: str = "fr"
    login_admin: str = "dircrise"
    mot_de_passe: str = ""
    nom_affiche_admin: str = "Directeur de Crise"
    fournisseur_ia: str = ""
    cle_api_ia: str = ""
    modele_ia: str = ""
    url_base_ia: str = ""
    federation_active: bool = False
    collecteur_url: str = ""
    collecteur_token: str = ""
    sync_crise: bool = False
    sync_sanitaire: bool = False
    # Liste des sites (lat/lon)
    sites: list[dict[str, Any]] = field(default_factory=list)
    # Onglet DIRECTEURS
    directeurs: list[dict[str, str]] = field(default_factory=list)
    # Onglet TELEPHONIE
    telephonie: list[dict[str, str]] = field(default_factory=list)
    # Onglet UF_INCIDENTS
    unites_fonctionnelles: list[dict[str, Any]] = field(default_factory=list)
    # Onglet SERVICES_CAPACITE
    services_capacite: list[dict[str, Any]] = field(default_factory=list)
    # Erreurs/avertissements
    warnings: list[str] = field(default_factory=list)


def parse_xlsx(file_bytes: bytes) -> ParsedConfig:
    """Parse un fichier xlsx (en bytes) et retourne une ParsedConfig.

    Lève ValueError si le fichier est invalide (onglets manquants, etc.).
    """
    try:
        import openpyxl
    except ImportError:
        raise ValueError("openpyxl n'est pas installé. Lancez : pip install openpyxl")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Fichier xlsx invalide : {e}") from e

    cfg = ParsedConfig()
    expected_sheets = {
        "ETABLISSEMENT", "DIRECTEURS", "TELEPHONIE",
        "UF_INCIDENTS", "SERVICES_CAPACITE",
    }
    found_sheets = set(wb.sheetnames)
    missing = expected_sheets - found_sheets
    if missing:
        raise ValueError(
            f"Onglets manquants dans le xlsx : {', '.join(sorted(missing))}. "
            f"Téléchargez le template depuis le master."
        )

    # ── ETABLISSEMENT ────────────────────────────────────────────────────────
    ws = wb["ETABLISSEMENT"]
    params: dict[str, str] = {}
    in_sites = False
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        # Détection passage à la section sites
        low = key.lower().replace("é", "e")
        if "site" in low and ("principal" in low or "geographique" in low or "📍" in key):
            in_sites = True
            continue
        if in_sites:
            # Ligne site : nom, adresse, lat, lon, téléphone
            if row[0] and row[2]:  # a un nom et une lat
                try:
                    lat = float(str(row[2]).strip())
                    lon = float(str(row[3]).strip()) if row[3] else None
                    cfg.sites.append({
                        "nom":       str(row[0]).strip(),
                        "adresse":   str(row[1] or "").strip(),
                        "latitude":  lat,
                        "longitude": lon,
                        "telephone_garde": str(row[4] or "").strip() if len(row) > 4 else "",
                    })
                except (ValueError, TypeError):
                    cfg.warnings.append(f"Site '{row[0]}' : lat/lon invalide, ignoré")
        else:
            # Ligne paramètre clé/valeur
            if row[1] is not None and str(row[1]).strip():
                params[key] = str(row[1]).strip()

    cfg.nom            = params.get("NOM_ETABLISSEMENT", "")
    cfg.sigle          = params.get("SIGLE", "").upper()
    cfg.finess         = params.get("FINESS", "")
    cfg.langue         = params.get("LANGUE", "fr").lower()[:2]
    cfg.login_admin    = params.get("LOGIN_ADMIN", "dircrise")
    cfg.mot_de_passe   = params.get("MOT_DE_PASSE", "")
    cfg.nom_affiche_admin = params.get("NOM_AFFICHE_ADMIN", "Directeur de Crise")
    cfg.fournisseur_ia = params.get("FOURNISSEUR_IA", "").lower()
    cfg.cle_api_ia     = params.get("CLE_API_IA", "")
    cfg.modele_ia      = params.get("MODELE_IA", "")
    cfg.url_base_ia    = params.get("URL_BASE_IA", "")
    cfg.federation_active = params.get("FEDERATION_ACTIVE", "").lower() == "true"
    cfg.collecteur_url    = params.get("COLLECTEUR_URL", "")
    cfg.collecteur_token  = params.get("COLLECTEUR_TOKEN", "")
    cfg.sync_crise        = params.get("SYNC_CRISE", "").lower() == "true"
    cfg.sync_sanitaire    = params.get("SYNC_SANITAIRE", "").lower() == "true"

    if not cfg.sigle:
        raise ValueError("SIGLE obligatoire dans l'onglet ETABLISSEMENT")
    if not cfg.nom:
        raise ValueError("NOM_ETABLISSEMENT obligatoire dans l'onglet ETABLISSEMENT")

    # ── DIRECTEURS ───────────────────────────────────────────────────────────
    ws = wb["DIRECTEURS"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        cfg.directeurs.append({
            "nom":       str(row[0]).strip(),
            "fonction":  str(row[1] or "").strip(),
            "abrev":     str(row[2] or "").strip(),
            "telephone": str(row[3] or "").strip(),
            "note":      str(row[4] or "").strip() if len(row) > 4 else "",
        })

    # ── TELEPHONIE ───────────────────────────────────────────────────────────
    ws = wb["TELEPHONIE"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[0]:
            continue
        # Skip les en-têtes intermédiaires
        if str(row[0]).strip().startswith(("CONTACTS", "Service", "📞")):
            continue
        cfg.telephonie.append({
            "service":  str(row[0]).strip(),
            "interne":  str(row[1] or "").strip(),
            "direct":   str(row[2] or "").strip(),
            "mobile":   str(row[3] or "").strip(),
            "site":     str(row[4] or "").strip() if len(row) > 4 else "",
            "note":     str(row[5] or "").strip() if len(row) > 5 else "",
        })

    # ── UF_INCIDENTS ─────────────────────────────────────────────────────────
    ws = wb["UF_INCIDENTS"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        if str(row[0]).strip() in ("Code UF", "🏥"):
            continue
        actif_raw = str(row[4] or "O").strip().upper() if len(row) > 4 else "O"
        cfg.unites_fonctionnelles.append({
            "code":   str(row[0]).strip(),
            "libelle": str(row[1] or "").strip(),
            "pole":   str(row[2] or "").strip(),
            "site":   str(row[3] or "").strip() if len(row) > 3 else "",
            "actif":  actif_raw == "O",
            "note":   str(row[5] or "").strip() if len(row) > 5 else "",
        })

    # ── SERVICES_CAPACITE ────────────────────────────────────────────────────
    ws = wb["SERVICES_CAPACITE"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        if str(row[0]).strip() in ("Service", "🛏️"):
            continue
        try:
            capa = int(row[4]) if row[4] else 0
        except (ValueError, TypeError):
            capa = 0
        try:
            seuil1 = int(row[5]) if len(row) > 5 and row[5] else 0
        except (ValueError, TypeError):
            seuil1 = 0
        try:
            seuil2 = int(row[6]) if len(row) > 6 and row[6] else 0
        except (ValueError, TypeError):
            seuil2 = 0
        try:
            ordre = int(row[11]) if len(row) > 11 and row[11] else 0
        except (ValueError, TypeError):
            ordre = 0
        def _bool_oui(v): return str(v or "N").strip().upper() == "O"
        cfg.services_capacite.append({
            "service":      str(row[0]).strip(),
            "code_uf":      str(row[1] or "").strip(),
            "pole":         str(row[2] or "").strip(),
            "site":         str(row[3] or "").strip(),
            "capacite":     capa,
            "seuil_t1":     seuil1,
            "seuil_t2":     seuil2,
            "accepte_h":    _bool_oui(row[7]) if len(row) > 7 else True,
            "accepte_f":    _bool_oui(row[8]) if len(row) > 8 else True,
            "accepte_i":    _bool_oui(row[9]) if len(row) > 9 else False,
            "tel_cadre":    str(row[10] or "").strip() if len(row) > 10 else "",
            "ordre":        ordre,
        })

    return cfg


def get_first_site(cfg: ParsedConfig) -> dict[str, Any] | None:
    """Retourne le premier site (utile pour pré-remplir l'instance master)."""
    return cfg.sites[0] if cfg.sites else None
