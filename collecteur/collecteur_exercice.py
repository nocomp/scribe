"""
collecteur_exercice.py — Collecteur SCRIBE Exercice de Crise v2
Port 8565 — Supervision + Animation complète

Contient :
- Supervision des 7 instances exercice
- Interface animateur (scénarios, stimuli, bilan)
- Injecteur asyncio (timing compressé)
- Génération scénario via Albert IA
"""

import asyncio
import json
import logging
import os
import secrets
import time
import glob
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional

import httpx
import uvicorn
from fastapi import FastAPI, HTTPException, Request, Depends, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.exceptions import HTTPException as StarletteHTTPException
from fastapi.responses import JSONResponse as _JSONResponse
from fastapi import Request as _Request
from pydantic import BaseModel

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("collecteur_exercice")

BASE_DIR   = Path(__file__).parent
ROOT_DIR   = BASE_DIR.parent  # répertoire racine SCRIBE
DATA_FILE  = str(BASE_DIR / "collecteur_exo_data.json")
TOKENS_FILE= str(BASE_DIR / "collecteur_exo_tokens.json")
ADMIN_FILE = str(BASE_DIR / "collecteur_exo_admin.json")
UI_AUTH_FILE = str(BASE_DIR / "collecteur_exo_ui_auth.json")
SCENARIOS_DIR = ROOT_DIR / "scenarios"

# Instances exercice — v2316 : configurable par variable d'environnement
# Permet de surcharger pour Docker / build public / déploiements custom.
# Format de SCRIBE_EXO_INSTANCES : "SIGLE1:PORT1,SIGLE2:PORT2,..."
# Ex Docker public : "CH_NORD:8660,CH_SUD:8661,CHU_CENTRE:8662,..."
def _parse_exo_instances() -> dict:
    raw = os.environ.get("SCRIBE_EXO_INSTANCES", "").strip()
    if not raw:
        # Défaut hardcodé : les 7 GHT du déploiement initial CHAG/Arc Alpin
        return {
            "CHAG":8660,"GHTLMB":8661,"CHRUMILLY":8662,
            "HDLEMAN":8663,"HPMB":8664,"CHB":8665,"CHPG":8666,
        }
    out = {}
    for entry in raw.split(","):
        entry = entry.strip()
        if not entry or ":" not in entry: continue
        sigle, port = entry.split(":", 1)
        try:
            out[sigle.strip()] = int(port.strip())
        except ValueError:
            continue
    return out or {
        "CHAG":8660,"GHTLMB":8661,"CHRUMILLY":8662,
        "HDLEMAN":8663,"HPMB":8664,"CHB":8665,"CHPG":8666,
    }

EXO_INSTANCES = _parse_exo_instances()
EXO_HOST = os.environ.get("SCRIBE_EXO_HOST","http://localhost")

app = FastAPI(title="SCRIBE Collecteur Exercice",version="2.3.102")
security = HTTPBearer(auto_error=False)

# ── État en mémoire ────────────────────────────────────────────────────────────
etablissements: dict = {}
tokens: dict = {}
ui_sessions: dict = {}

# ── État injecteur ─────────────────────────────────────────────────────────────
_inj = {
    "running":False,"paused":False,"session_id":None,
    "scenario":None,"t_start":None,"t_paused_s":0.0,
    "pause_start":None,"done":[],"errors":[],"t_frozen":0,
}
_inj_task: Optional[asyncio.Task] = None

# ── Scénario actif ─────────────────────────────────────────────────────────────
_scenario_actif: Optional[dict] = None
_token_instances: dict = {}   # sigle → token JWT animateur sur l'instance

# v2202 — Buffer circulaire des 100 derniers logs d'injection, exposé via
# /api/exercice/injection-log pour diagnostic rapide quand un stimulus
# semble "vert" côté animateur mais n'arrive pas côté joueur. On peut ainsi
# voir EXACTEMENT ce que l'instance a répondu sans avoir à grep les logs VPS.
from collections import deque
_injection_log: deque = deque(maxlen=100)

def _log_injection(entry: dict):
    """Ajoute une entrée au buffer de logs d'injection."""
    entry["t"] = datetime.now(timezone.utc).isoformat()
    _injection_log.append(entry)

# ── Chat inter-GHT (porté depuis collecteur/collecteur.py) ─────────────────────
CHAT_SALONS_DEFAULT = ["général", "coordination", "transferts", "logistique", "direction"]
chat_messages: dict = {s: [] for s in CHAT_SALONS_DEFAULT}
chat_presence: dict = {}          # sigle → [{user_id, display_name, last_seen}]
chat_pj_store: dict = {}          # pj_id → {nom, dataUrl|remote_url, taille}
_chat_msg_counter: int = 0
_chat_pj_counter: int = 0

# ── Messagerie inter-GHT et transferts (porté depuis collecteur/collecteur.py) ─
MESSAGES_FILE   = str(BASE_DIR / "collecteur_exo_messages.json")
TRANSFERTS_FILE = str(BASE_DIR / "collecteur_exo_transferts.json")
messages_inter:   list = []
transferts_inter: list = []

def load_messages_inter():
    global messages_inter
    if Path(MESSAGES_FILE).exists():
        try: messages_inter = json.loads(Path(MESSAGES_FILE).read_text())
        except Exception: messages_inter = []

def save_messages_inter():
    try:
        Path(MESSAGES_FILE).write_text(json.dumps(messages_inter, ensure_ascii=False, indent=2))
    except Exception as e:
        logger.warning(f"save_messages_inter: {e}")

def load_transferts_inter():
    global transferts_inter
    if Path(TRANSFERTS_FILE).exists():
        try: transferts_inter = json.loads(Path(TRANSFERTS_FILE).read_text())
        except Exception: transferts_inter = []

def save_transferts_inter():
    try:
        Path(TRANSFERTS_FILE).write_text(json.dumps(transferts_inter, ensure_ascii=False, indent=2))
    except Exception as e:
        logger.warning(f"save_transferts_inter: {e}")

# Helpers d'auth étendus ──────────────────────────────────────────────────────
def get_etab_from_token(credentials) -> Optional[str]:
    """Retourne le sigle associé au token établissement, ou 'SUPERVISION' pour l'admin."""
    if not credentials:
        return None
    tok = credentials.credentials
    if tok == ADMIN_TOKEN:
        return "SUPERVISION"
    return tokens.get(tok)

def _check_any_auth(credentials) -> bool:
    """Accepte token établissement OU session UI valide OU token admin."""
    if not credentials:
        return False
    tok = credentials.credentials
    return tok == ADMIN_TOKEN or tok in tokens or tok in ui_sessions

# ── CORS ───────────────────────────────────────────────────────────────────────
@app.middleware("http")
async def cors_mw(request: Request, call_next):
    origin = request.headers.get("origin") or "*"
    h = {"Access-Control-Allow-Origin":origin,
         "Access-Control-Allow-Credentials":"true",
         "Access-Control-Allow-Methods":"GET,POST,PUT,DELETE,OPTIONS,PATCH",
         "Access-Control-Allow-Headers":"Authorization,Content-Type,Accept"}
    if request.method == "OPTIONS":
        return Response(status_code=200, headers=h)
    try: resp = await call_next(request)
    except Exception: resp = _JSONResponse({"detail":"error"},status_code=500)
    for k,v in h.items(): resp.headers[k] = v
    return resp

@app.exception_handler(StarletteHTTPException)
async def http_err(req,exc):
    orig = req.headers.get("origin") or "*"
    return _JSONResponse({"detail":exc.detail},status_code=exc.status_code,
        headers={"Access-Control-Allow-Origin":orig,"Access-Control-Allow-Credentials":"true"})

@app.exception_handler(Exception)
async def gen_err(req,exc):
    orig = req.headers.get("origin") or "*"
    logger.error(f"Exception: {exc}")
    return _JSONResponse({"detail":str(exc)},status_code=500,
        headers={"Access-Control-Allow-Origin":orig,"Access-Control-Allow-Credentials":"true"})

# ── Admin token ─────────────────────────────────────────────────────────────────
def _make_admin_token() -> str:
    if os.environ.get("ADMIN_TOKEN"): return os.environ["ADMIN_TOKEN"]
    p = Path(ADMIN_FILE)
    if p.exists():
        try: return json.loads(p.read_text())["admin_token"]
        except: pass
    t = "exo_admin_" + secrets.token_hex(16)
    p.write_text(json.dumps({"admin_token":t},indent=2))
    return t

ADMIN_TOKEN = _make_admin_token()

# v2315-public — Tokens d'authentification fédération.
# AVANT (privé) : tokens prévisibles `token_exo_<sigle>_2026`.
# APRÈS (public) : tokens dérivés d'une seed locale unique générée au
# 1er lancement (~/.scribe_federation_seed) et persistante. Un attaquant
# qui ne connaît pas la seed ne peut pas forger de token.
# Si tu veux tes propres tokens explicites, ajoute-les dans
# collecteur_exo_tokens.json (overwrite la map ci-dessous).
def _get_or_make_seed():
    """Retourne (et crée si besoin) une seed locale pour dériver les tokens.
    Stockée en clair dans un fichier home — non secrète, mais imprévisible."""
    import os, hashlib, secrets
    seed_file = Path(os.path.expanduser("~/.scribe_federation_seed"))
    if seed_file.exists():
        try: return seed_file.read_text().strip()
        except Exception: pass
    seed = secrets.token_urlsafe(32)
    try:
        seed_file.write_text(seed)
        seed_file.chmod(0o600)
    except Exception as e:
        logger.warning(f"Impossible d'écrire la seed fédération: {e}")
    return seed

_FED_SEED = _get_or_make_seed()

def _derive_token(sigle: str) -> str:
    """Dérive un token déterministe (sigle, seed) → token. Pas réversible."""
    import hashlib
    h = hashlib.sha256(f"{_FED_SEED}:{sigle.upper()}".encode()).hexdigest()
    return f"fed_{h[:32]}"

def load_tokens():
    global tokens
    for sigle in EXO_INSTANCES:
        tokens[_derive_token(sigle)] = sigle
    if Path(TOKENS_FILE).exists():
        try: tokens.update(json.loads(Path(TOKENS_FILE).read_text()))
        except: pass

def load_data():
    global etablissements
    if Path(DATA_FILE).exists():
        try: etablissements = json.loads(Path(DATA_FILE).read_text())
        except: etablissements = {}

def save_data():
    Path(DATA_FILE).write_text(json.dumps(etablissements,ensure_ascii=False,indent=2,default=str))

# ── Chargement initial au niveau module ─────────────────────────────────────
# Important : ces appels DOIVENT être au top-level (pas dans __main__), sinon
# le worker uvicorn réimporte le module sans passer par __main__ et les tokens
# restent vides → toutes les routes qui dépendent de `tokens` renvoient 401.
load_tokens()
load_data()
load_messages_inter()
load_transferts_inter()

def _auth_ok(credentials) -> bool:
    if not credentials: return False
    t = credentials.credentials
    return t == ADMIN_TOKEN or t in ui_sessions

def require_auth(credentials=Depends(security)):
    if not _auth_ok(credentials): raise HTTPException(401,"Non authentifié")
    return True

# ── Auth UI ─────────────────────────────────────────────────────────────────────
def check_creds(login:str, pwd:str) -> bool:
    import hashlib
    auth_data = {"users":[
        {"login":"animateur","password_hash":hashlib.sha256(b"Animateur2026!").hexdigest(),"role":"admin"},
        {"login":"supervision","password_hash":hashlib.sha256(b"Scribe2026!").hexdigest(),"role":"admin"},
    ]}
    if Path(UI_AUTH_FILE).exists():
        try: auth_data = json.loads(Path(UI_AUTH_FILE).read_text())
        except: pass
    h = hashlib.sha256(pwd.encode()).hexdigest()
    return any(u["login"]==login and u["password_hash"]==h for u in auth_data.get("users",[]))

@app.post("/api/ui/login")
async def ui_login(request: Request):
    body = await request.json()
    if not check_creds(body.get("login",""),body.get("password","")):
        raise HTTPException(401,"Identifiants incorrects")
    t = secrets.token_hex(24)
    ui_sessions[t] = {"login":body["login"],"at":time.time()}
    return {"ok":True,"token":t}

@app.get("/api/ui/verify")
def verify(credentials=Depends(security)):
    # v2198 — Retourner 200 {ok:false} au lieu de 401 pour ne pas polluer
    # la console Chrome avec des "GET ... 401 (Unauthorized)" à chaque
    # chargement de page avec un token expiré/absent. Le client lit
    # simplement la clé `ok` et réagit en conséquence.
    if not _auth_ok(credentials):
        return {"ok": False}
    return {"ok": True}

# ── Push depuis instances ──────────────────────────────────────────────────────
# Note : on accepte tout champ supplémentaire (demandes, declarations, etablissement…)
#        via Request.body() pour rester compatible avec le client fédération SCRIBE
#        qui envoie un payload riche dont le PushPayload restrictif ignorerait des champs.

@app.post("/api/push")
async def receive_push(request: Request, credentials=Depends(security)):
    if not credentials: raise HTTPException(401)
    t = credentials.credentials
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(400, "JSON invalide")
    sigle = tokens.get(t) or (payload.get("sigle") if t == ADMIN_TOKEN else None)
    if not sigle: raise HTTPException(401, "Token inconnu")
    etablissements[sigle] = {
        "sigle": sigle,
        "timestamp": payload.get("timestamp") or datetime.now(timezone.utc).isoformat(),
        "_received_at": datetime.now(timezone.utc).isoformat(),
        "niveau_global": payload.get("niveau_global", "VERT"),
        "nb_incidents_actifs": payload.get("nb_incidents_actifs", 0),
        "nb_transferts_actifs": payload.get("nb_transferts_actifs", 0),
        "sites": payload.get("sites") or [],
        "incidents": payload.get("incidents") or [],
        "demandes": payload.get("demandes") or [],
        "declarations": payload.get("declarations") or [],
        "etablissement": payload.get("etablissement") or {"sigle": sigle, "nom": sigle},
        "version": payload.get("version"),
        "port": EXO_INSTANCES.get(sigle, 8660),
    }
    save_data()
    return {"status": "ok", "sigle": sigle}

# ── Summary ────────────────────────────────────────────────────────────────────
@app.get("/api/summary")
async def get_summary(credentials=Depends(security)):
    # v2200 — Renvoie liste vide au lieu de 401 si pas d'auth. Les instances
    # joueur appelaient /api/summary sans token et pollaient la console avec
    # "GET /api/summary 401" à chaque refresh carte SOINS (30s).
    # Les instances avec token fédération valide continuent de recevoir les
    # vraies données. C'est cohérent avec le fix /verify en 2199.
    if not _check_any_auth(credentials):
        return []
    now = datetime.now(timezone.utc)
    result = []
    for sigle,port in EXO_INSTANCES.items():
        etab = etablissements.get(sigle,{})
        online = False
        ts = etab.get("timestamp")
        if ts:
            try:
                last = datetime.fromisoformat(ts.replace("Z","+00:00"))
                if last.tzinfo is None: last = last.replace(tzinfo=timezone.utc)
                online = (now-last).total_seconds() < 90
            except: pass
        result.append({
            "sigle":sigle,"port":port,"url":f"{EXO_HOST}:{port}","online":online,
            "niveau_global":etab.get("niveau_global","INCONNU"),
            "nb_incidents_actifs":etab.get("nb_incidents_actifs",0),
            "timestamp":ts,
        })
    return result

# ── Scénarios ──────────────────────────────────────────────────────────────────
@app.get("/api/exercice/scenarios")
async def list_scenarios(auth=Depends(require_auth)):
    """v2316 — Liste les scénarios disponibles dans le dossier.
    Log explicitement les fichiers ignorés (parsing impossible, JSON invalide…)
    pour éviter qu'un scénario disparaisse silencieusement."""
    SCENARIOS_DIR.mkdir(exist_ok=True)
    result = []
    ignored = []
    for f in sorted(SCENARIOS_DIR.glob("*.json")):
        try:
            sc = json.loads(f.read_text(encoding="utf-8"))
            meta = sc.get("meta",{})
            result.append({
                "filename":f.name,
                "id":meta.get("id",f.stem),
                "titre":meta.get("titre","Sans titre"),
                "duree_min":meta.get("duree_min",60),
                "duree_reel_min":meta.get("duree_reel_min",240),
                "complexite":meta.get("complexite","MOYEN"),
                "type_crise":meta.get("type_crise","SANITAIRE"),
                "nb_stimuli":len(sc.get("stimuli",[])),
                "nb_sites":len(sc.get("acteurs",[])),
            })
        except json.JSONDecodeError as e:
            # v2316 — Log explicite (avant: try/except: pass silencieux)
            logger.warning(f"[SCENARIOS] Fichier '{f.name}' ignoré (JSON invalide): {e}")
            ignored.append({"filename": f.name, "error": f"JSON invalide: {e}"})
        except UnicodeDecodeError as e:
            logger.warning(f"[SCENARIOS] Fichier '{f.name}' ignoré (encodage non UTF-8): {e}")
            ignored.append({"filename": f.name, "error": f"Encodage non UTF-8: {e}"})
        except Exception as e:
            logger.warning(f"[SCENARIOS] Fichier '{f.name}' ignoré: {type(e).__name__}: {e}")
            ignored.append({"filename": f.name, "error": f"{type(e).__name__}: {e}"})
    if ignored:
        logger.warning(f"[SCENARIOS] {len(ignored)} fichier(s) JSON ignoré(s) au scan")
    return result

@app.get("/api/exercice/scenarios/{filename}")
async def get_scenario(filename:str, auth=Depends(require_auth)):
    p = SCENARIOS_DIR / filename
    if not p.exists(): raise HTTPException(404)
    return json.loads(p.read_text(encoding="utf-8"))

# v2197 — Téléchargement direct d'un scénario, sans passer par un Blob côté
# client. Chrome bloque les blob: URLs servis depuis une page http:// non
# sécurisée ("was loaded over an insecure connection"). En servant
# directement le fichier avec Content-Disposition: attachment, on évite
# totalement le mécanisme blob et donc ce blocage.
@app.get("/api/exercice/scenarios/{filename}/download")
async def download_scenario(filename:str, token: str = None, auth=None):
    # Sécurité : mêmes garde-fous que delete (pas de traversée)
    if ("/" in filename) or ("\\" in filename) or filename.startswith(".") or not filename.endswith(".json"):
        raise HTTPException(400, "Nom de fichier invalide")
    # Authentification : accepter soit le header Authorization (géré via Depends),
    # soit un token en query string (pour les <a href> de download direct qui ne
    # peuvent pas envoyer de header).
    if not token:
        # Si pas de token en query, on exige l'auth header (via Depends côté appel
        # fetch classique). On refait un require_auth manuel ici car le default
        # auth=None contourne le depend.
        raise HTTPException(401, "Token requis")
    if token != ADMIN_TOKEN and token not in tokens and token not in ui_sessions:
        raise HTTPException(401, "Token invalide")
    p = SCENARIOS_DIR / filename
    if not p.exists():
        raise HTTPException(404, "Scénario introuvable")
    # Nom de fichier proposé au téléchargement (slug du titre si dispo)
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        titre = (data.get("meta", {}) or {}).get("titre") or filename[:-5]
    except Exception:
        titre = filename[:-5]
    import re as _re
    slug = _re.sub(r"[^a-z0-9]+", "_", titre.lower()).strip("_")[:50] or "export"
    proposed = f"scenario_{slug}.json"
    from fastapi.responses import FileResponse
    return FileResponse(str(p), media_type="application/json", filename=proposed)


# v2194 — Suppression d'un scénario (garde-fou sur le nom de fichier : pas de
# traversée de chemin, extension .json obligatoire, pas de préfixe caché).
@app.delete("/api/exercice/scenarios/{filename}")
async def delete_scenario(filename:str, auth=Depends(require_auth)):
    # Sécurité : valider le nom de fichier contre toute traversée
    if ("/" in filename) or ("\\" in filename) or filename.startswith(".") or not filename.endswith(".json"):
        raise HTTPException(400, "Nom de fichier invalide")
    p = SCENARIOS_DIR / filename
    if not p.exists():
        raise HTTPException(404, "Scénario introuvable")
    try:
        p.unlink()
    except Exception as e:
        raise HTTPException(500, f"Échec suppression : {e}")
    logger.info(f"Scénario supprimé : {filename}")
    return {"ok": True, "filename": filename}

@app.post("/api/exercice/scenarios/import-xml")
async def import_xml_scenario(request:Request, auth=Depends(require_auth)):
    """Convertit un scénario XML en JSON et le sauvegarde."""
    import xml.etree.ElementTree as ET
    body = await request.body()
    try:
        root = ET.fromstring(body.decode("utf-8"))
    except Exception as e:
        raise HTTPException(400, f"XML invalide: {e}")
    
    def t(el, tag, default=""):
        n = el.find(tag)
        return n.text.strip() if n is not None and n.text else default
    
    meta_el = root.find("meta")
    stimuli_el = root.find("stimuli")
    acteurs_el = root.find("acteurs")
    dec_el = root.find("decisions_attendues")
    deb_el = root.find("debriefing")
    
    if meta_el is None or stimuli_el is None:
        raise HTTPException(400, "XML doit contenir <meta> et <stimuli>")
    
    scenario = {
        "meta": {
            "id": t(meta_el, "id") or f"exo_import_{datetime.now().strftime('%Y%m%d_%H%M')}",
            "titre": t(meta_el, "titre", "Scénario importé"),
            "description": t(meta_el, "description", ""),
            "type_crise": t(meta_el, "type_crise", "SANITAIRE"),
            "complexite": t(meta_el, "complexite", "MOYEN"),
            "duree_min": int(t(meta_el, "duree_exercice_min", "60")),
            "duree_reel_min": int(t(meta_el, "duree_reel_simulee_min", "240")),
            "ratio_compression": round(int(t(meta_el, "duree_reel_simulee_min", "240")) / max(1, int(t(meta_el, "duree_exercice_min", "60"))), 1),
            "objectifs_pedagogiques": [o.text.strip() for o in meta_el.findall(".//objectif") if o.text],
        },
        "acteurs": [],
        "stimuli": [],
        "decisions_attendues": [],
        "debriefing_guide": {}
    }
    
    # Acteurs
    if acteurs_el is not None:
        for a in acteurs_el.findall("acteur"):
            joueurs = []
            for j in a.findall(".//joueur"):
                joueurs.append({
                    "username": t(j, "username"),
                    "display_name": t(j, "display_name"),
                    "role_exercice": t(j, "role_exercice"),
                    "responsabilites": [r.text.strip() for r in j.findall(".//resp") if r.text],
                })
            scenario["acteurs"].append({
                "sigle": a.get("sigle", "CHAG"),
                "nom_etablissement": t(a, "nom_etablissement"),
                "role": a.get("role", "participant"),
                "port": int(a.get("port", "8660")),
                "joueurs": joueurs,
            })
    
    # Stimuli
    for s in stimuli_el.findall("stimulus"):
        pl = s.find("payload")
        payload = {}
        if pl is not None:
            payload = {
                "fait": t(pl, "fait"),
                "urgency": int(t(pl, "urgency", "2")),
                "type_crise": t(pl, "type_crise", "SANITAIRE"),
                "site_id": t(pl, "site_id"),
                "unite_fonctionnelle": t(pl, "unite_fonctionnelle"),
                "declarant_nom": t(pl, "declarant_nom"),
                "analyse": t(pl, "analyse"),
                "contenu": t(pl, "contenu"),  # pour type message
                "jalons_labels": [j.text.strip() for j in pl.findall(".//jalon") if j.text],
            }
        scenario["stimuli"].append({
            "id": t(s, "id", f"S{len(scenario['stimuli'])+1:02d}"),
            "t_min": float(t(s, "t_min", "0")),
            "cible": t(s, "cible", "CHAG"),
            "type": t(s, "type", "incident"),
            "titre": t(s, "titre", ""),
            "description_animateur": t(s, "description_animateur", ""),
            "action_attendue": t(s, "action_attendue", ""),
            "payload": payload,
        })
    
    # Décisions
    if dec_el is not None:
        for d in dec_el.findall("decision"):
            scenario["decisions_attendues"].append({
                "t_min": float(d.get("t_min", "0")),
                "contenu": d.text.strip() if d.text else "",
                "responsable": d.get("responsable", ""),
                "obligatoire": d.get("obligatoire", "false") == "true",
            })
    
    # Debriefing
    if deb_el is not None:
        scenario["debriefing_guide"] = {
            "points_cles": [p.text.strip() for p in deb_el.findall(".//point") if p.text],
            "questions_debriefing": [q.text.strip() for q in deb_el.findall(".//question") if q.text],
            "pieges_frequents": [p.text.strip() for p in deb_el.findall(".//piege") if p.text],
        }
    
    # Sauvegarder
    sid = scenario["meta"]["id"]
    SCENARIOS_DIR.mkdir(exist_ok=True)
    (SCENARIOS_DIR / f"{sid}.json").write_text(json.dumps(scenario, ensure_ascii=False, indent=2))
    return {"ok": True, "scenario_id": sid, "filename": f"{sid}.json", "scenario": scenario}

@app.post("/api/exercice/scenarios")
async def save_scenario(request:Request, auth=Depends(require_auth)):
    body = await request.json()

    # v2316 — Validation et auto-correction du scénario importé.
    # Même logique que pour la génération IA : on corrige les types
    # incorrects, on signale en warnings, on rejette si bloquant.
    from collecteur_exercice.scenario_validator import validate_and_fix
    ok, warns, errs, body = validate_and_fix(body, source="import")
    if not ok:
        raise HTTPException(
            422,
            "Le scénario importé est invalide : " + "; ".join(errs[:3]),
        )

    meta = body.get("meta",{})
    # v2192 — Pour les imports JSON, l'id du fichier source peut déjà exister
    # dans la bibliothèque. On suffixe systématiquement avec un timestamp pour
    # éviter d'écraser silencieusement un scénario existant.
    def _slug(s: str, maxlen: int = 40) -> str:
        import re as _re
        s = _re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")
        return (s[:maxlen] or "scenario").strip("_")
    source_id = meta.get("id") or _slug(meta.get("titre",""))
    if not source_id:
        source_id = "exo_" + datetime.now().strftime("%Y%m%d_%H%M")
    sid = f"{source_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    body.setdefault("meta", {})["id"] = sid  # id mis à jour dans le fichier
    filename = f"{sid}.json"
    SCENARIOS_DIR.mkdir(exist_ok=True)
    (SCENARIOS_DIR/filename).write_text(json.dumps(body,ensure_ascii=False,indent=2))
    return {"ok":True,"filename":filename,"id":sid,"validation_warnings":warns}

# ── Génération IA ──────────────────────────────────────────────────────────────
class GenRequest(BaseModel):
    sujet:str; nb_sites:int=1; sites:list=["CHAG"]
    duree_exercice_min:int=60; duree_reel_min:int=240
    complexite:str="MOYEN"; type_crise:str="SANITAIRE"; langue:str="fr"
    nb_joueurs:int=4; nb_stimuli:int=8
    stimuli_externes:str="none"
    valeurs_metiers:list=[]
    services_supports:list=[]
    perturbations:str=""
    # v2192 — Identité saisie par l'utilisateur avant la génération.
    # Appliquée APRÈS l'appel Albert sur le scénario reçu. Le prompt envoyé
    # à Albert n'est PAS modifié (on ne touche pas à ce qui marche).
    nom_scenario:str=""
    categorie:str=""
    tags:list=[]

ALBERT_URL   = "https://albert.api.etalab.gouv.fr/v1/chat/completions"
ALBERT_MODEL = "mistralai/Ministral-3-8B-Instruct-2512"
ALBERT_KEY   = os.environ.get("SCRIBE_IA_KEY",
    os.getenv("ALBERT_API_KEY", ""))

SYSTEM_EXO = """Tu es expert en gestion de crise hospitalière française. Tu génères des scénarios d'exercice réalistes pour équipes GHT. Tu réponds UNIQUEMENT en JSON valide, sans texte autour."""

def _prompt_scenario(b:GenRequest) -> str:
    ratio = round(b.duree_reel_min/b.duree_exercice_min,1)
    sid = datetime.now().strftime("%Y%m%d_%H%M")
    ports = {"CHAG":8660,"GHTLMB":8661,"CHRUMILLY":8662,"HDLEMAN":8663,"HPMB":8664,"CHB":8665,"CHPG":8666}
    
    # Construire les instructions spécifiques
    type_instructions = {
        "SANITAIRE": "Crise médicale/chirurgicale. Inclure stimuli cliniques réalistes.",
        "CYBER": "Crise cyber hospitalière. SIH, DPI, PACS potentiellement touchés. Inclure stimuli CERT Santé, isolation réseaux, continuité sans outils numériques.",
        "MIXTE": "Crise mixte cyber+sanitaire. Commencer par l'aspect cyber puis impact sur la prise en charge.",
        "RH": "Crise de continuité de service RH. Manque de personnel, remplacement urgent, gestion des plannings. Ex: trouver anesthésiste disponible en urgence.",
        "TERTIAIRE": "Crise accueil/logistique. Afflux de patients, gestion des lits, restauration, transport, hébergement.",
    }.get(b.type_crise, "")
    
    # v2186a — supporte la liste CSV multi-sélection venant des chips UI.
    # Ancien format compat : "samu" / "prefecture" / "medias" / "all" / "none".
    # Nouveau format : "samu,prefecture,medias,famille,cert_sante,ght_voisin"
    _ext_labels = {
        "samu":       "SAMU 15 / régulation médicale (appels coordination, demandes de transport)",
        "prefecture": "préfecture / ARS (demandes de points de situation, instructions sanitaires)",
        "medias":     "médias / presse (sollicitations communication, risque de désinformation)",
        "famille":    "familles / patients (appels inquiets, demandes d'information)",
        "cert_sante": "CERT Santé / ANSSI (alertes cyber, demandes de remontée d'incident)",
        "ght_voisin": "établissement voisin hors GHT (demandes d'entraide, transferts inverses)",
    }
    raw = (b.stimuli_externes or "").strip()
    if raw in ("none", "", None):
        ext_keys = []
    elif raw == "all":
        ext_keys = ["samu", "prefecture", "medias"]
    else:
        ext_keys = [k.strip() for k in raw.split(",") if k.strip() in _ext_labels]
    if ext_keys:
        stimuli_ext_instructions = (
            "Inclure des stimuli externes provenant de : " +
            ", ".join(_ext_labels[k] for k in ext_keys) +
            ". Chaque acteur externe doit générer 1 à 2 stimuli minimum."
        )
    else:
        stimuli_ext_instructions = ""

    valeurs_str = ", ".join(b.valeurs_metiers) if b.valeurs_metiers else "coordination, communication"
    services_str = ", ".join(b.services_supports) if b.services_supports else ""
    
    return f"""Génère un scénario d'exercice de crise hospitalière :
SUJET: {b.sujet}
SITES GHT: {", ".join(b.sites)} ({b.nb_sites} site(s))
DURÉE EXERCICE: {b.duree_exercice_min} minutes (ratio compression: {ratio}x → 1 min exercice = {ratio} min réelles)
DURÉE RÉELLE SIMULÉE: {b.duree_reel_min} minutes
TYPE DE CRISE: {b.type_crise} — {type_instructions}
COMPLEXITÉ: {b.complexite}
NOMBRE DE PARTICIPANTS: {b.nb_joueurs} personnes (générer un joueur avec rôle pour chacun)
NOMBRE DE STIMULI: {b.nb_stimuli} (espacés progressivement)
VALEURS MÉTIERS À TRAVAILLER: {valeurs_str}
SERVICES SUPPORTS IMPLIQUÉS: {services_str if services_str else "selon pertinence clinique"}
STIMULI EXTERNES: {stimuli_ext_instructions if stimuli_ext_instructions else "aucun stimulus externe"}
PERTURBATIONS FONCTIONNELLES: {b.perturbations if b.perturbations else "aucune panne/perturbation"}

CONTRAINTE DE SITE — IMPORTANT :
- {b.nb_sites} site(s) sélectionné(s). TOUS les stimuli doivent cibler UNIQUEMENT ces sites : {", ".join(b.sites)}.
- La complexité "{b.complexite}" se traduit par la RICHESSE des événements (nb de stimuli, gravité, dilemmes),
  PAS par le nombre de sites. Un exercice complexe mono-site est parfaitement valide : crise majeure concentrée
  sur un seul établissement (ex. panne SI généralisée, afflux massif, coupure énergie...).
- NE JAMAIS cibler un site non listé. Les "cible" des stimuli doivent strictement appartenir à : {", ".join(b.sites)}.

NOMS DES ÉTABLISSEMENTS — pour la rédaction des textes :
- Les sigles à utiliser dans les champs "cible", "sigle" et les identifiants sont : CHAG, GHTLMB, CHRUMILLY, HDLEMAN, HPMB, CHB, CHPG.
- MAIS dans les textes en langage naturel (titres, descriptions, faits, contenus de messages), utilisez le VRAI nom :
  * CHAG = "CHANGE" (Centre Hospitalier ANnecy-GEnevois)
  * GHTLMB = "Hôpitaux du Léman" ou "Thonon"
  * CHRUMILLY = "CH Rumilly"
  * HDLEMAN = "Hôpitaux du Pays du Mont-Blanc"
  * HPMB = "Hôpital Privé Mont-Blanc"
  * CHB = "CH Bonneville"
  * CHPG = "CH Pays de Gex"
- Exemple CORRECT : titre "Afflux massif au CHANGE", cible "CHAG".
- Exemple INCORRECT : titre "Afflux massif au CHAG" (on utilise le nom long dans les textes).

Retourne UNIQUEMENT ce JSON valide (rien d'autre, pas de texte avant ou après):
{{"meta":{{"id":"exo_{sid}","titre":"<titre court et évocateur>","description":"<2-3 phrases>","duree_min":{b.duree_exercice_min},"duree_reel_min":{b.duree_reel_min},"ratio_compression":{ratio},"complexite":"{b.complexite}","type_crise":"{b.type_crise}","objectifs_pedagogiques":["<obj1>","<obj2>","<obj3>"]}},"acteurs":[{{"sigle":"{b.sites[0] if b.sites else "CHAG"}","nom_etablissement":"<nom complet>","role":"coordinateur","port":{ports.get(b.sites[0] if b.sites else "CHAG",8660)},"joueurs":[{{"username":"dircrise","display_name":"<prénom NOM — Directeur de Crise>","role_exercice":"<rôle précis dans l'exercice>","responsabilites":["<resp1>","<resp2>"]}}]}}],"stimuli":[{{"id":"S01","t_min":0,"cible":"{b.sites[0] if b.sites else "CHAG"}","type":"incident","titre":"<titre court>","description_animateur":"<contexte pour animateur — ce que les joueurs doivent faire>","payload":{{"fait":"<description précise pour les joueurs>","urgency":3,"type_crise":"{b.type_crise}","site_id":"<SIGLE>","unite_fonctionnelle":"<service>","declarant_nom":"<qui déclare>","analyse":"","jalons_labels":["<jalon1>","<jalon2>","<jalon3>"]}},"action_attendue":"<décision ou action attendue des joueurs>"}}],"decisions_attendues":[{{"t_min":5,"contenu":"<décision>","responsable":"Directeur de crise","obligatoire":true}}],"debriefing_guide":{{"points_cles":["<point1>","<point2>"],"questions_debriefing":["<question1>","<question2>"],"pieges_frequents":["<piège1>"]}}}}

RÈGLES STRICTES:
- EXACTEMENT {b.nb_stimuli} stimuli, espacés progressivement (T+0, T+5, T+10...)
- Types stimuli disponibles: incident, message, transfert, chat, decision
- Ports fixes: CHAG=8660 GHTLMB=8661 CHRUMILLY=8662 HDLEMAN=8663 HPMB=8664 CHB=8665 CHPG=8666
- EXACTEMENT {b.nb_joueurs} joueurs répartis sur les sites
- JSON VALIDE UNIQUEMENT — pas de backtick, pas de commentaire, pas de texte hors JSON"""


@app.post("/api/exercice/generate")
async def generate_scenario_ia(body:GenRequest, auth=Depends(require_auth)):
    """v2199 — Retry automatique si Albert produit un JSON invalide.

    Historique :
    - 2193 a fixé la troncature (max_tokens 6000→12000) via diagnostic runtime
    - Mais Albert peut encore produire occasionnellement un JSON mal formé
      en plein milieu (virgule oubliée, char de contrôle, guillemet non
      échappé dans une string). Pas une question de longueur, juste de
      non-déterminisme du LLM.
    - Fix : 2 tentatives. Si la 1re échoue avec json.JSONDecodeError,
      on relance avec temperature=0.2 (plus déterministe) et un hint dans
      le prompt pour insister sur le JSON strict.
    - Chaque raw est sauvegardé dans /tmp/albert_raw_*.txt pour diag.
    """
    import time as _t
    ts = _t.strftime("%Y%m%d_%H%M%S")
    last_err = None
    scenario = None

    for attempt in (1, 2):
        try:
            temp = 0.7 if attempt == 1 else 0.2
            prompt_user = _prompt_scenario(body)
            if attempt == 2:
                prompt_user += ("\n\nIMPORTANT: Réponds UNIQUEMENT avec un JSON "
                               "strictement valide. Pas de commentaires, pas de "
                               "markdown. Toutes les virgules séparatrices doivent "
                               "être présentes. Toutes les strings doivent fermer "
                               "leurs guillemets.")
                logger.warning(f"generate_scenario_ia : retry attempt {attempt} "
                              f"après échec JSON (temp={temp})")

            async with httpx.AsyncClient(timeout=120.0) as client:
                r = await client.post(ALBERT_URL,
                    headers={"Authorization":f"Bearer {ALBERT_KEY}","Content-Type":"application/json"},
                    json={"model":ALBERT_MODEL,"messages":[
                        {"role":"system","content":SYSTEM_EXO},
                        {"role":"user","content":prompt_user},
                    ],"max_tokens":30000,"temperature":temp}
                )
            if r.status_code != 200:
                raise HTTPException(500,f"Albert {r.status_code}: {r.text[:200]}")

            payload = r.json()
            finish_reason = payload.get("choices",[{}])[0].get("finish_reason","")
            usage = payload.get("usage",{})
            if finish_reason == "length":
                raise HTTPException(
                    500,
                    f"Albert a atteint la limite de tokens "
                    f"({usage.get('completion_tokens','?')} tokens output). "
                    f"Réduisez le nombre de stimuli ou la complexité du scénario."
                )

            raw = payload["choices"][0]["message"]["content"].strip()

            # Sauvegarder le raw pour diag (utile si JSON casse)
            try:
                debug_path = Path(f"/tmp/albert_raw_{ts}_try{attempt}.txt")
                debug_path.write_text(raw, encoding="utf-8")
            except Exception:
                pass

            # Nettoyer markdown (Albert encapsule parfois dans ```json ... ```)
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                raw = raw[4:] if raw.startswith("json") else raw
            if raw.endswith("```"):
                raw = raw[:-3]
            scenario = json.loads(raw.strip())
            # Succès
            if attempt > 1:
                logger.info(f"generate_scenario_ia : JSON valide à la tentative {attempt}")
            break
        except json.JSONDecodeError as e:
            last_err = e
            logger.warning(f"generate_scenario_ia tentative {attempt} : "
                          f"JSON invalide — {e} (raw sauvé dans /tmp/albert_raw_{ts}_try{attempt}.txt)")
            continue
        except HTTPException:
            raise
        except Exception as e:
            last_err = e
            logger.error(f"generate_scenario_ia tentative {attempt} : {e}")
            continue

    if scenario is None:
        raise HTTPException(500,
            f"Albert a produit un JSON invalide après 2 tentatives. "
            f"Dernière erreur : {last_err}. "
            f"Raw disponibles dans /tmp/albert_raw_{ts}_try*.txt pour diag.")

    try:
        # v2192 — Appliquer l'identité saisie par l'utilisateur AU RÉSULTAT
        # (on ne modifie PAS le prompt envoyé à Albert, pour préserver
        # le fonctionnement qui marchait en 2187).
        scenario.setdefault("meta", {})
        if body.nom_scenario.strip():
            scenario["meta"]["titre"] = body.nom_scenario.strip()
        if body.categorie.strip():
            scenario["meta"]["categorie"] = body.categorie.strip()
        if body.tags:
            scenario["meta"]["tags"] = body.tags

        # v2316 — Validation et auto-correction du scénario généré par l'IA.
        # Albert produit parfois des typages incorrects (ex: impact_fonctionnel
        # en string au lieu de bool, t_min tous à 0, payloads incomplets).
        # Le validateur corrige ce qu'il peut et signale le reste en warnings.
        # Si erreurs bloquantes, on rejette avec une 400 explicite (l'animateur
        # peut alors relancer la génération).
        from collecteur_exercice.scenario_validator import validate_and_fix
        ok, warns, errs, scenario = validate_and_fix(scenario, source="generate_ia")
        if not ok:
            logger.error(f"[GENERATE] Scénario IA invalide ({len(errs)} erreurs) : {errs}")
            raise HTTPException(
                422,
                "Le scénario généré par l'IA est invalide : "
                + "; ".join(errs[:3])
                + (". Relancer la génération." if errs else ""),
            )
        if warns:
            logger.info(f"[GENERATE] {len(warns)} corrections appliquées sur scénario IA")

        # Sauvegarder — nom de fichier basé sur titre saisi (slugifié),
        # suffixé d'un timestamp pour éviter les collisions.
        def _slug(s: str, maxlen: int = 40) -> str:
            import re as _re
            s = _re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")
            return (s[:maxlen] or "scenario").strip("_")
        base = _slug(body.nom_scenario) if body.nom_scenario else scenario.get("meta", {}).get("id", "")
        if not base:
            base = "exo_" + datetime.now().strftime("%Y%m%d_%H%M")
        sid = base + "_" + datetime.now().strftime("%Y%m%d_%H%M%S")
        scenario["meta"]["id"] = sid
        SCENARIOS_DIR.mkdir(exist_ok=True)
        (SCENARIOS_DIR/f"{sid}.json").write_text(json.dumps(scenario,ensure_ascii=False,indent=2))

        return {
            "ok": True,
            "scenario": scenario,
            "filename": f"{sid}.json",
            # v2316 — Renvoyer les warnings au frontend pour information
            "validation_warnings": warns,
        }
    except json.JSONDecodeError as e:
        raise HTTPException(500,f"Réponse IA non valide JSON: {e}")
    except Exception as e:
        logger.error(f"Erreur génération IA: {e}")
        raise HTTPException(500,str(e))

# ── Injecteur ──────────────────────────────────────────────────────────────────
def _elapsed_s() -> int:
    if not _inj["t_start"] or not _inj["running"]: return _inj["t_frozen"]
    now = datetime.now(timezone.utc)
    total = (now - _inj["t_start"]).total_seconds()
    paused = _inj["t_paused_s"]
    if _inj["paused"] and _inj["pause_start"]:
        paused += (now - _inj["pause_start"]).total_seconds()
    return max(0,int(total-paused))

async def _find_chat_salon_id(client, base: str, hdr: dict, preferred: str = "général-local") -> int:
    """v2195 — Trouver dynamiquement l'ID du salon où poster les stimuli.

    Historique : en 2194 on cherchait "général" (territorial) en premier,
    mais les joueurs ouvrent spontanément "général-local" (leur canal
    d'établissement). Les messages arrivaient donc bien côté serveur mais
    dans un salon que les joueurs ne regardaient pas → "vert animateur,
    rien joueur".

    Ordre de préférence :
      1. "général-local" (canal local par défaut, ouvert par les joueurs)
      2. "général" (territorial, fallback)
      3. Premier salon local disponible
      4. Premier salon tout court
      5. 1 en dernier recours (ne devrait jamais arriver)

    Logs complets pour voir ce qu'on trouve.
    """
    try:
        rs = await client.get(f"{base}/api/v1/chat/salons", headers=hdr)
        if rs.status_code >= 300:
            logger.warning(f"_find_chat_salon_id GET /salons → HTTP {rs.status_code} sur {base} — "
                          f"fallback id=1. Réponse : {rs.text[:200]}")
            return 1
        salons = rs.json()
        if not salons:
            logger.warning(f"_find_chat_salon_id {base} : aucun salon retourné — fallback id=1")
            return 1
        # Log tous les salons vus pour diag
        noms = [f"{s.get('nom')}(id={s.get('id')},type={s.get('type')})" for s in salons]
        logger.info(f"_find_chat_salon_id {base} : salons disponibles = {noms}")

        # 1. general-local d'abord (celui que les joueurs ouvrent)
        for s in salons:
            if s.get("nom") == preferred:
                return int(s.get("id", 1))
        # 2. général (territorial)
        for s in salons:
            if s.get("nom") == "général":
                return int(s.get("id", 1))
        # 3. Premier salon local
        for s in salons:
            if s.get("type") == "local":
                return int(s.get("id", 1))
        # 4. Premier salon tout court
        return int(salons[0].get("id", 1))
    except Exception as e:
        logger.warning(f"_find_chat_salon_id échec sur {base}: {e} — fallback id=1")
    return 1


async def _do_inject(stimulus:dict, tok_instances:dict):
    sigle = stimulus.get("cible","")

    # v2.3.86 — Gestion des cibles "acteurs externes" (CERT_SANTE, ANSSI, ARS,
    # SAMU, préfecture, médias...). Ces acteurs narratifs n'ont pas d'instance
    # dédiée, mais leurs stimuli (messages "reçus de l'ARS") doivent quand même
    # arriver dans l'UI joueur. On les redirige vers la première instance active
    # en préservant l'identité de l'expéditeur dans le texte du message.
    if sigle not in EXO_INSTANCES:
        # Sigle inconnu : acteur narratif externe. On redirige vers la 1re
        # instance active (pour que le stimulus arrive chez les joueurs).
        pilote_sigles = list(tok_instances.keys())
        if not pilote_sigles:
            logger.error(f"[INJECT] Cible '{sigle}' inconnue et aucune instance active. "
                         f"Stimulus {stimulus.get('id','?')} abandonné.")
            _inj["errors"].append(stimulus.get("id","?"))
            _log_injection({
                "stimulus_id": stimulus.get("id","?"),
                "type": stimulus.get("type","?"),
                "cible": sigle,
                "status": "SIGLE_INCONNU",
                "http": 0,
                "titre": stimulus.get("titre","")[:100],
                "response": f"Sigle '{sigle}' inconnu. Sigles valides : {sorted(EXO_INSTANCES.keys())}. "
                            f"Aucune instance active pour rediriger ce message.",
                "salon_id": None, "route": "",
            })
            return {"ok":False, "status_code":0, "error":"sigle_inconnu"}
        nouveau_sigle = pilote_sigles[0]
        logger.info(f"[INJECT] Cible externe '{sigle}' → redirigée vers instance '{nouveau_sigle}' "
                    f"(acteur narratif, préservation du contexte dans le texte)")
        # Préfixer le payload pour préserver l'identité externe
        pl_rewrite = dict(stimulus.get("payload", {}))
        stype_rw = stimulus.get("type", "incident")
        if stype_rw == "message" and "expediteur" not in pl_rewrite:
            pl_rewrite["expediteur"] = sigle  # ex: "CERT_SANTE", "ARS"
        elif stype_rw == "incident" and not pl_rewrite.get("declarant_nom","").startswith(sigle):
            pl_rewrite["declarant_nom"] = f"[{sigle}] " + pl_rewrite.get("declarant_nom", "Acteur externe")
        # On mute le stimulus pour la suite du traitement
        stimulus = dict(stimulus)
        stimulus["cible"] = nouveau_sigle
        stimulus["payload"] = pl_rewrite
        sigle = nouveau_sigle

    port  = EXO_INSTANCES.get(sigle, 8660)
    base  = f"{EXO_HOST}:{port}"
    tok   = tok_instances.get(sigle,"")

    # v2.3.85 — Si pas de token (login raté au start, ou sigle ajouté à la volée),
    # on retente un login. Sans ça le stimulus part avec "Bearer " (vide) → 401,
    # l'UI le voit vert mais l'instance ne reçoit jamais rien.
    if not tok:
        logger.warning(f"[INJECT] Pas de token pour {sigle}:{port}, relogin à la volée...")
        tok = await _login_instance(sigle, port)
        if tok:
            tok_instances[sigle] = tok  # cache pour stimuli suivants
            logger.info(f"[INJECT] Relogin {sigle} OK")
        else:
            logger.error(f"[INJECT] Relogin {sigle}:{port} ECHEC — stimulus {stimulus.get('id','?')} abandonné")
            _inj["errors"].append(stimulus.get("id","?"))
            _log_injection({
                "stimulus_id": stimulus.get("id","?"),
                "type": stimulus.get("type","?"),
                "cible": sigle,
                "status": "NO_TOKEN",
                "http": 0,
                "titre": stimulus.get("titre","")[:100],
                "response": f"Login animateur impossible sur {sigle}:{port}. "
                            f"Vérifier que l'instance tourne et que dircrise/Exercice2026! existe.",
                "salon_id": None,
                "route": "",
            })
            return {"ok":False, "status_code":0, "error":"no_token"}

    hdr   = {"Authorization":f"Bearer {tok}","Content-Type":"application/json"}
    pl    = dict(stimulus.get("payload",{}))  # v2192 : copie pour ne pas muter le scénario
    stype = stimulus.get("type","incident")

    # v2192 — injection de défauts robustes pour éviter les 422 silencieux quand
    # le scénario (notamment ceux générés par IA) n'a pas tous les champs
    # obligatoires des Pydantic models côté API SCRIBE.
    if stype == "incident":
        pl.setdefault("site_id", sigle)
        pl.setdefault("declarant_nom", f"Animateur (stimulus {stimulus.get('id','?')})")
        pl.setdefault("fait", stimulus.get("titre") or "Stimulus exercice")
        pl.setdefault("analyse", "")
        pl.setdefault("type_crise", "TECHNIQUE")
        pl.setdefault("urgency", 2)
    elif stype == "transfert":
        pl.setdefault("etablissement_origine", sigle)
        pl.setdefault("redacteur", f"Animateur (stimulus {stimulus.get('id','?')})")
        pl.setdefault("statut", "EN_COURS")
        pl.setdefault("nom", "Patient exercice")
        pl.setdefault("ipp", f"EXO-{stimulus.get('id','?')}")
        if "motif" in pl and "commentaire" not in pl:
            pl["commentaire"] = "Motif: " + pl["motif"]
        # v2196 — Sans eta + horodatage_depart, le transfert est créé côté
        # instance mais n'apparaît PAS sur la carte SOINS (qui filtre
        # `statut==='EN_COURS' && eta`). On fournit des valeurs par défaut
        # réalistes : départ = maintenant, ETA = T+30min (ordre de grandeur
        # transfert inter-hospitalier typique). L'animateur peut toujours
        # surcharger via le payload du scénario.
        now = datetime.now(timezone.utc)
        if pl.get("statut") == "EN_COURS":
            pl.setdefault("horodatage_depart", now.isoformat())
            # ETA paramétrable via stimulus.payload.eta_min (minutes)
            eta_min = int(stimulus.get("payload", {}).get("eta_min", 30))
            pl.setdefault("eta", (now + timedelta(minutes=eta_min)).isoformat())
    elif stype == "message":
        # v2.3.88 — Les stimuli "message" sont désormais routés vers la
        # messagerie interne (broadcast-externe) et non plus vers le chat.
        # Raison : un message externe (ARS, CERT, SAMU) sémantiquement EST
        # un message dans la boîte mail des dirigeants, pas un chat public.
        # Cela correspond à la demande explicite utilisateur v2.3.88 :
        #   "stimulus type=message → notification inbox + message (vrai)"
        #   "stimulus type=incident → pas de message inbox, juste badge"
        # Fallback chat si broadcast-externe indisponible (ancienne instance).
        pl.setdefault("contenu", stimulus.get("titre") or "Message exercice")
        expediteur = pl.pop("expediteur", None) or pl.pop("emetteur", None) or "Acteur externe"
        sujet = pl.pop("sujet", None) or (stimulus.get("titre") or "Message exercice")[:80]
        # Format final : expediteur externe + sujet + contenu
        # On préserve expediteur dans un champ dédié pour la nouvelle route.
        pl["_expediteur_externe"] = expediteur
        pl["_sujet_externe"] = sujet
        # Rétro-compat : si on doit fallback vers chat, on préfixe.
        pl["_chat_fallback_contenu"] = f"[{expediteur}] {sujet} : {pl['contenu']}"
    elif stype == "decision":
        pl.setdefault("contenu", stimulus.get("titre") or "Décision exercice")
        pl.setdefault("responsable", "Cellule de crise")
        pl.setdefault("base_reglementaire", "Plan Blanc")

    actual_salon_id = None
    route_called = ""
    # v2204 — Log le payload EXACT envoyé pour diagnostic rapide. Sans ça,
    # on ne voit jamais ce qui part vraiment quand un stimulus échoue
    # silencieusement. Tronqué à 500 chars pour éviter flood.
    try:
        _payload_preview = json.dumps(pl, ensure_ascii=False)[:500]
    except Exception:
        _payload_preview = str(pl)[:500]
    logger.info(f"[INJECT] {stimulus.get('id','?')} type={stype} cible={sigle} "
                f"port={port} token_ok={'oui' if tok else 'NON'} "
                f"payload={_payload_preview}")
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            if stype == "incident":
                route_called = "/api/v1/sitrep/post"
                r = await client.post(f"{base}{route_called}", json=pl, headers=hdr)
            elif stype == "message":
                # v2.3.88 — Routage vers messagerie interne (broadcast-externe).
                # Avantage : le message atterrit dans la VRAIE inbox des
                # dirigeants, avec notification, exactement comme un mail
                # entrant depuis un acteur externe (ARS, CERT, SAMU).
                route_called = "/api/v1/messagerie/broadcast-externe"
                body_msg = {
                    "expediteur_nom": pl.get("_expediteur_externe", "Acteur externe"),
                    "sujet":          pl.get("_sujet_externe", "Message exercice")[:200],
                    "contenu":        pl.get("contenu", ""),
                }
                logger.info(f"Stimulus {stimulus.get('id','?')} message → "
                            f"messagerie externe ({body_msg['expediteur_nom']})")
                r = await client.post(f"{base}{route_called}",
                                      json=body_msg, headers=hdr)
                # Fallback chat si 404 (route pas encore déployée sur cette instance)
                if r.status_code == 404:
                    logger.warning(f"Route broadcast-externe absente, fallback chat salon")
                    actual_salon_id = await _find_chat_salon_id(client, base, hdr)
                    route_called = f"/api/v1/chat/salons/{actual_salon_id}/messages"
                    r = await client.post(f"{base}{route_called}",
                        json={"contenu": pl.get("_chat_fallback_contenu",
                                                pl.get("contenu","")),
                              "mentions":[]}, headers=hdr)
            elif stype == "transfert":
                route_called = "/api/v1/transferts"
                r = await client.post(f"{base}{route_called}", json=pl, headers=hdr)
                # v2203 — Dupliquer le transfert côté destinataire pour qu'il
                # apparaisse à la fois sur l'émetteur (transfert sortant) et
                # sur le récepteur (transfert entrant). Auparavant seul le
                # côté "cible" du stimulus voyait le transfert.
                dest_sigle = pl.get("etablissement_destination")
                if dest_sigle and dest_sigle != sigle and dest_sigle in tok_instances:
                    try:
                        dest_port = EXO_INSTANCES.get(dest_sigle)
                        dest_tok = tok_instances.get(dest_sigle, "")
                        if dest_port and dest_tok:
                            dest_base = f"{EXO_HOST}:{dest_port}"
                            dest_hdr = {"Authorization":f"Bearer {dest_tok}","Content-Type":"application/json"}
                            # Payload miroir : l'émetteur de l'origine devient destination
                            pl_mirror = dict(pl)
                            pl_mirror["etablissement_destination"] = dest_sigle
                            # Le destinataire voit le transfert comme entrant
                            # On pose un flag pour éventuelle distinction UI
                            pl_mirror.setdefault("sens", "ENTRANT")
                            r_mirror = await client.post(f"{dest_base}/api/v1/transferts",
                                                        json=pl_mirror, headers=dest_hdr)
                            logger.info(f"Stimulus {stimulus.get('id','?')} transfert "
                                       f"aussi créé côté destinataire {dest_sigle}: "
                                       f"HTTP {r_mirror.status_code}")
                    except Exception as e:
                        logger.warning(f"Transfert miroir {dest_sigle} échec: {e}")
            elif stype == "chat":
                actual_salon_id = pl.get("salon_id")
                if not actual_salon_id:
                    actual_salon_id = await _find_chat_salon_id(client, base, hdr)
                route_called = f"/api/v1/chat/salons/{actual_salon_id}/messages"
                logger.info(f"Stimulus {stimulus.get('id','?')} chat → salon_id={actual_salon_id}")
                r = await client.post(f"{base}{route_called}",
                    json={"contenu":pl.get("contenu",""),"mentions":[]}, headers=hdr)
            elif stype == "decision":
                route_called = "/api/v1/cellule/decisions"
                r = await client.post(f"{base}{route_called}", json=pl, headers=hdr)
            elif stype == "capacite":
                # v2309 — Stimulus capacité : crée des tensions réalistes
                # sur les lits / RH / matériel d'un service. Approche :
                # on déclare une nouvelle situation capacitaire via
                # /api/v1/capacite/declaration, qui elle-même peut
                # déclencher un incident automatique si alerte_* = true.
                # L'intérêt pédagogique : le directeur de crise VOIT son
                # indicateur capacité basculer au rouge, il doit réagir
                # (reporter chirurgies, transferts, renforts...).
                #
                # Payload attendu dans le scénario :
                #   unite:           nom de l'UF (doit exister dans le
                #                    référentiel capacité, ex: "Orthopédie")
                #   statut_lits:     normal|tension|critique|ferme
                #   statut_rh:       complet|tension|insuffisant|degrade
                #   statut_materiel: ok|degrade|hs
                #   lits_vides_h/f/i: nombre de lits disponibles (optionnel)
                #   besoin_renfort:  int (agents manquants) (optionnel)
                #   peut_preter:     int (agents prêtables, ex: coursiers
                #                    dispo pour un autre service)
                #   commentaire_*:   contexte narratif pour les joueurs
                #
                # Le collecteur résout le nom d'unité → referentiel_id en
                # interrogeant /api/v1/capacite/referentiel avant d'envoyer
                # la déclaration.
                route_called = "/api/v1/capacite/declaration"
                unite_nom = (pl.get("unite") or pl.get("uf") or "").strip()
                ref_id = None
                try:
                    r_ref = await client.get(f"{base}/api/v1/capacite/referentiel",
                                             headers=hdr)
                    if r_ref.status_code == 200:
                        refs = r_ref.json() if isinstance(r_ref.json(), list) else r_ref.json().get("items", [])
                        # v2309-hotfix — Le référentiel expose "service_nom"
                        # (et optionnellement "uf_code"), pas "nom". Avant
                        # le match ne trouvait JAMAIS rien → fallback sur
                        # ref_id=1 ou erreur. Maintenant :
                        #   1. Match exact service_nom (prioritaire)
                        #   2. Match exact uf_code (code d'unité)
                        #   3. Match partiel service_nom (inclusion)
                        u_low = unite_nom.lower()
                        # 1. Match exact sur service_nom
                        for ref in refs:
                            nom = (ref.get("service_nom") or ref.get("nom") or "").strip()
                            if nom.lower() == u_low:
                                ref_id = ref.get("id")
                                break
                        # 2. Match exact sur uf_code
                        if ref_id is None:
                            for ref in refs:
                                code = (ref.get("uf_code") or "").strip()
                                if code and code.lower() == u_low:
                                    ref_id = ref.get("id")
                                    logger.info(
                                        f"Stimulus capacité : match par uf_code "
                                        f"'{unite_nom}' → '{ref.get('service_nom')}' (id={ref_id})"
                                    )
                                    break
                        # 3. Match partiel service_nom
                        if ref_id is None:
                            for ref in refs:
                                nom = (ref.get("service_nom") or "").strip().lower()
                                if nom and (nom.startswith(u_low) or u_low in nom or nom in u_low):
                                    ref_id = ref.get("id")
                                    logger.info(
                                        f"Stimulus capacité : match partiel "
                                        f"'{unite_nom}' → '{ref.get('service_nom')}' (id={ref_id})"
                                    )
                                    break
                except Exception as e:
                    logger.warning(f"Impossible de résoudre unité capacité '{unite_nom}': {e}")
                if (ref_id is None):
                    # v2309-hotfix — Avant : fallback ref_id=1 qui polluait
                    # une unité aléatoire du référentiel. Désormais : on
                    # échoue proprement avec un message d'erreur visible
                    # dans le log d'injection. L'animateur voit tout de
                    # suite qu'il faut créer l'unité ou ajuster le nom.
                    err_msg = (
                        f"Stimulus capacité : unité '{unite_nom}' introuvable "
                        f"dans le référentiel de l'instance {sigle}. "
                        f"Créer l'unité dans Capacité > Référentiel ou ajuster "
                        f"le nom dans le scénario. Unités disponibles : "
                        + ", ".join(sorted(set([
                            (ref.get("service_nom") or ref.get("nom") or "")
                            for ref in refs[:20]
                            if (ref.get("service_nom") or ref.get("nom"))
                        ])))[:300]
                    )
                    logger.warning(err_msg)
                    _inj["errors"].append(stimulus.get("id","?"))
                    _log_injection({
                        "stimulus_id": stimulus.get("id","?"),
                        "type": stype, "cible": sigle,
                        "status": "ERR", "http": 0,
                        "titre": stimulus.get("titre","")[:100],
                        "response": err_msg[:500],
                        "salon_id": None, "route": route_called,
                    })
                    return {"ok": False, "error": err_msg}
                body_cap = {
                    "referentiel_id": ref_id,
                    "redacteur":      pl.get("redacteur", "Stimulus exercice"),
                    "point":          pl.get("point", "matin"),
                    "lits_vides_h":   int(pl.get("lits_vides_h", 0)),
                    "lits_vides_f":   int(pl.get("lits_vides_f", 0)),
                    "lits_vides_i":   int(pl.get("lits_vides_i", 0)),
                    "tension_activee":int(pl.get("tension_activee", 0)),
                    "lits_sup":       int(pl.get("lits_sup", 0)),
                    "statut_lits":    pl.get("statut_lits", "normal"),
                    "statut_rh":      pl.get("statut_rh", "complet"),
                    "statut_materiel":pl.get("statut_materiel", "ok"),
                    "alerte_lits":    bool(pl.get("alerte_lits", False)),
                    "alerte_rh":      bool(pl.get("alerte_rh", False)),
                    "alerte_materiel":bool(pl.get("alerte_materiel", False)),
                    "commentaire_lits":     pl.get("commentaire_lits"),
                    "commentaire_rh":       pl.get("commentaire_rh"),
                    "commentaire_materiel": pl.get("commentaire_materiel"),
                    "commentaire_general":  pl.get("commentaire_general"),
                    "mode_degrade":   bool(pl.get("mode_degrade", False)),
                    "besoin_renfort": int(pl.get("besoin_renfort", 0)),
                    "peut_preter":    int(pl.get("peut_preter", 0)),
                }
                logger.info(
                    f"Stimulus {stimulus.get('id','?')} capacité → {unite_nom} "
                    f"(ref={ref_id}) lits={body_cap['statut_lits']} "
                    f"rh={body_cap['statut_rh']} mat={body_cap['statut_materiel']}"
                )
                r = await client.post(f"{base}{route_called}",
                                      json=body_cap, headers=hdr)
            elif stype == "brancardage":
                # v2309 — Stimulus brancardage : crée une mission de
                # transport patient. Permet de simuler des pics
                # d'activité, des priorités P1 en pleine crise, et
                # d'observer comment l'équipe gère les urgences
                # simultanées. Payload attendu dans le scénario :
                #   ref_patient:      identifiant / pseudonyme patient
                #   uf_origine:       service de départ
                #   uf_destination:   service d'arrivée
                #   priorite:         P1|P2|P3 (P1 = urgence vitale)
                #   motif:            raison médicale
                #   type_transport:   BRANCARD|FAUTEUIL|MARCHE|LIT
                #   commentaire:      contexte narratif animateur
                route_called = "/api/v1/brancardage/missions"
                body_br = {
                    "ref_type":       pl.get("ref_type", "IPP"),
                    "ref_patient":    pl.get("ref_patient", "EXO-" + stimulus.get("id","?")),
                    "uf_origine":     pl.get("uf_origine", "?"),
                    "uf_destination": pl.get("uf_destination", "?"),
                    "chambre_depart":  pl.get("chambre_depart"),
                    "chambre_arrivee": pl.get("chambre_arrivee"),
                    "type_transport": pl.get("type_transport", "BRANCARD"),
                    "priorite":       pl.get("priorite", "P2"),
                    "motif":          pl.get("motif", stimulus.get("titre","")),
                    "commentaire":    pl.get("commentaire"),
                    "programmee":     int(pl.get("programmee", 0)),
                    "heure_prevue":   pl.get("heure_prevue"),
                    "avec_retour":    int(pl.get("avec_retour", 0)),
                    "heure_retour":   pl.get("heure_retour"),
                }
                logger.info(
                    f"Stimulus {stimulus.get('id','?')} brancardage → "
                    f"{body_br['uf_origine']} → {body_br['uf_destination']} "
                    f"priorité {body_br['priorite']}"
                )
                r = await client.post(f"{base}{route_called}",
                                      json=body_br, headers=hdr)
            else:
                logger.warning(f"[INJECT] Type stimulus inconnu: {stype} (stimulus {stimulus.get('id','?')})")
                _inj["errors"].append(stimulus.get("id","?"))
                _log_injection({
                    "stimulus_id": stimulus.get("id","?"),
                    "type": stype, "cible": sigle,
                    "status": "TYPE_INCONNU", "http": 0,
                    "titre": stimulus.get("titre","")[:100],
                    "response": f"Type '{stype}' non géré. Types valides: incident, message, transfert, chat, decision, capacite, brancardage.",
                    "salon_id": None, "route": "",
                })
                return {"ok":False,"error":f"Type inconnu: {stype}"}

        ok = r.status_code < 300
        try:
            resp_preview = r.json() if r.headers.get("content-type","").startswith("application/json") else r.text[:200]
        except Exception:
            resp_preview = r.text[:200]
        if ok:
            _inj["done"].append(stimulus["id"])
            logger.info(
                f"Stimulus {stimulus['id']} ({stype}) → {sigle} OK HTTP {r.status_code} — "
                f"réponse instance : {str(resp_preview)[:300]}"
            )
        else:
            _inj["errors"].append(stimulus["id"])
            # v2316 — Diagnostic enrichi pour HTTP 422 (Pydantic).
            # Avant : le log indiquait juste "HTTP 422 — payload: ..." sans
            # détailler quel champ posait problème. Maintenant on extrait
            # les erreurs Pydantic structurées pour pouvoir corriger.
            err_detail = ""
            if r.status_code == 422:
                try:
                    pyd_errors = r.json().get("detail", [])
                    if isinstance(pyd_errors, list):
                        # Format Pydantic : [{"loc":["body","field"],"msg":"...","type":"..."}]
                        bad_fields = []
                        for pe in pyd_errors[:3]:
                            loc = ".".join(str(x) for x in pe.get("loc", [])[1:])  # skip "body"
                            msg = pe.get("msg", "")
                            bad_fields.append(f"{loc}: {msg}")
                        err_detail = " — Champs invalides : " + " | ".join(bad_fields)
                except Exception:
                    pass
                logger.warning(
                    f"[INJECT 422] Stimulus {stimulus['id']} ({stype}) → {sigle} : "
                    f"validation Pydantic échouée.{err_detail} — payload envoyé: {pl}"
                )
            else:
                logger.warning(
                    f"Stimulus {stimulus['id']} ({stype}) → {sigle}: "
                    f"HTTP {r.status_code} — réponse: {r.text[:300]} — payload: {pl}"
                )
        # v2202 — buffer de diag pour injection-log
        _log_injection({
            "stimulus_id": stimulus.get("id","?"),
            "type": stype,
            "cible": sigle,
            "status": "OK" if ok else ("PYDANTIC_422" if r.status_code == 422 else "ERR"),
            "http": r.status_code,
            "titre": stimulus.get("titre","")[:100],
            "response": str(resp_preview)[:300],
            "salon_id": actual_salon_id,
            "route": route_called,
        })
        detail = r.text[:300] if not ok else ""
        return {"ok":ok, "status_code":r.status_code, "error": detail if not ok else None}
    except Exception as e:
        _inj["errors"].append(stimulus["id"])
        logger.error(f"Inject {stimulus['id']} → {sigle}: {e}")
        _log_injection({
            "stimulus_id": stimulus.get("id","?"),
            "type": stype, "cible": sigle,
            "status": "EXC", "http": 0,
            "titre": stimulus.get("titre","")[:100],
            "response": f"Exception: {e}",
        })
        return {"ok":False,"error":str(e)}

async def _run_injector(scenario:dict, tok_instances:dict):
    """Scheduler principal d'injection des stimuli selon t_min + ratio.

    v2.3.86 — Robustesse majeure :
    - try/except PAR stimulus : une erreur sur S02 ne bloque plus S03+.
      Avant : asyncio mangeait silencieusement les exceptions, la tâche
      mourait après S01, l'UI restait figée à "En cours..." sur tous les
      suivants. Symptôme exact : "que le premier arrive".
    - Log détaillé des erreurs de scheduler (distinct des erreurs HTTP).
    - On marque le stimulus en erreur même si l'exception est dans le
      scheduler lui-même (pas juste dans _do_inject).
    """
    stimuli = sorted(scenario.get("stimuli",[]), key=lambda s:s["t_min"])
    ratio   = scenario.get("meta",{}).get("ratio_compression",4.0)
    disabled = set(_inj.get("disabled_stimuli",[]))
    logger.info(f"[SCHEDULER] Démarrage : {len(stimuli)} stimuli, ratio={ratio}, "
                f"instances actives={list(tok_instances.keys())}")
    for stimulus in stimuli:
        if not _inj["running"]: break
        sid = stimulus.get("id","?")
        if sid in disabled:
            _inj["done"].append(sid)
            logger.info(f"[SCHEDULER] {sid} désactivé, skip")
            continue
        try:
            target_s = int(stimulus["t_min"] * 60 / ratio)
            while True:
                if not _inj["running"]: return
                while _inj["paused"]: await asyncio.sleep(2)
                if _elapsed_s() >= target_s: break
                await asyncio.sleep(3)
            if sid not in _inj["done"]:
                logger.info(f"[SCHEDULER] Injection {sid} (t_min={stimulus['t_min']}, "
                            f"target={target_s}s, elapsed={_elapsed_s()}s)")
                await _do_inject(stimulus, tok_instances)
        except Exception as e:
            # v2.3.86 — Filet de sécurité : une erreur sur 1 stimulus ne doit
            # pas arrêter tout le scénario.
            logger.exception(f"[SCHEDULER] ERREUR sur {sid}: {e}")
            if sid not in _inj["errors"]:
                _inj["errors"].append(sid)
            try:
                _log_injection({
                    "stimulus_id": sid,
                    "type": stimulus.get("type","?"),
                    "cible": stimulus.get("cible","?"),
                    "status": "SCHEDULER_EXC",
                    "http": 0,
                    "titre": stimulus.get("titre","")[:100],
                    "response": f"Exception scheduler: {type(e).__name__}: {e}",
                    "salon_id": None, "route": "",
                })
            except Exception:
                pass
            # On continue avec le stimulus suivant
            continue
    _inj["running"] = False
    logger.info(f"[SCHEDULER] Scénario terminé. "
                f"Injectés OK: {len(_inj['done'])}, Erreurs: {len(_inj['errors'])}")

class StartRequest(BaseModel):
    filename:str; sites:list; login_instances:Optional[dict]=None
    disabled_stimuli:list=[]

async def _login_instance(sigle:str, port:int) -> str:
    """Login animateur sur l'instance exercice, retourne le JWT."""
    try:
        async with httpx.AsyncClient(timeout=8.0) as client:
            r = await client.post(f"{EXO_HOST}:{port}/api/v1/auth/login",
                json={"username":"dircrise","password":"Exercice2026!"},
                headers={"Content-Type":"application/json"})
        if r.status_code == 200: return r.json().get("token","")
        logger.warning(f"Login {sigle}:{port} → {r.status_code}")
        return ""
    except Exception as e:
        logger.error(f"Login {sigle}:{port}: {e}"); return ""

@app.post("/api/exercice/start")
async def start_exercice(body:StartRequest, auth=Depends(require_auth)):
    global _inj_task, _scenario_actif, _token_instances
    if _inj["running"]: raise HTTPException(409,"Exercice déjà en cours")
    
    p = SCENARIOS_DIR / body.filename
    if not p.exists(): raise HTTPException(404,"Scénario non trouvé")
    scenario = json.loads(p.read_text(encoding="utf-8"))
    _scenario_actif = scenario
    
    # Login sur toutes les instances actives
    _token_instances = {}
    for sigle in body.sites:
        port = EXO_INSTANCES.get(sigle)
        if port:
            tok = await _login_instance(sigle, port)
            if tok:
                _token_instances[sigle] = tok
                logger.info(f"Token {sigle}:{port} OK")
            else:
                # v2.3.85 — NE PAS ajouter la clé si login raté. Sinon _do_inject
                # envoie "Bearer " (vide) → 401, stimulus vert côté UI mais
                # jamais reçu par l'instance. Symptôme classique.
                logger.error(f"Token {sigle}:{port} ECHEC — "
                             f"stimuli vers {sigle} échoueront. "
                             f"Vérifier que l'instance tourne et que "
                             f"dircrise/Exercice2026! existe.")
    
    _inj.update({"running":True,"paused":False,"session_id":f"exo_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
        "scenario":scenario,"t_start":datetime.now(timezone.utc),
        "t_paused_s":0.0,"pause_start":None,"done":[],"errors":[],"t_frozen":0,
        "disabled_stimuli":body.disabled_stimuli})
    
    loop = asyncio.get_event_loop()
    _inj_task = loop.create_task(_run_injector(scenario, _token_instances))
    return {"ok":True,"session_id":_inj["session_id"],"tokens_ok":list(_token_instances.keys())}

@app.post("/api/exercice/pause")
async def pause_exercice(auth=Depends(require_auth)):
    if not _inj["running"] or _inj["paused"]: raise HTTPException(400)
    _inj["paused"] = True; _inj["pause_start"] = datetime.now(timezone.utc)
    return {"ok":True}

@app.post("/api/exercice/resume")
async def resume_exercice(auth=Depends(require_auth)):
    if not _inj["paused"]: raise HTTPException(400)
    if _inj["pause_start"]:
        _inj["t_paused_s"] += (datetime.now(timezone.utc)-_inj["pause_start"]).total_seconds()
    _inj["paused"] = False; _inj["pause_start"] = None
    return {"ok":True}

@app.post("/api/exercice/stop")
async def stop_exercice(auth=Depends(require_auth)):
    global _inj_task, _scenario_actif
    _inj["t_frozen"] = _elapsed_s()
    _inj["running"] = False; _inj["paused"] = False
    if _inj_task and not _inj_task.done(): _inj_task.cancel()
    # v2196 — Reset complet de l'état d'injection pour que le prochain
    # scénario sélectionné parte d'une ardoise propre. Sans ça,
    # get_exercice_status continue à retourner les stimuli de l'ancien
    # scénario avec leurs états done/error, et quand on Utilise un
    # nouveau scénario on voyait les stimuli fantômes.
    _inj["scenario"] = None
    _inj["done"] = []
    _inj["errors"] = []
    _inj["session_id"] = None
    _inj["t_start"] = None
    _inj["t_frozen"] = 0
    _scenario_actif = None
    return {"ok":True}

@app.post("/api/exercice/inject/{stimulus_id}")
async def inject_manual(stimulus_id:str, auth=Depends(require_auth)):
    if not _scenario_actif: raise HTTPException(400,"Aucun scénario actif")
    st = next((s for s in _scenario_actif.get("stimuli",[]) if s["id"]==stimulus_id),None)
    if not st: raise HTTPException(404,f"Stimulus {stimulus_id} non trouvé")
    result = await _do_inject(st, _token_instances)
    return result


# v2.3.102 — Reset complet d'exercice. Efface tout sur toutes les instances
# et sur le collecteur lui-même. Appelé par le bouton "🧹 RESET TOUT" de
# l'UI animateur. N'a PAS besoin que l'exercice soit en cours.
@app.post("/api/exercice/reset-all")
async def reset_all_exercice(auth=Depends(require_auth)):
    global _inj_task, _scenario_actif, transferts_inter, _injection_log

    errors = []
    instances_ok = 0
    instances_total = 0

    # 1. Arrêter le scheduler s'il tourne
    try:
        _inj["t_frozen"] = 0
        _inj["running"] = False
        _inj["paused"] = False
        if _inj_task and not _inj_task.done():
            _inj_task.cancel()
    except Exception as e:
        errors.append(f"Arrêt scheduler: {e}")

    # 2. Pour chaque instance active : login + appel /admin/reset-exercice
    #    (on refait un login frais au cas où les tokens en cache seraient
    #    expirés — robuste pour un reset post-crash).
    for sigle, port in EXO_INSTANCES.items():
        instances_total += 1
        try:
            tok = await _login_instance(sigle, port)
            if not tok:
                errors.append(f"{sigle}: login impossible")
                continue
            async with httpx.AsyncClient(timeout=20.0) as client:
                r = await client.delete(
                    f"{EXO_HOST}:{port}/api/v1/admin/reset-exercice",
                    headers={"Authorization": f"Bearer {tok}"},
                )
                if r.status_code < 300:
                    instances_ok += 1
                    logger.info(f"[RESET-ALL] {sigle}:{port} ✓ "
                                f"({r.json() if r.headers.get('content-type','').startswith('application/json') else 'ok'})")
                else:
                    errors.append(f"{sigle}: HTTP {r.status_code}")
        except Exception as e:
            errors.append(f"{sigle}: {type(e).__name__}: {str(e)[:80]}")

    # 3. Nettoyer l'état du collecteur lui-même
    collecteur_cleaned = False
    try:
        transferts_inter.clear()
        save_transferts_inter()
        _injection_log.clear()
        # Reset injection state
        _inj["scenario"] = None
        _inj["done"] = []
        _inj["errors"] = []
        _inj["session_id"] = None
        _inj["t_start"] = None
        _inj["t_frozen"] = 0
        _inj["disabled_stimuli"] = []
        _scenario_actif = None
        collecteur_cleaned = True
        logger.info("[RESET-ALL] Collecteur vidé")
    except Exception as e:
        errors.append(f"Collecteur: {e}")

    return {
        "ok": True,
        "instances_ok": instances_ok,
        "instances_total": instances_total,
        "collecteur_cleaned": collecteur_cleaned,
        "errors": errors,
    }


# v2202 — Log d'injection en direct pour diagnostic rapide. Retourne les
# 100 derniers événements avec la route exacte appelée, le status HTTP,
# le salon_id utilisé (pour les messages/chat) et la réponse de l'instance.
# Permet de détecter immédiatement :
#   - Stimulus "vert" mais HTTP 500 caché côté serveur
#   - Salon_id qui n'est pas celui qu'on attendait
#   - Réponse instance vide/bizarre
@app.get("/api/exercice/injection-log")
async def get_injection_log(auth=Depends(require_auth)):
    return {"entries": list(_injection_log), "count": len(_injection_log)}


# v2201 — Injection d'un stimulus ad-hoc créé à la volée par l'animateur.
# Permet de déstabiliser les équipes avec un événement qui n'était pas dans
# le scénario (ex: pénurie de carburant pendant un transfert programmé).
# Pas besoin d'un scénario actif pour utiliser cette route : l'animateur
# peut injecter des stimuli même en "exercice libre".
class InjectAdhocRequest(BaseModel):
    type: str  # incident | message | transfert | chat | decision | capacite
    cible: str  # sigle d'établissement (doit être dans _token_instances)
    titre: str = ""
    payload: dict = {}
    description_animateur: str = ""
    action_attendue: str = ""


@app.post("/api/exercice/inject-adhoc")
async def inject_adhoc(body: InjectAdhocRequest, auth=Depends(require_auth)):
    """Injecte un stimulus créé à la volée, sans passer par un scénario."""
    # v2204 — Permettre l'injection même sans exercice démarré.
    # Si _token_instances est vide (pas d'exercice actif), on fait un
    # login "à la volée" sur l'instance cible. Ça permet de tester en
    # mode G7 nominal ou mode démo sans avoir à démarrer un scénario.
    global _token_instances
    if body.cible not in _token_instances:
        # Tenter un login à la volée
        port = EXO_INSTANCES.get(body.cible)
        if not port:
            raise HTTPException(
                400,
                f"Sigle '{body.cible}' inconnu. Sigles valides : {sorted(EXO_INSTANCES.keys())}"
            )
        logger.info(f"inject-adhoc : login à la volée sur {body.cible}:{port}")
        tok = await _login_instance(body.cible, port)
        if not tok:
            raise HTTPException(
                503,
                f"Instance '{body.cible}' (port {port}) injoignable ou login échoué. "
                f"Vérifier que l'instance tourne et que le compte 'dircrise' existe."
            )
        _token_instances[body.cible] = tok

    allowed_types = {"incident", "message", "transfert", "chat", "decision", "capacite", "brancardage"}
    if body.type not in allowed_types:
        raise HTTPException(400, f"Type invalide. Autorisés : {sorted(allowed_types)}")

    # Construire un stimulus synthétique conforme à la structure scénario
    import uuid as _uuid
    stimulus = {
        "id": f"MANUEL-{_uuid.uuid4().hex[:8].upper()}",
        "type": body.type,
        "t_min": 0,
        "cible": body.cible,
        "titre": body.titre or f"Stimulus manuel {body.type}",
        "description_animateur": body.description_animateur,
        "action_attendue": body.action_attendue,
        "payload": body.payload or {},
    }
    logger.info(f"Stimulus manuel injecté par animateur : {stimulus['id']} "
                f"type={body.type} cible={body.cible}")

    result = await _do_inject(stimulus, _token_instances)

    # v2204 — Si on a créé le token à la volée (pas d'exercice actif),
    # on ne le garde pas en mémoire pour ne pas polluer le prochain start
    if not _inj.get("running"):
        _token_instances.pop(body.cible, None)

    # Ajouter aussi le stimulus au scénario courant s'il y en a un
    if _inj.get("scenario"):
        try:
            _inj["scenario"].setdefault("stimuli", []).append(stimulus)
        except Exception:
            pass
    return {**result, "stimulus_id": stimulus["id"]}


# v2309-hotfix — Route publique exposant la liste des sites engagés dans
# l'exercice courant. Utilisée par les instances joueur pour filtrer la
# liste des destinataires de la messagerie : plutôt que d'afficher tout
# l'univers SCRIBE, on ne montre que les sites concernés par l'exo, ce
# qui colle mieux au périmètre de décision du directeur de crise.
# Pas d'auth : la simple liste des sigles n'est pas sensible.
# v2309-hotfix — Relay du référentiel capacité d'une instance cible vers
# l'UI animateur, pour alimenter le dropdown multi-sélection des stimuli
# capacité. Requiert que la cible soit joignable et que dircrise/Exercice2026!
# y existe. Fail-safe : retourne une liste vide + message si indisponible.
@app.get("/api/exercice/unites-capacite")
async def get_unites_capacite(cible: str, auth=Depends(require_auth)):
    """Liste les unités du référentiel capacité d'une instance cible.
    Utilisé par le modal stimulus manuel pour offrir la multi-sélection."""
    import httpx as _httpx
    # Résolution port depuis le scénario actif (meilleure source)
    scenario = _inj.get("scenario") or {}
    port = None
    for a in scenario.get("acteurs", []):
        if (a.get("sigle") or "").upper() == cible.upper():
            port = int(a.get("port", 8660))
            break
    # Fallback : si pas de scénario actif, essayer 8660 par défaut
    if port is None:
        port = 8660
    base = f"http://localhost:{port}"
    # Token instance : essayer le cache, sinon login
    tok = _token_instances.get(cible.upper()) or _token_instances.get(cible)
    try:
        async with _httpx.AsyncClient(timeout=5.0) as client:
            if not tok:
                r_login = await client.post(
                    f"{base}/api/v1/auth/login",
                    json={"username": "dircrise", "password": "Exercice2026!"},
                )
                if r_login.status_code != 200:
                    return {"unites": [], "error": f"Login échoué sur {cible} (port {port})"}
                tok = r_login.json().get("token", "")
                _token_instances[cible.upper()] = tok
            r = await client.get(
                f"{base}/api/v1/capacite/referentiel",
                headers={"Authorization": f"Bearer {tok}"},
            )
            if r.status_code != 200:
                return {"unites": [], "error": f"HTTP {r.status_code}"}
            data = r.json()
            items = data if isinstance(data, list) else data.get("items", [])
            # Ne garder que les champs utiles
            unites = [
                {
                    "id": u.get("id"),
                    "service_nom": u.get("service_nom") or u.get("nom"),
                    "uf_code": u.get("uf_code"),
                    "pole": u.get("pole"),
                    "site": u.get("site"),
                }
                for u in items
                if (u.get("service_nom") or u.get("nom"))
            ]
            return {"unites": unites}
    except Exception as e:
        return {"unites": [], "error": str(e)[:200]}


@app.get("/api/exercice/sites-actifs")
async def get_sites_actifs_public():
    """Liste publique des sites actuellement engagés dans l'exercice.
    Réponse minimale pour limiter la surface d'exposition :
      {
        "running": bool,
        "sites": ["CHAG", "HDLEMAN", ...]
      }
    Si aucun exercice actif : running=False, sites=[].
    """
    scenario = _inj.get("scenario")
    sites_scenario = []
    if scenario:
        for a in scenario.get("acteurs", []):
            sigle = a.get("sigle") or ""
            if sigle and sigle not in sites_scenario:
                sites_scenario.append(sigle)
    return {
        "running": _inj.get("running", False),
        "sites": sites_scenario or list(_token_instances.keys()),
    }


# v2312 — Agrégation des statuts publics de tous les sites engagés dans
# l'exercice. Permet à l'animateur d'observer comment les joueurs
# utilisent leur page de statut public, qui est une fonction
# stratégique (décision de publier tel ou tel niveau d'alerte vers
# l'extérieur, mise à jour de la chronologie, etc.).
# Accessible uniquement au collecteur authentifié (require_auth).
@app.get("/api/exercice/statuts-publics")
async def get_statuts_publics_sites(auth=Depends(require_auth)):
    """Agrège le statut public de chaque site exercice pour vue pédagogique
    animateur. Appel direct /api/v1/status/public sur chaque instance
    (pas d'auth requise côté instance pour cet endpoint, statut public
    par nature).
    """
    import httpx as _httpx
    scenario = _inj.get("scenario") or {}
    acteurs = scenario.get("acteurs", [])
    if not acteurs:
        return {"running": False, "sites": []}

    result = []
    async with _httpx.AsyncClient(timeout=3.0) as client:
        for a in acteurs:
            sigle = (a.get("sigle") or "").upper()
            port = int(a.get("port", 0))
            if not sigle or not port:
                continue
            site = {
                "sigle": sigle,
                "nom": a.get("nom_etablissement") or sigle,
                "port": port,
                "published": False,
                "etat": "—",
                "message": "",
                "derniere_maj": None,
                "nb_chronologie": 0,
                "reachable": False,
                "error": None,
            }
            try:
                r = await client.get(f"http://localhost:{port}/api/v1/status/public")
                if r.status_code == 200:
                    d = r.json()
                    site["reachable"] = True
                    site["published"] = bool(d.get("published"))
                    site["etat"] = d.get("etat") or d.get("niveau") or "—"
                    site["message"] = (d.get("message") or "")[:200]
                    site["derniere_maj"] = d.get("updated_at") or d.get("derniere_maj")
                    site["nb_chronologie"] = len(d.get("chronologie") or [])
                elif r.status_code == 404:
                    site["error"] = "Route /status non exposée (instance ancienne)"
                else:
                    site["error"] = f"HTTP {r.status_code}"
            except Exception as e:
                site["error"] = str(e)[:80]
            result.append(site)
    return {
        "running": _inj.get("running", False),
        "sites": result,
    }


# v2312 — Liste des instances actuellement joignables, indexée par sigle.
# Utilisée par l'UI animateur pour afficher un état "actif / inactif /
# externe" dans la modal de détail de scénario. Pas de secret exposé.
# Parcourt les ports standards 8660-8666 + les ports déclarés dans le
# scénario actif s'il y en a, et teste leur disponibilité.
@app.get("/api/exercice/instances-up")
async def get_instances_up(auth=Depends(require_auth)):
    """Retourne la liste des instances joignables avec leur sigle et port.
    Combine les ports standards avec ceux déclarés dans le scénario.
    """
    import httpx as _httpx
    # Ports à tester : standards v2311 + ceux du scénario courant
    ports_to_test = {8660, 8661, 8662, 8663, 8664, 8665, 8666}
    scenario = _inj.get("scenario") or {}
    for a in scenario.get("acteurs", []):
        p = a.get("port")
        if p:
            try: ports_to_test.add(int(p))
            except Exception: pass

    instances = []
    async with _httpx.AsyncClient(timeout=2.0) as client:
        for port in sorted(ports_to_test):
            up = False
            sigle = None
            try:
                # /api/v1/status/public est sans auth + léger
                r = await client.get(f"http://localhost:{port}/api/v1/status/public")
                if r.status_code == 200:
                    up = True
                    d = r.json()
                    # Essayer d'extraire le sigle depuis l'établissement
                    etab = d.get("etablissement") or {}
                    sigle = etab.get("sigle") or etab.get("code") or None
            except Exception:
                pass
            # Fallback sigle : via le scénario actif (mapping port → sigle)
            if not sigle:
                for a in scenario.get("acteurs", []):
                    if int(a.get("port") or 0) == port:
                        sigle = a.get("sigle")
                        break
            instances.append({
                "port": port,
                "sigle": (sigle or "").upper() or None,
                "up": up,
            })
    return {"instances": instances}


@app.get("/api/exercice/status")
async def get_exercice_status(auth=Depends(require_auth)):
    elapsed = _elapsed_s()
    scenario = _inj.get("scenario")
    ratio = (scenario or {}).get("meta",{}).get("ratio_compression",4.0) if scenario else 4.0
    stimuli_status = []
    if scenario:
        for s in scenario.get("stimuli",[]):
            done = s["id"] in _inj["done"]
            err  = s["id"] in _inj["errors"]
            target_s = int(s["t_min"]*60/ratio)
            stimuli_status.append({
                "id":s["id"],"t_min":s["t_min"],"type":s["type"],
                "cible":s["cible"],"titre":s.get("titre",""),
                "description_animateur":s.get("description_animateur",""),
                "action_attendue":s.get("action_attendue",""),
                "done":done,"error":err,
                "t_restant_s":max(0,target_s-elapsed) if not done else 0,
            })
    m,sec = divmod(elapsed,60)
    return {
        "running":_inj["running"],"paused":_inj["paused"],
        "session_id":_inj["session_id"],
        "scenario_titre":(scenario or {}).get("meta",{}).get("titre","") if scenario else "",
        "t_elapsed_s":elapsed,"t_elapsed_display":f"{m:02d}:{sec:02d}",
        "ratio":ratio,"stimuli":stimuli_status,
        "sites_actifs":list(_token_instances.keys()),
    }

# ── Santé ───────────────────────────────────────────────────────────────────────
@app.get("/api/exercice/template-xml")
async def get_template_xml(auth=Depends(require_auth)):
    template = ROOT_DIR / "scenarios" / "TEMPLATE_EXERCICE.xml"
    if template.exists():
        from fastapi.responses import FileResponse
        return FileResponse(str(template), media_type="application/xml", filename="TEMPLATE_EXERCICE.xml")
    raise HTTPException(404, "Template non trouvé")


# v2201 — Templates dans différents formats pour faciliter la rédaction
# manuelle de scénarios (sans passer par l'IA).
def _template_json_skeleton() -> dict:
    """Squelette JSON utilisable comme point de départ pour un scénario."""
    return {
        "meta": {
            "titre": "Mon scénario d'exercice",
            "description": "Description courte du scénario",
            "type_crise": "CYBER",
            "complexite": "MOYEN",
            "duree_min": 60,
            "sites": ["CHAG"],
            "nb_joueurs": 5,
            "auteur": "À compléter",
            "date_creation": "2026-04-21"
        },
        "joueurs": [
            {"nom": "Martin DUPONT", "role": "Directeur général", "site": "CHAG"},
            {"nom": "Claire MOREAU", "role": "RSSI", "site": "CHAG"}
        ],
        "stimuli": [
            {
                "id": "S01",
                "type": "incident",
                "t_min": 0,
                "cible": "CHAG",
                "titre": "Premier incident",
                "description_animateur": "Ce que l'animateur doit savoir",
                "action_attendue": "Ce que les joueurs doivent faire",
                "payload": {
                    "fait": "Description publique de l'incident",
                    "urgency": 3,
                    "type_crise": "CYBER",
                    "declarant_nom": "Technicien informatique",
                    "unite_fonctionnelle": "DSI"
                }
            },
            {
                "id": "S02",
                "type": "message",
                "t_min": 10,
                "cible": "CHAG",
                "titre": "Message chat",
                "description_animateur": "Contexte du message",
                "action_attendue": "Réaction des joueurs",
                "payload": {
                    "contenu": "Texte du message qui arrivera dans le chat"
                }
            },
            {
                "id": "S03",
                "type": "transfert",
                "t_min": 25,
                "cible": "CHAG",
                "titre": "Transfert patient urgent",
                "description_animateur": "Contexte du transfert",
                "action_attendue": "Coordination du transfert",
                "payload": {
                    "unite_origine": "Urgences",
                    "etablissement_destination": "GHTLMB",
                    "site_destination": "Hôpitaux du Léman",
                    "unite_destination": "Réanimation",
                    "motif": "Saturation locale",
                    "mode_transport": "SMUR",
                    "urgence": "URGENT",
                    "eta_min": 30
                }
            }
        ]
    }


@app.get("/api/exercice/template-json")
async def get_template_json(auth=Depends(require_auth)):
    """v2201 — Template JSON téléchargeable, commenté avec des exemples."""
    from fastapi.responses import Response
    content = json.dumps(_template_json_skeleton(), indent=2, ensure_ascii=False)
    return Response(
        content=content,
        media_type="application/json",
        headers={"Content-Disposition": 'attachment; filename="TEMPLATE_SCENARIO.json"'}
    )


@app.get("/api/exercice/template-csv")
async def get_template_csv(auth=Depends(require_auth)):
    """v2201 — Template CSV simple (stimuli seulement, pas de métadonnées).
    Utile pour une saisie rapide en tableur. Les métadonnées du scénario
    devront être saisies ensuite dans l'UI ou via un fichier JSON/XML."""
    from fastapi.responses import Response
    rows = [
        # En-tête explicite des colonnes attendues
        "id,type,t_min,cible,titre,description_animateur,action_attendue,contenu_ou_fait,urgency,declarant_nom,unite_fonctionnelle,etablissement_destination,unite_destination,motif,mode_transport,urgence",
        # Exemple incident
        'S01,incident,0,CHAG,"Panne SI","Le SI tombe","Activer PCA","DPI inaccessible depuis 5 min",3,"Technicien DSI","DSI","","","","",""',
        # Exemple message
        'S02,message,10,CHAG,"Message ARS","L\'ARS demande info","Répondre sous 30 min","Demande urgente d\'information ARS",2,"","","","","","",""',
        # Exemple transfert
        'S03,transfert,25,CHAG,"Transfert SMUR","Saturation","Coordonner avec GHTLMB","",3,"","",GHTLMB,"Réanimation","Saturation locale",SMUR,URGENT',
    ]
    content = "\n".join(rows) + "\n"
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="TEMPLATE_SCENARIO.csv"'}
    )


@app.get("/api/exercice/template-xlsx")
async def get_template_xlsx(auth=Depends(require_auth)):
    """v2201 — Template XLSX avec 2 feuilles : Métadonnées + Stimuli."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "openpyxl non installé — utiliser CSV ou JSON")

    wb = Workbook()

    # Feuille 1 : Métadonnées
    ws_meta = wb.active
    ws_meta.title = "Metadonnees"
    ws_meta.append(["Champ", "Valeur", "Exemples / Remarques"])
    meta_rows = [
        ("titre", "Mon scénario", "Titre affiché dans la liste"),
        ("description", "Description courte", "Quelques phrases"),
        ("type_crise", "CYBER", "CYBER / SANITAIRE / MIXTE / RH / TERTIAIRE"),
        ("complexite", "MOYEN", "FACILE / MOYEN / DIFFICILE / EXPERT"),
        ("duree_min", 60, "Durée en minutes"),
        ("sites", "CHAG,GHTLMB", "Sites séparés par virgule"),
        ("nb_joueurs", 5, ""),
        ("auteur", "Prénom NOM", ""),
    ]
    for row in meta_rows:
        ws_meta.append(row)

    # Feuille 2 : Stimuli
    ws_stim = wb.create_sheet("Stimuli")
    headers = ["id", "type", "t_min", "cible", "titre",
               "description_animateur", "action_attendue",
               "contenu_ou_fait", "urgency", "declarant_nom",
               "unite_fonctionnelle", "etablissement_destination",
               "unite_destination", "motif", "mode_transport", "urgence"]
    ws_stim.append(headers)
    ws_stim.append(["S01", "incident", 0, "CHAG", "Panne SI",
                    "Le SI tombe", "Activer PCA",
                    "DPI inaccessible depuis 5 min", 3,
                    "Technicien DSI", "DSI", "", "", "", "", ""])
    ws_stim.append(["S02", "message", 10, "CHAG", "Message ARS",
                    "L'ARS demande info", "Répondre sous 30 min",
                    "Demande urgente d'information ARS", 2,
                    "", "", "", "", "", "", ""])
    ws_stim.append(["S03", "transfert", 25, "CHAG", "Transfert SMUR",
                    "Saturation", "Coordonner avec GHTLMB",
                    "", 3, "", "", "GHTLMB", "Réanimation",
                    "Saturation locale", "SMUR", "URGENT"])

    # Feuille 3 : Joueurs
    ws_joueurs = wb.create_sheet("Joueurs")
    ws_joueurs.append(["nom", "role", "site"])
    ws_joueurs.append(["Martin DUPONT", "Directeur général", "CHAG"])
    ws_joueurs.append(["Claire MOREAU", "RSSI", "CHAG"])

    # Écrire en mémoire
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="TEMPLATE_SCENARIO.xlsx"'}
    )


# ── v2185 — Bilan post-exercice : agrégation + rapport HTML ─────────────────

async def _collect_instance_data(sigle: str, port: int, timeout: float = 5.0) -> dict:
    """Récupère les données d'une instance via ses endpoints publics.
    Utilise le token exo_{sigle}_2026 (fixe, connu du collecteur).
    Retourne un dict avec incidents, décisions, tâches, transferts."""
    base_url = f"http://localhost:{port}"
    # Login animateur côté instance
    data = {"sigle": sigle, "port": port, "incidents": [], "decisions": [],
            "tasks": [], "transferts": [], "presences": [], "erreurs": []}
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            # Login côté instance pour obtenir un JWT (même mécanique que /autologin)
            r_login = await client.post(f"{base_url}/api/v1/auth/login",
                json={"username": "animateur", "password": "Animateur2026!"})
            if r_login.status_code >= 300:
                # Fallback : dircrise
                r_login = await client.post(f"{base_url}/api/v1/auth/login",
                    json={"username": "dircrise", "password": "Exercice2026!"})
            if r_login.status_code >= 300:
                data["erreurs"].append(f"login failed HTTP {r_login.status_code}")
                return data
            jwt = r_login.json().get("token", "")
            if not jwt:
                data["erreurs"].append("JWT absent")
                return data
            h = {"Authorization": f"Bearer {jwt}"}
            # Charger en parallèle
            async def _fetch(path):
                try:
                    rr = await client.get(f"{base_url}{path}", headers=h)
                    return rr.json() if rr.status_code < 300 else None
                except Exception as e:
                    return None
            res = await asyncio.gather(
                _fetch("/api/v1/sitrep/history"),
                _fetch("/api/v1/cellule/decisions"),
                _fetch("/api/v1/tasks/"),
                _fetch("/api/v1/transferts"),
                _fetch("/api/v1/cellule/presences"),
            )
            data["incidents"]  = res[0] or []
            data["decisions"]  = res[1] or []
            data["tasks"]      = res[2] or []
            data["transferts"] = res[3] or []
            data["presences"]  = res[4] or []
    except Exception as e:
        data["erreurs"].append(str(e))
    return data


def _compute_competences(site_data: dict, stimuli_status: list) -> dict:
    """v2185 — Calcule les scores de compétences (0-100) par heuristiques.
    6 axes : Réactivité, Communication, Décision, Coordination inter-GHT,
    Méthode, Documentation.
    """
    incs = site_data.get("incidents", [])
    decs = site_data.get("decisions", [])
    tasks = site_data.get("tasks", [])
    transferts = site_data.get("transferts", [])
    presences = site_data.get("presences", [])

    # Réactivité : ratio incidents traités (pas en SIGNALÉ) + vitesse de traitement
    #   On comptabilise : % incidents non-SIGNALÉ sur total
    total_inc = len(incs)
    if total_inc:
        traites = sum(1 for i in incs if (i.get("status") or "SIGNALÉ") != "SIGNALÉ")
        reactivite = min(100, int((traites / total_inc) * 100))
    else:
        reactivite = 0

    # Communication : nombre de décisions saisies + nombre de tâches créées
    #   (trace d'une activité collaborative)
    n_communication = len(decs) + len(presences)
    communication = min(100, n_communication * 10)

    # Décision : nombre de décisions formelles prises par la cellule
    decision = min(100, len(decs) * 15)

    # Coordination inter-GHT : nombre de transferts initiés
    coord = min(100, len(transferts) * 25)

    # Méthode : % de jalons cochés sur l'ensemble des incidents
    tot_jalons = 0
    done_jalons = 0
    for i in incs:
        jl = i.get("jalons")
        if jl:
            try:
                js = json.loads(jl) if isinstance(jl, str) else jl
                tot_jalons += len(js)
                done_jalons += sum(1 for j in js if j.get("done"))
            except Exception:
                pass
    methode = int(100 * done_jalons / tot_jalons) if tot_jalons else 0

    # Documentation : % d'incidents avec analyse remplie + % tâches avec description
    docd = 0
    if total_inc:
        n_doc = sum(1 for i in incs if (i.get("analyse") or "").strip())
        docd = int(100 * n_doc / total_inc)

    return {
        "Réactivité":            reactivite,
        "Communication":         communication,
        "Décision":              decision,
        "Coordination inter-GHT": coord,
        "Méthode":               methode,
        "Documentation":         docd,
    }


@app.get("/api/exercice/bilan-data")
async def get_bilan_data(auth=Depends(require_auth)):
    """Agrège les données de toutes les instances pour le rapport bilan.
    Retourne : scénario, timeline stimuli, par site (incidents, décisions,
    tâches, transferts, compétences calculées)."""
    scenario = _inj.get("scenario") or {}
    stimuli = scenario.get("stimuli", [])
    # État injecteur
    status = {
        "running":       _inj.get("running", False),
        "t_elapsed_s":   _elapsed_s(),
        "done":          list(_inj.get("done", [])),
        "errors":        list(_inj.get("errors", [])),
        "session_id":    _inj.get("session_id"),
    }
    # Timeline stimuli avec done/error enrichi
    timeline_stimuli = []
    for s in stimuli:
        timeline_stimuli.append({
            "id":       s.get("id"),
            "t_min":    s.get("t_min"),
            "type":     s.get("type"),
            "cible":    s.get("cible"),
            "titre":    s.get("titre") or _stimulus_titre(s),
            "action_attendue": s.get("action_attendue", ""),
            "payload":  s.get("payload", {}),
            "done":     s.get("id") in _inj.get("done", []),
            "error":    s.get("id") in _inj.get("errors", []),
        })
    # Collecter en parallèle les données de chaque instance active
    actifs = list(_token_instances.keys()) or list(EXO_INSTANCES.keys())
    tasks_coll = [_collect_instance_data(sig, EXO_INSTANCES[sig]) for sig in actifs if sig in EXO_INSTANCES]
    sites_data = await asyncio.gather(*tasks_coll) if tasks_coll else []
    # Enrichir avec compétences
    for sd in sites_data:
        sd["competences"] = _compute_competences(sd, timeline_stimuli)
    return {
        "meta":            scenario.get("meta", {}),
        "acteurs":         scenario.get("acteurs", []),
        "solution_attendue": scenario.get("solution_attendue", scenario.get("meta", {}).get("solution_attendue", "")),
        "status":          status,
        "stimuli":         timeline_stimuli,
        "sites":           sites_data,
        "generated_at":    datetime.now(timezone.utc).isoformat(),
    }


def _stimulus_titre(s: dict) -> str:
    """Titre lisible pour un stimulus en l'absence de champ 'titre' explicite."""
    p = s.get("payload", {}) or {}
    t = p.get("fait") or p.get("contenu") or p.get("motif") or ""
    return (t[:80] + ("…" if len(t) > 80 else "")) if t else f"Stimulus {s.get('id','?')}"


@app.get("/api/exercice/rapport.html", response_class=HTMLResponse)
async def generate_rapport_html(auth=Depends(require_auth)):
    """Génère un rapport HTML autonome téléchargeable avec :
    - Métadonnées exercice
    - Timeline chronologique (stimuli + décisions + actions)
    - KPIs par site
    - Radar de compétences (Chart.js)
    - Scénario idéal"""
    data = await get_bilan_data(auth)  # réutilise l'agrégation
    meta = data.get("meta", {})
    stimuli = data.get("stimuli", [])
    sites = data.get("sites", [])
    solution = data.get("solution_attendue", "") or "_(aucun scénario idéal défini pour ce scénario)_"

    # Construire la timeline fusionnée triée chronologiquement
    events = []
    for s in stimuli:
        if s.get("done"):
            events.append({
                "t_min": s.get("t_min", 0),
                "kind":  "stimulus",
                "site":  s.get("cible", "?"),
                "type":  s.get("type", ""),
                "titre": s.get("titre", ""),
                "id":    s.get("id", ""),
            })
    # Incidents, décisions, tâches, transferts de chaque site
    now_ts = datetime.now(timezone.utc).timestamp()
    for sd in sites:
        sig = sd.get("sigle", "?")
        for i in sd.get("incidents", []):
            ts = i.get("timestamp")
            if ts:
                try:
                    dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
                    # On affiche juste une date formatée, pas un t_min relatif
                    events.append({
                        "t_min":  None,
                        "horodatage": dt.strftime("%d/%m %H:%M"),
                        "kind":   "incident",
                        "site":   sig,
                        "type":   i.get("type_crise", ""),
                        "titre":  f"#{i.get('id','?')} {(i.get('fait') or '')[:80]}",
                        "urgency": i.get("urgency", 1),
                    })
                except Exception:
                    pass
        for d in sd.get("decisions", []):
            ts = d.get("timestamp")
            if ts:
                try:
                    dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
                    events.append({
                        "t_min":  None,
                        "horodatage": dt.strftime("%d/%m %H:%M"),
                        "kind":   "decision",
                        "site":   sig,
                        "titre":  (d.get("contenu") or "")[:120],
                        "responsable": d.get("responsable", ""),
                    })
                except Exception:
                    pass
        for t in sd.get("transferts", []):
            events.append({
                "t_min":  None,
                "horodatage": "—",
                "kind":   "transfert",
                "site":   sig,
                "titre":  f"{t.get('unite_origine','?')} → {t.get('unite_destination','?')} ({t.get('statut','?')})",
            })

    # Construire les datasets du radar Chart.js
    radar_labels = ["Réactivité", "Communication", "Décision",
                    "Coordination inter-GHT", "Méthode", "Documentation"]
    colors = [
        "rgba(37,99,235,0.2)",   "rgba(239,68,68,0.2)",
        "rgba(34,197,94,0.2)",   "rgba(249,115,22,0.2)",
        "rgba(139,92,246,0.2)",  "rgba(14,165,233,0.2)",
        "rgba(236,72,153,0.2)",
    ]
    border_colors = [
        "rgba(37,99,235,1)",   "rgba(239,68,68,1)",
        "rgba(34,197,94,1)",   "rgba(249,115,22,1)",
        "rgba(139,92,246,1)",  "rgba(14,165,233,1)",
        "rgba(236,72,153,1)",
    ]
    radar_datasets = []
    for idx, sd in enumerate(sites):
        comp = sd.get("competences", {})
        radar_datasets.append({
            "label":           sd.get("sigle", "?"),
            "data":            [comp.get(l, 0) for l in radar_labels],
            "backgroundColor": colors[idx % len(colors)],
            "borderColor":     border_colors[idx % len(border_colors)],
            "borderWidth":     2,
            "pointRadius":     4,
        })

    # KPIs par site
    kpis_rows = []
    for sd in sites:
        sig = sd.get("sigle", "?")
        incs = sd.get("incidents", [])
        kpis_rows.append({
            "sigle":     sig,
            "nb_inc":    len(incs),
            "nb_res":    sum(1 for i in incs if i.get("status") == "RÉSOLU"),
            "nb_crit":   sum(1 for i in incs if (i.get("urgency") or 0) >= 3),
            "nb_dec":    len(sd.get("decisions", [])),
            "nb_tsk":    len(sd.get("tasks", [])),
            "nb_tra":    len(sd.get("transferts", [])),
            "nb_pres":   len(sd.get("presences", [])),
        })

    # Échappement HTML
    def _esc(s):
        if s is None: return ""
        return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

    # Lignes timeline
    events_sorted = sorted(
        events,
        key=lambda e: (e.get("horodatage") or "", e.get("t_min") or 0)
    )
    timeline_html = ""
    for e in events_sorted:
        col = {"stimulus":"#6366f1","incident":"#ef4444","decision":"#22c55e","transfert":"#f59e0b"}.get(e["kind"], "#64748b")
        when = e.get("horodatage") or (f"T+{e['t_min']}min" if e.get("t_min") is not None else "—")
        site = _esc(e.get("site", ""))
        titre = _esc(e.get("titre", ""))
        extra = ""
        if e["kind"] == "stimulus":
            extra = f' <span class="tl-tag" style="background:{col}22;color:{col}">{_esc(e.get("type",""))}</span>'
        elif e["kind"] == "incident":
            extra = f' <span class="tl-tag" style="background:#fee2e2;color:#b91c1c">urg.{e.get("urgency",0)}</span>'
        timeline_html += (
            f'<tr>'
            f'<td class="tl-when">{_esc(when)}</td>'
            f'<td class="tl-site">{site}</td>'
            f'<td class="tl-kind" style="color:{col}">● {e["kind"]}</td>'
            f'<td>{titre}{extra}</td>'
            f'</tr>'
        )
    if not timeline_html:
        timeline_html = '<tr><td colspan="4" style="text-align:center;padding:20px;color:#888">Aucun événement enregistré</td></tr>'

    # Tableau KPIs
    kpi_html = ""
    for r in kpis_rows:
        kpi_html += (
            f'<tr>'
            f'<td><b>{_esc(r["sigle"])}</b></td>'
            f'<td>{r["nb_inc"]}</td>'
            f'<td>{r["nb_res"]}</td>'
            f'<td style="color:#dc2626">{r["nb_crit"]}</td>'
            f'<td>{r["nb_dec"]}</td>'
            f'<td>{r["nb_tsk"]}</td>'
            f'<td>{r["nb_tra"]}</td>'
            f'<td>{r["nb_pres"]}</td>'
            f'</tr>'
        )
    if not kpi_html:
        kpi_html = '<tr><td colspan="8" style="text-align:center;padding:20px;color:#888">Aucun site actif</td></tr>'

    # Acteurs
    acteurs_html = ""
    for a in data.get("acteurs", []):
        jrs = a.get("joueurs", [])
        joueurs_txt = ", ".join([(j.get("display_name") or j.get("username") or "?") for j in jrs]) or "—"
        acteurs_html += (
            f'<tr><td><b>{_esc(a.get("sigle",""))}</b></td>'
            f'<td>{_esc(a.get("nom_etablissement","—"))}</td>'
            f'<td>{_esc(a.get("role","—"))}</td>'
            f'<td style="font-size:11px;color:#555">{_esc(joueurs_txt)}</td></tr>'
        )
    if not acteurs_html:
        acteurs_html = '<tr><td colspan="4" style="text-align:center;padding:20px;color:#888">—</td></tr>'

    # Génération
    html_out = _render_rapport_html(
        titre=_esc(meta.get("titre", "Exercice SCRIBE")),
        description=_esc(meta.get("description", "")),
        complexite=_esc(meta.get("complexite", "—")),
        type_crise=_esc(meta.get("type_crise", "—")),
        duree_min=meta.get("duree_min", "?"),
        ratio=meta.get("ratio_compression", "?"),
        nb_stimuli=len(stimuli),
        generated_at=data.get("generated_at", ""),
        session_id=data.get("status", {}).get("session_id", "—"),
        radar_labels_json=json.dumps(radar_labels, ensure_ascii=False),
        radar_datasets_json=json.dumps(radar_datasets, ensure_ascii=False),
        timeline_html=timeline_html,
        kpi_html=kpi_html,
        acteurs_html=acteurs_html,
        solution=_esc(solution),
    )
    return HTMLResponse(content=html_out, headers={
        "Content-Disposition": f'attachment; filename="rapport_scribe_{data.get("status",{}).get("session_id") or "exercice"}.html"'
    })


def _render_rapport_html(**kw) -> str:
    """Rapport HTML autonome avec Chart.js embarqué depuis CDN."""
    return """<!DOCTYPE html>
<html lang="fr"><head>
<meta charset="UTF-8">
<title>Rapport SCRIBE — """ + kw["titre"] + """</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: system-ui, -apple-system, sans-serif; background: #f8fafc;
  color: #0f172a; line-height: 1.5; padding: 0; }
.container { max-width: 1100px; margin: 0 auto; padding: 30px 24px 60px; }
header { background: linear-gradient(135deg, #003189, #0050c7); color: #fff;
  padding: 28px 30px; border-radius: 10px; margin-bottom: 24px;
  box-shadow: 0 4px 20px rgba(0,50,137,.25); }
header h1 { font-size: 22px; margin-bottom: 6px; }
header .meta { font-family: monospace; font-size: 11px; opacity: .85; letter-spacing: .5px; }
header .description { margin-top: 12px; font-size: 13px; opacity: .95; line-height: 1.6;}
h2 { font-size: 16px; color: #003189; margin: 30px 0 12px;
  padding-bottom: 6px; border-bottom: 2px solid #003189; }
h3 { font-size: 13px; color: #334155; margin: 16px 0 8px; }
section { background: #fff; border: 1px solid #e2e8f0; border-radius: 8px;
  padding: 20px 24px; margin-bottom: 20px; }
table { width: 100%; border-collapse: collapse; font-size: 12px; }
table th { text-align: left; padding: 8px 10px; background: #f1f5f9;
  border-bottom: 2px solid #cbd5e1; font-weight: 600; color: #475569; }
table td { padding: 7px 10px; border-bottom: 1px solid #e2e8f0; }
table tr:hover td { background: #fafbfc; }
.meta-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(150px,1fr));
  gap: 10px; margin-bottom: 12px; }
.meta-pill { background: #f1f5f9; border: 1px solid #e2e8f0; border-radius: 6px;
  padding: 8px 12px; font-size: 12px; }
.meta-pill b { color: #003189; display: block; font-size: 14px; margin-top: 2px; }
.tl-when { font-family: monospace; font-size: 11px; color: #64748b; white-space: nowrap; }
.tl-site { font-family: monospace; font-size: 11px; font-weight: 700; color: #003189;
  white-space: nowrap; }
.tl-kind { font-family: monospace; font-size: 11px; font-weight: 700; white-space: nowrap; }
.tl-tag { font-family: monospace; font-size: 10px; padding: 1px 6px; border-radius: 3px;
  margin-left: 5px; font-weight: 700; }
.radar-wrap { max-width: 720px; margin: 0 auto; padding: 10px; }
.solution { background: #f0fdf4; border-left: 4px solid #16a34a; padding: 12px 16px;
  border-radius: 0 6px 6px 0; font-size: 12px; line-height: 1.6; white-space: pre-wrap; }
.solution.empty { background: #fefce8; border-left-color: #f59e0b; color: #78716c;
  font-style: italic; }
footer { text-align: center; padding: 24px; color: #94a3b8; font-size: 11px;
  font-family: monospace; }
@media print {
  body { background: #fff; }
  section { page-break-inside: avoid; border: 1px solid #ddd; box-shadow: none; }
  header { box-shadow: none; }
}
</style>
</head><body>
<div class="container">

<header>
  <h1>📊 Rapport d'exercice SCRIBE</h1>
  <div class="meta">""" + kw["titre"] + """ · Session """ + str(kw["session_id"]) + """ · Généré le """ + kw["generated_at"] + """</div>
  <div class="description">""" + kw["description"] + """</div>
</header>

<section>
  <h2>1. Caractéristiques de l'exercice</h2>
  <div class="meta-grid">
    <div class="meta-pill">Type de crise<b>""" + kw["type_crise"] + """</b></div>
    <div class="meta-pill">Complexité<b>""" + kw["complexite"] + """</b></div>
    <div class="meta-pill">Durée jeu<b>""" + str(kw["duree_min"]) + """ min</b></div>
    <div class="meta-pill">Compression<b>x""" + str(kw["ratio"]) + """</b></div>
    <div class="meta-pill">Stimuli<b>""" + str(kw["nb_stimuli"]) + """</b></div>
  </div>
</section>

<section>
  <h2>2. Acteurs &amp; joueurs</h2>
  <table>
    <thead><tr><th>Sigle</th><th>Établissement</th><th>Rôle</th><th>Joueurs</th></tr></thead>
    <tbody>""" + kw["acteurs_html"] + """</tbody>
  </table>
</section>

<section>
  <h2>3. KPIs par site</h2>
  <table>
    <thead><tr>
      <th>Site</th>
      <th>Incidents</th>
      <th>Résolus</th>
      <th>Critiques</th>
      <th>Décisions</th>
      <th>Tâches</th>
      <th>Transferts</th>
      <th>Présences</th>
    </tr></thead>
    <tbody>""" + kw["kpi_html"] + """</tbody>
  </table>
</section>

<section>
  <h2>4. Radar des compétences</h2>
  <p style="font-size:11px;color:#64748b;margin-bottom:10px">
    Scores calculés par heuristiques (0-100) sur 6 axes : Réactivité, Communication,
    Décision, Coordination inter-GHT, Méthode, Documentation.
  </p>
  <div class="radar-wrap">
    <canvas id="radarChart"></canvas>
  </div>
</section>

<section>
  <h2>5. Timeline chronologique</h2>
  <table>
    <thead><tr><th>Moment</th><th>Site</th><th>Type</th><th>Événement</th></tr></thead>
    <tbody>""" + kw["timeline_html"] + """</tbody>
  </table>
</section>

<section>
  <h2>6. Scénario idéal</h2>
  <div class="solution""" + (" empty" if not kw["solution"] or kw["solution"].startswith("_(") else "") + """">""" + kw["solution"] + """</div>
</section>

<footer>SCRIBE Crisis OS · Rapport généré automatiquement · Sans données nominatives patient</footer>
</div>

<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<script>
const radarData = {
  labels: """ + kw["radar_labels_json"] + """,
  datasets: """ + kw["radar_datasets_json"] + """
};
new Chart(document.getElementById('radarChart'), {
  type: 'radar', data: radarData,
  options: {
    responsive: true, maintainAspectRatio: true,
    scales: { r: { min: 0, max: 100, ticks: { stepSize: 20, color: '#64748b' },
      pointLabels: { color: '#334155', font: { size: 12, weight: 'bold' } },
      grid: { color: '#e2e8f0' }, angleLines: { color: '#cbd5e1' } } },
    plugins: { legend: { position: 'top', labels: { color: '#334155',
      font: { size: 11 }, padding: 12 } } }
  }
});
</script>
</body></html>"""


# ── Joueurs prêts ──────────────────────────────────────────────────────────────
_joueurs_prets: dict = {}  # sigle → {username, at}

class PretPayload(BaseModel):
    sigle: str
    username: str

@app.post("/api/exercice/joueur-pret")
async def declare_pret(body: PretPayload):
    _joueurs_prets[body.sigle] = {"username": body.username, "at": datetime.now(timezone.utc).isoformat()}
    logger.info(f"Joueur prêt: {body.username} @ {body.sigle}")
    return {"ok": True}

@app.get("/api/exercice/prets")
async def get_prets(auth=Depends(require_auth)):
    return _joueurs_prets

@app.delete("/api/exercice/prets")
async def reset_prets(auth=Depends(require_auth)):
    _joueurs_prets.clear()
    return {"ok": True}

@app.get("/api/exercice/autologin")
async def autologin(port: int, sigle: str, auth=Depends(require_auth)):
    """Login côté serveur sur l'instance exercice → retourne le JWT au collecteur."""
    try:
        async with httpx.AsyncClient(timeout=8.0) as client:
            r = await client.post(
                f"http://localhost:{port}/api/v1/auth/login",
                json={"username": "dircrise", "password": "Exercice2026!"},
                headers={"Content-Type": "application/json"}
            )
        if r.status_code == 200:
            data = r.json()
            return {"ok": True, "token": data.get("token", "")}
        return {"ok": False, "status": r.status_code}
    except Exception as e:
        logger.error(f"Autologin {sigle}:{port}: {e}")
        return {"ok": False, "error": str(e)}

# ═══════════════════════════════════════════════════════════════════════════════
# ROUTES PORTÉES DEPUIS collecteur/collecteur.py (v2178+)
# — nécessaires pour que les instances SCRIBE exercice communiquent entre elles
#   (chat inter-GHT, transferts, messages, demandes, présence).
# ═══════════════════════════════════════════════════════════════════════════════

# ── Helpers /api/coll/* (attendus par index.html / chat.html des instances) ───
@app.get("/api/coll/me")
async def coll_me(credentials=Depends(security)):
    if not credentials: raise HTTPException(401)
    tok = credentials.credentials
    if tok == ADMIN_TOKEN:
        return {"id": 0, "username": "supervision", "display_name": "Supervision", "role": "admin"}
    sess = ui_sessions.get(tok)
    if sess:
        return {"id": 0, "username": sess.get("login", "supervision"),
                "display_name": sess.get("login", "supervision").capitalize(),
                "role": sess.get("role", "admin")}
    # Token établissement : on retourne une identité minimale
    sigle = tokens.get(tok)
    if sigle:
        return {"id": 0, "username": sigle.lower(), "display_name": sigle, "role": "etablissement"}
    raise HTTPException(401)

@app.get("/api/coll/users")
async def coll_users(credentials=Depends(security)):
    if not _check_any_auth(credentials): raise HTTPException(401)
    # Liste statique des utilisateurs animateur (suffisant pour l'autocomplétion @mention)
    return [
        {"id": 1, "username": "animateur",   "display_name": "Animateur",   "active": True},
        {"id": 2, "username": "supervision", "display_name": "Supervision", "active": True},
    ]

@app.get("/api/coll/fed-status")
async def coll_fed_status(credentials=Depends(security)):
    """Endpoint minimal — empêche les erreurs côté instance quand elle interroge le collecteur."""
    if not _check_any_auth(credentials): raise HTTPException(401)
    return {"ready": False, "enabled": False, "etablissement": "SUPERVISION"}

# ── Messagerie inter-GHT ─────────────────────────────────────────────────────
class MessageInterGHTBody(BaseModel):
    destinataire: str
    sujet: str
    contenu: str
    expediteur_nom: Optional[str] = None

@app.post("/api/messages")
async def send_message_interght(body: MessageInterGHTBody, credentials=Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle: raise HTTPException(401, "Non autorisé")
    msg = {
        "id": len(messages_inter) + 1,
        "expediteur": sigle,
        "expediteur_nom": body.expediteur_nom or sigle,
        "destinataire": (body.destinataire or "").upper(),
        "sujet": body.sujet,
        "contenu": body.contenu,
        "lu_par": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    messages_inter.append(msg)
    save_messages_inter()

    # v2311 — Fan-out : propager aussi le message dans la messagerie interne
    # de chaque destinataire concerné. Avant v2311 les messages inter-GHT
    # n'apparaissaient que dans l'onglet INTER-GHT, désactivé par défaut
    # depuis v2307 → messages invisibles côté joueur. En pushant vers
    # /api/v1/messagerie/broadcast-externe, les messages atterrissent
    # dans l'inbox interne avec notification, canal unifié.
    #
    # Règle : on exclut l'expéditeur (il n'a pas besoin de se voir) et
    # on envoie à tous les autres sites participants à l'exercice.
    try:
        scenario = _inj.get("scenario") or {}
        # Déterminer la liste des destinataires
        dest_upper = (body.destinataire or "").upper()
        if dest_upper == "TOUS":
            cibles = [
                a for a in scenario.get("acteurs", [])
                if (a.get("sigle") or "").upper() != sigle.upper()
                and a.get("port")  # doit avoir une instance joignable
            ]
        else:
            cibles = [
                a for a in scenario.get("acteurs", [])
                if (a.get("sigle") or "").upper() == dest_upper
                and a.get("port")
            ]
        if cibles:
            import httpx as _httpx
            async with _httpx.AsyncClient(timeout=5.0) as client:
                for a in cibles:
                    cible_sigle = (a.get("sigle") or "").upper()
                    port = int(a.get("port", 8660))
                    base = f"http://localhost:{port}"
                    # Token instance (cache ou login)
                    tok = _token_instances.get(cible_sigle) or _token_instances.get(a.get("sigle"))
                    if not tok:
                        try:
                            r_login = await client.post(
                                f"{base}/api/v1/auth/login",
                                json={"username": "dircrise", "password": "Exercice2026!"},
                            )
                            if r_login.status_code == 200:
                                tok = r_login.json().get("token", "")
                                _token_instances[cible_sigle] = tok
                        except Exception:
                            continue
                    if not tok:
                        continue
                    # Push en tant que message externe dans la messagerie interne
                    try:
                        await client.post(
                            f"{base}/api/v1/messagerie/broadcast-externe",
                            headers={"Authorization": f"Bearer {tok}",
                                     "Content-Type": "application/json"},
                            json={
                                "expediteur_nom": body.expediteur_nom or sigle,
                                "sujet":          body.sujet[:200],
                                "contenu":        body.contenu,
                            },
                        )
                    except Exception as e:
                        logger.warning(f"Fan-out message inter-GHT vers {cible_sigle} échoué: {e}")
        logger.info(f"Message inter-GHT {sigle} → {dest_upper} propagé sur {len(cibles)} instance(s)")
    except Exception as e:
        logger.warning(f"Fan-out inter-GHT global échoué: {e}")

    return {"ok": True, "id": msg["id"], "propage": True}

@app.get("/api/messages")
async def get_messages_interght(credentials=Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle: raise HTTPException(401, "Non autorisé")
    received = [m for m in messages_inter
                if m["destinataire"] == sigle or m["destinataire"] == "TOUS"]
    sent = [m for m in messages_inter if m["expediteur"] == sigle]
    return {"received": received, "sent": sent}

@app.put("/api/messages/{msg_id}/lire")
async def marquer_lu_interght(msg_id: int, credentials=Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle: raise HTTPException(401, "Non autorisé")
    for m in messages_inter:
        if m["id"] == msg_id:
            if sigle not in m["lu_par"]:
                m["lu_par"].append(sigle)
            save_messages_inter()
            return {"ok": True}
    raise HTTPException(404, "Message non trouvé")

@app.get("/api/messages/non-lus")
async def non_lus_interght(credentials=Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle: raise HTTPException(401, "Non autorisé")
    count = sum(1 for m in messages_inter
                if (m["destinataire"] == sigle or m["destinataire"] == "TOUS")
                and sigle not in m.get("lu_par", []))
    return {"count": count}

# ── Demandes inter-GHT ───────────────────────────────────────────────────────
@app.get("/api/declarations")
async def get_declarations_interght(credentials=Depends(security)):
    result = []
    for sigle, payload in etablissements.items():
        for d in payload.get("declarations", []):
            result.append({**d, "ght_emetteur": sigle})
    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return result

@app.get("/api/demandes")
async def get_demandes_interght(credentials=Depends(security)):
    result = []
    for sigle, payload in etablissements.items():
        for d in payload.get("demandes", []):
            result.append({**d, "ght_emetteur": d.get("ght_emetteur") or sigle})
    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return result

@app.patch("/api/demandes/{dem_id}")
async def update_demande_response(dem_id, request: Request, credentials=Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle: raise HTTPException(401, "Non autorisé")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "JSON invalide")
    reponse     = body.get("reponse", "")
    statut      = body.get("statut", "traite")
    repondu_par = body.get("repondu_par", sigle)
    for etab_sigle, payload in etablissements.items():
        for d in payload.get("demandes", []):
            if str(d.get("id")) == str(dem_id):
                d["reponse"]     = reponse
                d["statut"]      = statut
                d["repondu_par"] = repondu_par
                save_data()
                return {"ok": True}
    raise HTTPException(404, "Demande non trouvée")

# ── Transferts inter-établissements ──────────────────────────────────────────
@app.post("/api/push-transfert")
async def push_transfert_interght(request: Request, credentials=Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle: raise HTTPException(401, "Non autorisé")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "JSON invalide")
    transfert = {
        "id_local": body.get("id_local"),
        "ght_emetteur": sigle,
        "ght_emetteur_nom": body.get("ght_emetteur_nom", sigle),
        "ght_destinataire": body.get("ght_destinataire", ""),
        "unite_origine": body.get("unite_origine", ""),
        "etablissement_origine": body.get("etablissement_origine", ""),
        "unite_destination": body.get("unite_destination", ""),
        "etablissement_destination": body.get("etablissement_destination", ""),
        "site_destination": body.get("site_destination", ""),
        "statut": body.get("statut", "EN_PREPARATION"),
        "eta": body.get("eta"),
        "horodatage_depart": body.get("horodatage_depart"),
        "commentaire": body.get("commentaire"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    existing = next((t for t in transferts_inter
                     if t.get("id_local") == transfert["id_local"]
                     and t.get("ght_emetteur") == sigle), None)
    if transfert["statut"] in ("ARRIVE", "ANNULE"):
        transferts_inter[:] = [t for t in transferts_inter
                               if not (t.get("id_local") == transfert["id_local"]
                                       and t.get("ght_emetteur") == sigle)]
    elif existing:
        existing.update(transfert)
    else:
        transferts_inter.append(transfert)
    save_transferts_inter()
    return {"ok": True}

@app.get("/api/transferts-en-cours")
async def get_transferts_en_cours(credentials=Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle: raise HTTPException(401, "Non autorisé")
    from datetime import timedelta
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat()
    result = [t for t in transferts_inter
              if t.get("ght_destinataire", "").upper() == sigle.upper()
              and (t.get("statut") in ("EN_PREPARATION", "EN_COURS")
                   or (t.get("statut") == "ARRIVE" and t.get("created_at", "") >= cutoff))]
    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return result

@app.get("/api/transfert-statut/{id_local}")
async def get_transfert_statut(id_local: str, credentials=Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle: raise HTTPException(401, "Non autorisé")
    t = next((t for t in transferts_inter
              if str(t.get("id_local")) == str(id_local)
              and t.get("ght_emetteur", "").upper() == sigle.upper()), None)
    if t:
        return {"statut": t.get("statut"), "found": True}
    return {"statut": "ARRIVE", "found": False}

# ── Chat inter-GHT : messages ────────────────────────────────────────────────
@app.post("/api/chat/messages")
async def chat_push_message(request: Request, credentials=Depends(security)):
    """Une instance pousse un message dans un salon territorial."""
    sigle = get_etab_from_token(credentials)
    if not sigle: raise HTTPException(401)
    global _chat_msg_counter, _chat_pj_counter
    try:
        body = await request.json()
    except Exception as e:
        raise HTTPException(400, f"Corps JSON invalide: {e}")
    salon_nom = body.get("salon_nom", "général")
    if salon_nom not in chat_messages:
        chat_messages[salon_nom] = []
    # Anti-doublon : même sigle + même contenu dans les 5 dernières secondes
    contenu = body.get("contenu", "")
    recent = [m for m in chat_messages[salon_nom][-10:]
              if m.get("auteur_sigle") == sigle and m.get("contenu") == contenu]
    if recent:
        try:
            last_ts = datetime.fromisoformat(recent[-1]["horodatage"].replace("Z", ""))
            delta = (datetime.now(timezone.utc) - last_ts.replace(tzinfo=timezone.utc)).total_seconds()
            if delta < 5:
                return {"ok": True, "detail": "doublon ignoré"}
        except Exception:
            pass
    _chat_msg_counter += 1
    pjs = []
    # Nouveau format : URL directe vers l'instance source
    for pm in body.get("pj_meta", []):
        _chat_pj_counter += 1
        pj_key = f"ght-{_chat_pj_counter}"
        if pm.get("remote_url"):
            chat_pj_store[pj_key] = {
                "nom":        pm.get("nom", "fichier"),
                "remote_url": pm["remote_url"],
                "taille":     pm.get("taille", 0),
            }
        elif pm.get("dataUrl"):
            chat_pj_store[pj_key] = {
                "nom":     pm.get("nom", "fichier"),
                "dataUrl": pm["dataUrl"],
                "taille":  pm.get("taille", 0),
            }
        else:
            continue
        pjs.append({"id": pj_key, "nom": pm.get("nom", ""), "taille": pm.get("taille", 0)})
    # Legacy : base64 direct
    for pj_data in body.get("pj_data", []) + body.get("pj_inline", []):
        dataUrl = pj_data.get("dataUrl", "")
        if dataUrl:
            _chat_pj_counter += 1
            pj_key = f"ght-{_chat_pj_counter}"
            chat_pj_store[pj_key] = {
                "nom":     pj_data.get("nom", "fichier"),
                "dataUrl": dataUrl,
                "taille":  pj_data.get("taille", 0),
            }
            pjs.append({"id": pj_key, "nom": pj_data.get("nom", ""), "taille": pj_data.get("taille", 0)})
    msg = {
        "id":           _chat_msg_counter,
        "salon_nom":    salon_nom,
        "auteur_nom":   body.get("auteur_nom", sigle),
        "auteur_sigle": sigle,
        "contenu":      contenu,
        "mentions":     body.get("mentions", []),
        "reply_to_id":  body.get("reply_to_id"),
        "horodatage":   datetime.now(timezone.utc).isoformat(),
        "origine":      "ght",
        "pj":           pjs,
    }
    chat_messages[salon_nom].append(msg)
    if len(chat_messages[salon_nom]) > 500:
        chat_messages[salon_nom] = chat_messages[salon_nom][-500:]
    return {"ok": True, "id": _chat_msg_counter}

@app.get("/api/chat/messages")
async def chat_get_messages(salon_nom: str = "général", since_id: int = 0,
                            credentials=Depends(security)):
    if not _check_any_auth(credentials): raise HTTPException(401)
    msgs = chat_messages.get(salon_nom, [])
    filtered = [m for m in msgs if m.get("id", 0) > since_id]
    return filtered[-100:]

# ── Chat inter-GHT : salons ──────────────────────────────────────────────────
@app.post("/api/chat/salons")
async def chat_create_salon(request: Request, credentials=Depends(security)):
    if not _check_any_auth(credentials): raise HTTPException(401)
    body = await request.json()
    nom = (body.get("nom", "") or "").strip().lower()
    if nom and nom not in chat_messages:
        chat_messages[nom] = []
    salon_list = list(chat_messages.keys())
    idx = salon_list.index(nom) if nom in salon_list else 0
    return {"id": idx + 1, "nom": nom, "type": "territorial",
            "icone": body.get("icone", "💬"),
            "couleur": body.get("couleur", "#7c3aed"),
            "description": body.get("description", f"Salon #{nom}")}

@app.get("/api/chat/salons")
async def chat_list_salons(credentials=Depends(security)):
    if not _check_any_auth(credentials): raise HTTPException(401)
    SALON_ICONS = {
        "général": "💬", "coordination": "🎯", "transferts": "🚑",
        "logistique": "📦", "direction": "🏛️",
    }
    result = []
    for i, nom in enumerate(chat_messages.keys()):
        result.append({
            "id":          i + 1,
            "nom":         nom,
            "description": f"Salon territorial #{nom}",
            "couleur":     "#7c3aed",
            "icone":       SALON_ICONS.get(nom, "💬"),
            "type":        "territorial",
            "ordre":       i,
        })
    return result

@app.get("/api/chat/salons/{salon_id}/messages")
async def get_salon_messages_by_id(salon_id: int, limit: int = 50,
                                   credentials=Depends(security)):
    if not _check_any_auth(credentials): raise HTTPException(401)
    salon_list = list(chat_messages.keys())
    idx = salon_id - 1
    if idx < 0 or idx >= len(salon_list):
        raise HTTPException(404, f"Salon {salon_id} non trouvé")
    salon_nom = salon_list[idx]
    return chat_messages.get(salon_nom, [])[-limit:]

@app.post("/api/chat/salons/{salon_id}/messages")
async def post_salon_message_by_id(salon_id: int, request: Request,
                                   credentials=Depends(security)):
    global _chat_msg_counter
    if not _check_any_auth(credentials): raise HTTPException(401)
    salon_list = list(chat_messages.keys())
    idx = salon_id - 1
    if idx < 0 or idx >= len(salon_list):
        raise HTTPException(404, f"Salon {salon_id} non trouvé")
    salon_nom = salon_list[idx]
    body = await request.json()
    tok = credentials.credentials if credentials else ""
    if tok == ADMIN_TOKEN:
        auteur_nom = "Animateur"
    else:
        sess = ui_sessions.get(tok, {})
        auteur_nom = sess.get("login", "Animateur").capitalize()
    _chat_msg_counter += 1
    msg = {
        "id":           _chat_msg_counter,
        "salon_nom":    salon_nom,
        "auteur_nom":   auteur_nom,
        "auteur_sigle": "SUPERVISION",
        "contenu":      body.get("contenu", ""),
        "mentions":     body.get("mentions", []),
        "reply_to_id":  body.get("reply_to_id"),
        "reply_to":     None,
        "horodatage":   datetime.now(timezone.utc).isoformat(),
        "origine":      "ght",
        "pj":           [],
        "supprime":     False,
    }
    if salon_nom not in chat_messages:
        chat_messages[salon_nom] = []
    chat_messages[salon_nom].append(msg)
    return msg

# ── Chat inter-GHT : présence ────────────────────────────────────────────────
@app.post("/api/chat/presence")
async def chat_push_presence(request: Request, credentials=Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle: raise HTTPException(401)
    body = await request.json()
    users = body.get("users", [])
    now = datetime.now(timezone.utc).isoformat()
    chat_presence[sigle] = [
        {"user_id": u.get("user_id"), "display_name": u.get("display_name"), "last_seen": now}
        for u in users
    ]
    return {"ok": True}

@app.get("/api/chat/presence")
async def chat_get_presence(credentials=Depends(security)):
    if not _check_any_auth(credentials): raise HTTPException(401)
    from datetime import timedelta
    cutoff = datetime.now(timezone.utc) - timedelta(minutes=2)
    result = {}
    for etab_sigle, users in chat_presence.items():
        active = []
        for u in users:
            try:
                ts = datetime.fromisoformat(u["last_seen"].replace("Z", "")).replace(tzinfo=timezone.utc)
                if ts >= cutoff:
                    active.append(u)
            except Exception:
                pass
        if active:
            result[etab_sigle] = active
    return result

@app.post("/api/chat/presence/ping")
async def chat_presence_ping(request: Request, credentials=Depends(security)):
    """Ping de présence — accepte tokens établissements, admin, session UI ou JWT SCRIBE."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    tok = credentials.credentials if credentials else ""
    sigle = tokens.get(tok)
    if sigle:
        display = body.get("display_name", sigle)
        uid = body.get("user_id", 0)
    else:
        if tok == ADMIN_TOKEN:
            sigle, display, uid = "SUPERVISION", "Animateur", 0
        elif tok in ui_sessions:
            sess = ui_sessions[tok]
            sigle = "SUPERVISION"
            display = sess.get("login", "supervision").capitalize()
            uid = abs(hash(tok)) % 100000
        else:
            sigle  = body.get("sigle", "INCONNU")
            display = body.get("display_name", sigle)
            uid = body.get("user_id", 0)
    if not sigle:
        return {"ok": True}
    now = datetime.now(timezone.utc).isoformat()
    existing = chat_presence.get(sigle, [])
    existing = [u for u in existing if u.get("user_id") != uid]
    existing.append({"user_id": uid, "display_name": display, "last_seen": now})
    chat_presence[sigle] = existing
    return {"ok": True}

# ── Chat inter-GHT : pièces jointes ──────────────────────────────────────────
@app.get("/api/chat/pj/{pj_id}")
async def serve_chat_pj(pj_id: str, token: str = "", credentials=Depends(security)):
    """Sert une PJ stockée dans le collecteur (redirect si remote_url, sinon base64)."""
    if not _check_any_auth(credentials) and not (token and (token == ADMIN_TOKEN or token in ui_sessions or token in tokens)):
        raise HTTPException(401)
    pj = chat_pj_store.get(str(pj_id))
    if not pj:
        raise HTTPException(404, "PJ non trouvée")
    if pj.get("remote_url"):
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=pj["remote_url"], status_code=302)
    data_url = pj.get("dataUrl", "")
    if not data_url:
        raise HTTPException(404, "Données PJ manquantes")
    if "," in data_url:
        header, b64 = data_url.split(",", 1)
        mime = header.split(":")[1].split(";")[0] if ":" in header else "application/octet-stream"
    else:
        b64 = data_url
        mime = "application/octet-stream"
    import base64
    from fastapi.responses import Response as FResponse
    content = base64.b64decode(b64)
    return FResponse(content=content, media_type=mime,
                     headers={"Content-Disposition": f"inline; filename={pj['nom']}"})

# ── Annuaire inter-GHT (lookup minimal pour compat instances) ────────────────
@app.get("/api/annuaire")
async def get_annuaire_interght():
    """Retourne les infos annuaire publiées par chaque établissement via push."""
    result = []
    for sigle, data in etablissements.items():
        etab_info = data.get("etablissement", {}) or {}
        result.append({
            "sigle":    sigle,
            "nom":      etab_info.get("nom", sigle),
            "contacts": etab_info.get("contacts", []),
        })
    return result

# ═══════════════════════════════════════════════════════════════════════════════
# FIN DES ROUTES PORTÉES
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/health")
def health():
    return {"status":"ok","port":8565,"exercice":True,
            "etabs":len(etablissements),"running":_inj["running"]}

@app.post("/api/admin/pending/accept-all")
async def accept_all():
    load_tokens()
    return {"accepted":list(set(tokens.values()))}

# ── Interface HTML ──────────────────────────────────────────────────────────────
@app.get("/static/logo-scribe.png")
async def serve_logo():
    for p in [BASE_DIR.parent/"app"/"static"/"logo-scribe.png",
              BASE_DIR/"logo-scribe.png"]:
        if p.exists():
            from fastapi.responses import FileResponse
            return FileResponse(str(p))
    raise HTTPException(404)

@app.get("/static/favicon.svg")
async def serve_favicon():
    """Sert le favicon SVG (même visuel que les instances SCRIBE)."""
    for p in [BASE_DIR.parent/"app"/"static"/"favicon.svg",
              BASE_DIR/"favicon.svg"]:
        if p.exists():
            from fastapi.responses import FileResponse
            return FileResponse(str(p), media_type="image/svg+xml")
    raise HTTPException(404)

@app.get("/favicon.ico")
async def serve_favicon_ico():
    """Redirige /favicon.ico vers le SVG pour compat navigateurs."""
    for p in [BASE_DIR.parent/"app"/"static"/"favicon.svg",
              BASE_DIR/"favicon.svg"]:
        if p.exists():
            from fastapi.responses import FileResponse
            return FileResponse(str(p), media_type="image/svg+xml")
    raise HTTPException(404)

@app.get("/", response_class=HTMLResponse)
async def dashboard():
    html_path = BASE_DIR / "collecteur_exercice.html"
    if html_path.exists():
        return HTMLResponse(html_path.read_text(encoding="utf-8"))
    return HTMLResponse(_HTML)

# ── HTML interface animateur ───────────────────────────────────────────────────

# ── Démarrage ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Les load_*() sont désormais au niveau module (voir plus haut)
    port = int(os.environ.get("COLLECTEUR_PORT", os.environ.get("SCRIBE_EXO_PORT",8565)))
    print(f"\n  SCRIBE Collecteur Exercice — port {port}")
    print(f"  Admin: animateur / Animateur2026!")
    print(f"  Token: {ADMIN_TOKEN[:20]}...\n")
    uvicorn.run("collecteur_exercice:app",host="0.0.0.0",port=port,reload=False)
