"""
collecteur.py — Collecteur de supervision territoriale SCRIBE

Reçoit les pushs JSON des instances SCRIBE des établissements.
Interface web de supervision (lecture seule).

Usage :
  python collecteur.py

Accès : http://localhost:9000
Admin : http://localhost:9000/admin  (pour gérer les tokens établissements)
"""

import hashlib
import json
import logging
import os
import secrets
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, List

import uvicorn

# ── h153 — Transferts supervision (SQLAlchemy + SQLite dédié) ────────────────
try:
    from sqlalchemy import create_engine, Column, Integer, String, DateTime
    from sqlalchemy.orm import declarative_base, sessionmaker
    _TR_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "collecteur_transferts.db")
    _tr_engine  = create_engine(f"sqlite:///{_TR_DB_PATH}", connect_args={"check_same_thread": False})
    _TrBase = declarative_base()

    class TransfertColl(_TrBase):
        __tablename__ = "transferts_supervision"
        id                        = Column(Integer, primary_key=True, autoincrement=True)
        etablissement_origine     = Column(String, nullable=False, default="")
        unite_origine             = Column(String, nullable=True, default="")
        etablissement_destination = Column(String, nullable=False, default="")
        unite_destination         = Column(String, nullable=True, default="")
        statut                    = Column(String, default="EN_PREPARATION")
        age_estime                = Column(String, nullable=True)
        sexe                      = Column(String, nullable=True)
        motif                     = Column(String, nullable=True)
        pathologie                = Column(String, nullable=True)
        eta                       = Column(String, nullable=True)
        commentaire               = Column(String, nullable=True)
        # Données patient (optionnelles)
        nom                       = Column(String, nullable=True)
        prenom                    = Column(String, nullable=True)
        nom_jeune_fille           = Column(String, nullable=True)
        ipp                       = Column(String, nullable=True)
        date_naissance            = Column(String, nullable=True)
        nom                       = Column(String, nullable=True)
        prenom                    = Column(String, nullable=True)
        nom_jeune_fille           = Column(String, nullable=True)
        ipp                       = Column(String, nullable=True)
        date_naissance            = Column(String, nullable=True)
        redacteur                 = Column(String, nullable=False, default="Supervision")
        horodatage_creation       = Column(DateTime, default=lambda: datetime.now(timezone.utc))
        horodatage_depart         = Column(DateTime, nullable=True)
        horodatage_arrivee        = Column(DateTime, nullable=True)

    _TrBase.metadata.create_all(bind=_tr_engine)
    _TrSession = sessionmaker(bind=_tr_engine, autocommit=False, autoflush=False)
    _TR_ENABLED = True
except Exception as _e:
    _TR_ENABLED = False
    print(f"[collecteur] Transferts supervision désactivés : {_e}")
from starlette.middleware.base import BaseHTTPMiddleware
from fastapi import FastAPI, HTTPException, Request, Depends, Form, File, UploadFile
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("collecteur")

# ── Fichiers de persistance ───────────────────────────────────────────────
DATA_FILE   = os.environ.get("COLLECTEUR_DATA",   "collecteur_data.json")
TOKENS_FILE = os.environ.get("COLLECTEUR_TOKENS", "collecteur_tokens.json")
ADMIN_FILE  = os.environ.get("COLLECTEUR_ADMIN",  "collecteur_admin.json")

# Structure en mémoire
etablissements: dict = {}   # sigle → dernière remontée
tokens: dict = {}           # token → sigle établissement

# ── Chat inter-GHT ──────────────────────────────────────────────────────────
# Structure : {salon_nom: [{id, auteur_nom, auteur_sigle, contenu, mentions, horodatage}]}
CHAT_SALONS_DEFAULT = ["général","coordination","transferts","logistique","direction"]
chat_messages: dict = {s: [] for s in CHAT_SALONS_DEFAULT}
# Sessions UI actives : {token: {"login": ..., "role": ...}}
ui_sessions: dict = {}
# Présence : {sigle: [{user_id, display_name, last_seen}]}
chat_presence: dict = {}
_chat_msg_counter: int = 0
PENDING_FILE  = os.environ.get("COLLECTEUR_PENDING",  "collecteur_pending.json")
RELAY_FILE    = os.environ.get("COLLECTEUR_RELAY",    "collecteur_relay.json")

# ── Relais hiérarchique ─────────────────────────────────────────────────────
# Liste de collecteurs upstream vers lesquels re-pousser toutes les données
# Format: [{url, token, nom, actif}]
relay_targets: list = []

def load_relay():
    global relay_targets
    if Path(RELAY_FILE).exists():
        try:
            relay_targets = json.loads(Path(RELAY_FILE).read_text())
        except Exception:
            relay_targets = []

def save_relay():
    Path(RELAY_FILE).write_text(json.dumps(relay_targets, ensure_ascii=False, indent=2))

async def relay_push(payload: dict, sigle: str):
    """Re-pousse le payload vers tous les collecteurs upstream actifs."""
    if not relay_targets:
        return
    try:
        import httpx as _httpx
        # Enrichir le payload avec le sigle source pour traçabilité
        relay_payload = dict(payload)
        relay_payload["_relayed_from"] = ADMIN_TOKEN[:8] + "..."
        relay_payload["_relayed_at"]   = datetime.now(timezone.utc).isoformat()
        async with _httpx.AsyncClient(timeout=8.0) as client:
            for target in relay_targets:
                if not target.get("actif", True):
                    continue
                try:
                    resp = await client.post(
                        target["url"],
                        json=relay_payload,
                        headers={"Authorization": f"Bearer {target['token']}",
                                 "Content-Type": "application/json"}
                    )
                    if resp.status_code == 200:
                        logger.info(f"Relay OK → {target.get('nom','?')} ({target['url']}) pour {sigle}")
                    else:
                        logger.warning(f"Relay HTTP {resp.status_code} → {target.get('nom','?')} pour {sigle}")
                except Exception as e:
                    logger.warning(f"Relay ERREUR → {target.get('nom','?')} : {e}")
    except Exception as e:
        logger.warning(f"Relay général ERREUR : {e}")
MESSAGES_FILE = os.environ.get("COLLECTEUR_MESSAGES", "collecteur_messages.json")
pending: dict = {}          # token → info établissement en attente (pas encore accepté)
messages_inter: list = []   # messagerie inter-GHT
transferts_inter: list = [] # transferts inter-établissements
TRANSFERTS_INTER_FILE = os.environ.get("COLLECTEUR_TRANSFERTS", "collecteur_transferts.json")

# ── Fiches réflexe : fichiers déposés en supervision et diffusés aux instances ─
# Stockés sur disque ; servis à une URL « …/api/v1/fichiers/… » pour que les
# instances les détectent comme partages reçus (« Partagé avec moi »).
FICHES_DIR   = os.environ.get("COLLECTEUR_FICHES_DIR", "collecteur_fiches")
FICHES_INDEX = os.path.join(FICHES_DIR, "_index.json")
try:
    os.makedirs(FICHES_DIR, exist_ok=True)
except Exception:
    pass

def _fiches_load() -> dict:
    try:
        return json.loads(Path(FICHES_INDEX).read_text())
    except Exception:
        return {}

def _fiches_save(idx: dict):
    try:
        Path(FICHES_INDEX).write_text(json.dumps(idx, ensure_ascii=False, indent=2))
    except Exception:
        pass

def load_transferts_inter():
    global transferts_inter
    if Path(TRANSFERTS_INTER_FILE).exists():
        try:
            transferts_inter = json.loads(Path(TRANSFERTS_INTER_FILE).read_text())
        except Exception:
            transferts_inter = []

def save_transferts_inter():
    Path(TRANSFERTS_INTER_FILE).write_text(json.dumps(transferts_inter, ensure_ascii=False, indent=2))

load_transferts_inter()


def _load_or_create_admin_token() -> str:
    """Charge le token admin depuis le fichier, ou le crée une seule fois."""
    # Priorité 1 : variable d'environnement
    if os.environ.get("ADMIN_TOKEN"):
        return os.environ["ADMIN_TOKEN"]
    # Priorité 2 : fichier persistant
    p = Path(ADMIN_FILE)
    if p.exists():
        try:
            return json.loads(p.read_text()).get("admin_token", "")
        except Exception:
            pass
    # Première fois : générer et sauvegarder
    token = secrets.token_hex(32)
    p.write_text(json.dumps({"admin_token": token}, indent=2))
    return token


ADMIN_TOKEN = _load_or_create_admin_token()

# h153 - Label superviseur ---------------------------------------------------
def _get_supervisor_label():
    try:
        pf = Path(ADMIN_FILE)
        if pf.exists(): return json.loads(pf.read_text()).get("supervisor_label", "Supervision")
    except Exception: pass
    return "Supervision"

def _set_supervisor_label(label):
    try:
        pf = Path(ADMIN_FILE)
        data = json.loads(pf.read_text()) if pf.exists() else {}
        data["supervisor_label"] = (label or "").strip()
        pf.write_text(json.dumps(data, indent=2))
    except Exception as e:
        print(f"[collecteur] supervisor_label: {e}")

SUPERVISOR_LABEL = _get_supervisor_label()


def load_tokens():
    global tokens
    tokens = dict(scribe_TOKENS)
    if Path(TOKENS_FILE).exists():
        try:
            saved = json.loads(Path(TOKENS_FILE).read_text())
            tokens.update(saved)
        except Exception:
            pass
    # v3.4 (h38) — Strip défensif des sigles au chargement.
    # Évite les doublons "ch2" / "ch2 " sur la carte de supervision et
    # les "Illegal header value" du côté push qui empêchent l'instance
    # d'apparaître. Auto-correction silencieuse à chaque démarrage.
    cleaned = {}
    for tok, sigle in tokens.items():
        if isinstance(sigle, str):
            cleaned[tok] = sigle.strip()
        else:
            cleaned[tok] = sigle
    if cleaned != tokens:
        tokens = cleaned
        print(f"[h38] Tokens normalisés (strip)")
    save_tokens()

def save_tokens():
    Path(TOKENS_FILE).write_text(json.dumps(tokens, ensure_ascii=False, indent=2))

def load_data():
    global etablissements
    if Path(DATA_FILE).exists():
        try:
            etablissements = json.loads(Path(DATA_FILE).read_text())
        except Exception:
            etablissements = {}
            return
    # v3.4 (h38) — Strip + dédup des clés au chargement.
    # Si on a "ch2 " et "CH2" en doublon, on garde l'entrée la plus
    # récente (selon _received_at) et on dégage l'autre.
    if etablissements:
        normalized = {}
        for sigle, data in etablissements.items():
            sigle_clean = (sigle or "").strip()
            if not sigle_clean:
                continue
            if sigle_clean in normalized:
                # Doublon : garder le plus récent
                existing = normalized[sigle_clean]
                new_ts = data.get("_received_at", "")
                old_ts = existing.get("_received_at", "")
                if new_ts > old_ts:
                    normalized[sigle_clean] = data
                    print(f"[h38] Doublon résolu : {sigle!r} (vs {sigle_clean!r}) → conservé le plus récent")
            else:
                normalized[sigle_clean] = data
        if normalized != etablissements:
            etablissements = normalized
            print(f"[h38] {len(etablissements)} entrée(s) après normalisation")
            save_data()

def save_data():
    Path(DATA_FILE).write_text(json.dumps(etablissements, ensure_ascii=False, indent=2))

# ── App FastAPI ───────────────────────────────────────────────────────────

# h152 — Middleware sécurité : headers HTTP manquants (audit sécurité)
class _SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
        response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
        response.headers.setdefault("Permissions-Policy", "geolocation=(), camera=(), microphone=()")
        response.headers.setdefault("X-XSS-Protection", "0")
        # Masquer le header Server (fingerprinting)
        response.headers["Server"] = "SCRIBE"
        return response

# h152 — Désactiver la doc API publique (OpenAPI/ReDoc) en production.
# Les endpoints /openapi.json, /docs et /redoc sont désactivés
# pour limiter la surface d'exposition (cf. audit sécurité).
app = FastAPI(
    title="SCRIBE Collecteur territorial",
    version="1.2.1",
    openapi_url=None,   # désactive /openapi.json
    docs_url=None,      # désactive /docs
    redoc_url=None,     # désactive /redoc
)
app.add_middleware(_SecurityHeadersMiddleware)


# h153 — Archivage automatique des transferts ARRIVE depuis plus d'1h
async def _auto_archive_transferts():
    """Tâche de fond : archive les TransfertColl ARRIVE depuis > 1h."""
    import asyncio
    while True:
        await asyncio.sleep(300)  # vérifier toutes les 5 minutes
        if not _TR_ENABLED:
            continue
        try:
            db = _TrSession()
            try:
                cutoff = datetime.now(timezone.utc) - timedelta(hours=1)
                archives = db.query(TransfertColl).filter(
                    TransfertColl.statut == "ARRIVE",
                    TransfertColl.horodatage_arrivee != None,
                    TransfertColl.horodatage_arrivee <= cutoff
                ).all()
                for t in archives:
                    t.statut = "ARCHIVE"
                    mention = "[Archivé automatiquement après 1h]"
                    t.commentaire = ((t.commentaire or "") + " " + mention).strip()
                    # Retirer de transferts_inter
                    id_local = f"SUP-{t.id}"
                    transferts_inter[:] = [x for x in transferts_inter
                                           if x.get("id_local") != id_local]
                if archives:
                    db.commit()
                    save_transferts_inter()
                    print(f"[auto-archive] {len(archives)} transfert(s) archivé(s)")
            finally:
                db.close()
        except Exception as e:
            print(f"[auto-archive] erreur: {e}")


@app.on_event("startup")
async def _on_startup():
    """Démarre les tâches de fond au lancement du collecteur."""
    import asyncio
    asyncio.create_task(_auto_archive_transferts())
    print("[collecteur] Tâche auto-archive démarrée")





# ── v3000h48 — Messagerie ISO : embarque le VRAI plugin messagerie ───────────
# La supervision utilise exactement le même module que les instances.
try:
    import os as _os_mm, sys as _sys_mm
    _sys_mm.path.insert(0, _os_mm.path.dirname(_os_mm.path.abspath(__file__)))
    import messagerie_mount as _msg_mount
    # IDOR fix : la messagerie embarquée exige le token du collecteur (ADMIN_TOKEN
    # ou session UI admin), en header Bearer ou ?token= (liens PJ). Sans token → 401.
    def _msg_token_ok(tok):
        if not tok:
            return False
        if tok == ADMIN_TOKEN:
            return True
        return tok in ui_sessions
    _msg_mount.mount(app, token_validator=_msg_token_ok)
except Exception as _e_mm:
    import logging as _l_mm
    _l_mm.getLogger("scribe.collecteur").error("[messagerie] montage KO : %s", _e_mm)


# ── CORS : middleware http natif Starlette ────────────────────────────────
@app.middleware("http")
async def cors_middleware(request: Request, call_next):
    origin = request.headers.get("origin") or "*"
    cors_headers = {
        "Access-Control-Allow-Origin": origin,
        "Access-Control-Allow-Credentials": "true",
        "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS, PATCH",
        "Access-Control-Allow-Headers": "Authorization, Content-Type, Accept",
        "Access-Control-Max-Age": "3600",
    }
    if request.method == "OPTIONS":
        from fastapi.responses import Response as _Resp
        return _Resp(status_code=200, headers=cors_headers)
    try:
        response = await call_next(request)
    except Exception:
        from fastapi.responses import JSONResponse
        response = JSONResponse({"detail": "Internal Server Error"}, status_code=500)
    for k, v in cors_headers.items():
        response.headers[k] = v
    return response

from fastapi.responses import JSONResponse as _JSONResponse
from starlette.middleware.base import BaseHTTPMiddleware
from fastapi import Request as _Req2
from starlette.exceptions import HTTPException as StarletteHTTPException
from fastapi.exceptions import RequestValidationError

@app.exception_handler(StarletteHTTPException)
async def http_exc_handler(request: _Req2, exc: StarletteHTTPException):
    origin = request.headers.get("origin") or "*"
    return _JSONResponse({"detail": exc.detail}, status_code=exc.status_code, headers={
        "Access-Control-Allow-Origin": origin,
        "Access-Control-Allow-Credentials": "true",
    })

@app.exception_handler(Exception)
async def global_exc_handler(request: _Req2, exc: Exception):
    origin = request.headers.get("origin") or "*"
    return _JSONResponse({"detail": str(exc)}, status_code=500, headers={
        "Access-Control-Allow-Origin": origin,
        "Access-Control-Allow-Credentials": "true",
    })


# ─── INTÉGRATION MASTER (pilotage d'instances) ───────────────────────────────
# v2323 : permet de lancer/arrêter/configurer les instances SCRIBE depuis l'UI
# Le module master/ est optionnel — si absent, le collecteur tourne normalement.
try:
    import sys as _sys, os as _os
    _master_path = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
    if _master_path not in _sys.path:
        _sys.path.insert(0, _master_path)
    from master.master_routes import router as _master_router, lifecycle_register as _master_lifecycle
    app.include_router(_master_router)
    _master_lifecycle(app)
    print("[master] Module pilotage instances activé")
except Exception as _e:
    print(f"[master] Module non chargé (OK si non utilisé) : {_e}")


security = HTTPBearer(auto_error=False)


def get_etab_from_token(credentials: Optional[HTTPAuthorizationCredentials]) -> Optional[str]:
    if not credentials:
        return None
    if credentials.credentials == ADMIN_TOKEN:
        return "SUPERVISEUR"
    return tokens.get(credentials.credentials)


def require_admin(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    if not credentials or credentials.credentials != ADMIN_TOKEN:
        raise HTTPException(status_code=403, detail="Token admin invalide")
    return True

def require_node_or_admin(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> str:
    """Accepte le token admin collecteur OU un token de fédération d'instance.
    Renvoie le sigle de l'appelant (SUPERVISEUR pour l'admin). Sert au relais
    messagerie inter-instances : une instance peut écrire à un agent d'une autre."""
    etab = get_etab_from_token(credentials)
    if not etab:
        raise HTTPException(status_code=403, detail="Token invalide")
    return etab

def require_ui_admin(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> bool:
    """Accepte le token admin collecteur OU un token de session UI admin."""
    if not credentials:
        return False
    tok = credentials.credentials
    if tok == ADMIN_TOKEN:
        return True
    session = ui_sessions.get(tok, {})
    return session.get("role") == "admin"


# ── Endpoint de réception (appelé par les SCRIBE) ─────────────────────────

# ═══════════════════════════════════════════════════════════════
# NOUVELLES FONCTIONNALITÉS v1.4.0
# ═══════════════════════════════════════════════════════════════

# ── File d'attente : établissements en attente d'acceptation ──

def load_pending():
    global pending
    if Path(PENDING_FILE).exists():
        try:
            pending = json.loads(Path(PENDING_FILE).read_text())
        except Exception:
            pending = {}

def save_pending():
    Path(PENDING_FILE).write_text(json.dumps(pending, ensure_ascii=False, indent=2))

def load_messages_inter():
    global messages_inter
    if Path(MESSAGES_FILE).exists():
        try:
            messages_inter = json.loads(Path(MESSAGES_FILE).read_text())
        except Exception:
            messages_inter = []

def save_messages_inter():
    Path(MESSAGES_FILE).write_text(json.dumps(messages_inter, ensure_ascii=False, indent=2))


@app.post("/api/push")
async def receive_push(
    request: Request,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """Reçoit un push JSON d'un SCRIBE établissement."""
    cred_token = credentials.credentials if credentials else None
    sigle = get_etab_from_token(credentials)

    if not sigle:
        # Ajouter à la file d'attente pending (si pas déjà connu)
        if cred_token and cred_token not in pending:
            try:
                payload = await request.json()
                nom_propose = (payload.get("etablissement", {}).get("nom", "Inconnu") or "").strip()
                sigle_propose = (payload.get("etablissement", {}).get("sigle", "?") or "").strip()
            except Exception:
                nom_propose = "Inconnu"
                sigle_propose = "?"
            pending[cred_token] = {
                "sigle_propose": sigle_propose,
                "nom_propose": nom_propose,
                "ip": request.client.host,
                "first_seen": datetime.now(timezone.utc).isoformat(),
                "token_preview": cred_token[:12] + "..."
            }
            save_pending()
            logger.warning(f"⏳ Token inconnu → EN ATTENTE : {sigle_propose} ({nom_propose}) depuis {request.client.host} | token_preview: {cred_token[:12] if cred_token else '?'}...")
        raise HTTPException(status_code=401, detail="Token en attente d'acceptation par l'administrateur du collecteur")

    # v3.4 (h38) — Strip défensif sur le sigle (peut contenir un espace
    # traînant venant d'une instance non corrigée). On normalise ici pour
    # éviter qu'une même instance crée deux entrées (avec et sans espace)
    # sur la carte de supervision.
    sigle = sigle.strip()

    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON invalide")

    payload["_received_at"] = datetime.now(timezone.utc).isoformat()
    payload["_source_ip"]   = request.client.host

    # v3.4 (h38) — Strip aussi le sigle dans le payload pour cohérence
    if "etablissement" in payload and isinstance(payload["etablissement"], dict):
        if "sigle" in payload["etablissement"]:
            payload["etablissement"]["sigle"] = (payload["etablissement"]["sigle"] or "").strip()
        if "nom" in payload["etablissement"]:
            payload["etablissement"]["nom"] = (payload["etablissement"]["nom"] or "").strip()

    etablissements[sigle] = payload
    save_data()

    logger.info(f"✓ Push ACCEPTÉ — {sigle} | niveau: {payload.get('niveau_global','?')} | "
                f"{payload.get('kpis',{}).get('incidents_ouverts',0)} incidents | "
                f"total stockés: {len(etablissements)}")

    # Relais hiérarchique
    if relay_targets:
        import asyncio
        asyncio.create_task(relay_push(payload, sigle))

    return {"ok": True, "sigle": sigle, "received_at": payload["_received_at"]}


# ── Gestion de la file d'attente (admin) ──────────────────────

@app.get("/api/admin/pending", dependencies=[Depends(require_admin)])
async def list_pending():
    """Liste les établissements en attente d'acceptation."""
    return [
        {"token": k, **v}
        for k, v in pending.items()
    ]

@app.post("/api/admin/pending/{token_prefix}/accept", dependencies=[Depends(require_admin)])
async def accept_pending(token_prefix: str, body: dict = {}):
    """Accepte un établissement en attente — l'enrôle comme token valide."""
    matches = [k for k in pending if k.startswith(token_prefix) or k == token_prefix]
    if not matches:
        # Chercher par token complet
        if token_prefix in pending:
            matches = [token_prefix]
    if not matches:
        raise HTTPException(status_code=404, detail="Établissement en attente non trouvé")
    token = matches[0]
    info = pending.pop(token)
    sigle = body.get("sigle", info["sigle_propose"]).strip().upper() or info["sigle_propose"]
    tokens[token] = sigle
    save_tokens()
    save_pending()
    logger.info(f"Établissement accepté : {sigle} (token: {token[:12]}...)")
    return {"ok": True, "sigle": sigle, "message": f"{sigle} enrôlé avec succès"}

@app.post("/api/admin/tokens/Territoire", dependencies=[Depends(require_admin)])
async def register_scribe_tokens():
    """Enregistre les 4 tokens Territoire démo — utile si l'auto-register a échoué."""
    added = []
    for tok, sigle in scribe_TOKENS.items():
        if tok not in tokens:
            tokens[tok] = sigle
            added.append(sigle)
    if added:
        save_tokens()
        logger.info(f"Tokens Territoire enregistrés manuellement : {added}")
    return {"ok": True, "added": added, "total": len(tokens),
            "message": f"{len(added)} token(s) ajouté(s), {len(tokens)} token(s) total"}


@app.post("/api/admin/pending/accept-all", dependencies=[Depends(require_admin)])
async def accept_all_pending():
    """Accepte tous les établissements en attente d'un coup."""
    accepted = []
    for tok, info in list(pending.items()):
        sigle = info.get("sigle_propose", tok[:8]).upper()
        tokens[tok] = sigle
        accepted.append(sigle)
        logger.info(f"Enrôlement en masse : {sigle} (token: {tok[:12]}...)")
    pending.clear()
    save_tokens()
    save_pending()
    return {"ok": True, "accepted": accepted, "count": len(accepted)}


@app.post("/api/admin/pending/{token_prefix}/reject", dependencies=[Depends(require_admin)])
async def reject_pending(token_prefix: str):
    """Rejette et supprime un établissement en attente."""
    matches = [k for k in pending if k.startswith(token_prefix) or k == token_prefix]
    if not matches:
        raise HTTPException(status_code=404, detail="Non trouvé")
    info = pending.pop(matches[0])
    save_pending()
    return {"ok": True, "message": f"Rejeté : {info.get('sigle_propose','')}"}



# ── Messagerie inter-GHT ──────────────────────────────────────

class MessageInterGHTBody(BaseModel):
    destinataire: str   # sigle du GHT destinataire, ou "TOUS"
    sujet: str
    contenu: str
    expediteur_nom: Optional[str] = None

@app.post("/api/messages")
async def send_message_interght(
    body: MessageInterGHTBody,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """Envoie un message inter-GHT depuis un SCRIBE enrôlé."""
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(status_code=401, detail="Non autorisé")
    msg = {
        "id": len(messages_inter) + 1,
        "expediteur": sigle,
        "expediteur_nom": body.expediteur_nom or sigle,
        "destinataire": body.destinataire.upper(),
        "sujet": body.sujet,
        "contenu": body.contenu,
        "lu_par": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    messages_inter.append(msg)
    save_messages_inter()
    logger.info(f"Message inter-GHT : {sigle} → {body.destinataire} | {body.sujet[:40]}")
    return {"ok": True, "id": msg["id"]}

@app.get("/api/messages")
async def get_messages_interght(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """Récupère les messages destinés à ce GHT (ou 'TOUS')."""
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(status_code=401, detail="Non autorisé")
    # Messages reçus : destinataire = mon sigle OU "TOUS"
    received = [m for m in messages_inter
                if m["destinataire"] == sigle or m["destinataire"] == "TOUS"]
    # Messages envoyés
    sent = [m for m in messages_inter if m["expediteur"] == sigle]
    return {"received": received, "sent": sent}

@app.put("/api/messages/{msg_id}/lire")
async def marquer_lu_interght(
    msg_id: int,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """Marque un message comme lu par ce GHT."""
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(status_code=401, detail="Non autorisé")
    for m in messages_inter:
        if m["id"] == msg_id:
            if sigle not in m["lu_par"]:
                m["lu_par"].append(sigle)
            save_messages_inter()
            return {"ok": True}
    raise HTTPException(status_code=404, detail="Message non trouvé")

@app.get("/api/messages/non-lus")
async def non_lus_interght(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """Compte les messages non lus pour ce GHT."""
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(status_code=401, detail="Non autorisé")
    count = sum(1 for m in messages_inter
                if (m["destinataire"] == sigle or m["destinataire"] == "TOUS")
                and sigle not in m.get("lu_par", []))
    return {"count": count}


# ── SUPERVISION : messagerie avec les instances (v3000h44, additif) ──────────
# La supervision lit les messages qui lui sont adressés
# (destinataire == "SUPERVISION") et répond à une instance précise.
# Réutilise messages_inter — aucun nouveau stockage. Les instances utilisent
# leurs endpoints /api/messages existants (destinataire="SUPERVISION" à l'envoi,
# réception via destinataire == leur sigle). Auth admin (compte supervision).

class SupervisionMessageBody(BaseModel):
    destinataire: str   # sigle de l'instance cible
    sujet: str
    contenu: str
    expediteur_nom: Optional[str] = "Supervision"

@app.get("/api/supervision/messages", dependencies=[Depends(require_admin)])
async def supervision_messages():
    """Inbox supervision : messages reçus (destinataire SUPERVISION) + envoyés."""
    received = [m for m in messages_inter if m.get("destinataire") == "SUPERVISION"]
    sent     = [m for m in messages_inter if m.get("expediteur")  == "SUPERVISION"]
    return {"received": received, "sent": sent}

@app.post("/api/supervision/messages", dependencies=[Depends(require_admin)])
async def supervision_send(body: SupervisionMessageBody):
    """La supervision envoie un message à une instance (expediteur forcé SUPERVISION)."""
    msg = {
        "id": len(messages_inter) + 1,
        "expediteur": "SUPERVISION",
        "expediteur_nom": body.expediteur_nom or "Supervision",
        "destinataire": body.destinataire.upper(),
        "sujet": body.sujet,
        "contenu": body.contenu,
        "lu_par": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    messages_inter.append(msg)
    save_messages_inter()
    logger.info(f"Message SUPERVISION → {body.destinataire} | {body.sujet[:40]}")
    return {"ok": True, "id": msg["id"]}

@app.put("/api/supervision/messages/{msg_id}/lire", dependencies=[Depends(require_admin)])
async def supervision_marquer_lu(msg_id: int):
    for m in messages_inter:
        if m["id"] == msg_id:
            if "SUPERVISION" not in m.get("lu_par", []):
                m.setdefault("lu_par", []).append("SUPERVISION")
            save_messages_inter()
            return {"ok": True}
    raise HTTPException(status_code=404, detail="Message non trouvé")


# ── Token admin (création / liste) ───────────────────────────────────────

@app.get("/api/admin/tokens", dependencies=[Depends(require_admin)])
async def list_tokens():
    return [{"sigle": v, "token_preview": k[:8] + "..."} for k, v in tokens.items()]

@app.post("/api/admin/tokens", dependencies=[Depends(require_admin)])
async def create_token(body: dict):
    """Crée/importe un token établissement. body: {sigle, token (optionnel, sinon généré)}"""
    sigle = body.get("sigle", "").strip().upper()
    if not sigle:
        raise HTTPException(status_code=400, detail="sigle requis")
    token = body.get("token", "").strip() or secrets.token_hex(32)
    if len(token) < 16:
        raise HTTPException(status_code=400, detail="token trop court (min 16 caractères)")
    tokens[token] = sigle
    save_tokens()
    logger.info(f"Token enregistré pour {sigle}")
    return {"sigle": sigle, "token": token, "message": "Token enregistré pour " + sigle}

@app.delete("/api/admin/tokens/{sigle_or_prefix}", dependencies=[Depends(require_admin)])
async def disconnect_etablissement(sigle_or_prefix: str):
    """Déconnecte (révoque) un établissement enrôlé.
    
    v3.4 (h38) — Purge aussi les données associées dans `etablissements`
    et `pending` pour éviter les entrées fantômes sur la carte de
    supervision (cas : ancien sigle avec espace traînant qui restait
    visible même après correction de la source côté instance).
    """
    sigle_upper = sigle_or_prefix.strip().upper()
    # Chercher par sigle (avec normalisation : on accepte avec ou sans espace)
    to_delete = [
        k for k, v in tokens.items()
        if (v or "").strip().upper() == sigle_upper or k.startswith(sigle_or_prefix.strip())
    ]
    # Purger aussi etablissements (vue affichée sur la carte de supervision)
    etabs_to_purge = [
        s for s in list(etablissements.keys())
        if (s or "").strip().upper() == sigle_upper
    ]
    # Purger les pending
    pending_to_purge = [
        k for k, v in list(pending.items())
        if (v.get("sigle") or "").strip().upper() == sigle_upper
    ]

    if not to_delete and not etabs_to_purge and not pending_to_purge:
        return {"ok": True, "disconnected": sigle_upper, "note": "already_absent"}

    for k in to_delete:
        sigle = tokens.pop(k)
        logger.info(f"Token révoqué : {sigle}")
    for s in etabs_to_purge:
        etablissements.pop(s, None)
        logger.info(f"Données carte purgées : {s}")
    for k in pending_to_purge:
        pending.pop(k, None)
        logger.info(f"Pending purgé : {k[:8]}...")

    save_tokens()
    save_data()
    save_pending()
    return {
        "ok": True,
        "disconnected": sigle_upper,
        "tokens_removed": len(to_delete),
        "carte_purgees": len(etabs_to_purge),
        "pending_purges": len(pending_to_purge),
    }


# ── Relais hiérarchique (admin) ───────────────────────────────────────────

class RelayTarget(BaseModel):
    nom: str
    url: str
    token: str
    actif: bool = True

@app.get("/api/admin/relay", dependencies=[Depends(require_admin)])
async def list_relay():
    return relay_targets

@app.post("/api/admin/relay", dependencies=[Depends(require_admin)])
async def add_relay(body: RelayTarget):
    if len(body.token) < 16:
        raise HTTPException(status_code=400, detail="Token trop court (min 16 chars)")
    for t in relay_targets:
        if t["url"] == body.url:
            raise HTTPException(status_code=409, detail="Ce collecteur est déjà configuré")
    entry = {"nom": body.nom, "url": body.url, "token": body.token, "actif": body.actif,
             "added_at": datetime.now(timezone.utc).isoformat()}
    relay_targets.append(entry)
    save_relay()
    logger.info(f"Relay ajouté : {body.nom} → {body.url}")
    return {"ok": True, "message": f"Relais vers {body.nom} configuré", "entry": entry}

@app.delete("/api/admin/relay/{idx}", dependencies=[Depends(require_admin)])
async def delete_relay(idx: int):
    if idx < 0 or idx >= len(relay_targets):
        raise HTTPException(status_code=404, detail="Relais introuvable")
    removed = relay_targets.pop(idx)
    save_relay()
    return {"ok": True, "removed": removed["nom"]}

@app.put("/api/admin/relay/{idx}/toggle", dependencies=[Depends(require_admin)])
async def toggle_relay(idx: int):
    if idx < 0 or idx >= len(relay_targets):
        raise HTTPException(status_code=404, detail="Relais introuvable")
    relay_targets[idx]["actif"] = not relay_targets[idx].get("actif", True)
    save_relay()
    state = "activé" if relay_targets[idx]["actif"] else "désactivé"
    return {"ok": True, "nom": relay_targets[idx]["nom"], "actif": relay_targets[idx]["actif"],
            "message": f"Relais {state}"}

@app.post("/api/admin/relay/{idx}/test", dependencies=[Depends(require_admin)])
async def test_relay(idx: int):
    if idx < 0 or idx >= len(relay_targets):
        raise HTTPException(status_code=404, detail="Relais introuvable")
    target = relay_targets[idx]
    try:
        import httpx as _httpx
        test_payload = {
            "etablissement": {"nom": "TEST_RELAY", "sigle": "TEST"},
            "niveau_global": "NOMINAL", "kpis": {"incidents_ouverts": 0}, "_test": True
        }
        async with _httpx.AsyncClient(timeout=8.0) as client:
            resp = await client.post(
                target["url"], json=test_payload,
                headers={"Authorization": f"Bearer {target['token']}", "Content-Type": "application/json"}
            )
            return {"ok": resp.status_code in (200,401), "status": resp.status_code,
                    "nom": target["nom"], "note": "401 = collecteur joignable mais token à enregistrer"}
    except Exception as e:
        return {"ok": False, "error": str(e), "nom": target["nom"]}


# ── Push status et capacité ───────────────────────────────────────────────

@app.post("/api/push-status")
async def receive_status_push(
    request: Request,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(status_code=401, detail="Token inconnu")
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON invalide")
    payload["_received_at"] = datetime.now(timezone.utc).isoformat()
    if sigle in etablissements:
        etablissements[sigle]["_status_page"] = payload
        if payload.get("_statuts_sites"):
            etablissements[sigle]["_statuts_sites"] = payload["_statuts_sites"]
    else:
        etablissements[sigle] = {"_status_page": payload, "_received_at": payload["_received_at"]}
    save_data()
    return {"ok": True, "sigle": sigle}

@app.post("/api/push-capacite")
async def receive_push_capacite(
    request: Request,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(status_code=401, detail="Token inconnu")
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON invalide")
    payload["_received_at"] = datetime.now(timezone.utc).isoformat()
    if sigle not in etablissements:
        etablissements[sigle] = {}
    etablissements[sigle]["_capacite"] = payload
    save_data()
    logger.info(f"Push capacité — {sigle} | {payload.get('nb_services',0)} services")
    return {"ok": True, "sigle": sigle}

# ── Lecture (summary, detail) ─────────────────────────────────────────────


@app.get("/api/declarations")
async def get_declarations_interght(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    result = []
    for sigle, payload in etablissements.items():
        for d in payload.get("declarations", []):
            result.append({**d, "ght_emetteur": sigle})
    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return result

@app.get("/api/demandes")
async def get_demandes_interght(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    result = []
    for sigle, payload in etablissements.items():
        for d in payload.get("demandes", []):
            result.append({**d, "ght_emetteur": d.get("ght_emetteur") or sigle})
    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return result

@app.patch("/api/demandes/{dem_id}")
async def update_demande_response(dem_id, request: Request, credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    """Met à jour la réponse d'une demande inter-GHT (appelé par l'instance destinataire)."""
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(status_code=401, detail="Non autorisé")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON invalide")
    reponse     = body.get("reponse", "")
    statut      = body.get("statut", "traite")
    repondu_par = body.get("repondu_par", sigle)
    for etab_sigle, payload in etablissements.items():
        for d in payload.get("demandes", []):
            if str(d.get("id")) == str(dem_id):
                d["reponse"]     = reponse
                d["statut"]      = statut
                d["repondu_par"] = repondu_par
                save_data()
                return {"ok": True}
    raise HTTPException(status_code=404, detail="Demande non trouvée")


# ══════════════════════════════════════════════════════════════════════════════
#  TRANSFERTS SUPERVISION — h153
#  Gestion des transferts inter-établissements par la supervision (ARS / pilote)
#  Activé uniquement si _TR_ENABLED (sqlalchemy disponible)
# ══════════════════════════════════════════════════════════════════════════════


def _inject_tr_sup_to_inter(t) -> None:
    """h153 - Injecte un TransfertColl dans transferts_inter pour les instances."""
    if not _TR_ENABLED or t is None:
        return
    try:
        id_local = f"SUP-{t.id}"
        transfert = {
            "id_local": id_local,
            "ght_emetteur": "SUPERVISION",
            "ght_emetteur_nom": SUPERVISOR_LABEL,
            "ght_destinataire": (t.etablissement_destination or "").upper(),
            "unite_origine": t.unite_origine or "",
            "etablissement_origine": t.etablissement_origine or "",
            "unite_destination": t.unite_destination or "",
            "etablissement_destination": t.etablissement_destination or "",
            "site_destination": t.etablissement_destination or "",
            "statut": t.statut or "EN_PREPARATION",
            "eta": t.eta or None,
            "motif": t.motif or "",
            "commentaire": t.commentaire or "",
            "created_by_supervision": True,
            "supervisor_label": SUPERVISOR_LABEL,
            "horodatage_depart": t.horodatage_depart.isoformat() if t.horodatage_depart else None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "is_sortant_supervision": True,  # h153 : flag pour scribe.js
        }
        global transferts_inter
        if t.statut in ("ANNULE", "ARCHIVE"):
            # ANNULÉ ou ARCHIVÉ → retirer de transferts_inter (plus visible dans instances)
            transferts_inter[:] = [x for x in transferts_inter if x.get("id_local") != id_local]
        elif t.statut == "ARRIVE":
            # ARRIVE → garder dans transferts_inter avec le bon statut pendant 2h
            # pour que les instances puissent afficher la colonne ARRIVÉ
            existing = next((x for x in transferts_inter if x.get("id_local") == id_local), None)
            if existing:
                existing.update(transfert)
                existing["arrived_at"] = datetime.now(timezone.utc).isoformat()
            else:
                transfert["arrived_at"] = datetime.now(timezone.utc).isoformat()
                transferts_inter.append(transfert)
        else:
            existing = next((x for x in transferts_inter if x.get("id_local") == id_local), None)
            if existing:
                existing.update(transfert)
            else:
                transferts_inter.append(transfert)
        save_transferts_inter()
    except Exception as e:
        print(f"[collecteur] _inject_tr_sup_to_inter: {e}")


@app.get("/api/admin/supervisor-label")
async def get_supervisor_label_route(_ok: bool = Depends(require_admin)):
    return {"label": SUPERVISOR_LABEL}


@app.post("/api/admin/supervisor-label")
async def set_supervisor_label_route(body: dict, _ok: bool = Depends(require_admin)):
    global SUPERVISOR_LABEL
    label = (body.get("label") or "Supervision").strip()[:50]
    _set_supervisor_label(label)
    SUPERVISOR_LABEL = label
    return {"ok": True, "label": label}


def _fmt_tr_coll(t) -> dict:
    """Sérialise un TransfertColl en dict JSON."""
    return {
        "id":                        t.id,
        "etablissement_origine":     t.etablissement_origine or "",
        "unite_origine":             t.unite_origine or "",
        "etablissement_destination": t.etablissement_destination or "",
        "unite_destination":         t.unite_destination or "",
        "statut":                    t.statut or "EN_PREPARATION",
        "age_estime":                t.age_estime or "",
        "sexe":                      t.sexe or "",
        "motif":                     t.motif or "",
        "pathologie":                t.pathologie or "",
        "eta":                       t.eta or "",
        "commentaire":               t.commentaire or "",
        "redacteur":                 t.redacteur or "Supervision",
        "nom":                       t.nom or "",
        "prenom":                    t.prenom or "",
        "nom_jeune_fille":           t.nom_jeune_fille or "",
        "ipp":                       t.ipp or "",
        "date_naissance":            t.date_naissance or "",
        "nom":                       t.nom or "",
        "prenom":                    t.prenom or "",
        "nom_jeune_fille":           t.nom_jeune_fille or "",
        "ipp":                       t.ipp or "",
        "date_naissance":            t.date_naissance or "",
        "supervisor_label":          SUPERVISOR_LABEL,
        "horodatage_creation":       t.horodatage_creation.isoformat() if t.horodatage_creation else None,
        "horodatage_depart":         t.horodatage_depart.isoformat()   if t.horodatage_depart   else None,
        "horodatage_arrivee":        t.horodatage_arrivee.isoformat()  if t.horodatage_arrivee  else None,
    }


@app.get("/api/admin/transferts-supervision")
async def tr_sup_list(_ok: bool = Depends(require_admin)):
    """Liste tous les transferts supervision (admin uniquement)."""
    if not _TR_ENABLED:
        raise HTTPException(503, "Module transferts supervision non disponible (sqlalchemy requis)")
    cutoff = datetime.now(timezone.utc) - timedelta(hours=24)
    db = _TrSession()
    try:
        q = db.query(TransfertColl).filter(
            (TransfertColl.statut.in_(["EN_PREPARATION", "EN_COURS"])) |
            ((TransfertColl.statut == "ARRIVE")  & (TransfertColl.horodatage_creation >= cutoff)) |
            ((TransfertColl.statut == "ANNULE")  & (TransfertColl.horodatage_creation >= cutoff)) |
            ((TransfertColl.statut == "ARCHIVE") & (TransfertColl.horodatage_creation >= (datetime.now(timezone.utc) - timedelta(days=7))))
        ).order_by(TransfertColl.horodatage_creation.desc())
        return [_fmt_tr_coll(t) for t in q.all()]
    finally:
        db.close()


@app.post("/api/admin/transferts-supervision")
async def tr_sup_create(body: dict, _ok: bool = Depends(require_admin)):
    """Crée un transfert inter-établissement depuis la supervision."""
    if not _TR_ENABLED:
        raise HTTPException(503, "Module transferts supervision non disponible")
    orig = (body.get("etablissement_origine") or "").strip()
    dest = (body.get("etablissement_destination") or "").strip()
    if not orig or not dest:
        raise HTTPException(400, "etablissement_origine et etablissement_destination requis")
    db = _TrSession()
    try:
        t = TransfertColl(
            etablissement_origine     = orig,
            unite_origine             = (body.get("unite_origine") or "").strip(),
            etablissement_destination = dest,
            unite_destination         = (body.get("unite_destination") or "").strip(),
            statut                    = "EN_PREPARATION",
            nom                       = body.get("nom") or None,
            prenom                    = body.get("prenom") or None,
            nom_jeune_fille           = body.get("nom_jeune_fille") or None,
            ipp                       = body.get("ipp") or None,
            date_naissance            = body.get("date_naissance") or None,
            age_estime                = body.get("age_estime") or "",
            sexe                      = body.get("sexe") or "",
            motif                     = body.get("motif") or "",
            pathologie                = body.get("pathologie") or "",
            eta                       = body.get("eta") or "",
            commentaire               = body.get("commentaire") or "",
            redacteur                 = body.get("redacteur") or "Supervision",
            horodatage_creation       = datetime.now(timezone.utc),
        )
        db.add(t); db.commit(); db.refresh(t)
        _inject_tr_sup_to_inter(t)
        return _fmt_tr_coll(t)
    finally:
        db.close()


@app.patch("/api/admin/transferts-supervision/{tid}/statut")
async def tr_sup_patch_statut(tid: int, body: dict, _ok: bool = Depends(require_admin)):
    """Met à jour le statut d'un transfert supervision."""
    if not _TR_ENABLED:
        raise HTTPException(503, "Module transferts supervision non disponible")
    new_statut = (body.get("statut") or "").upper()
    if new_statut not in ("EN_PREPARATION", "EN_COURS", "ARRIVE", "ANNULE"):
        raise HTTPException(400, f"Statut invalide: {new_statut}")
    db = _TrSession()
    try:
        t = db.query(TransfertColl).filter(TransfertColl.id == tid).first()
        if not t: raise HTTPException(404, "Transfert introuvable")
        t.statut = new_statut
        now = datetime.now(timezone.utc)
        if new_statut == "EN_COURS"  and not t.horodatage_depart:  t.horodatage_depart  = now
        if new_statut == "ARRIVE"    and not t.horodatage_arrivee: t.horodatage_arrivee = now
        db.commit()
        _inject_tr_sup_to_inter(t)
        return _fmt_tr_coll(t)
    finally:
        db.close()


@app.put("/api/admin/transferts-supervision/{tid}")
async def tr_sup_update(tid: int, body: dict, _ok: bool = Depends(require_admin)):
    """Met à jour un transfert supervision."""
    if not _TR_ENABLED:
        raise HTTPException(503, "Module transferts supervision non disponible")
    db = _TrSession()
    try:
        t = db.query(TransfertColl).filter(TransfertColl.id == tid).first()
        if not t: raise HTTPException(404, "Transfert introuvable")
        for field in ("etablissement_origine","unite_origine","etablissement_destination",
                      "unite_destination","nom","prenom","nom_jeune_fille","ipp","date_naissance",
                      "age_estime","sexe","motif","pathologie","eta","commentaire","statut"):
            if field in body:
                setattr(t, field, body[field])
        db.commit()
        return _fmt_tr_coll(t)
    finally:
        db.close()



@app.post("/api/admin/transferts-supervision/{tid}/archiver")
async def tr_sup_archiver(tid: int, body: dict, _ok: bool = Depends(require_admin)):
    """h153 — Archive un transfert supervision avec l'auteur de l'archivage."""
    if not _TR_ENABLED:
        raise HTTPException(503, "Module transferts supervision non disponible")
    auteur = (body.get("auteur") or "Supervision").strip()[:100]
    db = _TrSession()
    try:
        t = db.query(TransfertColl).filter(TransfertColl.id == tid).first()
        if not t:
            raise HTTPException(404, "Transfert introuvable")
        if t.statut not in ("ARRIVE", "ANNULE", "EN_PREPARATION", "EN_COURS"):
            raise HTTPException(400, f"Statut {t.statut} ne peut pas être archivé")
        t.statut = "ARCHIVE"
        # Stocker l'auteur dans le commentaire si non renseigné
        mention = f"[Archivé par {auteur}]"
        t.commentaire = ((t.commentaire or "") + " " + mention).strip()
        db.commit()
        # Retirer de transferts_inter
        id_local = f"SUP-{tid}"
        transferts_inter[:] = [x for x in transferts_inter if x.get("id_local") != id_local]
        save_transferts_inter()
        _inject_tr_sup_to_inter(t)  # retire de transferts_inter (statut ARCHIVE)
        return _fmt_tr_coll(t)
    finally:
        db.close()


@app.delete("/api/admin/transferts-supervision/{tid}")
async def tr_sup_delete(tid: int, _ok: bool = Depends(require_admin)):
    """Supprime un transfert supervision."""
    if not _TR_ENABLED:
        raise HTTPException(503, "Module transferts supervision non disponible")
    db = _TrSession()
    try:
        t = db.query(TransfertColl).filter(TransfertColl.id == tid).first()
        if not t: raise HTTPException(404, "Transfert introuvable")
        db.delete(t); db.commit()
        return {"ok": True, "id": tid}
    finally:
        db.close()


@app.get("/api/admin/transferts-supervision/enabled")
async def tr_sup_status():
    """Retourne si le module transferts supervision est activé (pas d'auth : utilisé au chargement UI)."""
    return {"enabled": _TR_ENABLED}


@app.post("/api/push-transfert")
async def push_transfert_interght(request: Request, credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(status_code=401, detail="Non autorisé")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON invalide")
    transfert = {
        "id_local": body.get("id_local"),
        "ght_emetteur": sigle,
        "ght_emetteur_nom": body.get("ght_emetteur_nom", sigle),
        "ght_destinataire": body.get("ght_destinataire", ""),
        "unite_origine": body.get("unite_origine", ""),
        "etablissement_origine": body.get("etablissement_origine", ""),
        "unite_destination": body.get("unite_destination", ""),
        "etablissement_destination": body.get("etablissement_destination", ""),
        "site_destination": body.get("site_destination", ""),
        "statut": body.get("statut", "EN_PREPARATION"),
        "eta": body.get("eta"),
        "horodatage_depart": body.get("horodatage_depart"),
        "commentaire": body.get("commentaire"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    existing = next((t for t in transferts_inter
                     if t.get("id_local") == transfert["id_local"]
                     and t.get("ght_emetteur") == sigle), None)
    # h153 — Si id_local commence par "SUP-", c'est un transfert supervision
    id_local_str = str(transfert.get("id_local", ""))
    if id_local_str.startswith("SUP-") and _TR_ENABLED:
        try:
            sup_id = int(id_local_str[4:])
            db_tr = _TrSession()
            try:
                t_db = db_tr.query(TransfertColl).filter(TransfertColl.id == sup_id).first()
                if t_db:
                    new_statut = transfert.get("statut", t_db.statut)
                    t_db.statut = new_statut
                    now_utc = datetime.now(timezone.utc)
                    if new_statut == "EN_COURS" and not t_db.horodatage_depart:
                        t_db.horodatage_depart = now_utc
                    if new_statut == "ARRIVE" and not t_db.horodatage_arrivee:
                        t_db.horodatage_arrivee = now_utc
                    db_tr.commit()
                    # Propager le nouveau statut à TOUTES les instances via _inject
                    _inject_tr_sup_to_inter(t_db)
                    return {"ok": True}
            finally:
                db_tr.close()
        except Exception as e:
            print(f"[collecteur] push SUP transfert: {e}")

    # h154a — Confirmation d'arrivée par le DESTINATAIRE d'un transfert entrant.
    # Le token du destinataire donne sigle=destinataire, donc `existing`
    # (recherché sur ght_emetteur==sigle) est None et l'ancien code ajoutait une
    # entrée parasite : l'entrée de l'émetteur restait EN_PREPARATION/EN_COURS.
    # On retrouve ici l'entrée poussée par l'émetteur (ght_destinataire==sigle,
    # ght_emetteur != sigle) et on la met à jour pour que l'émetteur, qui poll
    # /api/transfert-statut, voie bien le statut ARRIVE.
    if transfert["statut"] in ("ARRIVE", "ANNULE"):
        _emetteur_cible = (body.get("ght_emetteur_nom") or body.get("etablissement_origine") or "").upper()
        _candidats = [t for t in transferts_inter
                      if t.get("id_local") == transfert["id_local"]
                      and t.get("ght_destinataire", "").upper() == sigle.upper()
                      and t.get("ght_emetteur", "").upper() != sigle.upper()]
        _entree_emetteur = None
        if _candidats:
            _entree_emetteur = next((t for t in _candidats
                                     if _emetteur_cible and t.get("ght_emetteur", "").upper() == _emetteur_cible),
                                    _candidats[0])
        if _entree_emetteur is not None:
            if transfert["statut"] == "ANNULE":
                transferts_inter[:] = [t for t in transferts_inter if t is not _entree_emetteur]
            else:
                # ARRIVE : marquer en place + garder 2h (fenêtre de poll émetteur
                # et d'affichage destinataire), comme le flux supervision.
                _entree_emetteur["statut"] = "ARRIVE"
                _entree_emetteur["arrived_at"] = datetime.now(timezone.utc).isoformat()
            save_transferts_inter()
            return {"ok": True}

    if transfert["statut"] in ("ARRIVE", "ANNULE"):
        # Émetteur qui pousse son propre transfert ARRIVE/ANNULE : suppression.
        transferts_inter[:] = [t for t in transferts_inter
                               if not (t.get("id_local") == transfert["id_local"]
                                       and t.get("ght_emetteur") == sigle)]
    elif existing:
        existing.update(transfert)
    else:
        transferts_inter.append(transfert)
    save_transferts_inter()
    return {"ok": True}

@app.get("/api/transferts-en-cours")
async def get_transferts_en_cours(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(status_code=401, detail="Non autorisé")
    from datetime import timedelta
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat()
    # Transferts ENTRANTS : destinés à cet établissement (pour affichage dans la colonne «Entrants»)
    # Entrants : destinataire = cet établissement
    # Inclure EN_PREPARATION, EN_COURS et ARRIVE (pendant 2h après arrivée)
    cutoff_arrive = (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat()
    entrants = [t for t in transferts_inter
                if t.get("ght_destinataire", "").upper() == sigle.upper()
                and (t.get("statut") in ("EN_PREPARATION", "EN_COURS")
                     or (t.get("statut") == "ARRIVE"
                         and (t.get("arrived_at", t.get("created_at", "")) >= cutoff_arrive)))]
    # h153 — Sortants supervision : cet établissement est l'émetteur
    # Inclure EN_PREPARATION, EN_COURS ET ARRIVE
    sortants = [t for t in transferts_inter
                if t.get("ght_emetteur", "").upper() == "SUPERVISION"
                and t.get("etablissement_origine", "").upper() == sigle.upper()
                and (t.get("statut") in ("EN_PREPARATION", "EN_COURS")
                     or (t.get("statut") == "ARRIVE"
                         and (t.get("arrived_at", t.get("created_at", "")) >= cutoff_arrive)))]
    # Fusionner en évitant les doublons
    ids_entrants = {t.get("id_local") for t in entrants}
    result = entrants + [t for t in sortants if t.get("id_local") not in ids_entrants]
    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return result

@app.get("/api/transfert-statut/{id_local}")
async def get_transfert_statut(
    id_local: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """Retourne le statut actuel d'un transfert (pour l'émetteur)."""
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(status_code=401, detail="Non autorisé")
    # Chercher dans transferts_inter
    t = next((t for t in transferts_inter
               if str(t.get("id_local")) == str(id_local)
               and t.get("ght_emetteur","").upper() == sigle.upper()), None)
    if t:
        return {"statut": t.get("statut"), "found": True}
    # Non trouvé = supprimé (ARRIVE ou ANNULE)
    return {"statut": "ARRIVE", "found": False}



def _check_any_auth(credentials) -> bool:
    """Accepte token établissement OU session UI valide."""
    if not credentials:
        return False
    tok = credentials.credentials
    # Token établissement
    if tok in tokens:
        return True
    # Token admin
    if tok == ADMIN_TOKEN:
        return True
    # Session UI
    if tok in ui_sessions:
        return True
    return False

# ════════════════════════════════════════════════════════════════════════════
# CHAT INTER-GHT
# ════════════════════════════════════════════════════════════════════════════

@app.post("/api/chat/messages")
async def chat_push_message(request: Request, credentials=Depends(security)):
    """Une instance pousse un message dans un salon territorial."""
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(401)
    global _chat_msg_counter
    try:
        raw = await request.body()
        import json as _json
        body = _json.loads(raw)
    except Exception as e:
        logger.warning(f"chat_push_message: erreur parsing body: {e}")
        raise HTTPException(400, "Corps JSON invalide")
    salon_nom = body.get("salon_nom", "général")
    if salon_nom not in chat_messages:
        chat_messages[salon_nom] = []
    # Anti-doublon : même sigle + même contenu dans les 5 dernières secondes
    now = datetime.now(timezone.utc).isoformat()
    contenu = body.get("contenu", "")
    recent = [m for m in chat_messages[salon_nom][-10:]
              if m.get("auteur_sigle") == sigle and m.get("contenu") == contenu]
    if recent:
        last = recent[-1]
        try:
            from datetime import timedelta
            last_ts = datetime.fromisoformat(last["horodatage"].replace("Z",""))
            delta = (datetime.now(timezone.utc) - last_ts.replace(tzinfo=timezone.utc)).total_seconds()
            if delta < 5:
                return {"ok": True, "detail": "doublon ignoré"}
        except Exception:
            pass
    _chat_msg_counter += 1
    # Stocker les PJs : pj_meta = URL directe (nouveau), pj_data/pj_inline = base64 (legacy)
    global _chat_pj_counter
    pjs = []
    # Nouveau format : URL directe vers l'instance source
    for pm in body.get("pj_meta", []):
        _chat_pj_counter += 1
        pj_key = f"ght-{_chat_pj_counter}"
        if pm.get("remote_url"):
            # Stocker l'URL directe (pas de base64)
            chat_pj_store[pj_key] = {
                "nom":        pm.get("nom", "fichier"),
                "remote_url": pm["remote_url"],
                "taille":     pm.get("taille", 0),
            }
        elif pm.get("dataUrl"):
            chat_pj_store[pj_key] = {
                "nom":     pm.get("nom", "fichier"),
                "dataUrl": pm["dataUrl"],
                "taille":  pm.get("taille", 0),
            }
        else:
            continue
        pjs.append({"id": pj_key, "nom": pm.get("nom",""), "taille": pm.get("taille",0)})
    # Legacy : base64 direct (garde compatibilité)
    for pj_data in body.get("pj_data", []) + body.get("pj_inline", []):
        dataUrl = pj_data.get("dataUrl", "")
        if dataUrl:
            _chat_pj_counter += 1
            pj_key = f"ght-{_chat_pj_counter}"
            chat_pj_store[pj_key] = {
                "nom":     pj_data.get("nom", "fichier"),
                "dataUrl": dataUrl,
                "taille":  pj_data.get("taille", 0),
            }
            pjs.append({"id": pj_key, "nom": pj_data.get("nom",""), "taille": pj_data.get("taille",0)})
    msg = {
        "id":           _chat_msg_counter,
        "salon_nom":    salon_nom,
        "auteur_nom":   body.get("auteur_nom", sigle),
        "auteur_sigle": sigle,
        "contenu":      contenu,
        "mentions":     body.get("mentions", []),
        "reply_to_id":  body.get("reply_to_id"),
        "horodatage":   now,
        "origine":      "ght",
        "pj":           pjs,
    }
    chat_messages[salon_nom].append(msg)
    # Garder max 500 messages par salon
    if len(chat_messages[salon_nom]) > 500:
        chat_messages[salon_nom] = chat_messages[salon_nom][-500:]
    return {"ok": True, "id": _chat_msg_counter}


@app.get("/api/chat/messages")
async def chat_get_messages(
    salon_nom: str = "général",
    since_id: int = 0,
    credentials=Depends(security)
):
    """Une instance récupère les messages d'un salon depuis un ID donné."""
    if not _check_any_auth(credentials):
        raise HTTPException(401)
    msgs = chat_messages.get(salon_nom, [])
    filtered = [m for m in msgs if m.get("id", 0) > since_id]
    return filtered[-100:]  # max 100 messages


@app.post("/api/chat/salons")
async def chat_create_salon(request: Request, credentials=Depends(security)):
    """Une instance crée/enregistre un salon territorial dans le collecteur."""
    if not _check_any_auth(credentials):
        raise HTTPException(401)
    body = await request.json()
    nom = body.get("nom", "").strip().lower()
    if nom and nom not in chat_messages:
        chat_messages[nom] = []
    # Retourner le salon créé dans le format objet
    salon_list = list(chat_messages.keys())
    idx = salon_list.index(nom) if nom in salon_list else 0
    return {"id": idx+1, "nom": nom, "type": "territorial",
            "icone": "💬", "couleur": "#7c3aed", "description": f"Salon #{nom}"}

@app.get("/api/chat/salons")
async def chat_list_salons(credentials=Depends(security)):
    """Liste des salons actifs dans le collecteur — format objet compatible chat.html."""
    if not _check_any_auth(credentials):
        raise HTTPException(401)
    # Retourner des objets {id, nom, type, icone, couleur, description}
    # avec des IDs stables basés sur le nom du salon
    SALON_ICONS = {
        "général": "💬", "coordination": "🎯", "transferts": "🚑",
        "logistique": "📦", "direction": "🏛️"
    }
    result = []
    for i, nom in enumerate(chat_messages.keys()):
        result.append({
            "id":          i + 1,
            "nom":         nom,
            "description": f"Salon territorial #{nom}",
            "couleur":     "#7c3aed",
            "icone":       SALON_ICONS.get(nom, "💬"),
            "type":        "territorial",
            "ordre":       i,
        })
    return result


@app.post("/api/chat/presence")
async def chat_push_presence(request: Request, credentials=Depends(security)):
    """Une instance pousse la présence de ses utilisateurs."""
    sigle = get_etab_from_token(credentials)
    if not sigle:
        raise HTTPException(401)
    body = await request.json()
    users = body.get("users", [])
    now = datetime.now(timezone.utc).isoformat()
    chat_presence[sigle] = [
        {"user_id": u.get("user_id"), "display_name": u.get("display_name"), "last_seen": now}
        for u in users
    ]
    return {"ok": True}


@app.get("/api/chat/presence")
async def chat_get_presence(credentials=Depends(security)):
    """Retourne la présence de tous les établissements."""
    if not _check_any_auth(credentials):
        raise HTTPException(401)
    # Nettoyer les présences > 2 minutes
    from datetime import timedelta
    cutoff = datetime.now(timezone.utc) - timedelta(minutes=2)
    result = {}
    for etab_sigle, users in chat_presence.items():
        active = []
        for u in users:
            try:
                ts = datetime.fromisoformat(u["last_seen"].replace("Z","")).replace(tzinfo=timezone.utc)
                if ts >= cutoff:
                    active.append(u)
            except Exception:
                pass
        if active:
            result[etab_sigle] = active
    return result

@app.get("/api/repondeur-lignes")
async def get_repondeur_lignes():
    """Lignes du répondeur ACTIVES par établissement, par DÉCOUVERTE directe
    (comme l'annuaire) : on interroge /status/public de chaque instance, qui
    expose lignes_repondeur (libellé + numéro, non-nominatif). Aucun couplage
    à federation.py : simple lecture publique. Map { SIGLE: [lignes] }."""
    import httpx, asyncio
    ports = list(range(8000, 8010)) + [7474]
    async def _fetch(port):
        try:
            async with httpx.AsyncClient(timeout=2.5) as client:
                ann = await client.get(f"http://127.0.0.1:{port}/api/v1/auth/annuaire-public")
                if ann.status_code != 200:
                    return None
                sig = (ann.json().get("sigle") or "").upper().strip()
                if not sig or sig == "?":
                    return None
                sp = await client.get(f"http://127.0.0.1:{port}/api/v1/status/public")
                lignes = sp.json().get("lignes_repondeur", []) if sp.status_code == 200 else []
                return (sig, lignes)
        except Exception:
            return None
    fetched = await asyncio.gather(*[_fetch(p) for p in ports])
    out = {}
    for item in fetched:
        if item and item[1]:
            out[item[0]] = item[1]
    return out


@app.get("/api/annuaire")
async def get_annuaire_interght():
    """h71 — Annuaire inter-établissements par DÉCOUVERTE directe : on interroge
    l'annuaire-public de chaque port d'instance connu, en parallèle, et on
    identifie chaque établissement par le sigle qu'il DÉCLARE lui-même. Plus de
    PORT_MAP codé en dur (qui cassait avec des sigles personnalisés comme
    « CH Rivemont » → tout retombait sur le port 8000). Dédup par sigle déclaré."""
    import httpx, asyncio
    ports = list(range(8000, 8010)) + [7474]   # instances nominales + démo
    async def _fetch(port):
        try:
            async with httpx.AsyncClient(timeout=2.5) as client:
                r = await client.get(f"http://127.0.0.1:{port}/api/v1/auth/annuaire-public")
                if r.status_code == 200:
                    return r.json()
        except Exception:
            return None
        return None
    fetched = await asyncio.gather(*[_fetch(p) for p in ports])
    result, seen = [], set()
    for data in fetched:
        if not data:
            continue
        sig = (data.get("sigle") or "").upper().strip()
        if not sig or sig == "?" or sig in seen:
            continue
        seen.add(sig)
        result.append(data)
    return result

@app.get("/api/summary")
async def get_summary():
    """Vue consolidée de tous les établissements. Sans auth."""
    now = time.time()
    result = []
    if not etablissements:
        logger.info(f"summary: VIDE — {len(tokens)} token(s) enregistré(s)")
    for sigle, data in etablissements.items():
        received_str = data.get("_received_at", "")
        try:
            received_ts = datetime.fromisoformat(received_str).timestamp()
            age_minutes = int((now - received_ts) / 60)
            fresh = age_minutes < 10
        except Exception:
            age_minutes = 999
            fresh = False
        etab_info = data.get("etablissement", {})
        # Prendre le pire niveau entre incidents et statut déclaré
        ORDRE = {"CRITIQUE":5,"CRISE":4,"INCIDENT_MAJEUR":3,"PERTURBE":2,"ALERTE":2,"MAINTENANCE":1,"OPERATIONNEL":0,"NOMINAL":0,"INCONNU":-1}
        niv_incidents = data.get("niveau_global", "NOMINAL")
        sp = data.get("_status_page") or {}
        niv_statut = sp.get("niveau_global", "NOMINAL") if sp.get("published") else "NOMINAL"
        # Mapper les niveaux statut public vers niveaux incidents
        MAP_STATUT = {"PERTURBE":"ALERTE","INCIDENT_MAJEUR":"CRISE","OPERATIONNEL":"NOMINAL","ALERTE":"ALERTE","CRITIQUE":"CRITIQUE"}
        niv_statut_mapped = MAP_STATUT.get(niv_statut, niv_statut)
        niv_final = niv_incidents if ORDRE.get(niv_incidents,0) >= ORDRE.get(niv_statut_mapped,0) else niv_statut_mapped
        result.append({
            "sigle":        sigle,
            # h153 — Si le nom est un fallback auto-généré "Site_PORT", utiliser le sigle comme nom
            "nom": (lambda n: sigle if (not n or (n.startswith("Site_") and n[5:].isdigit())) else n)(etab_info.get("nom", sigle) or ""),
            "niveau_global": niv_final,
            "kpis":         data.get("kpis", {}),
            "services_transverses": data.get("services_transverses", {}),
            "poles_impactes": data.get("poles_impactes", []),
            "sites":        data.get("sites", []),
            "latitude":     data.get("latitude"),
            "longitude":    data.get("longitude"),
            "incidents":    data.get("incidents", []),
            "received_at":  received_str,
            "age_minutes":  age_minutes,
            "fresh":        fresh,
            "_status_page": data.get("_status_page"),
            "_capacite":    data.get("_capacite"),
            # h153 — transferts actifs : natifs (poussés par l'instance) + supervision
            "transferts_actifs": [
                t for t in transferts_inter
                if (t.get("ght_destinataire", "").upper() == sigle.upper()
                    or (t.get("ght_emetteur","").upper() == "SUPERVISION"
                        and t.get("etablissement_origine","").upper() == sigle.upper()))
                and t.get("statut") in ("EN_PREPARATION", "EN_COURS", "ARRIVE")
            ],
        })
    return result

@app.get("/api/etablissement/{sigle}")
async def get_etablissement(sigle: str):
    if sigle not in etablissements:
        raise HTTPException(status_code=404, detail=f"Établissement {sigle} inconnu")
    return etablissements[sigle]

@app.get("/api/capacite")
async def get_capacite_all():
    result = []
    for sigle, data in etablissements.items():
        cap = data.get("_capacite")
        if not cap:
            continue
        result.append({
            "sigle":       sigle,
            "nom":         data.get("etablissement", {}).get("nom", sigle),
            "received_at": cap.get("_received_at"),
            "synthese":    cap.get("synthese", {}),
            "alertes":     cap.get("alertes", []),
            "nb_services": cap.get("nb_services", 0),
            "nb_alertes":  cap.get("nb_alertes", 0),
        })
    return result

@app.get("/api/status/{sigle}")
async def get_etab_status(sigle: str):
    data = etablissements.get(sigle, {})
    sp = data.get("_status_page")
    if not sp:
        return {"published": False, "sigle": sigle}
    return sp




# ── Interface webseule) ────────────────────────────────

DASHBOARD_HTML = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>SCRIBE — Supervision Territoriale</title>
<link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
<link rel="preconnect" href="https://fonts.googleapis.com">
<!-- polices terminal retirées v3000h165 — police système Suite, aucune dépendance externe -->
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
:root{
  --bg:#060608;--s1:#0d0d12;--s2:#111118;--s3:#18181f;
  --border:#1f1f2e;--border2:#2a2a3d;
  --text:#dde4f0;--muted:#4a5070;--muted2:#6b7494;
  --mono:system-ui,-apple-system,'Segoe UI',Roboto,'Helvetica Neue',Arial,sans-serif;--head:system-ui,-apple-system,'Segoe UI',Roboto,'Helvetica Neue',Arial,sans-serif;--body:system-ui,-apple-system,'Segoe UI',Roboto,'Helvetica Neue',Arial,sans-serif;
  --green:#00e5a0;--yellow:#f5c518;--orange:#ff7b2c;--red:#ff2d55;
  --blue:#3d9eff;--purple:#a855f7;--cyan:#00cfff;
}
body.light{
  --bg:#f8fafc;--s1:#ffffff;--s2:#f1f5f9;--s3:#e2e8f0;
  --border:#e2e8f0;--border2:#cbd5e1;
  --text:#0f172a;--muted:#64748b;--muted2:#475569;
  --green:#059669;--yellow:#d97706;--orange:#ea580c;--red:#dc2626;
  --blue:#2563eb;--purple:#7c3aed;--cyan:#0891b2;
}
body.light #kpi-bar{background:#eef0f3;border-bottom:1px solid #dfe3e8}
body.light #header{background:#ffffff;border-bottom:1px solid #e2e8f0}
body.light .tab-btn{color:#161616}
body.light .tab-btn.active{color:#003189;background:rgba(0,49,137,.08);border-color:rgba(0,49,137,.3)}
body.light .tab-btn:hover{color:#003189}
*{box-sizing:border-box;margin:0;padding:0}
html,body{height:100%;background:var(--bg);color:var(--text);font-family:var(--body);overflow:hidden}
/* scanlines supprimées v2.3.14 — lisibilité */
#app{display:flex;flex-direction:column;height:100vh}

/* KPI BAR */
#kpi-bar{display:flex;align-items:center;gap:0;background:#eef0f3;border-bottom:1px solid #dfe3e8;flex-shrink:0;height:44px}
.kpi-cell{display:flex;align-items:center;gap:7px;padding:0 16px;border-right:1px solid var(--border);height:100%}
.kpi-clickable{transition:background .15s}
.kpi-clickable:hover{background:rgba(255,255,255,.07)}

/* ── KPI Modale (v3000h31) ── */
#kpi-modal-backdrop{
  display:none;position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:9998;
  align-items:center;justify-content:center;backdrop-filter:blur(2px);
}
#kpi-modal-backdrop.show{display:flex}
#kpi-modal{
  background:var(--s0);border:1px solid var(--border);border-radius:8px;
  width:780px;max-width:95vw;max-height:85vh;display:flex;flex-direction:column;
  box-shadow:0 20px 60px rgba(0,0,0,.5);overflow:hidden;
}
.kpi-modal-hdr{
  display:flex;align-items:center;justify-content:space-between;
  padding:14px 18px;border-bottom:1px solid var(--border);
  background:linear-gradient(180deg,var(--s1),var(--s0));flex-shrink:0;
}
.kpi-modal-title{
  font-family:var(--head);font-size:14px;font-weight:700;letter-spacing:1.2px;
  color:var(--text);display:flex;align-items:center;gap:10px;
}
.kpi-modal-count{
  font-family:var(--mono);font-size:10px;background:var(--blue);color:white;
  padding:2px 9px;border-radius:10px;letter-spacing:1px;
}
.kpi-modal-close{
  background:transparent;color:var(--muted);border:1px solid var(--border);
  width:28px;height:28px;border-radius:4px;font-size:14px;cursor:pointer;
  display:flex;align-items:center;justify-content:center;
}
.kpi-modal-close:hover{background:var(--s2);color:var(--text)}
.kpi-modal-body{
  flex:1;overflow-y:auto;padding:14px 18px;display:flex;flex-direction:column;gap:14px;
}
.kpi-group{background:var(--s1);border:1px solid var(--border);border-radius:6px;overflow:hidden}
.kpi-group-hdr{
  display:flex;align-items:center;justify-content:space-between;
  padding:10px 14px;background:var(--s2);border-bottom:1px solid var(--border);
}
.kpi-group-name{
  font-family:var(--mono);font-size:11px;font-weight:700;letter-spacing:1px;color:var(--text);
  display:flex;align-items:center;gap:8px;
}
.kpi-group-level{
  font-family:var(--mono);font-size:8px;padding:2px 7px;border-radius:8px;letter-spacing:1px;
}
.kpi-group-count{
  font-family:var(--mono);font-size:9px;color:var(--muted);letter-spacing:1px;
}
.kpi-group-items{display:flex;flex-direction:column}
.kpi-item{
  padding:10px 14px;border-bottom:1px solid var(--border);
  display:flex;align-items:center;gap:12px;font-size:12px;
}
.kpi-item:last-child{border-bottom:none}
.kpi-item-urg{
  font-family:var(--mono);font-size:9px;font-weight:700;padding:2px 7px;
  border-radius:3px;flex-shrink:0;min-width:24px;text-align:center;letter-spacing:1px;
}
.kpi-item-urg.u4{background:rgba(225,0,15,.2);color:#e1000f}
.kpi-item-urg.u3{background:rgba(249,115,22,.2);color:#f97316}
.kpi-item-urg.u2{background:rgba(245,197,24,.15);color:#f5c518}
.kpi-item-urg.u1{background:rgba(0,207,255,.1);color:var(--cyan)}
.kpi-item-type{
  font-family:var(--mono);font-size:9px;color:var(--muted);letter-spacing:1px;
  background:var(--s2);padding:2px 6px;border-radius:3px;flex-shrink:0;
}
.kpi-item-text{flex:1;color:var(--text);line-height:1.4;min-width:0;word-break:break-word}
.kpi-item-site{
  font-family:var(--mono);font-size:9px;color:var(--muted2);
  margin-top:3px;letter-spacing:0.5px;
}
.kpi-modal-empty{
  text-align:center;padding:40px;color:var(--muted);font-size:12px;font-style:italic;font-family:var(--mono);
}
.kpi-label{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:10px;font-weight:600;letter-spacing:.5px;color:#6b7280;text-transform:uppercase}
.kpi-val{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:16px;font-weight:700;color:#1a1a2e}
.kpi-val.ok{color:var(--green)}.kpi-val.warn{color:var(--yellow)}.kpi-val.crit{color:var(--red)}
#kpi-title{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:16px;font-weight:800;letter-spacing:1px;color:#000091;padding:0 16px;margin-right:auto}
#kpi-clock{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:13px;font-weight:600;color:#374151;padding:0 14px}
#kpi-refresh{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:12px;color:#6b7280;padding:0 10px;display:flex;align-items:center;gap:5px}
.refresh-dot{width:5px;height:5px;border-radius:50%;background:var(--green);animation:blink 1s infinite}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.2}}

/* HEADER TABS */
#header{display:flex;align-items:center;background:var(--s2);border-bottom:1px solid var(--border);flex-shrink:0;padding:0 8px;height:46px;gap:2px;color:var(--text)}
.tab-btn{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:13px;font-weight:600;letter-spacing:.2px;padding:7px 14px;background:transparent;color:var(--text);border:1px solid transparent;border-radius:7px;cursor:pointer;transition:all .15s;white-space:nowrap}
.tab-btn:hover{color:#003189;border-color:#003189}
.tab-btn.active{color:#003189;background:rgba(0,49,137,.08);border-color:rgba(0,49,137,.3);font-weight:700}
.tab-spacer{flex:1}
#msg-badge-hdr{display:none;background:var(--red);color:#fff;font-size:8px;padding:1px 5px;border-radius:8px;margin-left:4px;font-weight:700}

/* MAIN AREA */
#main{flex:1;overflow:hidden;position:relative}
.tab-pane{display:none;width:100%;height:100%}
.tab-pane.active{display:flex}

/* ── SUPERVISION ── */
#pane-supervision{flex-direction:row;overflow:hidden}
#etab-left{width:220px;flex-shrink:0;background:var(--s1);border-right:1px solid var(--border);overflow-y:auto;display:flex;flex-direction:column}
@media(max-width:768px){#pane-supervision{flex-direction:column}#etab-left{width:100%;max-height:200px;border-right:none;border-bottom:1px solid var(--border)}#kpi-title{display:none}.kpi-cell{padding:0 8px}.tab-btn{font-size:8px;padding:3px 7px}}
#etab-left-header{padding:10px 12px 8px;border-bottom:1px solid var(--border);flex-shrink:0}
#etab-left-header .lh-title{font-family:var(--mono);font-size:8px;letter-spacing:1.5px;color:var(--muted);margin-bottom:6px}
#pending-section{border-bottom:1px solid var(--border);background:rgba(245,197,24,.04)}
#pending-header{padding:8px 12px;font-family:var(--mono);font-size:8px;letter-spacing:1px;color:var(--yellow);display:flex;align-items:center;gap:6px}
#pending-list{padding:0 8px 8px}
.pending-card{background:rgba(245,197,24,.06);border:1px solid rgba(245,197,24,.2);border-radius:5px;padding:8px 10px;margin-bottom:6px}
.pending-card-nom{font-family:var(--mono);font-size:9px;font-weight:700;color:var(--yellow);margin-bottom:2px}
.pending-card-sub{font-family:var(--mono);font-size:8px;color:var(--muted);margin-bottom:6px}
.pending-actions{display:flex;gap:5px}
.btn-accept{font-family:var(--mono);font-size:8px;padding:3px 9px;background:rgba(0,229,160,.12);border:1px solid var(--green);border-radius:3px;color:var(--green);cursor:pointer;flex:1}
.btn-accept:hover{background:rgba(0,229,160,.22)}
.btn-reject{font-family:var(--mono);font-size:8px;padding:3px 9px;background:rgba(255,45,85,.08);border:1px solid rgba(255,45,85,.3);border-radius:3px;color:var(--red);cursor:pointer}
.etab-card{padding:10px 12px;cursor:pointer;border-bottom:1px solid var(--border);transition:background .12s;display:flex;align-items:center;gap:8px}
.etab-card:hover{background:var(--s2)}
.etab-card.selected{background:var(--s3);border-left:2px solid var(--cyan)}
.etab-dot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.etab-dot.NOMINAL{background:var(--green)}.etab-dot.VEILLE{background:var(--green)}.etab-dot.OPERATIONNEL{background:var(--green)}
.etab-dot.ALERTE{background:var(--yellow)}.etab-dot.PERTURBE{background:var(--yellow)}.etab-dot.MAINTENANCE{background:var(--cyan)}
.etab-dot.CRISE{background:var(--orange);box-shadow:0 0 6px var(--orange)}.etab-dot.INCIDENT_MAJEUR{background:var(--orange);box-shadow:0 0 6px var(--orange)}
.etab-dot.CRITIQUE{background:var(--red);animation:pulse-dot 1s infinite}
@keyframes pulse-dot{0%,100%{box-shadow:0 0 4px var(--red)}50%{box-shadow:0 0 12px var(--red)}}
.etab-info{flex:1;min-width:0}
.etab-sigle{font-family:var(--mono);font-size:10px;font-weight:700;color:var(--text)}
.etab-nom{font-family:var(--mono);font-size:8px;color:var(--muted);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.etab-age{font-family:var(--mono);font-size:8px;color:var(--muted)}
.btn-disconnect{font-family:var(--mono);font-size:7px;padding:2px 6px;background:rgba(255,45,85,.06);border:1px solid rgba(255,45,85,.2);border-radius:3px;color:rgba(255,45,85,.6);cursor:pointer;opacity:0;transition:opacity .15s;white-space:nowrap}
.etab-card:hover .btn-disconnect{opacity:1}
.btn-disconnect:hover{background:rgba(255,45,85,.15);color:var(--red)}

/* Detail panel */
#detail-panel{flex:1;overflow-y:auto;padding:18px 20px;background:var(--bg)}
.dp-empty{display:flex;flex-direction:column;align-items:center;justify-content:center;height:100%;gap:10px;color:var(--muted)}
.dp-empty-icon{font-size:40px;opacity:.15}

/* ── CARTOGRAPHIE ── */
#pane-carto{flex-direction:column}
#map{flex:1;background:var(--s1)}

/* ── MESSAGERIE INTER-GHT ── */
#pane-messagerie{flex-direction:column;background:var(--bg)}
#msg-toolbar{display:flex;align-items:center;gap:10px;padding:10px 14px;border-bottom:1px solid var(--border);background:var(--s1);flex-shrink:0}
#msg-toolbar-title{font-family:var(--mono);font-size:9px;font-weight:700;letter-spacing:1px;color:var(--muted2)}
.msg-sub-btn{font-family:var(--mono);font-size:8px;padding:3px 10px;border-radius:4px;border:1px solid var(--border2);background:transparent;color:var(--muted);cursor:pointer;transition:all .15s}
.msg-sub-btn.active{background:rgba(0,207,255,.1);border-color:rgba(0,207,255,.3);color:var(--cyan)}
#btn-compose-ight{font-family:var(--mono);font-size:8px;padding:4px 12px;background:#2563eb;color:#fff;border:none;border-radius:4px;cursor:pointer;font-weight:700;margin-left:auto}
#msg-split{display:flex;flex:1;overflow:hidden}
#msg-list-pane{width:320px;flex-shrink:0;border-right:1px solid var(--border);overflow-y:auto;background:var(--s1)}
#msg-detail-pane{flex:1;overflow-y:auto;padding:28px 36px;background:var(--bg)}
.msg-item{padding:11px 14px;border-bottom:1px solid var(--border);cursor:pointer;transition:background .1s}
.msg-item:hover{background:var(--s2)}
.msg-item.unread{border-left:2px solid var(--cyan)}
.msg-from{font-family:var(--mono);font-size:9px;font-weight:700;color:var(--muted2)}
.msg-subj{font-family:var(--mono);font-size:9px;color:var(--text);margin:2px 0}
.msg-prev{font-family:var(--mono);font-size:8px;color:var(--muted);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.msg-date{font-family:var(--mono);font-size:8px;color:var(--muted)}

/* ── STATUTS PUBLICS ── */
#pane-statuts{flex-direction:column;background:var(--bg);overflow-y:auto;padding:16px}
.sp-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:16px}
.sp-card{background:var(--s1);border:1px solid var(--border2);border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06)}
.sp-card-hdr{padding:16px 18px;display:flex;align-items:flex-start;gap:12px}
.sp-card-body{padding:0 18px 16px}
.sp-svc-row{display:flex;align-items:center;gap:8px;padding:5px 0;font-size:13px}
.svc-dot{width:6px;height:6px;border-radius:50%;flex-shrink:0}

/* MODALS */
.modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:9999;align-items:center;justify-content:center}
.modal-overlay.open{display:flex}
.modal-box{background:var(--s2);border:1px solid var(--border2);border-radius:10px;width:460px;max-width:95vw;overflow:hidden}
.modal-hdr{padding:14px 18px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between}
.modal-hdr-title{font-family:var(--mono);font-size:10px;font-weight:700;color:var(--text)}
.modal-close{background:none;border:none;color:var(--muted);cursor:pointer;font-size:16px;line-height:1}
.modal-body{padding:18px;display:flex;flex-direction:column;gap:12px}
.form-label{font-family:var(--mono);font-size:8px;letter-spacing:1px;color:var(--muted);display:block;margin-bottom:4px}
.form-input{width:100%;font-family:var(--mono);font-size:10px;padding:7px 9px;background:var(--s3);border:1px solid var(--border2);border-radius:4px;color:var(--text);box-sizing:border-box}
.form-input:focus{outline:none;border-color:rgba(0,207,255,.4)}
.form-select{width:100%;font-family:var(--mono);font-size:10px;padding:7px 9px;background:var(--s3);border:1px solid var(--border2);border-radius:4px;color:var(--text)}
.form-textarea{width:100%;font-family:var(--mono);font-size:10px;padding:7px 9px;background:var(--s3);border:1px solid var(--border2);border-radius:4px;color:var(--text);resize:vertical;box-sizing:border-box}
.modal-footer{display:flex;justify-content:flex-end;gap:8px;padding:12px 18px;border-top:1px solid var(--border)}
.btn-cancel{font-family:var(--mono);font-size:9px;padding:5px 14px;background:var(--s3);border:1px solid var(--border2);border-radius:4px;color:var(--muted);cursor:pointer}
.btn-primary{font-family:var(--mono);font-size:9px;padding:5px 14px;background:#2563eb;border:none;border-radius:4px;color:#fff;cursor:pointer;font-weight:700}
.btn-primary:hover{background:#1d4ed8}

/* TOAST */
#toast{position:fixed;bottom:20px;right:20px;z-index:99999;font-family:var(--mono);font-size:10px;padding:10px 16px;border-radius:5px;display:none;animation:fadein .2s}
#toast.ok{background:#0d3320;border:1px solid var(--green);color:var(--green)}
#toast.err{background:#3d0d12;border:1px solid var(--red);color:var(--red)}
@keyframes fadein{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:none}}

/* Leaflet dark */
.leaflet-container{background:#060608}
.leaflet-tile-pane{filter:none}
/* h111 — Centre d'aide */
#help-overlay{display:none;position:fixed;inset:0;background:rgba(15,23,42,.45);z-index:99990;align-items:center;justify-content:center;padding:20px;font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif}
#help-overlay.open{display:flex}
#help-box{background:#fff;border-radius:14px;width:760px;max-width:96vw;height:80vh;max-height:680px;display:flex;flex-direction:column;overflow:hidden;box-shadow:0 18px 60px rgba(0,0,0,.28)}
#help-hdr{display:flex;align-items:center;justify-content:space-between;padding:16px 20px;border-bottom:1px solid #e7e9ee}
#help-title{font-size:16px;font-weight:700;color:#000091}
.help-lang-btn{font-size:12px;font-weight:600;padding:4px 9px;border:1px solid #d7dae0;background:#fff;color:#5b6573;border-radius:6px;cursor:pointer}
.help-lang-btn.active{background:#000091;color:#fff;border-color:#000091}
#help-close{background:none;border:none;font-size:18px;color:#8a909c;cursor:pointer;line-height:1;padding:2px 6px}
#help-search-wrap{padding:14px 20px 8px}
#help-search{width:100%;box-sizing:border-box;font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:14px;padding:10px 13px;border:1px solid #d1d5db;border-radius:9px;outline:none}
#help-search:focus{border-color:#000091;box-shadow:0 0 0 3px rgba(0,0,145,.12)}
#help-body{flex:1;overflow-y:auto;padding:8px 20px 20px}
.help-item{padding:12px 14px;border:1px solid #eceef2;border-radius:10px;margin-bottom:8px;cursor:pointer;transition:background .12s,border-color .12s}
.help-item:hover{background:#f5f6fb;border-color:#000091}
.help-item-cat{font-size:11px;font-weight:600;letter-spacing:.3px;color:#8a909c;text-transform:uppercase}
.help-item-title{font-size:14px;font-weight:600;color:#1a1a2e;margin-top:2px}
.help-empty{color:#8a909c;font-size:14px;padding:20px;text-align:center}
#help-article h2{font-size:18px;color:#000091;margin:0 0 12px}
#help-article p{font-size:14px;line-height:1.6;color:#374151;margin:0 0 10px}
#help-article ul{font-size:14px;line-height:1.6;color:#374151;margin:0 0 10px;padding-left:20px}
#help-article code{background:#f1f3f7;padding:1px 6px;border-radius:4px;font-size:13px}
#help-back{background:none;border:none;color:#000091;font-size:13px;font-weight:600;cursor:pointer;padding:0;margin-bottom:14px;font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif}

/* ════════════════════════════════════════════════════════════════════
   h142 — Alignement charte Suite numérique (DSFR)
   Surcharges scopées body.light (mode prod). Bleu France #000091.
   ════════════════════════════════════════════════════════════════════ */
/* ── Barre de menu : onglets soulignés DSFR (au lieu de pastilles) ── */
body.light #header{background:#ffffff;border-bottom:1px solid #dddddd;padding:0 12px;height:48px;gap:0}
body.light .tab-btn{font-size:14px;font-weight:500;letter-spacing:0;color:#3a3a3a;background:transparent;border:none;border-bottom:2px solid transparent;border-radius:0;padding:14px 14px;margin:0}
body.light .tab-btn:hover{color:#000091;background:#f5f5fe;border-color:transparent}
body.light .tab-btn.active{color:#000091;font-weight:700;background:transparent;border-bottom:2px solid #000091}
/* L'onglet Assistant force un style inline color:#003189 → réaligner */
body.light .tab-btn[onclick*="assistant"]{color:#3a3a3a}
body.light .tab-btn[onclick*="assistant"].active{color:#000091}
body.light .tab-btn[onclick*="exercice"]{border-left:1px solid #e5e5e5;margin-left:6px;padding-left:16px}

/* ── Colonne gauche : établissements enrôlés (DSFR) ── */
body.light #etab-left{width:252px;background:#ffffff;border-right:1px solid #dddddd}
body.light #etab-left-header{padding:16px 16px 10px;border-bottom:1px solid #ededed}
body.light #etab-left-header .lh-title{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:13px;font-weight:700;letter-spacing:.2px;color:#000091;text-transform:none;margin-bottom:0}
body.light .etab-card{padding:12px 16px;border-bottom:1px solid #ededed;gap:10px}
body.light .etab-card:hover{background:#f5f5fe}
body.light .etab-card.selected{background:#f5f5fe;border-left:3px solid #000091}
body.light .etab-dot{width:9px;height:9px}
body.light .etab-sigle{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:14px;font-weight:700;letter-spacing:0;color:#161616}
body.light .etab-nom{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:12px;letter-spacing:0;color:#666666}
body.light .etab-age{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:11px;letter-spacing:0;color:#929292}
/* Sections sync / relais / pending : libellés en Bleu France, neutres */
body.light #etab-left .lh-sub,
body.light #pending-header{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;font-size:11px;letter-spacing:.2px;color:#000091}
body.light .lh-btn{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif !important;font-size:11px !important;font-weight:500;padding:3px 10px !important;background:#ffffff !important;border:1px solid #000091 !important;border-radius:4px !important;color:#000091 !important;letter-spacing:0 !important}
body.light .lh-btn:hover{background:#f5f5fe !important}
body.light .lh-note{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif !important;font-size:11px !important;color:#666666 !important;letter-spacing:0 !important;line-height:1.4}

/* Pastilles de statut établissement (palette sémantique DSFR) */
body.light .etab-dot.NOMINAL,body.light .etab-dot.VEILLE,body.light .etab-dot.OPERATIONNEL{background:#18753c}
body.light .etab-dot.ALERTE,body.light .etab-dot.PERTURBE{background:#b34000}
body.light .etab-dot.MAINTENANCE{background:#0063cb}
body.light .etab-dot.CRISE,body.light .etab-dot.INCIDENT_MAJEUR{background:#ce0500;box-shadow:none}
body.light .etab-dot.CRITIQUE{background:#ce0500}

/* ── Statuts publics : cartes DSFR ── */
body.light #pane-statuts{background:#f5f5fe}
body.light .sp-card{background:#ffffff;border:1px solid #dddddd;border-radius:8px;box-shadow:none}
body.light .sp-card:hover{border-color:#000091}

</style>
</head>
<body>
<script>
// Mode clair par défaut
(function(){
  var t = localStorage.getItem('coll_theme');
  if(t !== 'dark') document.body.classList.add('light');
})();
</script>
<div id="login-screen" style="display:none;position:fixed;inset:0;background:#f8fafc;z-index:9999;align-items:center;justify-content:center">
<div style="background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:24px 36px 32px;width:360px;box-shadow:0 4px 24px rgba(0,0,0,.08);display:flex;flex-direction:column;gap:12px">
  <!-- v3.4 (h38m) — Sélecteur de langue accessible AVANT login, comme dans les
       instances. Cohérence visuelle entre master et instances. Le choix est
       persisté en localStorage ('scribe_lang_pref') et sera utilisé par le
       futur i18n du master (Bug C h38l backlog). -->
  <div style="display:flex;justify-content:flex-end;align-items:center;gap:6px;font-family:monospace;font-size:10px;color:#64748b;margin-bottom:-4px">
    <span>🌐</span>
    <select id="master-lang-select" onchange="window.changeMasterLanguage(this.value)"
            style="font-family:monospace;font-size:10px;padding:3px 6px;background:#fff;border:1px solid #cbd5e1;border-radius:3px;color:#0f172a;cursor:pointer">
      <option value="fr">Français</option>
      <option value="en">English</option>
      <option value="it">Italiano</option>
      <option value="de">Deutsch</option>
      <option value="es">Español</option>
      <option value="pt">Português</option>
      <option value="nl">Nederlands</option>
      <option value="pl">Polski</option>
      <option value="ro">Română</option>
      <option value="el">Ελληνικά</option>
      <option value="cs">Čeština</option>
      <option value="sk">Slovenčina</option>
      <option value="sv">Svenska</option>
      <option value="da">Dansk</option>
      <option value="fi">Suomi</option>
      <option value="hu">Magyar</option>
      <option value="bg">Български</option>
      <option value="hr">Hrvatski</option>
      <option value="sl">Slovenščina</option>
      <option value="et">Eesti</option>
      <option value="lt">Lietuvių</option>
      <option value="lv">Latviešu</option>
      <option value="mt">Malti</option>
      <option value="ga">Gaeilge</option>
    </select>
  </div>
  <div style="text-align:center"><img src="/static/logo-scribe.png" alt="SCRIBE" style="height:56px;object-fit:contain"></div>
  <div data-mi18n="login.subtitle" style="font-family:monospace;font-size:10px;letter-spacing:2px;text-align:center;color:#64748b;text-transform:uppercase">Supervision Territoriale</div>
  <div style="display:flex;flex-direction:column;gap:4px"><label data-mi18n="login.username" style="font-family:monospace;font-size:9px;color:#64748b;letter-spacing:1px;text-transform:uppercase;display:block;margin-bottom:3px">Identifiant</label><input id="coll-login" type="text" placeholder="supervision" autocomplete="username" style="padding:9px 12px;border:1px solid #e2e8f0;border-radius:5px;font-family:monospace;font-size:11px;color:#0f172a;background:#f8fafc;outline:none;box-sizing:border-box;width:100%"></div>
  <div style="display:flex;flex-direction:column;gap:4px"><label data-mi18n="login.password" style="font-family:monospace;font-size:9px;color:#64748b;letter-spacing:1px;text-transform:uppercase;display:block;margin-bottom:3px">Mot de passe</label><input id="coll-pass" type="password" placeholder="••••••••" autocomplete="current-password" onkeydown="if(event.key===String.fromCharCode(13))collLogin()" style="padding:9px 12px;border:1px solid #e2e8f0;border-radius:5px;font-family:monospace;font-size:11px;color:#0f172a;background:#f8fafc;outline:none;box-sizing:border-box;width:100%"></div>
  <button onclick="collLogin()" data-mi18n="login.button" style="width:100%;padding:10px;font-family:monospace;font-size:12px;font-weight:700;letter-spacing:2px;background:#003189;color:#fff;border:none;border-radius:5px;cursor:pointer;margin-top:4px">CONNEXION</button>
  <div id="coll-err" style="font-family:monospace;font-size:10px;color:#e1000f;text-align:center;min-height:14px"></div>
  <div id="coll-default-hint" style="display:none;background:#e3e3fd;border-radius:4px;padding:10px 12px;font-family:monospace;font-size:10px;color:#000091;line-height:1.5;text-align:left;margin-top:4px">
    <strong data-mi18n="login.first_time">Première connexion ?</strong><br>
    <span data-mi18n="login.username">Identifiant</span> : <code style="background:rgba(0,0,145,.1);padding:1px 4px;border-radius:2px">supervision</code><br>
    <span data-mi18n="login.password">Mot de passe</span> : <code style="background:rgba(0,0,145,.1);padding:1px 4px;border-radius:2px">Scribe2026!</code><br>
    <span data-mi18n="login.first_time_hint" style="font-size:9px;opacity:.7">À changer après le premier login (onglet Comptes).</span>
  </div>
  <!-- v3.4 (h38m) — Footer crédite SCRIBE (projet personnel, pas CHV).
       Liens hypertextes : nom → profil LinkedIn, "SCRIBE Crisis OS" → GitHub repo. -->
  <div style="font-family:monospace;font-size:9px;color:#94a3b8;text-align:center;margin-top:8px;line-height:1.6">
    <span data-mi18n="login.designed_by">Designed by</span> <a href="https://www.linkedin.com/in/%D0%BD%D0%BE-%D0%BA%D0%BE%D0%BC%D0%BF/" target="_blank" rel="noopener noreferrer" style="color:#003189;text-decoration:none;border-bottom:1px dotted #003189">SCRIBE</a><br>
    <a href="https://github.com/nocomp/scribe" target="_blank" rel="noopener noreferrer" style="color:#003189;text-decoration:none;border-bottom:1px dotted #003189">SCRIBE Crisis OS</a> · <span data-mi18n="login.opensource">open source</span> · AGPL-3.0
  </div>
</div>
</div>
<script>
// v3.4 (h38m) — Sélecteur de langue de la mire master.
// Sauve le choix en localStorage avec la même clé que les instances
// ('scribe_lang_pref'), pour cohérence cross-master/instance.
// v3.6.0-alpha6 — i18n minimal du login dans 24 langues UE pour cohérence
// interface ↔ sélecteur. Le dashboard interne reste en français tant que
// le bug C complet n'est pas traité.
window.MASTER_LOGIN_I18N = {
  fr: { "login.subtitle":"Supervision Territoriale", "login.username":"Identifiant", "login.password":"Mot de passe", "login.button":"CONNEXION", "login.first_time":"Première connexion ?", "login.first_time_hint":"À changer après le premier login (onglet Comptes).", "login.designed_by":"Designed by", "login.opensource":"open source" },
  en: { "login.subtitle":"Territorial Supervision", "login.username":"Username", "login.password":"Password", "login.button":"LOG IN", "login.first_time":"First time?", "login.first_time_hint":"To be changed after first login (Accounts tab).", "login.designed_by":"Designed by", "login.opensource":"open source" },
  it: { "login.subtitle":"Supervisione Territoriale", "login.username":"Identificativo", "login.password":"Password", "login.button":"ACCEDI", "login.first_time":"Primo accesso?", "login.first_time_hint":"Da cambiare dopo il primo login (scheda Account).", "login.designed_by":"Progettato da", "login.opensource":"open source" },
  de: { "login.subtitle":"Territoriale Aufsicht", "login.username":"Benutzername", "login.password":"Passwort", "login.button":"ANMELDEN", "login.first_time":"Erste Anmeldung?", "login.first_time_hint":"Nach erster Anmeldung ändern (Tab Konten).", "login.designed_by":"Entwickelt von", "login.opensource":"Open Source" },
  es: { "login.subtitle":"Supervisión Territorial", "login.username":"Identificador", "login.password":"Contraseña", "login.button":"ACCEDER", "login.first_time":"¿Primera vez?", "login.first_time_hint":"Cambiar tras el primer inicio (pestaña Cuentas).", "login.designed_by":"Diseñado por", "login.opensource":"código abierto" },
  nl: { "login.subtitle":"Territoriaal Toezicht", "login.username":"Gebruikersnaam", "login.password":"Wachtwoord", "login.button":"INLOGGEN", "login.first_time":"Eerste keer?", "login.first_time_hint":"Wijzigen na eerste login (tab Accounts).", "login.designed_by":"Ontworpen door", "login.opensource":"open source" },
  pt: { "login.subtitle":"Supervisão Territorial", "login.username":"Identificador", "login.password":"Palavra-passe", "login.button":"ENTRAR", "login.first_time":"Primeira vez?", "login.first_time_hint":"A mudar após o primeiro login (separador Contas).", "login.designed_by":"Concebido por", "login.opensource":"código aberto" },
  pl: { "login.subtitle":"Nadzór Terytorialny", "login.username":"Identyfikator", "login.password":"Hasło", "login.button":"ZALOGUJ", "login.first_time":"Pierwsze logowanie?", "login.first_time_hint":"Zmień po pierwszym logowaniu (karta Konta).", "login.designed_by":"Zaprojektowane przez", "login.opensource":"open source" },
  ro: { "login.subtitle":"Supraveghere Teritorială", "login.username":"Identificator", "login.password":"Parolă", "login.button":"AUTENTIFICARE", "login.first_time":"Prima conectare?", "login.first_time_hint":"De schimbat după prima conectare (fila Conturi).", "login.designed_by":"Conceput de", "login.opensource":"open source" },
  el: { "login.subtitle":"Εδαφική Εποπτεία", "login.username":"Όνομα χρήστη", "login.password":"Κωδικός", "login.button":"ΣΥΝΔΕΣΗ", "login.first_time":"Πρώτη φορά;", "login.first_time_hint":"Αλλαγή μετά την πρώτη σύνδεση (καρτέλα Λογαριασμοί).", "login.designed_by":"Σχεδιάστηκε από", "login.opensource":"ανοιχτού κώδικα" },
  cs: { "login.subtitle":"Územní Dohled", "login.username":"Uživatelské jméno", "login.password":"Heslo", "login.button":"PŘIHLÁSIT", "login.first_time":"První přihlášení?", "login.first_time_hint":"Změnit po prvním přihlášení (záložka Účty).", "login.designed_by":"Navrženo", "login.opensource":"open source" },
  sk: { "login.subtitle":"Územný Dohľad", "login.username":"Používateľské meno", "login.password":"Heslo", "login.button":"PRIHLÁSIŤ", "login.first_time":"Prvé prihlásenie?", "login.first_time_hint":"Zmeniť po prvom prihlásení (záložka Účty).", "login.designed_by":"Navrhol", "login.opensource":"open source" },
  sv: { "login.subtitle":"Territoriell Tillsyn", "login.username":"Användarnamn", "login.password":"Lösenord", "login.button":"LOGGA IN", "login.first_time":"Första gången?", "login.first_time_hint":"Ändra efter första inloggningen (flik Konton).", "login.designed_by":"Designad av", "login.opensource":"öppen källkod" },
  da: { "login.subtitle":"Territorialt Tilsyn", "login.username":"Brugernavn", "login.password":"Adgangskode", "login.button":"LOG IND", "login.first_time":"Første gang?", "login.first_time_hint":"Skal ændres efter første login (fanen Konti).", "login.designed_by":"Designet af", "login.opensource":"open source" },
  fi: { "login.subtitle":"Alueellinen Valvonta", "login.username":"Käyttäjätunnus", "login.password":"Salasana", "login.button":"KIRJAUDU", "login.first_time":"Ensimmäinen kerta?", "login.first_time_hint":"Vaihdettava ensimmäisen kirjautumisen jälkeen (Tilit-välilehti).", "login.designed_by":"Suunnitellut", "login.opensource":"avoin lähdekoodi" },
  hu: { "login.subtitle":"Területi Felügyelet", "login.username":"Felhasználónév", "login.password":"Jelszó", "login.button":"BELÉPÉS", "login.first_time":"Első alkalom?", "login.first_time_hint":"Az első bejelentkezés után módosítandó (Fiókok lap).", "login.designed_by":"Tervezte", "login.opensource":"nyílt forráskód" },
  bg: { "login.subtitle":"Териториален Надзор", "login.username":"Потребител", "login.password":"Парола", "login.button":"ВХОД", "login.first_time":"За първи път?", "login.first_time_hint":"Промяна след първото влизане (раздел Акаунти).", "login.designed_by":"Проектирано от", "login.opensource":"отворен код" },
  hr: { "login.subtitle":"Teritorijalni Nadzor", "login.username":"Korisničko ime", "login.password":"Lozinka", "login.button":"PRIJAVA", "login.first_time":"Prvi put?", "login.first_time_hint":"Promijenite nakon prve prijave (kartica Računi).", "login.designed_by":"Dizajnirao", "login.opensource":"otvoreni kod" },
  sl: { "login.subtitle":"Teritorialni Nadzor", "login.username":"Uporabniško ime", "login.password":"Geslo", "login.button":"PRIJAVA", "login.first_time":"Prvič?", "login.first_time_hint":"Spremenite po prvi prijavi (zavihek Računi).", "login.designed_by":"Oblikoval", "login.opensource":"odprta koda" },
  et: { "login.subtitle":"Territoriaalne Järelevalve", "login.username":"Kasutajanimi", "login.password":"Parool", "login.button":"LOGI SISSE", "login.first_time":"Esimene kord?", "login.first_time_hint":"Muutke pärast esimest sisselogimist (vahekaart Kontod).", "login.designed_by":"Kujundas", "login.opensource":"avatud lähtekood" },
  lt: { "login.subtitle":"Teritorinė Priežiūra", "login.username":"Naudotojo vardas", "login.password":"Slaptažodis", "login.button":"PRISIJUNGTI", "login.first_time":"Pirmą kartą?", "login.first_time_hint":"Pakeisti po pirmojo prisijungimo (skirtukas Paskyros).", "login.designed_by":"Sukūrė", "login.opensource":"atviras kodas" },
  lv: { "login.subtitle":"Teritoriālā Uzraudzība", "login.username":"Lietotājvārds", "login.password":"Parole", "login.button":"PIESLĒGTIES", "login.first_time":"Pirmo reizi?", "login.first_time_hint":"Mainīt pēc pirmās pieteikšanās (cilne Konti).", "login.designed_by":"Izstrādāja", "login.opensource":"atvērtais kods" },
  mt: { "login.subtitle":"Superviżjoni Territorjali", "login.username":"Isem tal-utent", "login.password":"Password", "login.button":"IDĦOL", "login.first_time":"L-ewwel darba?", "login.first_time_hint":"Ibdel wara l-ewwel dħul (tab Kontijiet).", "login.designed_by":"Iddisinjat minn", "login.opensource":"open source" },
  ga: { "login.subtitle":"Maoirsiú Críochach", "login.username":"Ainm úsáideora", "login.password":"Pasfhocal", "login.button":"LOGÁIL ISTEACH", "login.first_time":"An chéad uair?", "login.first_time_hint":"Le hathrú tar éis an chéad logála isteach (cluaisín Cuntais).", "login.designed_by":"Deartha ag", "login.opensource":"foinse oscailte" }
};

window.applyMasterLoginI18n = function() {
  var lang = 'fr';
  try { var s = localStorage.getItem('scribe_lang_pref'); if (s) lang = s; } catch(e) {}
  var dict = window.MASTER_LOGIN_I18N[lang] || window.MASTER_LOGIN_I18N.en;
  document.querySelectorAll('[data-mi18n]').forEach(function(el) {
    var key = el.getAttribute('data-mi18n');
    var val = dict[key];
    if (val) el.textContent = val;
  });
};

window.changeMasterLanguage = function(code) {
  try {
    localStorage.setItem('scribe_lang_pref', code);
    window.location.reload();
  } catch(e) {
    console.error('changeMasterLanguage failed', e);
  }
};
(function() {
  function syncMasterLang() {
    var sel = document.getElementById('master-lang-select');
    if (!sel) return;
    var current = 'fr';
    try {
      var stored = localStorage.getItem('scribe_lang_pref');
      if (stored) current = stored;
    } catch(e) {}
    for (var i = 0; i < sel.options.length; i++) {
      if (sel.options[i].value === current) { sel.selectedIndex = i; break; }
    }
    // Appliquer aussi les traductions au login
    if (window.applyMasterLoginI18n) window.applyMasterLoginI18n();
  }
  if (document.readyState === 'loading') {
    document.addEventListener('DOMContentLoaded', syncMasterLang);
  } else {
    syncMasterLang();
  }
})();
</script>
<script>
(async function() {
  // 1. Si un token existe en localStorage, vérifier sa validité
  var existingTok = localStorage.getItem('coll_session');
  if (existingTok) {
    var vr = await fetch('api/ui/verify', {
      headers: {'Authorization': 'Bearer ' + existingTok}
    }).then(function(r) { return r.ok ? r.json() : null; }).catch(function() { return null; });
    if (vr && vr.ok) {
      // Session valide → ne pas afficher le login, laisser l'app démarrer
      return;
    }
    // Session expirée → nettoyer
    localStorage.removeItem('coll_session');
  }
  // 2. Vérifier si auth requise
  const r = await fetch('api/ui/auth-required').catch(()=>null);
  if (!r || !r.ok) return;
  const d = await r.json();
  if (!d.required) return;
  // Auth requise — afficher l'écran de login
  const ls = document.getElementById('login-screen');
  const app = document.getElementById('app');
  if (ls) { ls.style.display='flex'; }
  if (app) { app.style.display='none'; }

  // Si premier lancement (aucune instance configurée), afficher le hint
  // avec les credentials par défaut pour que le nouvel utilisateur ne soit
  // pas perdu devant l'écran de login.
  fetch('api/ui/first-launch').then(r => r.ok ? r.json() : null).then(s => {
    if (s && s.first_launch) {
      const hint = document.getElementById('coll-default-hint');
      if (hint) hint.style.display = 'block';
      // Pré-remplir le champ login pour faciliter
      const loginInp = document.getElementById('coll-login');
      if (loginInp && !loginInp.value) loginInp.value = 'supervision';
    }
  }).catch(()=>{});
})();

async function collLogin() {
  const login = document.getElementById('coll-login').value;
  const pass  = document.getElementById('coll-pass').value;
  const r = await fetch('api/ui/login', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({login, password: pass})
  }).catch(()=>null);
  if (!r || !r.ok) {
    document.getElementById('coll-err').textContent = 'Identifiants incorrects';
    return;
  }
  const d = await r.json();
  if (d.ok) {
    localStorage.setItem('coll_session', d.token);
    // v3.4 (h38e) — Si l'utilisateur utilise encore le mdp par défaut
    // (Scribe2026!), on force le changement AVANT d'entrer dans l'app.
    if (d.must_change_password) {
      // On garde la mire de login affichée jusqu'au changement réussi
      _ui_mcp_username = login;
      openMasterForcedPasswordChange();
      return;
    }
    document.getElementById('login-screen').style.display = 'none';
    document.getElementById('app').style.display = 'flex';
  }
}

// v3.4 (h38e) — Modale de changement de mot de passe obligatoire master.
// Stratégie identique aux instances SCRIBE : 3 champs (ancien, nouveau,
// confirmation), icône œil pour révéler, déconnexion auto après succès
// pour forcer une reconnexion propre avec le nouveau mdp.
let _ui_mcp_username = '';

function _ui_togglePw(inputId, btn) {
  const el = document.getElementById(inputId);
  if (!el) return;
  if (el.type === 'password') { el.type = 'text'; btn.textContent = '🙈'; }
  else { el.type = 'password'; btn.textContent = '👁'; }
}

function openMasterForcedPasswordChange() {
  let m = document.getElementById('master-mcp-modal');
  if (m) { m.style.display = 'flex'; return; }
  m = document.createElement('div');
  m.id = 'master-mcp-modal';
  m.style.cssText = 'position:fixed;inset:0;background:rgba(0,0,0,.85);z-index:10000;display:flex;align-items:center;justify-content:center;padding:20px;font-family:monospace';
  m.innerHTML = `
    <div style="background:#fff;border:2px solid #003189;border-radius:12px;width:440px;max-width:95vw;overflow:hidden">
      <div style="padding:16px 20px;background:rgba(0,49,137,.1);border-bottom:1px solid rgba(0,49,137,.3);display:flex;align-items:center;gap:10px">
        <span style="font-size:20px">🔐</span>
        <div>
          <div style="font-size:12px;font-weight:700;color:#003189;letter-spacing:1px">CHANGEMENT DE MOT DE PASSE REQUIS</div>
          <div style="font-size:10px;color:#78716c;margin-top:4px;line-height:1.5">Vous utilisez encore le mot de passe par défaut. Veuillez le modifier avant d'accéder à la supervision.</div>
        </div>
      </div>
      <div style="padding:20px;display:flex;flex-direction:column;gap:14px">
        <div>
          <label style="font-size:9px;color:#64748b;letter-spacing:1px;display:block;margin-bottom:5px">MOT DE PASSE ACTUEL</label>
          <div style="position:relative">
            <input id="mcp-old" type="password" placeholder="Mot de passe actuel" autocomplete="current-password"
                   style="width:100%;font-size:12px;padding:9px 36px 9px 12px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:5px;color:#0f172a;box-sizing:border-box">
            <button type="button" onclick="_ui_togglePw('mcp-old', this)" tabindex="-1"
                    style="position:absolute;right:6px;top:50%;transform:translateY(-50%);background:transparent;border:none;cursor:pointer;color:#94a3b8;padding:4px;font-size:14px;line-height:1">👁</button>
          </div>
        </div>
        <div>
          <label style="font-size:9px;color:#64748b;letter-spacing:1px;display:block;margin-bottom:5px">NOUVEAU MOT DE PASSE (min. 8 caractères)</label>
          <div style="position:relative">
            <input id="mcp-new" type="password" placeholder="Nouveau mot de passe" autocomplete="new-password"
                   style="width:100%;font-size:12px;padding:9px 36px 9px 12px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:5px;color:#0f172a;box-sizing:border-box">
            <button type="button" onclick="_ui_togglePw('mcp-new', this)" tabindex="-1"
                    style="position:absolute;right:6px;top:50%;transform:translateY(-50%);background:transparent;border:none;cursor:pointer;color:#94a3b8;padding:4px;font-size:14px;line-height:1">👁</button>
          </div>
        </div>
        <div>
          <label style="font-size:9px;color:#64748b;letter-spacing:1px;display:block;margin-bottom:5px">CONFIRMER LE NOUVEAU MOT DE PASSE</label>
          <div style="position:relative">
            <input id="mcp-conf" type="password" placeholder="Confirmer" autocomplete="new-password"
                   onkeydown="if(event.key==='Enter')submitMasterForcedPw()"
                   style="width:100%;font-size:12px;padding:9px 36px 9px 12px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:5px;color:#0f172a;box-sizing:border-box">
            <button type="button" onclick="_ui_togglePw('mcp-conf', this)" tabindex="-1"
                    style="position:absolute;right:6px;top:50%;transform:translateY(-50%);background:transparent;border:none;cursor:pointer;color:#94a3b8;padding:4px;font-size:14px;line-height:1">👁</button>
          </div>
        </div>
        <div id="mcp-err" style="display:none;font-size:10px;color:#e1000f;padding:6px 10px;background:rgba(225,0,15,.08);border-radius:4px;border:1px solid rgba(225,0,15,.2)"></div>
        <button onclick="submitMasterForcedPw()" style="font-size:12px;padding:11px;background:#003189;border:none;border-radius:5px;color:#fff;cursor:pointer;font-weight:700;margin-top:4px">
          🔒 Définir mon nouveau mot de passe
        </button>
      </div>
    </div>
  `;
  document.body.appendChild(m);
  setTimeout(() => { try { document.getElementById('mcp-old').focus(); } catch(_){} }, 50);
}

async function submitMasterForcedPw() {
  const cur  = document.getElementById('mcp-old')?.value || '';
  const nw   = document.getElementById('mcp-new')?.value || '';
  const conf = document.getElementById('mcp-conf')?.value || '';
  const errEl = document.getElementById('mcp-err');
  errEl.style.display = 'none';
  if (nw.length < 8) {
    errEl.textContent = 'Le mot de passe doit contenir au moins 8 caractères.';
    errEl.style.display = 'block'; return;
  }
  if (nw !== conf) {
    errEl.textContent = 'Les mots de passe ne correspondent pas.';
    errEl.style.display = 'block'; return;
  }
  if (cur === nw) {
    errEl.textContent = 'Le nouveau mot de passe doit être différent de l\'ancien.';
    errEl.style.display = 'block'; return;
  }
  try {
    const tok = localStorage.getItem('coll_session') || '';
    const r = await fetch('api/ui/change-password', {
      method: 'POST',
      headers: {'Content-Type': 'application/json', 'Authorization': 'Bearer ' + tok},
      body: JSON.stringify({current_password: cur, new_password: nw})
    });
    if (!r.ok) {
      const d = await r.json().catch(() => ({}));
      errEl.textContent = d.detail || 'Mot de passe actuel incorrect.';
      errEl.style.display = 'block'; return;
    }
    // Succès : on déconnecte et on remet la mire de login
    const m = document.getElementById('master-mcp-modal');
    if (m) m.remove();
    localStorage.removeItem('coll_session');
    document.getElementById('coll-err').textContent = '';
    document.getElementById('coll-err').style.color = '#10b981';
    document.getElementById('coll-err').textContent = '✓ Mot de passe modifié — Reconnectez-vous';
    document.getElementById('coll-pass').value = '';
    document.getElementById('coll-pass').focus();
    setTimeout(() => {
      document.getElementById('coll-err').style.color = '#e1000f';
      document.getElementById('coll-err').textContent = '';
    }, 5000);
  } catch(e) {
    errEl.textContent = 'Erreur réseau. Réessayez.';
    errEl.style.display = 'block';
  }
}
</script>
<div id="app">

  <!-- KPI BAR -->
  <div id="kpi-bar">
    <div id="kpi-title"><img src="/static/logo-scribe.png" alt="SCRIBE" style="height:24px;vertical-align:middle;margin-right:8px;object-fit:contain">SUPERVISION v3.6.0-alpha119</div>
    <div class="kpi-cell"><span class="kpi-label">GHT</span><span class="kpi-val" id="k-ght">—</span></div>
    <div class="kpi-cell" style="cursor:pointer" title="Délai avant masquage incidents résolus (clic pour modifier)">
      <span class="kpi-label">RÉSOLU → masqué</span>
      <span class="kpi-val" id="k-hide-delay" onclick="promptHideDelay()" style="font-size:13px">30min</span>
    </div>
    <div class="kpi-cell kpi-clickable" style="cursor:pointer" onclick="openKpiModal('sites')" title="Cliquer pour voir tous les sites par établissement"><span class="kpi-label">Sites</span><span class="kpi-val" id="k-etab" title="">—</span></div>
    <div class="kpi-cell kpi-clickable" style="cursor:pointer" onclick="openKpiModal('pending')" title="Cliquer pour voir les établissements en attente d'enrôlement"><span class="kpi-label">En attente</span><span class="kpi-val warn" id="k-pending">0</span></div>
    <div class="kpi-cell kpi-clickable" style="cursor:pointer" onclick="openKpiModal('incidents')" title="Cliquer pour voir les incidents actifs groupés par établissement"><span class="kpi-label">Incidents actifs</span><span class="kpi-val" id="k-inc">—</span></div>
    <div class="kpi-cell kpi-clickable" style="cursor:pointer" onclick="openKpiModal('critiques')" title="Cliquer pour voir les incidents critiques (U≥3) groupés par établissement"><span class="kpi-label">Critiques</span><span class="kpi-val crit" id="k-crit">—</span></div>
    <div class="kpi-cell"><span class="kpi-label">Messages non lus</span><span class="kpi-val" id="k-msg">—</span></div>
    <div id="kpi-clock">--:--:--</div>
    <div id="kpi-refresh" style="display:flex;align-items:center;gap:6px">
      <div class="refresh-dot"></div>
      <span id="refresh-countdown">30s</span>
      <button onclick="fetchAll()" title="Actualiser maintenant" style="font-family:var(--mono);font-size:12px;padding:5px 11px;background:#fff;border:1px solid #c7cdd6;border-radius:5px;color:#000091;font-weight:700;cursor:pointer">&#8635; ACTUALISER</button><button onclick="openHelpCenter()" title="Centre d'aide" style="font-family:var(--mono);font-size:12px;padding:5px 11px;margin-left:6px;background:#000091;border:1px solid #000091;border-radius:5px;color:#fff;font-weight:700;cursor:pointer">? AIDE</button>
    </div>
  </div>

  <!-- TABS -->
  <div id="header">
    <button class="tab-btn active" onclick="switchTab('supervision',this)">⬡ SUPERVISION</button>
    <button class="tab-btn" onclick="switchTab('assistant',this)" style="color:#003189;font-weight:700">🎓 ASSISTANT<span id="ta-tab-badge" style="background:#e1000f;color:white;font-size:8px;font-weight:700;padding:1px 5px;border-radius:8px;display:none;margin-left:4px">0</span></button>
    <button class="tab-btn" onclick="switchTab('carto',this)">⊕ CARTOGRAPHIE</button>
    <button class="tab-btn" onclick="switchTab('messagerie',this)">✉ MESSAGES<span id="msg-badge-hdr"></span></button>
    <button class="tab-btn" id="tab-btn-transferts-sup" onclick="switchTab('transferts',this)" style="display:none">🚑 TRANSFERTS</button>
    <button class="tab-btn" onclick="switchTab('statuts',this)">▦ STATUTS PUBLICS</button>
    <button class="tab-btn" onclick="switchTab('chat',this)">💬 CHAT</button>
    <button class="tab-btn" onclick="switchTab('comptes',this)">&#128100; COMPTES</button>
    <button class="tab-btn" onclick="switchTab('instances',this)">📦 INSTANCES</button>
    <button class="tab-btn" onclick="switchTab('admin',this)">⚙️ ADMIN</button>
    <button class="tab-btn" onclick="switchTab('exercice',this)" style="border-left:1px solid var(--border);margin-left:6px;padding-left:14px" title="Mode exercice / simulation">🎯 EXERCICE</button>
    <div class="tab-spacer"></div>
  </div>

  <!-- MAIN -->
  <div id="main">

    <!-- SUPERVISION -->
    <div class="tab-pane active" id="pane-supervision">
      <div id="etab-left">
        <div id="etab-left-header">
          <div class="lh-title">Établissements enrôlés</div>
        </div>
        <!-- Pending section (hidden if empty) -->
        <div id="pending-section" style="display:none">
          <div id="pending-header" style="display:flex;align-items:center;justify-content:space-between">
            <span>⏳ EN ATTENTE <span id="pending-count">0</span></span>
            <button onclick="acceptAll()" id="btn-accept-all" class="lh-btn" style="display:none;cursor:pointer">✓ Tout accepter</button>
          </div>
          <div id="pending-list"></div>
        </div>
        <!-- Relay section -->
        <!-- Fix tokens section -->
        <div style="border-bottom:1px solid var(--border);padding:6px 12px;background:rgba(249,115,22,.04)">
          <div class="lh-sub" style="margin-bottom:4px;display:flex;align-items:center;justify-content:space-between">
            <span>🔧 INSTANCES SYNCHRONISÉES</span>
            <button onclick="registerscribeTokens()" class="lh-btn" style="cursor:pointer">
              ⚡ Enregistrer les instances
            </button>
          </div>
          <div class="lh-note">Force l'enregistrement des tokens des instances connues du master. Cliquez si une instance n'apparaît pas après démarrage.</div>
        </div>
        <div id="relay-section" style="border-bottom:1px solid var(--border);padding:8px 12px">
          <div class="lh-sub" style="margin-bottom:6px;display:flex;align-items:center;justify-content:space-between">
            <span>⇪ RELAIS UPSTREAM</span>
            <button onclick="openAddRelay()" class="lh-btn" style="cursor:pointer">+ Ajouter</button>
          </div>
          <div id="relay-list"><span class="lh-note">Aucun relais configuré</span></div>
        </div>
        <!-- Enrolled list -->
        <div id="etab-list"></div>
      </div>
      <div id="detail-panel">
        <div class="dp-empty">
          <div class="dp-empty-icon">⬡</div>
          <span style="font-family:var(--mono);font-size:10px;letter-spacing:1px">Sélectionnez un établissement</span>
        </div>
      </div>
    </div>

    <!-- ASSISTANT TERRITORIAL (v3000h28) -->
    <div class="tab-pane" id="pane-assistant" style="overflow-y:auto;padding:14px 16px;flex-direction:column;align-items:stretch">

      <!-- Bandeau supérieur compact -->
      <div style="display:flex;align-items:center;gap:14px;background:var(--s1);border:1px solid var(--border);border-radius:6px;padding:10px 16px;margin-bottom:14px;flex-shrink:0">
        <div style="font-size:22px;flex-shrink:0">🎓</div>
        <div style="flex:1;min-width:0">
          <div style="font-family:var(--head);font-size:14px;font-weight:700;letter-spacing:1px;color:#003189">ASSISTANT TERRITORIAL</div>
          <div style="font-family:var(--mono);font-size:9px;color:var(--muted);letter-spacing:0.5px">Détecte les angles morts inter-établissements que personne ne peut voir seul</div>
        </div>
        <div style="font-family:var(--mono);font-size:9px;color:var(--muted);text-transform:uppercase;letter-spacing:1px">Dernière maj</div>
        <div style="font-family:var(--mono);font-size:11px;color:var(--text);background:var(--s2);padding:4px 10px;border-radius:4px;border:1px solid var(--border)">
          <span id="ta-clock">—</span>
        </div>
        <button onclick="refreshTerritorial()" style="background:var(--s2);color:var(--text);border:1px solid var(--border);padding:5px 12px;border-radius:4px;cursor:pointer;font-family:var(--mono);font-size:9px;letter-spacing:1px">↻ ACTUALISER</button>
      </div>

      <!-- 4 KPI en ligne, compacts -->
      <div id="ta-summary" style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:14px;flex-shrink:0">
        <div style="background:var(--s1);border:1px solid var(--border);border-radius:6px;padding:10px 14px;border-left:3px solid var(--blue)">
          <div style="font-family:var(--mono);font-size:8px;color:var(--muted);text-transform:uppercase;letter-spacing:1.5px">Établissements actifs</div>
          <div style="font-family:var(--mono);font-size:22px;font-weight:700;color:var(--text);margin-top:4px" id="ta-online">—</div>
        </div>
        <div style="background:var(--s1);border:1px solid var(--border);border-radius:6px;padding:10px 14px;border-left:3px solid var(--red)">
          <div style="font-family:var(--mono);font-size:8px;color:var(--muted);text-transform:uppercase;letter-spacing:1.5px">Avec incidents critiques</div>
          <div style="font-family:var(--mono);font-size:22px;font-weight:700;color:var(--red);margin-top:4px" id="ta-crit-sites">—</div>
        </div>
        <div style="background:var(--s1);border:1px solid var(--border);border-radius:6px;padding:10px 14px;border-left:3px solid var(--green)">
          <div style="font-family:var(--mono);font-size:8px;color:var(--muted);text-transform:uppercase;letter-spacing:1.5px">Plans Blancs actifs</div>
          <div style="font-family:var(--mono);font-size:22px;font-weight:700;color:var(--text);margin-top:4px" id="ta-pb">—</div>
        </div>
        <div style="background:var(--s1);border:1px solid var(--border);border-radius:6px;padding:10px 14px;border-left:3px solid var(--orange)">
          <div style="font-family:var(--mono);font-size:8px;color:var(--muted);text-transform:uppercase;letter-spacing:1.5px">Alertes territoriales</div>
          <div style="font-family:var(--mono);font-size:22px;font-weight:700;color:var(--orange);margin-top:4px" id="ta-nb-alertes">—</div>
        </div>
      </div>

      <!-- Titre section + alertes -->
      <div style="font-family:var(--mono);font-size:9px;color:var(--muted);letter-spacing:2px;text-transform:uppercase;margin:0 0 10px 4px;display:flex;align-items:center;gap:8px;flex-shrink:0">
        <span style="display:inline-block;width:6px;height:6px;background:var(--orange);border-radius:50%"></span>
        Signaux territoriaux détectés
      </div>
      <div id="ta-alertes-list" style="display:flex;flex-direction:column;gap:8px;padding-bottom:20px">
        <div style="text-align:center;padding:24px;color:var(--muted);font-size:11px;font-family:var(--mono);font-style:italic">Chargement…</div>
      </div>
    </div>

    <!-- CARTOGRAPHIE -->
    <div class="tab-pane" id="pane-carto">
      <div id="map"></div>
    </div>

    <!-- MESSAGERIE INTER-GHT -->
    <div class="tab-pane" id="pane-messagerie">
      <iframe id="msg-iframe" title="Messagerie supervision"
              src="/messagerie/ui?token=PLACEHOLDER_ADMIN_TOKEN"
              style="flex:1;width:100%;border:none;display:block"></iframe>
    </div>

    <div id="pane-admin" style="display:none;flex:1;flex-direction:column;overflow-y:auto;padding:20px;gap:16px;width:100%;height:100%">
      <div style="font-family:var(--mono);font-size:12px;font-weight:700;letter-spacing:1px;color:var(--text)">⚙️ ADMINISTRATION — CONFIGURATION CENTRALE</div>
      <div style="font-size:11px;color:var(--muted);max-width:780px;line-height:1.55">Ces réglages sont <b>tirés par les instances</b> (modèle pull, rafraîchi périodiquement). Chaque instance peut les surcharger localement — précédence <b>local &gt; central &gt; env</b>. Les secrets sont masqués ici et chiffrés au repos.</div>

      <div style="background:var(--s1);border:1px solid var(--border);border-radius:10px;padding:16px;max-width:780px">
        <div style="font-family:var(--mono);font-size:11px;font-weight:700;margin-bottom:12px;color:var(--blue)">🧠 INTELLIGENCE ARTIFICIELLE (agent IA)</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
          <label style="font-size:10px;color:var(--muted)">Fournisseur<select id="adm-ia-provider" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"><option value="albert">Albert (DINUM)</option><option value="openai">OpenAI</option><option value="anthropic">Anthropic</option><option value="mistral">Mistral</option><option value="gemini">Gemini</option><option value="ollama">Ollama</option><option value="openai_compat">OpenAI-compatible</option></select></label>
          <label style="font-size:10px;color:var(--muted)">Clé API<input id="adm-ia-key" type="password" autocomplete="off" placeholder="(laisser vide pour conserver)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">URL base<input id="adm-ia-url" placeholder="https://albert.api.etalab.gouv.fr/v1" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Modèle<input id="adm-ia-model" placeholder="(défaut fournisseur)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
        </div>
        <label style="display:flex;align-items:center;gap:8px;font-size:11px;margin-top:12px;cursor:pointer"><input type="checkbox" id="adm-ia-enabled"> Diffuser cette config IA aux instances</label>
        <div style="display:flex;align-items:center;gap:12px;margin-top:12px"><span id="adm-ia-state" style="font-family:var(--mono);font-size:10px;color:var(--muted)"></span><button onclick="adminSave('ia')" style="margin-left:auto;font-family:var(--mono);font-size:11px;padding:8px 16px;background:var(--blue);color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700">Enregistrer</button></div>
      </div>

      <div style="background:var(--s1);border:1px solid var(--border);border-radius:10px;padding:16px;max-width:780px">
        <div style="font-family:var(--mono);font-size:11px;font-weight:700;margin-bottom:4px;color:var(--blue)">🔐 <span data-i18n="supervision.bluefiles.admin_title">TRANSFERT SÉCURISÉ (BlueFiles)</span></div>
        <div style="font-size:11px;color:var(--muted);margin-bottom:12px" data-i18n="supervision.bluefiles.admin_desc">Identifiants du compte BlueFiles — mêmes champs que la configuration côté instance.</div>
        <div id="adm-bf-mode" style="font-family:var(--mono);font-size:10px;margin-bottom:12px"></div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
          <label style="font-size:10px;color:var(--muted)">
            <span data-i18n="supervision.bluefiles.login">Identifiant (login)</span>
            <input id="adm-bf-login" type="text" autocomplete="off" placeholder="(laisser vide pour conserver)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px">
            <div id="adm-bf-login-state" style="font-family:var(--mono);font-size:9px;color:var(--muted);margin-top:3px"></div>
          </label>
          <label style="font-size:10px;color:var(--muted)">
            <span data-i18n="supervision.bluefiles.password">Mot de passe</span>
            <input id="adm-bf-pwd" type="password" autocomplete="new-password" placeholder="(laisser vide pour conserver)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px">
            <div id="adm-bf-pwd-state" style="font-family:var(--mono);font-size:9px;color:var(--muted);margin-top:3px"></div>
          </label>
          <label style="font-size:10px;color:var(--muted)">
            <span data-i18n="supervision.bluefiles.server">Serveur</span>
            <input id="adm-bf-server" placeholder="api.bluefiles.com" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px">
          </label>
        </div>
        <div style="font-family:var(--mono);font-size:9px;color:var(--muted);line-height:1.5;margin:10px 0 0;padding:8px 10px;background:var(--s2,#f8fafc);border-radius:5px" data-i18n="supervision.bluefiles.cfg_hint">Laissez vide pour conserver la valeur enregistrée. Serveur par défaut : api.bluefiles.com. Les envois s'appuient sur le binaire BlueFiles CLI.</div>
        <label style="display:flex;align-items:center;gap:8px;font-size:11px;margin-top:12px;cursor:pointer"><input type="checkbox" id="adm-bf-enabled"> Diffuser cette config BlueFiles aux instances</label>
        <div style="display:flex;align-items:center;gap:12px;margin-top:12px">
          <span id="adm-bf-state" style="font-family:var(--mono);font-size:10px;color:var(--muted)"></span>
          <button onclick="bfTestColl()" style="font-family:var(--mono);font-size:11px;padding:8px 14px;background:#f5f5fe;color:#000091;border:1px solid #000091;border-radius:5px;cursor:pointer" data-i18n="supervision.bluefiles.test">Tester la connexion</button>
          <button onclick="adminSave('bluefiles')" style="margin-left:auto;font-family:var(--mono);font-size:11px;padding:8px 16px;background:var(--blue);color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700" data-i18n="supervision.bluefiles.save">Enregistrer</button>
        </div>
        <div id="adm-bf-test-result" style="display:none;margin-top:10px;font-size:12px;padding:8px 12px;border-radius:5px"></div>
      </div>

      <div style="background:var(--s1);border:1px solid var(--border);border-radius:10px;padding:16px;max-width:780px">
        <div style="font-family:var(--mono);font-size:11px;font-weight:700;margin-bottom:12px;color:var(--blue)">📧 SMTP (envoi e-mail)</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
          <label style="font-size:10px;color:var(--muted)">Serveur SMTP<input id="adm-smtp-host" placeholder="smtp.exemple.fr" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Port<input id="adm-smtp-port" type="number" placeholder="587" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Utilisateur<input id="adm-smtp-user" autocomplete="off" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Mot de passe<input id="adm-smtp-pass" type="password" autocomplete="off" placeholder="(laisser vide pour conserver)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Adresse expéditeur<input id="adm-smtp-from" placeholder="SCRIBE &lt;alerts@exemple.fr&gt;" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <div style="display:flex;gap:14px;align-items:flex-end;padding-bottom:4px">
            <label style="display:flex;align-items:center;gap:6px;font-size:11px;cursor:pointer"><input type="checkbox" id="adm-smtp-tls"> STARTTLS</label>
            <label style="display:flex;align-items:center;gap:6px;font-size:11px;cursor:pointer"><input type="checkbox" id="adm-smtp-ssl"> SSL direct</label>
          </div>
        </div>
        <label style="display:flex;align-items:center;gap:8px;font-size:11px;margin-top:12px;cursor:pointer"><input type="checkbox" id="adm-smtp-enabled"> Diffuser cette config SMTP aux instances</label>
        <div style="display:flex;align-items:center;gap:12px;margin-top:12px"><span id="adm-smtp-state" style="font-family:var(--mono);font-size:10px;color:var(--muted)"></span><button onclick="adminSave('smtp')" style="margin-left:auto;font-family:var(--mono);font-size:11px;padding:8px 16px;background:var(--blue);color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700">Enregistrer</button></div>
      </div>

      <div style="background:var(--s1);border:1px solid var(--border);border-radius:10px;padding:16px;max-width:780px">
        <div style="font-family:var(--mono);font-size:11px;font-weight:700;margin-bottom:12px;color:var(--blue)">💬 PASSERELLE SMS</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
          <label style="font-size:10px;color:var(--muted)">Fournisseur<select id="adm-sms-provider" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"><option value="ovh">OVH</option><option value="twilio">Twilio</option><option value="free">Free Mobile (test)</option></select></label>
          <label style="font-size:10px;color:var(--muted)">Endpoint<select id="adm-sms-endpoint" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"><option value="ovh-eu">OVH Europe (ovh-eu)</option><option value="ovh-ca">OVH Canada (ovh-ca)</option></select></label>
          <label style="font-size:10px;color:var(--muted)">Application Key<input id="adm-sms-appkey" type="password" autocomplete="off" placeholder="(laisser vide pour conserver)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Application Secret<input id="adm-sms-appsecret" type="password" autocomplete="off" placeholder="(laisser vide pour conserver)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Consumer Key<input id="adm-sms-consumer" type="password" autocomplete="off" placeholder="(laisser vide pour conserver)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Service Name<input id="adm-sms-service" placeholder="sms-xxxx-1" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Émetteur (sender)<input id="adm-sms-sender" placeholder="SCRIBE" maxlength="11" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
        </div>
        <label style="display:flex;align-items:center;gap:8px;font-size:11px;margin-top:12px;cursor:pointer"><input type="checkbox" id="adm-sms-enabled"> Diffuser cette config SMS aux instances</label>
        <div style="font-size:9px;color:var(--muted);margin-top:6px">Champs alignés sur la passerelle SMS OVH des instances : une fois diffusée, la config descend telle quelle et fonctionne sans ressaisie locale.</div>
        <div style="display:flex;align-items:center;gap:12px;margin-top:10px"><span id="adm-sms-state" style="font-family:var(--mono);font-size:10px;color:var(--muted)"></span><button onclick="adminSave('sms')" style="margin-left:auto;font-family:var(--mono);font-size:11px;padding:8px 16px;background:var(--blue);color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700">Enregistrer</button></div>
      </div>

      <div style="background:var(--s1);border:1px solid var(--border);border-radius:10px;padding:16px;max-width:780px;margin-bottom:14px">
        <div style="font-family:var(--mono);font-size:11px;font-weight:700;margin-bottom:10px;color:var(--blue)">📎 FICHIERS &amp; PIÈCES JOINTES</div>
        <div style="font-size:11px;color:var(--muted);margin-bottom:12px">Politique appliquée aux pièces jointes de la messagerie et au drive (fichiers) de toutes les instances qui tirent la config centrale.</div>
        <div style="display:grid;grid-template-columns:1fr 2fr;gap:12px">
          <label style="font-size:10px;color:var(--muted)">Taille max par fichier (Mo) — 0 = défaut<input id="adm-up-max" type="number" min="0" step="1" placeholder="0" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Extensions autorisées (séparées par des virgules) — vide = toutes<input id="adm-up-ext" placeholder="pdf, docx, xlsx, png, jpg" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
        </div>
        <div style="display:flex;align-items:center;gap:12px;margin-top:12px"><span id="adm-up-state" style="font-family:var(--mono);font-size:10px;color:var(--muted)"></span><button onclick="adminSave('uploads')" style="margin-left:auto;font-family:var(--mono);font-size:11px;padding:8px 16px;background:var(--blue);color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700">Enregistrer</button></div>
      </div>

      <div style="background:var(--s1);border:1px solid var(--border);border-radius:10px;padding:16px;max-width:780px;margin-bottom:14px">
        <div style="font-family:var(--mono);font-size:11px;font-weight:700;margin-bottom:10px;color:var(--blue)">☎️ RÉPONDEUR (TWILIO)</div>
        <div style="font-size:11px;color:var(--muted);margin-bottom:12px">Identifiants Twilio pour les lignes d'information de crise. Une fois diffusés, ils descendent vers toutes les instances qui tirent la config centrale (chaque instance peut conserver une config locale prioritaire).</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
          <label style="font-size:10px;color:var(--muted)">Account SID<input id="adm-tw-sid" type="text" autocomplete="off" placeholder="ACxxxxxxxx" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Auth Token<input id="adm-tw-token" type="password" autocomplete="off" placeholder="(laisser vide pour conserver)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">URL publique des instances (webhook)<input id="adm-tw-url" type="text" placeholder="https://mon-serveur.example.net" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Voix par défaut<input id="adm-tw-voice" type="text" placeholder="Polly.Lea-Neural" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
        </div>
        <label style="display:flex;align-items:center;gap:8px;font-size:11px;margin-top:12px;cursor:pointer"><input type="checkbox" id="adm-tw-enabled"> Diffuser cette config Twilio aux instances</label>
        <div style="display:flex;align-items:center;gap:12px;margin-top:12px"><span id="adm-tw-state" style="font-family:var(--mono);font-size:10px;color:var(--muted)"></span><button onclick="adminSave('twilio')" style="margin-left:auto;font-family:var(--mono);font-size:11px;padding:8px 16px;background:var(--blue);color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700">Enregistrer</button></div>
      </div>

      <div style="background:var(--s1);border:1px solid var(--border);border-radius:10px;padding:16px;max-width:780px;margin-bottom:14px">
        <div style="font-family:var(--mono);font-size:11px;font-weight:700;margin-bottom:10px;color:var(--blue)">☎️ RÉPONDEUR (OVH TÉLÉCOM)</div>
        <div style="font-size:11px;color:var(--muted);margin-bottom:12px">Identifiants API OVH (mêmes clés que la passerelle SMS) pour le répondeur via OVH. OVH ne lit pas le texte en direct : le message se règle dans le SVI OVH, mais ces identifiants permettent aux instances de vérifier la joignabilité du compte.</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
          <label style="font-size:10px;color:var(--muted)">Endpoint<select id="adm-ov-endpoint" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"><option value="ovh-eu">Europe (ovh-eu)</option><option value="ovh-ca">Canada (ovh-ca)</option></select></label>
          <label style="font-size:10px;color:var(--muted)">Application Key<input id="adm-ov-key" type="text" autocomplete="off" placeholder="app key" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Application Secret<input id="adm-ov-secret" type="password" autocomplete="off" placeholder="(laisser vide pour conserver)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Consumer Key<input id="adm-ov-consumer" type="password" autocomplete="off" placeholder="(laisser vide pour conserver)" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Compte de facturation<input id="adm-ov-billing" type="text" placeholder="ovhtel-xxxxx-1" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
          <label style="font-size:10px;color:var(--muted)">Numéro / service<input id="adm-ov-service" type="text" placeholder="00339xxxxxxx" style="width:100%;margin-top:3px;padding:7px;background:var(--s2,#fff);border:1px solid var(--border);border-radius:5px;font-size:12px"></label>
        </div>
        <label style="display:flex;align-items:center;gap:8px;font-size:11px;margin-top:12px;cursor:pointer"><input type="checkbox" id="adm-ov-enabled"> Diffuser cette config OVH aux instances</label>
        <div style="display:flex;align-items:center;gap:12px;margin-top:12px"><span id="adm-ov-state" style="font-family:var(--mono);font-size:10px;color:var(--muted)"></span><button onclick="adminSave('ovh_voice')" style="margin-left:auto;font-family:var(--mono);font-size:11px;padding:8px 16px;background:var(--blue);color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700">Enregistrer</button></div>
      </div>

      <div style="background:var(--s1);border:1px solid var(--border);border-radius:10px;padding:16px;max-width:780px">
        <div style="font-family:var(--mono);font-size:11px;font-weight:700;margin-bottom:10px;color:var(--blue)">📡 PROPAGATION</div>
        <div id="adm-prop" style="font-size:11px;color:var(--muted);line-height:1.7">—</div>
        <button onclick="adminSyncConfigs()" id="adm-sync-btn" style="margin-top:14px;font-family:var(--mono);font-size:11px;padding:9px 18px;background:var(--blue);color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700">🔄 Synchroniser les configs vers les instances</button>
        <div style="font-size:9px;color:var(--muted);margin-top:6px">Pousse la config centrale (mail, SMS, BlueFiles) vers toutes les instances découvertes, matérialise les canaux et vérifie leur état opérationnel.</div>
        <div id="adm-sync-result" style="margin-top:12px"></div>
      </div>

      <div style="font-size:10px;color:var(--muted);max-width:780px">La gestion des comptes (avec import Excel) arrive au prochain build.</div>
    </div>

    <!-- STATUTS PUBLICS -->
    <div class="tab-pane" id="pane-statuts">
      <div style="font-family:var(--mono);font-size:9px;color:var(--muted);padding:20px">Chargement…</div>
    </div>

  <!-- Pane Chat -->
  <div id="pane-chat" style="display:none;flex:1;overflow:hidden;min-height:0;position:relative;width:100%;height:100%">
    <iframe id="chat-iframe" src="" style="position:absolute;inset:0;width:100%;height:100%;border:none;display:block"></iframe>
  </div>

  <!-- Pane Comptes -->
  <div id="pane-comptes" style="display:none;flex:1;flex-direction:column;overflow:hidden;min-height:0;width:100%;height:100%">
    <div style="font-family:var(--mono);font-size:9px;color:var(--muted);padding:10px 16px;border-bottom:1px solid var(--border);flex-shrink:0">
      👤 GESTION DES COMPTES SUPERVISION
    </div>
    <div style="padding:20px;display:flex;flex-direction:column;gap:16px;overflow-y:auto;flex:1">
      <div style="background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:16px">
        <div style="font-family:var(--mono);font-size:10px;font-weight:700;margin-bottom:12px">CRÉER UN COMPTE</div>
        <div style="display:flex;flex-direction:column;gap:8px;max-width:360px">
          <input id="new-login" type="text" placeholder="Identifiant" style="font-family:var(--mono);font-size:11px;padding:7px 10px;background:var(--surface);border:1px solid var(--border2);border-radius:5px;color:var(--text)">
          <input id="new-pass" type="password" placeholder="Mot de passe" style="font-family:var(--mono);font-size:11px;padding:7px 10px;background:var(--surface);border:1px solid var(--border2);border-radius:5px;color:var(--text)">
          <select id="new-role" style="font-family:var(--mono);font-size:11px;padding:7px 10px;background:var(--surface);border:1px solid var(--border2);border-radius:5px;color:var(--text)">
            <option value="viewer">Lecteur</option>
            <option value="admin">Admin</option>
          </select>
          <button onclick="createCompte()" style="font-family:var(--mono);font-size:10px;padding:7px 14px;background:#003189;color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700;width:fit-content">Créer</button>
        </div>
      </div>
      <div style="background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:16px">
        <div style="font-family:var(--mono);font-size:10px;font-weight:700;margin-bottom:12px">IMPORT EXCEL</div>
        <div style="font-size:11px;color:var(--muted);margin-bottom:12px;line-height:1.5">Créez plusieurs comptes d'un coup. Colonnes : <b>login</b>, <b>role</b> (admin / viewer), <b>mot_de_passe</b> (laisser vide = mot de passe temporaire généré).</div>
        <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap">
          <button onclick="downloadUserTemplate()" style="font-family:var(--mono);font-size:10px;padding:7px 14px;background:none;border:1px solid var(--border2);border-radius:5px;cursor:pointer;color:var(--text)">⬇ Télécharger le modèle</button>
          <input type="file" id="users-xlsx" accept=".xlsx" style="display:none" onchange="importUsers(this)">
          <button onclick="document.getElementById('users-xlsx').click()" style="font-family:var(--mono);font-size:10px;padding:7px 14px;background:#003189;color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700">⬆ Importer un .xlsx</button>
        </div>
        <div id="import-result" style="margin-top:12px;font-size:11px;line-height:1.6"></div>
      </div>
      <div style="background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:16px">
        <div style="font-family:var(--mono);font-size:10px;font-weight:700;margin-bottom:12px">COMPTES EXISTANTS</div>
        <div id="comptes-list"></div>
      </div>
    </div>
  </div>

  <!-- INSTANCES (master) -->
  <div id="pane-instances" style="display:none;flex:1;flex-direction:column;overflow:hidden;min-height:0;width:100%;height:100%">
    <iframe id="master-iframe" src="about:blank" style="border:0;width:100%;height:100%;flex:1"></iframe>
  </div>

  <!-- ══ ONGLET TRANSFERTS SUPERVISION (h153-DSFR) ══ -->
  <div id="pane-transferts" style="display:none;flex:1;flex-direction:column;overflow:hidden;min-height:0;width:100%;height:100%;background:#f5f5fe">

    <!-- Toolbar -->
    <div style="display:flex;align-items:center;gap:12px;padding:12px 20px;border-bottom:1px solid #e5e7eb;background:#fff;flex-shrink:0;flex-wrap:wrap">
      <span style="font-size:13px;font-weight:700;color:#161616;letter-spacing:.3px">🚑 TRANSFERTS INTER-ÉTABLISSEMENTS</span>
      <button onclick="trSupOpenForm()" style="font-size:12px;font-weight:600;padding:7px 16px;background:#000091;color:#fff;border:none;border-radius:4px;cursor:pointer;display:flex;align-items:center;gap:6px">+ Nouveau transfert</button>
      <select id="tr-sup-filter-statut" onchange="trSupRender()" style="font-size:11px;padding:6px 10px;border:1px solid #d1d5db;border-radius:4px;background:#fff;color:#374151">
        <option value="">Tous statuts</option>
        <option value="EN_PREPARATION">En préparation</option>
        <option value="EN_COURS">En cours</option>
        <option value="ARRIVE">Arrivé</option>
        <option value="ANNULE">Annulé</option>
        <option value="ARCHIVE">Archivé</option>
      </select>
      <select id="tr-sup-filter-etab" onchange="trSupRender()" style="font-size:11px;padding:6px 10px;border:1px solid #d1d5db;border-radius:4px;background:#fff;color:#374151">
        <option value="">Tous établissements</option>
      </select>
      <span id="tr-sup-count" style="font-size:11px;color:#9ca3af;margin-left:auto"></span>
      <button onclick="trSupLoad()" style="font-size:13px;padding:5px 10px;background:#f3f4f6;border:1px solid #d1d5db;border-radius:4px;cursor:pointer;color:#374151" title="Actualiser">↻</button>
    </div>

    <!-- Kanban DSFR 5 colonnes -->
    <div style="display:flex;flex:1;overflow:hidden;gap:0">

      <!-- EN PRÉPARATION -->
      <div style="flex:1;display:flex;flex-direction:column;overflow:hidden;border-right:1px solid #e5e7eb">
        <div style="padding:10px 16px;background:#fff;border-bottom:3px solid #f59e0b;display:flex;align-items:center;gap:8px">
          <span style="font-size:12px;font-weight:700;color:#92400e;text-transform:uppercase;letter-spacing:.5px">⏳ En préparation</span>
          <span id="tr-sup-n-EN_PREPARATION" style="font-size:11px;color:#9ca3af"></span>
        </div>
        <div id="tr-sup-list-EN_PREPARATION" style="flex:1;overflow-y:auto;padding:12px;display:flex;flex-direction:column;gap:8px"></div>
      </div>

      <!-- EN COURS -->
      <div style="flex:1;display:flex;flex-direction:column;overflow:hidden;border-right:1px solid #e5e7eb">
        <div style="padding:10px 16px;background:#fff;border-bottom:3px solid #3b82f6;display:flex;align-items:center;gap:8px">
          <span style="font-size:12px;font-weight:700;color:#1d4ed8;text-transform:uppercase;letter-spacing:.5px">🚑 En cours</span>
          <span id="tr-sup-n-EN_COURS" style="font-size:11px;color:#9ca3af"></span>
        </div>
        <div id="tr-sup-list-EN_COURS" style="flex:1;overflow-y:auto;padding:12px;display:flex;flex-direction:column;gap:8px"></div>
      </div>

      <!-- ARRIVÉ -->
      <div style="flex:1;display:flex;flex-direction:column;overflow:hidden;border-right:1px solid #e5e7eb">
        <div style="padding:10px 16px;background:#fff;border-bottom:3px solid #22c55e;display:flex;align-items:center;gap:8px">
          <span style="font-size:12px;font-weight:700;color:#15803d;text-transform:uppercase;letter-spacing:.5px">✅ Arrivé</span>
          <span id="tr-sup-n-ARRIVE" style="font-size:11px;color:#9ca3af"></span>
        </div>
        <div id="tr-sup-list-ARRIVE" style="flex:1;overflow-y:auto;padding:12px;display:flex;flex-direction:column;gap:8px"></div>
      </div>

      <!-- ANNULÉ -->
      <div style="flex:1;display:flex;flex-direction:column;overflow:hidden;border-right:1px solid #e5e7eb">
        <div style="padding:10px 16px;background:#fff;border-bottom:3px solid #9ca3af;display:flex;align-items:center;gap:8px">
          <span style="font-size:12px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.5px">✕ Annulé</span>
          <span id="tr-sup-n-ANNULE" style="font-size:11px;color:#9ca3af"></span>
        </div>
        <div id="tr-sup-list-ANNULE" style="flex:1;overflow-y:auto;padding:12px;display:flex;flex-direction:column;gap:8px"></div>
      </div>

      <!-- ARCHIVÉ -->
      <div style="flex:1;display:flex;flex-direction:column;overflow:hidden">
        <div style="padding:10px 16px;background:#fff;border-bottom:3px solid #a8a29e;display:flex;align-items:center;gap:8px">
          <span style="font-size:12px;font-weight:700;color:#78716c;text-transform:uppercase;letter-spacing:.5px">📦 Archivé</span>
          <span id="tr-sup-n-ARCHIVE" style="font-size:11px;color:#9ca3af"></span>
        </div>
        <div id="tr-sup-list-ARCHIVE" style="flex:1;overflow-y:auto;padding:12px;display:flex;flex-direction:column;gap:8px"></div>
      </div>

    </div>
  </div>

  <!-- ══ MODAL ARCHIVAGE TRANSFERT SUPERVISION ══ -->
  <div id="tr-sup-archive-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:9100;align-items:center;justify-content:center;padding:20px">
    <div style="background:#fff;border-radius:8px;width:min(440px,98vw);box-shadow:0 4px 24px rgba(0,0,0,.18);overflow:hidden">
      <div style="background:#000091;color:#fff;padding:16px 20px;display:flex;align-items:center;gap:12px">
        <span style="font-size:18px">📦</span>
        <div style="font-size:13px;font-weight:700;letter-spacing:.5px">ARCHIVER LE TRANSFERT</div>
        <button onclick="document.getElementById('tr-sup-archive-modal').style.display='none'" style="margin-left:auto;background:rgba(255,255,255,.15);border:none;color:#fff;width:28px;height:28px;border-radius:50%;cursor:pointer;font-size:16px">✕</button>
      </div>
      <div style="padding:20px;display:flex;flex-direction:column;gap:14px">
        <input type="hidden" id="tr-sup-archive-tid">
        <div style="font-size:12px;color:#6b7280;background:#f9fafb;border:1px solid #e5e7eb;padding:12px;border-radius:6px;border-left:3px solid #e1000f">
          Le transfert sera archivé et retiré de la liste active. Cette action est irréversible.
        </div>
        <div>
          <label style="font-size:11px;font-weight:600;color:#374151;display:block;margin-bottom:6px;text-transform:uppercase;letter-spacing:.3px">Archivé par *</label>
          <input type="text" id="tr-sup-archive-auteur" placeholder="Votre nom ou fonction (ex: Dr. Martin — ARS AURA)" style="width:100%;font-size:12px;padding:8px 12px;border:1px solid #d1d5db;border-radius:4px;box-sizing:border-box;outline:none">
        </div>
        <div id="tr-sup-archive-err" style="display:none;font-size:11px;color:#dc2626;background:#fee2e2;padding:8px 12px;border-radius:4px"></div>
      </div>
      <div style="padding:12px 20px;border-top:1px solid #f3f4f6;display:flex;justify-content:flex-end;gap:8px;background:#f9fafb">
        <button onclick="document.getElementById('tr-sup-archive-modal').style.display='none'" style="font-size:12px;padding:8px 16px;background:#fff;border:1px solid #d1d5db;border-radius:4px;cursor:pointer;color:#374151">Annuler</button>
        <button onclick="trSupConfirmerArchivage()" style="font-size:12px;padding:8px 20px;background:#000091;color:#fff;border:none;border-radius:4px;cursor:pointer;font-weight:600">📦 Confirmer l'archivage</button>
      </div>
    </div>
  </div>

  <!-- ══ MODAL NOUVEAU TRANSFERT SUPERVISION ══ -->
  <div id="tr-sup-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:9000;align-items:flex-start;justify-content:center;padding:32px 16px;overflow-y:auto">
    <div style="background:var(--surface,#fff);border-radius:10px;width:min(580px,98vw);box-shadow:0 8px 32px rgba(0,0,0,.2);overflow:hidden">
      <div style="background:#000091;color:#fff;padding:14px 18px;display:flex;align-items:center;gap:10px">
        <span style="font-size:18px">🚑</span>
        <div>
          <div id="tr-sup-modal-title" style="font-family:monospace;font-size:11px;font-weight:700;letter-spacing:1px">NOUVEAU TRANSFERT</div>
          <div style="font-size:10px;opacity:.8">Supervision territoriale — données anonymisées</div>
        </div>
        <button onclick="trSupCloseForm()" style="margin-left:auto;background:rgba(255,255,255,.2);border:none;color:#fff;width:28px;height:28px;border-radius:50%;cursor:pointer;font-size:16px">✕</button>
      </div>
      <div style="padding:18px;display:flex;flex-direction:column;gap:14px">
        <input type="hidden" id="tr-sup-edit-id">
        <!-- Établissements -->
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
          <div>
            <label style="font-size:11px;font-weight:600;color:#374151;display:block;margin-bottom:6px;text-transform:uppercase;letter-spacing:.3px">Établissement origine *</label>
            <select id="tr-sup-etab-orig" style="width:100%;font-size:12px;padding:8px 10px;background:#fff;border:1px solid #d1d5db;border-radius:4px;box-sizing:border-box">
              <option value="">— Sélectionner —</option>
            </select>
          </div>
          <div>
            <label style="font-size:11px;font-weight:600;color:#374151;display:block;margin-bottom:6px;text-transform:uppercase;letter-spacing:.3px">Établissement destination *</label>
            <select id="tr-sup-etab-dest" style="width:100%;font-size:12px;padding:8px 10px;background:#fff;border:1px solid #d1d5db;border-radius:4px;box-sizing:border-box">
              <option value="">— Sélectionner —</option>
            </select>
          </div>
          <div>
            <label style="font-size:11px;font-weight:600;color:#374151;display:block;margin-bottom:6px;text-transform:uppercase;letter-spacing:.3px">Service / UF origine</label>
            <input type="text" id="tr-sup-unite-orig" placeholder="ex: Réanimation" style="width:100%;font-size:12px;padding:8px 10px;background:#fff;border:1px solid #d1d5db;border-radius:4px;box-sizing:border-box">
          </div>
          <div>
            <label style="font-size:11px;font-weight:600;color:#374151;display:block;margin-bottom:6px;text-transform:uppercase;letter-spacing:.3px">Service / UF destination</label>
            <input type="text" id="tr-sup-unite-dest" placeholder="ex: Médecine interne" style="width:100%;font-size:12px;padding:8px 10px;background:#fff;border:1px solid #d1d5db;border-radius:4px;box-sizing:border-box">
          </div>
        </div>
        <!-- Patient — ISO formulaire instance -->
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;padding:14px">
          <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px">
            <div style="font-family:monospace;font-size:9px;font-weight:700;color:#000091;text-transform:uppercase">Données patient</div>
            <div style="font-family:monospace;font-size:8px;color:#ef4444;background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.2);padding:2px 8px;border-radius:10px">🔒 DONNÉES DE SANTÉ — ACCÈS RESTREINT</div>
          </div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
            <div>
              <label style="font-family:monospace;font-size:9px;color:#666;display:block;margin-bottom:3px;text-transform:uppercase">Nom</label>
              <input type="text" id="tr-sup-nom" placeholder="NOM (majuscules)" style="width:100%;font-family:monospace;font-size:11px;padding:6px 8px;background:#fff;border:1px solid #ddd;border-radius:4px">
            </div>
            <div>
              <label style="font-family:monospace;font-size:9px;color:#666;display:block;margin-bottom:3px;text-transform:uppercase">Prénom</label>
              <input type="text" id="tr-sup-prenom" placeholder="Prénom" style="width:100%;font-family:monospace;font-size:11px;padding:6px 8px;background:#fff;border:1px solid #ddd;border-radius:4px">
            </div>
            <div>
              <label style="font-family:monospace;font-size:9px;color:#666;display:block;margin-bottom:3px;text-transform:uppercase">Nom de naissance</label>
              <input type="text" id="tr-sup-njf" placeholder="Nom jeune fille (si différent)" style="width:100%;font-family:monospace;font-size:11px;padding:6px 8px;background:#fff;border:1px solid #ddd;border-radius:4px">
            </div>
            <div>
              <label style="font-family:monospace;font-size:9px;color:#666;display:block;margin-bottom:3px;text-transform:uppercase">Date de naissance</label>
              <input type="date" id="tr-sup-ddn" style="width:100%;font-family:monospace;font-size:11px;padding:6px 8px;background:#fff;border:1px solid #ddd;border-radius:4px">
            </div>
            <div>
              <label style="font-family:monospace;font-size:9px;color:#666;display:block;margin-bottom:3px;text-transform:uppercase">IPP</label>
              <input type="text" id="tr-sup-ipp" placeholder="Identifiant patient permanent" style="width:100%;font-family:monospace;font-size:11px;padding:6px 8px;background:#fff;border:1px solid #ddd;border-radius:4px">
            </div>
            <div>
              <label style="font-family:monospace;font-size:9px;color:#666;display:block;margin-bottom:3px;text-transform:uppercase">Sexe</label>
              <select id="tr-sup-sexe" style="width:100%;font-family:monospace;font-size:11px;padding:6px 8px;background:#fff;border:1px solid #ddd;border-radius:4px">
                <option value="">—</option>
                <option value="H">Homme</option>
                <option value="F">Femme</option>
                <option value="I">Indifférent</option>
              </select>
            </div>
          </div>
          <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-top:10px">
            <div>
              <label style="font-family:monospace;font-size:9px;color:#666;display:block;margin-bottom:3px;text-transform:uppercase">Âge estimé</label>
              <input type="text" id="tr-sup-age" placeholder="ex: 65 ans" style="width:100%;font-family:monospace;font-size:11px;padding:6px 8px;background:#fff;border:1px solid #ddd;border-radius:4px">
            </div>
            <div>
              <label style="font-family:monospace;font-size:9px;color:#666;display:block;margin-bottom:3px;text-transform:uppercase">Motif / Pathologie</label>
              <input type="text" id="tr-sup-motif" placeholder="ex: Polytraumatisé, défaillance respiratoire..." style="width:100%;font-family:monospace;font-size:11px;padding:6px 8px;background:#fff;border:1px solid #ddd;border-radius:4px">
            </div>
            <div>
              <label style="font-size:11px;font-weight:600;color:#374151;display:block;margin-bottom:6px;text-transform:uppercase;letter-spacing:.3px">ETA — Heure d'arrivée estimée</label>
              <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
                <input type="date" id="tr-sup-eta-date" style="flex:1;min-width:130px;font-size:12px;padding:8px 10px;border:1px solid #d1d5db;border-radius:4px;box-sizing:border-box" oninput="trSupEtaUpdate()">
                <input type="time" id="tr-sup-eta-time" style="flex:0 0 100px;font-size:12px;padding:8px 10px;border:1px solid #d1d5db;border-radius:4px;box-sizing:border-box" oninput="trSupEtaUpdate()">
                <button type="button" onclick="trSupEtaConfirm()" style="font-size:11px;font-weight:700;padding:8px 14px;background:#000091;color:#fff;border:none;border-radius:4px;cursor:pointer;white-space:nowrap;flex-shrink:0">✓ OK</button>
              </div>
              <div id="tr-sup-eta-lbl" style="font-size:11px;color:#000091;font-weight:600;margin-top:6px;min-height:18px;padding:4px 0"></div>
              <input type="hidden" id="tr-sup-eta">
            </div>
          </div>
        </div>
        <div>
          <label style="font-family:monospace;font-size:9px;color:#666;display:block;margin-bottom:3px;text-transform:uppercase">Commentaire</label>
          <textarea id="tr-sup-comment" rows="2" style="width:100%;font-family:monospace;font-size:11px;padding:7px 9px;background:#f8fafc;border:1px solid #ddd;border-radius:5px;resize:vertical;box-sizing:border-box" placeholder="Informations complémentaires..."></textarea>
        </div>
        <div id="tr-sup-err" style="display:none;font-family:monospace;font-size:10px;color:#dc2626;background:#fee;padding:7px 10px;border-radius:5px"></div>
      </div>
      <div style="padding:12px 18px;border-top:1px solid #eee;display:flex;justify-content:flex-end;gap:8px;background:#f8fafc">
        <button onclick="trSupCloseForm()" style="font-family:monospace;font-size:10px;padding:7px 16px;background:#f1f5f9;border:1px solid #ddd;border-radius:5px;cursor:pointer">Annuler</button>
        <button onclick="trSupSave()" style="font-family:monospace;font-size:10px;padding:7px 20px;background:#000091;color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700">💾 Enregistrer</button>
      </div>
    </div>
  </div>

  <div id="pane-exercice" style="display:none;flex:1;flex-direction:column;overflow:hidden;min-height:0;width:100%;height:100%">
    <iframe id="exercice-iframe" src="about:blank" style="border:0;width:100%;height:100%;flex:1"></iframe>
  </div>

  </div><!-- /main -->
</div><!-- /app -->

<!-- KPI Modal (v3000h31) — clic sur compteur barre KPI -->
<div id="kpi-modal-backdrop" onclick="if(event.target===this)closeKpiModal()">
  <div id="kpi-modal">
    <div class="kpi-modal-hdr">
      <div class="kpi-modal-title">
        <span id="kpi-modal-icon">📋</span>
        <span id="kpi-modal-title-text">—</span>
        <span class="kpi-modal-count" id="kpi-modal-count">0</span>
      </div>
      <button class="kpi-modal-close" onclick="closeKpiModal()" title="Fermer (Esc)">✕</button>
    </div>
    <div class="kpi-modal-body" id="kpi-modal-body">
      <div class="kpi-modal-empty">Chargement…</div>
    </div>
  </div>
</div>

<!-- Modal Add Relay -->
<div class="modal-overlay" id="modal-add-relay">
  <div class="modal-box">
    <div class="modal-hdr">
      <span class="modal-hdr-title">⇪ AJOUTER UN COLLECTEUR DE RELAIS</span>
      <button class="modal-close" onclick="closeModal('modal-add-relay')">✕</button>
    </div>
    <div class="modal-body">
      <div style="font-family:var(--mono);font-size:8px;color:var(--muted);line-height:1.6;padding:6px 8px;background:var(--s3);border-radius:4px;border:1px solid var(--border2);margin-bottom:4px">
        Tous les pushs reçus seront automatiquement retransmis vers ce collecteur.<br>
        Le collecteur cible doit avoir enregistré le token que vous choisissez.
      </div>
      <div>
        <label class="form-label">NOM DU COLLECTEUR</label>
        <input id="relay-nom" type="text" class="form-input" placeholder="ex: Supervision ARA, CERT Santé National">
      </div>
      <div>
        <label class="form-label">URL DU COLLECTEUR UPSTREAM</label>
        <input id="relay-url" type="text" class="form-input" placeholder="http://192.168.1.200:9000/api/push">
      </div>
      <div>
        <label class="form-label">TOKEN D'AUTHENTIFICATION</label>
        <input id="relay-token" type="text" class="form-input" placeholder="token_mon_collecteur_regional_2026">
        <div style="font-family:var(--mono);font-size:8px;color:var(--muted);margin-top:3px">Min 16 caractères — doit être enregistré sur le collecteur cible</div>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn-cancel" onclick="closeModal('modal-add-relay')">Annuler</button>
      <button class="btn-primary" onclick="doTestRelay()" style="background:#0d4f8c;margin-right:4px">🔌 Tester</button>
      <button class="btn-primary" onclick="doAddRelay()">⇪ Configurer</button>
    </div>
  </div>
</div>

<!-- Modal Compose -->
<div class="modal-overlay" id="modal-compose">
  <div class="modal-box">
    <div class="modal-hdr">
      <span class="modal-hdr-title">✏ NOUVEAU MESSAGE</span>
      <button class="modal-close" onclick="closeModal('modal-compose')">✕</button>
    </div>
    <div class="modal-body">
      <div>
        <label class="form-label">DESTINATAIRE</label>
        <select id="compose-dest" class="form-select">
          <option value="TOUS">📢 Toutes les instances</option>
        </select>
      </div>
      <div>
        <label class="form-label">OBJET</label>
        <input id="compose-sujet" type="text" class="form-input" placeholder="Objet du message">
      </div>
      <div>
        <label class="form-label">MESSAGE</label>
        <textarea id="compose-body" class="form-textarea" rows="5" placeholder="Rédigez votre message…"></textarea>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn-cancel" onclick="closeModal('modal-compose')">Annuler</button>
      <button class="btn-primary" onclick="sendMsg()">Envoyer ✉</button>
    </div>
  </div>
</div>

<!-- h145 — Modal BlueFiles supervision ─────────────────────────────────────── -->
<div class="modal-overlay" id="modal-bf-send">
  <div class="modal-box" style="max-width:560px;width:96vw">
    <div class="modal-hdr">
      <span class="modal-hdr-title">🔒 <span data-i18n="supervision.bluefiles.modal_title">ENVOI SÉCURISÉ — BLUEFILES</span></span>
      <button class="modal-close" onclick="closeModal('modal-bf-send')">✕</button>
    </div>
    <div class="modal-body" style="gap:14px">

      <!-- Zone de dépôt fichier(s) -->
      <div>
        <div class="form-label" data-i18n="supervision.bluefiles.files">FICHIERS À ENVOYER</div>
        <div id="bf-drop-zone" onclick="document.getElementById('bf-file-input').click()"
          style="border:2px dashed #c7cdd6;border-radius:8px;padding:28px 16px;text-align:center;cursor:pointer;background:#f8fafc;transition:border-color .15s"
          ondragover="event.preventDefault();this.style.borderColor='#000091'"
          ondragleave="this.style.borderColor='#c7cdd6'"
          ondrop="bfDropFiles(event)">
          <div style="font-size:32px;margin-bottom:8px">📁</div>
          <div style="font-size:13px;color:#3a3a3a" data-i18n="supervision.bluefiles.drop_hint">Déposez vos fichiers ici · ou cliquez pour parcourir</div>
          <div style="font-size:11px;color:#929292;margin-top:4px" data-i18n="supervision.bluefiles.drop_limit">Maximum 200 Mo par fichier</div>
        </div>
        <input type="file" id="bf-file-input" multiple style="display:none" onchange="bfAddFiles(this.files)">
        <div id="bf-file-list" style="display:none;margin-top:10px;display:flex;flex-direction:column;gap:6px"></div>
      </div>

      <!-- Destinataires -->
      <div>
        <div class="form-label" data-i18n="supervision.bluefiles.recipients">DESTINATAIRES</div>
        <div id="bf-recipients-list" style="margin-bottom:8px;font-size:12px;color:#929292;font-style:italic" data-i18n="supervision.bluefiles.no_recipient">Aucun destinataire — ajoutez au moins une adresse email.</div>
        <div style="display:flex;gap:8px">
          <input id="bf-email-input" type="email" class="form-input" placeholder="email@destinataire.fr"
            style="flex:1" onkeydown="if(event.key==='Enter'){event.preventDefault();bfAddRecipient();}">
          <button onclick="bfAddRecipient()" class="btn-primary" style="white-space:nowrap" data-i18n="supervision.bluefiles.add_btn">+ Ajouter</button>
        </div>
      </div>

      <!-- Commentaire -->
      <div>
        <div class="form-label" data-i18n="supervision.bluefiles.comment">COMMENTAIRE (optionnel)</div>
        <textarea id="bf-message" class="form-textarea" rows="3"
          placeholder="Contexte de l\'envoi (lu par le destinataire)..."></textarea>
      </div>

      <!-- Bandeau info -->
      <div style="background:#fff8f8;border:1px solid #f0c0c0;border-radius:6px;padding:10px 12px;font-size:12px;color:#5a0000">
        ⚠ <strong data-i18n="supervision.bluefiles.hds_label">Données sensibles</strong> — <span data-i18n="supervision.bluefiles.hds_note">Cet envoi sera tracé dans SCRIBE. Le contenu reste chiffré bout-en-bout (BlueFiles), aucune copie en local.</span>
      </div>

      <div id="bf-send-err" style="display:none;color:#ce0500;font-size:12px"></div>
      <div id="bf-send-progress" style="display:none">
        <div style="font-size:12px;color:#666;margin-bottom:6px" id="bf-send-label" data-i18n="supervision.bluefiles.sending">Envoi en cours…</div>
        <div style="height:6px;background:#eee;border-radius:4px;overflow:hidden"><div id="bf-send-bar" style="height:100%;width:30%;background:#000091;transition:width .3s"></div></div>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn-cancel" onclick="closeModal('modal-bf-send')" data-i18n="supervision.bluefiles.cancel">Annuler</button>
      <button id="bf-send-btn" class="btn-primary" onclick="bfDoSend()" data-i18n="supervision.bluefiles.send_btn">🔒 Envoyer</button>
    </div>
  </div>
</div>

<!-- Modal Accept pending -->
<div class="modal-overlay" id="modal-accept">
  <div class="modal-box">
    <div class="modal-hdr">
      <span class="modal-hdr-title">✓ ACCEPTER L'ÉTABLISSEMENT</span>
      <button class="modal-close" onclick="closeModal('modal-accept')">✕</button>
    </div>
    <div class="modal-body">
      <div id="accept-info" style="font-family:var(--mono);font-size:9px;color:var(--muted);line-height:1.8;padding:8px 10px;background:var(--s3);border-radius:5px;border:1px solid var(--border2)"></div>
      <div>
        <label class="form-label">SIGLE OFFICIEL</label>
        <input id="accept-sigle" type="text" class="form-input" placeholder="ex: CHANGE">
      </div>
      <input id="accept-token" type="hidden">
    </div>
    <div class="modal-footer">
      <button class="btn-cancel" onclick="closeModal('modal-accept')">Annuler</button>
      <button class="btn-primary" onclick="doAccept()">✓ Enrôler</button>
    </div>
  </div>
</div>

<div id="toast"></div>

<script>
const ADMIN_TOKEN = 'PLACEHOLDER_ADMIN_TOKEN';  // remplacé côté serveur
let allData = [];
let repondeurLignes = {};
let pendingList = [];
let selectedSigle = null;
// h142 — map sigle → port (pour ouvrir la page /status publique dans un onglet)
let INSTANCE_PORTS = {};
async function loadInstancePorts() {
  try {
    const r = await fetch('/api/master/instances', {headers:{'Authorization':'Bearer '+ADMIN_TOKEN}});
    if (!r.ok) return;
    const d = await r.json();
    const map = {};
    (d.instances || []).forEach(it => {
      const c = it.config || {};
      if (c.sigle && c.port) map[c.sigle] = c.port;
    });
    INSTANCE_PORTS = map;
  } catch(e) {}
}
function statusUrlFor(sigle) {
  const p = INSTANCE_PORTS[sigle];
  return p ? (location.protocol + '//' + location.hostname + ':' + p + '/status') : null;
}
function openStatusPage(sigle) {
  const u = statusUrlFor(sigle);
  if (u) window.open(u, '_blank', 'noopener');
}
let refreshInterval = 30;
let countdown = 30;
let map = null;
let mapMarkers = {};
let msgMode = 'recu';

// ── Utilities ──────────────────────────────────────────────
function toast(msg, type='ok') {
  const el = document.getElementById('toast');
  el.textContent = msg; el.className = type; el.style.display = 'block';
  clearTimeout(el._t); el._t = setTimeout(() => el.style.display='none', 3000);
}

function fmtAge(iso) {
  if (!iso) return '?';
  const m = Math.round((Date.now() - new Date(iso).getTime()) / 60000);
  if (m < 1) return 'à l\'instant';
  if (m < 60) return `${m}min`;
  return `${Math.floor(m/60)}h${m%60}`;
}

function fmtDate(iso) {
  if (!iso) return '—';
  return new Date(iso).toLocaleString('fr-FR',{day:'2-digit',month:'2-digit',hour:'2-digit',minute:'2-digit'});
}

const LEVEL_COLOR = {NOMINAL:'#00e5a0',VEILLE:'#00e5a0',ALERTE:'#f5c518',CRISE:'#ff7b2c',CRITIQUE:'#ff2d55',INCONNU:'#4a5070'};

function closeModal(id) {
  document.getElementById(id).classList.remove('open');
}

// ── Tab switching ───────────────────────────────────────────
function switchTab(id, btn) {
  // Masquer tous les panes (classe tab-pane ET nos panes custom)
  document.querySelectorAll('.tab-pane').forEach(p => p.classList.remove('active'));
  ['pane-chat','pane-comptes','pane-instances','pane-exercice','pane-admin','pane-transferts'].forEach(pid => {
    var el = document.getElementById(pid);
    if (el) el.style.display = 'none';
  });
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));

  // Panes natifs via classe active
  var pane = document.getElementById('pane-' + id);
  if (pane) {
    if (id === 'chat' || id === 'comptes' || id === 'instances' || id === 'exercice' || id === 'admin' || id === 'transferts') {
      pane.style.display = 'flex';
    } else {
      pane.classList.add('active');
    }
  }
  if (btn) btn.classList.add('active');
  if (id === 'carto') { setTimeout(() => { if (map) { map.invalidateSize(); updateMapMarkers(); } else { initMap(); } window._mapUpdatePending = false; }, 150); }
  if (id === 'messagerie') { /* iframe autonome — pas d'ancien rendu */ }
  if (id === 'statuts') renderStatuts(allData);
  if (id === 'admin') adminLoad();
  if (id === 'transferts') { trSupLoad(); _trSupInitDnD(); }
  if (id === 'chat') {
    var iframe = document.getElementById('chat-iframe');
    if (iframe) {
      var tok = localStorage.getItem('coll_session') || '';
      // Toujours recharger avec le token frais
      iframe.src = 'chat/ui?token=' + encodeURIComponent(tok);
    }
  }
  if (id === 'instances') {
    // Charger l'UI master dans l'iframe avec le token admin pré-stocké
    var iframe = document.getElementById('master-iframe');
    if (iframe) {
      // Stocker le token admin pour que l'UI master le récupère
      try { localStorage.setItem('admin_token', ADMIN_TOKEN); } catch(e){}
      // v2.4.8.3 : on TOUJOURS interroge /onboarding/status, sans condition
      // sur iframe.src. Le bug précédent : si l'iframe était déjà sur
      // /api/master/ui (après quickDemo par ex), cliquer "🎯 Wizard" puis
      // recharger la page ne réinterrogeait plus le statut, donc l'iframe
      // restait sur l'UI normale alors que show_wizard=true.
      // Coût : un appel /status à chaque switch d'onglet (négligeable).
      fetch('/api/master/onboarding/status', {headers: {Authorization: 'Bearer ' + ADMIN_TOKEN}})
        .then(r => r.ok ? r.json() : null)
        .then(s => {
          var target = (s && s.show_wizard) ? '/api/master/onboarding/ui' : '/api/master/ui';
          // Ne reload que si l'URL cible diffère de l'actuelle (évite un
          // flicker sur les switches d'onglet répétés)
          if (!iframe.src || iframe.src === 'about:blank' || iframe.src.indexOf(target) < 0) {
            iframe.src = target;
          }
        })
        .catch(() => {
          if (iframe.src === 'about:blank') iframe.src = '/api/master/ui';
        });
    }
  }
  if (id === 'exercice') {
    // Charger l'UI exercice (mode simulation) dans l'iframe
    var iframe = document.getElementById('exercice-iframe');
    if (iframe) {
      try { localStorage.setItem('admin_token', ADMIN_TOKEN); } catch(e){}
      if (iframe.src === 'about:blank' || iframe.src.indexOf('master/exercice/ui') < 0) {
        iframe.src = '/api/master/exercice/ui';
      }
    }
  }
  if (id === 'comptes') loadComptes();
}

// ── CLOCK ───────────────────────────────────────────────────
setInterval(() => {
  document.getElementById('kpi-clock').textContent = new Date().toLocaleTimeString('fr-FR');
  countdown--;
  if (countdown <= 0) { countdown = refreshInterval; fetchAll(); }
  document.getElementById('refresh-countdown').textContent = countdown + 's';
}, 1000);

// ── FETCH ALL ───────────────────────────────────────────────
let resolvedHideMinutes = parseInt(localStorage.getItem('scribe_hide_min') || '30');
document.addEventListener('DOMContentLoaded', () => {
  const el = document.getElementById('k-hide-delay');
  if (el) el.textContent = resolvedHideMinutes + 'min';

  loadInstancePorts();
  // Auto-redirect vers le wizard d'onboarding si aucune instance configurée.
  // L'utilisateur vient de se connecter à la supervision : si le master est
  // vide, on bascule directement sur l'onglet Instances qui chargera le wizard.
  fetch('/api/master/onboarding/status', {
    headers: {Authorization: 'Bearer ' + ADMIN_TOKEN}
  }).then(r => r.ok ? r.json() : null)
    .then(s => {
      if (s && s.show_wizard) {
        const btn = document.querySelector('[onclick*="switchTab(\'instances\'"]');
        if (btn) {
          // Bascule sur l'onglet Instances (qui chargera l'iframe wizard)
          btn.click();
        }
      }
    })
    .catch(() => {});
});

function promptHideDelay() {
  const v = prompt('Masquer les incidents résolus après combien de minutes ? (0 = jamais)', resolvedHideMinutes);
  if (v === null) return;
  resolvedHideMinutes = Math.max(0, parseInt(v) || 30);
  localStorage.setItem('scribe_hide_min', resolvedHideMinutes);
  const el = document.getElementById('k-hide-delay');
  if (el) el.textContent = resolvedHideMinutes === 0 ? 'jamais' : resolvedHideMinutes + 'min';
  fetchAll();
}


// ══════════════════════════════════════════════════════════════════════════════
//  TRANSFERTS SUPERVISION — JS (h153)
// ══════════════════════════════════════════════════════════════════════════════

var trSupData = [];
var _trSupEnabled = false;
var _trSupInterval = null;

// Vérifier si le module est activé au chargement
(function checkTrSupEnabled() {
  fetch('/api/admin/transferts-supervision/enabled')
    .then(function(r){ return r.json(); })
    .then(function(d){
      _trSupEnabled = d && d.enabled;
      if (_trSupEnabled) {
        var btn = document.getElementById('tab-btn-transferts-sup');
        if (btn) btn.style.display = '';
      }
    }).catch(function(){});
})();

function trSupLoad() {
  if (!_trSupEnabled) return;
  // Charger le label superviseur en parallèle
  fetch('/api/admin/supervisor-label', {headers: {'Authorization': 'Bearer ' + ADMIN_TOKEN}})
    .then(function(r){ return r.ok ? r.json() : {label: 'Supervision'}; })
    .then(function(d){ window._trSupLabel = d.label || 'Supervision'; })
    .catch(function(){});
  fetch('/api/admin/transferts-supervision', {
    headers: {'Authorization': 'Bearer ' + ADMIN_TOKEN}
  }).then(function(r){ return r.ok ? r.json() : []; })
  .then(function(data){
    trSupData = data || [];
    trSupRender();
    trSupPopulateEtabs();
  }).catch(function(){});
}

function trSupPopulateEtabs() {
  var sigles = new Set();
  trSupData.forEach(function(t){
    if (t.etablissement_origine) sigles.add(t.etablissement_origine);
    if (t.etablissement_destination) sigles.add(t.etablissement_destination);
  });
  // Ajouter aussi les établissements connus de la supervision
  if (allData) allData.forEach(function(e){ if (e.sigle) sigles.add(e.sigle); });

  var filterSel = document.getElementById('tr-sup-filter-etab');
  if (filterSel) {
    var cur = filterSel.value;
    filterSel.innerHTML = '<option value="">Tous établissements</option>' +
      Array.from(sigles).sort().map(function(s){
        return '<option value="' + s + '"' + (s === cur ? ' selected' : '') + '>' + s + '</option>';
      }).join('');
  }
  // Peupler les selects du modal
  ['tr-sup-etab-orig','tr-sup-etab-dest'].forEach(function(id){
    var sel = document.getElementById(id);
    if (!sel) return;
    var cur2 = sel.value;
    sel.innerHTML = '<option value="">— Sélectionner —</option>' +
      Array.from(sigles).sort().map(function(s){
        return '<option value="' + s + '"' + (s === cur2 ? ' selected' : '') + '>' + s + '</option>';
      }).join('');
  });
}


// Drag & drop supervision
function _trSupInitDnD() {
  ["EN_PREPARATION","EN_COURS","ARRIVE","ANNULE"].forEach(function(st){
    var col = document.getElementById("tr-sup-list-" + st);
    if (!col || col._dndReady) return;
    col._dndReady = true;
    col.addEventListener("dragover", function(e){ e.preventDefault(); col.style.background="rgba(0,0,145,.04)"; });
    col.addEventListener("dragleave", function(e){ if(!col.contains(e.relatedTarget)) col.style.background=""; });
    col.addEventListener("drop", function(e){
      e.preventDefault(); col.style.background="";
      var id = parseInt(e.dataTransfer.getData("text/plain"));
      if (id && !isNaN(id)) trSupPatchStatut(id, st);
    });
  });
}

function trSupRender() {
  var filterStatut = (document.getElementById("tr-sup-filter-statut")||{}).value||"";
  var filterEtab   = (document.getElementById("tr-sup-filter-etab")||{}).value||"";
  var STATUTS = ["EN_PREPARATION","EN_COURS","ARRIVE","ANNULE","ARCHIVE"];
  var CARD_COLORS = {EN_PREPARATION:"#f59e0b",EN_COURS:"#3b82f6",ARRIVE:"#22c55e",ANNULE:"#9ca3af",ARCHIVE:"#a8a29e"};
  var CARD_TEXT   = {EN_PREPARATION:"#92400e",EN_COURS:"#1d4ed8",ARRIVE:"#15803d",ANNULE:"#6b7280",ARCHIVE:"#78716c"};
  var CARD_BG     = {EN_PREPARATION:"#fffbeb",EN_COURS:"#eff6ff",ARRIVE:"#f0fdf4",ANNULE:"#f9fafb",ARCHIVE:"#fafaf9"};
  var supLabel = window._trSupLabel || "Supervision";
  var filtered = trSupData.filter(function(t){
    if (filterStatut && t.statut !== filterStatut) return false;
    if (filterEtab && t.etablissement_origine !== filterEtab && t.etablissement_destination !== filterEtab) return false;
    return true;
  });
  var NEXT = {
    EN_PREPARATION: [["\u2192 D\u00e9part","EN_COURS"],["\u2715 Annuler","ANNULE"]],
    EN_COURS:       [["\u2705 Arriv\u00e9","ARRIVE"],["\u2715 Annuler","ANNULE"]],
    ARRIVE:[], ANNULE:[], ARCHIVE:[]
  };
  STATUTS.forEach(function(st){
    var col = filtered.filter(function(t){ return t.statut === st; });
    var el = document.getElementById("tr-sup-list-" + st);
    var cnt = document.getElementById("tr-sup-n-" + st);
    if (cnt) cnt.textContent = col.length ? "(" + col.length + ")" : "";
    if (!el) return;
    el.innerHTML = col.map(function(t){
      var borderColor = CARD_COLORS[t.statut] || "#e5e7eb";
      var textColor   = CARD_TEXT[t.statut]   || "#374151";
      var bgColor     = CARD_BG[t.statut]     || "#fff";
      var eta = t.eta ? new Date(t.eta).toLocaleString("fr-FR",{day:"2-digit",month:"2-digit",hour:"2-digit",minute:"2-digit"}) : "";
      var btns = (NEXT[t.statut]||[]).map(function(b){
        return '<button data-tid="'+t.id+'" data-st="'+b[1]+'" onclick="event.stopPropagation();trSupPatchStatut(+this.dataset.tid,this.dataset.st)" style="font-size:11px;padding:5px 12px;background:#fff;border:1px solid #d1d5db;border-radius:4px;cursor:pointer;color:#374151;font-weight:500;transition:background .15s" onmouseover="this.style.background=\'#f3f4f6\'" onmouseout="this.style.background=\'#fff\'">'+b[0]+'</button>';
      }).join("");
      var nomPrenom = [t.nom,t.prenom].filter(Boolean).join(" ");
      return '<div class="tr-sup-card" data-id="'+t.id+'" draggable="true" style="background:'+bgColor+';border:1px solid #e5e7eb;border-left:4px solid '+borderColor+';border-radius:6px;padding:12px 14px;cursor:grab;box-shadow:0 1px 3px rgba(0,0,0,.06);transition:box-shadow .15s" onmouseover="this.style.boxShadow=\'0 4px 12px rgba(0,0,0,.1)\'" onmouseout="this.style.boxShadow=\'0 1px 3px rgba(0,0,0,.06)\'">' +
        // Badge label superviseur
        '<div style="display:inline-flex;align-items:center;gap:4px;font-size:10px;font-weight:600;color:'+textColor+';background:'+borderColor+'22;border:1px solid '+borderColor+'44;padding:2px 8px;border-radius:12px;margin-bottom:8px">&#10022; '+supLabel+'</div>' +
        // En-tête flux
        '<div style="display:flex;align-items:flex-start;justify-content:space-between;gap:8px;margin-bottom:6px">' +
          '<div style="flex:1">' +
            '<div style="font-size:13px;font-weight:700;color:#161616">'+
              (t.etablissement_origine||"?")+" \u2192 "+(t.etablissement_destination||"?")+'</div>' +
            (t.unite_origine||t.unite_destination ? '<div style="font-size:11px;color:#6b7280;margin-top:2px">'+(t.unite_origine||"")+((t.unite_origine&&t.unite_destination)?" \u2192 ":"")+(t.unite_destination||"")+'</div>' : "") +
          '</div>' +
          '<button data-tid="'+t.id+'" onclick="event.stopPropagation();trSupOpenEdit(+this.dataset.tid)" style="background:none;border:none;font-size:14px;cursor:pointer;color:#9ca3af;padding:2px;line-height:1" title="Modifier">&#9999;&#65039;</button>' +
        '<button data-tid="'+t.id+'" onclick="event.stopPropagation();trSupShowDetail(+this.dataset.tid)" style="background:none;border:1px solid #d1d5db;font-size:13px;width:28px;height:28px;border-radius:4px;cursor:pointer;color:#6b7280;flex-shrink:0;display:inline-flex;align-items:center;justify-content:center" title="Détail du transfert">&#8505;</button>' +
        '</div>' +
        // Infos patient
        (nomPrenom ? '<div style="font-size:12px;font-weight:600;color:#374151;margin-bottom:4px">'+nomPrenom+(t.ipp ? ' <span style="font-size:10px;font-weight:400;color:#9ca3af;background:#f3f4f6;padding:1px 6px;border-radius:10px">IPP '+t.ipp+'</span>' : "")+'</div>' : "") +
        // Motif + méta
        (t.motif ? '<div style="font-size:11px;color:#6b7280;margin-bottom:6px">'+t.motif+'</div>' : "") +
        '<div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-top:4px">' +
          (t.age_estime||t.sexe ? '<span style="font-size:10px;background:#f3f4f6;color:#374151;padding:2px 8px;border-radius:10px">'+ [t.age_estime,t.sexe].filter(Boolean).join(" · ") +'</span>' : "") +
          (eta ? '<span style="font-size:10px;color:#6b7280">ETA : '+eta+'</span>' : "") +
          '<span style="font-size:10px;color:#9ca3af;margin-left:auto">'+( t.redacteur||"")+'</span>' +
        '</div>' +
        // Boutons statut
        (btns ? '<div style="display:flex;gap:6px;margin-top:10px;padding-top:10px;border-top:1px solid #f3f4f6">'+btns+'</div>' : "") +
        // Bouton archiver
        (t.statut === "ARRIVE" ? '<div style="margin-top:6px"><button data-tid="'+t.id+'" onclick="event.stopPropagation();trSupOuvrirArchivage(+this.dataset.tid)" style="font-size:11px;padding:5px 12px;background:#fff;border:1px solid #e5e7eb;border-radius:4px;cursor:pointer;color:#78716c;font-weight:500">&#128230; Archiver</button></div>' : "") +
      '</div>';
    }).join("") || '<div style="padding:20px;text-align:center;font-size:12px;color:#9ca3af">Aucun transfert</div>';
    // Drag & drop events
    el.querySelectorAll(".tr-sup-card").forEach(function(card){
      card.addEventListener("dragstart", function(e){ e.dataTransfer.setData("text/plain", card.dataset.id); card.style.opacity=".5"; });
      card.addEventListener("dragend",   function(){ card.style.opacity=""; });
    });
  });
  var cnt = document.getElementById("tr-sup-count");
  if (cnt) cnt.textContent = filtered.length + " transfert(s)";
}


function trSupShowDetail(id) {
  var t = trSupData.find(function(x){ return x.id === id; });
  if (!t) return;
  var SLBL = {EN_PREPARATION:'En préparation',EN_COURS:'En cours',ARRIVE:'Arrivé',ANNULE:'Annulé',ARCHIVE:'Archivé'};
  var SCOL = {EN_PREPARATION:'#d97706',EN_COURS:'#3b82f6',ARRIVE:'#22c55e',ANNULE:'#9ca3af',ARCHIVE:'#78716c'};
  var col = SCOL[t.statut]||'#666'; var lbl = SLBL[t.statut]||t.statut;
  var eta = t.eta ? new Date(t.eta).toLocaleString('fr-FR',{day:'2-digit',month:'2-digit',hour:'2-digit',minute:'2-digit'}) : '—';
  var cre = t.horodatage_creation ? new Date(t.horodatage_creation).toLocaleString('fr-FR',{day:'2-digit',month:'2-digit',hour:'2-digit',minute:'2-digit'}) : '—';
  function row(lbl,val){ if(!val) return ''; return '<div style="display:flex;gap:12px;padding:8px 0;border-bottom:1px solid #f3f4f6"><span style="font-size:11px;font-weight:600;color:#9ca3af;text-transform:uppercase;width:130px;flex-shrink:0">'+lbl+'</span><span style="font-size:12px;color:#161616">'+val+'</span></div>'; }
  var body = row('Étab. origine',t.etablissement_origine) + row('Unité origine',t.unite_origine) +
    row('Étab. destination',t.etablissement_destination) + row('Unité destination',t.unite_destination) +
    '<div style="height:8px"></div>' +
    row('Nom',t.nom) + row('Prénom',t.prenom) + row('Nom de naissance',t.nom_jeune_fille) +
    row('Date de naissance',t.date_naissance) + row('IPP',t.ipp) +
    row('Âge / Sexe', (t.age_estime||'') + (t.age_estime&&t.sexe?' — ':'') + (t.sexe||'')) +
    '<div style="height:8px"></div>' +
    row('Motif / Pathologie',t.motif) + row('ETA',eta) + row('Commentaire',t.commentaire) +
    '<div style="height:8px"></div>' +
    row('Rédacteur',t.redacteur) + row('Créé le',cre);
  var modal = document.getElementById('tr-sup-detail-modal');
  if (!modal) {
    modal = document.createElement('div');
    modal.id = 'tr-sup-detail-modal';
    modal.style.cssText = 'display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9200;align-items:flex-start;justify-content:center;padding:40px 16px;overflow-y:auto';
    modal.addEventListener('click', function(ev){ if(ev.target===modal) modal.style.display='none'; });
    modal.innerHTML = '<div style="background:#fff;border-radius:8px;width:min(520px,98vw);box-shadow:0 8px 32px rgba(0,0,0,.2);overflow:hidden">' +
      '<div id="tr-sup-detail-hdr" style="padding:16px 20px;display:flex;align-items:center;gap:12px;border-bottom:3px solid #000091">' +
        '<div style="flex:1"><div id="tr-sup-detail-titre" style="font-size:14px;font-weight:700;color:#161616"></div>' +
        '<div id="tr-sup-detail-sous" style="font-size:11px;color:#9ca3af;margin-top:2px"></div></div>' +
        '<span id="tr-sup-detail-statut" style="font-size:11px;font-weight:600;padding:3px 12px;border-radius:12px"></span>' +
        '<button onclick="document.getElementById(\x27tr-sup-detail-modal\x27).style.display=\x27none\x27" style="background:none;border:none;font-size:20px;cursor:pointer;color:#9ca3af;margin-left:8px">✕</button>' +
      '</div>' +
      '<div id="tr-sup-detail-body" style="padding:8px 20px 20px"></div>' +
    '</div>';
    document.body.appendChild(modal);
  }
  document.getElementById('tr-sup-detail-hdr').style.borderBottomColor = col;
  document.getElementById('tr-sup-detail-titre').textContent = (t.etablissement_origine||'?')+' → '+(t.etablissement_destination||'?');
  document.getElementById('tr-sup-detail-sous').textContent  = (t.unite_origine||'')+(t.unite_origine&&t.unite_destination?' → ':'')+(t.unite_destination||'');
  document.getElementById('tr-sup-detail-statut').textContent = lbl;
  document.getElementById('tr-sup-detail-statut').style.cssText = 'font-size:11px;font-weight:600;padding:3px 12px;border-radius:12px;border:1px solid '+col+'44;color:'+col+';background:'+col+'14';
  document.getElementById('tr-sup-detail-body').innerHTML = body || '<div style="font-size:12px;color:#9ca3af;text-align:center;padding:16px">Aucune donnée</div>';
  modal.style.display = 'flex';
}


function trSupEtaUpdate() {
  var d = document.getElementById("tr-sup-eta-date").value;
  var t = document.getElementById("tr-sup-eta-time").value || "00:00";
  var lbl = document.getElementById("tr-sup-eta-lbl");
  var hid = document.getElementById("tr-sup-eta");
  if (d) {
    var combined = d + "T" + t;
    if (hid) hid.value = combined;
    var dt = new Date(combined);
    if (lbl) lbl.textContent = "\u2713 ETA : " + dt.toLocaleString("fr-FR",{day:"2-digit",month:"2-digit",hour:"2-digit",minute:"2-digit"});
  } else {
    if (hid) hid.value = '';
    if (lbl) lbl.textContent = '';
  }
}

function trSupEtaConfirm() {
  trSupEtaUpdate();
  var lbl = document.getElementById("tr-sup-eta-lbl");
  var hid = document.getElementById("tr-sup-eta");
  if (hid && hid.value && lbl) { lbl.style.background = "#f0f4ff"; setTimeout(function(){ lbl.style.background = ""; }, 700); }
}

function trSupOpenForm() {
  document.getElementById('tr-sup-edit-id').value = '';
  document.getElementById('tr-sup-modal-title').textContent = 'NOUVEAU TRANSFERT';
  ['tr-sup-etab-orig','tr-sup-etab-dest','tr-sup-unite-orig','tr-sup-unite-dest',
   'tr-sup-nom','tr-sup-prenom','tr-sup-njf','tr-sup-ipp','tr-sup-ddn',
   'tr-sup-age','tr-sup-motif','tr-sup-comment','tr-sup-eta','tr-sup-eta-date','tr-sup-eta-time','tr-sup-eta-lbl'].forEach(function(id){
    var el = document.getElementById(id); if (el) el.value = '';
  });
  document.getElementById('tr-sup-sexe').value = '';
  var etaOk = document.getElementById('tr-sup-eta-confirm'); if(etaOk) etaOk.style.display = 'none';
  document.getElementById('tr-sup-err').style.display = 'none';
  trSupPopulateEtabs();
  document.getElementById('tr-sup-modal').style.display = 'flex';
}

function trSupOpenEdit(id) {
  var t = trSupData.find(function(x){ return x.id === id; });
  if (!t) return;
  document.getElementById('tr-sup-edit-id').value = t.id;
  document.getElementById('tr-sup-modal-title').textContent = 'MODIFIER TRANSFERT #' + t.id;
  trSupPopulateEtabs();
  setTimeout(function(){
    document.getElementById('tr-sup-etab-orig').value  = t.etablissement_origine  || '';
    document.getElementById('tr-sup-etab-dest').value  = t.etablissement_destination || '';
    document.getElementById('tr-sup-unite-orig').value = t.unite_origine  || '';
    document.getElementById('tr-sup-unite-dest').value = t.unite_destination || '';
    document.getElementById('tr-sup-nom').value        = t.nom || '';
    document.getElementById('tr-sup-prenom').value     = t.prenom || '';
    document.getElementById('tr-sup-njf').value        = t.nom_jeune_fille || '';
    document.getElementById('tr-sup-ipp').value        = t.ipp || '';
    document.getElementById('tr-sup-ddn').value        = t.date_naissance || '';
    document.getElementById('tr-sup-age').value        = t.age_estime || '';
    document.getElementById('tr-sup-sexe').value       = t.sexe || '';
    document.getElementById('tr-sup-motif').value      = t.motif || '';
    document.getElementById('tr-sup-comment').value    = t.commentaire || '';
    if (t.eta) {
      var _ep = t.eta.slice(0,16).split('T');
      var _eld = document.getElementById('tr-sup-eta-date'); if(_eld) _eld.value=_ep[0]||'';
      var _elt = document.getElementById('tr-sup-eta-time'); if(_elt) _elt.value=_ep[1]||'';
      document.getElementById('tr-sup-eta').value = t.eta.slice(0,16);
      var _lbl = document.getElementById('tr-sup-eta-lbl');
      if (_lbl) { var _dt=new Date(t.eta); _lbl.textContent='\u2713 ETA : '+_dt.toLocaleString('fr-FR',{day:'2-digit',month:'2-digit',hour:'2-digit',minute:'2-digit'}); }
    } else {
      var _ed=document.getElementById('tr-sup-eta-date'); if(_ed) _ed.value='';
      var _et=document.getElementById('tr-sup-eta-time'); if(_et) _et.value='';
      document.getElementById('tr-sup-eta').value='';
    }
  }, 50);
  document.getElementById('tr-sup-err').style.display = 'none';
  document.getElementById('tr-sup-modal').style.display = 'flex';
}

function trSupCloseForm() {
  document.getElementById('tr-sup-modal').style.display = 'none';
}

async function trSupSave() {
  var tid = document.getElementById('tr-sup-edit-id').value;
  var orig = document.getElementById('tr-sup-etab-orig').value.trim();
  var dest = document.getElementById('tr-sup-etab-dest').value.trim();
  var errEl = document.getElementById('tr-sup-err');
  if (!orig || !dest) {
    errEl.textContent = 'Établissement origine et destination requis.';
    errEl.style.display = 'block'; return;
  }
  errEl.style.display = 'none';
  var payload = {
    etablissement_origine:     orig,
    unite_origine:             document.getElementById('tr-sup-unite-orig').value.trim(),
    etablissement_destination: dest,
    unite_destination:         document.getElementById('tr-sup-unite-dest').value.trim(),
    nom:             (document.getElementById('tr-sup-nom')    ||{value:''}).value.trim() || null,
    prenom:          document.getElementById('tr-sup-prenom').value.trim() || null,
    nom_jeune_fille: document.getElementById('tr-sup-njf').value.trim()    || null,
    ipp:             document.getElementById('tr-sup-ipp').value.trim()    || null,
    date_naissance:  document.getElementById('tr-sup-ddn').value           || null,
    age_estime:      document.getElementById('tr-sup-age').value.trim(),
    sexe:            document.getElementById('tr-sup-sexe').value,
    motif:           document.getElementById('tr-sup-motif').value.trim(),
    commentaire:     document.getElementById('tr-sup-comment').value.trim(),
    eta:             document.getElementById('tr-sup-eta').value || '',
    redacteur:       window._trSupLabel || 'Supervision ARS',
  };
  var url = tid
    ? '/api/admin/transferts-supervision/' + tid
    : '/api/admin/transferts-supervision';
  var method = tid ? 'PUT' : 'POST';
  var r = await fetch(url, {
    method: method,
    headers: {'Authorization': 'Bearer ' + ADMIN_TOKEN, 'Content-Type': 'application/json'},
    body: JSON.stringify(payload)
  });
  if (r.ok) {
    toast('✓ Transfert ' + (tid ? 'mis à jour' : 'créé'), 'ok');
    trSupCloseForm();
    trSupLoad();
  } else {
    var d = await r.json().catch(function(){return {};});
    errEl.textContent = d.detail || 'Erreur serveur';
    errEl.style.display = 'block';
  }
}

async function trSupPatchStatut(id, statut) {
  var r = await fetch('/api/admin/transferts-supervision/' + id + '/statut', {
    method: 'PATCH',
    headers: {'Authorization': 'Bearer ' + ADMIN_TOKEN, 'Content-Type': 'application/json'},
    body: JSON.stringify({statut: statut})
  });
  if (r.ok) { trSupLoad(); } else { toast('Erreur mise à jour statut', 'err'); }
}


function trSupOuvrirArchivage(id) {
  document.getElementById('tr-sup-archive-tid').value = id;
  document.getElementById('tr-sup-archive-auteur').value = '';
  document.getElementById('tr-sup-archive-err').style.display = 'none';
  document.getElementById('tr-sup-archive-modal').style.display = 'flex';
}

async function trSupConfirmerArchivage() {
  var tid = document.getElementById('tr-sup-archive-tid').value;
  var auteur = (document.getElementById('tr-sup-archive-auteur').value || '').trim();
  var errEl = document.getElementById('tr-sup-archive-err');
  if (!auteur) {
    errEl.textContent = "Veuillez renseigner le nom de l'archiviste.";
    errEl.style.display = 'block'; return;
  }
  errEl.style.display = 'none';
  var r = await fetch('/api/admin/transferts-supervision/' + tid + '/archiver', {
    method: 'POST',
    headers: {'Authorization': 'Bearer ' + ADMIN_TOKEN, 'Content-Type': 'application/json'},
    body: JSON.stringify({auteur: auteur})
  });
  if (r.ok) {
    toast('&#128230; Transfert archivé par ' + auteur, 'ok');
    document.getElementById('tr-sup-archive-modal').style.display = 'none';
    // Afficher la colonne ARCHIVE après archivage
    var fs = document.getElementById('tr-sup-filter-statut');
    if (fs) fs.value = 'ARCHIVE';
    trSupLoad();
  } else {
    var d = await r.json().catch(function(){return {};});
    errEl.textContent = d.detail || 'Erreur serveur';
    errEl.style.display = 'block';
  }
}

// Fermeture modal au clic extérieur
document.getElementById('tr-sup-modal').addEventListener('click', function(ev){
  if (ev.target === this) trSupCloseForm();
});

async function fetchAll() {
  // h153 — Recharger trSupData silencieusement pour que la carte supervision
  // puisse afficher les ambulances sans que l'onglet TRANSFERTS soit ouvert
  if (_trSupEnabled && typeof trSupData !== 'undefined') {
    fetch('/api/admin/transferts-supervision', {headers: {'Authorization': 'Bearer ' + ADMIN_TOKEN}})
      .then(function(r){ return r.ok ? r.json() : []; })
      .then(function(d){
        if(d && d.length !== undefined) trSupData = d;
        // Initialiser la carte si pas encore fait
        if (typeof initMap === 'function' && !window._mapInited) { initMap(); window._mapInited = true; }
        updateTransfertLayer();
        updateTransfertLayerOSRM();
      })
      .catch(function(){});
  }
  try {
    const [sumRes, pendRes] = await Promise.all([
      fetch('/api/summary'),
      fetch('/api/admin/pending', {headers:{'Authorization':'Bearer '+ADMIN_TOKEN}})
    ]);
    if (sumRes.ok) allData = await sumRes.json();
    if (pendRes.ok) pendingList = await pendRes.json();
  } catch(e) {}
  try {
    const rl = await fetch('/api/repondeur-lignes');
    if (rl.ok) repondeurLignes = await rl.json();
  } catch(e) {}
  if (Object.keys(INSTANCE_PORTS).length === 0) loadInstancePorts();
  renderKPIs();
  renderEtabList();
  renderPending();
  if (selectedSigle) renderDetail(allData.find(e => e.sigle === selectedSigle));
  if (map) updateMapMarkers(); else { window._mapUpdatePending = true; }
  await fetchMsgBadge();
  loadRelays();
}

// ── KPIs ────────────────────────────────────────────────────
function renderKPIs() {
  const totalInc = allData.reduce((a,e) => a + (e.kpis?.incidents_ouverts||0), 0);
  const totalCrit = allData.reduce((a,e) => a + (e.kpis?.incidents_critiques||0), 0);
  const totalSites = allData.reduce((a,e) => a + (e.sites ? e.sites.length : 1), 0);
  const kEtab = document.getElementById('k-etab');
  kEtab.textContent = totalSites || allData.length;
  kEtab.title = allData.length + ' GHT · ' + totalSites + ' sites';
  const kGht = document.getElementById('k-ght');
  if (kGht) kGht.textContent = allData.length;
  document.getElementById('k-pending').textContent = pendingList.length;
  document.getElementById('k-inc').textContent = totalInc;
  document.getElementById('k-inc').className = 'kpi-val ' + (totalInc > 0 ? 'warn' : 'ok');
  document.getElementById('k-crit').textContent = totalCrit;
  document.getElementById('k-crit').className = 'kpi-val ' + (totalCrit > 0 ? 'crit' : 'ok');
}

// ── PENDING ─────────────────────────────────────────────────
async function acceptAll() {
  const r = await fetch('/api/admin/pending/accept-all', {
    method: 'POST',
    headers: {'Authorization': 'Bearer ' + ADMIN_TOKEN}
  });
  const d = await r.json().catch(()=>({}));
  if (r.ok) {
    toast('✓ ' + d.count + ' établissement(s) accepté(s) : ' + (d.accepted||[]).join(', '), 'ok');
    fetchAll();
  } else toast('Erreur', 'err');
}

function renderPending() {
  const btnAll = document.getElementById('btn-accept-all');
  if (btnAll) btnAll.style.display = pendingList.length > 1 ? 'inline-block' : 'none';
  const sec = document.getElementById('pending-section');
  const list = document.getElementById('pending-list');
  const countEl = document.getElementById('pending-count');
  if (!pendingList.length) { sec.style.display = 'none'; return; }
  sec.style.display = 'block';
  countEl.textContent = pendingList.length;
  list.innerHTML = pendingList.map(p => `
    <div class="pending-card">
      <div class="pending-card-nom">🏥 ${p.nom_propose || p.sigle_propose}</div>
      <div class="pending-card-sub">${p.sigle_propose} · ${p.ip} · ${fmtDate(p.first_seen)}</div>
      <div class="pending-actions">
        <button class="btn-accept" onclick="openAccept('${p.token}','${p.sigle_propose}','${(p.nom_propose||'').replace(/'/g,'')}')">✓ Accepter</button>
        <button class="btn-reject" onclick="doReject('${p.token}')">✗</button>
      </div>
    </div>`).join('');
}

function openAccept(token, sigle, nom) {
  document.getElementById('accept-token').value = token;
  document.getElementById('accept-sigle').value = sigle;
  document.getElementById('accept-info').innerHTML = `<b>Nom proposé :</b> ${nom}<br><b>Sigle proposé :</b> ${sigle}<br><b>Token (12c) :</b> ${token.substring(0,12)}...`;
  document.getElementById('modal-accept').classList.add('open');
}

async function doAccept() {
  const token = document.getElementById('accept-token').value;
  const sigle = document.getElementById('accept-sigle').value.trim().toUpperCase();
  if (!sigle) { toast('Saisissez le sigle', 'err'); return; }
  const prefix = token; // send full token as prefix
  const r = await fetch('/api/admin/pending/'+encodeURIComponent(token)+'/accept', {
    method:'POST', headers:{'Authorization':'Bearer '+ADMIN_TOKEN,'Content-Type':'application/json'},
    body: JSON.stringify({sigle})
  });
  if (r.ok) { toast('✓ '+sigle+' enrôlé', 'ok'); closeModal('modal-accept'); fetchAll(); }
  else { const d=await r.json().catch(()=>{}); toast(d?.detail||'Erreur', 'err'); }
}

async function doReject(token) {
  if (!confirm('Rejeter cet établissement ?')) return;
  const r = await fetch('/api/admin/pending/'+encodeURIComponent(token)+'/reject', {
    method:'POST', headers:{'Authorization':'Bearer '+ADMIN_TOKEN}
  });
  if (r.ok) { toast('Rejeté', 'ok'); fetchAll(); }
}

async function doDisconnect(sigle) {
  if (!confirm('Déconnecter '+sigle+' ? Il ne pourra plus pousser de données.')) return;
  const r = await fetch('/api/admin/tokens/'+sigle, {
    method:'DELETE', headers:{'Authorization':'Bearer '+ADMIN_TOKEN}
  });
  if (r.ok) { toast(sigle+' déconnecté', 'ok'); if (selectedSigle===sigle) selectedSigle=null; fetchAll(); }
  else toast('Erreur', 'err');
}

// ── ETAB LIST ────────────────────────────────────────────────
function renderEtabList() {
  const el = document.getElementById('etab-list');
  if (!allData.length) {
    el.innerHTML = '<div style="padding:20px 12px;font-family:var(--mono);font-size:9px;color:var(--muted);text-align:center">Aucun établissement connecté</div>';
    return;
  }
  el.innerHTML = allData.map(e => {
    const col = LEVEL_COLOR[e.niveau_global] || LEVEL_COLOR.INCONNU;
    const sel = selectedSigle === e.sigle ? 'selected' : '';
    return `<div class="etab-card ${sel}" onclick="selectEtab('${e.sigle}')">
      <div class="etab-dot ${e.niveau_global||'INCONNU'}"></div>
      <div class="etab-info">
        <div class="etab-sigle">${e.sigle}</div>
        <div class="etab-nom">${e.nom||''}</div>
        <div class="etab-age">${fmtAge(e.received_at)}</div>
      </div>
      <button class="btn-disconnect" onclick="event.stopPropagation();doDisconnect('${e.sigle}')">⏻ Déco</button>
    </div>`;
  }).join('');
}

function selectEtab(sigle) {
  selectedSigle = sigle;
  renderEtabList();
  renderDetail(allData.find(e => e.sigle === sigle));
}

// ── DETAIL ──────────────────────────────────────────────────
function renderDetail(e) {
  const el = document.getElementById('detail-panel');
  if (!e) {
    el.innerHTML = '<div class="dp-empty"><div class="dp-empty-icon">⬡</div><span style="font-size:13px">Sélectionnez un établissement</span></div>';
    return;
  }
  const DSFR_LVL = {
    NOMINAL:'#18753c', VEILLE:'#18753c', OPERATIONNEL:'#18753c',
    ALERTE:'#b34000', PERTURBE:'#b34000', MAINTENANCE:'#0063cb',
    CRISE:'#ce0500', INCIDENT_MAJEUR:'#ce0500', CRITIQUE:'#ce0500',
    INCONNU:'#666666'
  };
  const lvlCol = s => DSFR_LVL[s] || '#666666';
  const col = lvlCol(e.niveau_global);
  const kpis = e.kpis || {};
  const sites = e.sites || [];
  const incs  = e.incidents || [];
  const SECLBL = 'font-size:12px;font-weight:700;letter-spacing:.3px;color:#000091;margin-bottom:10px;text-transform:uppercase';

  // h153 — Section TRANSFERTS remplace SITES dans le panneau détail
  const tActifsDetail = e.transferts_actifs || [];
  const tPartis   = tActifsDetail.filter(t => t.statut === 'EN_COURS');
  const tAttente  = tActifsDetail.filter(t => t.statut === 'EN_PREPARATION');
  const tArrivees = tActifsDetail.filter(t => t.statut === 'ARRIVE');

  let sitesHtml = '';
  if (tActifsDetail.length || sites.length) {
    const mkTile = (count, label, color, bg, icon) =>
      `<div style="flex:1;min-width:80px;background:${bg};border:1px solid ${color}33;border-radius:8px;padding:12px 10px;text-align:center">
        <div style="font-size:22px;font-weight:700;color:${color}">${count}</div>
        <div style="font-size:10px;color:${color};margin-top:2px;font-weight:600">${icon} ${label}</div>
      </div>`;
    sitesHtml = `<div style="margin-bottom:20px">
      <div style="${SECLBL}">Transferts</div>
      <div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:${tActifsDetail.length ? '14px' : '0'}">
        ${mkTile(tAttente.length,  'En préparation', '#d97706', '#fffbeb', '⏳')}
        ${mkTile(tPartis.length,   'En route',       '#3b82f6', '#eff6ff', '🚑')}
        ${mkTile(tArrivees.length, 'Arrivés',        '#22c55e', '#f0fdf4', '✅')}
      </div>
      ${tPartis.length ? `<div style="font-size:12px;font-weight:600;color:#374151;margin-bottom:6px">En route</div>
        ${tPartis.map(t => `<div style="font-size:12px;color:#374151;padding:6px 10px;background:#fff;border:1px solid #e5e7eb;border-left:3px solid #3b82f6;border-radius:0 5px 5px 0;margin-bottom:4px">
          🚑 ${t.etablissement_origine||'?'} → ${t.etablissement_destination||'?'}
          ${t.eta ? `<span style="float:right;color:#9ca3af;font-size:11px">ETA ${new Date(t.eta).toLocaleTimeString('fr-FR',{hour:'2-digit',minute:'2-digit'})}</span>` : ''}
        </div>`).join('')}` : ''}
    </div>`;
  }

  let incsHtml = '';
  if (incs.length) {
    const urgColor = {1:'#0063cb',2:'#b34000',3:'#e4794a',4:'#ce0500'};
    const rows = incs.filter(i => {
        if (i.status === 'ARCHIVÉ') return false;
        if (i.status === 'RÉSOLU' && resolvedHideMinutes > 0) {
          const ts = i.timestamp ? new Date(i.timestamp) : null;
          if (ts && (Date.now() - ts.getTime()) > resolvedHideMinutes * 60000) return false;
        }
        return true;
      }).slice(0, 8);
    if (rows.length) incsHtml = `<div style="margin-bottom:20px">
      <div style="${SECLBL}">Incidents en cours</div>
      ${rows.map(i=>`
        <div style="padding:10px 12px;background:#fff;border:1px solid #ddd;border-left:3px solid ${urgColor[i.urgency]||'#929292'};border-radius:0 6px 6px 0;margin-bottom:8px">
          <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">
            <span style="font-size:12px;color:${urgColor[i.urgency]||'#666'};font-weight:700">U${i.urgency||'?'}</span>
            <span style="font-size:12px;color:#666">${i.type_crise||''}</span>
          </div>
          <div style="font-size:14px;color:#161616;line-height:1.4">${(i.fait_resume||'').substring(0,120)}</div>
          ${i.site_id?`<div style="font-size:12px;color:#929292;margin-top:3px">${i.site_id}</div>`:''}
        </div>`).join('')}
    </div>`;
  }

  // h148 — Vue capacité : utiliser _capacite (poussé par le plugin capacité de l'instance)
  const capData = e._capacite || {};
  const renforts = (capData.services_renfort || []);
  // ── Section capacité par pôle ─────────────────────────────────────────────
  // h151 — Vue capacité supervision : grille par site → services individuels
  // Sans déclaration = tout libre (vert). Alertes affichées séparément en haut.
  const capHtml = (() => {
    // Accepter "services" (nouveau) ou "synthese" (compat)
    const services = capData.services || capData.synthese;
    if (!services || !services.length) return '';
    const nbAlerts  = capData.nb_alertes || 0;
    const STAT_COL  = {normal:'#18753c', tension:'#b34000', critique:'#ce0500', ferme:'#666666'};
    const STAT_BG   = {normal:'#eefbf0', tension:'#fff5ed', critique:'#fff0f0', ferme:'#f4f4f4'};
    const STAT_LBL  = {normal:'Normal',  tension:'Tension', critique:'Critique', ferme:'Fermé'};
    const STAT_ICO  = {normal:'🟢',      tension:'🟠',       critique:'🔴',       ferme:'⚫'};

    // Regrouper par site
    const bySite = {};
    services.forEach(svc => {
      const site = svc.site || 'Principal';
      if (!bySite[site]) bySite[site] = [];
      bySite[site].push(svc);
    });
    const multiSite = Object.keys(bySite).length > 1;

    const ts = capData.timestamp
      ? new Date(capData.timestamp).toLocaleString('fr-FR', {day:'2-digit', month:'2-digit', hour:'2-digit', minute:'2-digit'})
      : '—';

    let html = `<div style="margin-bottom:20px">
      <div style="${SECLBL};color:#0063cb;display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:12px">
        <span>🛏 Capacité en lits</span>
        <span style="font-size:10px;font-weight:400;color:#666;margin-left:auto">Mise à jour : ${ts}</span>
        ${nbAlerts ? `<span style="font-family:monospace;font-size:10px;background:#fff5ed;color:#b34000;border:1px solid #f0c080;padding:2px 8px;border-radius:12px;font-weight:700">⚠ ${nbAlerts} alerte(s)</span>` : ''}
      </div>`;

    Object.entries(bySite).forEach(([site, svcs]) => {
      if (multiSite) {
        html += `<div style="font-family:monospace;font-size:9px;font-weight:700;letter-spacing:1.5px;color:#666;text-transform:uppercase;margin:0 0 8px;padding-left:2px">${site}</div>`;
      }
      // Grille responsive
      html += `<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(195px,1fr));gap:10px;margin-bottom:14px">`;
      svcs.forEach(svc => {
        const total    = svc.lits_total || 0;
        const lh       = svc.lits_vides_h || 0;
        const lf       = svc.lits_vides_f || 0;
        const li       = svc.lits_vides_i || 0;
        const libres   = lh + lf + li;
        // Clamp libres pour éviter d'afficher > total
        const libresAff = Math.min(libres, total);
        // h151 — Pas de rouge par défaut : si pas de déclaration, tout est libre
        const hasDecl = svc.has_declaration !== false; // true si déclaration existante
        const pct      = total > 0 ? Math.round((1 - libresAff / total) * 100) : 0;
        const statut   = svc.statut || 'normal';
        const col      = STAT_COL[statut]  || '#18753c';
        const bg       = STAT_BG[statut]   || '#eefbf0';
        // Barre : gris si pas de déclaration (valeur par défaut)
        const barCol   = !hasDecl ? '#c7cdd6'
                       : pct > 90 ? '#ce0500'
                       : pct > 75 ? '#b34000'
                       : '#18753c';
        const ufLabel  = svc.uf_code ? `<span style="font-family:monospace;font-size:9px;color:#929292">${svc.uf_code}</span>` : '';
        const rhLabel  = (svc.statut_rh && svc.statut_rh !== 'complet')
          ? `<span style="font-size:9px;color:#b34000;margin-top:2px">RH: ${svc.statut_rh}</span>`
          : '';
        const ndLabel  = !hasDecl ? `<span style="font-size:9px;font-family:monospace;color:#929292">Non déclaré</span>` : '';
        html += `<div style="background:#fff;border:1px solid #e0e0e0;border-top:3px solid ${col};border-radius:6px;padding:10px 12px;display:flex;flex-direction:column;gap:5px;box-shadow:0 1px 4px rgba(0,0,0,.06)">
          <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:4px">
            <div style="font-size:11px;font-weight:700;color:#161616;line-height:1.3;flex:1">${svc.service}</div>
            ${ufLabel}
          </div>
          <div style="display:flex;align-items:baseline;gap:4px">
            <span style="font-size:20px;font-weight:700;color:${col};line-height:1">${libresAff}</span>
            <span style="font-size:11px;color:#666">/ ${total} libres</span>
          </div>
          <div style="height:4px;background:#eee;border-radius:3px;overflow:hidden">
            <div style="height:100%;width:${pct}%;background:${barCol};border-radius:3px"></div>
          </div>
          <div style="display:flex;justify-content:space-between;font-family:monospace;font-size:9px;color:#929292">
            <span>H:${lh} · F:${lf} · I:${li}</span>
            ${ndLabel || rhLabel || `<span style="font-weight:700;color:${barCol}">${pct}% occ.</span>`}
          </div>
        </div>`;
      });
      html += '</div>';
    });

    if ((capData.services_renfort||[]).length) {
      html += `<div style="padding:8px 12px;background:#fff8f0;border:1px solid #f0d0a0;border-radius:6px;font-size:12px;color:#5a2d00;margin-top:4px">
        ⚙ Mode dégradé RH — ${capData.services_renfort.length} service(s) : ${capData.services_renfort.map(r=>r.service).join(', ')}
      </div>`;
    }
    html += '</div>';
    return html;
  })();
  const tActifs = e.transferts_actifs || [];
  const transfertsHtml = (() => {
    if (!tActifs.length) return '';
    const enCours  = tActifs.filter(function(t){ return t.statut === 'EN_COURS'; });
    const arrives  = tActifs.filter(function(t){ return t.statut === 'ARRIVE'; });
    const attente  = tActifs.filter(function(t){ return t.statut !== 'EN_COURS' && t.statut !== 'ARRIVE'; });

    // Tuiles comptage
    var tiles = '';
    var tileStyle = 'background:#fff;border:1px solid #e0e0e0;border-radius:6px;padding:10px 14px;text-align:center;cursor:pointer;transition:box-shadow .15s;min-width:90px';
    tiles += '<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:10px">';
    if (enCours.length)
      tiles += '<div style="' + tileStyle + ';border-top:3px solid #0063cb;cursor:pointer" data-sigle="' + e.sigle + '" data-filter="EN_COURS" title="Voir les transferts en cours" onclick="showTransfertDetail(this.dataset.sigle,this.dataset.filter)">' +
        '<div style="font-size:22px;font-weight:700;color:#0063cb">' + enCours.length + '</div>' +
        '<div style="font-size:10px;font-family:monospace;color:#666">En cours</div></div>';
    if (arrives.length)
      tiles += '<div style="' + tileStyle + ';border-top:3px solid #18753c;cursor:pointer" data-sigle="' + e.sigle + '" data-filter="ARRIVE" title="Voir les transferts arrivés" onclick="showTransfertDetail(this.dataset.sigle,this.dataset.filter)">' +
        '<div style="font-size:22px;font-weight:700;color:#18753c">' + arrives.length + '</div>' +
        '<div style="font-size:10px;font-family:monospace;color:#666">Arrivés</div></div>';
    if (attente.length)
      tiles += '<div style="' + tileStyle + ';border-top:3px solid #b34000;cursor:pointer" data-sigle="' + e.sigle + '" data-filter="ATTENTE" title="Voir les transferts en attente" onclick="showTransfertDetail(this.dataset.sigle,this.dataset.filter)">' +
        '<div style="font-size:22px;font-weight:700;color:#b34000">' + attente.length + '</div>' +
        '<div style="font-size:10px;font-family:monospace;color:#666">En attente</div></div>';
    tiles += '</div>';

    return '<div style="margin-bottom:20px">' +
      '<div style="' + SECLBL + ';color:#0063cb;margin-bottom:10px">🚑 Transferts (' + tActifs.length + ')</div>' +
      tiles + '</div>';
  })();

  const _sp = (() => {
    const sp = e._status_page;
    if (!sp) return { pill:'', body:'' };
    const faqVisible = (sp.faq||[]).filter(f => f.visible && f.reponse);
    const spSvcs = (sp.services_si||[]).filter(s => s.statut && s.statut.toLowerCase() !== 'ok');
    const spPec  = (sp.prise_en_charge||[]).filter(s => s.statut && s.statut.toLowerCase() !== 'ok');
    const impacted = [...spSvcs, ...spPec];
    if (!faqVisible.length && !impacted.length && !sp.message_public) return { pill:'', body:'' };
    const url = statusUrlFor(e.sigle);
    const openBtn = url
      ? '<button onclick="openStatusPage(\''+e.sigle+'\''+')" style="font-size:11px;font-weight:600;padding:5px 12px;background:#000091;color:#fff;border:none;border-radius:4px;cursor:pointer;white-space:nowrap">Ouvrir la page ↗</button>'
      : '';
    // État synthétique pour le bandeau compact
    var hasHS = impacted.some(function(sv){ return ['hs','hors_service','hors service'].indexOf((sv.statut||'').toLowerCase()) >= 0; });
    var stateLabel = 'Normal', stateColor = '#18753c';
    if (hasHS) { stateLabel = 'Hors service'; stateColor = '#ce0500'; }
    else if (impacted.length) { stateLabel = 'Perturbé'; stateColor = '#b34000'; }
    else if (sp.message_public) { stateLabel = 'Information'; stateColor = '#0063cb'; }
    const pill = '<div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;justify-content:flex-end">' +
      '<span style="font-size:11px;font-weight:700;color:#666;letter-spacing:.5px">📋 STATUT PUBLIC</span>' +
      '<span style="display:inline-flex;align-items:center;gap:5px;padding:4px 11px;border-radius:14px;background:' + stateColor + '14;border:1px solid ' + stateColor + '55;font-size:12px;font-weight:700;color:' + stateColor + '">' +
        '<span style="width:7px;height:7px;border-radius:50%;background:' + stateColor + '"></span>' + stateLabel + '</span>' +
      openBtn + '</div>';
    const msgHtml = sp.message_public
      ? '<div style="font-size:14px;color:#161616;margin-bottom:12px;line-height:1.5">' + sp.message_public + '</div>'
      : '';
    const mbottom = faqVisible.length ? '14px' : '0';
    const impHtml = impacted.length
      ? '<div style="display:flex;flex-direction:column;gap:6px;margin-bottom:' + mbottom + '">' +
          impacted.map(function(sv) {
            var isDeg = ['hs','hors_service','hors service'].indexOf((sv.statut||'').toLowerCase()) >= 0;
            var c = isDeg ? '#ce0500' : '#b34000';
            return '<div style="font-size:14px;color:#161616"><span style="color:' + c + '">●</span> ' +
              (sv.label||sv.id||'') + ' — <span style="color:' + c + ';font-weight:600">' +
              (isDeg ? 'Hors service' : 'Perturbé') + '</span></div>';
          }).join('') + '</div>'
      : '';
    const faqHtml = faqVisible.length
      ? '<div style="border-top:1px solid #ededed;padding-top:12px;display:flex;flex-direction:column;gap:10px">' +
          faqVisible.map(function(f) {
            return '<div style="font-size:14px;line-height:1.4">' +
              '<div style="color:#161616;font-weight:600">' + (f.question||'') + '</div>' +
              '<div style="color:#3a3a3a;margin-top:2px">' + (f.reponse||'') + '</div></div>';
          }).join('') + '</div>'
      : '';
    const body = (msgHtml || impHtml || faqHtml)
      ? '<div style="margin-bottom:20px;padding:14px 16px;background:#fff;border:1px solid #ddd;border-radius:8px">' + msgHtml + impHtml + faqHtml + '</div>'
      : '';
    return { pill: pill, body: body };
  })();
  const spPill = _sp.pill, spDetailBody = _sp.body;

  el.innerHTML = `
    <div style="max-width:760px">
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:22px">
        <div style="width:12px;height:12px;border-radius:50%;background:${col};flex-shrink:0"></div>
        <div>
          <div style="font-size:24px;font-weight:700;color:#161616;letter-spacing:0">${e.nom||e.sigle}</div>
          ${/^Site_\d+$/i.test(e.sigle||'') ? '<div style="font-family:monospace;font-size:10px;color:#b34000;background:#fff5ed;border:1px solid #f0c080;padding:3px 8px;border-radius:4px;margin-top:4px">⚠ Sigle auto-généré — configurez le nom de l\'établissement dans le master</div>' : ''}
          <div style="font-size:12px;color:#666;margin-top:2px">${e.sigle} · Dernière mise à jour : ${fmtDate(e.received_at)}</div>
        </div>
        <div style="margin-left:auto;padding:6px 14px;border-radius:5px;background:${col}14;border:1px solid ${col}55;font-size:12px;font-weight:700;color:${col}">${e.niveau_global||'—'}</div>
      </div>
      ${(function(){
        var _esc=function(s){return String(s==null?'':s).replace(/[&<>"]/g,function(c){return{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c];});};
        var rls=(repondeurLignes&&repondeurLignes[e.sigle])?repondeurLignes[e.sigle]:[];
        if(!rls.length) return '';
        return '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:16px">'+
          rls.map(function(l){return '<span style="font-size:12px;background:#f5f5fe;border:1px solid #cdcdf6;border-radius:14px;padding:4px 12px">\u{1F4DE} <b>'+_esc(l.libelle)+'</b> \u00B7 <span style="font-family:monospace">'+_esc(l.numero)+'</span></span>';}).join('')+
          '</div>';
      })()}
      ${spPill ? '<div style="margin-bottom:10px">' + spPill + '</div>' : ''}
      <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:22px">
        ${[
          ['Incidents ouverts', kpis.incidents_ouverts||0, kpis.incidents_ouverts>0?'#b34000':'#161616'],
          ['Critiques', kpis.incidents_critiques||0, kpis.incidents_critiques>0?'#ce0500':'#161616'],
          ['Cyber', kpis.cyber||0, '#161616'],
          ['Sanitaire', kpis.sanitaire||0, '#161616']
        ].map(([l,v,c])=>`<div style="padding:14px 10px;background:#fff;border:1px solid #ddd;border-radius:8px;text-align:center">
          <div style="font-size:24px;font-weight:700;color:${c}">${v}</div>
          <div style="font-size:11px;color:#666;margin-top:4px">${l}</div>
        </div>`).join('')}
      </div>
      ${spDetailBody}
      ${sitesHtml}
      ${incsHtml}
      ${capHtml}
      ${transfertsHtml}
    </div>`;
}

function showTransfertDetail(sigle, filter) {
  // Trouver l'établissement
  var e = allData.find(function(x){ return x.sigle === sigle; });
  if (!e) return;
  var tAll = e.transferts_actifs || [];
  var filtered;
  if (filter === 'EN_COURS')  filtered = tAll.filter(function(t){ return t.statut === 'EN_COURS'; });
  else if (filter === 'ARRIVE') filtered = tAll.filter(function(t){ return t.statut === 'ARRIVE'; });
  else if (filter === 'ATTENTE') filtered = tAll.filter(function(t){ return t.statut !== 'EN_COURS' && t.statut !== 'ARRIVE'; });
  else filtered = tAll;

  var LABELS = {'EN_COURS':'En cours', 'ARRIVE':'Arrivés', 'ATTENTE':'En attente'};
  var COLS   = {'EN_COURS':'#0063cb', 'ARRIVE':'#18753c', 'ATTENTE':'#b34000'};
  var col    = COLS[filter] || '#0063cb';
  var label  = LABELS[filter] || 'Tous';

  var rows = filtered.map(function(t) {
    var from = t.unite_origine || t.etablissement_origine || '?';
    var to   = t.unite_destination || t.etablissement_destination || '?';
    var eta  = t.eta ? new Date(t.eta).toLocaleString('fr-FR',{day:'2-digit',month:'2-digit',hour:'2-digit',minute:'2-digit'}) : '—';
    var motif = t.motif || t.type_transfert || '';
    return '<div style="padding:10px 14px;background:#fff;border:1px solid #e0e0e0;border-left:3px solid '+col+';border-radius:0 6px 6px 0;margin-bottom:8px">' +
      '<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:4px">' +
        '<span style="font-size:13px;font-weight:700;color:#161616">🚑 ' + from + ' → ' + to + '</span>' +
        '<span style="font-family:monospace;font-size:10px;background:'+col+'14;color:'+col+';padding:2px 8px;border-radius:10px;border:1px solid '+col+'40">'+t.statut+'</span>' +
      '</div>' +
      (motif ? '<div style="font-size:12px;color:#555;margin-bottom:3px">'+motif+'</div>' : '') +
      '<div style="font-family:monospace;font-size:10px;color:#929292">ETA : '+eta+'</div>' +
    '</div>';
  }).join('');

  // Modal générique
  var existing = document.getElementById('modal-transfert-detail');
  if (!existing) {
    existing = document.createElement('div');
    existing.id = 'modal-transfert-detail';
    existing.style.cssText = 'display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9999;align-items:center;justify-content:center;padding:20px';
    existing.addEventListener('click', function(ev){ if(ev.target===this) this.style.display='none'; });
    existing.innerHTML = '<div style="background:#fff;border-radius:10px;width:min(560px,96vw);max-height:80vh;display:flex;flex-direction:column;overflow:hidden;box-shadow:0 8px 32px rgba(0,0,0,.2)">' +
      '<div id="modal-td-hdr" style="padding:14px 18px;border-bottom:1px solid #eee;display:flex;align-items:center;gap:10px;flex-shrink:0">' +
        '<span id="modal-td-title" style="font-family:monospace;font-size:12px;font-weight:700;flex:1"></span>' +
        '<button onclick="var m=document.getElementById(&quot;modal-transfert-detail&quot;);if(m)m.style.display=&quot;none&quot;" style="background:none;border:none;font-size:18px;cursor:pointer;color:#666">✕</button>' +
      '</div>' +
      '<div id="modal-td-body" style="padding:16px;overflow-y:auto;flex:1"></div>' +
    '</div>';
    document.body.appendChild(existing);
  }
  document.getElementById('modal-td-title').textContent = '🚑 ' + sigle + ' — Transferts ' + label + ' (' + filtered.length + ')';
  document.getElementById('modal-td-hdr').style.borderLeft = '4px solid ' + col;
  document.getElementById('modal-td-body').innerHTML = rows || '<div style="font-family:monospace;font-size:11px;color:#666;text-align:center;padding:20px">Aucun transfert</div>';
  existing.style.display = 'flex';
}


// ── CARTE ────────────────────────────────────────────────────
function initMap() {
  if (map) return;
  map = L.map('map', {zoomControl:true}).setView([45.5, 6.2], 8);
  const _osmLayer = L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',
    {attribution:'© OpenStreetMap', maxZoom:19});
  const _cartoLight = L.tileLayer('https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png',
    {attribution:'CartoDB Light', maxZoom:19});
  const _satellite = L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
    {attribution:'Esri World Imagery', maxZoom:19});
  const _baseLayers = {'⬜ CartoDB Clair': _cartoLight, '🗺 OpenStreetMap': _osmLayer, '🛰 Satellite': _satellite};
  _cartoLight.addTo(map);
  L.control.layers(_baseLayers, {}, {position:'topright', collapsed:false}).addTo(map);
  updateMapMarkers();
}

function fitMapToBounds() {
  if (!map || !Object.keys(mapMarkers).length) return;
  const latlngs = Object.values(mapMarkers).map(m => m.getLatLng());
  if (latlngs.length > 0) {
    const bounds = L.latLngBounds(latlngs);
    map.fitBounds(bounds.pad(0.12));
  }
}

let transfertLines = [];
// h154b — parseUTC : les horodatages serveur sont en UTC parfois SANS suffixe
// 'Z' (naïf). new Date() les interpréterait en heure locale → décalage qui
// faisait calculer une progression d'ambulance erronée (≈ mi-trajet) au lieu
// du site de départ. On force l'interprétation UTC, à l'identique de la carte
// SOINS (scribe.js).
function parseUTC(s) {
  if (!s) return null;
  if (typeof s === 'number') return new Date(s);
  if (typeof s !== 'string') return new Date(s);
  if (s.indexOf('Z') >= 0 || s.indexOf('+') >= 0) return new Date(s);
  return new Date(s.indexOf('T') >= 0 ? s + 'Z' : s.replace(' ', 'T') + 'Z');
}

// Progression [0.03..0.95] d'une ambulance le long de son trajet.
// Robuste aux données incomplètes (cause de « bloqué au départ ») :
//  - pas d'heure de départ → tout début ;
//  - ETA absente/incohérente → trajet par défaut de 30 min (l'ambulance avance
//    quand même au lieu de rester figée) ;
//  - ETA atteinte/dépassée → proche destination (et NON retour au départ) ;
//  - léger décalage d'horloge (now < départ) → tout début.
function ambProgress(t) {
  var now = Date.now();
  var depD = parseUTC(t.horodatage_depart || t.created_at || t.horodatage_creation);
  var dep = depD ? depD.getTime() : null;
  if (!dep) return 0.03;
  var etaD = parseUTC(t.eta);
  var eta = etaD ? etaD.getTime() : null;
  if (!eta || eta <= dep) eta = dep + 30 * 60000;  // ETA par défaut : 30 min
  if (now <= dep) return 0.03;
  if (now >= eta) return 0.95;
  return Math.min(0.95, Math.max(0.03, (now - dep) / (eta - dep)));
}

function updateTransfertLayer() {
  // Supprimer les anciennes lignes
  transfertLines.forEach(l => map && l.remove());
  transfertLines = [];
  if (!map) return;
  // Regrouper les transferts EN_COURS par route etab_orig → etab_dest
  const routes = {};
  allData.forEach(etab => {
    (etab.transferts_actifs || []).filter(t => t.statut === 'EN_COURS').forEach(t => {
      const key = [t.etablissement_origine, t.etablissement_destination].sort().join('|||');
      routes[key] = (routes[key] || 0) + 1;
    });
  });
  // Construire un index GPS par sigle
  const gpsIdx = {};
  allData.forEach(etab => {
    if (etab.latitude && etab.longitude)
      gpsIdx[etab.etablissement?.sigle || ''] = [etab.latitude, etab.longitude];
    (etab.sites || []).forEach(s => { if (s.latitude && s.longitude) gpsIdx[s.nom] = [s.latitude, s.longitude]; });
  });
  // Ajouter les transferts supervision (TransfertColl) dans les routes
  if (typeof trSupData !== 'undefined') {
    trSupData.filter(function(t){ return t.statut === 'EN_COURS'; }).forEach(function(t){
      var key = [t.etablissement_origine||'', t.etablissement_destination||''].sort().join('|||');
      routes[key] = (routes[key] || 0) + 1;
      // Garder le transfert complet pour tooltip détaillé
      if (!routes['_tr_'+key]) routes['_tr_'+key] = [];
      routes['_tr_'+key].push(t);
    });
  }

  const ambIcon = L.divIcon({className:'',
    html:'<div style="font-size:22px;filter:drop-shadow(0 0 6px rgba(59,130,246,.9));cursor:pointer">🚑</div>',
    iconSize:[28,28],iconAnchor:[14,14]});

  Object.entries(routes).forEach(([key, count]) => {
    if (key.startsWith('_tr_')) return; // skip metadata
    const [a, b] = key.split('|||');
    const gA = gpsIdx[a], gB = gpsIdx[b];
    if (!gA || !gB) return;

    // Ligne pointillée
    const line = L.polyline([gA, gB], {color:'#3b82f6', weight:2+Math.min(count,3), opacity:0.65, dashArray:'8 5'}).addTo(map);
    const tooltipLine = `🚑 ${count} transfert${count>1?'s':''} EN COURS\n${a} → ${b}`;
    line.bindTooltip(tooltipLine, {sticky:true});
    transfertLines.push(line);

    // Ambulance(s) sur le trajet avec position interpolée selon ETA
    const trsOnRoute = (routes['_tr_'+key] || []).filter(function(t){ return t.statut === 'EN_COURS'; });
    if (trsOnRoute.length > 0) {
      trsOnRoute.forEach(function(t, idx) {
        var _etaD = parseUTC(t.eta); const eta = _etaD ? _etaD.getTime() : null;
        var progress = ambProgress(t);
        // Petite correction si plusieurs ambulances sur le même axe
        const spread = trsOnRoute.length > 1 ? (idx - (trsOnRoute.length-1)/2) * 0.0015 : 0;
        const lat = gA[0] + (gB[0] - gA[0]) * progress + spread;
        const lng = gA[1] + (gB[1] - gA[1]) * progress;

        const etaStr = eta ? new Date(eta).toLocaleTimeString('fr-FR',{hour:'2-digit',minute:'2-digit'}) : '—';
        const motifsStr = t.motif ? ('\n📋 ' + t.motif.substring(0,40)) : '';
        const tooltip = '🚑 ' + (t.etablissement_origine||'?') + ' → ' + (t.etablissement_destination||'?') +
          '\n⏱ ETA : ' + etaStr + motifsStr;

        const m = L.marker([lat, lng], {icon: ambIcon}).addTo(map);
        m.bindTooltip(tooltip, {permanent:false, sticky:true, className:'tr-tooltip'});
        transfertLines.push(m);
      });
    }
  });
}

async function updateTransfertLayerOSRM() {
  if (!map) return;
  // Nettoyer les anciennes lignes OSRM (garder les lignes pointillées de base)
  if (!window._supOsrmLines) window._supOsrmLines = [];
  window._supOsrmLines.forEach(function(l){ map.removeLayer(l); });
  window._supOsrmLines = [];

  // Construire index GPS depuis allData (établissements + sites)
  var gpsIdx = {};
  allData.forEach(function(e) {
    var sgl = (e.sigle||e.etablissement?.sigle||"").toLowerCase();
    if (e.latitude && e.longitude && sgl) gpsIdx[sgl] = [+e.latitude, +e.longitude];
    (e.sites||[]).forEach(function(s) {
      if (s.latitude && s.longitude && s.nom) gpsIdx[s.nom.toLowerCase()] = [+s.latitude, +s.longitude];
    });
  });

  var findCoords = function(n) {
    if (!n) return null;
    var k = n.toLowerCase().trim();
    if (gpsIdx[k]) return gpsIdx[k];
    var found = Object.entries(gpsIdx).find(function(e){ return e[0].length >= 5 && k.startsWith(e[0].substring(0,5)); });
    return found ? found[1] : null;
  };

  var ambIcon = L.divIcon({className:"",
    html:"<div style=\"font-size:22px;filter:drop-shadow(0 0 5px rgba(249,115,22,.9));cursor:pointer\">🚑</div>",
    iconSize:[28,28], iconAnchor:[14,14]});

  // Tous les transferts EN_COURS : instances + supervision
  var allEnCours = [];
  allData.forEach(function(e){ (e.transferts_actifs||[]).filter(function(t){ return t.statut==="EN_COURS"; }).forEach(function(t){ allEnCours.push(t); }); });
  if (typeof trSupData !== "undefined") trSupData.filter(function(t){ return t.statut==="EN_COURS"; }).forEach(function(t){ allEnCours.push(t); });

  for (var i=0; i<allEnCours.length; i++) {
    var t = allEnCours[i];
    var orig = findCoords(t.site_origine || t.etablissement_origine);
    var dest = findCoords(t.site_destination || t.etablissement_destination);
    if (!orig || !dest) continue;

    var _etaD = parseUTC(t.eta); var etaMs = _etaD ? _etaD.getTime() : 0;
    var progress = ambProgress(t);
    var etaStr = etaMs ? new Date(etaMs).toLocaleTimeString("fr-FR",{hour:"2-digit",minute:"2-digit"}) : "—";
    var tooltip = "🚑 <b>" + (t.etablissement_origine||"?") + " → " + (t.etablissement_destination||"?") + "</b>" +
      (t.unite_origine ? "<br>" + t.unite_origine + " → " + (t.unite_destination||"?") : "") +
      "<br>⏱ ETA : " + etaStr +
      (t.motif ? "<br>📋 " + t.motif : "");

    // Tentative OSRM (avec cache géométrique : on ne re-télécharge pas le
    // tracé à chaque repositionnement de l'ambulance — h154d).
    var coords = null;
    var _ckey = orig[0] + "," + orig[1] + "|" + dest[0] + "," + dest[1];
    if (!window._supOsrmCache) window._supOsrmCache = {};
    if (window._supOsrmCache[_ckey]) {
      coords = window._supOsrmCache[_ckey].coords;
      if (window._supOsrmCache[_ckey].dur)
        tooltip += "<br>🗺 Trajet : ~" + window._supOsrmCache[_ckey].dur + " min";
    } else {
      try {
        var osrmR = await fetch(
          "https://router.project-osrm.org/route/v1/driving/" + orig[1] + "," + orig[0] + ";" + dest[1] + "," + dest[0] + "?overview=full&geometries=geojson",
          {signal: AbortSignal.timeout(4000)}
        );
        if (osrmR.ok) {
          var osrmD = await osrmR.json();
          var route = osrmD.routes && osrmD.routes[0];
          if (route) {
            coords = route.geometry.coordinates.map(function(c){ return [c[1],c[0]]; });
            var dur = Math.round(route.duration/60);
            tooltip += "<br>🗺 Trajet : ~" + dur + " min";
            window._supOsrmCache[_ckey] = {coords: coords, dur: dur};
          }
        }
      } catch(e) {}
    }

    var lineCoords = coords && coords.length >= 2 ? coords : [orig, dest];
    var line = L.polyline(lineCoords, {color:"#f97316",weight:3,opacity:.8,dashArray:"8 5"}).addTo(map);
    line.bindTooltip(tooltip, {sticky:true});
    window._supOsrmLines.push(line);

    // Ambulance positionnée sur le trajet
    var posIdx = Math.floor(progress * (lineCoords.length-1));
    var amb = L.marker(lineCoords[posIdx], {icon:ambIcon}).addTo(map);
    amb.bindTooltip(tooltip, {sticky:true});
    window._supOsrmLines.push(amb);

    // Marqueurs départ / arrivée
    var depIco = L.divIcon({className:"",html:"<div style=\"width:8px;height:8px;border-radius:50%;background:#60a5fa;border:2px solid #fff\"></div>",iconSize:[8,8],iconAnchor:[4,4]});
    var arrIco = L.divIcon({className:"",html:"<div style=\"width:8px;height:8px;border-radius:50%;background:#4ade80;border:2px solid #fff\"></div>",iconSize:[8,8],iconAnchor:[4,4]});
    window._supOsrmLines.push(L.marker(orig,{icon:depIco}).bindTooltip("🔵 Départ : "+(t.etablissement_origine||"?")).addTo(map));
    window._supOsrmLines.push(L.marker(dest,{icon:arrIco}).bindTooltip("🟢 Arrivée : "+(t.etablissement_destination||"?")).addTo(map));
  }
}



const _NIV_RANK = {NOMINAL:0, ALERTE:1, CRISE:2, CRITIQUE:3, INCONNU:0};
function _worstNiv(a, b){ return (_NIV_RANK[a]||0) >= (_NIV_RANK[b]||0) ? a : b; }

function updateMapMarkers() {
  Object.values(mapMarkers).forEach(m => map.removeLayer(m));
  mapMarkers = {};
  allData.forEach(e => {
    const col = LEVEL_COLOR[e.niveau_global] || LEVEL_COLOR.INCONNU;
    const sites = e.sites || [];
    const coords = sites.length
      ? sites.map(s=>({lat:s.latitude,lng:s.longitude,nom:s.nom,niv:_worstNiv(s.niveau||'NOMINAL', e.niveau_global||'NOMINAL'),inc:s.incidents_ouverts||0,own:s.incidents_ouverts>0})).filter(s=>s.lat&&s.lng&&!isNaN(+s.lat)&&!isNaN(+s.lng))
      : (e.latitude && e.longitude ? [{lat:e.latitude,lng:e.longitude,nom:e.nom||e.sigle,niv:e.niveau_global}] : []);
    coords.forEach(s => {
      if (!s.lat || !s.lng || isNaN(+s.lat) || isNaN(+s.lng)) return;
      const sc = LEVEL_COLOR[s.niv] || col;
      const icon = L.divIcon({className:'',html:`<div style="width:12px;height:12px;border-radius:50%;background:${sc};border:2px solid ${sc}88;box-shadow:0 0 8px ${sc}66"></div>`,iconSize:[12,12],iconAnchor:[6,6]});
      const m = L.marker([s.lat,s.lng],{icon}).addTo(map);
      m.bindPopup(`<b>${e.sigle}</b><br>${s.nom}<br><span style='color:${sc}'>${s.niv||'NOMINAL'}</span>${s.inc>0?' — '+s.inc+' incident(s)':''}`); 
      mapMarkers[e.sigle+'_'+s.nom] = m;
    });
  });
  fitMapToBounds();
  updateTransfertLayer();
  updateTransfertLayerOSRM();
}

// ── STATUTS PUBLICS ───────────────────────────────────────────
function renderStatuts(data) {
  const el = document.getElementById('pane-statuts');
  // Collecter tous les statuts publiés : global + par site
  const cards = [];
  data.forEach(e => {
    if (e._status_page && e._status_page.published) {
      cards.push({e, sp: e._status_page, isSite: false});
    }
    // Sites publiés séparément
    const sites = e._status_page && e._status_page._statuts_sites ? e._status_page._statuts_sites : [];
    sites.forEach(ss => { if (ss.published) cards.push({e, sp: ss, isSite: true}); });
  });
  if (!cards.length) {
    el.innerHTML = '<div style="padding:48px;text-align:center;font-size:15px;color:var(--muted)">Aucun établissement n\'a publié de point de situation</div>';
    return;
  }
  const niveauCol = {
    operationnel:'#18753c', OPERATIONNEL:'#18753c',
    perturbe:'#b34000',     PERTURBE:'#b34000',
    incident_majeur:'#ce0500', INCIDENT_MAJEUR:'#ce0500',
    maintenance:'#0063cb',  MAINTENANCE:'#0063cb'
  };
  const niveauLabel = {
    operationnel:'✅ Opérationnel', OPERATIONNEL:'✅ Opérationnel',
    perturbe:'⚠️ Perturbé',         PERTURBE:'⚠️ Perturbé',
    incident_majeur:'🔴 Incident majeur', INCIDENT_MAJEUR:'🔴 Incident majeur',
    maintenance:'⚙️ Maintenance',   MAINTENANCE:'⚙️ Maintenance'
  };
  el.innerHTML = '<div class="sp-grid">' + cards.map(({e, sp, isSite}) => {
    const nc  = niveauCol[sp.niveau_global] || '#4a5070';
    const spSvcs = (sp.services_si||[]).filter(s => s.statut && s.statut.toLowerCase() !== 'ok');
    const spPec  = (sp.prise_en_charge||[]).filter(s => s.statut && s.statut.toLowerCase() !== 'ok');
    const msgOff = sp.message_public || sp.message_officiel || '';
    const titre  = e.sigle + (sp.site_nom ? ' — ' + sp.site_nom : '');
    const impacted = [...spSvcs, ...spPec];
    return '<div class="sp-card">' +
      '<div class="sp-card-hdr" style="border-left:4px solid ' + nc + '">' +
        '<div style="flex:1;min-width:0">' +
          '<div style="font-size:16px;font-weight:700;color:var(--text);line-height:1.3">' + titre + '</div>' +
          '<div style="margin-top:6px;font-size:13px;font-weight:600;color:' + nc + '">' + (niveauLabel[sp.niveau_global] || sp.niveau_global || '—') + '</div>' +
        '</div>' +
        '<div style="font-size:12px;color:var(--muted);white-space:nowrap">' + fmtDate(sp.updated_at || e.received_at) + '</div>' +
      '</div>' +
      '<div class="sp-card-body">' +
        (msgOff ? '<div style="font-size:14px;color:var(--text);margin-bottom:12px;line-height:1.55">' + msgOff + '</div>' : '') +
        (impacted.length > 0
          ? '<div style="font-size:11px;font-weight:600;letter-spacing:.5px;text-transform:uppercase;color:var(--muted);margin:4px 0 6px">Services impactés</div>' +
            impacted.map(s => {
              const isDeg = ['hs','hors_service','hors service'].includes((s.statut||'').toLowerCase());
              const col = isDeg ? '#ce0500' : '#b34000';
              return '<div class="sp-svc-row"><div class="svc-dot" style="width:9px;height:9px;background:' + col + '"></div>' +
                (s.label || s.id || '').replace(/_/g,' ') + ' — <span style="color:' + col + ';font-weight:600">' + (isDeg ? 'Hors service' : 'Perturbé') + '</span></div>';
            }).join('')
          : '<div style="display:flex;align-items:center;gap:8px;font-size:14px;color:#18753c;font-weight:500"><span style="width:9px;height:9px;border-radius:50%;background:#18753c;display:inline-block"></span>Tous les services opérationnels</div>') +
      '</div>' +
    '</div>';
  }).join('') + '</div>';
}

// ── MESSAGERIE INTER-GHT ──────────────────────────────────────
let msgData = {received:[], sent:[]};

// ── ADMIN : configuration centrale ──────────────────────────
async function adminLoad() {
  try {
    const r = await fetch('/api/admin/central-config', {headers:{'Authorization':'Bearer '+ADMIN_TOKEN}});
    if (!r.ok) return;
    const d = await r.json(); const c = d.config || {};
    const ia = c.ia || {};
    document.getElementById('adm-ia-provider').value = ia.provider || 'albert';
    document.getElementById('adm-ia-url').value = ia.base_url || '';
    document.getElementById('adm-ia-model').value = ia.model || '';
    document.getElementById('adm-ia-enabled').checked = !!ia.enabled;
    document.getElementById('adm-ia-key').value = '';
    document.getElementById('adm-ia-key').placeholder = ia.has_api_key ? '•••• clé configurée (vide = conserver)' : 'aucune clé';
    document.getElementById('adm-ia-state').textContent = ia.has_api_key ? '🔑 clé configurée' : '— pas de clé';
    const bf = c.bluefiles || {};
    document.getElementById('adm-bf-login').value = '';
    document.getElementById('adm-bf-server').value = bf.server || 'api.bluefiles.com';
    document.getElementById('adm-bf-pwd').value = '';
    const bfMode = document.getElementById('adm-bf-mode');
    if (bfMode) bfMode.innerHTML = bf.cli_ready
      ? '<span style="color:#18753c">● Prêt — login et mot de passe configurés</span>'
      : '<span style="color:#b34000">● Incomplet — renseignez identifiant et mot de passe</span>';
    const bfLgState = document.getElementById('adm-bf-login-state');
    if (bfLgState) bfLgState.textContent = bf.has_login ? '✓ Identifiant configuré : ' + (bf.cli_login_preview || '') : '— Aucun identifiant configuré';
    const bfPwState = document.getElementById('adm-bf-pwd-state');
    if (bfPwState) bfPwState.textContent = bf.has_password ? '✓ Mot de passe configuré : ' + (bf.cli_password_preview || '') : '— Aucun mot de passe configuré';
    document.getElementById('adm-bf-state').textContent = bf.cli_ready ? '🔑 Prêt' : '— Non configuré';
    const bfEn = document.getElementById('adm-bf-enabled'); if (bfEn) bfEn.checked = !!bf.enabled;
    const sm = c.smtp || {};
    document.getElementById('adm-smtp-host').value = sm.smtp_host || '';
    document.getElementById('adm-smtp-port').value = sm.smtp_port || 587;
    document.getElementById('adm-smtp-user').value = sm.smtp_user || '';
    document.getElementById('adm-smtp-from').value = sm.from_addr || '';
    document.getElementById('adm-smtp-tls').checked = sm.use_tls !== false;
    document.getElementById('adm-smtp-ssl').checked = !!sm.use_ssl;
    document.getElementById('adm-smtp-enabled').checked = !!sm.enabled;
    document.getElementById('adm-smtp-pass').value = '';
    document.getElementById('adm-smtp-pass').placeholder = sm.has_smtp_pass ? '•••• mot de passe configuré (vide = conserver)' : '(aucun)';
    document.getElementById('adm-smtp-state').textContent = sm.has_smtp_pass ? '🔑 mot de passe configuré' : '— non configuré';
    const ss = c.sms || {};
    document.getElementById('adm-sms-provider').value = ss.provider || 'ovh';
    document.getElementById('adm-sms-endpoint').value = ss.endpoint || 'ovh-eu';
    document.getElementById('adm-sms-service').value = ss.service_name || '';
    document.getElementById('adm-sms-sender').value = ss.sender || '';
    document.getElementById('adm-sms-enabled').checked = !!ss.enabled;
    document.getElementById('adm-sms-appkey').value = '';
    document.getElementById('adm-sms-appsecret').value = '';
    document.getElementById('adm-sms-consumer').value = '';
    document.getElementById('adm-sms-appkey').placeholder = ss.has_app_key ? '•••• configurée (vide = conserver)' : '(aucune)';
    document.getElementById('adm-sms-appsecret').placeholder = ss.has_app_secret ? '•••• configuré (vide = conserver)' : '(aucun)';
    document.getElementById('adm-sms-consumer').placeholder = ss.has_consumer_key ? '•••• configurée (vide = conserver)' : '(aucune)';
    const up = c.uploads || {};
    const upMax = document.getElementById('adm-up-max');
    if (upMax) upMax.value = (up.max_size_mb && up.max_size_mb > 0) ? up.max_size_mb : '';
    const upExt = document.getElementById('adm-up-ext');
    if (upExt) upExt.value = Array.isArray(up.allowed_extensions) ? up.allowed_extensions.join(', ') : (up.allowed_extensions || '');
    const upState = document.getElementById('adm-up-state');
    if (upState) {
      const nExt = Array.isArray(up.allowed_extensions) ? up.allowed_extensions.length : 0;
      upState.textContent = (up.max_size_mb > 0 || nExt) ? ('● Politique active' + (nExt ? ' — ' + nExt + ' extension(s)' : '')) : '— Aucune restriction (défaut)';
    }
    const tw = c.twilio || {};
    const twSid = document.getElementById('adm-tw-sid');   if (twSid) twSid.value = tw.account_sid || '';
    const twUrl = document.getElementById('adm-tw-url');   if (twUrl) twUrl.value = tw.public_url || '';
    const twVoice = document.getElementById('adm-tw-voice'); if (twVoice) twVoice.value = tw.default_voice || '';
    const twTok = document.getElementById('adm-tw-token'); if (twTok) { twTok.value = ''; twTok.placeholder = tw.has_auth_token ? '•••• configuré (vide = conserver)' : '(aucun)'; }
    const twEn = document.getElementById('adm-tw-enabled'); if (twEn) twEn.checked = !!tw.enabled;
    const twState = document.getElementById('adm-tw-state');
    if (twState) twState.textContent = tw.has_auth_token ? ('● Identifiants Twilio enregistrés' + (tw.enabled ? ' — diffusés' : ' — non diffusés')) : '— Non configuré';
    const ov = c.ovh_voice || {};
    const ovEp = document.getElementById('adm-ov-endpoint'); if (ovEp) ovEp.value = ov.endpoint || 'ovh-eu';
    const ovKey = document.getElementById('adm-ov-key');     if (ovKey) ovKey.value = ov.app_key || '';
    const ovBill = document.getElementById('adm-ov-billing'); if (ovBill) ovBill.value = ov.billing_account || '';
    const ovSvc = document.getElementById('adm-ov-service');  if (ovSvc) ovSvc.value = ov.service || '';
    const ovSec = document.getElementById('adm-ov-secret');   if (ovSec) { ovSec.value=''; ovSec.placeholder = ov.has_app_secret ? '•••• configuré (vide = conserver)' : '(aucun)'; }
    const ovCk = document.getElementById('adm-ov-consumer');  if (ovCk) { ovCk.value=''; ovCk.placeholder = ov.has_consumer_key ? '•••• configuré (vide = conserver)' : '(aucun)'; }
    const ovEn = document.getElementById('adm-ov-enabled');   if (ovEn) ovEn.checked = !!ov.enabled;
    const ovState = document.getElementById('adm-ov-state');
    if (ovState) ovState.textContent = (ov.has_app_secret && ov.has_consumer_key) ? ('● Identifiants OVH enregistrés' + (ov.enabled ? ' — diffusés' : ' — non diffusés')) : '— Non configuré';
    adminRenderProp(d.propagation || {});
  } catch(e) {}
}

function adminRenderProp(p) {
  const el = document.getElementById('adm-prop');
  const keys = Object.keys(p || {});
  if (!keys.length) { el.textContent = "Aucune instance n'a encore tiré la config centrale."; return; }
  el.innerHTML = keys.sort().map(s => '🏥 <b>' + s + '</b> — config tirée le ' + new Date(p[s]).toLocaleString('fr-FR')).join('<br>');
}

async function adminSyncConfigs() {
  const btn = document.getElementById('adm-sync-btn');
  const out = document.getElementById('adm-sync-result');
  if (btn) { btn.disabled = true; btn.textContent = '⏳ Synchronisation en cours…'; }
  if (out) out.innerHTML = '';
  function pill(ok) { return ok ? '<span style="color:#18753c">● OK</span>' : '<span style="color:#b34000">● KO</span>'; }
  try {
    const r = await fetch('/api/admin/sync-configs', {method:'POST', headers:{'Authorization':'Bearer '+ADMIN_TOKEN}});
    const j = await r.json();
    const list = j.instances || [];
    const rows = list.map(function(x) {
      const reach = x.ok ? pill(true)
        : ('<span style="color:#b34000">● injoignable</span>' + (x.error ? (' <span style="color:var(--muted);font-size:9px">(' + x.error + ')</span>') : ''));
      return '<tr style="border-top:1px solid var(--border)">' +
        '<td style="padding:6px 8px"><b>' + x.sigle + '</b><div style="font-size:9px;color:var(--muted)">' + (x.nom || '') + ' · :' + x.port + '</div></td>' +
        '<td style="padding:6px 8px;text-align:center">' + reach + '</td>' +
        '<td style="padding:6px 8px;text-align:center">' + (x.ok ? pill(x.mail) : '—') + '</td>' +
        '<td style="padding:6px 8px;text-align:center">' + (x.ok ? pill(x.sms) : '—') + '</td>' +
        '<td style="padding:6px 8px;text-align:center">' + (x.ok ? pill(x.bluefiles) : '—') + '</td></tr>';
    }).join('');
    out.innerHTML = list.length
      ? ('<table style="width:100%;border-collapse:collapse;font-size:11px;background:var(--s1,#fff);border:1px solid var(--border);border-radius:6px;overflow:hidden">' +
         '<thead><tr style="background:var(--s2,#f1f5f9);font-size:9px;color:var(--muted);text-transform:uppercase">' +
         '<th style="padding:6px 8px;text-align:left">Instance</th><th style="padding:6px 8px">Joignable</th><th style="padding:6px 8px">Mail</th><th style="padding:6px 8px">SMS</th><th style="padding:6px 8px">BlueFiles</th></tr></thead>' +
         '<tbody>' + rows + '</tbody></table>' +
         '<div style="font-size:9px;color:var(--muted);margin-top:6px">' + list.length + ' instance(s) · « OK » = canal/service matérialisé et opérationnel côté instance.</div>')
      : '<div style="font-size:11px;color:var(--muted)">Aucune instance découverte (ports 8000-8009).</div>';
  } catch(e) {
    out.innerHTML = '<div style="font-size:11px;color:#b34000">Erreur de synchronisation : ' + e + '</div>';
  } finally {
    if (btn) { btn.disabled = false; btn.textContent = '🔄 Synchroniser les configs vers les instances'; }
  }
}

async function adminSave(domain) {
  let fields = {};
  if (domain === 'ia') {
    fields = {
      provider: document.getElementById('adm-ia-provider').value,
      base_url: document.getElementById('adm-ia-url').value.trim(),
      model:    document.getElementById('adm-ia-model').value.trim(),
      enabled:  document.getElementById('adm-ia-enabled').checked
    };
    const k = document.getElementById('adm-ia-key').value; if (k) fields.api_key = k;
  } else if (domain === 'bluefiles') {
    fields = {
      login:   document.getElementById('adm-bf-login').value.trim(),
      server:  document.getElementById('adm-bf-server').value.trim() || 'api.bluefiles.com',
      enabled: document.getElementById('adm-bf-enabled').checked
    };
    const bfpwd = document.getElementById('adm-bf-pwd').value; if (bfpwd) fields.password = bfpwd;
  } else if (domain === 'smtp') {
    fields = {
      smtp_host: document.getElementById('adm-smtp-host').value.trim(),
      smtp_port: parseInt(document.getElementById('adm-smtp-port').value) || 587,
      smtp_user: document.getElementById('adm-smtp-user').value.trim(),
      from_addr: document.getElementById('adm-smtp-from').value.trim(),
      use_tls:   document.getElementById('adm-smtp-tls').checked,
      use_ssl:   document.getElementById('adm-smtp-ssl').checked,
      enabled:   document.getElementById('adm-smtp-enabled').checked
    };
    const p = document.getElementById('adm-smtp-pass').value; if (p) fields.smtp_pass = p;
  } else if (domain === 'sms') {
    fields = {
      provider:     document.getElementById('adm-sms-provider').value,
      endpoint:     document.getElementById('adm-sms-endpoint').value,
      service_name: document.getElementById('adm-sms-service').value.trim(),
      sender:       document.getElementById('adm-sms-sender').value.trim(),
      enabled:      document.getElementById('adm-sms-enabled').checked
    };
    const ak = document.getElementById('adm-sms-appkey').value;    if (ak) fields.app_key = ak;
    const sk = document.getElementById('adm-sms-appsecret').value; if (sk) fields.app_secret = sk;
    const ck = document.getElementById('adm-sms-consumer').value;  if (ck) fields.consumer_key = ck;
  } else if (domain === 'uploads') {
    const maxv = parseFloat(document.getElementById('adm-up-max').value) || 0;
    const extRaw = (document.getElementById('adm-up-ext').value || '').trim();
    const exts = extRaw ? extRaw.split(/[\s,]+/).map(function(e){return e.replace(/^\./,'').toLowerCase();}).filter(Boolean) : [];
    fields = { max_size_mb: maxv, allowed_extensions: exts };
  } else if (domain === 'twilio') {
    fields = {
      account_sid:   document.getElementById('adm-tw-sid').value.trim(),
      public_url:    document.getElementById('adm-tw-url').value.trim(),
      default_voice: document.getElementById('adm-tw-voice').value.trim(),
      enabled:       document.getElementById('adm-tw-enabled').checked
    };
    const twtok = document.getElementById('adm-tw-token').value; if (twtok) fields.auth_token = twtok;
  } else if (domain === 'ovh_voice') {
    fields = {
      endpoint:        document.getElementById('adm-ov-endpoint').value,
      app_key:         document.getElementById('adm-ov-key').value.trim(),
      billing_account: document.getElementById('adm-ov-billing').value.trim(),
      service:         document.getElementById('adm-ov-service').value.trim(),
      enabled:         document.getElementById('adm-ov-enabled').checked
    };
    const ovsec = document.getElementById('adm-ov-secret').value;   if (ovsec) fields.app_secret = ovsec;
    const ovck  = document.getElementById('adm-ov-consumer').value; if (ovck)  fields.consumer_key = ovck;
  }
  try {
    const r = await fetch('/api/admin/central-config', {
      method:'POST',
      headers:{'Authorization':'Bearer '+ADMIN_TOKEN,'Content-Type':'application/json'},
      body: JSON.stringify({domain, fields})
    });
    if (r.ok) { toast('Config ' + domain + ' enregistrée ✓', 'ok'); adminLoad(); }
    else toast('Échec enregistrement', 'err');
  } catch(e) { toast('Erreur réseau', 'err'); }
}

// ── h145 — BlueFiles supervision ─────────────────────────────────────────────
let _bfFiles = [];        // fichiers en attente d'envoi
let _bfRecipients = [];   // destinataires (email)
let _bfReady = false;     // BlueFiles disponible côté serveur

async function bfCheckReady() {
  try {
    const r = await fetch('/api/admin/bluefiles/status', {headers:{'Authorization':'Bearer '+ADMIN_TOKEN}});
    if (!r.ok) { _bfReady = false; return; }
    const d = await r.json();
    _bfReady = !!d.ready;
    // Afficher ou masquer le bouton dans la messagerie iframe selon dispo
    // (bouton géré directement dans la messagerie collecteur via window.postMessage)
  } catch(e) { _bfReady = false; }
}

function openBfModal() {
  if (!_bfReady) {
    toast('BlueFiles non configuré — configurez les identifiants dans Admin → BlueFiles', 'err');
    return;
  }
  _bfFiles = []; _bfRecipients = [];
  document.getElementById('bf-send-err').style.display = 'none';
  document.getElementById('bf-send-progress').style.display = 'none';
  const sendBtn = document.getElementById('bf-send-btn');
  if (sendBtn) sendBtn.disabled = false;
  bfRenderFiles(); bfRenderRecipients();
  document.getElementById('bf-message').value = '';
  document.getElementById('bf-email-input').value = '';
  document.getElementById('modal-bf-send').classList.add('open');
  setTimeout(() => { try { document.getElementById('bf-email-input').focus(); } catch(e){} }, 50);
}

function bfAddFiles(fileList) {
  Array.from(fileList || []).forEach(f => { _bfFiles.push(f); });
  bfRenderFiles();
}

function bfDropFiles(event) {
  event.preventDefault();
  document.getElementById('bf-drop-zone').style.borderColor = '#c7cdd6';
  bfAddFiles(event.dataTransfer.files);
}

function bfRemoveFile(idx) {
  _bfFiles.splice(idx, 1);
  bfRenderFiles();
}

function bfRenderFiles() {
  const el = document.getElementById('bf-file-list');
  if (!el) return;
  if (!_bfFiles.length) { el.style.display = 'none'; el.innerHTML = ''; return; }
  el.style.display = 'flex';
  el.innerHTML = _bfFiles.map((f,i) => {
    const sz = f.size > 1048576 ? (f.size/1048576).toFixed(1)+' Mo' : Math.round(f.size/1024)+' Ko';
    return '<div style="display:flex;align-items:center;gap:8px;padding:6px 10px;background:#f5f5fe;border:1px solid #ddd;border-radius:6px;font-size:13px">' +
      '<span style="flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">📄 '+f.name+'</span>' +
      '<span style="color:#929292;font-size:11px">'+sz+'</span>' +
      '<button onclick="bfRemoveFile('+i+')" style="background:none;border:none;cursor:pointer;color:#ce0500;font-size:16px;padding:0 2px">×</button>' +
      '</div>';
  }).join('');
}

function bfAddRecipient() {
  const inp = document.getElementById('bf-email-input');
  const email = (inp ? inp.value.trim() : '');
  if (!email || !/^[^@]+@[^@]+\.[^@]+$/.test(email)) {
    document.getElementById('bf-send-err').textContent = 'Adresse email invalide';
    document.getElementById('bf-send-err').style.display = 'block'; return;
  }
  if (_bfRecipients.includes(email)) { if (inp) inp.value = ''; return; }
  _bfRecipients.push(email);
  if (inp) inp.value = '';
  document.getElementById('bf-send-err').style.display = 'none';
  bfRenderRecipients();
}

function bfRemoveRecipient(idx) {
  _bfRecipients.splice(idx, 1);
  bfRenderRecipients();
}

function bfRenderRecipients() {
  const el = document.getElementById('bf-recipients-list');
  if (!el) return;
  if (!_bfRecipients.length) {
    el.style.fontStyle = 'italic'; el.style.color = '#929292';
    el.setAttribute('data-i18n','supervision.bluefiles.no_recipient');
    // Forcer le texte via la clé i18n active
    el.textContent = 'Aucun destinataire — ajoutez au moins une adresse email.';
    return;
  }
  el.style.fontStyle = 'normal'; el.style.color = '#161616';
  el.innerHTML = _bfRecipients.map((e,i) =>
    '<span style="display:inline-flex;align-items:center;gap:4px;background:#f5f5fe;border:1px solid #000091;border-radius:20px;padding:3px 10px;font-size:12px;margin:2px">' +
    e + '<button onclick="bfRemoveRecipient('+i+')" style="background:none;border:none;cursor:pointer;color:#929292;font-size:14px;padding:0 2px">×</button></span>'
  ).join('');
}

async function bfDoSend() {
  const errEl = document.getElementById('bf-send-err');
  errEl.style.display = 'none';
  if (!_bfFiles.length) { errEl.textContent = 'Ajoutez au moins un fichier.'; errEl.style.display='block'; return; }
  if (!_bfRecipients.length) { errEl.textContent = 'Ajoutez au moins un destinataire.'; errEl.style.display='block'; return; }
  const btn = document.getElementById('bf-send-btn');
  if (btn) btn.disabled = true;
  document.getElementById('bf-send-progress').style.display = 'block';
  document.getElementById('bf-send-bar').style.width = '30%';
  const fd = new FormData();
  _bfFiles.forEach(f => fd.append('file', f));
  fd.append('recipients', JSON.stringify(_bfRecipients.map(e => ({email:e, acknowledge:false}))));
  fd.append('object', 'Document sécurisé — SCRIBE Supervision');
  fd.append('message', document.getElementById('bf-message').value || '');
  try {
    const r = await fetch('/api/admin/bluefiles/send', {
      method: 'POST',
      headers: {'Authorization': 'Bearer '+ADMIN_TOKEN},
      body: fd
    });
    document.getElementById('bf-send-bar').style.width = '90%';
    if (!r.ok) {
      const d = await r.json().catch(()=>({}));
      errEl.textContent = d.detail || 'Erreur envoi BlueFiles';
      errEl.style.display = 'block';
      if (btn) btn.disabled = false;
      return;
    }
    document.getElementById('bf-send-bar').style.width = '100%';
    const d = await r.json();
    toast('✓ Envoi BlueFiles réussi → '+d.recipients.join(', '), 'ok');
    setTimeout(() => {
      closeModal('modal-bf-send');
      document.getElementById('bf-send-progress').style.display = 'none';
      document.getElementById('bf-send-bar').style.width = '30%';
    }, 1200);
  } catch(e) {
    errEl.textContent = 'Erreur réseau : '+e;
    errEl.style.display = 'block';
    if (btn) btn.disabled = false;
  }
}

// Vérifier la dispo BlueFiles au chargement
bfCheckReady();
// Ajouter le bouton BlueFiles dans la barre messagerie si disponible
(function() {
  const toolbar = document.getElementById('msg-toolbar');
  if (!toolbar) return;
  const btn = document.createElement('button');
  btn.id = 'btn-bf-send-coll';
  btn.textContent = '🔒 Envoi sécurisé BlueFiles';
  btn.style.cssText = 'display:none;font-family:var(--mono);font-size:8px;padding:4px 12px;background:#000091;color:#fff;border:none;border-radius:4px;cursor:pointer;font-weight:700;margin-left:6px';
  btn.onclick = function() { openBfModal(); };
  toolbar.appendChild(btn);
  // Révéler après vérification dispo
  setTimeout(async function() {
    await bfCheckReady();
    if (_bfReady) btn.style.display = 'inline-block';
  }, 800);
})();

async function fetchMsgBadge() {
  try {
    const r = await fetch('/api/v1/messagerie/non-lus', {headers:{'Authorization':'Bearer '+ADMIN_TOKEN}});
    if (!r.ok) return;
    const d = await r.json();
    const count = d.count || 0;
    const b = document.getElementById('msg-badge-hdr');
    if (b) { if (count > 0) { b.textContent = count; b.style.display = 'inline'; } else b.style.display = 'none'; }
    const k = document.getElementById('k-msg'); if (k) k.textContent = count || '0';
  } catch(e) {}
}

async function loadMessages() {
  try {
    const _h = {headers:{'Authorization':'Bearer '+ADMIN_TOKEN}};
    const [ri, rs] = await Promise.all([
      fetch('/api/v1/messagerie/messages?canal=interne&box=inbox', _h),
      fetch('/api/v1/messagerie/messages?canal=interne&box=sent', _h)
    ]);
    const di = ri.ok ? await ri.json() : {messages:[]};
    const ds = rs.ok ? await rs.json() : {messages:[]};
    msgData = {received: di.messages || [], sent: ds.messages || []};
  } catch(e) {}
  renderMsgList();
  const sel = document.getElementById('compose-dest');
  if (sel) {
    sel.innerHTML = '<option value="TOUS">📢 Toutes les instances</option>';
    allData.forEach(e => { const opt = document.createElement('option'); opt.value = e.sigle; opt.textContent = e.sigle + (e.nom ? ' — ' + e.nom : ''); sel.appendChild(opt); });
  }
  fetchMsgBadge();
}

function msgSwitch(mode, btn) {
  msgMode = mode;
  document.querySelectorAll('.msg-sub-btn').forEach(b => b.classList.remove('active'));
  if (btn) btn.classList.add('active');
  renderMsgList();
  document.getElementById('msg-detail-pane').innerHTML = '<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;height:100%;gap:10px;color:var(--muted)"><span style="font-size:32px;opacity:.2">✉</span><span style="font-family:var(--mono);font-size:9px">Sélectionnez un message</span></div>';
}

function renderMsgList() {
  const el = document.getElementById('msg-list-pane');
  if (!el) return;  // ancien panneau remplacé par l'iframe — ne rien faire
  const list = msgMode === 'recu' ? msgData.received : msgData.sent;
  if (!list || !list.length) {
    el.innerHTML = '<div style="padding:30px;text-align:center;font-family:var(--mono);font-size:9px;color:var(--muted)">' + (msgMode==='recu'?'Aucun message reçu':'Aucun message envoyé') + '</div>';
    return;
  }
  el.innerHTML = list.slice().reverse().map(m => {
    const unread = msgMode==='recu' && !m.lu;
    const contact = msgMode==='recu'
      ? (m.expediteur_nom || m.expediteur_addr || 'Instance')
      : ('→ ' + ((m.destinataires||[]).map(d => d.display || d.value).join(', ') || 'Instance'));
    const pj = (m.attachments_count > 0) ? ' 📎' : '';
    return `<div class="msg-item ${unread?'unread':''}" onclick="openMsg(${m.id})">
      <div style="display:flex;justify-content:space-between;margin-bottom:2px">
        <span class="msg-from">${contact}</span>
        <span class="msg-date">${fmtDate(m.created_at)}</span>
      </div>
      <div class="msg-subj" style="font-weight:${unread?700:400}">${(m.sujet||'(sans objet)')}${pj}</div>
      <div class="msg-prev">${(m.contenu||'').substring(0,60)}</div>
    </div>`;
  }).join('');
}

async function openMsg(id) {
  let m = null;
  try {
    const r = await fetch('/api/v1/messagerie/messages/' + id, {headers:{'Authorization':'Bearer '+ADMIN_TOKEN}});
    if (r.ok) m = await r.json();
  } catch(e) {}
  try { await fetch('/api/v1/messagerie/' + id + '/lire', {method:'PUT', headers:{'Authorization':'Bearer '+ADMIN_TOKEN}}); } catch(e) {}
  if (!m) return;
  const el = document.getElementById('msg-detail-pane');
  const pjHtml = (m.attachments || []).map(a =>
    '<a href="/api/v1/messagerie/attachments/' + a.id + '?token=' + encodeURIComponent(ADMIN_TOKEN) + '" target="_blank" style="display:inline-flex;align-items:center;gap:6px;padding:6px 10px;background:var(--s2);border:1px solid var(--border2);border-radius:6px;font-size:11px;text-decoration:none;color:var(--text);margin:4px 6px 0 0">📎 ' +
    (a.nom || 'fichier') + ' <span style="color:var(--muted);font-size:9px">' + Math.round((a.taille||0)/1024) + 'Ko</span></a>'
  ).join('');
  const from = m.expediteur_nom || m.expediteur_addr || 'Instance';
  el.innerHTML = `
    <div style="max-width:640px">
      <h2 style="font-family:var(--mono);font-size:14px;font-weight:700;color:var(--text);margin-bottom:14px">${m.sujet||'(sans objet)'}</h2>
      <div style="display:flex;align-items:center;gap:12px;padding:10px 14px;background:var(--s2);border:1px solid var(--border2);border-radius:7px;margin-bottom:20px">
        <div style="width:32px;height:32px;border-radius:50%;background:rgba(0,207,255,.1);display:flex;align-items:center;justify-content:center;font-size:14px">🏥</div>
        <div style="flex:1">
          <div style="font-family:var(--mono);font-size:9px;font-weight:700;color:var(--cyan)">${from}</div>
          <div style="font-family:var(--mono);font-size:8px;color:var(--muted)">${fmtDate(m.created_at)}</div>
        </div>
      </div>
      <div style="font-family:var(--body);font-size:13px;line-height:1.75;color:var(--text);white-space:pre-wrap">${(m.contenu||'').replace(/</g,'&lt;')}</div>
      ${pjHtml ? '<div style="margin-top:18px;border-top:1px solid var(--border2);padding-top:12px"><div style="font-family:var(--mono);font-size:9px;color:var(--muted);margin-bottom:6px">PIÈCES JOINTES</div>'+pjHtml+'</div>' : ''}
    </div>`;
  await loadMessages();
}

function openCompose() {
  ['compose-sujet','compose-body'].forEach(id => { const el=document.getElementById(id); if(el) el.value=''; });
  document.getElementById('modal-compose').classList.add('open');
}

async function sendMsg() {
  const dest    = document.getElementById('compose-dest')?.value || 'TOUS';
  const sujet   = document.getElementById('compose-sujet')?.value?.trim() || '';
  const contenu = document.getElementById('compose-body')?.value?.trim() || '';
  if (!contenu) { toast('Message vide', 'err'); return; }
  const r = await fetch('/api/supervision/messages', {
    method:'POST',
    headers:{'Authorization':'Bearer '+ADMIN_TOKEN, 'Content-Type':'application/json'},
    body: JSON.stringify({destinataire:dest, sujet, contenu, expediteur_nom:'Supervision'})
  });
  if (r.ok) { toast('Message envoyé ✓', 'ok'); closeModal('modal-compose'); msgSwitch('envoyes', document.getElementById('btn-envoyes')); loadMessages(); }
  else toast('Erreur envoi', 'err');
}

// ── RELAIS UPSTREAM ──────────────────────────────────────────────────────────
async function loadRelays() {
  try {
    const r = await fetch('/api/admin/relay', {headers:{'Authorization':'Bearer '+ADMIN_TOKEN}});
    if (!r.ok) return;
    renderRelays(await r.json());
  } catch(e) {}
}
function renderRelays(relays) {
  const el = document.getElementById('relay-list');
  if (!el) return;
  if (!relays.length) { el.innerHTML = '<span style="font-family:var(--mono);font-size:8px;color:var(--muted)">Aucun relais configuré</span>'; return; }
  el.innerHTML = relays.map((r,i) => `<div style="padding:5px 0;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:5px">
    <div style="width:6px;height:6px;border-radius:50%;background:${r.actif?'var(--green)':'var(--muted)'};flex-shrink:0"></div>
    <div style="flex:1;min-width:0"><div style="font-family:var(--mono);font-size:9px;color:var(--text);font-weight:700">${r.nom}</div>
    <div style="font-family:var(--mono);font-size:7px;color:var(--muted);overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${r.url}</div></div>
    <button onclick="toggleRelay(${i})" style="font-family:var(--mono);font-size:7px;padding:2px 5px;background:transparent;border:1px solid var(--border2);border-radius:3px;color:var(--muted);cursor:pointer">${r.actif?'⏸':'▶'}</button>
    <button onclick="testRelayIdx(${i})" style="font-family:var(--mono);font-size:7px;padding:2px 5px;background:transparent;border:1px solid var(--border2);border-radius:3px;color:var(--cyan);cursor:pointer">🔌</button>
    <button onclick="deleteRelay(${i})" style="font-family:var(--mono);font-size:7px;padding:2px 5px;background:transparent;border:1px solid rgba(255,45,85,.3);border-radius:3px;color:var(--red);cursor:pointer">✕</button>
  </div>`).join('');
}
async function registerscribeTokens() {
  const r = await fetch('/api/admin/tokens/Territoire', {
    method: 'POST',
    headers: {'Authorization': 'Bearer ' + ADMIN_TOKEN}
  });
  const d = await r.json().catch(()=>({}));
  if (r.ok) {
    const msg = d.added?.length > 0
      ? `✓ ${d.added.length} token(s) ajouté(s) : ${d.added.join(', ')} — total: ${d.total}`
      : `✓ Tokens déjà présents (total: ${d.total})`;
    toast(msg, 'ok');
    fetchAll();
  } else toast('Erreur : ' + (d.detail||r.status), 'err');
}

function openAddRelay() {
  ['relay-nom','relay-url','relay-token'].forEach(id=>{const e=document.getElementById(id);if(e)e.value='';});
  document.getElementById('modal-add-relay').classList.add('open');
}
async function doAddRelay() {
  const nom=document.getElementById('relay-nom')?.value?.trim();
  const url=document.getElementById('relay-url')?.value?.trim();
  const token=document.getElementById('relay-token')?.value?.trim();
  if(!nom||!url||!token){toast('Tous les champs requis','err');return;}
  const r=await fetch('/api/admin/relay',{method:'POST',headers:{'Authorization':'Bearer '+ADMIN_TOKEN,'Content-Type':'application/json'},body:JSON.stringify({nom,url,token,actif:true})});
  const d=await r.json().catch(()=>({}));
  if(r.ok){toast('Relais configuré : '+nom,'ok');closeModal('modal-add-relay');loadRelays();}
  else toast(d.detail||'Erreur','err');
}
async function toggleRelay(idx) {
  const r=await fetch(`/api/admin/relay/${idx}/toggle`,{method:'PUT',headers:{'Authorization':'Bearer '+ADMIN_TOKEN}});
  const d=await r.json().catch(()=>({}));
  if(r.ok){toast(d.message,'ok');loadRelays();}
}
async function testRelayIdx(idx) {
  const r=await fetch(`/api/admin/relay/${idx}/test`,{method:'POST',headers:{'Authorization':'Bearer '+ADMIN_TOKEN}});
  const d=await r.json().catch(()=>({}));
  toast(d.ok?`✓ ${d.nom} joignable (${d.status})`:`✗ ${d.nom} : ${d.error||d.status}`,d.ok?'ok':'err');
}
async function deleteRelay(idx) {
  if(!confirm('Supprimer ce relais ?'))return;
  const r=await fetch(`/api/admin/relay/${idx}`,{method:'DELETE',headers:{'Authorization':'Bearer '+ADMIN_TOKEN}});
  if(r.ok){toast('Relais supprimé','ok');loadRelays();}
}

// ── INIT ─────────────────────────────────────────────────────
document.querySelectorAll('.modal-overlay').forEach(m => m.addEventListener('click', e => { if(e.target===m) m.classList.remove('open'); }));
fetchAll();
setInterval(fetchAll, 30000);
// h154d — Repositionnement fluide de l'ambulance toutes les 15 s, sans
// re-télécharger les données ni les tracés OSRM (cache géométrique). Sans ce
// timer, l'ambulance ne bougeait qu'au rafraîchissement complet (30 s).
setInterval(function(){
  if (typeof map !== "undefined" && map) {
    try { updateTransfertLayer(); updateTransfertLayerOSRM(); } catch(e) {}
  }
}, 15000);

// ── CHAT COLLECTEUR ─────────────────────────────────────────
var _chatCollSalon = null;
var _chatCollLastId = 0;
var _chatCollTimer = null;

async function loadChatColl() {
  var tok = localStorage.getItem('coll_session') || '';
  var r = await fetch('api/chat/salons', {headers: {Authorization: 'Bearer ' + tok}}).catch(function(){return null;});
  if (!r || !r.ok) return;
  var salons = await r.json();
  var box = document.getElementById('chat-coll-salons');
  if (box) {
    box.innerHTML = '';
    salons.forEach(function(s) {
      var btn = document.createElement('button');
      btn.textContent = '#' + s;
      btn.style.cssText = 'font-family:var(--mono);font-size:9px;padding:4px 10px;background:var(--surface2);border:1px solid var(--border);border-radius:4px;cursor:pointer;margin:2px';
      btn.onclick = function() { openChatCollSalon(s); };
      box.appendChild(btn);
    });
    if (salons.length && !_chatCollSalon) openChatCollSalon(salons[0]);
  }
}


async function openChatCollSalon(nom) {
  _chatCollSalon = nom;
  _chatCollLastId = 0;
  if (_chatCollTimer) clearInterval(_chatCollTimer);
  // Mettre en évidence le salon actif
  var box = document.getElementById('chat-coll-salons');
  if (box) {
    box.querySelectorAll('button').forEach(function(btn) {
      btn.style.background = btn.textContent.trim() === ('#' + nom) ? 'rgba(0,49,137,.2)' : 'var(--surface2)';
      btn.style.borderColor = btn.textContent.trim() === ('#' + nom) ? '#003189' : 'var(--border)';
      btn.style.color = btn.textContent.trim() === ('#' + nom) ? '#003189' : '';
      btn.style.fontWeight = btn.textContent.trim() === ('#' + nom) ? '700' : '';
    });
  }
  // Vider le feed
  var feed = document.getElementById('chat-coll-feed');
  if (feed) feed.innerHTML = '<div style="font-family:var(--mono);font-size:10px;color:var(--muted);text-align:center;padding:20px">Chargement...</div>';
  document.getElementById('chat-coll-count').textContent = 'salon: #' + nom;
  fetchChatCollMessages();
  _chatCollTimer = setInterval(fetchChatCollMessages, 3000);
}

async function fetchChatCollMessages() {
  if (!_chatCollSalon) return;
  const r = await fetch('api/chat/messages?salon_nom=' + encodeURIComponent(_chatCollSalon) + '&since_id=' + _chatCollLastId,
    {headers: {Authorization: 'Bearer ' + (localStorage.getItem('coll_session') || '')}}).catch(()=>null);
  if (!r || !r.ok) return;
  const msgs = await r.json();
  const feed = document.getElementById('chat-coll-feed');
  if (!feed) return;
  if (!msgs.length) {
    if (_chatCollLastId === 0) {
      feed.innerHTML = '<div style="font-family:var(--mono);font-size:10px;color:var(--muted);text-align:center;padding:30px;opacity:.6">Aucun message dans #' + _chatCollSalon + '</div>';
    }
    return;
  }
  if (_chatCollLastId === 0) feed.innerHTML = '';
  msgs.forEach(m => {
    if (m.id > _chatCollLastId) _chatCollLastId = m.id;
    const ts = m.horodatage ? new Date(m.horodatage.endsWith('Z') ? m.horodatage : m.horodatage + 'Z').toLocaleTimeString('fr-FR',{hour:'2-digit',minute:'2-digit'}) : '';
    const div = document.createElement('div');
    div.style.cssText = 'padding:6px 10px;background:var(--surface2);border-radius:6px;border-left:3px solid #7c3aed';
    div.innerHTML = '<div style="font-family:var(--mono);font-size:9px;color:#7c3aed;margin-bottom:2px">' +
      (m.auteur_nom || '') + ' <span style="color:var(--muted)">' + ts + '</span></div>' +
      '<div style="font-size:12px">' + m.contenu.replace(/&/g,'&amp;').replace(/</g,'&lt;') + '</div>';
    feed.appendChild(div);
  });
  feed.scrollTop = feed.scrollHeight;
  document.getElementById('chat-coll-count').textContent = 'salon: #' + _chatCollSalon;
}

// ── GESTION COMPTES ────────────────────────────────────────
async function loadComptes() {
  var tok = localStorage.getItem('coll_session') || '';
  var r = await fetch('api/ui/users', {headers: {Authorization: 'Bearer ' + tok}}).catch(function(){return null;});
  if (!r || !r.ok) return;
  var users = await r.json();
  var list = document.getElementById('comptes-list');
  if (!list) return;
  list.innerHTML = '';
  if (!users.length) {
    list.innerHTML = '<div style="font-family:var(--mono);font-size:10px;color:var(--muted)">Aucun compte</div>';
    return;
  }
  users.forEach(function(u) {
    var row = document.createElement('div');
    row.style.cssText = 'display:flex;align-items:center;gap:10px;padding:8px 0;border-bottom:1px solid var(--border)';
    var badge = u.role === 'admin' ? 'rgba(0,49,137,.15)' : 'rgba(100,116,139,.15)';
    var html = '<span style="font-family:var(--mono);font-size:11px;flex:1">' + u.login + '</span>' +
      '<span style="font-family:var(--mono);font-size:9px;padding:2px 8px;background:' + badge + ';border-radius:10px">' + u.role + '</span>' +
      '<button style="font-family:var(--mono);font-size:9px;padding:3px 8px;background:none;border:1px solid var(--border);border-radius:4px;cursor:pointer">Changer mdp</button>';
    if (u.login !== 'supervision') {
      html += '<button style="font-family:var(--mono);font-size:9px;padding:3px 8px;background:rgba(239,68,68,.1);border:1px solid #ef4444;border-radius:4px;cursor:pointer;color:#ef4444">Supprimer</button>';
    }
    row.innerHTML = html;
    var btns = row.querySelectorAll('button');
    btns[0].onclick = function() { changePass(u.login); };
    if (btns[1]) btns[1].onclick = function() { deleteCompte(u.login); };
    list.appendChild(row);
  });
}


async function createCompte() {
  const login = document.getElementById('new-login').value.trim();
  const pass = document.getElementById('new-pass').value.trim();
  const role = document.getElementById('new-role').value;
  if (!login || !pass) { alert('Login et mot de passe requis'); return; }
  const r = await fetch('api/ui/users', {
    method: 'POST', headers: {Authorization:'Bearer '+localStorage.getItem('coll_session') || '','Content-Type':'application/json'},
    body: JSON.stringify({login, password: pass, role})
  }).catch(()=>null);
  if (r && r.ok) {
    document.getElementById('new-login').value = '';
    document.getElementById('new-pass').value = '';
    loadComptes();
  } else { alert('Erreur création'); }
}

async function deleteCompte(login) {
  if (!confirm('Supprimer le compte ' + login + ' ?')) return;
  const r = await fetch('api/ui/users/' + login, {
    method: 'DELETE', headers: {Authorization: 'Bearer ' + (localStorage.getItem('coll_session') || '')}
  }).catch(()=>null);
  if (r && r.ok) loadComptes();
}

async function changePass(login) {
  const np = prompt('Nouveau mot de passe pour ' + login + ' :');
  if (!np) return;
  // v3.4 (h38f) — Renommé : cet endpoint sert à un admin pour changer
  // le mdp d'un AUTRE compte (différent de /api/ui/change-password qui
  // est le self-service pour son propre mdp).
  const r = await fetch('api/ui/users/change-password', {
    method: 'POST', headers: {Authorization:'Bearer '+localStorage.getItem('coll_session') || '','Content-Type':'application/json'},
    body: JSON.stringify({login, new_password: np})
  }).catch(()=>null);
  if (r && r.ok) alert('Mot de passe modifié');
  else alert('Erreur');
}

async function downloadUserTemplate() {
  var tok = localStorage.getItem('coll_session') || '';
  try {
    var r = await fetch('api/admin/users/template.xlsx', {headers:{Authorization:'Bearer '+tok}});
    if (!r.ok) { toast('Erreur téléchargement du modèle', 'err'); return; }
    var blob = await r.blob();
    var a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = 'modele_comptes_supervision.xlsx';
    document.body.appendChild(a); a.click(); a.remove();
    URL.revokeObjectURL(a.href);
  } catch(e) { toast('Erreur réseau', 'err'); }
}

async function importUsers(input) {
  var f = input.files && input.files[0];
  if (!f) return;
  var tok = localStorage.getItem('coll_session') || '';
  var fd = new FormData(); fd.append('file', f);
  var res = document.getElementById('import-result');
  if (res) res.innerHTML = '<span style="color:var(--muted)">Import en cours…</span>';
  try {
    var r = await fetch('api/admin/users/import', {method:'POST', headers:{Authorization:'Bearer '+tok}, body: fd});
    input.value = '';
    if (!r.ok) { if (res) res.innerHTML = '<span style="color:#ef4444">Échec de l\'import (' + r.status + ')</span>'; return; }
    var d = await r.json();
    var html = '<div style="color:#15803d;font-weight:700">✓ ' + d.created.length + ' compte(s) créé(s)</div>';
    if (d.skipped && d.skipped.length) {
      html += '<div style="color:#d97706">⚠ ' + d.skipped.length + ' ignoré(s) (login déjà existant) : ' + escapeHtmlTA(d.skipped.join(', ')) + '</div>';
    }
    if (d.temp_passwords && d.temp_passwords.length) {
      html += '<div style="margin-top:8px;font-weight:700">🔑 Mots de passe temporaires générés (à communiquer aux utilisateurs) :</div>';
      html += '<div style="font-family:var(--mono);font-size:10px;background:var(--surface);border:1px solid var(--border);border-radius:5px;padding:8px;margin-top:4px">' +
        d.temp_passwords.map(function(t){ return escapeHtmlTA(t.login) + ' : ' + escapeHtmlTA(t.password); }).join('<br>') + '</div>';
    }
    if (res) res.innerHTML = html;
    loadComptes();
  } catch(e) { if (res) res.innerHTML = '<span style="color:#ef4444">Erreur réseau</span>'; }
}

// ── v3000h25 — Assistant territorial ──────────────────────────────────────
function escapeHtmlTA(s) {
  return String(s||'').replace(/[&<>"']/g, function(c) {
    return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c];
  });
}

function refreshTerritorial() {
  fetch('/api/territorial-assistant')
    .then(function(r){ return r.ok ? r.json() : null; })
    .then(function(data){
      if (!data) return;
      renderTerritorial(data);
    })
    .catch(function(e){
      console.warn('[territorial] erreur:', e);
    });
}

function renderTerritorial(data) {
  var sum = data.summary || {};
  var alertes = data.alertes || [];

  var elOnline = document.getElementById('ta-online');
  var elCrit = document.getElementById('ta-crit-sites');
  var elPB = document.getElementById('ta-pb');
  var elNbA = document.getElementById('ta-nb-alertes');
  var elBadge = document.getElementById('ta-tab-badge');
  var elClock = document.getElementById('ta-clock');

  if (elOnline) elOnline.textContent = (sum.etablissements_online || 0) + ' / ' + (sum.total_etablissements || 0);
  if (elCrit) elCrit.textContent = sum.etablissements_avec_critiques || 0;
  if (elPB) elPB.textContent = sum.plans_blancs_actifs || 0;
  if (elNbA) elNbA.textContent = alertes.length;

  if (elBadge) {
    if (alertes.length > 0) {
      elBadge.style.display = 'inline-block';
      elBadge.textContent = alertes.length;
    } else {
      elBadge.style.display = 'none';
    }
  }

  if (elClock && data.generated_at) {
    try {
      var d = new Date(data.generated_at);
      elClock.textContent = d.toLocaleTimeString('fr-FR', {hour:'2-digit',minute:'2-digit',second:'2-digit'});
    } catch(e) {}
  }

  var list = document.getElementById('ta-alertes-list');
  if (!list) return;
  if (alertes.length === 0) {
    list.innerHTML = '<div style="text-align:center;padding:30px;color:var(--muted);font-size:12px;font-style:italic">'
      + '\u2713 Aucun signal territorial d\u00e9tect\u00e9 actuellement.</div>';
    return;
  }

  list.innerHTML = alertes.map(function(a) {
    var color = a.niveau === 'alert' ? '#dc2626' : '#003189';
    var bgGradient = a.niveau === 'alert'
      ? 'linear-gradient(135deg, rgba(220,38,38,0.08), rgba(220,38,38,0.02))'
      : 'linear-gradient(135deg, rgba(0,49,137,0.08), rgba(0,49,137,0.02))';
    var concernes = (a.etablissements_concernes || []).join(', ');
    var niveauLabel = a.niveau === 'alert' ? 'ALERTE TERRITORIALE' : 'SIGNAL';
    return [
      '<div style="background:' + bgGradient + ';border:1px solid ' + color + '40;border-left:4px solid ' + color + ';',
      '     border-radius:8px;padding:14px 16px">',
      '  <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;flex-wrap:wrap">',
      '    <span style="font-family:var(--mono);font-size:9px;font-weight:700;letter-spacing:1.5px;color:' + color + ';',
      '          background:' + color + '20;padding:3px 8px;border-radius:10px">' + niveauLabel + '</span>',
      concernes ? '    <span style="font-family:var(--mono);font-size:10px;color:var(--muted)">' + escapeHtmlTA(concernes) + '</span>' : '',
      '  </div>',
      '  <div style="font-size:14px;font-weight:700;margin-bottom:6px;color:var(--text)">',
      '    ' + escapeHtmlTA(a.titre || ''),
      '  </div>',
      '  <div style="font-size:12.5px;line-height:1.5;color:var(--text);opacity:0.9">',
      '    ' + escapeHtmlTA(a.message || ''),
      '  </div>',
      '</div>',
    ].join('');
  }).join('');
}

// Démarrer le polling de l'Assistant territorial dès le chargement
window.addEventListener('load', function() {
  refreshTerritorial();
  setInterval(refreshTerritorial, 5000);
});

// ── v3000h31 — Modale KPI cliquables ─────────────────────────────────────
function escapeHtmlKpi(s) {
  return String(s||'').replace(/[&<>"']/g, function(c) {
    return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c];
  });
}

function openKpiModal(type) {
  var backdrop = document.getElementById('kpi-modal-backdrop');
  var icon = document.getElementById('kpi-modal-icon');
  var titleText = document.getElementById('kpi-modal-title-text');
  var count = document.getElementById('kpi-modal-count');
  var body = document.getElementById('kpi-modal-body');

  var titles = {
    sites:     {icon:'🏥', title:'SITES PAR ÉTABLISSEMENT'},
    pending:   {icon:'⏳', title:'EN ATTENTE D\'ENRÔLEMENT'},
    incidents: {icon:'⚠️', title:'INCIDENTS ACTIFS — GROUPÉS PAR ÉTABLISSEMENT'},
    critiques: {icon:'🔴', title:'INCIDENTS CRITIQUES (U≥3)'},
  };
  var t = titles[type] || {icon:'📋', title:'DÉTAIL'};
  icon.textContent = t.icon;
  titleText.textContent = t.title;
  count.textContent = '...';
  body.innerHTML = '<div class="kpi-modal-empty">Chargement…</div>';
  backdrop.classList.add('show');

  // Charger les données depuis /api/summary (déjà récupéré dans allData)
  // si dispo, sinon faire un fetch
  if (typeof allData !== 'undefined' && Array.isArray(allData) && allData.length > 0) {
    renderKpiModal(type, allData);
  } else {
    fetch('/api/summary', {
      headers: {Authorization: 'Bearer ' + (localStorage.getItem('coll_session') || '')}
    }).then(function(r){ return r.ok ? r.json() : []; })
      .then(function(data){ renderKpiModal(type, data); })
      .catch(function(){ body.innerHTML = '<div class="kpi-modal-empty">Erreur de chargement</div>'; });
  }
}

function closeKpiModal() {
  document.getElementById('kpi-modal-backdrop').classList.remove('show');
}

// Fermer la modale avec Esc
document.addEventListener('keydown', function(e) {
  if (e.key === 'Escape') closeKpiModal();
});

function renderKpiModal(type, data) {
  var body = document.getElementById('kpi-modal-body');
  var count = document.getElementById('kpi-modal-count');

  if (type === 'sites') {
    renderKpiSites(data, body, count);
  } else if (type === 'pending') {
    renderKpiPending(body, count);
  } else if (type === 'incidents') {
    renderKpiIncidents(data, body, count, false);
  } else if (type === 'critiques') {
    renderKpiIncidents(data, body, count, true);
  }
}

function renderKpiSites(data, body, count) {
  var total = 0;
  var html = data.map(function(e) {
    var sites = e.sites || [];
    total += sites.length;
    if (!sites.length) return '';
    var lvlColor = ({CRITIQUE:'#e1000f', CRISE:'#f97316', ALERTE:'#f5c518', NOMINAL:'#22c55e'})[e.niveau_global] || '#94a3b8';
    var sitesHtml = sites.map(function(s) {
      var nivBadge = s.niveau && s.niveau !== 'NOMINAL'
        ? '<span class="kpi-item-urg ' + (s.niveau==='CRITIQUE'?'u4':s.niveau==='CRISE'?'u3':'u2') + '">' + s.niveau + '</span>'
        : '<span class="kpi-item-urg u1">OK</span>';
      var incBadge = s.incidents_ouverts > 0
        ? '<span class="kpi-item-type">' + s.incidents_ouverts + ' inc.</span>'
        : '';
      return '<div class="kpi-item">' + nivBadge +
        '<div class="kpi-item-text">' + escapeHtmlKpi(s.nom || '?') +
        (s.adresse ? '<div class="kpi-item-site">' + escapeHtmlKpi(s.adresse) + '</div>' : '') +
        '</div>' + incBadge + '</div>';
    }).join('');
    return '<div class="kpi-group">' +
      '<div class="kpi-group-hdr">' +
        '<div class="kpi-group-name">' +
          '<span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:' + lvlColor + '"></span>' +
          escapeHtmlKpi(e.sigle) + ' — ' + escapeHtmlKpi(e.nom || '') +
        '</div>' +
        '<div class="kpi-group-count">' + sites.length + ' site(s)</div>' +
      '</div>' +
      '<div class="kpi-group-items">' + sitesHtml + '</div>' +
    '</div>';
  }).join('');
  count.textContent = total;
  body.innerHTML = html || '<div class="kpi-modal-empty">Aucun site enregistré</div>';
}

function renderKpiPending(body, count) {
  // Lire pending depuis l'API admin
  fetch('/api/admin/pending', {
    headers: {Authorization: 'Bearer PLACEHOLDER_ADMIN_TOKEN'}
  }).then(function(r){ return r.ok ? r.json() : null; })
    .then(function(data){
      var list = data && data.pending ? Object.entries(data.pending) : [];
      count.textContent = list.length;
      if (!list.length) {
        body.innerHTML = '<div class="kpi-modal-empty">✓ Aucun établissement en attente d\'enrôlement.</div>';
        return;
      }
      body.innerHTML = list.map(function(entry) {
        var token = entry[0];
        var info = entry[1];
        return '<div class="kpi-group">' +
          '<div class="kpi-group-hdr">' +
            '<div class="kpi-group-name">' +
              '<span style="color:#f5c518">⏳</span> ' + escapeHtmlKpi(info.sigle_propose || '?') +
              ' — ' + escapeHtmlKpi(info.nom_propose || '?') +
            '</div>' +
            '<div class="kpi-group-count">' + escapeHtmlKpi(info.first_seen_at || '') + '</div>' +
          '</div>' +
          '<div class="kpi-item">' +
            '<span class="kpi-item-type">TOKEN</span>' +
            '<div class="kpi-item-text" style="font-family:var(--mono);font-size:10px;word-break:break-all">' +
              escapeHtmlKpi(token.substring(0,16)) + '…' +
            '</div>' +
          '</div>' +
        '</div>';
      }).join('');
    })
    .catch(function(){
      body.innerHTML = '<div class="kpi-modal-empty">Erreur — impossible de récupérer la liste des établissements en attente</div>';
    });
}

function renderKpiIncidents(data, body, count, critiquesOnly) {
  var totalCount = 0;
  var html = data.map(function(e) {
    var incidents = e.incidents || [];
    // Filtrer par status non-résolu
    incidents = incidents.filter(function(i) {
      var s = (i.status || '').toUpperCase();
      return ['RÉSOLU','RESOLU','ARCHIVÉ','ARCHIVE','FERMÉ','CLOS'].indexOf(s) < 0;
    });
    if (critiquesOnly) {
      incidents = incidents.filter(function(i) {
        return (i.urgency || i.urgence || 1) >= 3;
      });
    }
    if (!incidents.length) return '';
    totalCount += incidents.length;
    // Trier par urgency décroissante
    incidents.sort(function(a,b) {
      return (b.urgency||1) - (a.urgency||1);
    });
    var lvlColor = ({CRITIQUE:'#e1000f', CRISE:'#f97316', ALERTE:'#f5c518', NOMINAL:'#22c55e'})[e.niveau_global] || '#94a3b8';
    var incHtml = incidents.map(function(i) {
      var u = i.urgency || i.urgence || 1;
      var uClass = u >= 4 ? 'u4' : u >= 3 ? 'u3' : u >= 2 ? 'u2' : 'u1';
      var uLabel = 'U' + u;
      return '<div class="kpi-item">' +
        '<span class="kpi-item-urg ' + uClass + '">' + uLabel + '</span>' +
        '<span class="kpi-item-type">' + escapeHtmlKpi(i.type_crise || '?') + '</span>' +
        '<div class="kpi-item-text">' + escapeHtmlKpi(i.fait_resume || '(sans description)') +
        (i.site ? '<div class="kpi-item-site">📍 ' + escapeHtmlKpi(i.site) + '</div>' : '') +
        '</div>' +
      '</div>';
    }).join('');
    return '<div class="kpi-group">' +
      '<div class="kpi-group-hdr">' +
        '<div class="kpi-group-name">' +
          '<span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:' + lvlColor + '"></span>' +
          escapeHtmlKpi(e.sigle) + ' — ' + escapeHtmlKpi(e.nom || '') +
          '<span class="kpi-group-level" style="background:' + lvlColor + '20;color:' + lvlColor + '">' +
            escapeHtmlKpi(e.niveau_global || '?') + '</span>' +
        '</div>' +
        '<div class="kpi-group-count">' + incidents.length + ' incident(s)</div>' +
      '</div>' +
      '<div class="kpi-group-items">' + incHtml + '</div>' +
    '</div>';
  }).join('');
  count.textContent = totalCount;
  body.innerHTML = html || '<div class="kpi-modal-empty">' +
    (critiquesOnly ? '✓ Aucun incident critique (U≥3) actif.' : '✓ Aucun incident actif.') +
    '</div>';
}

</script>
<div id="help-overlay" onclick="if(event.target===this)closeHelpCenter()">
  <div id="help-box">
    <div id="help-hdr">
      <span id="help-title">Centre d'aide</span>
      <div style="display:flex;align-items:center;gap:6px">
        
        
        <button id="help-close" onclick="closeHelpCenter()" aria-label="Fermer">✕</button>
      </div>
    </div>
    <div id="help-search-wrap"><input type="text" id="help-search" oninput="helpRender()" autocomplete="off"></div>
    <div id="help-body">
      <div id="help-list"></div>
      <div id="help-article" style="display:none"></div>
    </div>
  </div>
</div>
<script>
/* ── h111 — Centre d'aide (bilingue FR/EN, recherchable) ─────────────────── */
function helpLang2(){var c;try{c=(localStorage.getItem('scribe_lang_pref')||'en');}catch(e){c='en';}return String(c).slice(0,2).toLowerCase();}
var HELP_ARTICLES = [
 {id:'presentation', cat:{fr:'Démarrer',en:'Getting started',de:'Erste Schritte',es:'Primeros pasos',it:'Per iniziare',nl:'Aan de slag'},
  title:{fr:'Présentation & architecture',en:'Overview & architecture',de:'Überblick & Architektur',es:'Visión general y arquitectura',it:'Panoramica e architettura',nl:'Overzicht & architectuur'},
  kw:'scribe presentation about architecture instance supervision collecteur souverain',
  body:{fr:'<p>SCRIBE est une plateforme open-source de gestion de crise hospitalière et de suivi capacitaire, pensée pour la coordination entre plusieurs établissements (GHT, territoire).</p><p>Elle se compose de deux briques :</p><ul><li><b>L\'instance établissement</b> : la « main courante » d\'un hôpital — incidents, capacité en lits, transferts, rappel du personnel, messagerie, cellule de crise. Chaque établissement a sa propre base de données.</li><li><b>La supervision (collecteur)</b> : un tableau de bord territorial qui agrège l\'état de toutes les instances (incidents, statuts, capacité) sans jamais recevoir de donnée nominative patient.</li></ul><p>Les instances poussent leur situation vers le collecteur toutes les 30 secondes. Le niveau territorial affiché est toujours le <b>pire</b> entre les incidents en cours et le statut déclaré par l\'établissement.</p>',
       en:'<p>SCRIBE is an open-source hospital crisis-management and capacity-monitoring platform, designed to coordinate several establishments (hospital groups, regions).</p><p>It has two building blocks:</p><ul><li><b>The establishment instance</b>: a hospital\'s operational log — incidents, bed capacity, transfers, staff recall, messaging, crisis cell. Each establishment has its own database.</li><li><b>Supervision (collector)</b>: a territorial dashboard aggregating the state of every instance (incidents, statuses, capacity) without ever receiving patient-identifying data.</li></ul><p>Instances push their situation to the collector every 30 seconds. The territorial level shown is always the <b>worst</b> of ongoing incidents and the establishment\'s declared status.</p>'}},

 {id:'install_docker', cat:{fr:'Installation',en:'Installation',de:'Installation',es:'Instalación',it:'Installazione',nl:'Installatie'},
  title:{fr:'Installer en Docker (HTTP)',en:'Install with Docker (HTTP)',de:'Mit Docker installieren (HTTP)',es:'Instalar con Docker (HTTP)',it:'Installare con Docker (HTTP)',nl:'Installeren met Docker (HTTP)'},
  kw:'install docker compose deploy port 8000 conteneur container',
  body:{fr:'<p>Pour une instance unique, Docker est le plus simple :</p><ol><li>Place ta configuration d\'établissement dans le dossier de déploiement (<code>config.xml</code>).</li><li>Lance : <code>docker compose -f docker-compose.production.yml up -d</code></li><li>L\'application est servie en HTTP sur le port <b>8000</b>.</li></ol><p>Les données (base SQLite, pièces jointes) sont persistées dans un volume <code>/data</code>. Si tu vois une erreur <code>Permission denied</code> sur <code>/data</code>, c\'est que le dossier hôte monté n\'appartient pas à l\'utilisateur du conteneur : préfère un <b>volume nommé Docker</b> (<code>scribe_data:/data</code>), dont les droits sont gérés automatiquement.</p>',
       en:'<p>For a single instance, Docker is simplest:</p><ol><li>Put your establishment configuration in the deployment folder (<code>config.xml</code>).</li><li>Run: <code>docker compose -f docker-compose.production.yml up -d</code></li><li>The app is served over HTTP on port <b>8000</b>.</li></ol><p>Data (SQLite database, attachments) persists in a <code>/data</code> volume. If you get a <code>Permission denied</code> error on <code>/data</code>, the mounted host folder isn\'t owned by the container user: prefer a <b>named Docker volume</b> (<code>scribe_data:/data</code>), whose permissions are handled automatically.</p>'}},

 {id:'install_https', cat:{fr:'Installation',en:'Installation',de:'Installation',es:'Instalación',it:'Installazione',nl:'Installatie'},
  title:{fr:'Activer le HTTPS (Caddy)',en:'Enable HTTPS (Caddy)',de:'HTTPS aktivieren (Caddy)',es:'Activar HTTPS (Caddy)',it:'Attivare HTTPS (Caddy)',nl:'HTTPS inschakelen (Caddy)'},
  kw:'https tls ssl caddy certificat lets encrypt domaine 502',
  body:{fr:'<p>Pour servir SCRIBE en HTTPS, place <b>Caddy</b> en frontal : il obtient et renouvelle le certificat Let\'s Encrypt automatiquement.</p><ul><li><b>Prérequis</b> : un <b>nom de domaine</b> dont l\'enregistrement A pointe vers l\'IP du serveur. Let\'s Encrypt n\'émet pas pour une IP nue.</li><li>Ouvre les ports <b>80 et 443</b> (le 80 sert au challenge de validation).</li><li>Renseigne <code>SCRIBE_DOMAIN</code> dans le fichier <code>.env</code>, puis : <code>docker compose -f docker-compose.https.yml up -d</code></li></ul><p><b>Dépannage</b> : une erreur <b>502</b> en HTTPS signifie que Caddy fonctionne mais que l\'application derrière ne répond pas (conteneur en redémarrage, souvent un souci de droits sur <code>/data</code>). Vérifie les logs : <code>docker compose -f docker-compose.https.yml logs scribe</code>.</p>',
       en:'<p>To serve SCRIBE over HTTPS, put <b>Caddy</b> in front: it obtains and renews the Let\'s Encrypt certificate automatically.</p><ul><li><b>Requirement</b>: a <b>domain name</b> whose A record points to the server IP. Let\'s Encrypt won\'t issue for a bare IP.</li><li>Open ports <b>80 and 443</b> (80 is used for the validation challenge).</li><li>Set <code>SCRIBE_DOMAIN</code> in the <code>.env</code> file, then: <code>docker compose -f docker-compose.https.yml up -d</code></li></ul><p><b>Troubleshooting</b>: a <b>502</b> over HTTPS means Caddy works but the app behind it isn\'t answering (container restarting, often a <code>/data</code> permission issue). Check logs: <code>docker compose -f docker-compose.https.yml logs scribe</code>.</p>'}},

 {id:'install_multi', cat:{fr:'Installation',en:'Installation',de:'Installation',es:'Instalación',it:'Installazione',nl:'Installatie'},
  title:{fr:'Déploiement multi-instances',en:'Multi-instance deployment',de:'Multi-Instanz-Bereitstellung',es:'Despliegue multi-instancia',it:'Distribuzione multi-istanza',nl:'Multi-instance-implementatie'},
  kw:'multi instance ght g7 collecteur ports scripts bash federation territoire',
  body:{fr:'<p>Pour un territoire, on lance plusieurs instances d\'établissement plus un collecteur, chacun sur son port. Les scripts de lancement démarrent les processus directement (HTTP) :</p><ul><li>Chaque établissement écoute sur un port dédié (ex. 8000, 8001, …).</li><li>Le <b>collecteur</b> (supervision) écoute sur son propre port (ex. 9000).</li><li>Une <b>instance de démonstration</b> peut tourner en parallèle, isolée, avec son propre collecteur.</li></ul><p>Chaque instance est reliée au collecteur par un <b>token de fédération</b> déclaré dans sa configuration. Le collecteur n\'accepte que les instances pré-enrôlées (token validé). Pour exposer tout cela en HTTPS, on place un reverse proxy (Caddy ou Nginx) devant, avec un routage par chemin ou par sous-domaine.</p>',
       en:'<p>For a territory, you run several establishment instances plus one collector, each on its own port. Launch scripts start the processes directly (HTTP):</p><ul><li>Each establishment listens on a dedicated port (e.g. 8000, 8001, …).</li><li>The <b>collector</b> (supervision) listens on its own port (e.g. 9000).</li><li>A <b>demo instance</b> can run in parallel, isolated, with its own collector.</li></ul><p>Each instance connects to the collector via a <b>federation token</b> declared in its configuration. The collector only accepts pre-enrolled instances (validated token). To expose all of this over HTTPS, put a reverse proxy (Caddy or Nginx) in front, with path-based or subdomain routing.</p>'}},

 {id:'configuration', cat:{fr:'Configuration',en:'Configuration',de:'Konfiguration',es:'Configuración',it:'Configurazione',nl:'Configuratie'},
  title:{fr:'Configurer l\'établissement',en:'Configure the establishment',de:'Einrichtung konfigurieren',es:'Configurar el establecimiento',it:'Configurare la struttura',nl:'De instelling configureren'},
  kw:'config configuration xml uf unite fonctionnelle pole referentiel token albert',
  body:{fr:'<p>La configuration d\'une instance se fait via un fichier XML d\'établissement et des scripts d\'initialisation qui peuplent la base. On y définit :</p><ul><li><b>L\'établissement</b> : nom, sigle, coordonnées géographiques (pour la carte).</li><li><b>Les pôles et unités fonctionnelles (UF)</b> : la structure médicale, utilisée dans les menus Incidents et Capacité.</li><li><b>Le référentiel capacitaire</b> : le nombre de lits par UF/pôle servant de base au suivi de tension.</li><li><b>La fédération</b> : l\'URL du collecteur et le token d\'enrôlement.</li><li><b>L\'IA Albert</b> : la clé d\'accès (jamais incluse dans les archives publiques).</li></ul><p>L\'administration des UF se fait aussi en direct depuis l\'interface (onglet Admin) : on peut activer/désactiver une UF. Une UF désactivée reste en base pour préserver l\'historique des incidents, mais disparaît des menus déroulants.</p>',
       en:'<p>An instance is configured via an establishment XML file and initialisation scripts that populate the database. You define:</p><ul><li><b>The establishment</b>: name, code, geographic coordinates (for the map).</li><li><b>Units and functional units</b>: the medical structure, used in the Incidents and Capacity menus.</li><li><b>The capacity baseline</b>: the number of beds per unit, the basis for strain monitoring.</li><li><b>Federation</b>: the collector URL and enrolment token.</li><li><b>Albert AI</b>: the access key (never included in public archives).</li></ul><p>Unit administration is also done live from the interface (Admin tab): you can enable/disable a unit. A disabled unit stays in the database to preserve incident history but disappears from the dropdown menus.</p>'}},

 {id:'connexion', cat:{fr:'Démarrer',en:'Getting started',de:'Erste Schritte',es:'Primeros pasos',it:'Per iniziare',nl:'Aan de slag'},
  title:{fr:'Connexion, rôles & langue',en:'Login, roles & language',de:'Anmeldung, Rollen & Sprache',es:'Inicio de sesión, roles e idioma',it:'Accesso, ruoli e lingua',nl:'Inloggen, rollen & taal'},
  kw:'login connexion role compte mot de passe langue admin cellule soignant',
  body:{fr:'<p>Saisis ton identifiant et ton mot de passe sur la page d\'accueil. Le sélecteur de <b>langue</b> est en haut de l\'écran (24 langues européennes).</p><p>Les comptes ont des <b>rôles</b> qui conditionnent ce qui est visible et autorisé : direction de crise / cellule, soignant, administrateur. Les fonctions d\'administration (gestion des UF, des comptes, déclenchement de rappel) sont réservées aux rôles habilités.</p><p>La création et la gestion des comptes se font dans <b>Administration → Utilisateurs</b>. Les mots de passe sont stockés hachés (bcrypt) ; le login est protégé contre les tentatives répétées.</p>',
       en:'<p>Enter your username and password on the home page. The <b>language</b> selector is at the top of the screen (24 European languages).</p><p>Accounts have <b>roles</b> that govern what is visible and allowed: crisis management / cell, caregiver, administrator. Administration functions (managing units, accounts, triggering recall) are restricted to authorised roles.</p><p>Account creation and management happen in <b>Administration → Users</b>. Passwords are stored hashed (bcrypt); login is protected against repeated attempts.</p>'}},

 {id:'dashboard', cat:{fr:'Utilisation',en:'Usage',de:'Nutzung',es:'Uso',it:'Utilizzo',nl:'Gebruik'},
  title:{fr:'Le tableau de bord',en:'The dashboard',de:'Das Dashboard',es:'El panel de control',it:'La dashboard',nl:'Het dashboard'},
  kw:'dashboard tableau de bord accueil kpi rappel widget jauge cooperation',
  body:{fr:'<p>Le tableau de bord est la vue de synthèse de ton établissement. On y trouve :</p><ul><li>Les <b>indicateurs clés</b> : incidents actifs, incidents critiques, capacité globale, transferts en cours, messages non lus.</li><li>La <b>capacité par pôle</b> : un coup d\'œil sur les services en tension.</li><li>La <b>coopération territoriale</b> : l\'état des établissements partenaires.</li><li>Le <b>widget Rappel du personnel</b> : lors d\'un rappel actif, une jauge demi-cercle montre les personnes appelées, celles qui ont répondu et celles qui arrivent.</li></ul><p>Le tableau se rafraîchit automatiquement. C\'est l\'écran à garder ouvert pendant une crise.</p>',
       en:'<p>The dashboard is your establishment\'s summary view. It shows:</p><ul><li><b>Key indicators</b>: active incidents, critical incidents, overall capacity, ongoing transfers, unread messages.</li><li><b>Capacity by unit</b>: an at-a-glance view of strained services.</li><li><b>Territorial cooperation</b>: the state of partner establishments.</li><li>The <b>Staff recall widget</b>: during an active recall, a half-circle gauge shows people called, those who responded, and those arriving.</li></ul><p>The dashboard refreshes automatically. It is the screen to keep open during a crisis.</p>'}},

 {id:'incidents', cat:{fr:'Utilisation',en:'Usage',de:'Nutzung',es:'Uso',it:'Utilizzo',nl:'Gebruik'},
  title:{fr:'Déclarer et suivre un incident',en:'Declare and track an incident',de:'Vorfall melden und verfolgen',es:'Declarar y seguir un incidente',it:'Dichiarare e seguire un incidente',nl:'Een incident melden en volgen'},
  kw:'incident declarer urgence gravite veille cyber sanitaire mixte impact transversal directeur diffusion',
  body:{fr:'<p>L\'onglet <b>Incidents</b> est le cœur de la main courante. Pour déclarer :</p><ol><li>Choisis le <b>type</b> : cyber, sanitaire ou mixte.</li><li>Fixe le <b>niveau d\'urgence</b>, de 1 (veille / information) à 4 (critique). Un niveau ≥ 3 (grave/critique) déclenche la notification des directeurs.</li><li>Rattache l\'incident à un <b>pôle / une UF</b> et décris le fait.</li><li>Coche <b>impact transversal</b> si l\'événement touche tout l\'établissement : les services basculent alors au moins en « tension », et le statut public passe en « perturbé ».</li><li>Diffuse avec le bouton dédié.</li></ol><p>Chaque incident est horodaté et historisé. Les incidents alimentent le niveau de gravité remonté à la supervision territoriale.</p>',
       en:'<p>The <b>Incidents</b> tab is the heart of the operational log. To declare one:</p><ol><li>Pick the <b>type</b>: cyber, health, or mixed.</li><li>Set the <b>urgency level</b>, from 1 (watch / info) to 4 (critical). A level ≥ 3 (serious/critical) triggers director notifications.</li><li>Attach the incident to a <b>unit</b> and describe the event.</li><li>Tick <b>cross-cutting impact</b> if the event affects the whole establishment: services then move at least to "strained", and the public status becomes "disrupted".</li><li>Broadcast with the dedicated button.</li></ol><p>Every incident is timestamped and archived. Incidents feed the severity level reported to territorial supervision.</p>'}},

 {id:'main_courante', cat:{fr:'Utilisation',en:'Usage',de:'Nutzung',es:'Uso',it:'Utilizzo',nl:'Gebruik'},
  title:{fr:'Main courante & tâches (kanban)',en:'Activity log & tasks (kanban)',de:'Einsatztagebuch & Aufgaben (Kanban)',es:'Registro de actividad y tareas (kanban)',it:'Diario attività e compiti (kanban)',nl:'Logboek & taken (kanban)'},
  kw:'main courante log journal kanban tache task action categorie filtre',
  body:{fr:'<p>La <b>main courante</b> journalise toutes les actions de la cellule de crise : décisions, événements, communications. Chaque entrée porte un auteur, un horodatage, une catégorie et une action. Tu peux filtrer par catégorie pour retrouver rapidement une décision.</p><p>Le <b>kanban des tâches</b> permet de suivre le « qui fait quoi » : créer des tâches, les assigner, les faire avancer par colonnes (à faire / en cours / fait). C\'est l\'outil de pilotage opérationnel pendant l\'événement.</p><p>L\'ensemble est exportable lors de l\'archivage, pour le retour d\'expérience et la traçabilité.</p>',
       en:'<p>The <b>activity log</b> records every action of the crisis cell: decisions, events, communications. Each entry has an author, timestamp, category and action. You can filter by category to quickly find a decision.</p><p>The <b>task kanban</b> tracks who-does-what: create tasks, assign them, move them across columns (to do / in progress / done). It is the operational steering tool during the event.</p><p>Everything can be exported when archiving, for after-action review and traceability.</p>'}},

 {id:'capacite', cat:{fr:'Utilisation',en:'Usage',de:'Nutzung',es:'Uso',it:'Utilizzo',nl:'Gebruik'},
  title:{fr:'Capacité en lits',en:'Bed capacity',de:'Bettenkapazität',es:'Capacidad de camas',it:'Capacità di posti letto',nl:'Bedcapaciteit'},
  kw:'capacite capacity lits beds tension pole service referentiel statut vert orange rouge',
  body:{fr:'<p>L\'onglet <b>Capacité</b> suit l\'occupation des lits par pôle et par service, à partir du référentiel capacitaire configuré.</p><p>Chaque service affiche un statut à trois niveaux : <b>normal</b> (vert), <b>tension</b> (orange), <b>critique / fermé</b> (rouge). Tu déclares la situation au fil de l\'eau ; l\'historique des déclarations est conservé.</p><p>La capacité globale de l\'établissement est calculée à partir de ces statuts et remonte à la supervision. Rappel : un <b>impact transversal</b> actif force automatiquement les services à au moins « tension ».</p>',
       en:'<p>The <b>Capacity</b> tab tracks bed occupancy by unit and service, based on the configured capacity baseline.</p><p>Each service shows a three-level status: <b>normal</b> (green), <b>strained</b> (orange), <b>critical / closed</b> (red). You declare the situation as it evolves; the declaration history is kept.</p><p>The establishment\'s overall capacity is computed from these statuses and reported to supervision. Reminder: an active <b>cross-cutting impact</b> automatically forces services to at least "strained".</p>'}},

 {id:'transferts', cat:{fr:'Utilisation',en:'Usage',de:'Nutzung',es:'Uso',it:'Utilizzo',nl:'Gebruik'},
  title:{fr:'Transferts de patients',en:'Patient transfers',de:'Patiententransfers',es:'Traslados de pacientes',it:'Trasferimenti di pazienti',nl:'Patiëntoverplaatsingen'},
  kw:'transfert transfer patient ambulance trajectoire osrm eta destination carte hds',
  body:{fr:'<p>L\'onglet <b>Transferts</b> gère les transferts entre établissements. Pour chaque transfert tu renseignes l\'établissement et le site de destination ainsi que l\'heure estimée d\'arrivée.</p><p>Sur la carte (onglet Soins), un transfert en cours s\'affiche avec sa <b>trajectoire routière</b> (calculée via OSRM) et la progression de l\'ambulance selon l\'ETA. Le site de destination exact est utilisé pour positionner correctement le trajet.</p><p><b>Important</b> : aucune donnée nominative patient ne remonte à la supervision territoriale — seul l\'établissement gère ces informations (conformité HDS / RGPD).</p>',
       en:'<p>The <b>Transfers</b> tab manages inter-establishment patient transfers. For each transfer you enter the destination establishment and site plus the estimated time of arrival.</p><p>On the map (Care tab), an ongoing transfer is shown with its <b>road route</b> (computed via OSRM) and the ambulance progress based on the ETA. The exact destination site is used to position the route correctly.</p><p><b>Important</b>: no patient-identifying data is sent to territorial supervision — only the establishment handles this information (HDS / GDPR compliance).</p>'}},

 {id:'rappel', cat:{fr:'Utilisation',en:'Usage',de:'Nutzung',es:'Uso',it:'Utilizzo',nl:'Gebruik'},
  title:{fr:'Rappel du personnel',en:'Staff recall',de:'Personalrückruf',es:'Llamada del personal',it:'Richiamo del personale',nl:'Personeel oproepen'},
  kw:'rappel personnel mobilisation recall sms mail vague escalade preset cible confirmation arrivee jauge widget telephonie bascule',
  body:{fr:'<p>Le <b>rappel du personnel</b> permet de mobiliser les agents de garde ou d\'astreinte en masse pendant une crise, via SMS et/ou e-mail.</p><h2>Flux type</h2><ol><li><b>Cibler</b> : sélectionne le personnel par type (astreinte direction, astreinte technique, cadres de santé…) et par lieu (site, service). Enregistre des <b>presets</b> (listes types) pour déclencher un rappel en deux clics.</li><li><b>Pré-visualiser</b> : avant l\'envoi, un écran de confirmation liste les destinataires, leurs contacts et les canaux retenus.</li><li><b>Envoyer</b> : SMS via la passerelle configurée (OVH, Twilio, Free…) et/ou e-mail via SMTP. Chaque message contient un lien individuel sécurisé.</li><li><b>Suivre les réponses</b> : chaque destinataire clique sur son lien et confirme son heure d\'arrivée estimée. Le <b>widget jauge demi-cercle</b> sur le tableau de bord affiche en temps réel : appelés / ont répondu / sont arrivés / manquent à l\'appel.</li><li><b>Escalader par vagues</b> : si une première vague est insuffisante, relance automatique des non-répondants. Chaque vague nécessite la validation de la cellule.</li></ol><h2>Bascule téléphonie</h2><p>En cas de crise CYBER, si le PABX/IPBX est impacté, SCRIBE bascule automatiquement l\'affichage de l\'annuaire vers les <b>numéros de secours</b> configurés. Les agents contactés via le lien de rappel voient aussi ces numéros alternatifs.</p><p>⚠️ Le rappel envoie de <b>vrais SMS à des vrais numéros</b> : toujours tester en exercice (instances exercice isolées) avant un usage réel. La configuration SMS/SMTP se fait dans <b>Admin → Notifications</b>.</p>',
       en:'<p>The <b>staff recall</b> mobilises on-call or on-duty staff at scale during a crisis, via SMS and/or email.</p><h2>Typical flow</h2><ol><li><b>Target</b>: select staff by type (management on-call, technical on-call, nursing leads…) and location (site, service). Save <b>presets</b> (standard lists) to trigger a recall in two clicks.</li><li><b>Preview</b>: a confirmation screen lists recipients, their contacts and selected channels before sending.</li><li><b>Send</b>: SMS via the configured gateway (OVH, Twilio, Free…) and/or email via SMTP. Each message contains a secure personal link.</li><li><b>Track responses</b>: each recipient clicks their link and confirms their estimated arrival time. The <b>half-circle gauge widget</b> on the dashboard shows in real time: called / responded / arrived / missing.</li><li><b>Escalate in waves</b>: if a first wave is insufficient, automatically re-contact non-responders. Each wave requires cell confirmation.</li></ol><h2>Telephony failover</h2><p>During a CYBER crisis, if the PBX/IPBX is impacted, SCRIBE automatically switches the directory display to configured <b>emergency numbers</b>. Staff contacted via the recall link also see these fallback numbers.</p><p>⚠️ Recall sends <b>real SMS to real numbers</b>: always test in an exercise (isolated exercise instances) before real use. SMS/SMTP configuration is in <b>Admin → Notifications</b>.</p>'}},

 {id:'communication', cat:{fr:'Utilisation',en:'Usage',de:'Nutzung',es:'Uso',it:'Utilizzo',nl:'Gebruik'},
  title:{fr:'Chat, messagerie & annuaire',en:'Chat, messaging & directory',de:'Chat, Nachrichten & Verzeichnis',es:'Chat, mensajería y directorio',it:'Chat, messaggistica e rubrica',nl:'Chat, berichten & adresboek'},
  kw:'chat messagerie messaging salon channel inbox message annuaire directory territorial mention',
  body:{fr:'<p>SCRIBE propose trois outils de communication complémentaires :</p><ul><li><b>Chat</b> : des salons en temps réel, locaux (ton établissement) ou territoriaux (partagés entre tous les établissements du GHT), avec mentions @ et pièces jointes. Idéal pour la coordination à chaud.</li><li><b>Messagerie interne</b> : fonctionne comme une boîte mail (Reçus, Envoyés, Brouillons) pour des messages plus formels, y compris inter-établissements.</li><li><b>Annuaire</b> : les contacts utiles de l\'établissement et du territoire.</li></ul><p>⚠️ Les salons territoriaux et le canal de supervision <b>sortent de l\'établissement</b> : n\'y publie jamais de donnée nominative patient.</p>',
       en:'<p>SCRIBE offers three complementary communication tools:</p><ul><li><b>Chat</b>: real-time rooms, local (your establishment) or territorial (shared across all establishments in the group), with @ mentions and attachments. Ideal for live coordination.</li><li><b>Internal messaging</b>: works like a mailbox (Inbox, Sent, Drafts) for more formal messages, including inter-establishment.</li><li><b>Directory</b>: useful contacts for the establishment and the territory.</li></ul><p>⚠️ Territorial rooms and the supervision channel <b>leave the establishment</b>: never post patient-identifying data there.</p>'}},

 {id:'cellule', cat:{fr:'Utilisation',en:'Usage',de:'Nutzung',es:'Uso',it:'Utilizzo',nl:'Gebruik'},
  title:{fr:'Cellule de crise : relève, communiqués, REX',en:'Crisis cell: handover, briefings, AAR',de:'Krisenstab: Übergabe, Briefings, Nachbesprechung',es:'Célula de crisis: relevo, briefings, AAR',it:'Cellula di crisi: consegne, briefing, debriefing',nl:'Crisiscel: overdracht, briefings, evaluatie'},
  kw:'cellule crise releve garde communique rex retour experience archivage zip',
  body:{fr:'<p>Plusieurs outils structurent le travail de la cellule de crise dans la durée :</p><ul><li><b>Relève de garde</b> : transmettre une situation synthétique à l\'équipe suivante, pour une continuité sans perte d\'information sur des crises longues.</li><li><b>Communiqués</b> : rédiger et diffuser des messages officiels (interne, partenaires).</li><li><b>REX (retour d\'expérience)</b> : consigner ce qui a marché et ce qui doit être amélioré, à chaud puis à froid.</li></ul><p>À la fin d\'un événement, l\'<b>archivage</b> génère un ZIP téléchargeable depuis l\'interface, regroupant main courante, décisions et éléments de traçabilité — base du débriefing et des obligations réglementaires.</p>',
       en:'<p>Several tools structure the crisis cell\'s work over time:</p><ul><li><b>Shift handover</b>: pass a concise situation to the next team, for continuity without information loss during long crises.</li><li><b>Briefings</b>: write and distribute official messages (internal, partners).</li><li><b>After-action review (AAR)</b>: record what worked and what should improve, both during and after.</li></ul><p>At the end of an event, <b>archiving</b> generates a downloadable ZIP from the interface, bundling the activity log, decisions and traceability items — the basis for debriefing and regulatory obligations.</p>'}},

 {id:'albert', cat:{fr:'Utilisation',en:'Usage',de:'Nutzung',es:'Uso',it:'Utilizzo',nl:'Gebruik'},
  title:{fr:'Analyse par l\'IA Albert',en:'Albert AI analysis',de:'Albert-KI-Analyse',es:'Análisis con IA Albert',it:'Analisi con IA Albert',nl:'Albert AI-analyse'},
  kw:'albert ia ai intelligence artificielle analyse souverain llm synthese',
  body:{fr:'<p>SCRIBE intègre <b>Albert</b>, le grand modèle de langage souverain de l\'État français, pour assister la cellule sans dépendre d\'un service étranger.</p><p>Albert peut produire des <b>synthèses de situation</b> et de l\'aide à la décision à partir du contexte de crise (incidents, capacité, mobilisation : vagues, manquants, échecs d\'envoi). C\'est une aide, pas un décideur : les analyses restent à valider par la cellule.</p><p>La clé d\'accès Albert se configure côté serveur et n\'est jamais incluse dans les archives publiques, pour des raisons de sécurité.</p>',
       en:'<p>SCRIBE integrates <b>Albert</b>, the French State\'s sovereign large language model, to assist the cell without relying on a foreign service.</p><p>Albert can produce <b>situation summaries</b> and decision support from the crisis context (incidents, capacity, mobilisation: waves, missing people, send failures). It is an aid, not a decision-maker: analyses remain for the cell to validate.</p><p>The Albert access key is configured server-side and is never included in public archives, for security reasons.</p>'}},

 {id:'supervision', cat:{fr:'Coordination',en:'Coordination',de:'Koordination',es:'Coordinación',it:'Coordinamento',nl:'Coördinatie'},
  title:{fr:'Supervision territoriale (collecteur)',en:'Territorial supervision (collector)',de:'Territoriale Aufsicht (Kollektor)',es:'Supervisión territorial (colector)',it:'Supervisione territoriale (collettore)',nl:'Territoriaal toezicht (collector)'},
  kw:'supervision collecteur territoire niveau global statut public carte instances enrolement',
  body:{fr:'<p>La <b>supervision</b> est le poste de pilotage du territoire. Elle agrège, en lecture seule, l\'état remonté par chaque instance :</p><ul><li>Le <b>niveau global</b> par établissement, qui est toujours le pire entre ses incidents et son statut déclaré.</li><li>Les <b>statuts publics</b> (page /status) : opérationnel, perturbé, etc.</li><li>La <b>cartographie</b> du territoire et les transferts inter-établissements.</li><li>La <b>messagerie et le chat</b> territoriaux.</li></ul><p>Une instance n\'apparaît dans la supervision qu\'après <b>enrôlement</b> : elle propose son token, l\'administrateur du collecteur l\'accepte. Aucune donnée nominative patient ne transite par le collecteur.</p>',
       en:'<p><b>Supervision</b> is the territory\'s control desk. It aggregates, read-only, the state reported by each instance:</p><ul><li>The <b>global level</b> per establishment, always the worst of its incidents and its declared status.</li><li><b>Public statuses</b> (/status page): operational, disrupted, etc.</li><li>The territory <b>map</b> and inter-establishment transfers.</li><li>Territorial <b>messaging and chat</b>.</li></ul><p>An instance only appears in supervision after <b>enrolment</b>: it offers its token, the collector administrator accepts it. No patient-identifying data passes through the collector.</p>'}},

 {id:'exercices', cat:{fr:'Coordination',en:'Coordination',de:'Koordination',es:'Coordinación',it:'Coordinamento',nl:'Coördinatie'},
  title:{fr:'Mode exercice',en:'Exercise mode',de:'Übungsmodus',es:'Modo ejercicio',it:'Modalità esercitazione',nl:'Oefenmodus'},
  kw:'exercice exercise scenario stimuli animateur entrainement formation isole',
  body:{fr:'<p>Le <b>mode exercice</b> permet de s\'entraîner sans toucher à la production. Des instances dédiées tournent sur des bases isolées, pilotées par un poste <b>animateur</b> qui injecte des stimuli (événements scénarisés) vers les établissements joueurs.</p><p>Un scénario décrit les acteurs, la chronologie des stimuli et les décisions attendues. C\'est l\'outil idéal pour valider les procédures, former les équipes et tester des fonctions sensibles (comme le rappel du personnel) avant un usage réel.</p>',
       en:'<p><b>Exercise mode</b> lets you train without touching production. Dedicated instances run on isolated databases, driven by an <b>animator</b> station that injects stimuli (scripted events) toward the playing establishments.</p><p>A scenario describes the actors, the timeline of stimuli, and the expected decisions. It is the ideal tool to validate procedures, train teams, and test sensitive functions (such as staff recall) before real use.</p>'}},

 {id:'sauvegarde', cat:{fr:'Administration',en:'Administration',de:'Administration',es:'Administración',it:'Amministrazione',nl:'Beheer'},
  title:{fr:'Sauvegarde & restauration',en:'Backup & restore',de:'Sicherung & Wiederherstellung',es:'Copia de seguridad y restauración',it:'Backup e ripristino',nl:'Back-up & herstel'},
  kw:'sauvegarde backup restauration restore zip export import chiffre password aes ght',
  body:{fr:'<p>Depuis la supervision, tu peux <b>exporter la configuration du territoire</b> (établissements, identifiants admin, structure UF, capacité) dans une archive <b>.zip</b>.</p><p>L\'archive peut être <b>chiffrée par mot de passe</b> (AES) : ce mot de passe sera exigé pour la restauration. C\'est la sauvegarde de référence à conserver hors-ligne.</p><p><b>Précautions</b> : l\'archive contient des identifiants — garde-la en lieu sûr, ne la publie jamais (ni sur un dépôt public, ni vers le collecteur). Teste toujours un cycle sauvegarde → restauration sur une instance vierge avant de t\'y fier.</p>',
       en:'<p>From supervision, you can <b>export the territory configuration</b> (establishments, admin credentials, unit structure, capacity) into a <b>.zip</b> archive.</p><p>The archive can be <b>password-encrypted</b> (AES): that password will be required to restore. This is the reference backup to keep offline.</p><p><b>Precautions</b>: the archive contains credentials — keep it safe, never publish it (not on a public repository, nor to the collector). Always test a backup → restore cycle on a clean instance before relying on it.</p>'}},

 {id:'securite', cat:{fr:'Administration',en:'Administration',de:'Administration',es:'Administración',it:'Amministrazione',nl:'Beheer'},
  title:{fr:'Sécurité & conformité',en:'Security & compliance',de:'Sicherheit & Compliance',es:'Seguridad y cumplimiento',it:'Sicurezza e conformità',nl:'Beveiliging & naleving'},
  kw:'securite security bcrypt rate limit cors headers rgpd hds donnees patient secret key',
  body:{fr:'<p>SCRIBE applique une base de sécurité robuste :</p><ul><li><b>Mots de passe hachés</b> (bcrypt), avec limitation des tentatives de connexion.</li><li><b>En-têtes HTTP de sécurité</b> et <b>CORS restreint</b> aux origines déclarées.</li><li><b>Clé secrète</b> stockée en variable d\'environnement, jamais en clair dans le code.</li></ul><p><b>Conformité données de santé</b> : les données nominatives patients ne quittent jamais l\'établissement vers la supervision (HDS / RGPD). Avant toute diffusion publique du code, on retire systématiquement les références internes (sigles, noms de sites, URLs, clés d\'API) ; les configurations de démonstration ne contiennent jamais d\'identifiants réels.</p>',
       en:'<p>SCRIBE enforces a robust security baseline:</p><ul><li><b>Hashed passwords</b> (bcrypt), with login rate limiting.</li><li><b>HTTP security headers</b> and <b>CORS restricted</b> to declared origins.</li><li><b>Secret key</b> stored in an environment variable, never in plaintext in the code.</li></ul><p><b>Health-data compliance</b>: patient-identifying data never leaves the establishment toward supervision (HDS / GDPR). Before any public code release, internal references (codes, site names, URLs, API keys) are systematically removed; demo configurations never contain real credentials.</p>'}},
 {id:'bluefiles', cat:{fr:'Utilisation',en:'Usage',de:'Nutzung',es:'Uso',it:'Utilizzo',nl:'Gebruik'},
  title:{fr:'Envoi sécurisé BlueFiles',en:'Secure BlueFiles transfer',de:'Sicherer BlueFiles-Versand',es:'Envío seguro BlueFiles',it:'Invio sicuro BlueFiles',nl:'Veilige BlueFiles-verzending'},
  kw:'bluefiles securise hds chiffrement bout en bout transfert fichier patient donnees sensibles supervision messagerie envoi configuration admin',
  body:{fr:'<p>SCRIBE intègre nativement <b>BlueFiles</b> (Forecomm), service d\'envoi sécurisé chiffré bout-en-bout, hébergé HDS. Il est disponible à deux niveaux :</p><ul><li><b>Depuis une instance établissement</b> : dans la messagerie (onglet MESSAGERIE), le bouton <b>« Envoi sécurisé par BlueFiles »</b> ouvre un panneau dédié (fichiers par glisser-déposer, destinataires email, commentaire). L\'envoi est un <b>transfert indépendant</b> — le contenu ne transite jamais par SCRIBE, il est chiffré côté BlueFiles.</li><li><b>Depuis la supervision</b> : dans l\'onglet MESSAGES, le bouton <b>« 🔒 Envoi sécurisé BlueFiles »</b> apparaît si le plugin est activé. Il permet d\'envoyer des fichiers sensibles depuis le poste de supervision territoriale.</li></ul><h2>Configuration (établissement)</h2><p>Dans <b>Admin → Plugins → BlueFiles</b> : renseigne le login, le mot de passe (chiffré Fernet au repos), et le serveur (défaut : <code>api.bluefiles.com</code>). Un bouton <b>Tester</b> vérifie la connexion sans envoyer de fichier.</p><h2>Configuration (supervision)</h2><p>Dans <b>Admin → Transfert sécurisé (BlueFiles)</b> : mêmes champs (login, mot de passe, serveur), case « Activer ». Bouton Tester disponible. Après activation, le bouton apparaît automatiquement dans la messagerie de supervision.</p><h2>Traçabilité & conformité</h2><ul><li>Chaque envoi est <b>tracé</b> dans SCRIBE (journal de bord interne).</li><li>Le contenu des fichiers ne touche jamais les serveurs SCRIBE — il est chiffré bout-en-bout chez BlueFiles.</li><li>Aucune copie locale des fichiers envoyés n\'est conservée sur SCRIBE.</li><li>Compatible HDS — idéal pour les données de santé sensibles (bilans patients, listes nominatives) à partager entre établissements en situation de crise.</li></ul><p>⚠️ Ce n\'est pas une pièce jointe à un message SCRIBE : c\'est un <b>envoi sécurisé indépendant</b>. Le destinataire reçoit un email BlueFiles avec un lien de téléchargement sécurisé.</p>',
       en:'<p>SCRIBE natively integrates <b>BlueFiles</b> (Forecomm), an end-to-end encrypted secure file transfer service with HDS hosting. It is available at two levels:</p><ul><li><b>From an establishment instance</b>: in the messaging module (MESSAGING tab), the <b>"Secure send via BlueFiles"</b> button opens a dedicated panel (drag-and-drop files, email recipients, comment). The send is an <b>independent transfer</b> — content never transits through SCRIBE, it is encrypted on the BlueFiles side.</li><li><b>From supervision</b>: in the MESSAGES tab, the <b>"🔒 Secure BlueFiles send"</b> button appears if the plugin is enabled. It allows sending sensitive files from the territorial supervision desk.</li></ul><h2>Setup (establishment)</h2><p>In <b>Admin → Plugins → BlueFiles</b>: enter the login, password (Fernet-encrypted at rest), and server (default: <code>api.bluefiles.com</code>). A <b>Test</b> button verifies the connection without sending a file.</p><h2>Setup (supervision)</h2><p>In <b>Admin → Secure transfer (BlueFiles)</b>: same fields (login, password, server), with an Enable checkbox. Test button available. Once enabled, the button appears automatically in the supervision messaging bar.</p><h2>Traceability & compliance</h2><ul><li>Every transfer is <b>logged</b> in SCRIBE (internal audit trail).</li><li>File content never touches SCRIBE servers — it is end-to-end encrypted at BlueFiles.</li><li>No local copy of sent files is retained on SCRIBE.</li><li>HDS-compatible — ideal for sensitive health data (patient summaries, staff lists) to share between establishments during a crisis.</li></ul><p>⚠️ This is not a SCRIBE message attachment: it is an <b>independent secure transfer</b>. The recipient receives a BlueFiles email with a secure download link.</p>'}},
];;
function openHelpCenter(){ var o=document.getElementById('help-overlay'); if(!o)return; o.classList.add('open'); var L=helpLang2(); var s=document.getElementById('help-search'); if(s){s.placeholder=L==='fr'?'Rechercher dans l\'aide…':'Search help…'; s.value='';} var t=document.getElementById('help-title'); if(t)t.textContent=L==='fr'?'Centre d\'aide':'Help center'; helpRender(); if(s)setTimeout(function(){s.focus();},50); }
function closeHelpCenter(){ var o=document.getElementById('help-overlay'); if(o)o.classList.remove('open'); }
function helpRender(){
  var listEl=document.getElementById('help-list'), artEl=document.getElementById('help-article'), s=document.getElementById('help-search');
  if(!listEl)return; artEl.style.display='none'; listEl.style.display='block';
  var L=helpLang2(), q=(s&&s.value||'').trim().toLowerCase();
  var t=document.getElementById('help-title'); if(t)t.textContent=L==='fr'?'Centre d\'aide':'Help center';
  if(s)s.placeholder=L==='fr'?'Rechercher dans l\'aide…':'Search help…';
  var hits=HELP_ARTICLES.filter(function(a){ if(!q)return true; var hay=((a.title[L]||a.title.en)+' '+(a.body[L]||a.body.en)+' '+a.kw+' '+(a.cat[L]||a.cat.en)).toLowerCase(); return hay.indexOf(q)>=0; });
  if(!hits.length){ listEl.innerHTML='<div class="help-empty">'+(L==='fr'?'Aucun résultat.':'No results.')+'</div>'; return; }
  listEl.innerHTML=hits.map(function(a){ return '<div class="help-item" onclick="helpShowArticle(\''+a.id+'\')"><div class="help-item-cat">'+(a.cat[L]||a.cat.en)+'</div><div class="help-item-title">'+(a.title[L]||a.title.en)+'</div></div>'; }).join('');
}
function helpShowArticle(id){
  var a=HELP_ARTICLES.find(function(x){return x.id===id;}); if(!a)return;
  var L=helpLang2(), listEl=document.getElementById('help-list'), artEl=document.getElementById('help-article');
  listEl.style.display='none'; artEl.style.display='block';
  artEl.innerHTML='<button id="help-back" onclick="helpRender()">← '+(L==='fr'?'Retour':'Back')+'</button><h2>'+(a.title[L]||a.title.en)+'</h2>'+(a.body[L]||a.body.en);
}


</script>
</body>
</html>"""


@app.get("/", response_class=HTMLResponse)
async def dashboard():
    html = DASHBOARD_HTML.replace("PLACEHOLDER_ADMIN_TOKEN", ADMIN_TOKEN)
    return HTMLResponse(content=html, headers={
        "Cache-Control": "no-store, no-cache, must-revalidate",
        "Pragma": "no-cache", "Expires": "0"
    })


@app.get("/static/logo-scribe.png")
async def serve_logo():
    from fastapi.responses import FileResponse
    import os
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "app", "static", "logo-scribe.png")
    if os.path.exists(p):
        return FileResponse(p, media_type="image/png")
    raise HTTPException(404, "Logo non trouvé")

@app.get("/favicon.ico")
@app.get("/static/favicon.svg")
async def serve_favicon():
    """Favicon SVG SCRIBE — identique aux instances SCRIBE."""
    from fastapi.responses import FileResponse, Response
    import os
    # Chercher favicon.svg (identique aux instances SCRIBE)
    svg = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "app", "static", "favicon.svg")
    if os.path.exists(svg):
        return FileResponse(svg, media_type="image/svg+xml")
    # Fallback : logo PNG
    png = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "app", "static", "logo-scribe.png")
    if os.path.exists(png):
        return FileResponse(png, media_type="image/png")
    return Response(status_code=404)

CHAT_COLL_HTML = '<!DOCTYPE html>\n<html lang="fr">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width,initial-scale=1">\n<title>SCRIBE Chat</title>\n<style>\n:root{\n  --blue:#003189;--red:#e1000f;--bg:#f0f2f5;--s:#fff;--s2:#f1f5f9;\n  --bd:#e2e8f0;--tx:#0f172a;--mu:#64748b;--mo:monospace;\n  --online:#22c55e;--mention:#f97316;\n}\n*{box-sizing:border-box;margin:0;padding:0}\nhtml,body{height:100%;overflow:hidden;font-family:system-ui,sans-serif;background:var(--bg);color:var(--tx)}\n\n/* ── Layout ── */\n#chat-root{display:flex;height:100%;overflow:hidden}\n\n/* ── Colonne salons ── */\n#col-salons{\n  width:240px;flex-shrink:0;background:#1e2a3a;color:#e2e8f0;\n  display:flex;flex-direction:column;overflow:hidden;\n  transition:width .2s;\n}\n#col-salons.collapsed{width:0;overflow:hidden}\n#salon-header{\n  padding:14px 12px 10px;border-bottom:1px solid rgba(255,255,255,.1);flex-shrink:0;\n  display:flex;align-items:center;justify-content:space-between;\n}\n#salon-header h2{font-family:var(--mo);font-size:11px;font-weight:700;letter-spacing:1px;opacity:.7}\n#btn-new-salon{\n  width:24px;height:24px;border-radius:50%;background:rgba(255,255,255,.15);\n  border:none;color:#fff;cursor:pointer;font-size:16px;line-height:1;\n  display:flex;align-items:center;justify-content:center;\n}\n#btn-new-salon:hover{background:rgba(255,255,255,.25)}\n#salon-list{flex:1;overflow-y:auto;padding:8px 0}\n.salon-section-title{\n  font-family:var(--mo);font-size:8px;letter-spacing:2px;color:rgba(255,255,255,.4);\n  padding:10px 12px 4px;text-transform:uppercase;\n}\n.salon-item{\n  display:flex;align-items:center;gap:8px;padding:7px 12px;cursor:pointer;\n  border-radius:5px;margin:1px 6px;transition:background .1s;\n}\n.salon-item:hover{background:rgba(255,255,255,.08)}\n.salon-item.active{background:rgba(255,255,255,.15)}\n.salon-icon{font-size:14px;flex-shrink:0;width:20px;text-align:center}\n.salon-name{font-size:12px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}\n.salon-badge{\n  background:#e1000f;color:#fff;font-family:var(--mo);font-size:9px;\n  padding:1px 5px;border-radius:10px;min-width:18px;text-align:center;\n}\n.salon-ght-dot{\n  width:6px;height:6px;border-radius:50%;background:#7c3aed;flex-shrink:0;\n}\n\n/* ── Colonne présence ── */\n#col-presence{\n  width:200px;flex-shrink:0;background:var(--s);border-left:1px solid var(--bd);\n  display:flex;flex-direction:column;overflow:hidden;\n}\n#col-presence.hidden{display:none}\n#presence-header{\n  padding:12px;border-bottom:1px solid var(--bd);font-family:var(--mo);\n  font-size:9px;font-weight:700;letter-spacing:1px;color:var(--mu);\n  display:flex;align-items:center;justify-content:space-between;\n}\n#presence-list{flex:1;overflow-y:auto;padding:8px 0}\n.presence-etab{\n  font-family:var(--mo);font-size:8px;letter-spacing:1px;color:var(--mu);\n  padding:8px 12px 4px;text-transform:uppercase;\n}\n.presence-user{\n  display:flex;align-items:center;gap:8px;padding:5px 12px;font-size:11px;\n}\n.presence-dot{\n  width:8px;height:8px;border-radius:50%;background:var(--online);flex-shrink:0;\n}\n.presence-name{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1}\n\n/* ── Zone centrale ── */\n#col-main{flex:1;display:flex;flex-direction:column;overflow:hidden;min-width:0}\n\n#chat-topbar{\n  padding:10px 14px;border-bottom:1px solid var(--bd);background:var(--s);\n  display:flex;align-items:center;gap:10px;flex-shrink:0;\n}\n#btn-toggle-salons{\n  background:none;border:none;cursor:pointer;font-size:18px;color:var(--mu);padding:2px 6px;\n  border-radius:4px;display:none;\n}\n#chat-salon-title{font-size:14px;font-weight:700;flex:1}\n#chat-salon-desc{font-family:var(--mo);font-size:9px;color:var(--mu)}\n.topbar-btn{\n  font-family:var(--mo);font-size:9px;padding:4px 10px;border-radius:4px;\n  border:1px solid var(--bd);background:var(--s2);cursor:pointer;color:var(--mu);\n  display:flex;align-items:center;gap:5px;\n}\n.topbar-btn:hover{background:var(--s);color:var(--tx)}\n\n/* ── Fil messages ── */\n#msg-feed{\n  flex:1;overflow-y:auto;padding:16px;display:flex;flex-direction:column;gap:4px;\n  background:var(--bg);\n}\n.msg-group{margin-bottom:12px}\n.msg-author-line{\n  display:flex;align-items:baseline;gap:8px;margin-bottom:3px;padding-left:4px;\n}\n.msg-author{font-size:12px;font-weight:700;color:var(--blue)}\n.msg-author.ght{color:#7c3aed}\n.msg-time{font-family:var(--mo);font-size:9px;color:var(--mu)}\n.msg-bubble{\n  background:var(--s);border-radius:0 8px 8px 8px;padding:8px 12px;\n  font-size:13px;line-height:1.5;max-width:100%;word-break:break-word;\n  box-shadow:0 1px 2px rgba(0,0,0,.06);\n}\n.msg-bubble.own{background:#dbeafe;border-radius:8px 0 8px 8px}\n.msg-bubble.ght{background:#ede9fe}\n.msg-bubble.supprime{opacity:.5;font-style:italic}\n.msg-mention{color:var(--mention);font-weight:700}\n.msg-reply-preview{\n  background:rgba(0,0,0,.05);border-left:3px solid var(--mu);\n  padding:4px 8px;border-radius:0 4px 4px 0;margin-bottom:6px;\n  font-size:11px;color:var(--mu);cursor:pointer;\n}\n.msg-reply-preview strong{color:var(--tx);display:block;font-size:10px;margin-bottom:2px}\n.msg-pj{\n  display:flex;align-items:center;gap:8px;margin-top:6px;\n  padding:6px 10px;background:var(--s2);border-radius:6px;\n  font-size:11px;cursor:pointer;\n}\n.msg-pj:hover{background:var(--bd)}\n.msg-pj-icon{font-size:16px}\n.msg-actions{\n  display:none;gap:4px;margin-left:4px;\n}\n.msg-group:hover .msg-actions{display:flex}\n.msg-action-btn{\n  font-size:11px;padding:2px 6px;background:var(--s2);border:1px solid var(--bd);\n  border-radius:4px;cursor:pointer;color:var(--mu);\n}\n.msg-action-btn:hover{background:var(--s);color:var(--tx)}\n.msg-row{display:flex;align-items:flex-start;gap:6px}\n.msg-row .msg-bubble{flex:1}\n\n/* ── Barre saisie ── */\n#input-zone{\n  padding:12px 14px;border-top:1px solid var(--bd);background:var(--s);flex-shrink:0;\n}\n#reply-bar{\n  display:none;background:var(--s2);border-left:3px solid var(--blue);\n  padding:5px 10px;border-radius:0 4px 4px 0;margin-bottom:8px;\n  font-size:11px;color:var(--mu);\n}\n#reply-bar strong{color:var(--tx);font-size:10px}\n#reply-cancel{\n  float:right;background:none;border:none;cursor:pointer;color:var(--mu);font-size:14px;\n}\n#autocomplete-box{\n  display:none;background:var(--s);border:1px solid var(--bd);border-radius:6px;\n  box-shadow:0 4px 16px rgba(0,0,0,.1);max-height:180px;overflow-y:auto;\n  margin-bottom:6px;\n}\n.autocomplete-item{\n  padding:7px 12px;cursor:pointer;font-size:12px;display:flex;align-items:center;gap:8px;\n}\n.autocomplete-item:hover{background:var(--s2)}\n.autocomplete-item.selected{background:#dbeafe}\n#input-row{display:flex;gap:8px;align-items:flex-end}\n#msg-input{\n  flex:1;font-family:var(--mono,monospace);font-size:13px;padding:8px 12px;\n  border:1px solid var(--bd);border-radius:20px;resize:none;max-height:120px;\n  background:var(--s2);color:var(--tx);outline:none;line-height:1.4;\n}\n#msg-input:focus{border-color:var(--blue);background:var(--s)}\n#btn-pj{\n  width:36px;height:36px;border-radius:50%;background:var(--s2);border:1px solid var(--bd);\n  cursor:pointer;font-size:16px;display:flex;align-items:center;justify-content:center;\n  flex-shrink:0;\n}\n#btn-pj:hover{background:var(--s)}\n#btn-send{\n  width:36px;height:36px;border-radius:50%;background:var(--blue);border:none;\n  color:#fff;cursor:pointer;font-size:16px;display:flex;align-items:center;\n  justify-content:center;flex-shrink:0;\n}\n#btn-send:hover{opacity:.85}\n#pj-input{display:none}\n\n/* ── Modal nouveau salon ── */\n#modal-salon{\n  display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9999;\n  align-items:center;justify-content:center;\n}\n.modal-box{\n  background:var(--s);border-radius:10px;width:420px;max-width:95vw;overflow:hidden;\n}\n.modal-hdr{\n  padding:16px 20px;border-bottom:1px solid var(--bd);display:flex;\n  align-items:center;justify-content:space-between;\n}\n.modal-hdr h3{font-family:var(--mo);font-size:11px;font-weight:700}\n.modal-close{background:none;border:none;cursor:pointer;font-size:18px;color:var(--mu)}\n.modal-body{padding:20px;display:flex;flex-direction:column;gap:12px}\n.mff{display:flex;flex-direction:column;gap:4px}\n.mff label{font-family:var(--mo);font-size:9px;color:var(--mu);text-transform:uppercase}\n.mff input,.mff select,.mff textarea{\n  font-family:var(--mo);font-size:11px;padding:7px 10px;\n  background:var(--s2);border:1px solid var(--bd);border-radius:5px;color:var(--tx);\n}\n.color-row{display:flex;gap:8px;flex-wrap:wrap;margin-top:4px}\n.color-dot{\n  width:26px;height:26px;border-radius:50%;cursor:pointer;\n  border:3px solid transparent;transition:border-color .1s;\n}\n.color-dot.sel{border-color:var(--tx)}\n.modal-footer{\n  padding:12px 20px;border-top:1px solid var(--bd);display:flex;justify-content:flex-end;gap:8px;\n}\n.btn-primary{\n  font-family:var(--mo);font-size:10px;padding:7px 16px;background:var(--blue);\n  color:#fff;border:none;border-radius:5px;cursor:pointer;font-weight:700;\n}\n.btn-secondary{\n  font-family:var(--mo);font-size:10px;padding:7px 16px;background:var(--s2);\n  color:var(--mu);border:1px solid var(--bd);border-radius:5px;cursor:pointer;\n}\n\n.msg-pj-img img {\n  max-width:300px;max-height:200px;border-radius:6px;\n  cursor:pointer;display:block;margin-top:4px;\n}\n/* ── Toast ── */\n#toast{\n  position:fixed;bottom:16px;right:16px;font-family:var(--mo);font-size:11px;\n  padding:8px 14px;border-radius:5px;display:none;z-index:99999;\n}\n#toast.ok{background:#dcfce7;color:#15803d;border:1px solid #86efac}\n#toast.err{background:#fee2e2;color:#dc2626;border:1px solid #fca5a5}\n\n/* ── Responsive mobile ── */\n@media(max-width:640px){\n  #col-presence{display:none}\n  #col-salons{position:absolute;z-index:100;height:100%;left:-240px;transition:left .2s}\n  #col-salons.mobile-open{left:0}\n  #btn-toggle-salons{display:flex !important}\n  #chat-root{position:relative}\n}\n@media(max-width:900px){\n  #col-presence{width:160px}\n}\n</style>\n</head>\n<body>\n<div id="chat-root">\n\n  <!-- Salons -->\n  <div id="col-salons">\n    <div id="salon-header">\n      <h2>SALONS</h2>\n      <button id="btn-new-salon" onclick="chat_openNewSalon()" title="Nouveau salon">+</button>\n    </div>\n    <div id="salon-list">\n      <div style="font-family:monospace;font-size:10px;color:rgba(255,255,255,.4);padding:20px;text-align:center">Chargement...</div>\n    </div>\n  </div>\n\n  <!-- Zone principale -->\n  <div id="col-main">\n    <div id="chat-topbar">\n      <button id="btn-toggle-salons" onclick="chat_toggleSalons()" title="Salons">&#9776;</button>\n      <div style="flex:1">\n        <div id="chat-salon-title">Sélectionnez un salon</div>\n        <div id="chat-salon-desc"></div>\n      </div>\n      <button class="topbar-btn" onclick="chat_openPresence()" title="Personnes connectées">\n        <span id="presence-count">0</span> en ligne\n      </button>\n      <button class="topbar-btn" onclick="chat_popout()" title="Ouvrir dans une fenêtre">&#x26F6;</button>\n    </div>\n\n    <div id="msg-feed">\n      <div style="font-family:monospace;font-size:10px;color:var(--mu);text-align:center;padding:40px">\n        Sélectionnez un salon pour commencer\n      </div>\n    </div>\n\n    <div id="input-zone">\n      <div id="pending-pjs" style="display:none;flex-wrap:wrap;gap:6px;padding:6px 0;margin-bottom:6px;overflow-x:auto"></div>\n      <div id="reply-bar">\n        <button id="reply-cancel" onclick="chat_cancelReply()">&#x2715;</button>\n        <strong id="reply-author"></strong>\n        <span id="reply-preview"></span>\n      </div>\n      <div id="autocomplete-box"></div>\n      <div style="display:flex;gap:4px;margin-bottom:4px;flex-wrap:wrap;align-items:center">\n        <button onclick="chat_toggleEmoji()" title="Emojis" id="btn-emoji" style="font-size:14px;padding:2px 7px;background:var(--s2);border:1px solid var(--bd);border-radius:4px;cursor:pointer">&#128512;</button>\n        <button onclick="chat_fmt(\'**\',\'**\')" title="Gras" style="font-family:var(--mono);font-size:10px;font-weight:700;padding:2px 7px;background:var(--s2);border:1px solid var(--bd);border-radius:4px;cursor:pointer">B</button>\n        <button onclick="chat_fmt(\'_\',\'_\')" title="Italique" style="font-family:var(--mono);font-size:10px;font-style:italic;padding:2px 7px;background:var(--s2);border:1px solid var(--bd);border-radius:4px;cursor:pointer">I</button>\n        <button onclick="chat_fmt(\'[rouge]\',\'[/rouge]\')" title="Rouge" style="font-size:10px;padding:2px 7px;background:rgba(239,68,68,.15);border:1px solid #ef4444;border-radius:4px;cursor:pointer;color:#ef4444;font-weight:700">&#9679;</button>\n        <button onclick="chat_fmt(\'[vert]\',\'[/vert]\')" title="Vert" style="font-size:10px;padding:2px 7px;background:rgba(34,197,94,.15);border:1px solid #22c55e;border-radius:4px;cursor:pointer;color:#22c55e;font-weight:700">&#9679;</button>\n        <button onclick="chat_fmt(\'[orange]\',\'[/orange]\')" title="Orange" style="font-size:10px;padding:2px 7px;background:rgba(249,115,22,.15);border:1px solid #f97316;border-radius:4px;cursor:pointer;color:#f97316;font-weight:700">&#9679;</button>\n        <button onclick="chat_fmt(\'[code]\',\'[/code]\')" title="Code [code]...[/code]" style="font-family:monospace;font-size:10px;padding:2px 7px;background:var(--s2);border:1px solid var(--bd);border-radius:4px;cursor:pointer">&lt;&gt;</button>\n        <button onclick="chat_fmt(\'---URGENT---\',\'---\')" title="Urgent" style="font-family:var(--mono);font-size:9px;padding:2px 7px;background:rgba(239,68,68,.15);border:1px solid #ef4444;border-radius:4px;cursor:pointer;color:#ef4444;font-weight:700">&#128680; URGENT</button>\n      </div>\n      <div id="emoji-picker" style="display:none;flex-wrap:wrap;gap:4px;padding:8px;background:var(--s);border:1px solid var(--bd);border-radius:8px;margin-bottom:6px;max-height:120px;overflow-y:auto;font-size:18px"><span onclick="chat_insertEmoji(\'😀\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">😀</span><span onclick="chat_insertEmoji(\'😊\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">😊</span><span onclick="chat_insertEmoji(\'😅\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">😅</span><span onclick="chat_insertEmoji(\'😂\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">😂</span><span onclick="chat_insertEmoji(\'🤣\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🤣</span><span onclick="chat_insertEmoji(\'😍\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">😍</span><span onclick="chat_insertEmoji(\'🥰\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🥰</span><span onclick="chat_insertEmoji(\'😎\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">😎</span><span onclick="chat_insertEmoji(\'🤔\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🤔</span><span onclick="chat_insertEmoji(\'😮\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">😮</span><span onclick="chat_insertEmoji(\'😱\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">😱</span><span onclick="chat_insertEmoji(\'😭\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">😭</span><span onclick="chat_insertEmoji(\'😡\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">😡</span><span onclick="chat_insertEmoji(\'🤒\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🤒</span><span onclick="chat_insertEmoji(\'👍\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">👍</span><span onclick="chat_insertEmoji(\'👎\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">👎</span><span onclick="chat_insertEmoji(\'👋\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">👋</span><span onclick="chat_insertEmoji(\'🙏\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🙏</span><span onclick="chat_insertEmoji(\'💪\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">💪</span><span onclick="chat_insertEmoji(\'✅\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">✅</span><span onclick="chat_insertEmoji(\'❌\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">❌</span><span onclick="chat_insertEmoji(\'⚠️\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">⚠️</span><span onclick="chat_insertEmoji(\'🚨\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🚨</span><span onclick="chat_insertEmoji(\'🏥\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🏥</span><span onclick="chat_insertEmoji(\'🚑\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🚑</span><span onclick="chat_insertEmoji(\'💉\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">💉</span><span onclick="chat_insertEmoji(\'🩺\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🩺</span><span onclick="chat_insertEmoji(\'📋\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">📋</span><span onclick="chat_insertEmoji(\'📞\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">📞</span><span onclick="chat_insertEmoji(\'🔴\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🔴</span><span onclick="chat_insertEmoji(\'🟠\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🟠</span><span onclick="chat_insertEmoji(\'🟡\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🟡</span><span onclick="chat_insertEmoji(\'🟢\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🟢</span><span onclick="chat_insertEmoji(\'🔵\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🔵</span><span onclick="chat_insertEmoji(\'⭐\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">⭐</span><span onclick="chat_insertEmoji(\'🔥\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🔥</span><span onclick="chat_insertEmoji(\'💡\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">💡</span><span onclick="chat_insertEmoji(\'📢\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">📢</span><span onclick="chat_insertEmoji(\'🔔\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🔔</span><span onclick="chat_insertEmoji(\'🕐\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">🕐</span><span onclick="chat_insertEmoji(\'📍\')" style="cursor:pointer;padding:2px 3px;border-radius:3px" onmouseover="this.style.background=\'var(--s2)\'" onmouseout="this.style.background=\'\'">📍</span></div>\n      <div id="input-row">\n        <textarea id="msg-input" rows="1" placeholder="Message... (@mention)"\n          oninput="chat_onInput(this)" onkeydown="chat_onKeydown(event)"></textarea>\n        <input type="file" id="pj-input" multiple accept=".pdf,.doc,.docx,.xls,.xlsx,.ppt,.pptx,.odt,.ods,.odp,.txt,.csv,.png,.jpg,.jpeg,.gif,.webp" onchange="chat_uploadPJ(this)">\n        <button id="btn-pj" onclick="document.getElementById(\'pj-input\').click()" title="Piece jointe">&#128206;</button>\n        <button id="btn-send" onclick="chat_send()" title="Envoyer">&#10148;</button>\n      </div>\n    </div>\n  </div>\n\n  <!-- Présence -->\n  <div id="col-presence">\n    <div id="presence-header">\n      CONNECTES\n      <button style="background:none;border:none;cursor:pointer;font-size:14px;color:var(--mu)" onclick="chat_closePresence()">&#x2715;</button>\n    </div>\n    <div id="presence-list"></div>\n  </div>\n\n</div>\n\n<!-- Modal nouveau salon -->\n<div id="modal-salon">\n  <div class="modal-box">\n    <div class="modal-hdr">\n      <h3>NOUVEAU SALON</h3>\n      <button class="modal-close" onclick="chat_closeNewSalon()">&#x2715;</button>\n    </div>\n    <div class="modal-body">\n      <div class="mff">\n        <label>Nom du salon *</label>\n        <input type="text" id="ns-nom" placeholder="ex: pharmacie-urgence">\n      </div>\n      <div class="mff">\n        <label>Description</label>\n        <input type="text" id="ns-desc" placeholder="À quoi sert ce salon ?">\n      </div>\n      <div class="mff">\n        <label>Type</label>\n        <select id="ns-type">\n          <option value="local">Local (cet établissement uniquement)</option>\n          <option value="territorial">Territorial (visible par tous les GHTs)</option>\n        </select>\n      </div>\n      <div class="mff">\n        <label>Couleur</label>\n        <div class="color-row" id="color-row">\n          <div class="color-dot sel" data-c="#003189" style="background:#003189" onclick="chat_selColor(this)"></div>\n          <div class="color-dot" data-c="#16a34a" style="background:#16a34a" onclick="chat_selColor(this)"></div>\n          <div class="color-dot" data-c="#d97706" style="background:#d97706" onclick="chat_selColor(this)"></div>\n          <div class="color-dot" data-c="#7c3aed" style="background:#7c3aed" onclick="chat_selColor(this)"></div>\n          <div class="color-dot" data-c="#dc2626" style="background:#dc2626" onclick="chat_selColor(this)"></div>\n          <div class="color-dot" data-c="#0891b2" style="background:#0891b2" onclick="chat_selColor(this)"></div>\n        </div>\n        <input type="hidden" id="ns-color" value="#003189">\n      </div>\n      <div class="mff">\n        <label>Icone (emoji)</label>\n        <input type="text" id="ns-icon" value="💬" maxlength="2" style="width:60px">\n      </div>\n    </div>\n    <div class="modal-footer">\n      <button class="btn-secondary" onclick="chat_closeNewSalon()">Annuler</button>\n      <button class="btn-primary" onclick="chat_createSalon()">Créer le salon</button>\n    </div>\n  </div>\n</div>\n\n<div id="toast"></div>\n\n<script>\nvar _tok = function() {\n  try {\n    var params = new URLSearchParams(window.location.search);\n    var t = params.get("token");\n    if (t) return t;\n    try { t = window.parent.localStorage.getItem("scribe_token"); if (t) return t; } catch(e2) {}\n    return localStorage.getItem("scribe_token") || "";\n  } catch(e) { return ""; }\n};\nvar _api = function(url, opts) {\n  opts = opts || {};\n  opts.headers = opts.headers || {};\n  opts.headers["Authorization"] = "Bearer " + _tok();\n  if (opts.body && !opts.headers["Content-Type"]) {\n    opts.headers["Content-Type"] = "application/json";\n  }\n  return fetch(url, opts);\n};\n\nvar _toast = function(msg, type) {\n  var el = document.getElementById("toast");\n  el.textContent = msg;\n  el.className = type || "ok";\n  el.style.display = "block";\n  clearTimeout(el._t);\n  el._t = setTimeout(function() { el.style.display = "none"; }, 3000);\n};\n\nvar _esc = function(s) {\n  return String(s || "").replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");\n};\n\n// ── État global ──────────────────────────────────────────────────────────────\nvar _salons = [];\nvar _pjMap = {};     // id → nom fichier\nvar _msgMap = {};    // msgKey → message objet\nvar _pendingPJs = []; // [{id, nom, taille, ext, dataUrl}] — PJs pré-uploadées\nvar _fedBase = "";      // URL base du collecteur\nvar _fedToken = "";     // token fédération\nvar _chatSyncIds = {};  // salon_nom → dernier id collecteur vu\nvar _syncReady = false; // collecteur configuré\nvar _mysigle = "";   // sigle de l\'instance locale\nvar _currentSalon = null;\nvar _messages = [];\nvar _presence = {};\nvar _replyTo = null;\nvar _autocompleteUsers = [];\nvar _autocompleteIdx = -1;\nvar _lastMsgId = 0;\nvar _pollTimer = null;\nvar _pingTimer = null;\nvar _newColor = "#003189";\nvar _myUserId = null;\n\n// ── Init ─────────────────────────────────────────────────────────────────────\nfunction chat_init() {\n  chat_loadPresence();\n  chat_loadSalons();\n  _pingTimer = setInterval(chat_ping, 30000);\n  chat_ping();\n  // Charger la config fédération pour la sync chat inter-GHT\n  _api("/api/coll/fed-status").then(function(r) {\n    return r.ok ? r.json() : null;\n  }).then(function(fed) {\n    if (fed && fed.ready && fed.collecteur_url) {\n      _fedBase = fed.collecteur_url.replace("/api/push", "");\n      _fedToken = fed.token || "";\n      _syncReady = true;\n      // Démarrer la sync toutes les 3 secondes\n      setInterval(chat_syncCollecteur, 3000);\n      // Présence inter-GHT toutes les 30 secondes\n      setInterval(chat_pushPresence, 30000);\n      chat_pushPresence();\n    }\n  }).catch(function(){});\n\n  // Vérifier que le token est disponible\n  if (!_tok()) {\n    var list = document.getElementById("salon-list");\n    if (list) list.innerHTML = "<div style=\'color:#f87171;font-size:10px;padding:12px;font-family:monospace\'>Token absent — rechargez la page SCRIBE</div>";\n    console.error("SCRIBE Chat: token JWT manquant");\n    return;\n  }\n\n  // Charger les utilisateurs pour l\'autocomplétion\n  _api("/api/coll/users").then(function(r) {\n    return r.ok ? r.json() : [];\n  }).then(function(users) {\n    _autocompleteUsers = users;\n  }).catch(function(){});\n  // Récupérer l\'utilisateur courant et le sigle local\n  _api("/api/coll/me").then(function(r) {\n    return r.ok ? r.json() : null;\n  }).then(function(u) {\n    if (u) _myUserId = u.id;\n  }).catch(function(){});\n  // Le sigle est dans /api/coll/fed-status → "etablissement"\n  _api("/api/coll/fed-status").then(function(r) {\n    return r.ok ? r.json() : null;\n  }).then(function(fed) {\n    if (fed && fed.etablissement) _mysigle = fed.etablissement || "";\n  }).catch(function(){});\n}\n\n// ── Salons ───────────────────────────────────────────────────────────────────\nfunction chat_loadSalons() {\n  var list = document.getElementById("salon-list");\n  _api("/api/chat/salons").then(function(r) {\n    if (!r.ok) {\n      if (list) list.innerHTML = "<div style=\'color:#f87171;font-size:10px;padding:12px;font-family:monospace\'>Erreur " + r.status + " — rechargement dans 5s</div>";\n      setTimeout(chat_loadSalons, 5000);\n      return [];\n    }\n    return r.json();\n  }).then(function(data) {\n    if (!data || !data.length) return;\n    _salons = data;\n    chat_renderSalons();\n    if (!_currentSalon && data.length > 0) {\n      chat_openSalon(data[0]);\n    }\n  }).catch(function(e) {\n    if (list) list.innerHTML = "<div style=\'color:#f87171;font-size:10px;padding:12px;font-family:monospace\'>Connexion impossible — rechargement dans 5s</div>";\n    setTimeout(chat_loadSalons, 5000);\n  });\n}\n\nfunction chat_renderSalons() {\n  var list = document.getElementById("salon-list");\n  var terr = _salons.filter(function(s) { return s.type === "territorial"; });\n  var local = _salons.filter(function(s) { return s.type === "local"; });\n\n  var html = "";\n  if (terr.length) {\n    html += "<div class=\'salon-section-title\'>&#127760; Territorial G7</div>";\n    html += terr.map(chat_renderSalonItem).join("");\n  }\n  if (local.length) {\n    html += "<div class=\'salon-section-title\'>&#127973; Mon établissement</div>";\n    html += local.map(chat_renderSalonItem).join("");\n  }\n  list.innerHTML = html;\n}\n\nfunction chat_renderSalonItem(s) {\n  var active = _currentSalon && _currentSalon.id === s.id ? " active" : "";\n  var col = s.couleur || "#7c3aed";\n  var ght = s.type === "territorial"\n    ? "<div class=\'salon-ght-dot\' style=\'background:" + col + "\'></div>"\n    : "";\n  var activeStyle = (active && s.couleur)\n    ? " style=\'border-left:3px solid " + col + ";background:rgba(0,0,0,.15)\'"\n    : "";\n  return "<div class=\'salon-item" + active + "\' data-salon-id=\'" + s.id + "\'" + activeStyle + " onclick=\'chat_openSalonById(" + s.id + ")\'>" +\n    "<span class=\'salon-icon\'>" + _esc(s.icone) + "</span>" +\n    "<span class=\'salon-name\' style=\'color:" + (active ? col : "") + "\'>#" + _esc(s.nom) + "</span>" +\n    ght +\n  "</div>";\n}\n\nfunction chat_openSalonById(id) {\n  var s = _salons.find(function(x) { return x.id === id; });\n  if (s) chat_openSalon(s);\n}\n\nfunction chat_openSalon(sJson) {\n  var s = typeof sJson === "string" ? JSON.parse(sJson) : sJson;\n  // Stopper le poll AVANT de changer de salon\n  if (_pollTimer) { clearInterval(_pollTimer); _pollTimer = null; }\n  _currentSalon = s;\n  _lastMsgId = 0;\n  _messages = [];\n\n  document.getElementById("chat-salon-title").textContent = "#" + s.nom;\n  document.getElementById("chat-salon-desc").textContent = s.description || "";\n\n  // Vider le feed immédiatement pour feedback visuel\n  var feed = document.getElementById("msg-feed");\n  if (feed) feed.innerHTML = "<div style=\'font-family:monospace;font-size:10px;color:var(--mu);text-align:center;padding:40px\'>Chargement...</div>";\n\n  chat_renderSalons();\n  // Charger les messages puis relancer le poll\n  chat_loadMessages(true, function() {\n    _pollTimer = setInterval(chat_pollMessages, 3000);\n  });\n\n  // Sur mobile, fermer le panneau salons\n  if (window.innerWidth <= 640) {\n    document.getElementById("col-salons").classList.remove("mobile-open");\n  }\n}\n\n// ── Messages ─────────────────────────────────────────────────────────────────\nfunction chat_loadMessages(reset, callback) {\n  if (!_currentSalon) { if (callback) callback(); return; }\n  var url = "/api/chat/salons/" + _currentSalon.id + "/messages?limit=50";\n  _api(url).then(function(r) {\n    return r.ok ? r.json() : [];\n  }).then(function(data) {\n    if (reset) {\n      _messages = data;\n      chat_renderMessages(true);\n    } else {\n      var newMsgs = data.filter(function(m) { return m.id > _lastMsgId; });\n      if (newMsgs.length) {\n        _messages = _messages.concat(newMsgs);\n        newMsgs.forEach(function(m) { chat_appendMessage(m); });\n      }\n    }\n    if (data.length) {\n      _lastMsgId = Math.max.apply(null, data.map(function(m) { return m.id; }));\n    }\n    if (callback) callback();\n  }).catch(function() { if (callback) callback(); });\n}\n\nfunction chat_pollMessages() {\n  if (!_currentSalon) return;\n  var url = "/api/chat/salons/" + _currentSalon.id + "/messages?limit=20";\n  _api(url).then(function(r) {\n    return r.ok ? r.json() : [];\n  }).then(function(data) {\n    var newMsgs = data.filter(function(m) { return m.id > _lastMsgId; });\n    if (newMsgs.length) {\n      newMsgs.forEach(function(m) {\n        _messages.push(m);\n        chat_appendMessage(m);\n      });\n      _lastMsgId = Math.max.apply(null, data.map(function(m) { return m.id; }));\n    }\n  }).catch(function(){});\n}\n\nfunction chat_renderMessages(scrollBottom) {\n  var feed = document.getElementById("msg-feed");\n  if (!_messages.length) {\n    feed.innerHTML = "<div style=\'font-family:monospace;font-size:10px;color:var(--mu);text-align:center;padding:40px\'>Aucun message — soyez le premier !</div>";\n    return;\n  }\n  feed.innerHTML = "";\n  _messages.forEach(function(m) { chat_appendMessage(m); });\n  if (scrollBottom) { feed.scrollTop = feed.scrollHeight; }\n}\n\nfunction chat_appendMessage(m) {\n  var feed = document.getElementById("msg-feed");\n  var isOwn = m.auteur_id && m.auteur_id === _myUserId;\n  var isGHT = m.origine === "ght";\n  var atBottom = feed.scrollHeight - feed.scrollTop - feed.clientHeight < 80;\n\n  var replyHtml = "";\n  if (m.reply_to) {\n    var replyText = (m.reply_to.contenu || "").replace(/<[^>]+>/g,"").substring(0, 100);\n    replyHtml = "<div class=\'msg-reply-preview\' onclick=\'chat_scrollTo(" + m.reply_to.id + ")\' style=\'cursor:pointer\'>" +\n      "<strong style=\'display:block;margin-bottom:2px;color:var(--blue)\'>" + _esc(m.reply_to.auteur_nom) + "</strong>" +\n      "<span style=\'opacity:.8\'>" + _esc(replyText) + (replyText.length >= 100 ? "…" : "") + "</span>" +\n    "</div>";\n  }\n\n  // Formater le contenu avec mentions\n  var contenu = _esc(m.contenu);\n  // Mise en forme : gras, italique, couleurs, code, urgent\n  contenu = contenu\n    .replace(/\\*\\*([^*]+)\\*\\*/g, "<strong>$1</strong>")\n    .replace(/_([^_]+)_/g, "<em>$1</em>")\n    .replace(/\\[code\\](.*?)\\[\\/code\\]/g, "<code style=\'background:rgba(0,0,0,.1);padding:1px 4px;border-radius:3px;font-family:monospace\'>$1</code>")\n    .replace(/\\[rouge\\](.*?)\\[\\/rouge\\]/g, "<span style=\'color:#ef4444;font-weight:600\'>$1</span>")\n    .replace(/\\[vert\\](.*?)\\[\\/vert\\]/g, "<span style=\'color:#22c55e;font-weight:600\'>$1</span>")\n    .replace(/\\[orange\\](.*?)\\[\\/orange\\]/g, "<span style=\'color:#f97316;font-weight:600\'>$1</span>")\n    .replace(/---URGENT---(.*?)---/g, "<span style=\'background:#ef4444;color:#fff;padding:2px 8px;border-radius:4px;font-weight:700\'>&#128680; URGENT $1</span>");\n  // URLs cliquables\n  contenu = contenu.replace(/(https?:\\/\\/[^\\s<>"]+)/g,\n    "<a href=\'$1\' target=\'_blank\' rel=\'noopener\' style=\'color:var(--blue);text-decoration:underline\'>$1</a>");\n  contenu = contenu.replace(/@([\\w\\-\\[\\]]+)/g, "<span class=\'msg-mention\'>@$1</span>");\n\n  var pjHtml = "";\n  if (m.pj && m.pj.length) {\n    pjHtml = m.pj.map(function(p) {\n      var ext = p.nom.split(".").pop().toUpperCase();\n      _pjMap[p.id] = p.nom;\n      var isImage = ["PNG","JPG","JPEG","GIF","WEBP","SVG"].indexOf(ext) >= 0;\n      // PJ inter-GHT (id commence par "ght-") → URL du collecteur\n      var isGhtPj = String(p.id).indexOf("ght-") === 0;\n      var pjUrl = isGhtPj\n        ? (_fedBase + "/api/chat/pj/" + p.id + "?token=" + _fedToken)\n        : ("/api/chat/pj/" + p.id + "?token=" + _tok());\n      if (isImage) {\n        return "<div class=\'msg-pj-img\' data-pj-id=\'" + p.id + "\'>" +\n          "<img src=\'" + pjUrl + "\' " +\n          "style=\'max-width:300px;max-height:200px;border-radius:6px;cursor:pointer;display:block;margin-top:4px\' " +\n          "onclick=\'chat_downloadPJExtUrl(" + JSON.stringify(pjUrl) + "," + JSON.stringify(p.nom) + ")\' " +\n          "onerror=\'this.remove()\' " +\n          "alt=\'" + _esc(p.nom) + "\'></div>";\n      }\n      return "<div class=\'msg-pj\' data-pj-id=\'" + p.id + "\' onclick=\'chat_downloadPJExtUrl(" + JSON.stringify(pjUrl) + "," + JSON.stringify(p.nom) + ")\'>" +\n        "<span class=\'msg-pj-icon\'>&#128206;</span>" +\n        "<span style=\'flex:1;font-size:11px\'>" + _esc(p.nom) + "</span>" +\n        "<span style=\'font-family:monospace;font-size:9px;color:var(--mu)\'>" + ext + " " + Math.round(p.taille/1024) + "Ko</span>" +\n      "</div>";\n    }).join("");\n  }\n\n  var bubbleClass = "msg-bubble" + (isOwn ? " own" : "") + (isGHT ? " ght" : "") + (m.supprime ? " supprime" : "");\n  var authorClass = "msg-author" + (isGHT ? " ght" : "");\n\n  var ts = "";\n  if (m.horodatage) {\n    try {\n      var tsStr = m.horodatage;\n      // Normaliser : ajouter Z si absent, remplacer espace par T\n      if (tsStr.indexOf("Z") < 0 && tsStr.indexOf("+") < 0) tsStr = tsStr.replace(" ", "T") + "Z";\n      var d = new Date(tsStr);\n      if (!isNaN(d.getTime())) {\n        ts = d.toLocaleTimeString("fr-FR", {hour:"2-digit",minute:"2-digit"});\n      }\n    } catch(e) {}\n  }\n\n  var actionsHtml = "";\n  if (!m.supprime) {\n    var msgKey = "mk_" + String(m.id).replace(/[^a-zA-Z0-9]/g, "_");\n    _msgMap[msgKey] = m;\n    actionsHtml = "<div class=\'msg-actions\'>" +\n      "<button class=\'msg-action-btn\' data-mkey=\'" + msgKey + "\' onclick=\'chat_replyToKey(this.dataset.mkey)\'>&#8617; R&eacute;pondre</button>";\n    if (isOwn) {\n      actionsHtml += "<button class=\'msg-action-btn\' style=\'color:#dc2626\' data-mkey=\'" + msgKey + "\' onclick=\'chat_deleteMsgKey(this.dataset.mkey)\'>Supprimer</button>";\n    }\n    actionsHtml += "</div>";\n  }\n\n  var div = document.createElement("div");\n  div.className = "msg-group";\n  div.id = "msg-" + m.id;\n  div.innerHTML = "<div class=\'msg-author-line\'>" +\n      "<span class=\'" + authorClass + "\'>" + _esc(m.auteur_nom) + "</span>" +\n      "<span class=\'msg-time\'>" + ts + (isGHT ? " &#127760;" : "") + "</span>" +\n      actionsHtml +\n    "</div>" +\n    "<div class=\'msg-row\'>" +\n      "<div class=\'" + bubbleClass + "\'>" +\n        replyHtml +\n        "<div>" + contenu + "</div>" +\n        pjHtml +\n      "</div>" +\n    "</div>";\n\n  feed.appendChild(div);\n  if (atBottom) { feed.scrollTop = feed.scrollHeight; }\n\n  // Notification si message mentionne l\'utilisateur courant\n  var myName = _autocompleteUsers && _autocompleteUsers.length\n    ? null : null; // On utilise display_name via presence\n  var myDisplay = "";\n  Object.values(_presence).forEach(function(arr) {\n    arr.forEach(function(u) { if (!myDisplay) myDisplay = u.display_name; });\n  });\n  var isMentioned = myDisplay && m.mentions && m.mentions.some(function(mn) {\n    return mn.toLowerCase() === myDisplay.toLowerCase();\n  });\n  if (isMentioned || (m.auteur_sigle && m.auteur_sigle !== (_mysigle || ""))) {\n    // Envoyer un event au parent (index.html) pour le badge\n    try {\n      window.parent.postMessage({type: "scribe-chat-new-msg", mentioned: isMentioned}, "*");\n    } catch(e) {}\n  }\n}\n\nfunction chat_scrollTo(msgId) {\n  var el = document.getElementById("msg-" + msgId);\n  if (el) { el.scrollIntoView({behavior:"smooth", block:"center"}); }\n}\n\n// ── Envoi ────────────────────────────────────────────────────────────────────\nfunction chat_send() {\n  if (!_currentSalon) { _toast("Sélectionnez un salon", "err"); return; }\n  var inp = document.getElementById("msg-input");\n  var contenu = inp.value.trim();\n  if (!contenu && !_pendingPJs.length) return; // Autoriser envoi si PJs\n\n  // Extraire les mentions\n  var mentions = [];\n  var re = /@([\\w\\-\\[\\]]+)/g;\n  var m;\n  while ((m = re.exec(contenu)) !== null) { mentions.push(m[1]); }\n\n  // Séparer les PJs inline (dataUrl) des PJs uploadées (id serveur)\n  var inlinePJs = _pendingPJs.filter(function(p) { return p.inline; });\n  var serverPJs = _pendingPJs.filter(function(p) { return !p.inline; });\n  var body = {\n    contenu: contenu,\n    mentions: mentions,\n    reply_to_id: (_replyTo && !isNaN(parseInt(_replyTo.id))) ? parseInt(_replyTo.id) : null,\n    pj_ids: serverPJs.map(function(p) { return p.id; }),\n    pj_inline: inlinePJs.map(function(p) {\n      return {nom: p.nom, taille: p.taille, dataUrl: p.dataUrl};\n    }),\n  };\n\n  _api("/api/chat/salons/" + _currentSalon.id + "/messages", {\n    method: "POST", body: JSON.stringify(body)\n  }).then(function(r) {\n    return r.ok ? r.json() : Promise.reject(r.status);\n  }).then(function(msg) {\n    inp.value = "";\n    inp.style.height = "auto";\n    chat_cancelReply();\n    // Si PJs → recharger le message depuis l\'API pour avoir les PJs à jour\n    if (msg.pj && msg.pj.length) {\n      _messages.push(msg);\n      if (msg.id > _lastMsgId) _lastMsgId = msg.id;\n      chat_appendMessage(msg);\n    } else {\n      _messages.push(msg);\n      if (msg.id > _lastMsgId) _lastMsgId = msg.id;\n      chat_appendMessage(msg);\n    }\n    document.getElementById("msg-feed").scrollTop = 9999999;\n    document.getElementById("autocomplete-box").style.display = "none";\n    // Sauvegarder les PJs avant vidage pour le push collecteur\n    var _pendingPJsSent = _pendingPJs.slice();\n    // Vider les PJs en attente\n    _pendingPJs = [];\n    chat_renderPendingPJs();\n    // Pousser vers le collecteur si salon territorial et sync active\n    if (_syncReady && _currentSalon && _currentSalon.type === "territorial") {\n      // NE PAS envoyer les images en base64 → trop lourd, cause CORS+500\n      // Envoyer uniquement les métadonnées avec URL directe vers cette instance\n      var srcOrigin = window.location.origin;\n      // URL directe vers chaque PJ sur cette instance (sans base64)\n      var pjsMeta = (msg.pj || []).map(function(p) {\n        var pjUrl = srcOrigin + "/api/chat/pj/" + p.id + "?token=" + _fedToken;\n        return {nom: p.nom, taille: p.taille, remote_url: pjUrl};\n      });\n      fetch(_fedBase + "/api/chat/messages", {\n        method: "POST",\n        headers: {Authorization: "Bearer " + _fedToken, "Content-Type": "application/json"},\n        body: JSON.stringify({\n          salon_nom:    _currentSalon.nom,\n          auteur_nom:   msg.auteur_nom,\n          auteur_sigle: msg.auteur_sigle,\n          contenu:      msg.contenu,\n          mentions:     msg.mentions || [],\n          pj_meta:      pjsMeta,\n        })\n      }).catch(function(){});\n   }\n  }).catch(function(e) {\n    _toast("Erreur envoi : " + e, "err");\n  });\n}\n\nfunction chat_onKeydown(e) {\n  var ac = document.getElementById("autocomplete-box");\n  var items = ac.querySelectorAll(".autocomplete-item");\n\n  if (ac.style.display !== "none" && items.length) {\n    if (e.key === "ArrowDown") {\n      e.preventDefault();\n      _autocompleteIdx = Math.min(_autocompleteIdx + 1, items.length - 1);\n      items.forEach(function(el, i) { el.classList.toggle("selected", i === _autocompleteIdx); });\n      return;\n    }\n    if (e.key === "ArrowUp") {\n      e.preventDefault();\n      _autocompleteIdx = Math.max(_autocompleteIdx - 1, 0);\n      items.forEach(function(el, i) { el.classList.toggle("selected", i === _autocompleteIdx); });\n      return;\n    }\n    if (e.key === "Tab" || e.key === "Enter") {\n      if (_autocompleteIdx >= 0 && items[_autocompleteIdx]) {\n        e.preventDefault();\n        items[_autocompleteIdx].click();\n        return;\n      }\n    }\n    if (e.key === "Escape") {\n      ac.style.display = "none";\n      return;\n    }\n  }\n\n  if (e.key === "Enter" && !e.shiftKey) {\n    e.preventDefault();\n    chat_send();\n  }\n}\n\nfunction chat_onInput(el) {\n  // Auto-resize\n  el.style.height = "auto";\n  el.style.height = Math.min(el.scrollHeight, 120) + "px";\n  // Autocomplétion @mention\n  var val = el.value;\n  var cursor = el.selectionStart;\n  var before = val.substring(0, cursor);\n  var atMatch = before.match(/@([\\w\\-]*)$/);\n  if (atMatch) {\n    var query = atMatch[1].toLowerCase();\n    chat_showAutocomplete(query, before.lastIndexOf("@"), cursor);\n  } else {\n    document.getElementById("autocomplete-box").style.display = "none";\n  }\n}\n\nfunction chat_showAutocomplete(query, atPos, cursor) {\n  var results = [];\n  var q = (query || "").toLowerCase();\n  // Uniquement les utilisateurs CONNECTES (présence locale + GHT)\n  var allConnected = {};\n  // Présence locale\n  Object.values(_presence).forEach(function(arr) {\n    arr.forEach(function(u) { allConnected[u.display_name] = {label: u.display_name, icon: "👤"}; });\n  });\n  // Présence inter-GHT\n  Object.values(_presenceGHT || {}).forEach(function(arr) {\n    arr.forEach(function(u) { allConnected[u.display_name] = {label: u.display_name, icon: "🌐"}; });\n  });\n  Object.values(allConnected).forEach(function(u) {\n    if (!q || u.label.toLowerCase().includes(q)) {\n      results.push({label: u.label, value: u.label, icon: u.icon});\n    }\n  });\n  // Établissements connectés\n  Object.keys(Object.assign({}, _presence, _presenceGHT || {})).forEach(function(s) {\n    if (!q || s.toLowerCase().includes(q)) {\n      results.push({label: s + " (établissement)", value: s, icon: "🏥"});\n    }\n  });\n\n  results = results.slice(0, 8);\n  var ac = document.getElementById("autocomplete-box");\n  if (!results.length) { ac.style.display = "none"; return; }\n\n  _autocompleteIdx = -1;\n  // Assigner aux variables GLOBALES (pas var → shadowing)\n  _acResults = results;\n  _acAtPos = atPos;\n  _acCursor = cursor;\n  ac.innerHTML = results.map(function(r, i) {\n    return "<div class=\'autocomplete-item\' data-ac-idx=\'" + i + "\' onclick=\'chat_pickMentionIdx(" + i + ")\'>" +\n      r.icon + " " + _esc(r.label) + "</div>";\n  }).join("");\n  ac.style.display = "block";\n}\n\nvar _acResults = [];\nvar _acAtPos = 0;\nvar _acCursor = 0;\n\nfunction chat_toggleEmoji() {\n  var p = document.getElementById("emoji-picker");\n  if (p) p.style.display = p.style.display === "none" ? "flex" : "none";\n}\n\nfunction chat_insertEmoji(e) {\n  var inp = document.getElementById("msg-input");\n  if (!inp) return;\n  var pos = inp.selectionStart || inp.value.length;\n  inp.value = inp.value.substring(0, pos) + e + inp.value.substring(pos);\n  inp.focus();\n  inp.selectionStart = inp.selectionEnd = pos + e.length;\n  inp.style.height = "auto";\n  inp.style.height = Math.min(inp.scrollHeight, 120) + "px";\n  document.getElementById("emoji-picker").style.display = "none";\n}\n\nfunction chat_fmt(before, after) {\n  var inp = document.getElementById("msg-input");\n  if (!inp) return;\n  var start = inp.selectionStart, end = inp.selectionEnd;\n  var sel = inp.value.substring(start, end) || "texte";\n  inp.value = inp.value.substring(0, start) + before + sel + after + inp.value.substring(end);\n  inp.focus();\n  inp.selectionStart = start + before.length;\n  inp.selectionEnd = start + before.length + sel.length;\n  inp.style.height = "auto";\n  inp.style.height = Math.min(inp.scrollHeight, 120) + "px";\n}\n\nfunction chat_pickMentionIdx(idx) {\n  var r = _acResults[idx];\n  if (r) chat_pickMention(r.value, _acAtPos, _acCursor);\n}\n\nfunction chat_pickMention(val, atPos, cursor) {\n  var inp = document.getElementById("msg-input");\n  var before = inp.value.substring(0, atPos);\n  var after = inp.value.substring(cursor);\n  inp.value = before + "@" + val + " " + after;\n  inp.focus();\n  document.getElementById("autocomplete-box").style.display = "none";\n  _autocompleteIdx = -1;\n}\n\n// ── Répondre ─────────────────────────────────────────────────────────────────\nfunction chat_replyToId(id) {\n  var m = _messages.find(function(x) { return String(x.id) === String(id); });\n  if (m) chat_replyTo(m);\n}\n\nfunction chat_replyToKey(key) {\n  var m = _msgMap[key];\n  if (m) chat_replyTo(m);\n}\n\nfunction chat_deleteMsgKey(key) {\n  var m = _msgMap[key];\n  if (!m) return;\n  if (String(m.id).indexOf("coll-") === 0) {\n    alert("Impossible de supprimer un message d un autre etablissement.");\n    return;\n  }\n  chat_deleteMsg(m.id);\n}\n\nfunction chat_replyTo(mJson) {\n  var m = typeof mJson === "string" ? JSON.parse(mJson) : mJson;\n  _replyTo = m;\n  var authorEl = document.getElementById("reply-author");\n  var previewEl = document.getElementById("reply-preview");\n  if (authorEl) authorEl.textContent = m.auteur_nom;\n  if (previewEl) {\n    var preview = (m.contenu || "").replace(/<[^>]+>/g, "").substring(0, 80);\n    previewEl.textContent = preview + (m.contenu && m.contenu.length > 80 ? "…" : "");\n  }\n  var bar = document.getElementById("reply-bar");\n  if (bar) bar.style.display = "block";\n  document.getElementById("msg-input").focus();\n}\n\nfunction chat_cancelReply() {\n  _replyTo = null;\n  document.getElementById("reply-bar").style.display = "none";\n  document.getElementById("reply-author").textContent = "";\n  document.getElementById("reply-preview").textContent = "";\n}\n\n// ── Supprimer message ─────────────────────────────────────────────────────────\nfunction chat_deleteMsg(msgId) {\n  if (!confirm("Supprimer ce message ?")) return;\n  _api("/api/chat/salons/" + _currentSalon.id + "/messages/" + msgId, {method:"DELETE"})\n    .then(function(r) {\n      if (r.ok) {\n        var el = document.getElementById("msg-" + msgId);\n        if (el) {\n          var bubble = el.querySelector(".msg-bubble");\n          if (bubble) { bubble.textContent = "_(message supprimé)_"; bubble.classList.add("supprime"); }\n        }\n      } else { _toast("Erreur suppression", "err"); }\n    });\n}\n\n// ── Pièces jointes ────────────────────────────────────────────────────────────\nfunction chat_uploadPJ(input) {\n  if (!_currentSalon) { _toast("Sélectionnez un salon", "err"); return; }\n  var files = Array.from(input.files || []);\n  if (!files.length) return;\n  input.value = "";\n\n  files.forEach(function(file) {\n    var reader = new FileReader();\n    reader.onload = function(e) {\n      var dataUrl = e.target.result;\n      var ext = (file.name.split(".").pop() || "").toUpperCase();\n      // Mode inline : stocker le dataUrl directement sans pré-upload\n      // (compatible collecteur ET instances SCRIBE avec images lourdes)\n      var fakePj = {\n        id: "inline-" + Date.now() + "-" + Math.random().toString(36).slice(2),\n        nom: file.name,\n        taille: file.size,\n        ext: ext,\n        dataUrl: dataUrl,\n        inline: true,\n      };\n      _pendingPJs.push(fakePj);\n      chat_renderPendingPJs();\n    };\n    reader.readAsDataURL(file);\n  });\n}\n\nfunction chat_renderPendingPJs() {\n  var zone = document.getElementById("pending-pjs");\n  if (!zone) return;\n  if (!_pendingPJs.length) { zone.style.display = "none"; zone.innerHTML = ""; return; }\n  zone.style.display = "flex";\n  zone.innerHTML = _pendingPJs.map(function(pj, i) {\n    var isImg = ["JPG","JPEG","PNG","GIF","WEBP","SVG"].indexOf((pj.ext||"").toUpperCase()) >= 0;\n    var preview = isImg\n      ? "<img src=\'" + pj.dataUrl + "\' style=\'height:60px;border-radius:4px;object-fit:cover\'>"\n      : "<span style=\'font-size:20px\'>&#128206;</span>";\n    return "<div style=\'position:relative;display:inline-flex;flex-direction:column;align-items:center;gap:2px;background:var(--s2);border:1px solid var(--bd);border-radius:6px;padding:6px 8px;font-size:10px;max-width:80px\'>" +\n      preview +\n      "<span style=\'overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:70px;font-family:monospace\'>" + _esc(pj.nom) + "</span>" +\n      "<button onclick=\'chat_removePendingPJ(" + i + ")\' style=\'position:absolute;top:-6px;right:-6px;width:16px;height:16px;border-radius:50%;background:#ef4444;color:#fff;border:none;cursor:pointer;font-size:10px;line-height:1;display:flex;align-items:center;justify-content:center\'>&#x2715;</button>" +\n    "</div>";\n  }).join("");\n}\n\nfunction chat_removePendingPJ(idx) {\n  _pendingPJs.splice(idx, 1);\n  chat_renderPendingPJs();\n}\n\nfunction chat_downloadPJExtUrl(url, nom) {\n  var a = document.createElement("a");\n  a.href = url;\n  a.download = nom || "fichier";\n  a.click();\n}\n\nfunction chat_downloadPJById(pjId) {\n  var nom = _pjMap[pjId] || "fichier";\n  chat_downloadPJ(pjId, nom);\n}\n\nfunction chat_downloadPJ(pjId, nom) {\n  var a = document.createElement("a");\n  a.href = "/api/chat/pj/" + pjId + "?token=" + _tok();\n  a.download = nom;\n  a.click();\n}\n\n// ── Présence ─────────────────────────────────────────────────────────────────\nfunction chat_ping() {\n  _api("/api/chat/presence/ping", {method:"POST"}).catch(function(){});\n}\n\nfunction chat_loadPresence() {\n  _api("/api/chat/presence").then(function(r) {\n    return r.ok ? r.json() : {};\n  }).then(function(data) {\n    _presence = data;\n    chat_renderPresenceMerged();\n  }).catch(function(){});\n  setTimeout(chat_loadPresence, 15000);\n  // Aussi pinger la présence locale\n  chat_ping();\n}\n\nfunction chat_renderPresence() {\n  var list = document.getElementById("presence-list");\n  var html = "";\n  Object.entries(_presence).sort().forEach(function(entry) {\n    var sigle = entry[0], users = entry[1];\n    html += "<div class=\'presence-etab\'>" + _esc(sigle) + " (" + users.length + ")</div>";\n    users.forEach(function(u) {\n      html += "<div class=\'presence-user\'><div class=\'presence-dot\'></div><div class=\'presence-name\'>" + _esc(u.display_name) + "</div></div>";\n    });\n  });\n  list.innerHTML = html || "<div style=\'font-family:monospace;font-size:10px;color:var(--mu);padding:16px;text-align:center\'>Aucun utilisateur connecté</div>";\n}\n\nfunction chat_openPresence() {\n  var col = document.getElementById("col-presence");\n  col.classList.toggle("hidden");\n}\n\nfunction chat_closePresence() {\n  document.getElementById("col-presence").classList.add("hidden");\n}\n\n// ── Nouveau salon ─────────────────────────────────────────────────────────────\nfunction chat_openNewSalon() {\n  document.getElementById("modal-salon").style.display = "flex";\n  document.getElementById("ns-nom").value = "";\n  document.getElementById("ns-desc").value = "";\n  document.getElementById("ns-type").value = "local";\n  document.getElementById("ns-icon").value = "💬";\n  document.getElementById("ns-color").value = "#003189";\n  document.querySelectorAll(".color-dot").forEach(function(d) { d.classList.remove("sel"); });\n  var first = document.querySelector(".color-dot");\n  if (first) first.classList.add("sel");\n}\n\nfunction chat_closeNewSalon() {\n  document.getElementById("modal-salon").style.display = "none";\n}\n\nfunction chat_selColor(el) {\n  document.querySelectorAll(".color-dot").forEach(function(d) { d.classList.remove("sel"); });\n  el.classList.add("sel");\n  document.getElementById("ns-color").value = el.dataset.c;\n  _newColor = el.dataset.c;\n}\n\nfunction chat_createSalon() {\n  var nom = document.getElementById("ns-nom").value.trim();\n  if (!nom) { _toast("Nom requis", "err"); return; }\n  var type = document.getElementById("ns-type").value;\n  var body = {\n    nom: nom,\n    description: document.getElementById("ns-desc").value.trim() || null,\n    type: type,\n    couleur: document.getElementById("ns-color").value,\n    icone: document.getElementById("ns-icon").value || "💬"\n  };\n  _api("/api/chat/salons", {method:"POST", body:JSON.stringify(body)})\n    .then(function(r) { return r.ok ? r.json() : Promise.reject(r.status); })\n    .then(function(s) {\n      _toast("Salon #" + s.nom + " créé !", "ok");\n      chat_closeNewSalon();\n      // Si salon territorial : pousser vers le collecteur (qu\'il soit syncReady ou pas)\n      if (type === "territorial") {\n        // Récupérer l\'URL du collecteur depuis fed-status\n        _api("/api/coll/fed-status").then(function(fr) {\n          return fr.ok ? fr.json() : null;\n        }).then(function(fed) {\n          if (fed && fed.collecteur_url) {\n            var fedBase = fed.collecteur_url.replace("/api/push", "");\n            var fedToken = fed.token || "";\n            fetch(fedBase + "/api/chat/salons", {\n              method: "POST",\n              headers: {Authorization: "Bearer " + fedToken, "Content-Type": "application/json"},\n              body: JSON.stringify({nom: s.nom})\n            }).catch(function(){});\n          }\n        }).catch(function(){});\n      }\n      chat_loadSalons();\n    }).catch(function(e) { _toast("Erreur création : " + e, "err"); });\n}\n\n// ── Mobile ────────────────────────────────────────────────────────────────────\nfunction chat_toggleSalons() {\n  document.getElementById("col-salons").classList.toggle("mobile-open");\n}\n\n// ── Popout ────────────────────────────────────────────────────────────────────\nfunction chat_popout() {\n  var token = _tok();\n  var salon = _currentSalon ? "&salon=" + _currentSalon.id : "";\n  window.open("/api/chat/ui/popout?token=" + encodeURIComponent(token) + salon, "_blank",\n    "width=900,height=700,menubar=no,toolbar=no,scrollbars=yes");\n}\n\n// ── Exposer dans window ───────────────────────────────────────────────────────\n\n// ── Sync inter-GHT via collecteur ─────────────────────────────────────────\nfunction chat_syncCollecteur() {\n  if (!_syncReady) return;\n  // Sync des salons territoriaux du collecteur\n  fetch(_fedBase + "/api/chat/salons", {\n    headers: {Authorization: "Bearer " + _fedToken}\n  }).then(function(r) { return r.ok ? r.json() : []; })\n  .then(function(collSalons) {\n    collSalons.forEach(function(cs) {\n      // Si salon territorial pas encore local → le créer\n      var exists = _salons.some(function(s) { return s.nom === cs.nom && s.type === "territorial"; });\n      if (!exists && cs.nom) {\n        _api("/api/chat/salons", {\n          method: "POST",\n          body: JSON.stringify({nom: cs.nom, type: "territorial", icone: cs.icone || "💬",\n                                couleur: cs.couleur || "#7c3aed", description: cs.description || ""})\n        }).then(function(r) {\n          if (r.ok) chat_loadSalons();\n        }).catch(function(){});\n      }\n    });\n  }).catch(function(){});\n  if (!_currentSalon) return;\n  var salonNom = _currentSalon.nom;\n  var sinceId = _chatSyncIds[salonNom] || 0;\n\n  // 1. Pull : récupérer les nouveaux messages du collecteur\n  fetch(_fedBase + "/api/chat/messages?salon_nom=" + encodeURIComponent(salonNom) + "&since_id=" + sinceId, {\n    headers: {Authorization: "Bearer " + _fedToken}\n  }).then(function(r) {\n    return r.ok ? r.json() : [];\n  }).then(function(msgs) {\n    msgs.forEach(function(m) {\n      // Ignorer nos propres messages (déjà dans le feed)\n      var sigleLocal = _mysigle || "";\n      if (m.auteur_sigle === sigleLocal) return;\n      // Vérifier si déjà affiché\n      if (document.getElementById("msg-coll-" + m.id)) return;\n      // Créer un pseudo-message au format local\n      var pseudo = {\n        id:           "coll-" + m.id,\n        salon_id:     _currentSalon ? _currentSalon.id : 0,\n        auteur_id:    null,\n        auteur_nom:   m.auteur_nom,\n        auteur_sigle: m.auteur_sigle,\n        contenu:      m.contenu,\n        mentions:     m.mentions || [],\n        reply_to:     null,\n        horodatage:   m.horodatage,\n        supprime:     false,\n        origine:      "ght",\n        pj:           m.pj || [],\n      };\n      _messages.push(pseudo);\n      chat_appendMessage(pseudo);\n      if (m.id > (_chatSyncIds[salonNom] || 0)) {\n        _chatSyncIds[salonNom] = m.id;\n      }\n    });\n  }).catch(function(){});\n\n  // 2. Pull présence inter-GHT\n  fetch(_fedBase + "/api/chat/presence", {\n    headers: {Authorization: "Bearer " + _fedToken}\n  }).then(function(r) {\n    return r.ok ? r.json() : {};\n  }).then(function(data) {\n    // Merger avec la présence locale\n    _presenceGHT = data;\n    chat_renderPresenceMerged();\n  }).catch(function(){});\n}\n\nvar _presenceGHT = {};\n\nfunction chat_pushPresence() {\n  if (!_syncReady) return;\n  // Pousser la présence locale vers le collecteur\n  var users = Object.values(_presence).reduce(function(acc, arr) {\n    return acc.concat(arr);\n  }, []);\n  if (!users.length) return;\n  fetch(_fedBase + "/api/chat/presence", {\n    method: "POST",\n    headers: {Authorization: "Bearer " + _fedToken, "Content-Type": "application/json"},\n    body: JSON.stringify({users: users})\n  }).catch(function(){});\n}\n\nfunction chat_renderPresenceMerged() {\n  // Fusionner présence locale + GHT\n  var merged = {};\n  // Locale\n  Object.entries(_presence).forEach(function(e) {\n    merged[e[0]] = (merged[e[0]] || []).concat(e[1]);\n  });\n  // GHT (autres établissements)\n  Object.entries(_presenceGHT).forEach(function(e) {\n    var sigle = e[0];\n    var users = e[1];\n    // Éviter les doublons avec le sigle local\n    var sigleLocal = _mysigle || "";\n    if (sigle === sigleLocal) return; // déjà dans la présence locale\n    merged[sigle] = (merged[sigle] || []).concat(users);\n  });\n\n  var list = document.getElementById("presence-list");\n  if (!list) return;\n  var total = 0;\n  var html = "";\n  Object.keys(merged).sort().forEach(function(sigle) {\n    var users = merged[sigle];\n    if (!users.length) return;\n    total += users.length;\n    html += "<div class=\'presence-etab\'>" + sigle + " (" + users.length + ")</div>";\n    users.forEach(function(u) {\n      html += "<div class=\'presence-user\'><div class=\'presence-dot\'></div><div class=\'presence-name\'>" + _esc(u.display_name) + "</div></div>";\n    });\n  });\n  list.innerHTML = html || "<div style=\'font-family:monospace;font-size:10px;color:var(--mu);padding:16px;text-align:center\'>Aucun utilisateur connecté</div>";\n  var el = document.getElementById("presence-count");\n  if (el) el.textContent = total;\n}\n\nwindow.chat_init            = chat_init;\nwindow.chat_openSalonById   = chat_openSalonById;\nwindow.chat_openSalon       = chat_openSalon;\nwindow.chat_send            = chat_send;\nwindow.chat_onInput         = chat_onInput;\nwindow.chat_onKeydown       = chat_onKeydown;\nwindow.chat_pickMention     = chat_pickMention;\nwindow.chat_replyToId       = chat_replyToId;\nwindow.chat_replyToKey      = chat_replyToKey;\nwindow.chat_deleteMsgKey    = chat_deleteMsgKey;\nwindow.chat_toggleEmoji     = chat_toggleEmoji;\nwindow.chat_insertEmoji     = chat_insertEmoji;\nwindow.chat_fmt             = chat_fmt;\nwindow.chat_pickMentionIdx  = chat_pickMentionIdx;\nwindow.chat_replyTo         = chat_replyTo;\nwindow.chat_cancelReply     = chat_cancelReply;\nwindow.chat_deleteMsg       = chat_deleteMsg;\nwindow.chat_uploadPJ        = chat_uploadPJ;\nwindow.chat_downloadPJExtUrl = chat_downloadPJExtUrl;\nwindow.chat_downloadPJById  = chat_downloadPJById;\nwindow.chat_removePendingPJ = chat_removePendingPJ;\nwindow.chat_renderPendingPJs = chat_renderPendingPJs;\nwindow.chat_downloadPJ      = chat_downloadPJ;\nwindow.chat_scrollTo        = chat_scrollTo;\nwindow.chat_openNewSalon    = chat_openNewSalon;\nwindow.chat_closeNewSalon   = chat_closeNewSalon;\nwindow.chat_selColor        = chat_selColor;\nwindow.chat_createSalon     = chat_createSalon;\nwindow.chat_toggleSalons    = chat_toggleSalons;\nwindow.chat_popout          = chat_popout;\nwindow.chat_syncCollecteur  = chat_syncCollecteur;\nwindow.chat_pushPresence    = chat_pushPresence;\nwindow.chat_openPresence    = chat_openPresence;\nwindow.chat_closePresence   = chat_closePresence;\n\nwindow.addEventListener("load", chat_init);\n</script>\n</body>\n</html>'

@app.get("/chat/ui", response_class=HTMLResponse)
async def chat_ui_coll(token: str = "", credentials=Depends(security)):
    """Sert le chat pour la supervision. Vérifie la session et injecte l'ADMIN_TOKEN."""
    # Vérifier que l'appelant est authentifié (session UI ou admin)
    # Accepter le token ADMIN_TOKEN OU un token de session valide
    # Si ui_sessions est vide (redémarrage), accepter quand même si token fourni
    # (la page supervisor est déjà derrière auth)
    ok = False
    tok_to_check = token or (credentials.credentials if credentials else "")
    if tok_to_check == ADMIN_TOKEN:
        ok = True
    elif tok_to_check in ui_sessions:
        ok = True
    elif tok_to_check:
        # Token fourni mais session expirée → accepter quand même
        # (l'utilisateur est déjà dans l'interface supervisor)
        ok = True
    if not ok:
        raise HTTPException(401, "Aucun token fourni")
    # Injecter l'ADMIN_TOKEN pour que le chat puisse appeler les routes /api/chat/*
    import html as _h
    tok_inject = (
        "<script>"
        "window._COLL_TOKEN=" + repr(ADMIN_TOKEN) + ";"
        "</script>"
    )
    patched = CHAT_COLL_HTML.replace(
        "var _tok = function() {",
        "var _tok = function() { try { if(window._COLL_TOKEN) return window._COLL_TOKEN; } catch(e){} "
    )
    return HTMLResponse(tok_inject + patched)


# ─────────────────────────────────────────────────────────────────────────────
# MESSAGERIE SUPERVISION — interface ISO instance (iframe) + livraison retour
# ─────────────────────────────────────────────────────────────────────────────
import os as _os_msgui
_MSG_UI_PATH = _os_msgui.path.join(_os_msgui.path.dirname(_os_msgui.path.abspath(__file__)), "messagerie_app.html")

@app.get("/messagerie/ui", response_class=HTMLResponse)
async def messagerie_ui_coll(token: str = "", credentials=Depends(security)):
    """Sert la page messagerie 3-panneaux (identique aux instances) pour l'iframe."""
    chk = token or (credentials.credentials if credentials else "")
    if not chk:
        raise HTTPException(401, "Token requis")
    try:
        html = open(_MSG_UI_PATH, encoding="utf-8").read()
    except Exception as e:
        return HTMLResponse(f"<p style='font-family:monospace'>Erreur chargement messagerie : {e}</p>", status_code=500)
    return HTMLResponse(html)


@app.get("/api/coll/instances")
async def coll_instances(credentials=Depends(security)):
    """Liste des instances enrôlées (pour le sélecteur destinataire du compose)."""
    if not _check_any_auth(credentials):
        raise HTTPException(401)
    seen = {}
    for _tok, sig in tokens.items():
        if not sig or sig in seen:
            continue
        nom = ""
        try:
            ed = etablissements.get(sig) or {}
            nom = ed.get("nom") or ed.get("etablissement") or ed.get("nom_etablissement") or ""
        except Exception:
            nom = ""
        seen[sig] = {"sigle": sig, "nom": nom or sig}
    return {"instances": list(seen.values())}


@app.post("/api/coll/msg-to-instance")
async def coll_msg_to_instance(
    target_sigle: str = Form(...),
    sujet:        str = Form(""),
    contenu:      str = Form(""),
    recipient_username: str = Form(""),
    origin_sigle: str = Form(""),
    origin_nom:   str = Form(""),
    origin_username: str = Form(""),
    fichiers:     list[UploadFile] = File([]),
    caller:       str = Depends(require_node_or_admin),
):
    """Livraison messagerie → instance : POST vers /api/v1/messagerie/ingest
    de l'instance cible (résolue via PORT_MAP). Utilisé par la supervision
    (broadcast cellule de crise) ET par une instance qui écrit à un agent
    nominatif d'une autre instance (recipient_username renseigné).
    `caller` = sigle de l'appelant (SUPERVISEUR si admin)."""
    # h72 — Résolution du port cible par DÉCOUVERTE (sigle réellement déclaré
    # par chaque instance), comme /api/annuaire. Le PORT_MAP codé en dur cassait
    # avec des sigles personnalisés (« CH Rivemont » → défaut 8000 → message livré
    # à la mauvaise instance). Repli PORT_MAP uniquement si la découverte échoue.
    import httpx, asyncio
    _want = (target_sigle or "").upper().strip()
    _ports = list(range(8000, 8010)) + [7474]

    # ── Diffusion à TOUTES les instances (communiqué) : target_sigle == "*" ──
    if _want == "*":
        _bcast_raw = []
        for up in (fichiers or []):
            try:
                _bcast_raw.append((up.filename or "fichier", await up.read(),
                                   up.content_type or "application/octet-stream"))
            except Exception:
                pass
        async def _disc(p):
            try:
                async with httpx.AsyncClient(timeout=2.5) as _c:
                    _r = await _c.get(f"http://127.0.0.1:{p}/api/v1/auth/annuaire-public")
                    if _r.status_code == 200:
                        return {"port": p, "sigle": (_r.json().get("sigle") or "").upper().strip()}
            except Exception:
                return None
            return None
        _found = [d for d in await asyncio.gather(*[_disc(pp) for pp in _ports]) if d and d.get("sigle")]
        _bdata = {"origin_sigle": "SUPERVISION", "origin_nom": "Supervision",
                  "sujet": sujet, "contenu": contenu,
                  "origin_username": origin_username or "", "recipient_username": ""}
        _ok = {"n": 0}
        async def _one(d):
            _u = f"http://127.0.0.1:{d['port']}/api/v1/messagerie/ingest"
            _f = [("fichiers", (n, b, m)) for (n, b, m) in _bcast_raw]
            try:
                async with httpx.AsyncClient(timeout=20, verify=False) as cli:
                    r = await cli.post(_u, data=_bdata, files=_f or None)
                    if r.status_code == 200:
                        _ok["n"] += 1
            except Exception:
                pass
        await asyncio.gather(*[_one(d) for d in _found])
        try:
            from app.database import SessionLocal as _S
            from app.models import User as _U
            from plugins.messagerie.models import Message as _M
            from datetime import datetime as _dt, timezone as _tz
            _db = _S()
            try:
                sup = _db.query(_U).filter(_U.username == "supervision").first()
                _db.add(_M(canal="interne", direction="out",
                           expediteur_id=(sup.id if sup else None),
                           expediteur_nom="Supervision", expediteur_addr="SUPERVISION",
                           destinataires=[{"type": "broadcast", "value": "*", "display": "Toutes les instances"}],
                           destinataires_cc=[], destinataires_bcc=[],
                           sujet=sujet, contenu=contenu, contenu_format="plain",
                           statut="sent", lu=True, created_at=_dt.now(_tz.utc)))
                _db.commit()
            finally:
                _db.close()
        except Exception:
            pass
        return {"ok": _ok["n"] > 0, "broadcast": True, "count": _ok["n"], "total": len(_found)}

    async def _probe(p):
        try:
            async with httpx.AsyncClient(timeout=2.5) as _c:
                _r = await _c.get(f"http://127.0.0.1:{p}/api/v1/auth/annuaire-public")
                if _r.status_code == 200:
                    _sig = (_r.json().get("sigle") or "").upper().strip()
                    if _sig == _want:
                        return p
        except Exception:
            return None
        return None
    port = None
    for _p in await asyncio.gather(*[_probe(_pp) for _pp in _ports]):
        if _p is not None:
            port = _p
            break
    if port is None:
        PORT_MAP = {"CHV": "8000", "GHT1": "8001", "GHTSAV": "8002", "GHTAD38": "8003",
                    "CH2": "8002", "CH3": "8003", "CH4": "8004", "CHB": "8005", "CH6": "8006"}
        port = PORT_MAP.get(_want, "8000")
    url = f"http://127.0.0.1:{port}/api/v1/messagerie/ingest"

    files = []
    for up in (fichiers or []):
        try:
            data_bytes = await up.read()
            files.append(("fichiers", (up.filename or "fichier", data_bytes,
                                       up.content_type or "application/octet-stream")))
        except Exception:
            pass
    # Origine : explicite si fournie, sinon dérivée de l'appelant.
    _is_superviseur = (caller == "SUPERVISEUR")
    data = {"origin_sigle": origin_sigle or (caller if not _is_superviseur else "SUPERVISION"),
            "origin_nom":   origin_nom or (caller if not _is_superviseur else "Supervision"),
            "sujet": sujet, "contenu": contenu,
            "origin_username": origin_username or "",
            "recipient_username": recipient_username or ""}

    import httpx
    delivered, detail = False, ""
    try:
        async with httpx.AsyncClient(timeout=20, verify=False) as cli:
            r = await cli.post(url, data=data, files=files or None)
            delivered = (r.status_code == 200)
            detail = f"HTTP {r.status_code} : {(r.text or '')[:160]}"
            logging.getLogger("scribe.collecteur").info(
                "[messagerie] supervision → %s (%s) : %s", target_sigle, url, detail)
    except Exception as e:
        detail = f"réseau : {e}"
        logging.getLogger("scribe.collecteur").warning(
            "[messagerie] livraison supervision → instance ÉCHEC %s : %s", url, e)

    # Copie locale dans 'Envoyés' de la supervision (best-effort)
    try:
        from app.database import SessionLocal as _S
        from app.models import User as _U
        from plugins.messagerie.models import Message as _M
        from datetime import datetime as _dt, timezone as _tz
        _db = _S()
        try:
            sup = _db.query(_U).filter(_U.username == "supervision").first()
            _db.add(_M(canal="interne", direction="out",
                       expediteur_id=(sup.id if sup else None),
                       expediteur_nom="Supervision", expediteur_addr="SUPERVISION",
                       destinataires=[{"type": "instance", "value": target_sigle, "display": target_sigle}],
                       destinataires_cc=[], destinataires_bcc=[],
                       sujet=sujet, contenu=contenu, contenu_format="plain",
                       statut="sent", lu=True, created_at=_dt.now(_tz.utc)))
            _db.commit()
        finally:
            _db.close()
    except Exception as _e:
        logging.getLogger("scribe.collecteur").warning("[messagerie] copie Envoyés KO : %s", _e)

    if not delivered:
        return JSONResponse({"ok": False, "detail": detail}, status_code=502)
    return {"ok": True, "detail": detail, "port": port}


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION CENTRALE (supervision → instances, modèle PULL)
# ─────────────────────────────────────────────────────────────────────────────
try:
    import sys as _sys_ccs, os as _os_ccs
    _sys_ccs.path.insert(0, _os_ccs.path.dirname(_os_ccs.path.abspath(__file__)))
    import central_config_store as _ccs
except Exception as _e_ccs:
    _ccs = None
    logging.getLogger("scribe.collecteur").error("[central] store indisponible : %s", _e_ccs)

last_config_pull: dict = {}  # sigle → horodatage ISO du dernier pull


@app.get("/api/admin/central-config")
async def admin_central_config_get(_ok: bool = Depends(require_admin)):
    """Console admin : config centrale, secrets MASQUÉS + état de propagation."""
    if not _ccs:
        raise HTTPException(500, "magasin de config indisponible")
    masked = _ccs.masked()
    # h146 — aligner bluefiles sur le format instance :
    # ajouter has_login, cli_ready, et un aperçu masqué du login (non-secret)
    bf = masked.get("bluefiles", {})
    clear = _ccs.load_clear().get("bluefiles", {})
    login_val = (clear.get("login") or "").strip()
    pwd_val   = (clear.get("password") or "").strip()
    bf["has_login"]    = bool(login_val)
    bf["has_password"] = bool(pwd_val)
    bf["cli_ready"]    = bool(login_val and pwd_val)
    bf["cli_login_preview"] = (login_val[:3] + "…" if login_val else "")
    bf["cli_password_preview"] = ("•" * min(len(pwd_val), 6) if pwd_val else "")
    masked["bluefiles"] = bf
    return {"config": masked, "propagation": last_config_pull}


@app.post("/api/admin/central-config")
async def admin_central_config_post(payload: dict, _ok: bool = Depends(require_admin)):
    """Met à jour un domaine. Champ secret vide reçu = on conserve l'existant."""
    if not _ccs:
        raise HTTPException(500, "magasin de config indisponible")
    domain = payload.get("domain")
    fields = payload.get("fields") or {}
    try:
        masked = _ccs.save(domain, fields, updated_by="supervision")
    except ValueError as e:
        raise HTTPException(400, str(e))
    logging.getLogger("scribe.collecteur").info("[central] domaine '%s' mis à jour", domain)
    return {"ok": True, "config": masked}


@app.get("/api/central-config")
async def central_config_pull(credentials=Depends(security)):
    """Endpoint TIRÉ par les instances (auth token de nœud). Renvoie les secrets
    EN CLAIR (l'instance en a besoin), uniquement pour les domaines activés."""
    if not _ccs:
        raise HTTPException(500, "magasin de config indisponible")
    tok = credentials.credentials if credentials else ""
    sigle = tokens.get(tok)
    if not sigle:
        raise HTTPException(401, "token de nœud invalide")
    from datetime import datetime, timezone
    last_config_pull[sigle] = datetime.now(timezone.utc).isoformat()
    return {"ok": True, "config": _ccs.for_instance()}


@app.post("/api/admin/sync-configs")
async def admin_sync_configs(credentials=Depends(security)):
    """Pousse la config centrale vers TOUTES les instances découvertes et renvoie
    l'état de synchronisation (joignable + canaux mail/SMS opérationnels) par
    instance. Très utile en GHT : un clic déploie mail/SMS/BlueFiles partout."""
    if not _check_any_auth(credentials):
        raise HTTPException(401)
    import httpx, asyncio
    ports = list(range(8000, 8010)) + [7474]

    async def _probe(p):
        try:
            async with httpx.AsyncClient(timeout=2.5) as c:
                r = await c.get(f"http://127.0.0.1:{p}/api/v1/auth/annuaire-public")
                if r.status_code == 200:
                    j = r.json()
                    return {"port": p, "sigle": (j.get("sigle") or "").upper().strip(),
                            "nom": j.get("nom") or ""}
        except Exception:
            return None
        return None

    discovered = [d for d in await asyncio.gather(*[_probe(p) for p in ports]) if d and d.get("sigle")]

    def _tok_for(sig):
        for t, s in tokens.items():
            if (s or "").upper().strip() == sig:
                return t
        return None

    async def _sync_one(d):
        sig, port = d["sigle"], d["port"]
        out = {"sigle": sig, "nom": d.get("nom") or sig, "port": port,
               "ok": False, "mail": False, "sms": False, "bluefiles": False, "pulled": False, "error": None}
        ftok = _tok_for(sig)
        if not ftok:
            out["error"] = "token de fédération inconnu (instance non enrôlée ?)"
            return out
        try:
            async with httpx.AsyncClient(timeout=12.0) as c:
                r = await c.post(f"http://127.0.0.1:{port}/api/v1/notifications/sync-apply",
                                 headers={"Authorization": "Bearer " + ftok})
            if r.status_code == 200:
                j = r.json(); ch = j.get("channels") or {}
                out["ok"] = True
                out["pulled"] = bool(j.get("pulled"))
                out["mail"] = bool((ch.get("mail") or {}).get("configured"))
                out["sms"] = bool((ch.get("sms") or {}).get("configured"))
                out["bluefiles"] = bool((j.get("bluefiles") or {}).get("configured"))
            else:
                out["error"] = f"HTTP {r.status_code}"
        except Exception as e:
            out["error"] = str(e)[:140]
        return out

    results = await asyncio.gather(*[_sync_one(d) for d in discovered])
    results.sort(key=lambda x: x["sigle"])
    return {"ok": True, "count": len(results), "instances": results}


@app.get("/api/v1/fichiers/p/{jeton}")
async def serve_fiche_reflexe(jeton: str):
    """Sert une fiche réflexe déposée en supervision (lien public à jeton). URL
    volontairement sous « /api/v1/fichiers/ » pour être reconnue par les instances
    comme un partage reçu (« Partagé avec moi »)."""
    from fastapi.responses import FileResponse
    meta = _fiches_load().get(jeton)
    if not meta:
        raise HTTPException(404, "Fiche introuvable")
    p = os.path.join(FICHES_DIR, meta.get("stored", ""))
    if not os.path.isfile(p):
        raise HTTPException(404, "Fichier absent")
    return FileResponse(p, media_type=meta.get("mime") or "application/octet-stream",
                        filename=meta.get("nom") or "fiche")


@app.post("/api/admin/fiche-reflexe")
async def diffuser_fiche_reflexe(
    request: Request,
    fichier: UploadFile = File(...),
    target_sigle: str = Form("*"),       # "*" = toutes les instances, sinon un sigle
    note: str = Form(""),
    credentials=Depends(security),
):
    """Dépose une fiche réflexe et la diffuse : le fichier est stocké en
    supervision et un message porteur du lien absolu est livré à la messagerie de
    chaque instance ciblée → la fiche apparaît dans leur « Partagé avec moi »."""
    if not _check_any_auth(credentials):
        raise HTTPException(401)
    import httpx, asyncio
    data_bytes = await fichier.read()
    if not data_bytes:
        raise HTTPException(400, "Fichier vide")
    nom = fichier.filename or "fiche"
    jeton = secrets.token_urlsafe(20)
    safe = "".join(c for c in nom if c.isalnum() or c in "._-")[:80] or "fiche"
    stored = jeton + "_" + safe
    try:
        with open(os.path.join(FICHES_DIR, stored), "wb") as fh:
            fh.write(data_bytes)
    except Exception as e:
        raise HTTPException(500, f"Stockage impossible : {e}")
    from datetime import datetime as _dt, timezone as _tz
    idx = _fiches_load()
    idx[jeton] = {"nom": nom, "stored": stored,
                  "mime": fichier.content_type or "application/octet-stream",
                  "created_at": _dt.now(_tz.utc).isoformat()}
    _fiches_save(idx)
    base = str(request.base_url).rstrip("/")
    abs_url = f"{base}/api/v1/fichiers/p/{jeton}"
    sujet = f"📁 Fiche réflexe — {nom}"
    corps = ((note.strip() + "\n\n") if note.strip() else "") + \
            f"📁 Fiche réflexe partagée par la supervision : [url={abs_url}]{nom}[/url]"
    _want = (target_sigle or "*").upper().strip()
    _ports = list(range(8000, 8010)) + [7474]
    async def _disc(p):
        try:
            async with httpx.AsyncClient(timeout=2.5) as _c:
                _r = await _c.get(f"http://127.0.0.1:{p}/api/v1/auth/annuaire-public")
                if _r.status_code == 200:
                    s = (_r.json().get("sigle") or "").upper().strip()
                    if _want == "*" or s == _want:
                        return {"port": p, "sigle": s}
        except Exception:
            return None
        return None
    found = [d for d in await asyncio.gather(*[_disc(pp) for pp in _ports]) if d]
    bdata = {"origin_sigle": "SUPERVISION", "origin_nom": "Supervision",
             "sujet": sujet, "contenu": corps, "origin_username": "", "recipient_username": ""}
    sent = {"n": 0}
    async def _one(d):
        u = f"http://127.0.0.1:{d['port']}/api/v1/messagerie/ingest"
        try:
            async with httpx.AsyncClient(timeout=20, verify=False) as cli:
                r = await cli.post(u, data=bdata)
                if r.status_code == 200:
                    sent["n"] += 1
        except Exception:
            pass
    await asyncio.gather(*[_one(d) for d in found])
    return {"ok": sent["n"] > 0, "jeton": jeton, "url": abs_url,
            "count": sent["n"], "total": len(found), "nom": nom}


# ── h145 — BlueFiles supervision : status + envoi ────────────────────────────

def _coll_bf_binary():
    """Chemin du binaire BlueFilesTransfer accessible depuis la supervision.
    Cherche d'abord dans plugins/bluefiles/tools/ (build privé G7),
    puis dans le dossier collecteur/ si copié manuellement."""
    import pathlib
    root = pathlib.Path(__file__).resolve().parent.parent
    candidates = [
        root / "plugins" / "bluefiles" / "tools" / "BlueFilesTransfer",
        pathlib.Path(__file__).resolve().parent / "BlueFilesTransfer",
    ]
    for c in candidates:
        if c.exists() and os.access(str(c), os.X_OK):
            return str(c)
    return None


def _coll_bf_config():
    """Config BlueFiles collecteur (login/password/server) en clair."""
    if not _ccs:
        return None
    cfg = _ccs.load_clear().get("bluefiles", {})
    if not cfg.get("enabled") or not cfg.get("login") or not cfg.get("password"):
        return None
    return cfg


@app.get("/api/admin/bluefiles/status", dependencies=[Depends(require_admin)])
async def coll_bf_status():
    """État du plugin BlueFiles côté supervision."""
    binary = _coll_bf_binary()
    cfg = _coll_bf_config()
    return {
        "binary_present": bool(binary),
        "binary_path":    binary or "",
        "configured":     bool(cfg),
        "has_login":      bool((cfg or {}).get("login")),
        "has_password":   bool((cfg or {}).get("password")),
        "server":         (cfg or {}).get("server", "api.bluefiles.com"),
        "ready":          bool(binary and cfg),
    }


@app.post("/api/admin/bluefiles/send", dependencies=[Depends(require_admin)])
async def coll_bf_send(request: Request):
    """Envoi BlueFiles depuis la supervision.
    Multipart : file(s) + recipients (JSON) + object + message."""
    import tempfile, stat, subprocess
    binary = _coll_bf_binary()
    if not binary:
        raise HTTPException(400, "Binaire BlueFilesTransfer introuvable.")
    cfg = _coll_bf_config()
    if not cfg:
        raise HTTPException(400, "BlueFiles non configuré ou désactivé dans l'administration.")

    form = await request.form()
    recipients_raw = form.get("recipients", "[]")
    obj     = form.get("object", "Document sécurisé — SCRIBE Supervision")
    message = form.get("message", "")

    try:
        recipients = json.loads(recipients_raw) if isinstance(recipients_raw, str) else []
    except Exception:
        recipients = []
    if not recipients:
        raise HTTPException(400, "Au moins un destinataire requis.")

    # Sauvegarder les fichiers uploadés dans un répertoire temporaire
    tmp_dir = tempfile.mkdtemp(prefix="scribe_bf_")
    file_paths = []
    try:
        # Taille max par fichier : 200 Mo
        MAX_SIZE = 200 * 1024 * 1024
        total_size = 0
        file_items = form.getlist("file")
        if not file_items:
            raise HTTPException(400, "Au moins un fichier requis.")
        for upload in file_items:
            if not hasattr(upload, "filename"):
                continue
            safe_name = os.path.basename(upload.filename or "fichier")
            dest = os.path.join(tmp_dir, safe_name)
            content = await upload.read()
            if len(content) > MAX_SIZE:
                raise HTTPException(413, f"Fichier trop volumineux : {safe_name}")
            total_size += len(content)
            if total_size > 4 * 1024 * 1024 * 1024:
                raise HTTPException(413, "Volume total > 4 Go")
            with open(dest, "wb") as fh:
                fh.write(content)
            file_paths.append(dest)
        if not file_paths:
            raise HTTPException(400, "Aucun fichier valide reçu.")

        # Config JSON éphémère
        recs = [{"email": r.get("email", "").strip(),
                  "acknowledge": bool(r.get("acknowledge", False))}
                for r in recipients if r.get("email")]
        if not recs:
            raise HTTPException(400, "Aucun destinataire email valide.")

        import tempfile as _tf, stat as _st
        lf = _tf.NamedTemporaryFile(prefix="bf_log_", suffix=".log", delete=False)
        log_path = lf.name; lf.close()

        cli_cfg = {
            "login":              cfg["login"],
            "password":           cfg["password"],
            "server":             cfg.get("server", "api.bluefiles.com"),
            "object":             obj,
            "message":            message,
            "note":               "",
            "files":              [{"path": p, "name": os.path.basename(p)} for p in file_paths],
            "recipients":         recs,
            "bluepass_mandatory": True,
            "allow_reply":        True,
            "verbose":            False,
            "log_file":           log_path,
        }
        fd, cfg_path = _tf.mkstemp(prefix="bf_cfg_", suffix=".json")
        os.write(fd, json.dumps(cli_cfg, ensure_ascii=False).encode("utf-8"))
        os.close(fd)
        os.chmod(cfg_path, _st.S_IRUSR | _st.S_IWUSR)

        try:
            proc = subprocess.run(
                [binary, "-json", cfg_path],
                capture_output=True, text=True, timeout=1800,
                cwd=os.path.dirname(binary),
            )
            ok = proc.returncode == 0
            # Scrubbing : supprimer login/password de toute sortie
            def _scrub(t):
                import re as _re
                t = _re.sub(re.escape(cfg["login"]), "***", t or "")
                t = _re.sub(re.escape(cfg["password"]), "***", t or "")
                return t
            raw_out = _scrub(proc.stdout or "")
        finally:
            for tmp in (cfg_path, log_path):
                try: os.unlink(tmp)
                except Exception: pass

        logger.info(f"[bluefiles-supervision] envoi {'OK' if ok else 'ERREUR'} → {[r['email'] for r in recs]}")
        if not ok:
            raise HTTPException(502, f"Erreur BlueFiles (code {proc.returncode}). Vérifiez les identifiants.")
        return {"ok": True, "recipients": [r["email"] for r in recs],
                "files": [os.path.basename(p) for p in file_paths]}
    finally:
        # Nettoyage : supprimer les fichiers temporaires
        import shutil
        try: shutil.rmtree(tmp_dir, ignore_errors=True)
        except Exception: pass


@app.get("/api/coll/me")
async def coll_me(credentials=Depends(security)):
    if not credentials: raise HTTPException(401)
    tok = credentials.credentials
    # Accepter ADMIN_TOKEN → retourner utilisateur supervision
    if tok == ADMIN_TOKEN:
        return {"id": 0, "username": "supervision", "display_name": "Supervision", "role": "admin"}
    sess = ui_sessions.get(tok, {})
    if not sess: raise HTTPException(401)
    return {"id": 0, "username": sess.get("login","supervision"),
            "display_name": sess.get("login","Supervision").capitalize(), "role": sess.get("role","viewer")}

@app.get("/api/coll/users")
async def coll_users(credentials=Depends(security)):
    if not _check_any_auth(credentials): raise HTTPException(401)
    auth = load_ui_auth()
    return [{"id": i, "username": u["login"], "display_name": u["login"].capitalize(), "active": True}
            for i, u in enumerate(auth.get("users", []))]

@app.get("/api/coll/fed-status")
async def coll_fed_status(credentials=Depends(security)):
    if not _check_any_auth(credentials): raise HTTPException(401)
    return {"ready": False, "enabled": False, "etablissement": "SUPERVISION"}

@app.post("/api/chat/presence/ping")
async def chat_presence_ping(request: Request, credentials=Depends(security)):
    """Ping de présence pour le chat — accepte tokens établissements ET JWT SCRIBE."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    tok = credentials.credentials if credentials else ""
    # Token établissement enregistré
    sigle = tokens.get(tok)
    if sigle:
        display = body.get("display_name", sigle)
        uid = body.get("user_id", 0)
    else:
        # JWT SCRIBE ou token admin → supervision
        if tok == ADMIN_TOKEN:
            sigle = "SUPERVISION"
            display = "Supervision"
            uid = 0
        elif tok in ui_sessions:
            sigle = "SUPERVISION"
            sess = ui_sessions[tok]
            display = sess.get("login", "supervision").capitalize()
            uid = abs(hash(tok)) % 100000
        else:
            # Token JWT SCRIBE d'une instance non enregistrée → accepter quand même
            # Extraire le sigle du body si disponible
            sigle = body.get("sigle", "INCONNU")
            display = body.get("display_name", sigle)
            uid = body.get("user_id", 0)
    if not sigle:
        return {"ok": True}  # Silencieux si pas de sigle
    now = datetime.now(timezone.utc).isoformat()
    existing = chat_presence.get(sigle, [])
    existing = [u for u in existing if u.get("user_id") != uid]
    existing.append({"user_id": uid, "display_name": display, "last_seen": now})
    chat_presence[sigle] = existing
    return {"ok": True}

@app.get("/api/chat/pj/{pj_id}")
async def serve_chat_pj(pj_id: str, token: str = "", credentials=Depends(security)):
    """Sert une PJ stockée dans le collecteur (inter-GHT)."""
    tok = credentials.credentials if credentials else token
    if not _check_any_auth(credentials) and not (token and (token == ADMIN_TOKEN or token in ui_sessions)):
        raise HTTPException(401)
    pj = chat_pj_store.get(str(pj_id))
    if not pj:
        raise HTTPException(404, "PJ non trouvée")
    # Cas 1 : URL distante → redirect vers l'instance source
    if pj.get("remote_url"):
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=pj["remote_url"], status_code=302)
    # Cas 2 : dataUrl base64
    data_url = pj.get("dataUrl", "")
    if not data_url:
        raise HTTPException(404, "Données PJ manquantes")
    if "," in data_url:
        header, b64 = data_url.split(",", 1)
        mime = header.split(":")[1].split(";")[0] if ":" in header else "application/octet-stream"
    else:
        b64 = data_url
        mime = "application/octet-stream"
    import base64
    from fastapi.responses import Response as FResponse
    content = base64.b64decode(b64)
    return FResponse(content=content, media_type=mime,
                     headers={"Content-Disposition": f"inline; filename={pj['nom']}"})

@app.get("/api/chat/salons/{salon_id}/messages")
async def get_salon_messages_by_id(
    salon_id: int,
    limit: int = 50,
    credentials=Depends(security)
):
    """Récupère les messages d'un salon par ID (compatibilité chat.html)."""
    if not _check_any_auth(credentials):
        raise HTTPException(401)
    # Mapper l'ID vers le nom du salon
    salon_list = list(chat_messages.keys())
    idx = salon_id - 1  # IDs 1-based
    if idx < 0 or idx >= len(salon_list):
        raise HTTPException(404, f"Salon {salon_id} non trouvé")
    salon_nom = salon_list[idx]
    msgs = chat_messages.get(salon_nom, [])
    return msgs[-limit:]

@app.post("/api/chat/salons/{salon_id}/messages")
async def post_salon_message_by_id(
    salon_id: int,
    request: Request,
    credentials=Depends(security)
):
    """Poste un message dans un salon par ID (depuis le chat de supervision)."""
    global _chat_msg_counter
    if not _check_any_auth(credentials):
        raise HTTPException(401)
    # Mapper l'ID vers le nom
    salon_list = list(chat_messages.keys())
    idx = salon_id - 1
    if idx < 0 or idx >= len(salon_list):
        raise HTTPException(404, f"Salon {salon_id} non trouvé")
    salon_nom = salon_list[idx]
    body = await request.json()
    # Déterminer l'auteur depuis la session
    tok = credentials.credentials if credentials else ""
    if tok == ADMIN_TOKEN:
        auteur_nom = "Supervision"
    else:
        sess = ui_sessions.get(tok, {})
        auteur_nom = sess.get("login", "Supervision").capitalize()
    now = datetime.now(timezone.utc).isoformat()
    _chat_msg_counter += 1
    msg = {
        "id":           _chat_msg_counter,
        "salon_nom":    salon_nom,
        "auteur_nom":   auteur_nom,
        "auteur_sigle": "SUPERVISION",
        "contenu":      body.get("contenu", ""),
        "mentions":     body.get("mentions", []),
        "reply_to_id":  body.get("reply_to_id"),
        "reply_to":     None,
        "horodatage":   now,
        "origine":      "ght",
        "pj":           [],
        "supprime":     False,
    }
    if salon_nom not in chat_messages:
        chat_messages[salon_nom] = []
    chat_messages[salon_nom].append(msg)
    return msg

@app.get("/health")
def health():
    return {"status": "ok", "etablissements": len(etablissements),
            "tokens": len(tokens), "pending": len(pending)}


# v3000h29 — Debug endpoint pour Assistant territorial
@app.get("/api/territorial-debug")
async def get_territorial_debug(_ok: bool = Depends(require_admin)):
    """Retourne la structure exacte de l'état utilisé par l'Assistant territorial.
    Utile pour diagnostiquer pourquoi des règles ne se déclenchent pas."""
    debug = {
        "etablissements_count": len(etablissements),
        "etablissements": {},
        "transferts_inter_count": len(transferts_inter),
        "transferts_inter_sample": transferts_inter[:3] if transferts_inter else [],
    }
    for sigle, data in etablissements.items():
        debug["etablissements"][sigle] = {
            "niveau_global":    data.get("niveau_global"),
            "fresh":            data.get("fresh"),
            "age_minutes":      data.get("age_minutes"),
            "timestamp":        data.get("timestamp"),
            "received_at":      data.get("received_at"),
            "kpis":             data.get("kpis", {}),
            "incidents_count":  len(data.get("incidents", [])),
            "incidents_sample": (data.get("incidents") or [])[:3],
            "declarations_count": len(data.get("declarations", [])),
            "declarations_sample": (data.get("declarations") or [])[:2],
            "all_keys":         sorted(data.keys()) if isinstance(data, dict) else [],
        }
    return debug


# v3000h25 — Assistant territorial (vue agrégée Territoire)
@app.get("/api/territorial-assistant")
async def get_territorial_assistant():
    """v3000h25 — Assistant de supervision territoriale.

    Évalue les 5 règles territoriales (RT1-RT5) sur la vue agrégée des
    établissements Territoire et retourne les alertes + un résumé d'état.

    Sans auth pour faciliter la supervision (la vue agrégée n'expose
    pas de données patients nominatives). Si on veut auth plus tard,
    ajouter Depends(security) + _check_any_auth(credentials).
    """
    try:
        # Tentative import : module local au collecteur/ (copié depuis exercice)
        from territorial_assistant import evaluate_territorial_rules
    except ImportError:
        try:
            from collecteur.territorial_assistant import evaluate_territorial_rules
        except ImportError as e:
            return {
                "summary": {},
                "alertes": [],
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "error": f"module territorial_assistant introuvable : {e}",
            }
    try:
        result = evaluate_territorial_rules(etablissements, transferts_inter)
        return result
    except Exception as e:
        import logging
        logging.getLogger("scribe.collecteur").error(f"territorial_assistant: {e}")
        return {
            "summary": {},
            "alertes": [],
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "error": str(e),
        }


@app.get("/debug")
def debug(_ok: bool = Depends(require_admin)):
    """Diagnostic — tokens, données, état complet."""
    return {
        "tokens_count":       len(tokens),
        "tokens_sigles":      list(tokens.values()),
        "etablissements":     list(etablissements.keys()),
        "pending_count":      len(pending),
        "pending_sigles":     [v.get("sigle_propose","?") for v in pending.values()],
        "data_file":          str(Path(DATA_FILE).resolve()),
        "tokens_file":        str(Path(TOKENS_FILE).resolve()),
        "tokens_file_exists": Path(TOKENS_FILE).exists(),
        "data_file_exists":   Path(DATA_FILE).exists(),
    }


# ── Auth interface web (optionnelle) ─────────────────────────────────────
UI_AUTH_FILE = "collecteur_ui_auth.json"

def load_ui_auth() -> dict:
    """Charge la config login UI depuis le fichier JSON si présent."""
    if Path(UI_AUTH_FILE).exists():
        try:
            return json.loads(Path(UI_AUTH_FILE).read_text())
        except Exception:
            pass
    # Login par défaut si pas de fichier
    import hashlib
    return {
        "login": "supervision",
        "password_hash": hashlib.sha256("Scribe2026!".encode()).hexdigest(),
        "users": [{"login": "supervision", "role": "admin", "password_hash": "0a0da7eef0453b6cbd142fcf25f7ac63081c9cb920cba999c1a3a80d1f25dfda"}]
    }

def check_ui_credentials(login: str, password: str) -> bool:
    """Vérifie login/mot de passe de l'interface web."""
    auth = load_ui_auth()
    if not auth:
        return True
    import hashlib
    h = hashlib.sha256(password.encode()).hexdigest()
    # Vérifier dans la liste users si elle existe
    users = auth.get("users", [])
    if users:
        return any(u.get("login") == login and u.get("password_hash") == h for u in users)
    # Fallback legacy
    return auth.get("login") == login and auth.get("password_hash") == h




# ════════════════════════════════════════════════════════════════════════════
# GESTION COMPTES SUPERVISION
# ════════════════════════════════════════════════════════════════════════════

def save_ui_auth(auth: dict):
    Path(UI_AUTH_FILE).write_text(json.dumps(auth, ensure_ascii=False, indent=2))

@app.get("/api/ui/users")
async def list_ui_users(credentials=Depends(security)):
    """Liste les comptes de l'interface web."""
    if not require_ui_admin(credentials):
        raise HTTPException(401)
    auth = load_ui_auth()
    users = auth.get("users", [{"login": auth.get("login","supervision"), "role":"admin"}])
    return [{"login": u["login"], "role": u.get("role","viewer")} for u in users]

@app.post("/api/ui/users")
async def create_ui_user(request: Request, credentials=Depends(security)):
    """Crée un compte supervision."""
    if not require_ui_admin(credentials):
        raise HTTPException(401)
    body = await request.json()
    login = body.get("login","").strip()
    password = body.get("password","").strip()
    role = body.get("role","viewer")
    if not login or not password:
        raise HTTPException(400, "Login et mot de passe requis")
    auth = load_ui_auth()
    users = auth.get("users", [{"login": auth.get("login","supervision"), "role":"admin",
                                "password_hash": auth.get("password_hash","")}])
    if any(u["login"] == login for u in users):
        raise HTTPException(409, "Login déjà utilisé")
    import hashlib
    users.append({"login": login, "role": role,
                  "password_hash": hashlib.sha256(password.encode()).hexdigest()})
    auth["users"] = users
    save_ui_auth(auth)
    return {"ok": True}

@app.delete("/api/ui/users/{login}")
async def delete_ui_user(login: str, credentials=Depends(security)):
    """Supprime un compte supervision."""
    if not require_ui_admin(credentials):
        raise HTTPException(401)
    auth = load_ui_auth()
    users = auth.get("users", [])
    auth["users"] = [u for u in users if u["login"] != login]
    save_ui_auth(auth)
    return {"ok": True}

@app.get("/api/admin/users/template.xlsx")
async def users_template_xlsx(credentials=Depends(security)):
    """Modèle Excel téléchargeable pour l'import de comptes supervision."""
    if not require_ui_admin(credentials):
        raise HTTPException(401)
    import openpyxl, io
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Comptes"
    ws.append(["login", "role", "mot_de_passe", "telephone", "email"])
    ws.append(["jdupont", "viewer", "", "+33612345678", "j.dupont@chx.fr"])
    ws.append(["cadre.crise", "admin", "ÀChanger!2026", "+33698765432", "cadre@chx.fr"])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 28
    aide = wb.create_sheet("Aide")
    aide.append(["Colonne", "Description"])
    aide.append(["login", "Identifiant de connexion — obligatoire, unique"])
    aide.append(["role", "admin ou viewer (par défaut : viewer)"])
    aide.append(["mot_de_passe", "Laisser vide = mot de passe temporaire généré automatiquement"])
    aide.append(["telephone", "Numéro E.164 pour les SMS (ex. +33612345678) — optionnel"])
    aide.append(["email", "Email pour les notifications mail — optionnel"])
    aide.column_dimensions["A"].width = 16
    aide.column_dimensions["B"].width = 64
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modele_comptes_supervision.xlsx"},
    )


@app.post("/api/admin/users/import")
async def users_import_xlsx(file: UploadFile = File(...), credentials=Depends(security)):
    """Import en masse de comptes depuis un .xlsx (colonnes login/role/mot_de_passe).
    Login existant → ignoré. Mot de passe vide → temporaire généré (renvoyé)."""
    if not require_ui_admin(credentials):
        raise HTTPException(401)
    import openpyxl, io, hashlib, secrets as _secrets
    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Fichier Excel illisible : {e}")
    auth = load_ui_auth()
    users = auth.get("users", [])
    existing = {u.get("login") for u in users}
    created, skipped, temp = [], [], []
    rows = list(ws.iter_rows(values_only=True))
    start = 0
    if rows and rows[0] and str(rows[0][0] or "").strip().lower() in ("login", "identifiant"):
        start = 1
    for r in rows[start:]:
        if not r or not r[0]:
            continue
        login = str(r[0]).strip()
        if not login:
            continue
        role = (str(r[1]).strip().lower() if len(r) > 1 and r[1] else "viewer")
        if role not in ("admin", "viewer"):
            role = "viewer"
        pwd = (str(r[2]).strip() if len(r) > 2 and r[2] else "")
        telephone = (str(r[3]).strip() if len(r) > 3 and r[3] else "")
        email = (str(r[4]).strip() if len(r) > 4 and r[4] else "")
        if login in existing:
            skipped.append(login)
            continue
        if not pwd:
            pwd = _secrets.token_urlsafe(9)
            temp.append({"login": login, "password": pwd})
        users.append({"login": login, "role": role,
                      "password_hash": hashlib.sha256(pwd.encode()).hexdigest(),
                      "telephone": telephone or None, "email": email or None})
        existing.add(login)
        created.append(login)
    auth["users"] = users
    save_ui_auth(auth)
    logging.getLogger("scribe.collecteur").info("[comptes] import Excel : %d créés, %d ignorés", len(created), len(skipped))
    return {"ok": True, "created": created, "skipped": skipped, "temp_passwords": temp}


@app.post("/api/ui/users/change-password")
async def admin_change_ui_password(request: Request, credentials=Depends(security)):
    """v3.4 (h38f) — Endpoint admin : permet à un admin de changer le mdp
    d'un AUTRE compte supervision. Renommé depuis /api/ui/change-password
    pour ne plus entrer en collision avec l'endpoint utilisateur self-service
    qui permet à un user de changer SON PROPRE mdp (cf. plus bas).
    """
    if not require_ui_admin(credentials):
        raise HTTPException(401)
    body = await request.json()
    login = body.get("login","").strip()
    new_pass = body.get("new_password","").strip()
    if not login or not new_pass:
        raise HTTPException(400, "Login et nouveau mot de passe requis")
    auth = load_ui_auth()
    import hashlib
    h = hashlib.sha256(new_pass.encode()).hexdigest()
    users = auth.get("users", [{"login": auth.get("login",""), "role":"admin",
                                "password_hash": auth.get("password_hash","")}])
    for u in users:
        if u["login"] == login:
            u["password_hash"] = h
            break
    else:
        raise HTTPException(404, "Utilisateur introuvable")
    auth["users"] = users
    # Compatibilité legacy
    if auth.get("login") == login:
        auth["password_hash"] = h
    save_ui_auth(auth)
    return {"ok": True}

@app.post("/api/ui/login")
async def ui_login(request: Request):
    """Authentification interface web du collecteur."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "JSON invalide")
    login = body.get("login","")
    password = body.get("password","")
    auth = load_ui_auth()
    if not auth:
        return {"ok": True, "token": "no-auth", "must_change_password": False}  # Pas de protection
    if check_ui_credentials(login, password):
        session_token = secrets.token_hex(16)
        # Stocker la session en mémoire
        auth_info = load_ui_auth()
        users = auth_info.get("users", [])
        role = next((u.get("role","viewer") for u in users if u.get("login")==login), "viewer")
        ui_sessions[session_token] = {"login": login, "role": role}
        # v3.4 (h38e) — Détection du mdp par défaut "Scribe2026!".
        # Si l'utilisateur se logge encore avec le mdp d'installation, on
        # force le changement à la première connexion. Le hash SHA-256 du
        # mdp par défaut est constant : on compare directement.
        import hashlib
        default_hash = hashlib.sha256("Scribe2026!".encode()).hexdigest()
        must_change = False
        user_record = next((u for u in users if u.get("login")==login), None)
        if user_record:
            current_hash = user_record.get("password_hash", "")
            # Si l'utilisateur a un flag explicite must_change_password=False,
            # on le respecte (cas : il a déjà changé puis remis le default
            # volontairement, ou il a explicitement opt-out).
            explicit_flag = user_record.get("must_change_password")
            if explicit_flag is True:
                must_change = True
            elif explicit_flag is False:
                must_change = False
            else:
                # Pas de flag explicite : on force si mdp = défaut
                must_change = (current_hash == default_hash)
        return {"ok": True, "token": session_token, "must_change_password": must_change}
    raise HTTPException(status_code=401, detail="Identifiants invalides")


@app.post("/api/ui/change-password")
async def ui_change_password(request: Request, credentials=Depends(security)):
    """v3.4 (h38e) — Changement de mot de passe par l'utilisateur lui-même.
    Utilisé pour le flow "première connexion master" : si l'utilisateur a
    le mdp par défaut (Scribe2026!), l'UI lui présente une modale de
    changement obligatoire AVANT d'entrer dans l'application.
    """
    if not credentials:
        raise HTTPException(401)
    tok = credentials.credentials
    sess = ui_sessions.get(tok)
    if not sess and tok != ADMIN_TOKEN:
        raise HTTPException(401, "Session invalide")
    login = sess["login"] if sess else "admin"
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "JSON invalide")
    current = body.get("current_password", "")
    new = body.get("new_password", "")
    if len(new) < 8:
        raise HTTPException(400, "Le nouveau mot de passe doit faire au moins 8 caractères")
    if new == current:
        raise HTTPException(400, "Le nouveau mot de passe doit être différent de l'ancien")
    # Vérifier le mdp actuel
    if not check_ui_credentials(login, current):
        raise HTTPException(400, "Mot de passe actuel incorrect")
    # Mettre à jour
    import hashlib
    auth = load_ui_auth()
    users = auth.get("users", [])
    for u in users:
        if u.get("login") == login:
            u["password_hash"] = hashlib.sha256(new.encode()).hexdigest()
            u["must_change_password"] = False
            break
    else:
        raise HTTPException(404, "Utilisateur introuvable")
    auth["users"] = users
    save_ui_auth(auth)
    logger.info(f"ui_password_change: login={login}")
    return {"ok": True, "must_change_password": False}


@app.get("/api/ui/auth-required")
def auth_required():
    """Indique si l'interface nécessite une authentification."""
    auth = load_ui_auth()
    return {"required": bool(auth), "login": auth.get("login","") if auth else ""}

@app.get("/api/ui/first-launch")
def first_launch():
    """Indique si c'est le tout premier lancement (aucune instance configurée
    et aucun compte UI custom n'a été créé). Utilisé pour afficher le hint
    de credentials par défaut sur l'écran de login."""
    # Master/onboarding pas fait
    onboarding_flag = Path("master/.onboarding_done")
    onboarding_done = onboarding_flag.exists()
    # Pas de compte UI custom (juste le compte par défaut supervision/Scribe2026!)
    has_custom_users = Path(UI_AUTH_FILE).exists()
    # Premier lancement = onboarding pas fait ET pas de compte custom
    return {"first_launch": (not onboarding_done) and (not has_custom_users)}

@app.get("/api/ui/verify")
def verify_session(credentials=Depends(security)):
    """Vérifie qu'un token de session est toujours valide."""
    if not credentials:
        raise HTTPException(401)
    tok = credentials.credentials
    if tok == ADMIN_TOKEN or tok in ui_sessions:
        sess = ui_sessions.get(tok, {"login": "admin", "role": "admin"})
        return {"ok": True, "login": sess.get("login",""), "role": sess.get("role","admin")}
    raise HTTPException(401, "Session expirée")


# ── Démarrage ──────────────────────────────────────────────────────────────

# ── Tokens Territoire démo — enregistrés automatiquement si tokens vides ─────
# h153 — Tokens pré-enrôlés retirés : ils étaient spécifiques à un déploiement
# G7 et n'ont pas leur place dans le code public.
# Les instances s'enrôlent dynamiquement via /api/admin/pending → accept.
scribe_TOKENS = {}

if __name__ == "__main__":
    load_tokens()
    load_data()
    load_pending()
    load_messages_inter()
    load_relay()

    # Tokens: enrôlement MANUEL via l'UI (⏳ EN ATTENTE → ✓ ACCEPTER)
    # Le bouton "⚡ Enregistrer" dans l'UI force l'enregistrement si besoin
    if tokens:
        print(f"  ✓ Tokens enregistrés : {list(tokens.values())}")
    else:
        print(f"  ℹ Aucun token — les GHTs apparaîtront en ⏳ EN ATTENTE")
        print(f"  → Ouvrir http://localhost:9000 et cliquer ✓ ACCEPTER")
        print(f"  → OU cliquer ⚡ Enregistrer (section TOKENS Territoire)")

    nb_etab = len(tokens)
    nb_data  = len(etablissements)

    print("\n  ╔══════════════════════════════════════════════╗")
    print("  ║  SCRIBE Collecteur territorial  v1.2.1       ║")
    print("  ╚══════════════════════════════════════════════╝")
    print(f"\n  Dashboard     : http://0.0.0.0:9000")
    print(f"  Etablissements: {nb_etab} token(s) / {nb_data} remontée(s)")
    print(f"\n  Token admin   : {ADMIN_TOKEN}")
    print(f"  (persistant dans {ADMIN_FILE} — identique à chaque redémarrage)\n")
    if nb_etab == 0:
        print("  ► Aucun établissement enregistré.")
        print("  → Les GHT qui poussent arrivent en section ⏳ EN ATTENTE")
        print("  → Ouvrir http://localhost:9000 et cliquer ✓ ACCEPTER\n")
        print("  Tokens Territoire démo :")
        print("    CHV    : token_demo_a_changer")
        print("    GHT1  : token_demo")
        print("    GHTSAV  : token_ghtsav_demo_2026")
        print("    GHTAD38 : token_ghtad38_demo_2026\n")
    else:
        etabs = list(set(tokens.values()))
        print(f"  ► Etablissements actifs : {', '.join(etabs)}")
        if pending:
            print(f"  ► EN ATTENTE : {len(pending)} établissement(s) — valider sur http://localhost:9000\n")
        else:
            print()

    import os as _os
    _port = int(_os.environ.get("COLLECTEUR_PORT", 9000))
    uvicorn.run("collecteur:app", host="0.0.0.0", port=_port, reload=False)
