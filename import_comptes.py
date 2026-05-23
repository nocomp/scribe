"""
import_comptes.py — Import de comptes utilisateurs en lot pour SCRIBE
Usage :
  python import_comptes.py <fichier.xlsx>                  # DB par défaut
  python import_comptes.py <fichier.xlsx> --port 8001      # DB de l'instance sur port 8001
  python import_comptes.py <fichier.xlsx> --db <path>      # DB explicite

Fonctionnement :
  - Lit le fichier Excel (onglet 'Comptes utilisateurs')
  - Crée les comptes absents, met à jour les existants
  - Mot de passe temporaire : changeme  (changement obligatoire à la première connexion)
  - Les comptes admin existants ne sont jamais écrasés

v2.4.8.2 — IMPORTANT pour les utilisateurs du master :
  Quand SCRIBE tourne via le master, CHAQUE INSTANCE A SA PROPRE DB sous
  data/instances/<PORT>/scribe.db. Sans option --port, ce script écrit dans
  scribe.db à la racine (DB orpheline, jamais lue par les instances).
  → Utiliser --port pour cibler une vraie instance.
"""

import hashlib
import sys
import os

# Ajout du répertoire courant au path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ─────────────────────────────────────────────────────────────────────────
# v2.4.8.2 : résolution de la DB cible AVANT d'importer app.database
# (sinon trop tard, l'env DATABASE_URL n'est plus pris en compte)
# ─────────────────────────────────────────────────────────────────────────
def _resolve_db_url(argv):
    """Cherche --port N ou --db PATH dans argv et set DATABASE_URL.
    Retourne (db_url, xlsx_path) ou affiche aide et sort."""
    xlsx_path = None
    db_path = None
    port = None
    i = 1
    while i < len(argv):
        a = argv[i]
        if a == "--port" and i + 1 < len(argv):
            port = int(argv[i + 1]); i += 2; continue
        if a == "--db" and i + 1 < len(argv):
            db_path = argv[i + 1]; i += 2; continue
        if a in ("-h", "--help"):
            print(__doc__); sys.exit(0)
        if not a.startswith("-"):
            xlsx_path = a; i += 1; continue
        print(f"  Option inconnue : {a}"); sys.exit(1)
    if not xlsx_path:
        print(__doc__); sys.exit(1)
    # Si --port, résoudre vers data/instances/<port>/scribe.db
    if port is not None and db_path is None:
        base = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base, "data", "instances", str(port), "scribe.db")
        if not os.path.exists(db_path):
            print(f"  ERREUR : DB introuvable pour le port {port} :")
            print(f"    {db_path}")
            print(f"  → L'instance a-t-elle été démarrée au moins une fois via le master ?")
            sys.exit(2)
    if db_path:
        db_url = "sqlite:///" + os.path.abspath(db_path)
        os.environ["DATABASE_URL"] = db_url
        print(f"  → DB cible : {db_path}")
    else:
        # Pas de --port ni --db : avertir si on est dans un contexte master
        base = os.path.dirname(os.path.abspath(__file__))
        instances_dir = os.path.join(base, "data", "instances")
        if os.path.isdir(instances_dir) and os.listdir(instances_dir):
            print()
            print("  ⚠  ATTENTION : des instances master existent sous data/instances/")
            print("     mais aucune option --port n'est fournie.")
            print("     Ce script va écrire dans scribe.db à la racine du projet,")
            print("     qui N'EST PAS la DB utilisée par les instances du master.")
            print("     → Utilisez --port <NUMERO> pour cibler une instance précise.")
            print("     → Ou --db <CHEMIN.db> pour un chemin explicite.")
            print()
            try:
                ans = input("  Continuer quand même (DB racine) ? [y/N] ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                ans = "n"
            if ans not in ("y", "yes", "o", "oui"):
                print("  Annulé."); sys.exit(0)
    return xlsx_path

XLSX_PATH = _resolve_db_url(sys.argv)

try:
    import openpyxl
except ImportError:
    print("  ERREUR: openpyxl requis — pip install openpyxl")
    sys.exit(1)

from app.database import SessionLocal, engine, Base
import app.models  # noqa — crée les tables
Base.metadata.create_all(bind=engine)

from app.models import User

# Mot de passe temporaire
TEMP_PASSWORD = "changeme"
TEMP_HASH     = hashlib.sha256(TEMP_PASSWORD.encode()).hexdigest()

VALID_ROLES = {"admin", "directeur", "observateur"}

COLUMNS = {
    "nom":              None,
    "prenom":           None,
    "username":         None,
    "role":             None,
    "unite_soin":       None,
    "telephone_fixe":   None,
    "telephone_mobile": None,
    "email":            None,
    "perimetre":        None,
}


def _h(s):
    return hashlib.sha256(str(s).encode()).hexdigest()


def read_excel(path: str):
    """Lit le fichier Excel et retourne une liste de dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if "compte" in name.lower() or "utilisateur" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    # Lire les en-têtes (ligne 1)
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            # Normaliser : mettre en minuscules, remplacer espaces/accents
            key = str(val).lower().strip()
            key = key.replace("é", "e").replace("è", "e").replace("ê", "e")
            key = key.replace("â", "a").replace("î", "i").replace("ô", "o").replace("û", "u")
            key = key.replace(" ", "_").replace("/", "_").replace(".", "_")
            # Correspondances flexibles
            if "prenom" in key or "pr\u00e9nom" in key:     key = "prenom"
            elif "nom" in key and "unite" not in key:        key = "nom"
            elif "username" in key or "identif" in key or "login" in key: key = "username"
            elif "role" in key:                              key = "role"
            elif "unite" in key or "service" in key:        key = "unite_soin"
            elif "mobile" in key or "gsm" in key:           key = "telephone_mobile"
            elif "fixe" in key or "interne" in key:         key = "telephone_fixe"
            elif "email" in key or "mail" in key:           key = "email"
            elif "perimetre" in key or "p\u00e9rim" in key: key = "perimetre"
            headers[col] = key

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row = {}
        has_data = False
        for col, key in headers.items():
            val = ws.cell(row=row_idx, column=col).value
            if val is not None:
                val = str(val).strip()
                has_data = True
            else:
                val = ""
            row[key] = val
        if has_data and row.get("username"):
            rows.append(row)

    return rows


def import_comptes(path: str):
    print(f"\n  SCRIBE — Import de comptes")
    print(f"  ════════════════════════════")
    print(f"  Fichier : {path}\n")

    if not os.path.exists(path):
        print(f"  ERREUR: Fichier introuvable : {path}")
        sys.exit(1)

    rows = read_excel(path)
    if not rows:
        print("  ERREUR: Aucune ligne trouvée. Vérifiez que l'onglet 'Comptes utilisateurs' existe et contient des données.")
        sys.exit(1)

    print(f"  {len(rows)} ligne(s) lue(s) dans le fichier.\n")

    db = SessionLocal()
    created = 0
    updated = 0
    skipped = 0
    errors  = []

    for i, row in enumerate(rows, 1):
        username = row.get("username", "").strip().lower().replace(" ", "")
        nom      = row.get("nom", "").strip()
        prenom   = row.get("prenom", "").strip()
        role     = row.get("role", "directeur").strip().lower()
        unite    = row.get("unite_soin", "")
        tel_fixe = row.get("telephone_fixe", "")
        tel_mob  = row.get("telephone_mobile", "")
        email    = row.get("email", "")
        perim    = row.get("perimetre", "")

        # Validations
        if not username:
            errors.append(f"Ligne {i+1}: username vide — ignorée")
            continue
        if not nom and not prenom:
            errors.append(f"Ligne {i+1}: nom et prénom vides pour '{username}' — ignorée")
            continue
        if role not in VALID_ROLES:
            role = "directeur"
            errors.append(f"Ligne {i+1}: rôle invalide pour '{username}' → 'directeur' utilisé")

        display_name = f"{prenom} {nom}".strip() or username

        # Vérifier si le compte existe
        existing = db.query(User).filter(User.username == username).first()

        if existing:
            # Ne jamais écraser un admin existant si on essaie de le dégrader
            if existing.role == "admin" and role != "admin":
                skipped += 1
                print(f"  ⊘  {username:20} — admin existant, rôle non modifié")
                continue
            # Mettre à jour sans toucher au mot de passe
            existing.display_name    = display_name
            existing.role            = role
            existing.perimetre       = perim or None
            existing.active          = True
            # Stocker infos contact dans perimetre si pas de champ dédié
            # (champs étendus stockés dans display_name enrichi pour compatibilité)
            db.commit()
            updated += 1
            print(f"  ↻  {username:20} — mis à jour ({role})")
        else:
            # Créer le compte avec mdp temporaire + flag must_change
            user = User(
                username=username,
                display_name=display_name,
                role=role,
                hashed_password=TEMP_HASH,
                perimetre=perim or None,
                active=True,
                must_change_password=True,  # forcer changement à la connexion
            )
            db.add(user)
            try:
                db.commit()
                created += 1
                print(f"  ✓  {username:20} — créé ({role}) — mdp temporaire: {TEMP_PASSWORD}")
            except Exception as e:
                db.rollback()
                errors.append(f"Ligne {i+1}: Erreur création '{username}': {e}")

    db.close()

    print(f"\n  ════════════════════════════")
    print(f"  ✓ Créés  : {created}")
    print(f"  ↻ Mis à jour : {updated}")
    print(f"  ⊘ Ignorés : {skipped}")
    if errors:
        print(f"\n  Avertissements :")
        for e in errors:
            print(f"    ! {e}")
    print(f"\n  Mot de passe temporaire : {TEMP_PASSWORD}")
    print(f"  → Les nouveaux comptes devront changer leur mot de passe à la première connexion.\n")


if __name__ == "__main__":
    # v2.4.8.2 : XLSX_PATH est résolu en début de fichier par _resolve_db_url
    import_comptes(XLSX_PATH)
