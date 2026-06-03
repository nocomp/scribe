"""
app/api/v140.py — SCRIBE v1.4.0
Nouvelles fonctionnalités :
  - Capacité médico-technique (blocs, dialyse, pharmacie)
  - Chat IA contextuel sur les analyses
  - Messagerie interne
  - Déclarations de situation inter-GHT
  - Demandes inter-GHT
  - Changement de mot de passe utilisateur
  - Endpoint supervision /api/v1/supervision/status (pour le collecteur)
"""

import hashlib
import json
import logging
from datetime import datetime, timezone
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    CapaciteMedicotech, IASession, MessageInterne,
    DeclarationSituation, DemandeInterGHT, User,
    SitrepEntry, CapaciteDeclaration, CapaciteReferentiel
)
from app.api.auth import get_current_user, require_admin

logger = logging.getLogger("scribe.v140")
router = APIRouter()

def _log_mc(db, user, categorie: str, action: str, detail: str = "", ref_id: int = None, ref_type: str = None, site: str = None, niveau: str = "INFO"):
    """Log un événement dans la main courante exhaustive."""
    try:
        from app.models import MainCouranteLog
        db.add(MainCouranteLog(
            auteur=getattr(user,'display_name',None) or getattr(user,'username','Système') if user else 'Système',
            auteur_role=getattr(user,'role','système') if user else 'système',
            categorie=categorie, action=action, detail=detail,
            ref_id=ref_id, ref_type=ref_type, site=site, niveau=niveau,
        ))
        db.flush()
    except Exception:
        pass


# ════════════════════════════════════════════════════════════
#  CHANGEMENT MOT DE PASSE UTILISATEUR
# ════════════════════════════════════════════════════════════

class ChangePasswordIn(BaseModel):
    ancien_mdp: str
    nouveau_mdp: str

@router.post("/auth/change-password")
def change_password(
    body: ChangePasswordIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not user:
        raise HTTPException(401, "Non authentifié")
    if len(body.nouveau_mdp) < 8:
        raise HTTPException(400, "Le mot de passe doit faire au moins 8 caractères")
    # Vérifier l'ancien mdp en supportant bcrypt ET sha256 (migration transparente)
    import bcrypt as _bcrypt, hashlib as _hl
    stored = user.hashed_password or ""
    def _sha(pw): return _hl.sha256(pw.encode()).hexdigest()
    if stored.startswith("$2b$") or stored.startswith("$2a$"):
        # bcrypt
        ok = _bcrypt.checkpw(body.ancien_mdp.encode(), stored.encode())
    else:
        ok = stored == _sha(body.ancien_mdp)
    if not ok:
        raise HTTPException(400, "Ancien mot de passe incorrect")
    # Hacher le nouveau mdp en bcrypt
    user.hashed_password = _bcrypt.hashpw(body.nouveau_mdp.encode(), _bcrypt.gensalt()).decode()
    if hasattr(user, 'must_change_password'):
        user.must_change_password = False
    db.commit()
    return {"status": "ok"}


# ════════════════════════════════════════════════════════════
#  CAPACITÉ MÉDICO-TECHNIQUE
# ════════════════════════════════════════════════════════════

class MedicotechIn(BaseModel):
    site_id: str
    blocs_total: int = 0
    blocs_operationnels: int = 0
    blocs_commentaire: Optional[str] = None
    dialyse_postes_total: int = 0
    dialyse_postes_actifs: int = 0
    dialyse_sessions_24h: int = 0
    dialyse_commentaire: Optional[str] = None
    pharmacie_statut: str = "normal"
    pharmacie_urgences_vitales: bool = True
    pharmacie_commentaire: Optional[str] = None

@router.get("/medicotech")
def get_medicotech(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not user:
        raise HTTPException(401)
    rows = db.query(CapaciteMedicotech).order_by(CapaciteMedicotech.updated_at.desc()).all()
    return [_mt_to_dict(r) for r in rows]

@router.post("/medicotech")
def upsert_medicotech(
    body: MedicotechIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    existing = db.query(CapaciteMedicotech).filter(
        CapaciteMedicotech.site_id == body.site_id
    ).first()
    if existing:
        existing.blocs_total = body.blocs_total
        existing.blocs_operationnels = body.blocs_operationnels
        existing.blocs_commentaire = body.blocs_commentaire
        existing.dialyse_postes_total = body.dialyse_postes_total
        existing.dialyse_postes_actifs = body.dialyse_postes_actifs
        existing.dialyse_sessions_24h = body.dialyse_sessions_24h
        existing.dialyse_commentaire = body.dialyse_commentaire
        existing.pharmacie_statut = body.pharmacie_statut
        existing.pharmacie_urgences_vitales = body.pharmacie_urgences_vitales
        existing.pharmacie_commentaire = body.pharmacie_commentaire
        existing.updated_by = user.display_name
    else:
        mt = CapaciteMedicotech(
            site_id=body.site_id,
            blocs_total=body.blocs_total,
            blocs_operationnels=body.blocs_operationnels,
            blocs_commentaire=body.blocs_commentaire,
            dialyse_postes_total=body.dialyse_postes_total,
            dialyse_postes_actifs=body.dialyse_postes_actifs,
            dialyse_sessions_24h=body.dialyse_sessions_24h,
            dialyse_commentaire=body.dialyse_commentaire,
            pharmacie_statut=body.pharmacie_statut,
            pharmacie_urgences_vitales=body.pharmacie_urgences_vitales,
            pharmacie_commentaire=body.pharmacie_commentaire,
            updated_by=user.display_name
        )
        db.add(mt)
    db.commit()
    return {"status": "ok"}

def _mt_to_dict(r):
    return {
        "id": r.id, "site_id": r.site_id,
        "blocs_total": r.blocs_total, "blocs_operationnels": r.blocs_operationnels,
        "blocs_commentaire": r.blocs_commentaire,
        "dialyse_postes_total": r.dialyse_postes_total,
        "dialyse_postes_actifs": r.dialyse_postes_actifs,
        "dialyse_sessions_24h": r.dialyse_sessions_24h,
        "dialyse_commentaire": r.dialyse_commentaire,
        "pharmacie_statut": r.pharmacie_statut,
        "pharmacie_urgences_vitales": r.pharmacie_urgences_vitales,
        "pharmacie_commentaire": r.pharmacie_commentaire,
        "updated_by": r.updated_by,
        "updated_at": r.updated_at.isoformat() if r.updated_at else None
    }


# ════════════════════════════════════════════════════════════
#  CHAT IA CONTEXTUEL
# ════════════════════════════════════════════════════════════

class IASaveIn(BaseModel):
    incident_id: Optional[int] = None
    statut_genere: str

class IAChatIn(BaseModel):
    session_id: int
    message: str

@router.post("/ia/session")
def save_ia_session(
    body: IASaveIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Enregistre ou remplace l'analyse IA générée, reset le chat."""
    if not user:
        raise HTTPException(401)
    if body.incident_id:
        existing = db.query(IASession).filter(
            IASession.incident_id == body.incident_id
        ).first()
        if existing:
            existing.statut_genere = body.statut_genere
            existing.historique_chat = "[]"
            db.commit()
            return {"session_id": existing.id}
    sess = IASession(
        incident_id=body.incident_id,
        statut_genere=body.statut_genere,
        historique_chat="[]"
    )
    db.add(sess)
    db.commit()
    db.refresh(sess)
    return {"session_id": sess.id}

@router.post("/ia/chat")
async def ia_chat(
    body: IAChatIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Échange de chat sur l'analyse IA générée."""
    if not user:
        raise HTTPException(401)
    sess = db.query(IASession).filter(IASession.id == body.session_id).first()
    if not sess:
        raise HTTPException(404, "Session IA introuvable")

    historique = json.loads(sess.historique_chat or "[]")
    historique.append({"role": "user", "content": body.message})

    # Appel IA via le routeur existant
    try:
        from app.api.ai_router import call_ai
        system = (
            "Tu es un assistant de gestion de crise hospitalière. "
            "Réponds de manière concise et opérationnelle en français."
        )
        ctx = f"Contexte — analyse initiale de la situation :\n{sess.statut_genere}\n\n"
        reponse, _ = await call_ai(system, ctx + body.message, max_tokens=600)
    except Exception as e:
        reponse = f"[Erreur IA : {e}]"

    historique.append({"role": "assistant", "content": reponse})
    sess.historique_chat = json.dumps(historique, ensure_ascii=False)
    db.commit()

    return {"reponse": reponse, "historique": historique}

@router.get("/ia/session/{session_id}")
def get_ia_session(
    session_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    sess = db.query(IASession).filter(IASession.id == session_id).first()
    if not sess:
        raise HTTPException(404)
    return {
        "id": sess.id,
        "statut_genere": sess.statut_genere,
        "historique": json.loads(sess.historique_chat or "[]")
    }


# ════════════════════════════════════════════════════════════
#  MESSAGERIE INTERNE
# ════════════════════════════════════════════════════════════

class MessageIn(BaseModel):
    destinataire_id: int
    sujet: Optional[str] = None
    contenu: str

@router.get("/messagerie")
def get_messages(
    boite: str = "reception",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    if boite == "envoi":
        rows = db.query(MessageInterne).filter(
            MessageInterne.expediteur_id == user.id
        ).order_by(MessageInterne.created_at.desc()).all()
    else:
        rows = db.query(MessageInterne).filter(
            MessageInterne.destinataire_id == user.id
        ).order_by(MessageInterne.created_at.desc()).all()

    result = []
    for m in rows:
        exp = db.query(User).filter(User.id == m.expediteur_id).first()
        dest = db.query(User).filter(User.id == m.destinataire_id).first()
        result.append({
            "id": m.id,
            "expediteur_id": m.expediteur_id,
            "expediteur_nom": exp.display_name if exp else "—",
            "destinataire_id": m.destinataire_id,
            "destinataire_nom": dest.display_name if dest else "—",
            "sujet": m.sujet,
            "contenu": m.contenu,
            "lu": m.lu,
            "lu_at": m.lu_at.isoformat() if m.lu_at else None,
            "created_at": m.created_at.isoformat() if m.created_at else None
        })
    return result

@router.get("/messagerie/non-lus")
def non_lus_count(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        return {"count": 0}
    n = db.query(MessageInterne).filter(
        MessageInterne.destinataire_id == user.id,
        MessageInterne.lu == False
    ).count()
    return {"count": n}

@router.post("/messagerie")
def send_message(
    body: MessageIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    dest = db.query(User).filter(User.id == body.destinataire_id, User.active == True).first()
    if not dest:
        raise HTTPException(404, "Destinataire introuvable")
    msg = MessageInterne(
        expediteur_id=user.id,
        destinataire_id=body.destinataire_id,
        sujet=body.sujet,
        contenu=body.contenu
    )
    db.add(msg)
    _log_mc(db, user, "MESSAGE", "ENVOYÉ", f"→ {dest.display_name or dest.username} | {body.sujet[:60]}")
    db.commit()
    return {"status": "ok"}

@router.put("/messagerie/{msg_id}/lire")
def marquer_lu(
    msg_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    msg = db.query(MessageInterne).filter(
        MessageInterne.id == msg_id,
        MessageInterne.destinataire_id == user.id
    ).first()
    if msg and not msg.lu:
        msg.lu = True
        msg.lu_at = datetime.now(timezone.utc)
        db.commit()
    return {"status": "ok"}


# ════════════════════════════════════════════════════════════
#  DÉCLARATIONS DE SITUATION (inter-GHT)
# ════════════════════════════════════════════════════════════

class DeclarationIn(BaseModel):
    site_id: str
    unite_fonct: Optional[str] = None
    type_crise: str
    niveau_tension: int = 1
    description: Optional[str] = None

@router.get("/declarations")
def get_declarations(
    actif_only: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    q = db.query(DeclarationSituation)
    if actif_only:
        q = q.filter(DeclarationSituation.actif == True)
    rows = q.order_by(DeclarationSituation.created_at.desc()).all()
    return [_decl_to_dict(r) for r in rows]

@router.post("/declarations")
def create_declaration(
    body: DeclarationIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    d = DeclarationSituation(
        site_id=body.site_id,
        unite_fonct=body.unite_fonct,
        type_crise=body.type_crise,
        niveau_tension=body.niveau_tension,
        description=body.description,
        created_by=user.display_name
    )
    db.add(d)
    db.commit()
    db.refresh(d)
    return _decl_to_dict(d)

@router.put("/declarations/{decl_id}/cloturer")
def cloturer_declaration(
    decl_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    d = db.query(DeclarationSituation).filter(DeclarationSituation.id == decl_id).first()
    if not d:
        raise HTTPException(404)
    d.actif = False
    db.commit()
    return {"status": "ok"}

def _decl_to_dict(d):
    return {
        "id": d.id, "site_id": d.site_id, "unite_fonct": d.unite_fonct,
        "type_crise": d.type_crise, "niveau_tension": d.niveau_tension,
        "description": d.description, "actif": d.actif,
        "created_by": d.created_by,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None
    }


# ════════════════════════════════════════════════════════════
#  DEMANDES INTER-GHT
# ════════════════════════════════════════════════════════════

class DemandeIn(BaseModel):
    type_situation: str
    unite_concernee: Optional[str] = None
    description: str
    ght_emetteur: str
    ght_destinataire: Optional[str] = None

class ReponseIn(BaseModel):
    reponse: str
    statut: str = "traite"

@router.get("/interght/demandes")
def get_demandes(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    rows = db.query(DemandeInterGHT).order_by(DemandeInterGHT.created_at.desc()).all()
    return [_dem_to_dict(r) for r in rows]

@router.post("/interght/demandes")
def create_demande(
    body: DemandeIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    d = DemandeInterGHT(
        type_situation=body.type_situation,
        unite_concernee=body.unite_concernee,
        description=body.description,
        ght_emetteur=body.ght_emetteur,
        ght_destinataire=body.ght_destinataire
    )
    db.add(d)
    db.commit()
    db.refresh(d)
    return _dem_to_dict(d)

@router.post("/interght/demandes/distant/repondre")
def repondre_demande_distante(
    body: dict,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Répond à une demande distante via un message collecteur."""
    if not user:
        raise HTTPException(401)
    # Pour les demandes distantes, on envoie un message dans la messagerie inter-GHT
    # La réponse sera visible dans les messages du collecteur
    from app.api.federation import FederationConfig
    import httpx, asyncio
    cfg = FederationConfig()
    if not cfg.is_ready:
        raise HTTPException(503, "Fédération non configurée")
    return {"ok": True, "message": "Réponse transmise via messagerie inter-GHT"}


@router.post("/interght/demandes/{dem_id}/repondre")
def repondre_demande(
    dem_id: int,
    body: ReponseIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(401)
    d = db.query(DemandeInterGHT).filter(DemandeInterGHT.id == dem_id).first()
    if not d:
        raise HTTPException(404)
    d.reponse = body.reponse
    d.statut = body.statut
    db.commit()
    return {"status": "ok"}

def _dem_to_dict(d):
    return {
        "id": d.id, "type_situation": d.type_situation,
        "unite_concernee": d.unite_concernee, "description": d.description,
        "ght_emetteur": d.ght_emetteur, "ght_destinataire": d.ght_destinataire,
        "statut": d.statut, "reponse": d.reponse,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None
    }


# ════════════════════════════════════════════════════════════
#  ENDPOINT SUPERVISION (pour le collecteur — polling)
#  Compatible avec l'architecture Option A (collecteur central)
# ════════════════════════════════════════════════════════════

import os

def _get_supervision_token() -> str:
    """Lit le token depuis config.js (injecté par setup.py)."""
    config_js = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        "app", "static", "config.js"
    )
    try:
        raw = open(config_js, encoding="utf-8").read()
        start = raw.find("const SCRIBE_CONFIG = ") + len("const SCRIBE_CONFIG = ")
        end = raw.rfind(";")
        cfg = json.loads(raw[start:end])
        return cfg.get("federation", {}).get("token", "")
    except Exception:
        return os.environ.get("SUPERVISION_TOKEN", "")

from fastapi import Header

@router.get("/supervision/status")
def supervision_status(
    x_supervision_token: Optional[str] = Header(None),
    db: Session = Depends(get_db)
):
    """
    Endpoint de supervision polling — appelé par le collecteur toutes les N secondes.
    Retourne un snapshot complet de l'état de l'instance.
    Authentification : header X-Supervision-Token
    """
    expected_token = _get_supervision_token()
    if expected_token and x_supervision_token != expected_token:
        raise HTTPException(403, "Token de supervision invalide")

    open_incidents = db.query(SitrepEntry).filter(
        SitrepEntry.status != "RÉSOLU"
    ).order_by(SitrepEntry.urgency.desc()).all()

    max_urgency = max((i.urgency for i in open_incidents), default=0)
    if max_urgency >= 4:    niveau = "CRITIQUE"
    elif max_urgency >= 3:  niveau = "CRISE"
    elif max_urgency >= 2:  niveau = "ALERTE"
    elif max_urgency >= 1:  niveau = "VEILLE"
    else:                   niveau = "NOMINAL"

    # Déclarations actives
    declarations = db.query(DeclarationSituation).filter(
        DeclarationSituation.actif == True
    ).all()

    # Médico-tech
    medicotech = db.query(CapaciteMedicotech).order_by(
        CapaciteMedicotech.updated_at.desc()
    ).first()

    # Sites avec leurs statuts
    from app.models import Hospital
    sites = db.query(Hospital).all()
    sites_statut = []
    for h in sites:
        site_incs = [i for i in open_incidents
                     if i.site_id == h.nom or i.site_id == str(h.id)]
        urg_site = max((i.urgency for i in site_incs), default=0)
        niv = (
            "CRITIQUE" if urg_site >= 4 else
            "CRISE" if urg_site >= 3 else
            "ALERTE" if urg_site >= 2 else
            "VEILLE" if urg_site >= 1 else "NOMINAL"
        )
        sites_statut.append({
            "site_id": h.nom, "nom": h.nom,
            "niveau": niv,
            "incidents_ouverts": len(site_incs),
            "latitude": h.latitude, "longitude": h.longitude
        })

    # Capacité lits (synthèse rapide)
    refs = db.query(CapaciteReferentiel).filter(CapaciteReferentiel.actif == True).all()
    lits_total = sum(r.capacite_totale or 0 for r in refs)
    alertes_cap = 0
    for ref in refs:
        last = (db.query(CapaciteDeclaration)
                  .filter(CapaciteDeclaration.referentiel_id == ref.id)
                  .order_by(CapaciteDeclaration.horodatage.desc()).first())
        if last and (last.alerte_lits or last.alerte_rh or last.alerte_materiel):
            alertes_cap += 1

    return {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "niveau_global": niveau,
        "incidents_ouverts": len(open_incidents),
        "incidents_critiques": sum(1 for i in open_incidents if i.urgency >= 4),
        "declarations_actives": len(declarations),
        "alertes_capacitaires": alertes_cap,
        "sites": sites_statut,
        "declarations": [_decl_to_dict(d) for d in declarations],
        "medicotech": _mt_to_dict(medicotech) if medicotech else None,
        "capacite_lits_total": lits_total,
    }


# ════════════════════════════════════════════════════════════
#  ENDPOINT RÉCEPTION DEMANDE INTER-GHT (depuis collecteur)
# ════════════════════════════════════════════════════════════

@router.post("/supervision/demande")
def receive_demande_interght(
    request_data: dict,
    x_supervision_token: Optional[str] = Header(None),
    db: Session = Depends(get_db)
):
    """Reçoit une demande inter-GHT routée par le collecteur."""
    expected_token = _get_supervision_token()
    if expected_token and x_supervision_token != expected_token:
        raise HTTPException(403, "Token invalide")
    d = DemandeInterGHT(
        type_situation=request_data.get("type_situation", ""),
        unite_concernee=request_data.get("unite_concernee"),
        description=request_data.get("description", ""),
        ght_emetteur=request_data.get("ght_emetteur", ""),
        ght_destinataire=request_data.get("ght_destinataire"),
        statut="recu"
    )
    db.add(d)
    db.commit()
    return {"status": "ok"}


# ════════════════════════════════════════════════════════════
#  SNAPSHOT AMBIANCE (injecté dans les prompts IA)
# ════════════════════════════════════════════════════════════

MOTS_TENSION = [
    "critique", "saturé", "saturation", "débordé", "rupture", "urgent",
    "alerte", "crise", "dégradé", "indisponible", "arrêt", "panne",
    "hors service", "transfert", "évacuation", "renfort", "aide",
    "surcharge", "plein", "complet", "fermé"
]

def build_ambiance_snapshot(db: Session) -> str:
    """Construit un texte d'ambiance plateforme pour enrichir les prompts IA."""
    lines = []

    # Incidents ouverts
    open_incs = db.query(SitrepEntry).filter(SitrepEntry.status != "RÉSOLU").all()
    n = len(open_incs)
    n_crit = sum(1 for i in open_incs if i.urgency >= 4)
    n_crise = sum(1 for i in open_incs if i.urgency == 3)
    if n > 0:
        lines.append(f"• {n} incident(s) ouvert(s) dont {n_crit} critique(s) et {n_crise} en état de crise")

    # Analyse tonale des textes libres
    textes = [(i.fait or "") + " " + (i.analyse or "") for i in open_incs]
    texte_global = " ".join(textes).lower()
    mots_trouves = [m for m in MOTS_TENSION if m in texte_global]
    if mots_trouves:
        lines.append(f"• Tonalité des écrits tendue — mots détectés : {', '.join(mots_trouves[:8])}")
    else:
        lines.append("• Tonalité des écrits : normale")

    # Médico-tech
    mt = db.query(CapaciteMedicotech).order_by(CapaciteMedicotech.updated_at.desc()).first()
    if mt:
        if mt.blocs_total > 0:
            pct = mt.blocs_operationnels / mt.blocs_total * 100
            etat = "critique" if pct < 40 else ("dégradé" if pct < 70 else "correct")
            lines.append(f"• Blocs : {mt.blocs_operationnels}/{mt.blocs_total} opérationnels ({pct:.0f}%) — état {etat}")
        if mt.dialyse_postes_total > 0:
            pct_d = mt.dialyse_postes_actifs / mt.dialyse_postes_total * 100
            lines.append(f"• Dialyse : {mt.dialyse_postes_actifs}/{mt.dialyse_postes_total} postes actifs ({pct_d:.0f}%)")
        pharma_label = {"normal": "fonctionnelle", "degrade": "dégradée", "arret": "À L'ARRÊT"}.get(mt.pharmacie_statut, mt.pharmacie_statut)
        lines.append(f"• Pharmacie : {pharma_label} — urgences vitales : {'OUI' if mt.pharmacie_urgences_vitales else 'NON'}")

    # Déclarations actives
    decls = db.query(DeclarationSituation).filter(DeclarationSituation.actif == True).all()
    for d in decls:
        lines.append(f"• Déclaration active : {d.type_crise} sur {d.unite_fonct or d.site_id} (niveau {d.niveau_tension}/3)")

    return "\n".join(lines) if lines else "Aucune donnée d'ambiance disponible."


# ═══════════════════════════════════════════════════════════════
# IMPORT DE COMPTES EN LOT (v1.4.0)
# ═══════════════════════════════════════════════════════════════

import hashlib as _hashlib
from fastapi import UploadFile, File
import io as _io

IMPORT_TEMP_PASSWORD = "Scribe2026!"
IMPORT_TEMP_HASH     = _hashlib.sha256(IMPORT_TEMP_PASSWORD.encode()).hexdigest()
IMPORT_VALID_ROLES   = {"admin", "directeur", "observateur"}


@router.post("/auth/import-comptes")
async def import_comptes_xlsx(
    file: UploadFile = File(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Import de comptes utilisateurs en lot depuis un fichier XLSX.
    Requiert le rôle admin. Retourne un rapport d'import."""
    try:
        import openpyxl as _xl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")

    content = await file.read()
    try:
        wb = _xl.load_workbook(_io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fichier Excel invalide : {e}")

    # Trouver le bon onglet
    ws = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "compte" in nl or "utilisateur" in nl:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    # Lire les en-têtes
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if not val:
            continue
        k = str(val).lower().strip()
        k = k.replace("é","e").replace("è","e").replace("ê","e").replace(" ","_").replace("/","_")
        if "prenom" in k:          col_map[col] = "prenom"
        elif "nom" in k and "unite" not in k: col_map[col] = "nom"
        elif "username" in k or "identif" in k or "login" in k: col_map[col] = "username"
        elif "role" in k:          col_map[col] = "role"
        elif "unite" in k or "service" in k: col_map[col] = "unite_soin"
        elif "mobile" in k or "gsm" in k: col_map[col] = "telephone_mobile"
        elif "fixe" in k or "interne" in k: col_map[col] = "telephone_fixe"
        elif "email" in k or "mail" in k: col_map[col] = "email"
        elif "perimetre" in k or "perim" in k: col_map[col] = "perimetre"

    created, updated, skipped = 0, 0, 0
    errors = []

    for row_idx in range(2, ws.max_row + 1):
        row = {}
        has_data = False
        for col, key in col_map.items():
            v = ws.cell(row=row_idx, column=col).value
            if v is not None:
                row[key] = str(v).strip()
                has_data = True
            else:
                row[key] = ""
        if not has_data:
            continue

        username = row.get("username", "").lower().replace(" ", "")
        nom      = row.get("nom", "")
        prenom   = row.get("prenom", "")
        role     = row.get("role", "directeur").lower()
        perim    = row.get("perimetre", "")

        if not username:
            errors.append(f"Ligne {row_idx}: username vide — ignorée")
            continue
        if role not in IMPORT_VALID_ROLES:
            errors.append(f"Ligne {row_idx}: rôle '{role}' invalide pour '{username}' → 'directeur' utilisé")
            role = "directeur"

        display_name = f"{prenom} {nom}".strip() or username

        existing = db.query(User).filter(User.username == username).first()
        if existing:
            if existing.role == "admin" and role != "admin":
                skipped += 1
                errors.append(f"Ligne {row_idx}: '{username}' est admin, rôle non modifié")
                continue
            existing.display_name   = display_name
            existing.role           = role
            existing.perimetre      = perim or None
            existing.active         = True
            db.commit()
            updated += 1
        else:
            u = User(
                username=username,
                display_name=display_name,
                role=role,
                hashed_password=IMPORT_TEMP_HASH,
                perimetre=perim or None,
                active=True,
                must_change_password=True,
            )
            db.add(u)
            try:
                db.commit()
                created += 1
            except Exception as e:
                db.rollback()
                errors.append(f"Ligne {row_idx}: erreur création '{username}': {e}")

    return {
        "ok": True,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "temp_password": IMPORT_TEMP_PASSWORD,
        "message": f"{created} créé(s), {updated} mis à jour, {skipped} ignoré(s)"
    }


@router.get("/auth/comptes-modele")
async def download_comptes_modele(user: User = Depends(get_current_user)):
    """Télécharge le fichier Excel modèle pour l'import de comptes.
    v2.4.8.2 : accessible à tout utilisateur authentifié — c'est juste un
    modèle vierge sans secret. La création de comptes nécessite toujours
    le rôle admin via /auth/import-comptes."""
    import os as _os
    from fastapi.responses import FileResponse as _FR
    # Chercher le fichier dans le répertoire racine du projet
    base = _os.path.dirname(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
    candidates = [
        _os.path.join(base, "comptes_modele.xlsx"),
        _os.path.join(base, "..", "comptes_modele.xlsx"),
    ]
    for path in candidates:
        if _os.path.exists(path):
            return _FR(path, filename="comptes_modele.xlsx",
                       media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    raise HTTPException(status_code=404, detail="Fichier modèle non trouvé. Placez comptes_modele.xlsx à la racine du projet.")


# ══════════════════════════════════════════════════════════════════════════════
# TRANSFERTS PATIENTS
# ══════════════════════════════════════════════════════════════════════════════

from app.models import TransfertPatient

class TransfertCreate(BaseModel):
    unite_origine:             str
    etablissement_origine:     str
    unite_destination:         str
    etablissement_destination: str
    site_destination: Optional[str] = None
    redacteur:                 str
    commentaire:               Optional[str] = None
    eta:                       Optional[str] = None
    nom:              Optional[str] = None
    prenom:           Optional[str] = None
    nom_jeune_fille:  Optional[str] = None
    ipp:              Optional[str] = None
    date_naissance:   Optional[str] = None

class TransfertUpdate(BaseModel):
    statut:            Optional[str]  = None
    commentaire:       Optional[str]  = None
    unite_destination: Optional[str]  = None
    etablissement_destination: Optional[str] = None
    site_destination:  Optional[str]  = None
    eta:               Optional[str]  = None

class StatutUpdate(BaseModel):
    statut: str
    reason: str | None = None  # v2.4.8 : motif obligatoire en cas de recul

def _fmt_transfert(t: TransfertPatient) -> dict:
    import json as _json
    try:
        history = _json.loads(t.historique_json) if getattr(t, "historique_json", None) else []
    except (ValueError, TypeError):
        history = []
    return {
        "id": t.id,
        "unite_origine": t.unite_origine,
        "etablissement_origine": t.etablissement_origine,
        "unite_destination": t.unite_destination,
        "etablissement_destination": t.etablissement_destination,
        "statut": t.statut,
        "redacteur": t.redacteur,
        "commentaire": t.commentaire,
        "nom": t.nom,
        "prenom": t.prenom,
        "nom_jeune_fille": t.nom_jeune_fille,
        "ipp": t.ipp,
        "date_naissance": t.date_naissance,
        "horodatage_creation": t.horodatage_creation.isoformat() if t.horodatage_creation else None,
        "horodatage_depart":   t.horodatage_depart.isoformat()   if t.horodatage_depart   else None,
        "horodatage_arrivee":  t.horodatage_arrivee.isoformat()  if t.horodatage_arrivee  else None,
        "eta":                 getattr(t, "eta", None),
        "site_destination":    getattr(t, "site_destination", None),
        "site_origine":        getattr(t, "site_origine", None),
        "historique":          history,  # v2.4.6
    }

@router.get("/transferts")
def list_transferts(
    statut: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(TransfertPatient)
    if statut:
        q = q.filter(TransfertPatient.statut == statut)
    else:
        from datetime import timedelta
        cutoff = datetime.now(timezone.utc) - timedelta(hours=24)
        q = q.filter(
            (TransfertPatient.statut.in_(["EN_PREPARATION", "EN_COURS"])) |
            ((TransfertPatient.statut == "ARRIVE") & (TransfertPatient.horodatage_arrivee >= cutoff)) |
            ((TransfertPatient.statut == "ANNULE") & (TransfertPatient.horodatage_creation >= cutoff))
        )
    items = q.order_by(TransfertPatient.horodatage_creation.desc()).all()
    return [_fmt_transfert(t) for t in items]


@router.post("/transferts")
def create_transfert(body: TransfertCreate, db: Session = Depends(get_db),
                     current_user=Depends(get_current_user)):
    redacteur_reel = (current_user.display_name or current_user.username) if current_user else (body.redacteur or "Inconnu")
    t = TransfertPatient(
        unite_origine=body.unite_origine,
        etablissement_origine=body.etablissement_origine,
        unite_destination=body.unite_destination,
        etablissement_destination=body.etablissement_destination,
            site_destination=body.site_destination,
        statut="EN_PREPARATION",
        redacteur=redacteur_reel,
        commentaire=body.commentaire,
        eta=body.eta,
        nom=body.nom, prenom=body.prenom,
        nom_jeune_fille=body.nom_jeune_fille,
        ipp=body.ipp, date_naissance=body.date_naissance,
    )
    db.add(t); db.commit(); db.refresh(t)
    return _fmt_transfert(t)


@router.patch("/transferts/{tid}/statut")
def patch_transfert_statut(tid: int, body: StatutUpdate, db: Session = Depends(get_db),
                            current_user=Depends(get_current_user)):
    STATUTS_VALIDES = {"EN_PREPARATION", "EN_COURS", "ARRIVE", "ANNULE"}
    if body.statut not in STATUTS_VALIDES:
        raise HTTPException(status_code=400, detail=f"Statut invalide : {body.statut}")
    t = db.query(TransfertPatient).filter(TransfertPatient.id == tid).first()
    if not t:
        raise HTTPException(status_code=404, detail="Transfert introuvable")
    now = datetime.now(timezone.utc)
    # v2.4.8 : si recul (ARRIVE→EN_COURS, EN_COURS→EN_PREPARATION), motif obligatoire
    ORDER = {"EN_PREPARATION": 0, "EN_COURS": 1, "ARRIVE": 2, "ANNULE": 99}
    is_regression = (
        t.statut in ORDER and body.statut in ORDER
        and ORDER[body.statut] < ORDER[t.statut]
        and body.statut != "ANNULE"
    )
    if is_regression and not (body.reason and body.reason.strip()):
        raise HTTPException(
            status_code=400,
            detail=f"Motif obligatoire pour passer de {t.statut} à {body.statut}"
        )
    if body.statut == "EN_COURS" and t.statut == "EN_PREPARATION":
        t.horodatage_depart = now
    elif body.statut == "ARRIVE" and t.statut != "ARRIVE":
        t.horodatage_arrivee = now
    # v2.4.6 : trace historique des changements de statut (+ motif v2.4.8)
    if body.statut != t.statut:
        _append_transfert_history(t, t.statut, body.statut, current_user, now,
                                   reason=(body.reason or "").strip() or None)
    t.statut = body.statut
    db.commit(); db.refresh(t)
    return _fmt_transfert(t)


def _append_transfert_history(t, old_statut, new_statut, current_user, now, reason=None):
    """Append une entrée à historique_json. v2.4.6 (+reason v2.4.8)"""
    import json as _json
    try:
        history = _json.loads(t.historique_json) if t.historique_json else []
    except (ValueError, TypeError):
        history = []
    user_label = "?"
    try:
        user_label = getattr(current_user, "login", None) or getattr(current_user, "username", None) or "?"
    except Exception:
        pass
    entry = {
        "ts":   now.isoformat() if now else "",
        "from": old_statut or "",
        "to":   new_statut or "",
        "user": user_label,
    }
    if reason:
        entry["reason"] = reason
    history.append(entry)
    t.historique_json = _json.dumps(history)


@router.put("/transferts/{tid}")
def update_transfert(tid: int, body: TransfertUpdate, db: Session = Depends(get_db),
                     current_user=Depends(get_current_user)):
    t = db.query(TransfertPatient).filter(TransfertPatient.id == tid).first()
    if not t:
        raise HTTPException(status_code=404, detail="Transfert introuvable")
    if body.statut:
        now = datetime.now(timezone.utc)
        if body.statut == "EN_COURS" and t.statut == "EN_PREPARATION":
            t.horodatage_depart = now
        elif body.statut == "ARRIVE" and t.statut != "ARRIVE":
            t.horodatage_arrivee = now
        # v2.4.6 : trace historique
        if body.statut != t.statut:
            _append_transfert_history(t, t.statut, body.statut, current_user, now)
        t.statut = body.statut
    if body.commentaire is not None: t.commentaire = body.commentaire
    if body.unite_destination is not None: t.unite_destination = body.unite_destination
    if body.etablissement_destination is not None: t.etablissement_destination = body.etablissement_destination
    if body.site_destination is not None: t.site_destination = body.site_destination
    if body.eta is not None: t.eta = body.eta
    db.commit(); db.refresh(t)
    return _fmt_transfert(t)


@router.get("/transferts/summary")
def transferts_summary(db: Session = Depends(get_db),
                       current_user=Depends(get_current_user)):
    en_cours = db.query(TransfertPatient).filter(TransfertPatient.statut == "EN_COURS").all()
    en_prep  = db.query(TransfertPatient).filter(TransfertPatient.statut == "EN_PREPARATION").count()
    return {
        "en_cours": len(en_cours),
        "en_preparation": en_prep,
        "flux": [{"origine": f"{t.unite_origine} ({t.etablissement_origine})",
                  "destination": f"{t.unite_destination} ({t.etablissement_destination})",
                  "statut": t.statut} for t in en_cours]
    }




# ── MAIN COURANTE EXHAUSTIVE ─────────────────────────────────────────────────

@router.get("/main-courante/logs")
def get_mc_logs(
    categorie: str = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Récupère les logs exhaustifs de la main courante."""
    from app.models import MainCouranteLog
    q = db.query(MainCouranteLog)
    if categorie:
        q = q.filter(MainCouranteLog.categorie == categorie)
    logs = q.order_by(MainCouranteLog.timestamp.desc()).limit(limit).all()
    return [{
        "id": l.id,
        "timestamp": l.timestamp.isoformat() if l.timestamp else None,
        "auteur": l.auteur,
        "auteur_role": l.auteur_role,
        "categorie": l.categorie,
        "action": l.action,
        "detail": l.detail,
        "ref_id": l.ref_id,
        "ref_type": l.ref_type,
        "site": l.site,
        "niveau": l.niveau,
    } for l in logs]
