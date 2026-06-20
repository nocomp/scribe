"""h79 — Chaîne d'alerte / mobilisation (Phase A).

Annuaire de personnes mobilisables (importé depuis l'Excel établissement),
ciblage par UF/pôle/site/fonction, envoi multi-canal (SMS + mail) avec un lien
unique tokenisé vers un formulaire ETA responsive, et vue « qui arrive ».

RGPD : ces données personnelles restent STRICTEMENT locales à l'instance — elles
ne sont jamais remontées au collecteur.
"""
from __future__ import annotations

import io
import os
import json
import html
import secrets
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse, HTMLResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import ContactMobilisation, AlerteMobilisation, AlerteCible

try:
    from app.api.auth import get_current_user, require_admin
except Exception:  # pragma: no cover
    def get_current_user(): pass
    def require_admin(): pass

router = APIRouter()

# Entêtes acceptés à l'import (normalisés : minuscule, sans accents légers, sans ":texte")
_FIELD_MAP = {
    "cle": "cle", "clé": "cle",
    "nom": "nom",
    "prenom": "prenom", "prénom": "prenom",
    "tel1": "tel1", "telephone": "tel1", "téléphone": "tel1", "tel": "tel1",
    "tel2": "tel2",
    "mail": "email", "email": "email", "e-mail": "email", "courriel": "email",
    "fonction": "fonction",
    "service": "service",
    "uf": "uf",
    "grade": "grade",
    "pole": "pole", "pôle": "pole",
    "site": "site",
}

_TEMPLATE_HEADERS = ["Clé", "Nom", "Prénom", "TEL1", "TEL2",
                     "Fonction", "Service", "Uf", "Grade", "Pole", "Site", "Mail"]


def _norm_header(h: str) -> str:
    h = (h or "").strip().lower()
    if ":" in h:
        h = h.split(":", 1)[0].strip()
    return h


# ── Import Excel ─────────────────────────────────────────────────────────────

@router.post("/import")
async def import_contacts(fichier: UploadFile = File(...),
                          remplacer: bool = Form(False),
                          db: Session = Depends(get_db),
                          admin=Depends(require_admin)):
    """Importe l'annuaire de mobilisation depuis un .xlsx. `remplacer=true` vide
    l'annuaire avant import ; sinon, upsert par clé (ou par nom+prénom si pas de clé)."""
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl indisponible côté serveur")
    raw = await fichier.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Fichier illisible : {e}")
    if not rows:
        raise HTTPException(400, "Fichier vide")

    headers = [_norm_header(str(c) if c is not None else "") for c in rows[0]]
    col_idx = {}
    for i, h in enumerate(headers):
        if h in _FIELD_MAP:
            col_idx[_FIELD_MAP[h]] = i
    if "nom" not in col_idx:
        raise HTTPException(400, "Colonne 'Nom' introuvable — vérifiez l'entête du fichier.")

    if remplacer:
        db.query(ContactMobilisation).delete()
        db.flush()

    def _cell(row, field):
        i = col_idx.get(field)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else str(v).strip()

    nb_imported, nb_updated = 0, 0
    for row in rows[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        nom = _cell(row, "nom")
        if not nom:
            continue
        cle = _cell(row, "cle")
        existing = None
        if not remplacer:
            if cle:
                existing = db.query(ContactMobilisation).filter(ContactMobilisation.cle == cle).first()
            if not existing:
                existing = (db.query(ContactMobilisation)
                              .filter(ContactMobilisation.nom == nom,
                                      ContactMobilisation.prenom == _cell(row, "prenom")).first())
        target = existing or ContactMobilisation()
        target.cle      = cle
        target.nom      = nom
        target.prenom   = _cell(row, "prenom")
        target.tel1     = _cell(row, "tel1")
        target.tel2     = _cell(row, "tel2")
        target.email    = _cell(row, "email")
        target.fonction = _cell(row, "fonction")
        target.service  = _cell(row, "service")
        target.uf       = _cell(row, "uf")
        target.grade    = _cell(row, "grade")
        target.pole     = _cell(row, "pole")
        target.site     = _cell(row, "site")
        target.actif    = True
        if existing:
            nb_updated += 1
        else:
            db.add(target)
            nb_imported += 1
    db.commit()
    total = db.query(ContactMobilisation).count()
    return {"ok": True, "importes": nb_imported, "mis_a_jour": nb_updated, "total": total}


# ── Template anonymisé ───────────────────────────────────────────────────────

@router.get("/template")
def download_template(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Génère un modèle .xlsx VIERGE (entêtes + 2 lignes d'exemple fictives) à
    diffuser aux établissements. Aucune donnée réelle."""
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl indisponible côté serveur")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mobilisation"
    ws.append(_TEMPLATE_HEADERS)
    ws.append(["001", "DUPONT", "Marie", "0600000001", "",
               "Cadre de santé", "Réanimation", "1200", "CADRE DE SANTÉ", "POLE A", "Site Principal", "marie.dupont@exemple.fr"])
    ws.append(["002", "MARTIN", "Paul", "0600000002", "",
               "Médecin", "Urgences", "1400", "PRATICIEN HOSPITALIER", "POLE B", "Site Annexe", "paul.martin@exemple.fr"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modele_mobilisation_scribe.xlsx"},
    )


# ── Facettes & contacts (ciblage) ────────────────────────────────────────────

@router.get("/facets")
def facets(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Valeurs distinctes pour le ciblage (UF, pôle, site, fonction)."""
    def _distinct(col):
        vals = [r[0] for r in db.query(col).filter(col.isnot(None), col != "").distinct().all()]
        return sorted({v.strip() for v in vals if v and v.strip()})
    return {
        "uf":       _distinct(ContactMobilisation.uf),
        "pole":     _distinct(ContactMobilisation.pole),
        "site":     _distinct(ContactMobilisation.site),
        "fonction": _distinct(ContactMobilisation.fonction),
        "total":    db.query(ContactMobilisation).filter(ContactMobilisation.actif == True).count(),
    }


def _resolve_contacts(db, criteres, contact_ids):
    """Union : contacts correspondant aux critères + contacts sélectionnés par
    identifiant (noms). Si NI critères NI identifiants → tout l'annuaire actif."""
    crit_has = any((criteres.get(k) or []) for k in ("uf", "pole", "site", "fonction"))
    ids = [int(x) for x in (contact_ids or []) if str(x).strip().isdigit()]
    if not crit_has and not ids:
        return _filter_contacts(db, {}).all()
    out = {}
    if crit_has:
        for c in _filter_contacts(db, criteres).all():
            out[c.id] = c
    if ids:
        for c in (db.query(ContactMobilisation)
                    .filter(ContactMobilisation.id.in_(ids),
                            ContactMobilisation.actif == True).all()):  # noqa: E712
            out[c.id] = c
    return list(out.values())


def _filter_contacts(db, criteres):
    q = db.query(ContactMobilisation).filter(ContactMobilisation.actif == True)
    for field, col in (("uf", ContactMobilisation.uf), ("pole", ContactMobilisation.pole),
                       ("site", ContactMobilisation.site), ("fonction", ContactMobilisation.fonction)):
        vals = [v for v in (criteres.get(field) or []) if v]
        if vals:
            q = q.filter(col.in_(vals))
    return q


class CibleQuery(BaseModel):
    uf:          Optional[list] = None
    pole:        Optional[list] = None
    site:        Optional[list] = None
    fonction:    Optional[list] = None
    contact_ids: Optional[list] = None


@router.get("/contacts")
def list_contacts(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Liste l'annuaire (pour la sélection nominative côté ciblage)."""
    rows = (db.query(ContactMobilisation)
              .filter(ContactMobilisation.actif == True)  # noqa: E712
              .order_by(ContactMobilisation.nom, ContactMobilisation.prenom).all())
    return [{
        "id": c.id, "nom": c.nom, "prenom": c.prenom, "fonction": c.fonction,
        "site": c.site, "pole": c.pole, "uf": c.uf,
        "tel": bool(c.tel1 or c.tel2), "mail": bool(c.email),
    } for c in rows]


@router.post("/preview")
def preview_cibles(body: CibleQuery, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Compte (et échantillonne) les contacts correspondant aux critères."""
    crit = body.dict()
    rows = _resolve_contacts(db, crit, body.contact_ids)
    avec_tel  = sum(1 for c in rows if (c.tel1 or c.tel2))
    avec_mail = sum(1 for c in rows if c.email)
    apercu = [{"nom": c.nom, "prenom": c.prenom, "fonction": c.fonction,
               "site": c.site, "uf": c.uf, "pole": c.pole,
               "tel": bool(c.tel1 or c.tel2), "mail": bool(c.email)} for c in rows[:50]]
    return {"total": len(rows), "avec_tel": avec_tel, "avec_mail": avec_mail, "apercu": apercu}


# ── Déclenchement d'une alerte ───────────────────────────────────────────────

class DeclenchementIn(BaseModel):
    titre:       str
    message:     str
    uf:          Optional[list] = None
    pole:        Optional[list] = None
    site:        Optional[list] = None
    fonction:    Optional[list] = None
    canaux:      Optional[list] = None      # ["sms","mail"]
    contact_ids: Optional[list] = None
    incident_id: Optional[int] = None


@router.post("/alerte")
async def declencher_alerte(body: DeclenchementIn, request: Request,
                            db: Session = Depends(get_db), admin=Depends(require_admin)):
    """Crée une alerte, résout les cibles et envoie SMS+mail avec un lien ETA
    unique par personne."""
    crit = {"uf": body.uf or [], "pole": body.pole or [], "site": body.site or [], "fonction": body.fonction or []}
    contacts = _resolve_contacts(db, crit, body.contact_ids)
    if not contacts:
        raise HTTPException(400, "Aucun contact ne correspond aux critères")
    canaux = [c for c in (body.canaux or ["sms", "mail"]) if c in ("sms", "mail")] or ["sms", "mail"]

    alerte = AlerteMobilisation(
        titre=(body.titre or "Mobilisation").strip()[:200],
        message=(body.message or "").strip(),
        criteres=json.dumps(crit, ensure_ascii=False),
        incident_id=body.incident_id,
        cree_par=getattr(admin, "username", "admin"),
    )
    db.add(alerte)
    db.flush()

    # Backends (best-effort)
    sms_backend = _load_backend(db, "sms") if "sms" in canaux else None
    mail_backend = _load_backend(db, "mail") if "mail" in canaux else None
    base = _public_base(request)

    nb_sms, nb_mail = 0, 0
    for c in contacts:
        token = secrets.token_urlsafe(16)
        tel = (c.tel1 or c.tel2 or "").strip()
        cible = AlerteCible(
            alerte_id=alerte.id, contact_id=c.id,
            nom=((c.prenom + " " + c.nom).strip() if c.prenom else c.nom),
            fonction=c.fonction, site=c.site, tel=tel, email=c.email,
            token=token, canaux=",".join(canaux), statut="envoye",
        )
        db.add(cible)
        link_rel = f"/api/v1/mobilisation/eta/{token}"
        # SMS
        if sms_backend and tel:
            ok = await _send_one(sms_backend, alerte, base, link_rel, tel)
            if ok:
                nb_sms += 1
        # Mail
        if mail_backend and c.email:
            ok = await _send_one(mail_backend, alerte, base, link_rel, c.email)
            if ok:
                nb_mail += 1
    db.commit()
    return {"ok": True, "alerte_id": alerte.id, "cibles": len(contacts),
            "sms_envoyes": nb_sms, "mails_envoyes": nb_mail}


def _public_base(request: Request) -> str:
    """h80 — URL publique fiable pour forger les liens : on privilégie l'entête
    Host (ce que le navigateur de l'admin a utilisé), car request.base_url peut
    renvoyer un host vide (lien cassé « http:/// »)."""
    host = request.headers.get("host") or request.url.netloc
    scheme = request.headers.get("x-forwarded-proto") or request.url.scheme or "http"
    if host:
        return f"{scheme}://{host}"
    return str(request.base_url).rstrip("/")


def _load_backend(db, kind):
    try:
        from plugins.notifications.models import NotifChannel
        from plugins.notifications.backends import BACKENDS
        from plugins.notifications.dispatcher import _apply_central_config
        ch = (db.query(NotifChannel)
                .filter(NotifChannel.kind == kind, NotifChannel.enabled == True)  # noqa: E712
                .first())
        if not ch:
            return None
        cfg = _apply_central_config(kind, json.loads(ch.config_json or "{}"))
        cls = BACKENDS.get(kind)
        if not cls:
            return None
        b = cls(cfg)
        return b if b.is_configured() else None
    except Exception:
        return None


async def _send_one(backend, alerte, base, link_rel, target):
    try:
        from plugins.notifications.backends import NotifPayload
        payload = NotifPayload(
            event_type="mobilisation", title=alerte.titre,
            body=alerte.message, urgency=4,
            context={"url": link_rel, "base_url": base},
        )
        res = await backend.send(payload, target)
        return bool(getattr(res, "success", False))
    except Exception:
        return False


# ── Suivi « qui arrive » ─────────────────────────────────────────────────────

@router.get("/alertes")
def list_alertes(db: Session = Depends(get_db), user=Depends(get_current_user)):
    rows = db.query(AlerteMobilisation).order_by(AlerteMobilisation.id.desc()).limit(80).all()
    rows = [a for a in rows if not getattr(a, "archived", 0)]
    out = []
    for a in rows:
        cibles = db.query(AlerteCible).filter(AlerteCible.alerte_id == a.id).all()
        out.append({
            "id": a.id, "titre": a.titre,
            "created_at": _iso(a.created_at),
            "cibles": len(cibles),
            "repondus": sum(1 for c in cibles if c.statut == "repondu"),
        })
    return out


@router.get("/alerte/{alerte_id}")
def detail_alerte(alerte_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    a = db.query(AlerteMobilisation).filter(AlerteMobilisation.id == alerte_id).first()
    if not a:
        raise HTTPException(404, "Alerte introuvable")
    cibles = db.query(AlerteCible).filter(AlerteCible.alerte_id == alerte_id).all()
    # UF résolue via le contact d'origine (pas de colonne dédiée sur la cible)
    uf_map = {}
    ids = [c.contact_id for c in cibles if c.contact_id]
    if ids:
        for cm in db.query(ContactMobilisation).filter(ContactMobilisation.id.in_(ids)).all():
            uf_map[cm.id] = cm.uf or ""
    return {
        "id": a.id, "titre": a.titre, "message": a.message,
        "created_at": _iso(a.created_at),
        "cibles": [{
            "id": c.id, "nom": c.nom, "fonction": c.fonction, "site": c.site,
            "uf": uf_map.get(c.contact_id, ""),
            "canaux": c.canaux, "statut": c.statut,
            "eta_choice": c.eta_choice, "commentaire": c.commentaire,
            "responded_at": _iso(c.responded_at),
            "has_tel": bool(c.tel), "has_mail": bool(c.email),
        } for c in cibles],
    }


def _iso(dt):
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


@router.post("/alerte/{alerte_id}/relancer")
async def relancer_alerte(alerte_id: int, request: Request, uf: Optional[str] = None,
                          db: Session = Depends(get_db), admin=Depends(require_admin)):
    """h80/h81 — Relance UNIQUEMENT les destinataires sans réponse, en réutilisant
    leur lien (token). `uf` optionnel : ne relancer que cette unité fonctionnelle."""
    a = db.query(AlerteMobilisation).filter(AlerteMobilisation.id == alerte_id).first()
    if not a:
        raise HTTPException(404, "Alerte introuvable")
    cibles = (db.query(AlerteCible)
                .filter(AlerteCible.alerte_id == alerte_id,
                        AlerteCible.statut != "repondu").all())
    if uf:
        ids = [c.contact_id for c in cibles if c.contact_id]
        uf_map = {}
        if ids:
            for cm in db.query(ContactMobilisation).filter(ContactMobilisation.id.in_(ids)).all():
                uf_map[cm.id] = cm.uf or ""
        cibles = [c for c in cibles if uf_map.get(c.contact_id, "") == uf]
    if not cibles:
        return {"ok": True, "relances": 0, "sms_envoyes": 0, "mails_envoyes": 0}
    sms_backend = _load_backend(db, "sms")
    mail_backend = _load_backend(db, "mail")
    base = _public_base(request)
    nb_sms, nb_mail = 0, 0
    for c in cibles:
        canaux = (c.canaux or "")
        link_rel = f"/api/v1/mobilisation/eta/{c.token}"
        if sms_backend and c.tel and "sms" in canaux:
            if await _send_one(sms_backend, a, base, link_rel, c.tel):
                nb_sms += 1
        if mail_backend and c.email and "mail" in canaux:
            if await _send_one(mail_backend, a, base, link_rel, c.email):
                nb_mail += 1
    return {"ok": True, "relances": len(cibles), "sms_envoyes": nb_sms, "mails_envoyes": nb_mail}


@router.post("/alerte/{alerte_id}/archive")
def archive_alerte(alerte_id: int, db: Session = Depends(get_db), admin=Depends(require_admin)):
    """h82 — Archive une campagne terminée (disparaît de la liste active)."""
    a = db.query(AlerteMobilisation).filter(AlerteMobilisation.id == alerte_id).first()
    if not a:
        raise HTTPException(404, "Alerte introuvable")
    a.archived = 1
    try:
        db.commit()
    except Exception:
        db.rollback()
    return {"ok": True}


class ReplyIn(BaseModel):
    message: str
    canal:   Optional[str] = None   # "mail" | "sms" | None(auto)


@router.post("/alerte/{alerte_id}/cible/{cible_id}/reply")
async def reply_cible(alerte_id: int, cible_id: int, body: ReplyIn, request: Request,
                      db: Session = Depends(get_db), admin=Depends(require_admin)):
    """h85 — Répond individuellement à une personne mobilisée, par mail ou SMS
    selon les coordonnées disponibles (ou le canal demandé)."""
    cible = (db.query(AlerteCible)
               .filter(AlerteCible.id == cible_id,
                       AlerteCible.alerte_id == alerte_id).first())
    if not cible:
        raise HTTPException(404, "Destinataire introuvable")
    msg = (body.message or "").strip()
    if not msg:
        raise HTTPException(400, "Message vide")
    a = db.query(AlerteMobilisation).filter(AlerteMobilisation.id == alerte_id).first()
    from plugins.notifications.backends import NotifPayload
    payload = NotifPayload(event_type="mobilisation_reply",
                           title=("Réponse — " + (a.titre if a else "Mobilisation")),
                           body=msg, urgency=3, context={})
    canal = (body.canal or "").strip().lower()
    # Choix du canal : demandé en priorité, sinon SMS si tél, sinon mail.
    order = []
    if canal in ("sms", "mail"):
        order = [canal]
    else:
        if cible.tel:
            order.append("sms")
        if cible.email:
            order.append("mail")
    for k in order:
        target = cible.tel if k == "sms" else cible.email
        if not target:
            continue
        backend = _load_backend(db, k)
        if not backend:
            continue
        try:
            res = await backend.send(payload, target)
            if getattr(res, "success", False):
                return {"ok": True, "via": k}
        except Exception:
            pass
    raise HTTPException(502, "Aucun canal disponible ou envoi échoué")


# ── Formulaire ETA public (tokenisé, responsive, multilingue) ────────────────

_ETA_FALLBACK = {
    "eta_title": "Mobilisation SCRIBE",
    "eta_intro": "Bonjour {nom}, vous êtes mobilisé(e).",
    "eta_context": "{fonction} — {site}",
    "eta_question": "Dans combien de temps pensez-vous être en salle de crise ?",
    "eta_15": "Moins de 15 min",
    "eta_30": "Environ 30 min",
    "eta_60": "Environ 1 h",
    "eta_indispo": "Indisponible",
    "eta_thanks": "Merci, votre réponse a bien été enregistrée.",
    "eta_already": "Vous avez déjà répondu :",
    "eta_choose": "Sélectionnez votre délai :",
    "eta_comment_label": "Précision (optionnel)",
    "eta_comment_ph": "Ex. j'arrive directement en réanimation…",
}


def _eta_strings(lang):
    """Charge les libellés ETA dans la langue demandée depuis app/lang/{lang}.json
    (clés `mobilisation.*`), avec repli FR puis fallback intégré."""
    out = dict(_ETA_FALLBACK)
    for code in (lang, "fr"):
        if not code:
            continue
        path = os.path.join(os.path.dirname(__file__), "..", "lang", f"{code}.json")
        try:
            with open(path, encoding="utf-8") as f:
                d = json.load(f)
            mob = d.get("mobilisation", {})
            for k in _ETA_FALLBACK:
                if mob.get(k):
                    out[k] = mob[k]
            break
        except Exception:
            continue
    return out


def _render_eta_page(cible, alerte, lang, done=False, choice=None):
    s = _eta_strings(lang)
    def e(x): return html.escape(str(x or ""))
    nom = e(cible.nom)
    intro = s["eta_intro"].replace("{nom}", nom)
    ctx = s["eta_context"].replace("{fonction}", e(cible.fonction)).replace("{site}", e(cible.site))
    choice_labels = {"15": s["eta_15"], "30": s["eta_30"], "60": s["eta_60"], "indispo": s["eta_indispo"]}
    if done:
        body_html = (
            f'<p class="ok">✅ {e(s["eta_thanks"])}</p>'
            f'<p class="chosen">{e(s["eta_already"])} <b>{e(choice_labels.get(choice, choice))}</b></p>'
        )
        if getattr(cible, "commentaire", None):
            body_html += f'<p class="cmt">💬 {e(cible.commentaire)}</p>'
    else:
        btns = "".join(
            f'<button type="submit" name="choix" value="{val}" class="eta-btn eta-{val}">{e(lbl)}</button>'
            for val, lbl in [("15", s["eta_15"]), ("30", s["eta_30"]),
                             ("60", s["eta_60"]), ("indispo", s["eta_indispo"])]
        )
        body_html = (
            f'<p class="q">{e(s["eta_question"])}</p>'
            f'<form method="post" class="eta-form">'
            f'<label class="cmt-label">{e(s["eta_comment_label"])}</label>'
            f'<textarea name="commentaire" class="cmt-input" rows="2" maxlength="500" '
            f'placeholder="{e(s["eta_comment_ph"])}"></textarea>'
            f'{btns}</form>'
        )
    msg = e(alerte.message) if alerte and alerte.message else ""
    title = e(alerte.titre if alerte else s["eta_title"])
    return f"""<!DOCTYPE html>
<html lang="{e(lang or 'fr')}"><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1">
<title>{e(s['eta_title'])}</title>
<style>
  :root {{ --bleu:#003189; --rouge:#e1000f; }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif;
          background:#f8fafc; color:#0f172a; padding:18px; }}
  .card {{ max-width:480px; margin:24px auto; background:#fff; border:1px solid #e2e8f0;
           border-radius:14px; overflow:hidden; box-shadow:0 8px 30px rgba(0,0,0,.08); }}
  .head {{ background:var(--bleu); color:#fff; padding:16px 20px; font-weight:700;
           font-size:16px; letter-spacing:.3px; }}
  .body {{ padding:20px; }}
  .intro {{ font-size:16px; margin:0 0 4px; }}
  .ctx {{ color:#64748b; font-size:13px; margin:0 0 14px; }}
  .alert-msg {{ background:#fff7ed; border-left:3px solid var(--rouge); padding:10px 12px;
               border-radius:6px; font-size:14px; color:#7c2d12; margin:0 0 18px; white-space:pre-wrap; }}
  .q {{ font-weight:600; font-size:15px; margin:0 0 12px; }}
  .eta-form {{ display:flex; flex-direction:column; gap:10px; }}
  .eta-btn {{ font-size:17px; padding:16px; border-radius:10px; border:2px solid var(--bleu);
              background:#fff; color:var(--bleu); font-weight:700; cursor:pointer;
              -webkit-tap-highlight-color:transparent; }}
  .eta-btn:active {{ transform:scale(.98); }}
  .eta-15 {{ border-color:#16a34a; color:#16a34a; }}
  .eta-30 {{ border-color:#0ea5e9; color:#0ea5e9; }}
  .eta-60 {{ border-color:#f59e0b; color:#b45309; }}
  .eta-indispo {{ border-color:#dc2626; color:#dc2626; }}
  .cmt-label {{ font-family:inherit; font-size:12px; color:#64748b; margin-bottom:-4px; }}
  .cmt-input {{ width:100%; box-sizing:border-box; padding:10px; border:1px solid #cbd5e1;
                border-radius:8px; font:inherit; font-size:14px; resize:vertical; }}
  .cmt {{ background:#f1f5f9; border-radius:8px; padding:10px 12px; font-size:14px; color:#334155; margin-top:10px; }}
  .ok {{ font-size:18px; color:#16a34a; font-weight:700; }}
  .chosen {{ font-size:15px; color:#334155; }}
  .foot {{ background:#f1f5f9; color:#94a3b8; font-size:11px; text-align:center; padding:10px; }}
</style></head>
<body>
  <div class="card">
    <div class="head">📣 {title}</div>
    <div class="body">
      <p class="intro">{intro}</p>
      <p class="ctx">{ctx}</p>
      {f'<div class="alert-msg">{msg}</div>' if msg else ''}
      {body_html}
    </div>
    <div class="foot">SCRIBE — Gestion de crise hospitalière</div>
  </div>
</body></html>"""


@router.get("/eta/{token}", response_class=HTMLResponse)
def eta_form(token: str, request: Request, lang: str = "fr", db: Session = Depends(get_db)):
    cible = db.query(AlerteCible).filter(AlerteCible.token == token).first()
    if not cible:
        return HTMLResponse("<h3 style='font-family:sans-serif;text-align:center;margin-top:40px'>Lien invalide ou expiré.</h3>", status_code=404)
    alerte = db.query(AlerteMobilisation).filter(AlerteMobilisation.id == cible.alerte_id).first()
    done = (cible.statut == "repondu")
    return HTMLResponse(_render_eta_page(cible, alerte, lang, done=done, choice=cible.eta_choice))


@router.post("/eta/{token}", response_class=HTMLResponse)
async def eta_submit(token: str, request: Request, lang: str = "fr", db: Session = Depends(get_db)):
    cible = db.query(AlerteCible).filter(AlerteCible.token == token).first()
    if not cible:
        return HTMLResponse("<h3 style='font-family:sans-serif;text-align:center;margin-top:40px'>Lien invalide.</h3>", status_code=404)
    form = await request.form()
    choix = str(form.get("choix") or "").strip()
    commentaire = str(form.get("commentaire") or "").strip()[:500]
    if choix in ("15", "30", "60", "indispo") and cible.statut != "repondu":
        cible.eta_choice = choix
        cible.commentaire = commentaire or None
        cible.statut = "repondu"
        cible.responded_at = datetime.now(timezone.utc)
        try:
            db.commit()
        except Exception:
            db.rollback()
    alerte = db.query(AlerteMobilisation).filter(AlerteMobilisation.id == cible.alerte_id).first()
    return HTMLResponse(_render_eta_page(cible, alerte, lang, done=True, choice=cible.eta_choice))
